import tkinter as tk
from tkinter import ttk

from tkinter.ttk import Treeview, Style

import need.tkutils as tku
from tkinter.filedialog import askdirectory
import re
import os
from PIL import Image

from time import time
from time import localtime
from time import strftime
import datetime
import pandas as pd
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl import load_workbook

import warnings
import pdfplumber
import PyPDF2

from collections import defaultdict

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors

from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Image, Table, TableStyle, NextPageTemplate, PageBreak
from cryptography.fernet import Fernet
import base64
import pyrfc
import traceback
import datetime
import logging
from need.custom_dialogs import CustomDialog, center_window, Tooltip, image_label

warnings.simplefilter(action='ignore', category=FutureWarning)

FilePath = ""  # 设置一个地址变量
pd.set_option('display.max_columns', None)  # 显示全部列
pd.set_option('display.max_row', None)  # 显示全部行
pd.set_option('display.width', 100)  # 设置数据的显示长度（解决自动换行）


def main(parent, w_ratio, h_ratio, file_path):
    global checklog_file_path
    checklog_file_path = file_path
    global open_table
    open_table = tk.PhotoImage(file="ico\\down.png")
    global close_table
    close_table = tk.PhotoImage(file="ico\\up.png")
    global open_folder
    open_folder = tk.PhotoImage(file="ico\\open_folder.png")
    global analyze_file
    analyze_file = tk.PhotoImage(file="ico\\read.png")
    global view_folder
    view_folder = tk.PhotoImage(file="ico\\view.png")
    global pdf_export
    pdf_export = tk.PhotoImage(file="ico\\export.png")

    global canvas
    canvas = tk.Canvas(parent, width=int(1750 * w_ratio), height=int(640 * h_ratio), bg="#C9DBE9")
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    canvas.update()
    canvas.bind("<MouseWheel>", on_mousewheel)

    scrollbar_v = tk.Scrollbar(master=parent)
    scrollbar_v.pack(side=tk.RIGHT, fill=tk.Y)
    scrollbar_v.config(command=canvas.yview)
    canvas.config(yscrollcommand=scrollbar_v.set)

    global content
    content = tk.Frame(canvas)
    # content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    canvas.create_window(0, 1, width=int(1750 * w_ratio), anchor=tk.NW, window=content)

    f1 = tk.Frame(content, bg="#c9dbe9", bd=0)
    tk.Label(f1, text="欢迎使用BOM检查功能", bg="#c9dbe9", fg="black", height=int(1 * h_ratio), font=("ABBvoice CNSG", int(20 * h_ratio), "bold")).pack(fill=tk.X)
    f1.pack(fill=tk.X)

    f2 = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Label(f2, text='   说明：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')
    tk.Label(f2, text='(1)EPLAN菜单Tools→Reports:Automated processing中选择Export Files V04导出；\n(2)选择BOM文件【C:/Temp/项目号-Files/项目号-BOM.xlsx】；\n(3)本功能需要导出整个项目pdf文件，且单线图为标准单线图，非标单线图无法正常识别，将会影响柜体参数的识别，进而后续与柜体参数相关的检查均无法执行。', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    f2.pack(fill=tk.X)

    tk.Frame(content, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f3 = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Label(f3, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f3, text='路径：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry  # 为了确保selectpath函数能正确调用entry,将其全局化
    entry = tk.Entry(f3, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)))
    entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

    global button_export_report
    button_export_report = tk.Button(f3, image=pdf_export, text="导出报表", font=("ABBvoice CNSG", int(13 * h_ratio)), bg="#eaf1f6", command=export_report, compound=tk.LEFT, state='disabled', activebackground='blue')
    button_export_report.pack(side=tk.RIGHT, padx=0)

    global button_open
    button_open = tk.Button(f3, image=view_folder, text="看文件夹", font=("ABBvoice CNSG", int(13 * h_ratio)), bg="#eaf1f6", command=open_filefolder, compound=tk.LEFT, state='disabled', activebackground='blue')
    button_open.pack(side=tk.RIGHT, padx=int(20 * w_ratio))

    global button2
    button2 = tk.Button(f3, text='检查', font=("ABBvoice CNSG", int(13 * h_ratio)), image=analyze_file, bg="#eaf1f6", compound=tk.LEFT, command=process1, activebackground='blue')
    button2.pack(side=tk.RIGHT, padx=0)
    button2['state'] = 'disabled'

    tk.Button(f3, text='选择', font=("ABBvoice CNSG", int(13 * h_ratio)), image=open_folder, bg="#eaf1f6", compound=tk.LEFT, command=selectpath, activebackground='blue').pack(side=tk.RIGHT, padx=int(20 * w_ratio))

    f3.pack(fill=tk.X)

    tk.Frame(content, height=int(30 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f40 = tk.Frame(content, bg="#eaf1f6", bd=0)
    f40L = tk.Frame(f40, bg="#eaf1f6", bd=0)
    f40L.pack(side=tk.LEFT, fill=tk.X, anchor='n')

    style = Style()
    style.configure('panel.Treeview', rowheight=30, font=("ABBvoice CNSG", int(13 * h_ratio)))
    style.configure('panel.Treeview.Heading', font=("ABBvoice CNSG", int(13 * h_ratio)), background="#EFF1F5")

    global table0_display_flag
    table0_display_flag = 1
    global table0_display_button
    tk.Label(f40L, text='   柜型：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(fill=tk.X, anchor='n')
    table0_display_button = tk.Button(f40L, image=open_table, bg="#eaf1f6", command=change_table0_display)
    table0_display_button.pack(pady=(0, 0))
    table0_display_button['state'] = 'disabled'
    Tooltip(table0_display_button, "展开/折叠")

    global Calculate_table0
    table_ybar0 = tk.Scrollbar(f40, orient="vertical")
    table_ybar0.pack(side=tk.RIGHT, fill=tk.Y)

    Calculate_table0 = Treeview(f40, show='headings', style='panel.Treeview', selectmode='browse', columns=('a', 'b', 'c', 'd', 'e', 'f'), yscrollcommand=table_ybar0.set, height=1)
    Calculate_table0.tag_configure('error_line', foreground='red')  # 多个尺寸时显示红色

    table_ybar0.config(command=Calculate_table0.yview)

    Calculate_table0.column('a', width=int(70 * w_ratio), anchor='center')
    Calculate_table0.column('b', width=int(70 * w_ratio), anchor='center')
    Calculate_table0.column('c', width=int(650 * w_ratio), anchor='center')
    Calculate_table0.column('d', width=int(200 * w_ratio), anchor='center')
    Calculate_table0.column('e', width=int(100 * w_ratio), anchor='center')
    Calculate_table0.column('f', width=int(350 * w_ratio), anchor='center')

    Calculate_table0.heading('a', text='站号', anchor='center')
    Calculate_table0.heading('b', text='柜型', anchor='center')
    Calculate_table0.heading('c', text='柜号', anchor='center')
    Calculate_table0.heading('d', text='尺寸', anchor='center')
    Calculate_table0.heading('e', text='产品类型', anchor='center')
    Calculate_table0.heading('f', text='主开关', anchor='center')

    Calculate_table0.pack(fill=tk.BOTH, padx=(0, 1), pady=0)

    f40.pack(fill=tk.BOTH)

    tk.Frame(content, height=int(30 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f40 = tk.Frame(content, bg="#eaf1f6", bd=0)
    f40L = tk.Frame(f40, bg="#eaf1f6", bd=0)
    f40L.pack(side=tk.LEFT, fill=tk.X, anchor='n')

    global table1_display_flag
    table1_display_flag = 1
    global table1_display_button
    tk.Label(f40L, text='   空开：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(fill=tk.X, anchor='n')
    table1_display_button = tk.Button(f40L, image=open_table, bg="#eaf1f6", command=change_table1_display)
    table1_display_button.pack(pady=(0, 0))
    table1_display_button['state'] = 'disabled'
    Tooltip(table1_display_button, "展开/折叠")

    global Calculate_table1
    table_ybar1 = tk.Scrollbar(f40, orient="vertical")
    table_ybar1.pack(side=tk.RIGHT, fill=tk.Y)

    Calculate_table1 = Treeview(f40, show='headings', style='panel.Treeview', selectmode='browse', columns=('a', 'b', 'c', 'd', 'e'), yscrollcommand=table_ybar1.set, height=1)
    table_ybar1.config(command=Calculate_table1.yview)

    Calculate_table1.column('a', width=int(100 * w_ratio), anchor='center')
    Calculate_table1.column('b', width=int(100 * w_ratio), anchor='center')
    Calculate_table1.column('c', width=int(100 * w_ratio), anchor='center')
    Calculate_table1.column('d', width=int(100 * w_ratio), anchor='center')
    Calculate_table1.column('e', width=int(100 * w_ratio), anchor='center')

    Calculate_table1.heading('a', text='柜型', anchor='center')
    Calculate_table1.heading('b', text='标签', anchor='center')
    Calculate_table1.heading('c', text='电压等级', anchor='center')
    Calculate_table1.heading('d', text='厂家', anchor='center')
    Calculate_table1.heading('e', text='辅助触点', anchor='center')

    Calculate_table1.pack(fill=tk.BOTH, padx=(0, 1), pady=0)

    f40.pack(fill=tk.BOTH)

    tk.Frame(content, height=int(30 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f41 = tk.Frame(content, bg="#eaf1f6", bd=0)
    f41L = tk.Frame(f41, bg="#eaf1f6", bd=0)
    f41L.pack(side=tk.LEFT, fill=tk.X, anchor='n')

    global table2_display_flag
    table2_display_flag = 1
    global table2_display_button

    tk.Label(f41L, text='   其他：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(fill=tk.X, anchor='n')
    table2_display_button = tk.Button(f41L, image=open_table, bg="#eaf1f6", command=change_table2_display)
    table2_display_button.pack(pady=(0, 0))
    table2_display_button['state'] = 'disabled'
    Tooltip(table2_display_button, "展开/折叠")

    global entry_typical
    entry_typical = tk.Entry(f41L, bg="#eaf1f6", font=("ABBvoice CNSG", int(10 * h_ratio)), justify='center', width=int(8 * w_ratio))
    entry_typical.pack(pady=(int(5 * h_ratio), 0))
    entry_typical['state'] = 'disabled'

    global Calculate_table2
    table_ybar2 = tk.Scrollbar(f41, orient="vertical")
    table_ybar2.pack(side=tk.RIGHT, fill=tk.Y)

    table_xbar2 = tk.Scrollbar(f41, orient="horizontal")
    table_xbar2.pack(side=tk.BOTTOM, fill=tk.X)

    Calculate_table2 = Treeview(f41, show='headings', style='panel.Treeview', selectmode='browse', columns=('a', 'c', 'd', 'e', 'f', 'g', 'h'), yscrollcommand=table_ybar2.set, xscrollcommand=table_xbar2.set, height=1)
    table_ybar2.config(command=Calculate_table2.yview)
    table_xbar2.config(command=Calculate_table2.xview)

    Calculate_table2.column('a', width=int(80 * w_ratio), anchor='center')
    Calculate_table2.column('c', width=int(120 * w_ratio), anchor='center')
    Calculate_table2.column('d', width=int(120 * w_ratio), anchor='center')
    Calculate_table2.column('e', width=int(120 * w_ratio), anchor='center')
    Calculate_table2.column('f', width=int(400 * w_ratio), anchor='center')
    Calculate_table2.column('g', width=int(400 * w_ratio), anchor='center')
    Calculate_table2.column('h', width=int(120 * w_ratio), anchor='center')
    Calculate_table2.heading('a', text='柜型', anchor='center')
    Calculate_table2.heading('c', text='低压室照明灯', anchor='center')
    Calculate_table2.heading('d', text='熔断器', anchor='center')
    Calculate_table2.heading('e', text='避雷器', anchor='center')
    Calculate_table2.heading('f', text='压板', anchor='center')
    Calculate_table2.heading('g', text='转换开关', anchor='center')
    Calculate_table2.heading('h', text='容性分压装置', anchor='center')
    Calculate_table2.pack(fill=tk.BOTH, padx=(0, 1), pady=0)
    Calculate_table2.bind('<<TreeviewSelect>>', on_treeview_select)

    Calculate_table2.tag_configure('fontsize', font=("ABBvoice CNSG", int(10 * h_ratio)))
    f41.pack(fill=tk.BOTH)

    tk.Frame(content, height=int(30 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f4 = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Label(f4, text='   结果：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')

    global text
    text = tk.Text(f4, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), height=int(50 * h_ratio), width=int(55 * w_ratio))
    text.pack(side=tk.LEFT, padx=(0, 1), pady=0, fill=tk.BOTH, expand=True)

    text.tag_configure('error', foreground='red')  # 设置tag

    scrollbar = tk.Scrollbar(f4)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    scrollbar.config(command=text.yview)
    text.config(yscrollcommand=scrollbar.set)
    f4.pack(fill=tk.X)

    tk.Frame(content, height=int(80 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    canvas.update_idletasks()
    # content.update_idletasks()
    canvas.config(scrollregion=canvas.bbox('all'))


def open_filefolder():
    os.startfile(os.path.dirname(FilePath))


def change_table0_display():
    global dataframe_groupby
    global table0_display_flag
    global table0_display_button
    global open_table
    global close_table
    if table0_display_flag:
        table0_display_button.configure(image=close_table)
        table0_display_flag = 0
        Calculate_table0.configure(height=len(dataframe_groupby['A']))
    else:
        table0_display_button.configure(image=open_table)
        table0_display_flag = 1
        Calculate_table0.configure(height=1)
    # 更新滚动条
    content.update_idletasks()
    canvas.config(scrollregion=canvas.bbox('all'))


def change_table1_display():
    global MCB_typical_list
    global table1_display_flag
    global table1_display_button
    global open_table
    global close_table
    if table1_display_flag:
        table1_display_button.configure(image=close_table)
        table1_display_flag = 0
        Calculate_table1.configure(height=len(MCB_typical_list))
    else:
        table1_display_button.configure(image=open_table)
        table1_display_flag = 1
        Calculate_table1.configure(height=1)
    # 更新滚动条
    content.update_idletasks()
    canvas.config(scrollregion=canvas.bbox('all'))


def change_table2_display():
    global LED_typical_list
    global table2_display_flag
    global table2_display_button
    global open_table
    global close_table
    if table2_display_flag:
        table2_display_button.configure(image=close_table)
        table2_display_flag = 0
        Calculate_table2.configure(height=len(LED_typical_list))
    else:
        table2_display_button.configure(image=open_table)
        table2_display_flag = 1
        Calculate_table2.configure(height=1)
    # 更新滚动条
    content.update_idletasks()
    canvas.config(scrollregion=canvas.bbox('all'))


def on_mousewheel(event):
    global canvas
    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")


def on_treeview_select(event):
    selected_item = Calculate_table2.focus()
    values = Calculate_table2.item(selected_item, 'values')

    # 清空Entry的数据
    entry_typical['state'] = 'normal'
    entry_typical.delete(0, tk.END)

    # 填入新数据
    if values:
        # print(values)
        entry_typical.insert(0, values[0])
    entry_typical['state'] = 'disabled'


def about_help(event):
    # tku.show_info("说明书")
    os.startfile(os.path.abspath('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\二次设计辅助工具【EBOM导SAP功能】答疑V1.2.pdf'))


def selectpath():
    filepath = tk.filedialog.askopenfilename(initialdir='C:/Temp/', title=u'请选择文件(XXX-BOM.xlsx)', filetypes=[("Excel", ".xlsx")])  # 选择打开什么文件，返回文件名
    stem, suffix = os.path.splitext(os.path.basename(filepath))  # stem是文件名,suffix是后缀
    # print(stem, suffix)

    if len(filepath) != 0:
        string_filename = ""
        for i in range(0, len(filepath)):
            string_filename += str(filepath[i])
        button2['state'] = 'normal'
        button_open['state'] = 'normal'

    else:
        button2['state'] = 'disabled'
        table0_display_button['state'] = 'disabled'
        table1_display_button['state'] = 'disabled'
        table2_display_button['state'] = 'disabled'
        button_open['state'] = 'disabled'
        button_export_report['state'] = 'disabled'

    text.delete(1.0, tk.END)  # 清空输出结果框

    # 清空treeview表格
    table_items = Calculate_table0.get_children()  # 在插入treeview数据时，需要先清空treeview
    [Calculate_table0.delete(table_item) for table_item in table_items]

    table_items = Calculate_table1.get_children()  # 在插入treeview数据时，需要先清空treeview
    [Calculate_table1.delete(table_item) for table_item in table_items]

    table_items = Calculate_table2.get_children()  # 在插入treeview数据时，需要先清空treeview
    [Calculate_table2.delete(table_item) for table_item in table_items]

    entry.delete(0, "end")  # 删除entry原始内容
    entry.insert(0, filepath)  # 重新填入地址

    entry_typical['state'] = 'normal'
    entry_typical.delete(0, tk.END)
    entry_typical['state'] = 'disabled'
    table0_display_button['state'] = 'disabled'
    table1_display_button['state'] = 'disabled'
    table2_display_button['state'] = 'disabled'
    button_export_report['state'] = 'disabled'
    global FilePath
    FilePath = filepath


# ”检查“按钮
def process1():
    try:
        error1_calculator = 0
        error2_calculator = 0
        error3_calculator = 0
        error4_calculator = 0
        error5_calculator = 0
        error6_calculator = 0
        error7_calculator = 0

        error15_calculator = 0
        error16_calculator = 0
        error17_calculator = 0
        error18_calculator = 0
        error19_calculator = 0
        error20_calculator = 0
        error21_calculator = 0
        error22_calculator = 0
        error23_calculator = 0
        error24_calculator = 0

        text.delete(1.0, tk.END)  # 清空输出结果框

        table_items = Calculate_table0.get_children()  # 在插入treeview数据时，需要先清空treeview
        [Calculate_table0.delete(table_item) for table_item in table_items]

        table_items = Calculate_table1.get_children()  # 在插入treeview数据时，需要先清空treeview
        [Calculate_table1.delete(table_item) for table_item in table_items]

        table_items = Calculate_table2.get_children()  # 在插入treeview数据时，需要先清空treeview
        [Calculate_table2.delete(table_item) for table_item in table_items]

        entry_typical['state'] = 'normal'
        entry_typical.delete(0, tk.END)
        entry_typical['state'] = 'disabled'

        global stem
        base_name = os.path.basename(FilePath)  # 获取路径的文件名（包括后缀）
        stem, suffix = os.path.splitext(base_name)  # stem是文件名,suffix是后缀

        if FilePath == "":
            tk.messagebox.showwarning("提示", "请选择文件！")
        elif '-BOM' not in stem:
            tk.messagebox.showwarning("提示", "请选择XXX-BOM.xlsx文件！")
        else:
            new_stem = stem.replace('-BOM', '')
            pdf_file_path = os.path.join(os.path.dirname(FilePath), new_stem + '.pdf')

            new_stem = stem.replace('-BOM', '-SLD tables')
            sld_table_path = os.path.join(os.path.dirname(FilePath), new_stem + '.xls')

            new_stem = stem.replace('-BOM', '-Panel')
            Panel_file_path = os.path.join(os.path.dirname(FilePath), new_stem + '.xlsx')

            new_stem = stem.replace('-BOM', '-Panel-replaced')
            Replaced_Panel_file_path = os.path.join(os.path.dirname(FilePath), new_stem + '.xlsx')

            new_stem = stem.replace('-BOM', '-Panel data')
            Panel_data_file_path = os.path.join(os.path.dirname(FilePath), new_stem + '.xlsx')

            new_stem = stem.replace('-BOM', '-Panel Size')
            Panel_Size_file_path = os.path.join(os.path.dirname(FilePath), new_stem + '.xls')

            new_stem = stem.replace('-BOM', '-PBOM')
            Panel_BOM_file_path = os.path.join(os.path.dirname(FilePath), new_stem + '.xlsx')

            new_stem = stem.replace('-BOM', '-DeviceLabel')
            Device_Label_file_path = os.path.join(os.path.dirname(FilePath), new_stem + '.xlsx')

            if not os.path.exists(pdf_file_path):
                tk.messagebox.showwarning("提示", "文件夹下缺少XXX.pdf文件！")
            # elif not os.path.exists(Panel_file_path):
            #     tk.messagebox.showwarning("提示", "文件夹下缺少XXX-Panel.xlsx文件！")

            elif not os.path.exists(Panel_data_file_path):
                tk.messagebox.showwarning("提示", "文件夹下缺少XXX-Panel data.xlsx文件！")
            elif not os.path.exists(Panel_BOM_file_path):
                tk.messagebox.showwarning("提示", "文件夹下缺少XXX-PBOM.xlsx文件！")
            elif not os.path.exists(Device_Label_file_path):
                tk.messagebox.showwarning("提示", "文件夹下缺少XXX-DeviceLabel.xlsx文件！")

            else:
                start = time()
                text.insert(tk.INSERT, '>>>EPLAN版SLD正在识别中...\n')
                sld_recognition(pdf_file_path, sld_table_path)
                sld_book = xlrd.open_workbook(sld_table_path, formatting_info=False)
                switchgear_number = []
                typical_type_list = []
                abb_panel_number = []
                abb_panel_number_ori = []
                switchgear_dimension = []
                product_type = []
                voltage_level = []
                cb_type = []
                control_source_voltage = []
                charge_source_voltage = []

                for sheet in sld_book.sheets():
                    sheet_name = sheet.name
                    if '&' in sheet_name:
                        separated_name = sheet_name.split('&')[0]

                    num_cols = sheet.ncols
                    num_rows = sheet.nrows

                    for row in range(1, num_rows):
                        if '二次图纸' in sheet.cell_value(row, 1) or 'SECONDARY DRAWING' in sheet.cell_value(row, 1):
                            row_typical = row
                            break
                        else:
                            row_typical = 2

                    for row in range(1, num_rows):
                        if 'ABB 柜号' in sheet.cell_value(row, 1) or 'ABB PANEL NO.' in sheet.cell_value(row, 1) or 'ABB NO.' in sheet.cell_value(row, 1):
                            row_abbpanelno = row
                            break
                        else:
                            row_abbpanelno = 3

                    for row in range(1, num_rows):
                        if '开关柜尺寸' in sheet.cell_value(row, 1) or 'DIMENSION' in sheet.cell_value(row, 1):
                            row_dimension = row
                            break
                        else:
                            row_dimension = 6

                    for row in range(1, num_rows):
                        if '主开关' in sheet.cell_value(row, 1) or 'MAIN SWITCH' in sheet.cell_value(row, 1):
                            row_cb = row
                            break
                        else:
                            row_cb = 10

                    for row in range(1, num_rows):
                        if '额定电压' in sheet.cell_value(row, 0) or 'RATED VOLTAGE' in sheet.cell_value(row, 0):
                            row_voltage = row
                            break
                        else:
                            row_voltage = 2

                    for row in range(1, num_rows):
                        if '控制电源' in sheet.cell_value(row, 0) or 'CONTROL POWER' in sheet.cell_value(row, 0):
                            row_control = row
                            break
                        else:
                            row_control = 7

                    for row in range(1, num_rows):
                        if '储能马达' in sheet.cell_value(row, 0) or 'CHARGE MOTOR' in sheet.cell_value(row, 0):
                            row_charge = row
                            break
                        else:
                            row_charge = 9

                    def is_valid_string(s):
                        if s == '':
                            return False
                        elif s[0] != 'A':
                            return False
                        elif not s[1:3].isdigit() or int(s[1:3]) < 1 or int(s[1:3]) > 99:
                            return False
                        elif s[3] not in ['/', '\\']:
                            return False
                        elif len(s.replace('/', '\\').split('\\')) != 2:
                            return False
                        elif len(s.replace('/', '\\').split('\\')) == 2 and s.replace('/', '\\').split('\\')[1] == '':
                            return False
                        return True

                    for col in range(2, num_cols):
                        if is_valid_string(sheet.cell_value(row_typical, col)):
                            cell_value_typical = sheet.cell_value(row_typical, col).replace(' ', '').replace('/', '\\')
                            cell_value_abbpanelno = sheet.cell_value(row_abbpanelno, col)
                            cell_value_dimension = sheet.cell_value(row_dimension, col)

                            product_type.append(sheet.cell_value(1, 0).replace('\n', '').split(':')[-1])
                            # print(product_type)

                            switchgear_number.append(separated_name)
                            typical_type_list.append(cell_value_typical.split('\\')[1])
                            abb_panel_number.append(cell_value_abbpanelno)
                            abb_panel_number_ori.append(cell_value_abbpanelno)
                            switchgear_dimension.append(cell_value_dimension.replace(' ', '').replace('  ', '').replace('   ', '').replace('M', '').replace('m', ''))

                            voltage_level.append(sheet.cell_value(row_voltage, 0))
                            cb_type.append(sheet.cell_value(row_cb, col))
                            control_source_voltage.append(sheet.cell_value(row_control, 0))
                            charge_source_voltage.append(sheet.cell_value(row_charge, 0))

                # print(product_type)
                # print(switchgear_number)
                # print(typical_type_list)
                # print(abb_panel_number)
                # print(switchgear_dimension)
                # print(voltage_level)
                # print(cb_type)
                # print(control_source_voltage)
                # print(charge_source_voltage)

                for i in range(0, len(abb_panel_number)):
                    abb_panel_number[i] = abb_panel_number[i] + ';'

                for i in range(0, len(product_type)):
                    if 'ZS1' in product_type[i]:
                        if '325' == switchgear_dimension[i].split('X')[0] or '375' == switchgear_dimension[i].split('X')[0]:
                            product_type[i] = 'ZVC'
                        elif '500' in switchgear_dimension[i].split('X')[0]:
                            product_type[i] = '500'
                        elif '550' in switchgear_dimension[i].split('X')[0]:
                            product_type[i] = '550'
                        elif ('650' in switchgear_dimension[i].split('X')[0] or '800' in switchgear_dimension[i].split('X')[0] or '1000' in switchgear_dimension[i].split('X')[0]) and ('1310' in switchgear_dimension[i].split('X')[2] or '1620' in switchgear_dimension[i].split('X')[2] or '1810' in switchgear_dimension[i].split('X')[2] or '1310+310' in switchgear_dimension[i].split('X')[2]):
                            product_type[i] = 'Beni'
                        elif '空柜' in typical_type_list[i] or 'DUMMY' in typical_type_list[i].upper():
                            product_type[i] = '空柜'
                        else:
                            product_type[i] = 'ZS1'
                    elif 'ZS3.2' in product_type[i]:
                        product_type[i] = 'ZS3.2'
                    elif 'ZX0' in product_type[i]:
                        if 'ZX0 Air' not in product_type[i] and 'ZX0.2' not in product_type[i]:
                            product_type[i] = 'ZX0'
                        elif 'ZX0 Air' in product_type[i]:
                            product_type[i] = 'ZX0 Air'
                        else:
                            product_type[i] = 'ZX0.2'

                    elif 'ZX1.2' in product_type[i]:
                        product_type[i] = 'ZX1.2'
                    elif 'ZX1.5-R' in product_type[i]:
                        product_type[i] = 'ZX1.5-R'
                    elif 'ZX2' in product_type[i]:
                        if 'ZX2 AirPlus' not in product_type[i]:
                            product_type[i] = 'ZX2'
                        else:
                            product_type[i] = 'ZX2 AirPlus'
                    elif 'PrimeGear ZX0' in product_type[i]:
                        product_type[i] = 'PrimeGear ZX0'

                # 定义一个函数用来将字符串进行去重拼接
                def unique_concatenate(series):
                    return ';'.join(set(series))

                def shorten_series(series, max_items=8):
                    result = []
                    for item in series:
                        items = item.split(';')
                        if len(items) > max_items:
                            shortened_items = items[:max_items]
                            truncated_str = ';'.join(shortened_items)
                            result.append(f"{truncated_str} (...共{len(items)}个)")
                        else:
                            result.append(';'.join(items))
                    return pd.Series(result)

                global dataframe
                global dataframe_groupby
                dataframe = pd.DataFrame({'A': switchgear_number, 'B': typical_type_list, 'C': abb_panel_number, 'D': switchgear_dimension, 'E': product_type, 'F': cb_type})
                dataframe_groupby = dataframe.groupby(['A', 'B']).agg({'C': 'sum', 'D': unique_concatenate, 'E': unique_concatenate, 'F': unique_concatenate}).reset_index()

                # print(dataframe_groupby)
                # 使用自定义函数缩短series C中的元素
                dataframe_groupby['C'] = shorten_series(dataframe_groupby['C'])
                # print(dataframe_groupby)

                global dataframe_groupby_rev
                dataframe_rev = pd.DataFrame({'A': switchgear_number, 'B': typical_type_list, 'C': abb_panel_number, 'D': switchgear_dimension, 'E': product_type, 'F': voltage_level, 'G': cb_type, 'H': control_source_voltage, 'I': charge_source_voltage})
                dataframe_groupby_rev = dataframe_rev.groupby(['A', 'B']).agg({'C': 'sum', 'D': unique_concatenate, 'E': unique_concatenate, 'F': unique_concatenate, 'G': unique_concatenate, 'H': unique_concatenate, 'I': unique_concatenate}).reset_index()

                # print(dataframe_groupby_rev)

                table_items = Calculate_table0.get_children()  # 在插入treeview数据时，需要先清空treeview
                [Calculate_table0.delete(table_item) for table_item in table_items]

                for i in range(0, len(dataframe_groupby['A'])):
                    if ';' in list(dataframe_groupby['E'][i]):
                        Calculate_table0.insert('', 'end', values=list(dataframe_groupby.loc[i]), tags="error_line")
                    else:
                        Calculate_table0.insert('', 'end', values=list(dataframe_groupby.loc[i]))
                table0_display_button['state'] = 'normal'

                # 改用PDF图纸中的单线图生成Panel-replaced.xlsx
                # 1. 组装只含三列的 DataFrame
                panel_df = pd.DataFrame({
                    'Order Line': switchgear_number,
                    'Typical': typical_type_list,
                    'Panel No': abb_panel_number_ori,
                })

                # 2. 写出到 Panel-replaced.xlsx，工作表名 Z2_xlsx
                with pd.ExcelWriter(Replaced_Panel_file_path, engine='openpyxl') as writer:
                    panel_df.to_excel(writer, sheet_name='Z2_xlsx', index=False)

                text.insert(tk.INSERT, 'EPLAN版SLD识别完成!  详见[柜型参数]表\n')

                text.insert(tk.INSERT, '\n>>>EPLAN属性正在读取中...\n')
                attribute_get(Panel_file_path, Panel_data_file_path, Panel_Size_file_path)
                text.insert(tk.INSERT, 'EPLAN属性读取完成!\n')

                text.insert(tk.INSERT, '\n>>>EPLAN属性与SLD的柜型参数正在对比中...\n')
                compare_sld_and_attribute(Panel_Size_file_path, switchgear_number, typical_type_list, abb_panel_number, switchgear_dimension)

                text.insert(tk.INSERT, 'EPLAN属性与SLD对比完成!\n')

                text.insert(tk.INSERT, '\n>>>BOM正在检查中...')
                book = load_workbook(FilePath)
                sheet = book['Z6_xlsx']
                A = []
                B = []
                C = []
                D = []
                E = []
                F = []
                G = []
                for i in range(2, sheet.max_row + 1):
                    A.append(str(sheet.cell(row=i, column=1).value))  # Hight-level列
                    B.append(str(sheet.cell(row=i, column=2).value))  # Zone列
                    C.append(str(sheet.cell(row=i, column=3).value))  # DT列
                    D.append(str(sheet.cell(row=i, column=4).value))  # PartNumber列
                    E.append(str(sheet.cell(row=i, column=5).value))  # Qty列
                    F.append(str(sheet.cell(row=i, column=6).value))  # Designation列
                    G.append(str(sheet.cell(row=i, column=7).value))  # Type列

                A_real = []
                B_real = []
                C_real = []
                D_real = []
                E_real = []
                F_real = []
                G_real = []

                for i in range(0, len(A)):
                    if B[i] != 'A' or C[i] != 'Panel':  # 剔除无效数据
                        A_real.append(A[i])
                        B_real.append(B[i])
                        C_real.append(C[i])
                        D_real.append(D[i])
                        E_real.append(E[i])
                        F_real.append(F[i])
                        G_real.append(G[i])

                # for i in range(0, len(A_real)):
                A_real_set = list(set(A_real))  # set去重后顺序会改变
                A_real_set.sort(key=list(A_real).index)  # 保证去重后顺序不变

                # 1、物料数量为0或空
                text.insert(tk.INSERT, '\n【1】EBOM中物料数量为0或空正在检查中...\n')
                Amount_zero_flag = 0
                for i in range(0, len(A_real)):
                    if E_real[i] == 'None':
                        Amount_zero_flag = 1
                        text.insert(tk.INSERT, '▲ %s    %s    数量为空\n' % (A_real[i], C_real[i]), 'error')
                        error1_calculator += 1
                    elif int(E_real[i]) == 0:
                        Amount_zero_flag = 1
                        text.insert(tk.INSERT, '▲ %s    %s    数量为0\n' % (A_real[i], C_real[i]), 'error')
                        error1_calculator += 1

                if not Amount_zero_flag:
                    text.insert(tk.INSERT, '无数量为0或空问题\n')

                # 2、物料定位缺失
                text.insert(tk.INSERT, '\n【2】EBOM中物料定位缺失正在检查中...\n')
                Position_empty_flag = 0
                for i in range(0, len(A_real)):
                    if B_real[i] == 'None':
                        Position_empty_flag = 1
                        text.insert(tk.INSERT, '▲ %s    %s    定位缺失\n' % (A_real[i], C_real[i]), 'error')
                        error2_calculator += 1
                if not Position_empty_flag:
                    text.insert(tk.INSERT, '无定位缺失问题\n')

                # 3、物料定位错误
                text.insert(tk.INSERT, '\n【3】EBOM中物料定位错误正在检查中...\n')
                Position_wrong_flag = 0
                for i in range(0, len(A_real)):
                    if B_real[i] != 'LV' and B_real[i] != 'MV' and B_real[i] != 'LVD' and B_real[i] != 'ITR' and B_real[i] != 'TR' \
                            and B_real[i] != 'None' and B_real[i] != 'LV.F' and B_real[i] != 'LV.M1' and B_real[i] != 'LV.M2' and B_real[i] != 'LV.R1' and B_real[i] != 'LV.R2' and B_real[i] != 'LV.L1' and B_real[i] != 'LV.L2':
                        Position_wrong_flag = 1
                        text.insert(tk.INSERT, '▲ %s    %s    %s    定位错误\n' % (A_real[i], C_real[i], B_real[i]), 'error')
                        error3_calculator += 1
                if not Position_wrong_flag:
                    text.insert(tk.INSERT, '无定位错误问题\n')

                # 4、同一物料多定位错误(同一ID、物料号，不同定位)
                text.insert(tk.INSERT, '\n【4】EBOM中同一物料多定位错误正在检查中...\n')
                Position_multi_flag = 0
                for i in range(0, len(A_real_set)):
                    dt_list = []
                    partnumber_list = []
                    zone_list = []

                    for j in range(0, len(A_real)):
                        if A_real_set[i] == A_real[j]:
                            zone_list.append(B_real[j])
                            dt_list.append(C_real[j])
                            partnumber_list.append(D_real[j])

                    for j in range(0, len(zone_list)):
                        for k in range(0, len(zone_list)):
                            if k > j and zone_list[j] != zone_list[k] and dt_list[j] == dt_list[k] and partnumber_list[j] == partnumber_list[k]:
                                text.insert(tk.INSERT, '▲ %s    %s包含不同定位%s和%s\n' % (A_real_set[i], dt_list[k], zone_list[k], zone_list[j]), 'error')
                                error4_calculator += 1
                                Position_multi_flag = 1
                if not Position_multi_flag:
                    text.insert(tk.INSERT, '无同一物料多定位问题\n')

                # 5、空开(ABB/人民电器)混用
                text.insert(tk.INSERT, '\n【5】EBOM中空开(ABB/人民电器)混用正在检查中...\n')
                FCM_A_number = 0
                FCM_R_number = 0

                for i in range(0, len(A_real)):
                    if D_real[i][0:5] == 'MCBR-':  # 人民电器空开
                        FCM_R_number += 1
                    if D_real[i][0:5] == 'MCBA-':  # ABB空开
                        FCM_A_number += 1
                if FCM_R_number != 0 and FCM_A_number != 0:
                    text.insert(tk.INSERT, '▲ 空开(ABB/人民电器)混用\n', 'error')
                    error5_calculator = min(FCM_R_number, FCM_A_number)
                else:
                    text.insert(tk.INSERT, '无空开(ABB/人民电器)混用问题\n')

                # 6、空开缺辅助触点或辅助触点缺空开
                text.insert(tk.INSERT, '\n【6】EBOM中空开缺辅助触点/辅助触点缺空开正在检查中...\n')

                Lack_MBD_flag = 0
                Lack_MCB_flag = 0
                for i in range(0, len(A_real_set)):
                    MCB_list = []
                    MBD_list = []
                    for j in range(0, len(A_real)):
                        if A_real_set[i] == A_real[j]:
                            if D_real[j][0:5] == 'MCBR-' or D_real[j][0:5] == 'MCBA-':
                                MCB_list.append(C_real[j])
                            if D_real[j][0:5] == 'MBDR-' or D_real[j][0:5] == 'MBDA-':
                                MBD_list.append(C_real[j])
                    diff_list1 = list(set(MCB_list) - set(MBD_list))
                    if len(diff_list1) > 0:
                        item_text1 = '；'.join(diff_list1)
                        text.insert(tk.INSERT, '▲ %s    空开缺辅助触点：%s\n' % (A_real_set[i], item_text1), 'error')
                        error6_calculator += 1
                        Lack_MBD_flag = 1

                    diff_list2 = list(set(MBD_list) - set(MCB_list))
                    if len(diff_list2) > 0:
                        item_text2 = '；'.join(diff_list2)
                        text.insert(tk.INSERT, '▲ %s    辅助触点缺空开：%s\n' % (A_real_set[i], item_text2), 'error')
                        error6_calculator += 1
                        Lack_MCB_flag = 1

                if not Lack_MBD_flag and not Lack_MCB_flag:
                    text.insert(tk.INSERT, '无空开缺辅助触点/辅助触点缺空开问题\n')

                # 7、除空开外，同一个设备标签名下有两个物料号检查
                text.insert(tk.INSERT, '\n【7】EBOM中多物料共用ID正在检查中...\n')
                Same_DT_flag = 0  # 有物料共用问题标志
                for i in range(0, len(A_real_set)):
                    DT_list = []
                    PartNumber_list = []
                    for j in range(0, len(A_real)):
                        if A_real_set[i] == A_real[j] and D_real[j][0:5] != 'MCBR-' and D_real[j][0:5] != 'MCBA-' and D_real[j][0:5] != 'MBDA-' and D_real[j][0:5] != 'MBDR-':
                            DT_list.append(C_real[j])
                            PartNumber_list.append(D_real[j])
                    index_DT = [a for a, b in enumerate(DT_list) if DT_list.count(b) > 1]  # n/n+1/n+2/n+3/n+4/n+5/n+6/n+7/n+8
                    content_DT = [b for a, b in enumerate(DT_list) if DT_list.count(b) > 1]  # AAABBCABC

                    if len(index_DT) > 1:
                        content_DT_set = list(set(content_DT))  # set去重后顺序会改变
                        content_DT_set.sort(key=list(content_DT).index)  # 保证去重后顺序不变,得到content_DT的集合(种类)

                        for k in range(0, len(content_DT_set)):  # 对于各种
                            index_one_type_DT = []  # 对于一种content_DT,用该变量记录下对应的index
                            for l in range(0, len(content_DT)):
                                if content_DT_set[k] == content_DT[l]:
                                    index_one_type_DT.append(index_DT[l])
                            if len(index_one_type_DT) == 2 and PartNumber_list[index_one_type_DT[0]] != PartNumber_list[index_one_type_DT[1]]:
                                Same_DT_flag = 1
                                text.insert(tk.INSERT, '▲ %s    %s对应物料号%s和%s\n' % (A_real_set[i], content_DT_set[k], PartNumber_list[index_one_type_DT[0]], PartNumber_list[index_one_type_DT[1]]), 'error')
                                error7_calculator += 1
                            if len(index_one_type_DT) == 3 and not (PartNumber_list[index_one_type_DT[0]] == PartNumber_list[index_one_type_DT[1]] and PartNumber_list[index_one_type_DT[1]] == PartNumber_list[index_one_type_DT[2]]):
                                Same_DT_flag = 1
                                text.insert(tk.INSERT, '▲ %s    %s对应物料号%s，%s，%s\n' % (A_real_set[i], content_DT_set[k], PartNumber_list[index_one_type_DT[0]], PartNumber_list[index_one_type_DT[1]], PartNumber_list[index_one_type_DT[2]]), 'error')
                                error7_calculator += 1
                            if len(index_one_type_DT) == 4 and not (PartNumber_list[index_one_type_DT[0]] == PartNumber_list[index_one_type_DT[1]] and PartNumber_list[index_one_type_DT[1]] == PartNumber_list[index_one_type_DT[2]] and PartNumber_list[index_one_type_DT[2]] == PartNumber_list[index_one_type_DT[3]]):
                                Same_DT_flag = 1
                                text.insert(tk.INSERT, '▲ %s    %s对应物料号%s，%s，%s，%s\n' % (A_real_set[i], content_DT_set[k], PartNumber_list[index_one_type_DT[0]], PartNumber_list[index_one_type_DT[1]], PartNumber_list[index_one_type_DT[2]], PartNumber_list[index_one_type_DT[3]]), 'error')
                                error7_calculator += 1

                for i in range(0, len(A_real_set)):
                    DT_list = []
                    PartNumber_list = []
                    for j in range(0, len(A_real)):
                        if A_real_set[i] == A_real[j] and (D_real[j][0:5] == 'MCBR-' or D_real[j][0:5] == 'MCBA-' or D_real[j][0:5] == 'MBDA-' or D_real[j][0:5] == 'MBDR-'):
                            DT_list.append(C_real[j])
                            PartNumber_list.append(D_real[j])

                    index_DT = [a for a, b in enumerate(DT_list) if DT_list.count(b) > 1]  # n/n+1/n+2/n+3/n+4/n+5/n+6/n+7/n+8
                    content_DT = [b for a, b in enumerate(DT_list) if DT_list.count(b) > 1]  # AAABBCABC

                    if len(index_DT) > 1:
                        content_DT_set = list(set(content_DT))  # set去重后顺序会改变
                        content_DT_set.sort(key=list(content_DT).index)  # 保证去重后顺序不变,得到content_DT的集合(种类)

                        for k in range(0, len(content_DT_set)):  # 对于各种
                            index_one_type_DT = []  # 对于一种content_DT,用该变量记录下对应的index
                            for l in range(0, len(content_DT)):
                                if content_DT_set[k] == content_DT[l]:
                                    index_one_type_DT.append(index_DT[l])
                            if len(index_one_type_DT) == 2:
                                prefix_0 = PartNumber_list[index_one_type_DT[0]][0:5]
                                prefix_1 = PartNumber_list[index_one_type_DT[1]][0:5]
                                if (prefix_0 == prefix_1) and (prefix_0 in ['MCBR-', 'MCBA-', 'MBDA-', 'MBDR-']):
                                    Same_DT_flag = 1
                                    text.insert(tk.INSERT, '▲ %s    %s对应物料号%s和%s(LVC布置图与原理图空开料号不一致)\n' % (A_real_set[i], content_DT_set[k], PartNumber_list[index_one_type_DT[0]], PartNumber_list[index_one_type_DT[1]]), 'error')
                                    error7_calculator += 1
                            if len(index_one_type_DT) == 3:
                                prefix_0 = PartNumber_list[index_one_type_DT[0]][0:5]
                                prefix_1 = PartNumber_list[index_one_type_DT[1]][0:5]
                                prefix_2 = PartNumber_list[index_one_type_DT[2]][0:5]
                                prefixes = [prefix_0, prefix_1, prefix_2]
                                if (prefixes.count('MCBR-') >= 2 or prefixes.count('MCBA-') >= 2 or prefixes.count('MBDA-') >= 2 or prefixes.count('MBDR-') >= 2):
                                    Same_DT_flag = 1
                                    text.insert(tk.INSERT, '▲ %s    %s对应物料号%s，%s，%s(LVC布置图与原理图空开料号不一致)\n' % (A_real_set[i], content_DT_set[k], PartNumber_list[index_one_type_DT[0]], PartNumber_list[index_one_type_DT[1]], PartNumber_list[index_one_type_DT[2]]), 'error')
                                    error7_calculator += 1

                if not Same_DT_flag:
                    text.insert(tk.INSERT, '无物料共用ID问题\n')

                # 8、空开电压等级统计
                text.insert(tk.INSERT, '\n【8】EBOM中空开电压等级正在统计中...\n')
                global MCB_typical_list
                MCB_typical_list = []
                MCB_DT_list = []
                MCB_level_list = []
                MCB_company_list = []
                MCB_aux_flag_list = []

                if not os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\MCB list.xlsx"):
                    text.insert(tk.INSERT, '▲ 失败，找不到空开数据表...\n', 'error')
                else:
                    workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\MCB list.xlsx")  # 加载空开数据表
                    worksheet = workbook['人民电器AC']
                    worksheet1 = workbook['人民电器DC']
                    worksheet2 = workbook['ABBAC']
                    worksheet3 = workbook['ABBDC']
                    worksheet4 = workbook['ABBACDC']

                    for i in range(0, len(A_real)):
                        if C_real[i][0:1] == 'F' and (D_real[i][0:5] == 'MCBR-' or D_real[i][0:5] == 'MCBA-'):
                            type_flag = 0
                            for j in range(2, worksheet.max_row + 1):  # 遍历空开数据表
                                if D_real[i] == worksheet.cell(row=j, column=1).value:
                                    type_flag = 1

                                    MCB_typical_list.append(A_real[i])
                                    MCB_DT_list.append(str(C_real[i]))
                                    MCB_level_list.append(worksheet.cell(row=j, column=3).value)
                                    MCB_company_list.append('人民电器')

                                    # text.insert(tk.INSERT, '%s    %s    %s    人民电器空开\n' % (A_real[i], C_real[i], worksheet.cell(row=j, column=3).value))
                                    continue
                            for j in range(2, worksheet1.max_row + 1):  # 遍历空开数据表
                                if D_real[i] == worksheet1.cell(row=j, column=1).value:
                                    type_flag = 1

                                    MCB_typical_list.append(A_real[i])
                                    MCB_DT_list.append(str(C_real[i]))
                                    MCB_level_list.append(worksheet1.cell(row=j, column=3).value)
                                    MCB_company_list.append('人民电器')

                                    # text.insert(tk.INSERT, '%s    %s    %s    人民电器空开\n' % (A_real[i], C_real[i], worksheet1.cell(row=j, column=3).value))
                                    continue
                            for j in range(2, worksheet2.max_row + 1):  # 遍历空开数据表
                                if D_real[i] == worksheet2.cell(row=j, column=1).value:
                                    type_flag = 1

                                    MCB_typical_list.append(A_real[i])
                                    MCB_DT_list.append(str(C_real[i]))
                                    MCB_level_list.append(worksheet2.cell(row=j, column=3).value)
                                    MCB_company_list.append('ABB')

                                    # text.insert(tk.INSERT, '%s    %s    %s    ABB空开\n' % (A_real[i], C_real[i], worksheet2.cell(row=j, column=3).value))
                                    continue
                            for j in range(2, worksheet3.max_row + 1):  # 遍历空开数据表
                                if D_real[i] == worksheet3.cell(row=j, column=1).value:
                                    type_flag = 1

                                    MCB_typical_list.append(A_real[i])
                                    MCB_DT_list.append(str(C_real[i]))
                                    MCB_level_list.append(worksheet3.cell(row=j, column=3).value)
                                    MCB_company_list.append('ABB')

                                    # text.insert(tk.INSERT, '%s    %s    %s    ABB空开\n' % (A_real[i], C_real[i], worksheet3.cell(row=j, column=3).value))
                                    continue
                            for j in range(2, worksheet4.max_row + 1):  # 遍历空开数据表
                                if D_real[i] == worksheet4.cell(row=j, column=1).value:
                                    type_flag = 1

                                    MCB_typical_list.append(A_real[i])
                                    MCB_DT_list.append(str(C_real[i]))
                                    MCB_level_list.append(worksheet4.cell(row=j, column=3).value)
                                    MCB_company_list.append('ABB')

                                    # text.insert(tk.INSERT, '%s    %s    %s    ABB空开\n' % (A_real[i], C_real[i], worksheet4.cell(row=j, column=3).value))
                                    continue
                            if not type_flag:
                                MCB_typical_list.append(A_real[i])
                                MCB_DT_list.append(str(C_real[i]))
                                MCB_level_list.append('无法识别' + D_real[i])
                                MCB_company_list.append('无法识别' + D_real[i])

                                # text.insert(tk.INSERT, '%s    %s   (物料号:%s)    不在空开数据表中,无法识别\n' % (A_real[i], C_real[i], D_real[i]), 'error')

                            if i < len(A_real) - 1:
                                if C_real[i + 1][0:1] == 'F' and (D_real[i + 1][0:5] == 'MBDR-' or D_real[i + 1][0:5] == 'MBDA-') and C_real[i] == C_real[i + 1]:
                                    MCB_aux_flag_list.append('有')
                                else:
                                    MCB_aux_flag_list.append('无')
                            if i == len(A_real) - 1:
                                MCB_aux_flag_list.append('无')

                    text.insert(tk.INSERT, '空开识别完成!  详见[空开统计]表\n')

                    if len(MCB_typical_list) > 0:
                        for i in range(0, len(MCB_typical_list)):
                            Calculate_table1.insert('', 'end', values=(MCB_typical_list[i], MCB_DT_list[i], MCB_level_list[i], MCB_company_list[i], MCB_aux_flag_list[i]))
                        table1_display_button['state'] = 'normal'
                # 9、低压室照明灯数量统计
                text.insert(tk.INSERT, '\n【9】EBOM中低压室照明灯正在统计中...\n')
                global LED_typical_list
                LED_typical_list = []
                LED_detail_list = []

                if not os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\LED lamp list.xlsx"):
                    text.insert(tk.INSERT, '▲ 失败，找不到照明灯数据表...\n', 'error')
                else:
                    workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\LED lamp list.xlsx")  # 加载照明灯数据表
                    worksheet = workbook['Sheet1']

                    for i in range(0, len(A_real_set)):
                        LED_number = 0
                        LED_DT = []

                        miss_LED_number = 0
                        miss_LED_DT = []
                        for j in range(0, len(A_real)):
                            if A_real_set[i] == A_real[j]:
                                if (C_real[j][0:2] == 'EA' or C_real[j][0:2] == 'EL') and B_real[j] in ['LV', 'LV.F', 'LV.M1', 'LV.M2', 'LV.R1', 'LV.R2', 'LV.L1', 'LV.L2']:
                                    EA_exist_flag = 0
                                    for k in range(2, worksheet.max_row + 1):  # 遍历照明灯数据表
                                        if D_real[j] == worksheet.cell(row=k, column=1).value:
                                            LED_number += int(E_real[j])
                                            LED_DT.append(C_real[j])
                                            EA_exist_flag = 1
                                            continue
                                    if not EA_exist_flag:
                                        miss_LED_number += int(E_real[j])
                                        miss_LED_DT.append(C_real[j])

                        if len(LED_DT) > 0 and len(miss_LED_DT) > 0:
                            LED_typical_list.append(A_real_set[i])
                            LED_detail_list.append('已识别:' + str(LED_number) + '个; ' + str(', '.join(LED_DT)) + ';' + '未识别:' + str(miss_LED_number) + '个; ' + str(', '.join(miss_LED_DT)) + ';')

                        if len(LED_DT) > 0 and len(miss_LED_DT) == 0:
                            LED_typical_list.append(A_real_set[i])
                            LED_detail_list.append('已识别:' + str(LED_number) + '个; ' + str(', '.join(LED_DT)) + ';')

                        if len(LED_DT) == 0 and len(miss_LED_DT) > 0:
                            LED_typical_list.append(A_real_set[i])
                            LED_detail_list.append('未识别:' + str(miss_LED_number) + '个; ' + str(', '.join(miss_LED_DT)) + ';')

                        if len(LED_DT) == 0 and len(miss_LED_DT) == 0:
                            LED_typical_list.append(A_real_set[i])
                            LED_detail_list.append('无照明灯' + ';')

                            # text.insert(tk.INSERT, '%s    无照明灯\n' % A_real_set[i])
                    text.insert(tk.INSERT, '低压室照明灯统计完成!  详见[其他统计]表\n')

                # 10、熔断器数量统计
                text.insert(tk.INSERT, '\n【10】EBOM中熔断器正在统计中...\n')

                FUSE_typical_list = []
                FUSE_detail_list = []

                if not os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\FUSE list.xlsx"):
                    text.insert(tk.INSERT, '▲ 失败，找不到熔断器数据表...\n', 'error')
                else:
                    workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\FUSE list.xlsx")  # 加载熔断器数据表
                    worksheet = workbook['熔丝物料号-FLS+FLE+FLZ']
                    worksheet1 = workbook['FLA']
                    worksheet2 = workbook['FLX']

                    max_sheet_rows = max(worksheet.max_row, worksheet1.max_row, worksheet2.max_row)
                    for i in range(0, len(A_real_set)):
                        FUSE_number = 0
                        FUSE_DT = []

                        miss_FUSE_number = 0
                        miss_FUSE_DT = []

                        for j in range(0, len(A_real)):
                            if A_real_set[i] == A_real[j]:
                                if C_real[j][0:3] == 'FCF':
                                    FCF_exist_flag = 0
                                    for k in range(2, max_sheet_rows + 1):  # 遍历熔断器数据表
                                        if D_real[j] == worksheet.cell(row=k, column=1).value:
                                            FUSE_number += int(E_real[j])
                                            FUSE_DT.append(C_real[j])
                                            FCF_exist_flag = 1
                                            continue
                                        if D_real[j] == worksheet1.cell(row=k, column=1).value:
                                            FUSE_number += int(E_real[j])
                                            FUSE_DT.append(C_real[j])
                                            FCF_exist_flag = 1
                                            continue
                                        if D_real[j] == worksheet2.cell(row=k, column=1).value:
                                            FUSE_number += int(E_real[j])
                                            FUSE_DT.append(C_real[j])
                                            FCF_exist_flag = 1
                                            continue
                                    if not FCF_exist_flag:
                                        miss_FUSE_number += int(E_real[j])
                                        miss_FUSE_DT.append(C_real[j])
                        if FUSE_number > 0 and miss_FUSE_number > 0:
                            FUSE_typical_list.append(A_real_set[i])
                            FUSE_detail_list.append('已识别:' + str(FUSE_number) + '支; ' + str(', '.join(FUSE_DT)) + ';' + '未识别:' + str(miss_FUSE_number) + '支; ' + str(', '.join(miss_FUSE_DT)) + ';')

                        if FUSE_number > 0 and miss_FUSE_number == 0:
                            FUSE_typical_list.append(A_real_set[i])
                            FUSE_detail_list.append('已识别:' + str(FUSE_number) + '支; ' + str(', '.join(FUSE_DT)) + ';')

                        if FUSE_number == 0 and miss_FUSE_number > 0:
                            FUSE_typical_list.append(A_real_set[i])
                            FUSE_detail_list.append('未识别:' + str(miss_FUSE_number) + '支; ' + str(', '.join(miss_FUSE_DT)) + ';')

                        if FUSE_number == 0 and miss_FUSE_number == 0:
                            FUSE_typical_list.append(A_real_set[i])
                            FUSE_detail_list.append('无熔断器' + ';')

                            # text.insert(tk.INSERT, '%s    无熔断器\n' % A_real_set[i])
                    text.insert(tk.INSERT, '熔断器统计完成!  详见[其他统计]表\n')

                # 11、避雷器数量统计
                text.insert(tk.INSERT, '\n【11】EBOM中避雷器正在统计中...\n')

                SA_typical_list = []
                SA_detail_list = []

                if not os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\Surge arrest list.xlsx"):
                    text.insert(tk.INSERT, '▲ 失败，找不到避雷器数据表...\n', 'error')
                else:
                    workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\Surge arrest list.xlsx")  # 加载避雷器数据表
                    worksheet = workbook['避雷器物料号SAS+SAG']
                    worksheet1 = workbook['SAY']
                    worksheet2 = workbook['SAA']

                    max_sheet_rows = max(worksheet.max_row, worksheet1.max_row, worksheet2.max_row)
                    for i in range(0, len(A_real_set)):
                        SA_number = 0
                        SA_DT = []

                        miss_SA_number = 0
                        miss_SA_DT = []
                        for j in range(0, len(A_real)):
                            if A_real_set[i] == A_real[j]:
                                if C_real[j][0:2] == 'FA':
                                    FA_exist_flag = 0
                                    for k in range(2, max_sheet_rows + 1):  # 遍历避雷器数据表
                                        if D_real[j] == worksheet.cell(row=k, column=1).value:
                                            SA_number += int(E_real[j])
                                            SA_DT.append(C_real[j])
                                            FA_exist_flag = 1
                                            continue
                                        if D_real[j] == worksheet1.cell(row=k, column=1).value:
                                            SA_number += int(E_real[j])
                                            SA_DT.append(C_real[j])
                                            FA_exist_flag = 1
                                            continue
                                        if D_real[j] == worksheet2.cell(row=k, column=1).value:
                                            SA_number += int(E_real[j])
                                            SA_DT.append(C_real[j])
                                            FA_exist_flag = 1
                                            continue
                                    if not FA_exist_flag:
                                        miss_SA_number += int(E_real[j])
                                        miss_SA_DT.append(C_real[j])
                        if SA_number > 0 and miss_SA_number > 0:
                            SA_typical_list.append(A_real_set[i])
                            SA_detail_list.append('已识别:' + str(SA_number) + '支; ' + str(', '.join(SA_DT)) + ';' + '未识别:' + str(miss_SA_number) + '支; ' + str(', '.join(miss_SA_DT)) + ';')

                        if SA_number > 0 and miss_SA_number == 0:
                            SA_typical_list.append(A_real_set[i])
                            SA_detail_list.append('已识别:' + str(SA_number) + '支; ' + str(', '.join(SA_DT)) + ';')

                        if SA_number == 0 and miss_SA_number > 0:
                            SA_typical_list.append(A_real_set[i])
                            SA_detail_list.append('未识别:' + str(miss_SA_number) + '支; ' + str(', '.join(miss_SA_DT)) + ';')

                        if SA_number == 0 and miss_SA_number == 0:
                            SA_typical_list.append(A_real_set[i])
                            SA_detail_list.append('无避雷器' + ';')

                            # text.insert(tk.INSERT, '%s    无避雷器\n' % A_real_set[i])
                    text.insert(tk.INSERT, '避雷器统计完成!  详见[其他统计]表\n')

                # 12、压板数量统计
                text.insert(tk.INSERT, '\n【12】EBOM中压板正在统计中...\n')

                LINK_typical_list = []
                LINK_detail_list = []

                if not os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\LINK list.xlsx"):
                    text.insert(tk.INSERT, '▲ 失败，找不到压板数据表...\n', 'error')
                else:
                    workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\LINK list.xlsx")  # 加载压板数据表
                    worksheet = workbook['Sheet1']

                    for i in range(0, len(A_real_set)):
                        LINK_number = 0
                        LINK_DT = []

                        miss_LINK_number = 0
                        miss_LINK_DT = []
                        for j in range(0, len(A_real)):
                            if A_real_set[i] == A_real[j]:
                                if C_real[j][0:3] == 'XLP':
                                    XLP_exist_flag = 0
                                    for k in range(2, worksheet.max_row + 1):  # 遍历压板数据表
                                        if D_real[j] == worksheet.cell(row=k, column=1).value:
                                            LINK_number += int(E_real[j])
                                            LINK_DT.append(C_real[j])
                                            XLP_exist_flag = 1
                                            continue
                                    if not XLP_exist_flag:
                                        miss_LINK_number += int(E_real[j])
                                        miss_LINK_DT.append(C_real[j])
                        if len(LINK_DT) > 0 and len(miss_LINK_DT) > 0:
                            LINK_typical_list.append(A_real_set[i])
                            LINK_detail_list.append('已识别:' + str(LINK_number) + '个; ' + str(', '.join(LINK_DT)) + ';' + '未识别:' + str(miss_LINK_number) + '个; ' + str(', '.join(miss_LINK_DT)) + ';')

                        if len(LINK_DT) > 0 and len(miss_LINK_DT) == 0:
                            LINK_typical_list.append(A_real_set[i])
                            LINK_detail_list.append('已识别:' + str(LINK_number) + '个; ' + str(', '.join(LINK_DT)) + ';')

                        if len(LINK_DT) == 0 and len(miss_LINK_DT) > 0:
                            LINK_typical_list.append(A_real_set[i])
                            LINK_detail_list.append('未识别:' + str(miss_LINK_number) + '个; ' + str(', '.join(miss_LINK_DT)) + ';')

                        if len(LINK_DT) == 0 and len(miss_LINK_DT) == 0:
                            LINK_typical_list.append(A_real_set[i])
                            LINK_detail_list.append('无压板' + ';')

                            # text.insert(tk.INSERT, '%s    无压板\n' % A_real_set[i])
                    text.insert(tk.INSERT, '压板统计完成!  详见[其他统计]表\n')

                # 13、转换开关数量统计
                text.insert(tk.INSERT, '\n【13】EBOM中转换开关正在统计中...\n')

                SFS_typical_list = []
                SFS_detail_list = []

                for i in range(0, len(A_real_set)):
                    SFS_APT_number = 0
                    SFS_JY_number = 0
                    SFS_KN_number = 0

                    SFS_APT_DT = []
                    SFS_JY_DT = []
                    SFS_KN_DT = []

                    miss_SFS_number = 0
                    miss_SFS_DT = []
                    for j in range(0, len(A_real)):
                        if A_real_set[i] == A_real[j]:
                            if C_real[j][0:3] == 'SFS' and D_real[j][0:3] == 'SSS':
                                SFS_APT_number += int(E_real[j])
                                SFS_APT_DT.append(C_real[j])
                            if C_real[j][0:3] == 'SFS' and D_real[j][0:3] == 'SSJ':
                                SFS_JY_number += int(E_real[j])
                                SFS_JY_DT.append(C_real[j])
                            if C_real[j][0:3] == 'SFS' and D_real[j][0:3] == 'SSK':
                                SFS_KN_number += int(E_real[j])
                                SFS_KN_DT.append(C_real[j])
                            if C_real[j][0:3] == 'SFS' and D_real[j][0:3] != 'SSS' and D_real[j][0:3] != 'SSJ' and D_real[j][0:3] != 'SSK':
                                miss_SFS_number += int(E_real[j])
                                miss_SFS_DT.append(C_real[j])

                    if SFS_APT_number != 0 and SFS_JY_number != 0 and SFS_KN_number != 0 and miss_SFS_number != 0:
                        SFS_typical_list.append(A_real_set[i])
                        SFS_detail_list.append(
                            '已识别:' + str(SFS_APT_number + SFS_JY_number + SFS_KN_number) + '个; 其中' + str(SFS_APT_number) + '个APT,' + str(', '.join(SFS_APT_DT)) + '; ' + str(SFS_JY_number) + '个江阴长江' + str(', '.join(SFS_JY_DT)) + '; ' + str(SFS_KN_number) + '个K&N' + str(', '.join(SFS_KN_DT)) + ';' + '未识别:' + str(miss_SFS_number) + '个,' + str(', '.join(miss_SFS_DT)) + ';')

                    if SFS_APT_number != 0 and SFS_JY_number != 0 and SFS_KN_number != 0 and miss_SFS_number == 0:
                        SFS_typical_list.append(A_real_set[i])
                        SFS_detail_list.append('已识别:' + str(SFS_APT_number + SFS_JY_number + SFS_KN_number) + '个; 其中' + str(SFS_APT_number) + '个APT,' + str(', '.join(SFS_APT_DT)) + '; ' + str(SFS_JY_number) + '个江阴长江' + str(', '.join(SFS_JY_DT)) + '; ' + str(SFS_KN_number) + '个K&N' + str(', '.join(SFS_KN_DT)) + ';')

                    if SFS_APT_number != 0 and SFS_JY_number != 0 and SFS_KN_number == 0 and miss_SFS_number != 0:
                        SFS_typical_list.append(A_real_set[i])
                        SFS_detail_list.append('已识别:' + str(SFS_APT_number + SFS_JY_number) + '个; 其中' + str(SFS_APT_number) + '个APT,' + str(', '.join(SFS_APT_DT)) + '; ' + str(SFS_JY_number) + '个江阴长江' + str(', '.join(SFS_JY_DT)) + ';' + '未识别:' + str(miss_SFS_number) + '个,' + str(', '.join(miss_SFS_DT)) + ';')

                    if SFS_APT_number != 0 and SFS_JY_number != 0 and SFS_KN_number == 0 and miss_SFS_number == 0:
                        SFS_typical_list.append(A_real_set[i])
                        SFS_detail_list.append('已识别:' + str(SFS_APT_number + SFS_JY_number) + '个; 其中' + str(SFS_APT_number) + '个APT,' + str(', '.join(SFS_APT_DT)) + '; ' + str(SFS_JY_number) + '个江阴长江' + str(', '.join(SFS_JY_DT)) + ';')

                    if SFS_APT_number != 0 and SFS_JY_number == 0 and SFS_KN_number != 0 and miss_SFS_number != 0:
                        SFS_typical_list.append(A_real_set[i])
                        SFS_detail_list.append('已识别:' + str(SFS_APT_number + SFS_KN_number) + '个; 其中' + str(SFS_APT_number) + '个APT,' + str(', '.join(SFS_APT_DT)) + '; ' + str(SFS_KN_number) + '个K&N' + str(', '.join(SFS_KN_DT)) + ';' + '未识别:' + str(miss_SFS_number) + '个; ' + str(', '.join(miss_SFS_DT)) + ';')

                    if SFS_APT_number != 0 and SFS_JY_number == 0 and SFS_KN_number != 0 and miss_SFS_number == 0:
                        SFS_typical_list.append(A_real_set[i])
                        SFS_detail_list.append('已识别:' + str(SFS_APT_number + SFS_KN_number) + '个; 其中' + str(SFS_APT_number) + '个APT,' + str(', '.join(SFS_APT_DT)) + '; ' + str(SFS_KN_number) + '个K&N' + str(', '.join(SFS_KN_DT)) + ';')

                    if SFS_APT_number == 0 and SFS_JY_number != 0 and SFS_KN_number != 0 and miss_SFS_number != 0:
                        SFS_typical_list.append(A_real_set[i])
                        SFS_detail_list.append('已识别:' + str(SFS_JY_number + SFS_KN_number) + '个; 其中' + str(SFS_JY_number) + '个江阴长江,' + str(', '.join(SFS_JY_DT)) + '; ' + str(SFS_KN_number) + '个K&N' + str(', '.join(SFS_KN_DT)) + ';' + '未识别:' + str(miss_SFS_number) + '个; ' + str(', '.join(miss_SFS_DT)) + ';')

                    if SFS_APT_number == 0 and SFS_JY_number != 0 and SFS_KN_number != 0 and miss_SFS_number == 0:
                        SFS_typical_list.append(A_real_set[i])
                        SFS_detail_list.append('已识别:' + str(SFS_JY_number + SFS_KN_number) + '个; 其中' + str(SFS_JY_number) + '个江阴长江,' + str(', '.join(SFS_JY_DT)) + '; ' + str(SFS_KN_number) + '个K&N' + str(', '.join(SFS_KN_DT)) + ';')

                    if SFS_APT_number == 0 and SFS_JY_number != 0 and SFS_KN_number == 0 and miss_SFS_number != 0:
                        SFS_typical_list.append(A_real_set[i])
                        SFS_detail_list.append('已识别:' + str(SFS_JY_number) + '个; ' + str(', '.join(SFS_JY_DT)) + ' 江阴长江' + ';' + '未识别:' + str(miss_SFS_number) + '个; ' + str(', '.join(miss_SFS_DT)) + ';')

                    if SFS_APT_number == 0 and SFS_JY_number != 0 and SFS_KN_number == 0 and miss_SFS_number == 0:
                        SFS_typical_list.append(A_real_set[i])
                        SFS_detail_list.append('已识别:' + str(SFS_JY_number) + '个; ' + str(', '.join(SFS_JY_DT)) + ' 江阴长江' + ';')

                    if SFS_APT_number == 0 and SFS_JY_number == 0 and SFS_KN_number != 0 and miss_SFS_number != 0:
                        SFS_typical_list.append(A_real_set[i])
                        SFS_detail_list.append('已识别:' + str(SFS_KN_number) + '个; ' + str(', '.join(SFS_KN_DT)) + ' K&N' + ';' + '未识别:' + str(miss_SFS_number) + '个; ' + str(', '.join(miss_SFS_DT)) + ';')

                    if SFS_APT_number == 0 and SFS_JY_number == 0 and SFS_KN_number != 0 and miss_SFS_number == 0:
                        SFS_typical_list.append(A_real_set[i])
                        SFS_detail_list.append('已识别:' + str(SFS_KN_number) + '个; ' + str(', '.join(SFS_KN_DT)) + ' K&N' + ';')

                    if SFS_APT_number != 0 and SFS_JY_number == 0 and SFS_KN_number == 0 and miss_SFS_number != 0:
                        SFS_typical_list.append(A_real_set[i])
                        SFS_detail_list.append('已识别:' + str(SFS_APT_number) + '个; ' + str(', '.join(SFS_APT_DT)) + ' APT' + ';' + '未识别:' + str(miss_SFS_number) + '个; ' + str(', '.join(miss_SFS_DT)) + ';')

                    if SFS_APT_number != 0 and SFS_JY_number == 0 and SFS_KN_number == 0 and miss_SFS_number == 0:
                        SFS_typical_list.append(A_real_set[i])
                        SFS_detail_list.append('已识别:' + str(SFS_APT_number) + '个; ' + str(', '.join(SFS_APT_DT)) + ' APT' + ';')

                    if SFS_APT_number == 0 and SFS_JY_number == 0 and SFS_KN_number == 0 and miss_SFS_number != 0:
                        SFS_typical_list.append(A_real_set[i])
                        SFS_detail_list.append('未识别:' + str(miss_SFS_number) + '个; ' + str(', '.join(miss_SFS_DT)) + ';')

                    if SFS_APT_number == 0 and SFS_JY_number == 0 and SFS_KN_number == 0 and miss_SFS_number == 0:
                        SFS_typical_list.append(A_real_set[i])
                        SFS_detail_list.append('无转换开关' + ';')

                text.insert(tk.INSERT, '转换开关统计完成!  详见[其他统计]表\n')

                # 14、容性分压装置数量统计
                text.insert(tk.INSERT, '\n【14】EBOM中容性分压装置正在统计中...\n')

                CAP_typical_list = []
                CAP_detail_list = []

                if not os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\CAP divider list.xlsx"):
                    text.insert(tk.INSERT, '▲ 失败，找不到容性分压装置数据表...\n', 'error')
                else:
                    workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\CAP divider list.xlsx")  # 加载容性分压装置数据表
                    worksheet = workbook['Sheet1']

                    for i in range(0, len(A_real_set)):
                        CAP_number = 0
                        CAP_DT = []

                        miss_CAP_number = 0
                        miss_CAP_DT = []
                        for j in range(0, len(A_real)):
                            if A_real_set[i] == A_real[j]:
                                if C_real[j][0:3] == 'BAD':
                                    CAP_exist_flag = 0
                                    for k in range(2, worksheet.max_row + 1):  # 遍历容性分压装置数据表
                                        if D_real[j] == worksheet.cell(row=k, column=1).value:
                                            CAP_number += int(E_real[j])
                                            CAP_DT.append(C_real[j])
                                            CAP_exist_flag = 1
                                            continue
                                    if not CAP_exist_flag:
                                        miss_CAP_number += int(E_real[j])
                                        miss_CAP_DT.append(C_real[j])
                        if CAP_number > 0 and miss_CAP_number > 0:
                            CAP_typical_list.append(A_real_set[i])
                            CAP_detail_list.append('已识别:' + str(CAP_number) + '个; ' + str(', '.join(CAP_DT)) + ';' + '未识别:' + str(miss_CAP_number) + '个; ' + str(', '.join(miss_CAP_DT)) + ';')

                        if CAP_number > 0 and miss_CAP_number == 0:
                            CAP_typical_list.append(A_real_set[i])
                            CAP_detail_list.append('已识别:' + str(CAP_number) + '个; ' + str(', '.join(CAP_DT)) + ';')

                        if CAP_number == 0 and miss_CAP_number > 0:
                            CAP_typical_list.append(A_real_set[i])
                            CAP_detail_list.append('未识别:' + str(miss_CAP_number) + '个; ' + str(', '.join(miss_CAP_DT)) + ';')

                        if CAP_number == 0 and miss_CAP_number == 0:
                            CAP_typical_list.append(A_real_set[i])
                            CAP_detail_list.append('无容性分压装置' + ';')

                            # text.insert(tk.INSERT, '%s    无容性分压装置\n' % A_real_set[i])
                    text.insert(tk.INSERT, '容性分压装置统计完成!  详见[其他统计]表\n')

                if len(LED_typical_list) > 0 and len(LED_detail_list) > 0 and len(FUSE_detail_list) > 0 and len(SA_detail_list) > 0 and len(LINK_detail_list) > 0 and len(SFS_detail_list) > 0 and len(CAP_detail_list) > 0:
                    for i in range(0, len(LED_typical_list)):
                        Calculate_table2.insert('', 'end', values=(LED_typical_list[i], LED_detail_list[i], FUSE_detail_list[i], SA_detail_list[i], LINK_detail_list[i], SFS_detail_list[i], CAP_detail_list[i]), tags='fontsize')  # dataframe逐行插入到表格中
                    table2_display_button['state'] = 'normal'

                # 15、AIS容性分压装置物料号检查
                text.insert(tk.INSERT, '\n【15】EBOM中AIS容性分压装置物料号正在检查中...\n')

                if not os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\CAP divider check list.xlsx"):
                    text.insert(tk.INSERT, '▲ 失败，找不到容性分压装置检查数据表...\n', 'error')
                else:
                    workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\CAP divider check list.xlsx")
                    worksheet = workbook['Sheet1']
                    error_CAP_divider_flag = False
                    spec_voltage_level = ''
                    for i in range(0, len(A_real_set)):
                        for j in range(0, len(A_real)):
                            if A_real_set[i] == A_real[j]:
                                if C_real[j][0:3] == 'BAD':
                                    spec_product_type = ''
                                    spec_panel_width = ''
                                    for k in range(0, len(dataframe_groupby['A'])):
                                        if A_real[j] == dataframe_groupby['B'][k]:
                                            spec_product_type = str(dataframe_groupby['E'][k])
                                            spec_panel_width = str(dataframe_groupby['D'][k])
                                            spec_voltage_level = str(dataframe_groupby_rev['F'][k])
                                            break
                                    for k in range(2, worksheet.max_row + 1):
                                        if str(worksheet.cell(row=k, column=1).value) in spec_voltage_level and spec_product_type == str(worksheet.cell(row=k, column=2).value):
                                            if D_real[j] == 'DIK-10-B10N1000C-1' and D_real[j] == str(worksheet.cell(row=k, column=3).value) and str(E_real[j]) != '3' and str(E_real[j]) != '6':
                                                text.insert(tk.INSERT, '▲ %s    容性分压装置物料号为%s，其数量应该为3的倍数\n' % (A_real_set[i], D_real[j]), 'error')
                                                error15_calculator += 1
                                                error_CAP_divider_flag = True
                                            if D_real[j] not in str(worksheet.cell(row=k, column=3).value) and '1400' not in spec_panel_width.split('X')[0]:
                                                text.insert(tk.INSERT, '▲ %s    容性分压装置物料号为%s，标准物料号为%s\n' % (A_real_set[i], D_real[j], worksheet.cell(row=k, column=3).value), 'error')
                                                error15_calculator += 1
                                                error_CAP_divider_flag = True
                                            elif D_real[j] in worksheet.cell(row=k, column=3).value and '1400' in spec_panel_width.split('X')[0]:
                                                text.insert(tk.INSERT, '▲ %s    容性分压装置物料号为%s，标准物料号为%s，且该物料应该在柜体BOM，而非EBOM中\n' % (A_real_set[i], D_real[j], '1YHT039153P0001'), 'error')
                                                error15_calculator += 1
                                                error_CAP_divider_flag = True
                                            break
                    if not error_CAP_divider_flag:
                        text.insert(tk.INSERT, '无AIS容性分压装置物料号问题\n')

                # 16、低压室照明灯物料号检查
                text.insert(tk.INSERT, '\n【16】EBOM中低压室照明灯物料号正在检查中...\n')
                error_LVC_LED_flag = False

                for i in range(0, len(A_real_set)):
                    for j in range(0, len(A_real)):
                        if A_real_set[i] == A_real[j]:
                            if (C_real[j][0:2] == 'EA' or C_real[j][0:2] == 'EL') and B_real[j] in ['LV', 'LV.F', 'LV.M1', 'LV.M2', 'LV.R1', 'LV.R2', 'LV.L1', 'LV.L2']:
                                spec_product_type = ''
                                for k in range(0, len(dataframe_groupby['A'])):
                                    if A_real[j] == dataframe_groupby['B'][k]:
                                        spec_product_type = str(dataframe_groupby['E'][k])
                                        break
                                if spec_product_type in ['ZVC', '500', '550', 'Beni', 'ZS1', 'ZS3.2'] and D_real[j] not in ['L-CL10-1C-W-AA', 'L-CL10-1C-W-AB', 'L-CL10-1C-W-AE']:
                                    text.insert(tk.INSERT, '▲ %s    AIS的低压室照明灯物料号应为%s\n' % (A_real_set[i], 'L-CL10-1C-W-AA，-AB，-AE'), 'error')
                                    error16_calculator += 1
                                    error_LVC_LED_flag = True
                                elif spec_product_type not in ['ZVC', '500', '550', 'Beni', 'ZS1', 'ZS3.2'] and D_real[j] in ['L-CL10-1C-W-AA', 'L-CL10-1C-W-AB', 'L-CL10-1C-W-AE']:
                                    text.insert(tk.INSERT, '▲ %s    GIS的低压室照明灯物料号不能为%s\n' % (A_real_set[i], 'L-CL10-1C-W-AA，-AB，-AE'), 'error')
                                    error16_calculator += 1
                                    error_LVC_LED_flag = True
                if not error_LVC_LED_flag:
                    text.insert(tk.INSERT, '无低压室照明灯物料号问题\n')

                book2 = load_workbook(Panel_BOM_file_path)
                sheet2 = book2['Z7_xlsx']
                A2 = []
                B2 = []
                C2 = []
                D2 = []
                E2 = []
                for i in range(2, sheet2.max_row + 1):
                    A2.append(str(sheet2.cell(row=i, column=1).value))  # Hight-level列
                    B2.append(str(sheet2.cell(row=i, column=2).value))  # Zone列
                    C2.append(str(sheet2.cell(row=i, column=3).value))  # DT列
                    D2.append(str(sheet2.cell(row=i, column=4).value))  # PartNumber列
                    E2.append(str(sheet2.cell(row=i, column=5).value))  # Qty列

                # 17、AIS柜体BOM微动开关配置检查1
                text.insert(tk.INSERT, '\n【17】AIS柜体BOM中微动开关配置1正在检查中...\n')
                error_microswitch_flag1 = False

                microswitch_name_list1 = ['压力释放板', '低压室照明', '航空插指示信号', '接地车航空插指示信号', '断路器手车运行位信号', '断路器手车试验位信号', '断路器手车试验位信号']
                microswitch_partnumber_list1 = ['EP-ITH', 'EP-LVMS', 'EP-S1NO-3NO2NC', 'EP-ESWTR-1NO4NC', 'EP-CBTR-BT4', 'EP-CBTR-BT5-55NO', 'EP-CBTR-BT5-ZS1']
                microswitch_qty_list1 = ['3', '1', '1', '1', '1', '1', '1']

                A2_real_set = list(set(A2))
                A2_real_set.sort(key=list(A2).index)

                # 遍历不重复的A2值
                for a2_value in A2_real_set:
                    for i in range(0, len(microswitch_partnumber_list1)):
                        microswitch_count = 0
                        for j in range(0, len(A2)):
                            if A2[j] == a2_value:
                                if D2[j] == microswitch_partnumber_list1[i]:
                                    microswitch_count += int(E2[j])
                        if microswitch_count > 0 and microswitch_count != int(microswitch_qty_list1[i]):
                            text.insert(tk.INSERT, '▲ %s    %s微动开关(物料号%s)的数量应为%s，实际数量为%s\n' % (a2_value, microswitch_name_list1[i], microswitch_partnumber_list1[i], microswitch_qty_list1[i], str(microswitch_count)), 'error')
                            error17_calculator += 1
                            error_microswitch_flag1 = True

                if not error_microswitch_flag1:
                    text.insert(tk.INSERT, '无AIS柜体BOM微动开关配置1问题\n')

                # 18、AIS柜体BOM微动开关配置检查2
                text.insert(tk.INSERT, '\n【18】AIS柜体BOM中微动开关配置2正在检查中...\n')
                error_microswitch_flag2 = False

                microswitch_name_list2 = ['电缆室前门关门信号', '电缆室后门关门信号(若为Beni电缆柜型)', '电缆室后门关门信号(若为Beni电缆柜型)', '电缆室后门关门信号(若为Beni电缆柜型)', '电缆室后门关门信号(若为Beni电缆柜型)',
                                          '电缆室前门关门信号(500/Beni)', '电缆室前门关门信号(500/Beni)', '电缆室后门关门信号(Beni，2节点)', '电缆室后门关门信号(Beni，2节点)', '电缆室后门关门信号(Beni，2节点)',
                                          '舌片信号', '断路器室关门信号', '断路器室关门信号(500/Beni)', '断路器手车位置信号', '断路器手车位置信号(500/Beni)', '断路器手车柜内外信号(500/Beni)', 'FailSafe', 'FailSafe',
                                          '活门信号', '活门信号(接近开关)', '电缆室门/断路器室门信号(接近开关)', '电缆室门/断路器室门信号(接近开关)']
                microswitch_partnumber_list2 = ['EP-CA-DOOR', 'EP-CA-DOOR', 'EP-CA-DOOR', 'EP-CA-DOOR', 'EP-CA-DOOR', 'EP-CA-DOOR-PG', 'EP-CA-DOOR-PG', 'EP-CA-DOOR-DOUBLE', 'EP-CA-DOOR-DOUBLE', 'EP-CA-DOOR-DOUBLE',
                                                'EP-ESWFLAP-5NC', 'EP-CB-DOOR', 'EP-CB-DOOR-PG', 'EP-CBTR-MC', 'EP-CBTR-MC-PG', 'EP-CB-TR-MB-PG', 'EP-FAILSAFE', 'EP-FAILSAFE', 'EP-SHUTTER', 'EP-SHUTTER-OE',
                                                'EP-DOOR-OE', 'EP-DOOR-OE']
                microswitch_qty_list2 = ['1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '2', '1', '4', '2', '1', '2']

                global MB585_flag
                global down_PT_truck
                global CB_eop
                global ES_exist
                global ES_eop

                for i in range(0, len(A_real_set)):
                    MB585_flag = False
                    down_PT_truck = False
                    CB_eop = False
                    ES_exist = False
                    ES_eop = False

                    space_flag = False

                    confirm_logic(A_real_set[i], A2, D2, A_real, D_real)

                    spec_product_type = ''
                    for j in range(0, len(dataframe_groupby['A'])):
                        if A_real_set[i] == dataframe_groupby['B'][j]:
                            spec_product_type = str(dataframe_groupby['E'][j])
                            break

                    # 进行微动检查
                    if spec_product_type in ['550', 'ZS1'] and ES_eop:
                        microswitch_exist0 = False
                        microswitch_0_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[0]:
                                    microswitch_exist0 = True
                                    microswitch_0_number += 1

                        if str(microswitch_0_number) == microswitch_qty_list2[0]:
                            microswitch_qty0 = True
                        else:
                            microswitch_qty0 = False

                        if not microswitch_exist0 and not microswitch_qty0:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[0], microswitch_partnumber_list2[0], microswitch_qty_list2[0]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist0 and not microswitch_qty0:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[0], microswitch_partnumber_list2[0], microswitch_qty_list2[0]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type == 'Beni' and ES_eop:
                        microswitch_exist1 = False
                        microswitch_1_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[1]:
                                    microswitch_exist1 = True
                                    microswitch_1_number += 1

                        if str(microswitch_1_number) == microswitch_qty_list2[1]:
                            microswitch_qty1 = True
                        else:
                            microswitch_qty1 = False

                        if not microswitch_exist1 and not microswitch_qty1:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list1[1], microswitch_partnumber_list1[1], microswitch_qty_list1[1]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist1 and not microswitch_qty1:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list1[1], microswitch_partnumber_list1[1], microswitch_qty_list1[1]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type == 'Beni' and not ES_exist:
                        microswitch_exist2 = False
                        microswitch_2_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[2]:
                                    microswitch_exist2 = True
                                    microswitch_2_number += 1

                        if str(microswitch_2_number) == microswitch_qty_list2[2]:
                            microswitch_qty2 = True
                        else:
                            microswitch_qty2 = False

                        if not microswitch_exist2 and not microswitch_qty2:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[2], microswitch_partnumber_list2[2], microswitch_qty_list2[2]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist2 and not microswitch_qty2:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[2], microswitch_partnumber_list2[2], microswitch_qty_list2[2]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type == 'Beni' and CB_eop:
                        microswitch_exist3 = False
                        microswitch_3_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[3]:
                                    microswitch_exist3 = True
                                    microswitch_3_number += 1

                        if str(microswitch_3_number) == microswitch_qty_list2[3]:
                            microswitch_qty3 = True
                        else:
                            microswitch_qty3 = False

                        if not microswitch_exist3 and not microswitch_qty3:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[3], microswitch_partnumber_list2[3], microswitch_qty_list2[3]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist3 and not microswitch_qty3:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[3], microswitch_partnumber_list2[3], microswitch_qty_list2[3]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type == 'Beni' and not CB_eop and not ES_eop and MB585_flag and ES_exist:
                        microswitch_exist4 = False
                        microswitch_4_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[4]:
                                    microswitch_exist4 = True
                                    microswitch_4_number += 1

                        if str(microswitch_4_number) == microswitch_qty_list2[4]:
                            microswitch_qty4 = True
                        else:
                            microswitch_qty4 = False

                        if not microswitch_exist4 and not microswitch_qty4:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[4], microswitch_partnumber_list2[4], microswitch_qty_list2[4]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist4 and not microswitch_qty4:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[4], microswitch_partnumber_list2[4], microswitch_qty_list2[4]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type in ['500', 'Beni'] and CB_eop and not ES_eop and down_PT_truck:
                        microswitch_exist5 = False
                        microswitch_5_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[5]:
                                    microswitch_exist5 = True
                                    microswitch_5_number += 1

                        if str(microswitch_5_number) == microswitch_qty_list2[5]:
                            microswitch_qty5 = True
                        else:
                            microswitch_qty5 = False

                        if not microswitch_exist5 and not microswitch_qty5:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[5], microswitch_partnumber_list2[5], microswitch_qty_list2[5]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist5 and not microswitch_qty5:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[5], microswitch_partnumber_list2[5], microswitch_qty_list2[5]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type in ['500', 'Beni'] and ES_eop and down_PT_truck:
                        microswitch_exist6 = False
                        microswitch_6_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[6]:
                                    microswitch_exist6 = True
                                    microswitch_6_number += 1

                        if str(microswitch_6_number) == microswitch_qty_list2[6]:
                            microswitch_qty6 = True
                        else:
                            microswitch_qty6 = False

                        if not microswitch_exist6 and not microswitch_qty6:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[6], microswitch_partnumber_list2[6], microswitch_qty_list2[6]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist6 and not microswitch_qty6:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[6], microswitch_partnumber_list2[6], microswitch_qty_list2[6]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type == 'Beni' and CB_eop and MB585_flag:
                        microswitch_exist7 = False
                        microswitch_7_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[7]:
                                    microswitch_exist7 = True
                                    microswitch_7_number += 1

                        if str(microswitch_7_number) == microswitch_qty_list2[7]:
                            microswitch_qty7 = True
                        else:
                            microswitch_qty7 = False

                        if not microswitch_exist7 and not microswitch_qty7:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[7], microswitch_partnumber_list2[7], microswitch_qty_list2[7]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist7 and not microswitch_qty7:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[7], microswitch_partnumber_list2[7], microswitch_qty_list2[7]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type == 'Beni' and ES_eop and MB585_flag:
                        microswitch_exist8 = False
                        microswitch_8_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[8]:
                                    microswitch_exist8 = True
                                    microswitch_8_number += 1

                        if str(microswitch_8_number) == microswitch_qty_list2[8]:
                            microswitch_qty8 = True
                        else:
                            microswitch_qty8 = False

                        if not microswitch_exist8 and not microswitch_qty8:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[8], microswitch_partnumber_list2[8], microswitch_qty_list2[8]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist8 and not microswitch_qty8:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[8], microswitch_partnumber_list2[8], microswitch_qty_list2[8]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type == 'Beni' and not ES_exist and MB585_flag:
                        microswitch_exist9 = False
                        microswitch_9_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[9]:
                                    microswitch_exist9 = True
                                    microswitch_9_number += 1

                        if str(microswitch_9_number) == microswitch_qty_list2[9]:
                            microswitch_qty9 = True
                        else:
                            microswitch_qty9 = False

                        if not microswitch_exist9 and not microswitch_qty9:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[9], microswitch_partnumber_list2[9], microswitch_qty_list2[9]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist9 and not microswitch_qty9:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[9], microswitch_partnumber_list2[9], microswitch_qty_list2[9]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type in ['Beni', 'ZS1'] and not ES_exist and CB_eop:
                        microswitch_exist10 = False
                        microswitch_10_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[10]:
                                    microswitch_exist10 = True
                                    microswitch_10_number += 1

                        if str(microswitch_10_number) == microswitch_qty_list2[10]:
                            microswitch_qty10 = True
                        else:
                            microswitch_qty10 = False

                        if not microswitch_exist10 and not microswitch_qty10:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[10], microswitch_partnumber_list2[10], microswitch_qty_list2[10]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist10 and not microswitch_qty10:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[10], microswitch_partnumber_list2[10], microswitch_qty_list2[10]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type in ['550', 'ZS1'] and CB_eop:
                        microswitch_exist11 = False
                        microswitch_11_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[11]:
                                    microswitch_exist11 = True
                                    microswitch_11_number += 1

                        if str(microswitch_11_number) == microswitch_qty_list2[11]:
                            microswitch_qty11 = True
                        else:
                            microswitch_qty11 = False

                        if not microswitch_exist11 and not microswitch_qty11:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[11], microswitch_partnumber_list2[11], microswitch_qty_list2[11]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist11 and not microswitch_qty11:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[11], microswitch_partnumber_list2[11], microswitch_qty_list2[11]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type in ['500', 'Beni'] and CB_eop:
                        microswitch_exist12 = False
                        microswitch_12_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[12]:
                                    microswitch_exist12 = True
                                    microswitch_12_number += 1

                        if str(microswitch_12_number) == microswitch_qty_list2[12]:
                            microswitch_qty12 = True
                        else:
                            microswitch_qty12 = False

                        if not microswitch_exist12 and not microswitch_qty12:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[12], microswitch_partnumber_list2[12], microswitch_qty_list2[12]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist12 and not microswitch_qty12:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[12], microswitch_partnumber_list2[12], microswitch_qty_list2[12]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type in ['550', 'ZS1'] and ES_eop:
                        microswitch_exist13 = False
                        microswitch_13_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[13]:
                                    microswitch_exist13 = True
                                    microswitch_13_number += 1

                        if str(microswitch_13_number) == microswitch_qty_list2[13]:
                            microswitch_qty13 = True
                        else:
                            microswitch_qty13 = False

                        if not microswitch_exist13 and not microswitch_qty13:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[13], microswitch_partnumber_list2[13], microswitch_qty_list2[13]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist13 and not microswitch_qty13:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[13], microswitch_partnumber_list2[13], microswitch_qty_list2[13]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type in ['500', 'Beni'] and ES_eop:
                        microswitch_exist14 = False
                        microswitch_14_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[14]:
                                    microswitch_exist14 = True
                                    microswitch_14_number += 1

                        if str(microswitch_14_number) == microswitch_qty_list2[14]:
                            microswitch_qty14 = True
                        else:
                            microswitch_qty14 = False

                        if not microswitch_exist14 and not microswitch_qty14:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[14], microswitch_partnumber_list2[14], microswitch_qty_list2[14]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist14 and not microswitch_qty14:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[14], microswitch_partnumber_list2[14], microswitch_qty_list2[14]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type in ['500', 'Beni'] and MB585_flag:
                        microswitch_exist15 = False
                        microswitch_15_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[15]:
                                    microswitch_exist15 = True
                                    microswitch_15_number += 1

                        if str(microswitch_15_number) == microswitch_qty_list2[15]:
                            microswitch_qty15 = True
                        else:
                            microswitch_qty15 = False

                        if not microswitch_exist15 and not microswitch_qty15:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[15], microswitch_partnumber_list2[15], microswitch_qty_list2[15]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist15 and not microswitch_qty15:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[15], microswitch_partnumber_list2[15], microswitch_qty_list2[15]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type == 'Beni' and MB585_flag:
                        microswitch_exist16 = False
                        microswitch_16_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[16]:
                                    microswitch_exist16 = True
                                    microswitch_16_number += 1

                        if str(microswitch_16_number) == microswitch_qty_list2[16]:
                            microswitch_qty16 = True
                        else:
                            microswitch_qty16 = False

                        if not microswitch_exist16 and not microswitch_qty16:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[16], microswitch_partnumber_list2[16], microswitch_qty_list2[16]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist16 and not microswitch_qty16:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[16], microswitch_partnumber_list2[16], microswitch_qty_list2[16]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type == '500' and MB585_flag:
                        microswitch_exist17 = False
                        microswitch_17_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[17]:
                                    microswitch_exist17 = True
                                    microswitch_17_number += 1

                        if str(microswitch_17_number) == microswitch_qty_list2[17]:
                            microswitch_qty17 = True
                        else:
                            microswitch_qty17 = False

                        if not microswitch_exist17 and not microswitch_qty17:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[17], microswitch_partnumber_list2[17], microswitch_qty_list2[17]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist17 and not microswitch_qty17:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[17], microswitch_partnumber_list2[17], microswitch_qty_list2[17]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type == 'Beni' and MB585_flag:
                        microswitch_exist18 = False
                        microswitch_18_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[18]:
                                    microswitch_exist18 = True
                                    microswitch_18_number += 1

                        if str(microswitch_18_number) == microswitch_qty_list2[18]:
                            microswitch_qty18 = True
                        else:
                            microswitch_qty18 = False

                        if not microswitch_exist18 and not microswitch_qty18:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[18], microswitch_partnumber_list2[18], microswitch_qty_list2[18]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist18 and not microswitch_qty18:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[18], microswitch_partnumber_list2[18], microswitch_qty_list2[18]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type == '500' and MB585_flag:
                        microswitch_exist19 = False
                        microswitch_19_number = 0
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[19]:
                                    microswitch_exist19 = True
                                    microswitch_19_number += 1

                        if str(microswitch_19_number) == microswitch_qty_list2[19]:
                            microswitch_qty19 = True
                        else:
                            microswitch_qty19 = False

                        if not microswitch_exist19 and not microswitch_qty19:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[19], microswitch_partnumber_list2[19], microswitch_qty_list2[19]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist19 and not microswitch_qty19:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[19], microswitch_partnumber_list2[19], microswitch_qty_list2[19]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type in ['500', 'Beni'] and MB585_flag and not down_PT_truck:
                        microswitch_exist20 = False
                        microswitch_20_number = 0

                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[20]:
                                    microswitch_exist20 = True
                                    microswitch_20_number += 1

                        if str(microswitch_20_number) == microswitch_qty_list2[20]:
                            microswitch_qty20 = True
                        else:
                            microswitch_qty20 = False

                        if not microswitch_exist20 and not microswitch_qty20:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[20], microswitch_partnumber_list2[20], microswitch_qty_list2[20]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist20 and not microswitch_qty20:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[20], microswitch_partnumber_list2[20], microswitch_qty_list2[20]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True

                    if spec_product_type in ['500', 'Beni'] and MB585_flag and down_PT_truck:
                        microswitch_exist21 = False
                        microswitch_21_number = 0
                        microswitch_qty21 = False
                        for k in range(0, len(A2)):
                            if A_real_set[i] == A2[k]:
                                if D2[k] == microswitch_partnumber_list2[21]:
                                    microswitch_exist21 = True
                                    microswitch_21_number += 1

                        if str(microswitch_21_number) == microswitch_qty_list2[21]:
                            microswitch_qty21 = True
                        else:
                            microswitch_qty21 = False

                        if not microswitch_exist21 and not microswitch_qty21:
                            text.insert(tk.INSERT, '▲ %s    %s，物料号应为%s，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[21], microswitch_partnumber_list2[21], microswitch_qty_list2[21]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                        elif microswitch_exist21 and not microswitch_qty21:
                            text.insert(tk.INSERT, '▲ %s    %s(物料号%s)，数量应为%s\n' % (A_real_set[i], microswitch_name_list2[21], microswitch_partnumber_list2[21], microswitch_qty_list2[21]), 'error')
                            error18_calculator += 1
                            error_microswitch_flag2 = True
                            space_flag = True
                    # if space_flag:
                    #     text.insert(tk.INSERT, '\n')

                if not error_microswitch_flag2:
                    text.insert(tk.INSERT, '无AIS柜体BOM微动开关配置2问题\n')

                # 19、EBOM中500柜CT型号150b/4检查
                text.insert(tk.INSERT, '\n【19】EBOM中500柜CT型号150b/4(比150b/2长120mm)正在检查中...\n')
                error_500_ct_flag = False

                for i in range(0, len(A_real_set)):
                    for j in range(0, len(A_real)):
                        if A_real_set[i] == A_real[j]:
                            if C_real[j][0:3] == 'BCT':
                                spec_product_type = ''
                                for k in range(0, len(dataframe_groupby['A'])):
                                    if A_real[j] == dataframe_groupby['B'][k]:
                                        spec_product_type = str(dataframe_groupby['E'][k])
                                        break
                                if spec_product_type == '500' and '150B/4' in G_real[j].upper():
                                    error_500_ct_flag = True
                                    text.insert(tk.INSERT, '▲ %s    500柜%s型号为%s，请复核\n' % (A_real_set[i], C_real[j], G_real[j]), 'error')
                                    error19_calculator += 1
                if not error_500_ct_flag:
                    text.insert(tk.INSERT, '无500柜CT型号150b/4问题\n')

                # 20、SAP柜体BOM中低压室门板宽度与SLD柜宽一致性检查
                text.insert(tk.INSERT, '\n【20】SAP柜体BOM柜宽与SLD柜宽一致性正在检查中...\n')
                error_sap_pbom_panel_width_flag = False

                # SAP登录
                sap_key_file = 'J:/Engineering/ShareFolder/new_ABB_Production_Tools/Ps/k/&%&^RD&^T&^WRE^@FEYUGYU@^E@DRYTRYTDFYTWF.txt'
                if not os.path.exists(sap_key_file):
                    tk.messagebox.showwarning('提示', 'SAP加密密钥文件不存在')
                else:
                    with open(sap_key_file, 'rb') as file:
                        sap_key = file.read()
                sap_cipher = Fernet(sap_key)

                sap_ciphertext_file = 'J:/Engineering/ShareFolder/new_ABB_Production_Tools/Ps/s/^&%&^R^YFT$%FTDTEYTKFY$^$FGKJGHE%#$%#EYD^%#^%DTYD.txt'
                if not os.path.exists(sap_ciphertext_file):
                    tk.messagebox.showwarning('提示', 'SAP加密密文文件不存在')
                else:
                    with open(sap_ciphertext_file, 'r') as file:
                        sap_ciphertext = file.read()

                sap_account = sap_decrypt_data(sap_ciphertext, sap_cipher, 'z_')
                sap_h = sap_account.split('+')[0]
                sap_c = sap_account.split('+')[1]
                sap_s = sap_account.split('+')[2]
                sap_u = sap_account.split('+')[3]
                sap_p = sap_account.split('+')[4]

                logging.info("Ready to connect to SAP")
                try:
                    conn = pyrfc.Connection(ashost=sap_h, sysnr=sap_s, client=sap_c, user=sap_u, passwd=sap_p, lang='EN')
                    if conn.alive:
                        logging.info("Connecting to SAP successfully")
                        project_number = stem.replace('-BOM', '').split('.')[0]
                        result = conn.call('ZY_SALES_ORDER_SHIFT', VBELN='0' + project_number)
                        # print(result)
                        if result['EX_CEPTION'] == '' and result['ITAB'][0]['WERKS'] == '1201':
                            item_data = []
                            for item in result['ITAB']:
                                posnr = item['POSNR'].lstrip('0') or '0'
                                arktx = item['ARKTX']
                                tptx1 = item['TPTX1']
                                tptx2 = item.get('TPTX2', '')
                                item_data.append((posnr, arktx, tptx1 + tptx2))

                            item_data_switchgear_id = []
                            item_data_width = []
                            item_data_panel_number = []
                            item_data_item_id = []

                            parent_nodes = {}
                            for item in item_data:
                                posnr = int(item[0])
                                if posnr % 1000 == 0:
                                    parent_nodes[posnr] = 1
                                else:
                                    parent_id = parent_nodes.get((int(posnr) // 1000) * 1000)
                                    if parent_id:
                                        # 获取 posnr 的千位数
                                        base = (posnr // 1000) * 1000
                                        # 计算 posnr 的偏移量
                                        offset = posnr - base

                                        if 1 <= offset <= 99:
                                            # print(offset, str(item[1]))
                                            match = re.findall(r'T\d+|pw\.\d+|p\.\d+|P\d+|P\.\d+|W\d+|p\d+', str(item[1]))

                                            if match:
                                                sap_panel_width = match[0].replace('T', '').replace('pw.', '').replace('p.', '').replace('P', '').replace('P.', '').replace('W', '').replace('p', '')
                                            else:
                                                sap_panel_width = '未知'  # 如果找不到匹配项，设置一个默认值

                                            for pi in (str(item[2]).replace('；', ';')).split(';')[0:-1]:
                                                code_number = base // 1000
                                                item_data_item_id.append(posnr)
                                                item_data_switchgear_id.append(f"A{code_number:02d}")
                                                item_data_width.append(sap_panel_width)
                                                item_data_panel_number.append(pi)

                        conn.close()
                        if not conn.alive:
                            logging.info("Disconnect from SAP")
                        for i in range(0, len(switchgear_number)):
                            for j in range(0, len(item_data_switchgear_id)):
                                if switchgear_number[i] == item_data_switchgear_id[j] and abb_panel_number_ori[i] == item_data_panel_number[j]:
                                    if switchgear_dimension[i].split('X')[0] != item_data_width[j]:
                                        error_sap_pbom_panel_width_flag = True
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜宽：%s(SAP柜体BOM%s)，%s(SLD)\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], item_data_width[j], item_data_item_id[j], switchgear_dimension[i].split('X')[0]), 'error')
                                        error20_calculator += 1
                                    break
                    if not error_sap_pbom_panel_width_flag:
                        text.insert(tk.INSERT, '无SAP柜体BOM柜宽与SLD柜宽不一致问题\n')

                except pyrfc.RFCError as e:
                    logging.info(e.key + ', ' + e.message)
                    tk.messagebox.showwarning("提示", traceback.format_exc())

                # 21、SLD中ZS1(500/Beni/ZS1)、ZX1.2/2、PrimeGear ZX0断路器相间距、电压等级检查
                text.insert(tk.INSERT, '\n【21】SLD中ZS1(500/Beni/ZS1)、ZX1.2/2、PrimeGear ZX0断路器相间距、电压正在检查中...\n')
                cb_phasedistance_error_flag = False
                cb_voltage_error_flag = False

                for i in range(0, len(cb_type)):
                    if ('VD4' in cb_type[i] or 'IT ' in cb_type[i]) and ' P' in cb_type[i]:

                        pattern1 = r'P\d{3}'  # 匹配 P 后面跟三个数字的模式
                        match1 = re.search(pattern1, cb_type[i])

                        pattern2 = r'(\d+)(?:\.\d+)?kV'  # 匹配 整数kV 或 整数.整数kV，并提取整数部分
                        match2 = re.search(pattern2, cb_type[i])

                        if product_type[i] == '500':
                            if len(cb_type[i].split(' ')) > 2:
                                if match1 and match1.group() != 'P135':
                                # if cb_type[i].split(' ')[2] != 'P135':
                                    text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], match1.group(), 'P135'), 'error')
                                    error21_calculator += 1
                                    cb_phasedistance_error_flag = True

                            if 'kV' in cb_type[i]:
                                if match2:
                                    matched_integer = match2.group(1)    # 获取匹配到的整数部分
                                    if matched_integer + 'kV' not in voltage_level[i].replace(' ', '') and matched_integer + '.5kV' not in voltage_level[i].replace(' ', '') and str(int(matched_integer)+2) + 'kV' not in voltage_level[i].replace(' ', '') and str(int(matched_integer)+2) + '.5kV' not in voltage_level[i].replace(' ', ''):
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器电压为%s，系统电压为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], matched_integer + 'kV', voltage_level[i].split('\n')[-1]), 'error')
                                        error21_calculator += 1
                                        cb_voltage_error_flag = True

                            elif cb_type[i].split(' ')[1][0:2] + 'kV' not in voltage_level[i].replace(' ', '') and cb_type[i].split(' ')[1][0:2] + '.5kV' not in voltage_level[i].replace(' ', '') and str(int(cb_type[i].split(' ')[1][0:2]) + 2) + 'kV' not in voltage_level[i].replace(' ', '') and str(int(cb_type[i].split(' ')[1][0:2]) + 2) + '.5kV' not in voltage_level[i].replace(' ', ''):
                                text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器电压为%s，系统电压为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(' ')[1][0:2] + 'kV', voltage_level[i].split('\n')[-1]), 'error')
                                error21_calculator += 1
                                cb_voltage_error_flag = True

                        elif product_type[i] in ['Beni', 'ZS1']:
                            if len(cb_type[i].split(' ')) > 2:
                                if switchgear_dimension[i].split('X')[0] == '650' and match1 and match1.group() != 'P150':
                                    text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], match1.group(), 'P150'), 'error')
                                    error21_calculator += 1
                                    cb_phasedistance_error_flag = True
                                elif switchgear_dimension[i].split('X')[0] == '800' and match1 and match1.group() != 'P210':
                                    text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], match1.group(), 'P210'), 'error')
                                    error21_calculator += 1
                                    cb_phasedistance_error_flag = True
                                elif switchgear_dimension[i].split('X')[0] == '1000' and match1 and match1.group() != 'P275':
                                    text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], match1.group(), 'P275'), 'error')
                                    error21_calculator += 1
                                    cb_phasedistance_error_flag = True

                            if 'kV' in cb_type[i]:
                                if match2:
                                    matched_integer = match2.group(1)  # 获取匹配到的整数部分
                                    if matched_integer + 'kV' not in voltage_level[i].replace(' ', '') and matched_integer + '.5kV' not in voltage_level[i].replace(' ', '') and str(int(matched_integer)+2) + 'kV' not in voltage_level[i].replace(' ', '') and str(int(matched_integer)+2) + '.5kV' not in voltage_level[i].replace(' ', ''):
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器电压为%s，系统电压为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], matched_integer + 'kV', voltage_level[i].split('\n')[-1]), 'error')
                                        error21_calculator += 1
                                        cb_voltage_error_flag = True
                            elif cb_type[i].split(' ')[1][0:2].isdigit():
                                if cb_type[i].split(' ')[1][0:2] + 'kV' not in voltage_level[i].replace(' ', '') and cb_type[i].split(' ')[1][0:2] + '.5kV' not in voltage_level[i].replace(' ', '') and str(int(cb_type[i].split(' ')[1][0:2]) + 2) + 'kV' not in voltage_level[i].replace(' ', '') and str(int(cb_type[i].split(' ')[1][0:2]) + 2) + '.5kV' not in voltage_level[i].replace(' ', ''):
                                    text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器电压为%s，系统电压为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(' ')[1][0:2] + 'kV', voltage_level[i].split('\n')[-1]), 'error')
                                    error21_calculator += 1
                                    cb_voltage_error_flag = True

                        elif product_type[i] in ['ZX1.2', 'ZX2']:
                            if len(cb_type[i].split(';')) > 1:
                                if len(cb_type[i].split(';')[-1].split(' ')) > 2:
                                    if switchgear_dimension[i].split('X')[0] == '600' and cb_type[i].split(';')[-1].split(' ')[2] != 'P150':
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(';')[-1].split(' ')[2], 'P150'), 'error')
                                        error21_calculator += 1
                                        cb_phasedistance_error_flag = True
                                    elif switchgear_dimension[i].split('X')[0] == '800' and cb_type[i].split(';')[-1].split(' ')[2] != 'P210':
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(';')[-1].split(' ')[2], 'P210'), 'error')
                                        error21_calculator += 1
                                        cb_phasedistance_error_flag = True
                                    elif switchgear_dimension[i].split('X')[0] == '840' and cb_type[i].split(';')[-1].split(' ')[2] != 'P210':
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(';')[-1].split(' ')[2], 'P210'), 'error')
                                        error21_calculator += 1
                                        cb_phasedistance_error_flag = True

                                if cb_type[i].split(';')[-1].split(' ')[1][0:2] + 'kV' not in voltage_level[i].replace(' ', '') and cb_type[i].split(';')[-1].split(' ')[1][0:2] + '.5kV' not in voltage_level[i].replace(' ', '') and str(int(cb_type[i].split(';')[-1].split(' ')[1][0:2]) + 2) + 'kV' not in voltage_level[i].replace(' ', '') and str(
                                        int(cb_type[i].split(';')[-1].split(' ')[1][0:2]) + 2) + '.5kV' not in voltage_level[i].replace(' ', ''):
                                    text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器电压为%s，系统电压为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(';')[-1].split(' ')[1][0:2] + 'kV', voltage_level[i].split('\n')[-1]), 'error')
                                    error21_calculator += 1
                                    cb_voltage_error_flag = True

                            else:
                                if len(cb_type[i].split(' ')) > 2:
                                    if switchgear_dimension[i].split('X')[0] == '600' and cb_type[i].split(' ')[2] != 'P150':
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(' ')[2], 'P150'), 'error')
                                        error21_calculator += 1
                                        cb_phasedistance_error_flag = True
                                    elif switchgear_dimension[i].split('X')[0] == '800' and cb_type[i].split(' ')[2] != 'P210':
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(' ')[2], 'P210'), 'error')
                                        error21_calculator += 1
                                        cb_phasedistance_error_flag = True
                                    elif switchgear_dimension[i].split('X')[0] == '840' and cb_type[i].split(' ')[2] != 'P210':
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(' ')[2], 'P210'), 'error')
                                        error21_calculator += 1
                                        cb_phasedistance_error_flag = True

                                if cb_type[i].split(' ')[1][0:2] + 'kV' not in voltage_level[i].replace(' ', '') and cb_type[i].split(' ')[1][0:2] + '.5kV' not in voltage_level[i].replace(' ', '') and str(int(cb_type[i].split(' ')[1][0:2]) + 2) + 'kV' not in voltage_level[i].replace(' ', '') and str(int(cb_type[i].split(' ')[1][0:2]) + 2) + '.5kV' not in voltage_level[i].replace(' ', ''):
                                    text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器电压为%s，系统电压为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(' ')[1][0:2] + 'kV', voltage_level[i].split('\n')[-1]), 'error')
                                    error21_calculator += 1
                                    cb_voltage_error_flag = True

                        elif product_type[i] == 'PrimeGear ZX0':
                            if len(cb_type[i].split(';')) > 1:
                                if len(cb_type[i].split(';')[-1].split(' ')) > 2:
                                    if switchgear_dimension[i].split('X')[0] == '450' and cb_type[i].split(';')[-1].split(' ')[2] != 'P125':
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(';')[-1].split(' ')[2], 'P125'), 'error')
                                        error21_calculator += 1
                                        cb_phasedistance_error_flag = True
                                    elif switchgear_dimension[i].split('X')[0] == '500' and cb_type[i].split(';')[-1].split(' ')[2] != 'P125':
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(';')[-1].split(' ')[2], 'P125'), 'error')
                                        error21_calculator += 1
                                        cb_phasedistance_error_flag = True
                                    elif switchgear_dimension[i].split('X')[0] == '600' and cb_type[i].split(';')[-1].split(' ')[2] != 'P150':
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(';')[-1].split(' ')[2], 'P150'), 'error')
                                        error21_calculator += 1
                                        cb_phasedistance_error_flag = True
                                    elif switchgear_dimension[i].split('X')[0] == '900' and cb_type[i].split(';')[-1].split(' ')[2] != 'P210':
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(';')[-1].split(' ')[2], 'P210'), 'error')
                                        error21_calculator += 1
                                        cb_phasedistance_error_flag = True
                                    elif switchgear_dimension[i].split('X')[0] == '1000' and cb_type[i].split(';')[-1].split(' ')[2] != 'P210':
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(';')[-1].split(' ')[2], 'P210'), 'error')
                                        error21_calculator += 1
                                        cb_phasedistance_error_flag = True

                                if cb_type[i].split(';')[-1].split(' ')[1][0:2] + 'kV' not in voltage_level[i].replace(' ', '') and cb_type[i].split(';')[-1].split(' ')[1][0:2] + '.5kV' not in voltage_level[i].replace(' ', '') and str(int(cb_type[i].split(';')[-1].split(' ')[1][0:2]) + 2) + 'kV' not in voltage_level[i].replace(' ', '') and str(
                                        int(cb_type[i].split(';')[-1].split(' ')[1][0:2]) + 2) + '.5kV' not in voltage_level[i].replace(' ', ''):
                                    text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器电压为%s，系统电压为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(';')[-1].split(' ')[1][0:2] + 'kV', voltage_level[i].split('\n')[-1]), 'error')
                                    error21_calculator += 1
                                    cb_voltage_error_flag = True

                            else:
                                if len(cb_type[i].split(' ')) > 2:
                                    if switchgear_dimension[i].split('X')[0] == '450' and cb_type[i].split(' ')[2] != 'P125':
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(' ')[2], 'P125'), 'error')
                                        error21_calculator += 1
                                        cb_phasedistance_error_flag = True
                                    elif switchgear_dimension[i].split('X')[0] == '500' and cb_type[i].split(' ')[2] != 'P125':
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(' ')[2], 'P125'), 'error')
                                        error21_calculator += 1
                                        cb_phasedistance_error_flag = True
                                    elif switchgear_dimension[i].split('X')[0] == '600' and cb_type[i].split(' ')[2] != 'P150':
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(' ')[2], 'P150'), 'error')
                                        error21_calculator += 1
                                        cb_phasedistance_error_flag = True
                                    elif switchgear_dimension[i].split('X')[0] == '900' and cb_type[i].split(' ')[2] != 'P210':
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(' ')[2], 'P210'), 'error')
                                        error21_calculator += 1
                                        cb_phasedistance_error_flag = True
                                    elif switchgear_dimension[i].split('X')[0] == '1000' and cb_type[i].split(' ')[2] != 'P210':
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器相间距为%s，应为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(' ')[2], 'P210'), 'error')
                                        error21_calculator += 1
                                        cb_phasedistance_error_flag = True

                                if cb_type[i].split(' ')[1][0:2] + 'kV' not in voltage_level[i].replace(' ', '') and cb_type[i].split(' ')[1][0:2] + '.5kV' not in voltage_level[i].replace(' ', '') and str(int(cb_type[i].split(' ')[1][0:2]) + 2) + 'kV' not in voltage_level[i].replace(' ', '') and str(int(cb_type[i].split(' ')[1][0:2]) + 2) + '.5kV' not in voltage_level[i].replace(' ', ''):
                                    text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中断路器电压为%s，系统电压为%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cb_type[i].split(' ')[1][0:2] + 'kV', voltage_level[i].split('\n')[-1]), 'error')
                                    error21_calculator += 1
                                    cb_voltage_error_flag = True

                if not cb_phasedistance_error_flag:
                    text.insert(tk.INSERT, '无SLD断路器相间距问题\n')
                if not cb_voltage_error_flag:
                    text.insert(tk.INSERT, '无SLD断路器电压等级问题\n')

                # 22、SLD中断路器描述与SAP断路器BOM描述一致性检查
                text.insert(tk.INSERT, '\n【22】SLD中断路器描述、柜号与SAP断路器BOM描述、柜号一致性正在检查中...\n')
                error_sld_sap_cbbom_description_flag = False
                error_sld_sap_cbbom_panel_number_flag = False

                logging.info("Ready to connect to SAP")
                try:
                    conn = pyrfc.Connection(ashost=sap_h, sysnr=sap_s, client=sap_c, user=sap_u, passwd=sap_p, lang='EN')
                    if conn.alive:
                        logging.info("Connecting to SAP successfully")
                        project_number = stem.replace('-BOM', '').split('.')[0]
                        result = conn.call('ZY_SALES_ORDER_SHIFT', VBELN='0' + project_number)
                        if result['EX_CEPTION'] == '' and result['ITAB'][0]['WERKS'] == '1201':
                            cbbom_item_data1 = []
                            for cbbom_item1 in result['ITAB']:
                                posnr_ori = cbbom_item1['POSNR']
                                posnr = cbbom_item1['POSNR'].lstrip('0') or '0'
                                matnr = cbbom_item1['MATNR']
                                arktx = cbbom_item1['ARKTX']
                                tptx1 = cbbom_item1['TPTX1']
                                tptx2 = cbbom_item1.get('TPTX2', '')
                                cbbom_item_data1.append((posnr_ori, posnr, matnr, arktx, tptx1 + tptx2))

                            cbbom_item_data_switchgear_id1 = []
                            cbbom_item_data_material1 = []
                            cbbom_item_data_description1 = []
                            cbbom_item_data_panel_number1 = []
                            cbbom_item_data_item_id1 = []

                            parent_nodes = {}
                            for cbbom_item in cbbom_item_data1:
                                posnr = int(cbbom_item[1])
                                if posnr % 1000 == 0:
                                    parent_nodes[posnr] = 1
                                else:
                                    parent_id = parent_nodes.get((int(posnr) // 1000) * 1000)
                                    if parent_id:
                                        # 获取 posnr 的千位数
                                        base = (posnr // 1000) * 1000
                                        # 计算 posnr 的偏移量
                                        offset = posnr - base

                                        if 501 <= offset <= 549 and ('CDX' in cbbom_item[2] or 'GCE' in cbbom_item[2] or 'HD4' in cbbom_item[2]):
                                            for pi in (str(cbbom_item[4]).replace('；', ';')).split(';')[0:-1]:
                                                code_number = base // 1000
                                                cbbom_item_data_switchgear_id1.append(f"A{code_number:02d}")
                                                cbbom_item_data_item_id1.append(posnr)
                                                cbbom_item_data_material1.append(cbbom_item[2])
                                                cbbom_item_data_description1.append(cbbom_item[3])
                                                cbbom_item_data_panel_number1.append(pi)

                    conn.close()
                    if not conn.alive:
                        logging.info("Disconnect from SAP")

                    for i in range(0, len(switchgear_number)):
                        for j in range(0, len(cbbom_item_data_switchgear_id1)):
                            if switchgear_number[i] == cbbom_item_data_switchgear_id1[j] and abb_panel_number_ori[i] == cbbom_item_data_panel_number1[j]:
                                # 处理 cb_type 中的描述
                                cb_description_parts = cb_type[i].split(';')
                                sld_description = cb_description_parts[-1] if len(cb_description_parts) > 1 else cb_type[i]
                                sld_description_parts = sld_description.split(' ')

                                # 标记用于检查所有部分是否都存在
                                all_parts_exist = True
                                for part in sld_description_parts:
                                    if part not in cbbom_item_data_description1[j]:
                                        all_parts_exist = False
                                        error_sld_sap_cbbom_description_flag = True

                                if not all_parts_exist:
                                    text.insert(tk.INSERT, '▲ %s    %s    %s柜：%s(SAP断路器BOM%s描述)，%s(SLD断路器描述)\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cbbom_item_data_description1[j], cbbom_item_data_item_id1[j], sld_description), 'error')
                                    error22_calculator += 1

                                break

                    if not error_sld_sap_cbbom_description_flag:
                        text.insert(tk.INSERT, '无SLD中断路器描述与SAP断路器BOM描述不一致问题\n')

                    for i in range(0, len(switchgear_number)):
                        if 'VD4' in cb_type[i] or 'HD4' in cb_type[i]:
                            sld_all_panel_exist = False
                            for j in range(0, len(cbbom_item_data_switchgear_id1)):
                                if switchgear_number[i] == cbbom_item_data_switchgear_id1[j] and abb_panel_number_ori[i] == cbbom_item_data_panel_number1[j]:
                                    sld_all_panel_exist = True
                                    break
                            if not sld_all_panel_exist:
                                error_sld_sap_cbbom_panel_number_flag = True
                                text.insert(tk.INSERT, '▲ %s    %s    %s柜：SLD中有断路器，但柜号没有维护在SAP断路器BOM中\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i]), 'error')
                                error22_calculator += 1

                    for i in range(0, len(cbbom_item_data_switchgear_id1)):
                        sap_all_panel_exist = False

                        for j in range(0, len(switchgear_number)):
                            if switchgear_number[j] == cbbom_item_data_switchgear_id1[i] and abb_panel_number_ori[j] == cbbom_item_data_panel_number1[i]:
                                sap_all_panel_exist = True
                                break
                        if not sap_all_panel_exist:
                            error_sld_sap_cbbom_panel_number_flag = True
                            text.insert(tk.INSERT, '▲ %s    %s柜：SLD中无断路器，但柜号却维护在SAP断路器BOM%s中\n' % (cbbom_item_data_switchgear_id1[i], cbbom_item_data_panel_number1[i], cbbom_item_data_item_id1[i]), 'error')
                            error22_calculator += 1

                    if not error_sld_sap_cbbom_panel_number_flag:
                        text.insert(tk.INSERT, '无SLD中断路器柜号与SAP断路器BOM柜号不一致问题\n')

                except pyrfc.RFCError as e:
                    logging.info(e.key + ', ' + e.message)
                    tk.messagebox.showwarning("提示", traceback.format_exc())

                # 23、ZS1(500/Beni/ZS1)中SAP断路器BOM选配与EPLAN所选Macro一致性检查
                text.insert(tk.INSERT, '\n【23】ZS1中SAP断路器BOM选配与EPLAN所选Macro一致性正在检查中...\n')
                error_sap_cbbom_flag = False

                logging.info("Ready to connect to SAP")
                try:
                    conn = pyrfc.Connection(ashost=sap_h, sysnr=sap_s, client=sap_c, user=sap_u, passwd=sap_p, lang='EN')
                    if conn.alive:
                        logging.info("Connecting to SAP successfully")
                        project_number = stem.replace('-BOM', '').split('.')[0]
                        result = conn.call('ZY_SALES_ORDER_SHIFT', VBELN='0' + project_number)
                        if result['EX_CEPTION'] == '' and result['ITAB'][0]['WERKS'] == '1201':
                            cbbom_item_data = []
                            for item in result['ITAB']:
                                posnr_ori = item['POSNR']
                                posnr = item['POSNR'].lstrip('0') or '0'
                                matnr = item['MATNR']
                                arktx = item['ARKTX']
                                tptx1 = item['TPTX1']
                                tptx2 = item.get('TPTX2', '')
                                cbbom_item_data.append((posnr_ori, posnr, matnr, arktx, tptx1 + tptx2))

                            cbbom_item_data_switchgear_id = []
                            cbbom_item_data_material = []
                            cbbom_item_data_description = []
                            cbbom_item_data_panel_number = []
                            cbbom_item_data_configuration = []
                            cbbom_item_data_item_id = []

                            parent_nodes = {}
                            for cbbom_item in cbbom_item_data:
                                posnr = int(cbbom_item[1])
                                if posnr % 1000 == 0:
                                    parent_nodes[posnr] = 1
                                else:
                                    parent_id = parent_nodes.get((int(posnr) // 1000) * 1000)
                                    if parent_id:
                                        # 获取 posnr 的千位数
                                        base = (posnr // 1000) * 1000
                                        # 计算 posnr 的偏移量
                                        offset = posnr - base

                                        if 501 <= offset <= 549 and ('CDX' in cbbom_item[2] or 'GCE' in cbbom_item[2] or 'HD4' in cbbom_item[2]):
                                            cb_configuration_result = conn.call('ZSD_GET_SO_CONFIGURATION', IM_VBELN='0' + project_number, IM_POSNR=cbbom_item[0])
                                            if cb_configuration_result['ET_CONF_OUT']:
                                                for pi in (str(cbbom_item[4]).replace('；', ';')).split(';')[0:-1]:
                                                    code_number = base // 1000
                                                    cbbom_item_data_item_id.append(posnr)
                                                    cbbom_item_data_switchgear_id.append(f"A{code_number:02d}")
                                                    cbbom_item_data_material.append(cbbom_item[2])
                                                    cbbom_item_data_description.append(cbbom_item[3])
                                                    cbbom_item_data_panel_number.append(pi)
                                                    cbbom_item_data_configuration.append(cb_configuration_result['ET_CONF_OUT'])

                    conn.close()
                    if not conn.alive:
                        logging.info("Disconnect from SAP")

                    Macro_cb = []
                    for i in range(0, len(switchgear_number)):
                        Macro_cb.append('')
                        for j in range(0, len(A2)):
                            if typical_type_list[i] == A2[j]:
                                if C2[j] == 'QAB' and 'VD4' in D2[j]:
                                    Macro_cb[i] = D2[j]

                    for i in range(0, len(switchgear_number)):
                        if Macro_cb[i] != '' and product_type[i] in ['500', 'Beni', 'ZS1']:
                            for s in Macro_cb[i].split('-')[1:]:
                                if s == '55' and 'A' not in Macro_cb[i].split('-')[1:]:
                                    error_aux_contact_flag = True
                                    sap_item_id = None
                                    for j in range(0, len(cbbom_item_data_switchgear_id)):
                                        if switchgear_number[i] == cbbom_item_data_switchgear_id[j] and abb_panel_number_ori[i] == cbbom_item_data_panel_number[j]:
                                            sap_item_id = j
                                            aux_contact = None
                                            for confi_item in cbbom_item_data_configuration[j]:
                                                Char_description = confi_item['ATBEZ']
                                                Char_value = confi_item['ATWTB']

                                                # if 'BB0' in Char_description and 'VDN_' not in Char_description:
                                                if ('AUXILIARY CONTACTS' in Char_description.upper() or 'BB0' in Char_description) and 'VDN_' not in Char_description:
                                                    aux_contact = Char_value
                                                    if '5NO-5NC' in Char_value or '5NO+5NC' in Char_value:
                                                        error_aux_contact_flag = False

                                    if error_aux_contact_flag and sap_item_id is not None:
                                        error_sap_cbbom_flag = True
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s选配中BB0应是5NO-5NC，实际是%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cbbom_item_data_item_id[sap_item_id], aux_contact), 'error')
                                        error23_calculator += 1

                                if (s == '55' and 'A' in Macro_cb[i].split('-')[1:]) or (s in ['46', '66', '67', '77']):
                                    error_aux_contact_flag = True
                                    sap_item_id = None
                                    for j in range(0, len(cbbom_item_data_switchgear_id)):
                                        if switchgear_number[i] == cbbom_item_data_switchgear_id[j] and abb_panel_number_ori[i] == cbbom_item_data_panel_number[j]:
                                            sap_item_id = j
                                            aux_contact = None
                                            for confi_item in cbbom_item_data_configuration[j]:
                                                Char_description = confi_item['ATBEZ']
                                                Char_value = confi_item['ATWTB']

                                                # if 'BB0' in Char_description and 'VDN_' not in Char_description:
                                                if ('AUXILIARY CONTACTS' in Char_description.upper() or 'BB0' in Char_description) and 'VDN_' not in Char_description:
                                                    aux_contact = Char_value
                                                    if '7NO-7NC' in Char_value or '7NO+7NC' in Char_value:
                                                        error_aux_contact_flag = False

                                    if error_aux_contact_flag and sap_item_id is not None:
                                        error_sap_cbbom_flag = True
                                        if s != '77':
                                            text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s选配中BB0应是7NO-7NC，实际是%s，修改后在Special requirement中备注实际的*NO*NC\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cbbom_item_data_item_id[sap_item_id], aux_contact), 'error')
                                            error23_calculator += 1
                                        if s == '77':
                                            text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s选配中BB0应是7NO-7NC，实际是%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cbbom_item_data_item_id[sap_item_id], aux_contact), 'error')
                                            error23_calculator += 1

                                if s in ['89', '99', '1010']:
                                    error_aux_contact_flag = True
                                    sap_item_id = ''
                                    for j in range(0, len(cbbom_item_data_switchgear_id)):
                                        if switchgear_number[i] == cbbom_item_data_switchgear_id[j] and abb_panel_number_ori[i] == cbbom_item_data_panel_number[j]:
                                            sap_item_id = j
                                            # for confi_item in cbbom_item_data_configuration[j]:
                                            # Char_description = confi_item['ATBEZ']
                                            # Char_value = confi_item['ATWTB']

                                            # aux_contact = ''
                                            # if 'BB0' in Char_description and '7NO-7NC' in Char_value:
                                            #     error_aux_contact_flag = False
                                            #     aux_contact = Char_value
                                    if error_aux_contact_flag and sap_item_id != '':
                                        error_sap_cbbom_flag = True
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s选配中BB0选7NO-7NC，然后在Special requirement中备注实际的*NO*NC，并通知一次将断路器BOM中旋转附开改为10NO-10NC\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cbbom_item_data_item_id[sap_item_id]), 'error')
                                        error23_calculator += 1

                            if 'RL' in Macro_cb[i].split('-')[1:]:
                                error_rl1_flag = True
                                sap_item_id = None
                                for k in range(0, len(cbbom_item_data_switchgear_id)):
                                    if switchgear_number[i] == cbbom_item_data_switchgear_id[k] and abb_panel_number_ori[i] == cbbom_item_data_panel_number[k]:
                                        sap_item_id = k
                                        voltage_rl1 = None
                                        for confi_item in cbbom_item_data_configuration[k]:
                                            Char_description = confi_item['ATBEZ']
                                            Char_value = confi_item['ATWTB']

                                            # if '-RL1' in Char_description and 'VDN_' not in Char_description:
                                            if (('LOCKING MAGNET' in Char_description.upper() and 'TRUCK' not in Char_description.upper()) or '-RL1' in Char_description) and 'VDN_' not in Char_description:
                                                voltage_rl1 = Char_value
                                                if 'DC' in control_source_voltage[i].split('\n')[-1]:
                                                    if voltage_rl1.replace(' ', '') == control_source_voltage[i].split('\n')[-1]:
                                                        error_rl1_flag = False
                                                if 'AC' in control_source_voltage[i].split('\n')[-1]:
                                                    if voltage_rl1.replace(' ', '') == control_source_voltage[i].split('\n')[-1].replace(' ', '').replace(',', ''):
                                                        error_rl1_flag = False
                                if error_rl1_flag and sap_item_id is not None:
                                    error_sap_cbbom_flag = True
                                    text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s选配中RL1应是%s，实际是%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cbbom_item_data_item_id[sap_item_id], control_source_voltage[i].split('\n')[-1], voltage_rl1), 'error')
                                    error23_calculator += 1

                            for s in Macro_cb[i].split('-')[1:]:
                                if s == 'R':
                                    error_rl2_flag = True
                                    sap_item_id = None
                                    for k in range(0, len(cbbom_item_data_switchgear_id)):
                                        if switchgear_number[i] == cbbom_item_data_switchgear_id[k] and abb_panel_number_ori[i] == cbbom_item_data_panel_number[k]:
                                            sap_item_id = k
                                            voltage_rl2 = None
                                            for confi_item in cbbom_item_data_configuration[k]:
                                                Char_description = confi_item['ATBEZ']
                                                Char_value = confi_item['ATWTB']

                                                # if '-RL2' in Char_description and 'VDN_' not in Char_description:
                                                if ('TRUCK LOCKING MAGNET' in Char_description.upper() or '-RL2' in Char_description) and 'VDN_' not in Char_description:
                                                    voltage_rl2 = Char_value
                                                    if 'DC' in control_source_voltage[i].split('\n')[-1]:
                                                        if voltage_rl2.replace(' ', '') == control_source_voltage[i].split('\n')[-1]:
                                                            error_rl2_flag = False
                                                    if 'AC' in control_source_voltage[i].split('\n')[-1]:
                                                        if voltage_rl2.replace(' ', '') == control_source_voltage[i].split('\n')[-1].replace(' ', '').replace(',', ''):
                                                            error_rl2_flag = False
                                    if error_rl2_flag and sap_item_id is not None:
                                        error_sap_cbbom_flag = True
                                        text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s选配中RL2应是%s，实际是%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cbbom_item_data_item_id[sap_item_id], control_source_voltage[i].split('\n')[-1], voltage_rl2), 'error')
                                        error23_calculator += 1

                            if 'O3' in Macro_cb[i].split('-')[1:]:
                                error_mo3_flag = True
                                sap_item_id = None
                                for k in range(0, len(cbbom_item_data_switchgear_id)):
                                    if switchgear_number[i] == cbbom_item_data_switchgear_id[k] and abb_panel_number_ori[i] == cbbom_item_data_panel_number[k]:
                                        sap_item_id = k
                                        for confi_item in cbbom_item_data_configuration[k]:
                                            Char_description = confi_item['ATBEZ']
                                            Char_value = confi_item['ATWTB']

                                            # if '-MO3' in Char_description and 'VDN_' not in Char_description and 'With MO3 (5A)' in Char_value:
                                            if ('OPENING SOLENOID' in Char_description.upper() or '-MO3' in Char_description) and 'VDN_' not in Char_description and 'With MO3 (5A)' in Char_value:
                                                error_mo3_flag = False

                                if error_mo3_flag and sap_item_id is not None:
                                    error_sap_cbbom_flag = True
                                    text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s选配中MO3应是%s，实际是%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cbbom_item_data_item_id[sap_item_id], 'With MO3 (5A)', None), 'error')
                                    error23_calculator += 1

                            for s in Macro_cb[i].split('-')[1:]:
                                if s == 'O':
                                    error_mo2_flag = True
                                    sap_item_id = None
                                    for k in range(0, len(cbbom_item_data_switchgear_id)):
                                        if switchgear_number[i] == cbbom_item_data_switchgear_id[k] and abb_panel_number_ori[i] == cbbom_item_data_panel_number[k]:
                                            sap_item_id = k
                                            voltage_mo2 = None
                                            mo2_flag = 0
                                            for confi_item in cbbom_item_data_configuration[k]:
                                                Char_description = confi_item['ATBEZ']
                                                Char_value = confi_item['ATWTB']

                                                # if '-MO2' in Char_description and 'VDN_' not in Char_description and '(PERM)' in Char_description:
                                                if ('RELEASE OFF' in Char_description.upper() or '-MO2' in Char_description) and 'O2' in Char_description and 'VDN_' not in Char_description and '(PERM)' in Char_description.upper():
                                                    mo2_flag = 1
                                                    voltage_mo2 = Char_value
                                                    if 'DC' in control_source_voltage[i].split('\n')[-1]:
                                                        if voltage_mo2.replace(' ', '') == control_source_voltage[i].split('\n')[-1]:
                                                            error_mo2_flag = False
                                                    if 'AC' in control_source_voltage[i].split('\n')[-1]:
                                                        if voltage_mo2.replace(' ', '') == control_source_voltage[i].split('\n')[-1].replace(' ', '').replace(',', ''):
                                                            error_mo2_flag = False

                                                # elif '-MO2' in Char_description and 'VDN_' not in Char_description and '(SMART)' in Char_description:
                                                elif ('RELEASE OFF' in Char_description.upper() or '-MO2' in Char_description) and 'O2' in Char_description and 'VDN_' not in Char_description and '(SMART)' in Char_description.upper():
                                                    mo2_flag = 2
                                                    voltage_mo2 = Char_value
                                                    if 'DC' in control_source_voltage[i].split('\n')[-1]:
                                                        if voltage_mo2.replace(' ', '') == control_source_voltage[i].split('\n')[-1]:
                                                            error_mo2_flag = False
                                                    if 'AC' in control_source_voltage[i].split('\n')[-1]:
                                                        if voltage_mo2.replace(' ', '') == control_source_voltage[i].split('\n')[-1].replace(' ', '').replace(',', ''):
                                                            error_mo2_flag = False

                                                # elif '-MO2' in Char_description and 'VDN_' not in Char_description and '(INST)' in Char_description:
                                                elif ('RELEASE OFF' in Char_description.upper()  or '-MO2' in Char_description) and 'O2' in Char_description and 'VDN_' not in Char_description and '(INST)' in Char_description.upper():
                                                    mo2_flag = 3
                                                    voltage_mo2 = Char_value
                                                    if 'DC' in control_source_voltage[i].split('\n')[-1]:
                                                        if voltage_mo2.replace(' ', '') == control_source_voltage[i].split('\n')[-1]:
                                                            error_mo2_flag = False
                                                    if 'AC' in control_source_voltage[i].split('\n')[-1]:
                                                        if voltage_mo2.replace(' ', '') == control_source_voltage[i].split('\n')[-1].replace(' ', '').replace(',', ''):
                                                            error_mo2_flag = False

                                    if error_mo2_flag and sap_item_id is not None:
                                        if mo2_flag == 0:
                                            error_sap_cbbom_flag = True
                                            text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s选配中MO2(PERM/SMART)应是%s，实际是%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cbbom_item_data_item_id[sap_item_id], control_source_voltage[i].split('\n')[-1], voltage_mo2), 'error')
                                            error23_calculator += 1
                                        elif mo2_flag == 1:
                                            error_sap_cbbom_flag = True
                                            text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s选配中MO2(PERM)应是%s，实际是%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cbbom_item_data_item_id[sap_item_id], control_source_voltage[i].split('\n')[-1], voltage_mo2), 'error')
                                            error23_calculator += 1
                                        elif mo2_flag == 2:
                                            error_sap_cbbom_flag = True
                                            text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s选配中MO2(SMART)应是%s，实际是%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cbbom_item_data_item_id[sap_item_id], control_source_voltage[i].split('\n')[-1], voltage_mo2), 'error')
                                            error23_calculator += 1
                                        elif mo2_flag == 3:
                                            error_sap_cbbom_flag = True
                                            text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s选配中MO2(INST)，请确认是否为INST，而不是PERM或SMART\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cbbom_item_data_item_id[sap_item_id]), 'error')
                                            error23_calculator += 1

                            if 'U' in Macro_cb[i].split('-')[1:]:
                                error_mu_flag = True
                                sap_item_id = None
                                for k in range(0, len(cbbom_item_data_switchgear_id)):
                                    if switchgear_number[i] == cbbom_item_data_switchgear_id[k] and abb_panel_number_ori[i] == cbbom_item_data_panel_number[k]:
                                        sap_item_id = k
                                        voltage_mu = None
                                        for confi_item in cbbom_item_data_configuration[k]:
                                            Char_description = confi_item['ATBEZ']
                                            Char_value = confi_item['ATWTB']

                                            # if '-MU' in Char_description and 'VDN_' not in Char_description:
                                            if ('UNDERVOL' in Char_description.upper() or '-MU' in Char_description) and 'VDN_' not in Char_description:
                                                voltage_mu = Char_value
                                                if 'DC' in control_source_voltage[i].split('\n')[-1]:
                                                    if voltage_mu.replace(' ', '') == control_source_voltage[i].split('\n')[-1]:
                                                        error_mu_flag = False
                                                if 'AC' in control_source_voltage[i].split('\n')[-1]:
                                                    if voltage_mu.replace(' ', '') == control_source_voltage[i].split('\n')[-1].replace(' ', '').replace(',', ''):
                                                        error_mu_flag = False
                                if error_mu_flag and sap_item_id is not None:
                                    error_sap_cbbom_flag = True
                                    text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s选配中MU应是%s，实际是%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cbbom_item_data_item_id[sap_item_id], control_source_voltage[i].split('\n')[-1], voltage_mu), 'error')
                                    error23_calculator += 1

                            if 'T' in Macro_cb[i].split('-')[1:]:
                                error_mt_flag = True
                                sap_item_id = None
                                for k in range(0, len(cbbom_item_data_switchgear_id)):
                                    if switchgear_number[i] == cbbom_item_data_switchgear_id[k] and abb_panel_number_ori[i] == cbbom_item_data_panel_number[k]:
                                        sap_item_id = k
                                        voltage_mt = None
                                        for confi_item in cbbom_item_data_configuration[k]:
                                            Char_description = confi_item['ATBEZ']
                                            Char_value = confi_item['ATWTB']

                                            if ('MOTORISED TRUCK' in Char_description.upper() or '-MT' in Char_description) and 'VDN_' not in Char_description:
                                                voltage_mt = Char_value
                                                if 'DC' in control_source_voltage[i].split('\n')[-1]:
                                                    if voltage_mt.replace(' ', '') == control_source_voltage[i].split('\n')[-1]:
                                                        error_mt_flag = False
                                                if 'AC' in control_source_voltage[i].split('\n')[-1]:
                                                    if voltage_mt.replace(' ', '') == control_source_voltage[i].split('\n')[-1].replace(' ', '').replace(',', ''):
                                                        error_mt_flag = False
                                if error_mt_flag and sap_item_id is not None:
                                    error_sap_cbbom_flag = True
                                    text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s选配中MT应是%s，实际是%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cbbom_item_data_item_id[sap_item_id], control_source_voltage[i].split('\n')[-1], voltage_mt), 'error')
                                    error23_calculator += 1

                            error_mc_flag = True
                            sap_item_id = None
                            for k in range(0, len(cbbom_item_data_switchgear_id)):
                                if switchgear_number[i] == cbbom_item_data_switchgear_id[k] and abb_panel_number_ori[i] == cbbom_item_data_panel_number[k]:
                                    sap_item_id = k
                                    voltage_mc = None
                                    for confi_item in cbbom_item_data_configuration[k]:
                                        Char_description = confi_item['ATBEZ']
                                        Char_value = confi_item['ATWTB']

                                        # if '-MC' in Char_description and 'VDN_' not in Char_description:
                                        if ('RELEASE ON' in Char_description.upper() or '-MC' in Char_description) and 'VDN_' not in Char_description:
                                            voltage_mc = Char_value
                                            if 'DC' in control_source_voltage[i].split('\n')[-1]:
                                                if voltage_mc.replace(' ', '') == control_source_voltage[i].split('\n')[-1]:
                                                    error_mc_flag = False
                                            if 'AC' in control_source_voltage[i].split('\n')[-1]:
                                                if voltage_mc.replace(' ', '') == control_source_voltage[i].split('\n')[-1].replace(' ', '').replace(',', ''):
                                                    error_mc_flag = False
                            if error_mc_flag and sap_item_id is not None:
                                error_sap_cbbom_flag = True
                                text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s选配中MC应是%s，实际是%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cbbom_item_data_item_id[sap_item_id], control_source_voltage[i].split('\n')[-1], voltage_mc), 'error')
                                error23_calculator += 1

                            error_mo_flag = True
                            sap_item_id = None
                            for k in range(0, len(cbbom_item_data_switchgear_id)):
                                if switchgear_number[i] == cbbom_item_data_switchgear_id[k] and abb_panel_number_ori[i] == cbbom_item_data_panel_number[k]:
                                    sap_item_id = k
                                    voltage_mo = None
                                    for confi_item in cbbom_item_data_configuration[k]:
                                        Char_description = confi_item['ATBEZ']
                                        Char_value = confi_item['ATWTB']

                                        # if '-MO' in Char_description and 'VDN_' not in Char_description:
                                        if 'O1' in Char_description and ('RELEASE OFF' in Char_description.upper() or '-MO1' in Char_description) and 'VDN_' not in Char_description:
                                            voltage_mo = Char_value
                                            if 'DC' in control_source_voltage[i].split('\n')[-1]:
                                                if voltage_mo.replace(' ', '') == control_source_voltage[i].split('\n')[-1]:
                                                    error_mo_flag = False
                                            if 'AC' in control_source_voltage[i].split('\n')[-1]:
                                                if voltage_mo.replace(' ', '') == control_source_voltage[i].split('\n')[-1].replace(' ', '').replace(',', ''):
                                                    error_mo_flag = False
                            if error_mo_flag and sap_item_id is not None:
                                error_sap_cbbom_flag = True
                                text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s选配中MO1应是%s，实际是%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cbbom_item_data_item_id[sap_item_id], control_source_voltage[i].split('\n')[-1], voltage_mo), 'error')
                                error23_calculator += 1

                            error_ms_flag = True
                            sap_item_id = None
                            for k in range(0, len(cbbom_item_data_switchgear_id)):
                                if switchgear_number[i] == cbbom_item_data_switchgear_id[k] and abb_panel_number_ori[i] == cbbom_item_data_panel_number[k]:
                                    sap_item_id = k
                                    voltage_ms = None
                                    for confi_item in cbbom_item_data_configuration[k]:
                                        Char_description = confi_item['ATBEZ']
                                        Char_value = confi_item['ATWTB']

                                        # if '-MS' in Char_description and 'VDN_' not in Char_description:
                                        if ('CHARGING MOTOR' in Char_description.upper() or '-MS' in Char_description) and 'VDN_' not in Char_description:
                                            voltage_ms = Char_value
                                            if 'DC' in charge_source_voltage[i].split('\n')[-1]:
                                                if voltage_ms.replace(' ', '') == charge_source_voltage[i].split('\n')[-1]:
                                                    error_ms_flag = False
                                            if 'AC' in charge_source_voltage[i].split('\n')[-1]:
                                                if voltage_ms.replace(' ', '') == charge_source_voltage[i].split('\n')[-1].replace(' ', '').replace(',', ''):
                                                    error_ms_flag = False
                            if error_ms_flag and sap_item_id is not None:
                                error_sap_cbbom_flag = True
                                text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s选配中MS应是%s，实际是%s\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cbbom_item_data_item_id[sap_item_id], charge_source_voltage[i].split('\n')[-1], voltage_ms), 'error')
                                error23_calculator += 1

                            if 'A' in Macro_cb[i].split('-')[1:]:
                                error_ma_flag = True
                                sap_item_id = None
                                for k in range(0, len(cbbom_item_data_switchgear_id)):
                                    if switchgear_number[i] == cbbom_item_data_switchgear_id[k] and abb_panel_number_ori[i] == cbbom_item_data_panel_number[k]:
                                        sap_item_id = k
                                if error_ma_flag and sap_item_id is not None:
                                    error_sap_cbbom_flag = True
                                    text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s选配中提醒一次进行角度传感器物料替换\n' % (switchgear_number[i], typical_type_list[i], abb_panel_number_ori[i], cbbom_item_data_item_id[sap_item_id]), 'error')
                                    error23_calculator += 1

                    for i in range(0, len(cbbom_item_data_switchgear_id)):
                        for j in range(0, len(switchgear_number)):
                            if cbbom_item_data_switchgear_id[i] == switchgear_number[j] and cbbom_item_data_panel_number[i] == abb_panel_number_ori[j]:
                                if Macro_cb[j] != '' and product_type[j] in ['500', 'Beni', 'ZS1']:
                                    for confi_item in cbbom_item_data_configuration[i]:
                                        if (('LOCKING MAGNET' in confi_item['ATBEZ'].upper() and 'TRUCK' not in confi_item['ATBEZ'].upper()) or '-RL1' in confi_item['ATBEZ']) and 'RL' not in Macro_cb[j]:
                                            error_sap_cbbom_flag = True
                                            text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s多选配了RL1\n' % (cbbom_item_data_switchgear_id[i], typical_type_list[j], cbbom_item_data_panel_number[i], cbbom_item_data_item_id[i]), 'error')
                                            error23_calculator += 1
                                        if ('TRUCK LOCKING MAGNET' in confi_item['ATBEZ'].upper() or '-RL2' in confi_item['ATBEZ']) and all(element != 'R' for element in Macro_cb[j].split('-')[1:]):
                                            error_sap_cbbom_flag = True
                                            text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s多选配了RL2\n' % (cbbom_item_data_switchgear_id[i], typical_type_list[j], cbbom_item_data_panel_number[i], cbbom_item_data_item_id[i]), 'error')
                                            error23_calculator += 1
                                        if ('RELEASE OFF' in confi_item['ATBEZ'].upper() or '-MO2' in confi_item['ATBEZ']) and 'O2' in confi_item['ATBEZ'] and all(element != 'O' for element in Macro_cb[j].split('-')[1:]):
                                            error_sap_cbbom_flag = True
                                            text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s多选配了MO2\n' % (cbbom_item_data_switchgear_id[i], typical_type_list[j], cbbom_item_data_panel_number[i], cbbom_item_data_item_id[i]), 'error')
                                            error23_calculator += 1
                                        if ('OPENING SOLENOID' in confi_item['ATBEZ'].upper() or '-MO3' in confi_item['ATBEZ']) and 'O3' not in Macro_cb[j]:
                                            error_sap_cbbom_flag = True
                                            text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s多选配了MO3\n' % (cbbom_item_data_switchgear_id[i], typical_type_list[j], cbbom_item_data_panel_number[i], cbbom_item_data_item_id[i]), 'error')
                                            error23_calculator += 1
                                        if ('UNDERVOL' in confi_item['ATBEZ'].upper() or '-MU' in confi_item['ATBEZ']) and 'U' not in Macro_cb[j]:
                                            error_sap_cbbom_flag = True
                                            text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s多选配了MU\n' % (cbbom_item_data_switchgear_id[i], typical_type_list[j], cbbom_item_data_panel_number[i], cbbom_item_data_item_id[i]), 'error')
                                            error23_calculator += 1
                                        if ('MOTORISED TRUCK' in confi_item['ATBEZ'].upper() or '-MT' in confi_item['ATBEZ']) and 'T' not in Macro_cb[j]:
                                            error_sap_cbbom_flag = True
                                            text.insert(tk.INSERT, '▲ %s    %s    %s柜：断路器BOM%s多选配了MT\n' % (cbbom_item_data_switchgear_id[i], typical_type_list[j], cbbom_item_data_panel_number[i], cbbom_item_data_item_id[i]), 'error')
                                            error23_calculator += 1

                    if not error_sap_cbbom_flag:
                        text.insert(tk.INSERT, '无ZS1中SAP断路器BOM选配与EPLAN所选Macro不一致问题\n')

                except pyrfc.RFCError as e:
                    logging.info(e.key + ', ' + e.message)
                    tk.messagebox.showwarning("提示", traceback.format_exc())

                # 24、EBOM中空开辅助触点数量校核
                text.insert(tk.INSERT, '\n【24】EBOM中空开辅助触点数量正在校核中...\n')
                error_MCB_aux_number_flag = False

                MBD_ebom_Typical = []
                MBD_ebom_DT = []
                MBD_ebom_Partnumber = []
                MBD_ebom_Qty = []

                for i in range(0, len(A_real_set)):
                    for j in range(0, len(A_real)):
                        if A_real_set[i] == A_real[j]:
                            if D_real[j][0:5] == 'MBDR-' or D_real[j][0:5] == 'MBDA-':
                                MBD_ebom_Typical.append(A_real[j])
                                MBD_ebom_DT.append(C_real[j])
                                MBD_ebom_Partnumber.append(D_real[j])
                                MBD_ebom_Qty.append(E_real[j])

                devicelabel_book = load_workbook(Device_Label_file_path)
                devicelabel_sheet = devicelabel_book['Z5_xlsx']
                devicelabel_A = []
                devicelabel_B = []
                devicelabel_D = []

                for i in range(3, devicelabel_sheet.max_row + 1):
                    devicelabel_A.append(str(devicelabel_sheet.cell(row=i, column=1).value))
                    devicelabel_B.append(str(devicelabel_sheet.cell(row=i, column=2).value))
                    devicelabel_D.append(str(devicelabel_sheet.cell(row=i, column=4).value))

                # 创建一个字典来存储合并后的数据
                merged_data = {}

                # 遍历原始列表并累加 MBD_ebom_Qty
                for i in range(len(MBD_ebom_Typical)):
                    key = (MBD_ebom_Typical[i], MBD_ebom_DT[i], MBD_ebom_Partnumber[i])
                    if key in merged_data:
                        merged_data[key] += int(MBD_ebom_Qty[i])
                    else:
                        merged_data[key] = int(MBD_ebom_Qty[i])

                # 清空原始列表并重新填充合并后的数据
                MBD_ebom_Typical.clear()
                MBD_ebom_DT.clear()
                MBD_ebom_Partnumber.clear()
                MBD_ebom_Qty.clear()

                for key, qty in merged_data.items():
                    MBD_ebom_Typical.append(key[0])
                    MBD_ebom_DT.append(key[1])
                    MBD_ebom_Partnumber.append(key[2])
                    MBD_ebom_Qty.append(qty)

                for i in range(0, len(MBD_ebom_Typical)):
                    for j in range(0, len(devicelabel_A)):
                        if MBD_ebom_Typical[i] == devicelabel_A[j] and MBD_ebom_DT[i] == devicelabel_B[j] and MBD_ebom_Partnumber[i] in devicelabel_D[j]:
                            MCB_aux_amount = 0

                            sub_devices = devicelabel_D[j].split()
                            for sub_device in sub_devices:
                                if sub_device == MBD_ebom_Partnumber[i]:
                                    MCB_aux_amount += 1

                            if MCB_aux_amount != int(MBD_ebom_Qty[i]):
                                error_MCB_aux_number_flag = True
                                text.insert(tk.INSERT, '▲ %s    %s空开辅助触点数量应为%s，导出的BOM报表中数量为%s，请复核\n' % (MBD_ebom_Typical[i], MBD_ebom_DT[i], MCB_aux_amount, MBD_ebom_Qty[i]), 'error')
                                error24_calculator += 1

                if not error_MCB_aux_number_flag:
                    text.insert(tk.INSERT, '无空开辅助触点数量问题\n')

                text.insert(tk.INSERT, '\n>>>保护LED灯非标标签正在检查中...\n')
                protection_led_check(Device_Label_file_path)

                end = time()
                text.insert(tk.INSERT, "\n>>>BOM配置信息检查完成!  用时%.3f秒\n" % (end - start))

                button_export_report['state'] = 'normal'

                global checklog_file_path
                checklogbook = load_workbook(checklog_file_path)
                checklogsheet = checklogbook['Sheet']
                project_no = stem.replace('-BOM', '')
                current_time = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
                data1 = [project_no, 'BOM检查-EBOM中物料数量为0或空', error1_calculator, current_time]
                data2 = [project_no, 'BOM检查-EBOM中物料定位缺失', error2_calculator, current_time]
                data3 = [project_no, 'BOM检查-EBOM中物料定位错误', error3_calculator, current_time]
                data4 = [project_no, 'BOM检查-EBOM中同一物料多定位错误', error4_calculator, current_time]
                data5 = [project_no, 'BOM检查-EBOM中空开(ABB/人民电器)混用', error5_calculator, current_time]
                data6 = [project_no, 'BOM检查-EBOM中空开缺辅助触点/辅助触点缺空开', error6_calculator, current_time]
                data7 = [project_no, 'BOM检查-EBOM中多物料共用ID', error7_calculator, current_time]
                data15 = [project_no, 'BOM检查-EBOM中AIS容性分压装置物料号', error15_calculator, current_time]
                data16 = [project_no, 'BOM检查-EBOM中低压室照明灯物料号', error16_calculator, current_time]
                data17 = [project_no, 'BOM检查-AIS柜体BOM中微动开关配置1', error17_calculator, current_time]
                data18 = [project_no, 'BOM检查-AIS柜体BOM中微动开关配置2', error18_calculator, current_time]
                data19 = [project_no, 'BOM检查-EBOM中500柜CT型号150b/4', error19_calculator, current_time]
                data20 = [project_no, 'BOM检查-SAP柜体BOM柜宽与SLD柜宽一致性', error20_calculator, current_time]
                data21 = [project_no, 'BOM检查-SLD中ZS1(500/Beni/ZS1)、ZX1.2/2、PrimeGear ZX0断路器相间距、电压', error21_calculator, current_time]
                data22 = [project_no, 'BOM检查-SLD中断路器描述、柜号与SAP断路器BOM描述、柜号一致性', error22_calculator, current_time]
                data23 = [project_no, 'BOM检查-ZS1中SAP断路器BOM选配与EPLAN所选Macro一致性', error23_calculator, current_time]
                data24 = [project_no, 'BOM检查-EBOM中空开辅助触点数量校核', error24_calculator, current_time]
                checklogsheet.append(data1)
                checklogsheet.append(data2)
                checklogsheet.append(data3)
                checklogsheet.append(data4)
                checklogsheet.append(data5)
                checklogsheet.append(data6)
                checklogsheet.append(data7)
                checklogsheet.append(data15)
                checklogsheet.append(data16)
                checklogsheet.append(data17)
                checklogsheet.append(data18)
                checklogsheet.append(data19)
                checklogsheet.append(data20)
                checklogsheet.append(data21)
                checklogsheet.append(data22)
                checklogsheet.append(data23)
                checklogsheet.append(data24)

                checklogbook.save(checklog_file_path)
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def sap_decrypt_data(data, data_cipher, prefix):
    # 移除列名前缀prefix，如果有的话
    if data.startswith(prefix):
        data = data[2:]

    # Base64 字符串中可能删除了等号，需要添加回原来的填充字符，因为 base64 编码的输出长度应该是 4 的倍数
    padding_needed = 4 - (len(data) % 4)
    if padding_needed:
        data += "=" * padding_needed

    # base64 解码以获取原始的加密字节串
    encrypted_data = base64.urlsafe_b64decode(data.encode())

    # 使用 data_cipher 对象解密，它应该有和加密时一样的密钥
    decrypted_data = data_cipher.decrypt(encrypted_data)

    # 将字节解码回字符串
    decrypted_string = decrypted_data.decode()
    return decrypted_string


def extract_table(page, bbox, table_settings):
    table = page.within_bbox(bbox).extract_tables(table_settings)
    table_row_number_count = [len(x) for x in table]
    sorted_lst = sorted(table_row_number_count, reverse=True)  # 表格行数降序
    unique_one_values = list(sorted_lst)
    top_one_values = unique_one_values[:1]
    tables = []
    for subtable in table:
        if len(subtable) in top_one_values:
            df = pd.DataFrame(subtable)
            tables.append(df)
    return tables


def sld_recognition(pdf_file_path, sld_table_path):
    global stem
    if not os.path.exists(sld_table_path):
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet('Sheet1')
        workbook.save(sld_table_path)
    sld_page = []
    sheet_name_list = []
    # 利用PyPDF2读取pdf文件的大纲
    with open(pdf_file_path, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)
        outlines = pdf_reader.outline
        if len(outlines) > 3:
            for idx, item in enumerate(outlines[3]):
                title = str(item.get("/Title", ""))

                if "&SLD/" in title and ("单线图" in title or "SINGLE LINE DIAGRAM" in title.upper()):
                    sld_page.append(idx)
                    sheet_name_list.append(title.replace('==', '').replace('/', '_'))

    # 利用pdfplumber读取pdf文件的表格
    with pdfplumber.open(pdf_file_path) as pdf:
        with pd.ExcelWriter(sld_table_path) as writer:
            for idx, page_num in enumerate(sld_page):
                page = pdf.pages[page_num]
                pw = page.width
                ph = page.height

                table_settings = {
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines",
                    "explicit_vertical_lines": [],
                    "explicit_horizontal_lines": [],
                    "snap_tolerance": 1,
                    "join_tolerance": 1,
                    "edge_min_length": 10,
                    "min_words_vertical": 3,
                    "min_words_horizontal": 1,
                    # "keep_blank_chars": True,    # 当前面策略为text时，才有用
                    "text_tolerance": 3,
                    # "text_x_tolerance": None,
                    # "text_y_tolerance": None,
                    "intersection_tolerance": 3,
                    # "intersection_x_tolerance": None,
                    # "intersection_y_tolerance": None,
                }

                # 提取当前页面的 bbox1 和 bbox2 中的表格
                bbox1 = (0, 0, int(pw / 4.0), int(2 * ph / 3.0))
                tables_bbox1 = extract_table(page, bbox1, table_settings)
                bbox2 = (0, int(ph / 2.0), pw, ph)
                tables_bbox2 = extract_table(page, bbox2, table_settings)

                # 创建一个空 DataFrame，用于存储当前页面的合并后的表格数据
                merged_tables = pd.DataFrame()

                # 将 bbox1 的表格添加到 merged_tables 的第一列
                for table in tables_bbox1:
                    merged_tables = pd.concat([merged_tables, table], axis=1)

                # 将 bbox2 的表格添加到 merged_tables 的后面列
                for table in tables_bbox2:
                    merged_tables = pd.concat([merged_tables, table], axis=1)

                # 将当前页面的合并后的表格数据保存到当前页面的工作表中
                sheet_name = sheet_name_list[idx]
                merged_tables.to_excel(writer, sheet_name=sheet_name, index=False)


def attribute_get(Panel_file_path, Panel_data_file_path, Panel_Size_file_path):
    stem, suffix = os.path.splitext(os.path.basename(Panel_file_path))  # stem是文件名,suffix是后缀
    # os.path.dirname()去掉文件名，返回目录
    # os.path.basename()去掉目录，返回文件名(含后缀)

    book = load_workbook(Panel_file_path)
    sheet = book['Z2_xlsx']
    A = []
    B = []
    C = []
    for i in range(2, sheet.max_row + 1):
        A.append(str(sheet.cell(row=i, column=1).value))  # Order Line列
        B.append(str(sheet.cell(row=i, column=2).value))  # Typical列
        C.append(str(sheet.cell(row=i, column=3).value))  # Panel No列

    if B[0] == 'None':  # 说明是旧版升版2024项目，它的Panel.xlsx报表B列为空，需要用Panel.xls，如果是纯2024项目，只需Panel.xlsx即可
        A = []
        B = []
        C = []
        book = xlrd.open_workbook(os.path.join(os.path.dirname(FilePath), stem + '.xls'))  # Panel.xls
        sheet = book.sheet_by_index(0)
        A = sheet.col_values(colx=0, start_rowx=1, end_rowx=None)  # Order Line列
        B = sheet.col_values(colx=1, start_rowx=1, end_rowx=None)  # Typical列
        C = sheet.col_values(colx=2, start_rowx=1, end_rowx=None)  # Panel No列

    book1 = load_workbook(Panel_data_file_path)
    sheet1 = book1['Z3_xlsx']
    A1 = []
    B1 = []
    C1 = []
    D1 = []
    for i in range(2, sheet1.max_row + 1):
        A1.append(str(sheet1.cell(row=i, column=1).value))  # Typical/Panel列
        B1.append(str(sheet1.cell(row=i, column=2).value))  # 柜宽（Typical）列
        C1.append(str(sheet1.cell(row=i, column=3).value))  # 柜深（Typical）列
        D1.append(str(sheet1.cell(row=i, column=4).value))  # 柜高（Panel）列

    book2 = xlwt.Workbook()  # 创建一个空文件对象
    book2.add_sheet('Sheet1')  # 创建一个Sheet页
    book2.save(Panel_Size_file_path)  # 创建Panel Size.xls文件
    book2 = xlrd.open_workbook(Panel_Size_file_path)  # 加载【项目号-Panel Size.xls】表格

    workbook = copy(book2)  # 使用xlutils.copy将xlrd读取的book对象转为xlwt可操作对象
    worksheet = workbook.get_sheet(0)  # 获取sheet
    worksheet.write(0, 0, '站号')  # 在sheet指定位置写入数据
    worksheet.write(0, 1, '柜型')  # 在sheet指定位置写入数据
    worksheet.write(0, 2, '柜号')  # 在sheet指定位置写入数据
    worksheet.write(0, 3, '宽(mm)')  # 在sheet指定位置写入数据
    worksheet.write(0, 4, '深(mm)')  # 在sheet指定位置写入数据
    worksheet.write(0, 5, '高(mm)')  # 在sheet指定位置写入数据
    for i in range(0, len(A)):  # 遍历Panel.xls，将Panel.xls数据复制到Panel Size.xls
        worksheet.write(i + 1, 0, A[i])
        worksheet.write(i + 1, 1, B[i])
        worksheet.write(i + 1, 2, C[i])

    flag = 0
    flag1 = 0
    NAN_flag = 0

    def custom_filter(char):
        return char.isdigit() or char == '+'

    for i in range(0, len(A)):  # 遍历Panel.xlsx
        if B[i] not in A1:
            worksheet.write(i + 1, 3, '--')
            worksheet.write(i + 1, 4, '--')
            NAN_flag += 1
        for j in range(0, len(A1)):  # 遍历Panel data.xls，逐行读取尺寸信息
            if B[i] == A1[j] and B1[j] != 'None' and C1[j] != 'None':
                worksheet.write(i + 1, 3, str(''.join(filter(custom_filter, B1[j]))))  # https://blog.csdn.net/weixin_44606217/article/details/100534834
                worksheet.write(i + 1, 4, str(''.join(filter(custom_filter, C1[j]))))
                flag = 1
            if B[i] == A1[j] and B1[j] == 'None' and C1[j] != 'None':
                NAN_flag += 1
                worksheet.write(i + 1, 3, '--')
                worksheet.write(i + 1, 4, str(''.join(filter(custom_filter, C1[j]))))
            if B[i] == A1[j] and B1[j] != 'None' and C1[j] == 'None':
                NAN_flag += 1
                worksheet.write(i + 1, 3, str(''.join(filter(custom_filter, B1[j]))))
                worksheet.write(i + 1, 4, '--')
            if B[i] == A1[j] and B1[j] == 'None' and C1[j] == 'None':
                NAN_flag += 1
                worksheet.write(i + 1, 3, '--')
                worksheet.write(i + 1, 4, '--')
            if C[i] == A1[j] and D1[j] != 'None':
                worksheet.write(i + 1, 5, str(''.join(filter(custom_filter, D1[j]))))
                flag1 = 1
            if C[i] == A1[j] and D1[j] == 'None':
                worksheet.write(i + 1, 5, '--')
                NAN_flag += 1
                flag1 = 1
            if flag and flag1:
                break  # 当宽高深数据都匹配到，就跳出，继续Panel.xls下一行
        flag = 0
        flag1 = 0
    workbook.save(Panel_Size_file_path)  # 将workbook保存到指定位置

    # 数据表为了方便与单线图核对，应将“宽，深，高”改成“宽，高，深”排列，交换5,6列即可
    book2 = xlrd.open_workbook(Panel_Size_file_path)  # 加载【项目号-Panel Size.xls】表格
    sheet2 = book2.sheet_by_index(0)
    workbook = copy(book2)  # 使用xlutils.copy将xlrd读取的book对象转为xlwt可操作对象
    worksheet = workbook.get_sheet(0)  # 获取sheet
    switchgear_number_panel_size = sheet2.col_values(colx=0, start_rowx=0, end_rowx=None)
    typical_type_panel_size = sheet2.col_values(colx=1, start_rowx=0, end_rowx=None)
    panel_number_panel_size = sheet2.col_values(colx=2, start_rowx=0, end_rowx=None)

    depth = sheet2.col_values(colx=4, start_rowx=0, end_rowx=None)
    for i in range(1, len(depth)):
        if '+' in depth[i]:
            text.insert(tk.INSERT, f'▲ {switchgear_number_panel_size[i]}    {typical_type_panel_size[i]}    {panel_number_panel_size[i]}柜深：{depth[i]}，请移除+号及其后尺寸，只保留+号前面的尺寸\n', 'error')
    height = sheet2.col_values(colx=5, start_rowx=0, end_rowx=None)
    for i in range(0, len(depth)):
        worksheet.write(i, 4, str(height[i]))  # 将”高“数据放到第5列
        worksheet.write(i, 5, str(depth[i]))  # 将”宽“数据放到第6列
    workbook.save(Panel_Size_file_path)  # 将workbook保存到指定位置

    # 为了方便与单线图核对，将表格按照站号对柜号进行排序
    df = pd.read_excel(Panel_Size_file_path, names=['站号', '柜型', '柜号', '宽(mm)', '高(mm)', '深(mm)'])
    df.sort_values(by=['站号', '柜号'], ascending=True, inplace=True)  # 升序
    df.to_excel(Panel_Size_file_path, index=False)  # 写入excel时不带索引


def compare_sld_and_attribute(Panel_Size_file_path, switchgear_number, typical_type_list, abb_panel_number, switchgear_dimension):
    book = xlrd.open_workbook(Panel_Size_file_path)
    sheet = book.sheet_by_index(0)

    def convert_cell(cell):
        if isinstance(cell, float):
            return str(int(cell))  # 对于实数，先取整，然后转换为字符串
        return str(cell)  # 其他类型直接转换为字符串

    A1 = [convert_cell(cell) for cell in sheet.col_values(colx=0, start_rowx=1)]
    B1 = [convert_cell(cell) for cell in sheet.col_values(colx=1, start_rowx=1)]
    C1 = [convert_cell(cell) for cell in sheet.col_values(colx=2, start_rowx=1)]
    D1 = [convert_cell(cell) for cell in sheet.col_values(colx=3, start_rowx=1)]
    E1 = [convert_cell(cell) for cell in sheet.col_values(colx=4, start_rowx=1)]
    F1 = [convert_cell(cell) for cell in sheet.col_values(colx=5, start_rowx=1)]

    # 清洗后的abb_panel_number列表去除了字符串中的分号
    cleaned_abb_panel_number = [num.replace(';', '') for num in abb_panel_number]

    overall_match = True

    for i in range(len(A1)):
        found = False
        for j in range(len(switchgear_number)):
            if A1[i] == switchgear_number[j] and B1[i] == typical_type_list[j] and C1[i] == cleaned_abb_panel_number[j]:
                found = True
                if '+' in switchgear_dimension[j]:
                    if '+' in F1[i]:  # 单线图深度尺寸有+,EPLAN属性有+
                        overall_match = False
                        text.insert(tk.INSERT, f'▲ {A1[i]}    {B1[i]}    {C1[i]}尺寸：{D1[i]}X{E1[i]}X{F1[i]}(EPLAN属性)，{switchgear_dimension[j]}(SLD)，请移除EPLAN属性+号及其后尺寸，只保留+号前面的尺寸\n', 'error')
                    else:  # 单线图深度尺寸有+,EPLAN属性无+
                        if not (D1[i] in switchgear_dimension[j] and E1[i] in switchgear_dimension[j] and F1[i] in switchgear_dimension[j]):
                            overall_match = False
                            text.insert(tk.INSERT, f'▲ {A1[i]}    {B1[i]}    {C1[i]}尺寸：{D1[i]}X{E1[i]}X{F1[i]}(EPLAN属性)，{switchgear_dimension[j]}(SLD)\n', 'error')

                else:
                    if not (D1[i] in switchgear_dimension[j] and E1[i] in switchgear_dimension[j] and F1[i] in switchgear_dimension[j]):
                        overall_match = False
                        text.insert(tk.INSERT, f'▲ {A1[i]}    {B1[i]}    {C1[i]}尺寸：{D1[i]}X{E1[i]}X{F1[i]}(EPLAN属性)，{switchgear_dimension[j]}(SLD)\n', 'error')

        if not found:
            overall_match = False
            text.insert(tk.INSERT, f'▲ {A1[i]}    {B1[i]}    {C1[i]}在EPLAN属性或SLD中不存在\n', 'error')

    return overall_match


def confirm_logic(typical, A2, D2, A_real, D_real):
    global MB585_flag
    global down_PT_truck
    global CB_eop
    global ES_exist
    global ES_eop

    for i in range(0, len(A_real)):
        if A_real[i] == typical:
            if 'RLY-MB585' in D_real[i]:
                MB585_flag = True

    for i in range(0, len(A2)):
        if A2[i] == typical:
            # print(D2[i])
            if 'QTV-A' in D2[i]:
                down_PT_truck = True

            if ('VD4/P' in D2[i] and '-T' in D2[i]) or ('VD4-' in D2[i] and '-T' in D2[i]) or ('VD4N-' in D2[i] and '-T' in D2[i]):
                CB_eop = True

            if 'EK6-' in D2[i] or 'ST-E-' in D2[i] or 'ST2-' in D2[i] or 'ST1-' in D2[i] or 'ET1-' in D2[i]:
                ES_exist = True

            if ('EK6-' in D2[i] and '-M' in D2[i]) or ('ST-E-' in D2[i] and '-M' in D2[i]) or ('ST2-' in D2[i] and '-M' in D2[i]) or ('ST1-' in D2[i] and '-M' in D2[i]) or ('ET1-' in D2[i] and '-M' in D2[i]):
                ES_eop = True

    # print(typical, MB585_flag, down_PT_truck, CB_eop, ES_exist, ES_eop)


def protection_led_check(Device_Label_file_path):
    book = load_workbook(Device_Label_file_path)
    sheet = book['Z5_xlsx']
    A = []
    B = []
    C = []
    D = []
    E = []
    F = []
    G = []
    H = []
    I = []
    J = []
    K = []
    L = []
    M = []
    N = []
    O = []
    for i in range(2, sheet.max_row + 1):
        A.append(str(sheet.cell(row=i, column=1).value))
        B.append(str(sheet.cell(row=i, column=2).value))
        C.append(str(sheet.cell(row=i, column=3).value))
        D.append(str(sheet.cell(row=i, column=4).value))
        E.append(str(sheet.cell(row=i, column=5).value))
        F.append(str(sheet.cell(row=i, column=6).value))
        G.append(str(sheet.cell(row=i, column=7).value))
        H.append(str(sheet.cell(row=i, column=8).value))
        I.append(str(sheet.cell(row=i, column=9).value))
        J.append(str(sheet.cell(row=i, column=10).value))
        K.append(str(sheet.cell(row=i, column=11).value))
        L.append(str(sheet.cell(row=i, column=12).value))
        M.append(str(sheet.cell(row=i, column=13).value))
        N.append(str(sheet.cell(row=i, column=14).value))
        O.append(str(sheet.cell(row=i, column=15).value))

    # print(E[0], F[0], G[0], H[0], I[0], J[0], K[0], L[0], M[0], N[0], O[0])
    if E[0].strip() == 'LED1' and F[0].strip() == 'LED2' and G[0].strip() == 'LED3' and H[0].strip() == 'LED4' and I[0].strip() == 'LED5' and J[0].strip() == 'LED6' and K[0].strip() == 'LED7' and L[0].strip() == 'LED8' and M[0].strip() == 'LED9' and N[0].strip() == 'LED10' and O[0].strip() == 'LED11':
        non_standard_led = False
        non_standard_led_typical_list = []
        for i in range(1, len(A)):
            if E[i].strip() != 'None' or F[i].strip() != 'None' or G[i].strip() != 'None' or H[i].strip() != 'None' or I[i].strip() != 'None' or J[i].strip() != 'None' or K[i].strip() != 'None' or L[i].strip() != 'None' or M[i].strip() != 'None' or N[i].strip() != 'None' or O[i].strip() != 'None':
                # print(i, E[i], F[i], G[i], H[i], I[i], J[i], K[i], L[i], M[i], N[i], O[i])
                non_standard_led = True
                non_standard_led_typical_list.append(A[i])
        if non_standard_led:
            unique_typical_list = "，".join(sorted(list(set(non_standard_led_typical_list))))
            text.insert(tk.INSERT, "▲ %s，存在保护LED灯非标标签，请在Datasheet中注明保护标签需要更换\n" % unique_typical_list, 'error')
        else:
            text.insert(tk.INSERT, "无保护LED灯非标标签\n")

    else:
        text.insert(tk.INSERT, "▲ DeviceLabel是旧版本，请重新导出新版本报表Z5\n", 'error')


def export_report():
    try:
        pdfmetrics.registerFont(TTFont('SimSun', 'simsun.ttc'))  # 注册字体
        home_path = os.path.expanduser("~")
        desktop_path = os.path.join(home_path, "Desktop")
        global project_number
        project_number = stem.replace('-BOM', '')
        desktop_path = askdirectory(title=u'请选择导出文件夹', initialdir=desktop_path)
        if not desktop_path:
            tk.messagebox.showwarning("提示", "未选择保存路径，导出操作已取消")
            return

        doc = SimpleDocTemplate(os.path.join(desktop_path, 'BOM检查报告-%s.pdf' % project_number), pagesize=A4, rightMargin=40, leftMargin=40, topMargin=100, bottomMargin=20)
        story = []
        # 获取样式
        styles = getSampleStyleSheet()
        style = styles['Normal']
        style.fontName = 'SimSun'
        style.wordWrap = 'CJK'
        style.leading = 15  # 行间距

        # 获取文本内容
        text_content = text_to_html(text)

        # 创建段落并添加到story
        p = Paragraph(text_content, style)
        story.append(p)

        # 定义表格名称
        table_names = ['附表1：柜型参数表', '附表2：空开统计表', '附表3：其他统计表']

        table_index = 0

        for treeview in [Calculate_table0, Calculate_table1, Calculate_table2]:
            # # 在表格前添加表格名称的段落
            table_name = table_names[table_index]
            table_header_style = styles['Title']
            table_header_style.leading = 22
            table_header_style.spaceAfter = 0
            table_header_style.fontName = 'SimSun'  # 使用SimSun字体
            table_header_style.fontSize = 10  # 可以设置更大的字体大小

            # 创建表格名称的段落并添加到story
            p = Paragraph(table_name, table_header_style)
            story.append(p)

            table_data = extract_treeview_content(treeview)

            # 创建表格实例
            table = Table(table_data, spaceAfter=10)
            # 可以通过 TableStyle 添加自定义样式
            if table_index == 0:  # 如果不是第三个表格, 用常规样式
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.white),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'SimSun'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                story.append(table)

            if table_index == 1:
                table_data = merge_table_content(table_data)
                for table_stack in table_data:
                    # 这里创建局部样式styleN以避免影响全局样式
                    styleN = styles['Normal'].clone('new_centered_style')  # 创建一个基于Normal的新样式副本
                    styleN.alignment = "JUSTIFY"  # 设置对齐为两端对齐
                    styleN.fontName = 'SimSun'
                    # 构建包含旋转文本段落的单元格数据
                    # table_data_rotated = [[Paragraph("<para align=justify>{}</para>".format(cell), styleN) for cell in row] for row in table_data]
                    # table = Table(table_data_rotated, spaceAfter=10)
                    table = Table(table_stack, spaceAfter=10)

                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.white),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, -1), 'SimSun'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ]))
                    story.append(table)

            if table_index == 2:
                # 这里创建局部样式styleN以避免影响全局样式
                styleN = styles['Normal'].clone('new_centered_style')  # 创建一个基于Normal的新样式副本
                styleN.alignment = "JUSTIFY"  # 设置对齐为两端对齐
                styleN.fontName = 'SimSun'
                # 构建包含旋转文本段落的单元格数据
                table_data_rotated = [[Paragraph("<para align=justify>{}</para>".format(cell), styleN) for cell in row] for row in table_data]
                table = Table(table_data_rotated, spaceAfter=10)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.white),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'SimSun'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))

                story.append(table)
            table_index += 1
        global page_number
        page_number = 1

        # 通过`SimpleDocTemplate`直接生成文档，传入`add_header`函数绘制每页的头部内容
        doc.build(story, onFirstPage=add_header1, onLaterPages=add_header2)
        tk.messagebox.showwarning("提示", "FAST BOM检查报告导出完成")
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def add_header1(rl_canvas, doc):
    rl_canvas.saveState()

    # 添加图片（请确保图片路径正确）
    image_path = 'ico/ABB_logo.png'
    img_width = 27.7 * mm
    img_height = 10.5 * mm
    img = Image(image_path, width=img_width, height=img_height)
    img.drawOn(rl_canvas, A4[0] / 2 - img_width / 2, A4[1] - img_height - 30)

    # 添加标题
    rl_canvas.setFont('SimSun', 16)
    title = "FAST BOM检查报告"
    rl_canvas.drawCentredString(A4[0] / 2, A4[1] - img_height - 60, title)

    rl_canvas.setFont('SimSun', 10)
    project_number_text = "项目号：" + project_number
    rl_canvas.drawString(45, A4[1] - img_height - 60 - 10, project_number_text)

    rl_canvas.setFont('SimSun', 8)
    global page_number
    page_number_text = "Page " + str(page_number)
    page_number += 1
    rl_canvas.drawString(A4[0] - 80, 5 * mm, page_number_text)

    rl_canvas.setFont('SimSun', 8)
    Timestamp = strftime('%Y-%m-%d %H:%M:%S', localtime())
    rl_canvas.drawCentredString(A4[0] / 2, 5 * mm, '制表时间：' + Timestamp)

    rl_canvas.restoreState()


def add_header2(rl_canvas, doc):
    rl_canvas.saveState()

    # 添加图片（请确保图片路径正确）
    image_path = 'ico/ABB_logo.png'
    img_width = 27.7 * mm
    img_height = 10.5 * mm
    img = Image(image_path, width=img_width, height=img_height)
    img.drawOn(rl_canvas, A4[0] / 2 - img_width / 2, A4[1] - img_height - 30)

    # 添加标题
    rl_canvas.setFont('SimSun', 16)
    title = "FAST BOM检查报告"
    rl_canvas.drawCentredString(A4[0] / 2, A4[1] - img_height - 60, title)

    rl_canvas.setFont('SimSun', 8)
    global page_number
    page_number_text = "Page " + str(page_number)
    page_number += 1
    rl_canvas.drawString(A4[0] - 80, 5 * mm, page_number_text)

    rl_canvas.setFont('SimSun', 8)
    Timestamp = strftime('%Y-%m-%d %H:%M:%S', localtime())
    rl_canvas.drawCentredString(A4[0] / 2, 5 * mm, '制表时间：' + Timestamp)

    rl_canvas.restoreState()


def text_to_html(text_widget):
    """将Tkinter Text widget内容转换为HTML."""

    # 获取文本及其所有“tag”属性
    text_content = text_widget.get("1.0", tk.END)
    text_tags = text_widget.tag_ranges(tk.ALL)

    html_output = "<html><body><p>"

    current_pos = "1.0"
    for i in range(0, len(text_tags), 2):
        # 获取当前标记的范围
        start, end = text_tags[i], text_tags[i + 1]
        tag_name = text_widget.tag_names(start)[0]
        # 获取标记的范围内的文本
        text_segment = text_widget.get(current_pos, start)
        # 将换行符转换为HTML的换行标签
        html_segment = text_segment.replace('\n', '<br />')
        # 添加到HTML输出中
        html_output += html_segment

        if tag_name == 'sel':  # 选中文本，保持默认处理
            text_segment = text_widget.get(start, end)
            html_segment = text_segment.replace('\n', '<br />')
        else:
            # 获取标签配置，例如颜色
            tag_config = text_widget.tag_configure(tag_name)
            foreground = tag_config['foreground'][-1]  # 获取前景色
            text_segment = text_widget.get(start, end)
            # 标记带颜色的文本
            html_segment = f'<span style="color: {foreground}">{text_segment}</span>'

        html_output += html_segment.replace('\n', '<br />')

        # 更新当前位置
        current_pos = end

    # 获取最后一段文本并添加到HTML输出
    text_segment = text_widget.get(current_pos, tk.END)
    html_segment = text_segment.replace('\n', '<br />')
    html_output += html_segment

    # 结束HTML文档
    html_output += "</p></body></html>"
    return html_output


def extract_treeview_content(treeview):
    # 提取 Treeview 表格内容
    columns = treeview["columns"]
    column_headings = [treeview.heading(col)['text'] for col in columns]  # 所有列的标题
    table_data = [column_headings]  # 包含表头的完整表格数据

    for child in treeview.get_children():
        row_data = [treeview.set(child, col) for col in columns]  # 忽略#0列，只从其他列收集数据
        table_data.append(row_data)

    return table_data


def merge_table_content(data):
    # # 获得所有唯一的柜型和标签
    # cabinet_types = sorted(set(row[0] for row in data[1:]))
    # labels = sorted(set(row[1] for row in data[1:]))
    #
    # # 创建初始透视表，行为唯一柜型加上一个空行，列为唯一标签加上一个空标签
    # pivot_table = [[''] + labels]
    #
    # # 为每个柜型创建一行，并用空字符串进行初始化
    # for cabinet in cabinet_types:
    #     pivot_table.append([cabinet] + ['' for _ in labels])
    #
    # # 创建一个简单的映射，以便可以通过标签找到对应的列索引
    # label_to_index = {label: index + 1 for index, label in enumerate(labels)}  # 列索引 +1 是为了跳过行头
    #
    # # 填充透视表数据
    # for row in data[1:]:
    #     cabinet, label, voltage, manufacturer, aux_contact = row
    #     # 找到对应柜型和标签的索引
    #     row_index = cabinet_types.index(cabinet) + 1  # 行索引 +1 是为了跳过标题行
    #     col_index = label_to_index[label]
    #     # 将电压等级，厂家，辅助触点组合为一个字符串，并分配给对应单元格
    #     pivot_table[row_index][col_index] = f"{voltage}\n{manufacturer}\n{aux_contact}"
    #
    # return pivot_table

    # 最大列数为10，-1 是为了让第一列空出来放柜型
    MAX_COLUMNS = 11

    # 提取各个柜型和标签，创建唯一的集合
    cabinet_types = sorted(set(row[0] for row in data[1:]))
    unique_labels = sorted(set(row[1] for row in data[1:]))

    # 准备数据结构
    pivot_data = {cabinet: {label: '' for label in unique_labels} for cabinet in cabinet_types}

    # 填充数据到透视表结构
    for row in data[1:]:
        cabinet = row[0]
        label = row[1]
        values = '\n'.join(row[2:])
        pivot_data[cabinet][label] = values

    # 拆分表头来适应最大列数，生成多个透视表
    tables = []
    labels_covered = 0
    while labels_covered < len(unique_labels):
        # 当前表覆盖的标签范围
        labels_for_this_table = unique_labels[labels_covered: labels_covered + MAX_COLUMNS - 1]
        labels_covered += MAX_COLUMNS - 1

        # 制作表头，包含柜型和当前表的标签
        header = [''] + labels_for_this_table
        tables.append([header])

        # 根据当前覆盖的标签填充每个柜型对应的数据行
        for cabinet in cabinet_types:
            row = [cabinet] + [pivot_data[cabinet].get(label, '') for label in labels_for_this_table]
            tables[-1].append(row)
    return tables


def draw_rotated_text(rl_canvas, text, x, y, angle, fontsize=10):
    rl_canvas.saveState()
    rl_canvas.translate(x, y)
    rl_canvas.rotate(angle)
    rl_canvas.setFont("Helvetica", fontsize)
    rl_canvas.drawString(0, 0, text)
    rl_canvas.restoreState()
