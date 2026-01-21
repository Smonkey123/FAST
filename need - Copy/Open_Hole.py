import tkinter as tk
import PyPDF2
from tkinter.filedialog import askdirectory
import os

import time
import datetime
from time import *

import xlrd
from openpyxl import load_workbook
import traceback

import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

FilePath = ""    # 设置一个地址变量
from need.custom_dialogs import CustomDialog, center_window, Tooltip, image_label
from need.LVD_pdf_extract import extract_device_from_pdf


def main(parent, w_rat, h_rat, file_path):
    global w_ratio
    w_ratio = w_rat

    global h_ratio
    h_ratio = h_rat

    global checklog_file_path
    checklog_file_path = file_path
    global open_folder
    open_folder = tk.PhotoImage(file="ico\\open_folder.png")
    global analyze_file
    analyze_file = tk.PhotoImage(file="ico\\read.png")

    tk.Label(parent, text="欢迎使用开孔图纸检查功能", bg="#c9dbe9", fg="black", height=int(1*h_ratio), font=("ABBvoice CNSG", int(20 * h_ratio), "bold")).pack(side=tk.TOP, fill=tk.X)
    f1 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f1, text='   说明：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')
    tk.Label(f1, text='(1)EPLAN菜单Tools→Reports:Automated processing中选择Export Files V04导出；\n(2)选择低压室面板开孔文件【C:/Temp/项目号-Files/项目号-LVD layout.xlsx】；\n(3)对EPLAN导出的开孔物料单和设备清单（低压室面板）进行对比检查。', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    f1.pack(fill=tk.X)

    tk.Frame(parent, height=int(20*h_ratio), bg="#eaf1f6").pack(fill=tk.X)    # 水平分割线

    f2 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f2, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2, text='路径：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry    # 为了确保selectpath函数能正确调用entry,将其全局化
    entry = tk.Entry(f2, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)))
    entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

    global button_check
    button_check = tk.Button(f2, text='检查', font=("ABBvoice CNSG", int(13 * h_ratio)), bg="#eaf1f6", image=analyze_file, compound=tk.LEFT, command=process, activebackground='blue')
    button_check.pack(side=tk.RIGHT, padx=(0, int(20*w_ratio)))
    button_check['state'] = 'disabled'

    tk.Button(f2, text='选择', font=("ABBvoice CNSG", int(13 * h_ratio)), image=open_folder, bg="#eaf1f6", compound=tk.LEFT, command=selectpath, activebackground='blue').pack(side=tk.RIGHT, padx=int(20*w_ratio))

    f2.pack(fill=tk.X)

    tk.Frame(parent, height=int(20*h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f3 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f3, text='   结果：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')

    global text
    text = tk.Text(f3, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), height=int(25 * h_ratio), width=int(65*w_ratio))
    text.pack(side=tk.LEFT, padx=(0, 1), pady=0, fill=tk.BOTH, expand=True)

    text.tag_configure('error', foreground='red')  # 设置tag

    scrollbar = tk.Scrollbar(f3)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, int(20*w_ratio)))
    scrollbar.config(command=text.yview)
    text.config(yscrollcommand=scrollbar.set)
    f3.pack(fill=tk.BOTH, expand=True)

    tk.Frame(parent, height=int(20*h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线


def selectpath():
    filepath = tk.filedialog.askopenfilename(initialdir='C:/Temp/', title=u'请选择文件(项目号-LVD layout.xlsx)', filetypes=[("Excel", ".xlsx")])    # 选择打开什么文件，返回文件名
    if len(filepath) != 0:
        string_filename = ""
        for i in range(0, len(filepath)):
            string_filename += str(filepath[i])
        button_check['state'] = 'normal'
    else:
        button_check['state'] = 'disabled'

    text.delete(1.0, "end")    # 清空输出结果框
    entry.delete(0, "end")    # 删除entry原始内容
    entry.insert(0, filepath)    # 重新填入地址
    global FilePath
    FilePath = filepath


def process():
    try:
        error_calculator = 0
        text.delete(1.0, tk.END)  # 清空输出结果框

        if FilePath == "":
            tk.messagebox.showwarning("提示", "请选择文件！")
        stem, suffix = os.path.splitext(os.path.basename(FilePath))    # stem是文件名,suffix是后缀
        if '-LVD' not in stem:
            tk.messagebox.showwarning("提示", "请选择项目号-LVD layout.xlsx文件！")

        else:
            new_stem = stem.replace('-LVD layout', '')
            pdf_file_path = os.path.join(os.path.dirname(FilePath), new_stem + '.pdf')

            info = None

            if os.path.exists(pdf_file_path):
                info = extract_device_from_pdf(pdf_file_path)

            text.insert(tk.INSERT, '>>>开孔信息正在读取中...\n')    # 进行-LVD layout.xlsx表格处理
            start = time()
            # print("正在处理中...")
            book = load_workbook(FilePath)
            sheet = book['Z4_xlsx']
            A = []
            B = []
            C = []

            for i in range(2, sheet.max_row + 1):
                A.append(str(sheet.cell(row=i, column=1).value))    # Typical列
                B.append(str(sheet.cell(row=i, column=2).value))    # DT列
                C.append(str(sheet.cell(row=i, column=3).value))    # Part Number列

            stem, suffix = os.path.splitext(os.path.basename(FilePath))    # stem是文件名,suffix是后缀
            # os.path.dirname()去掉文件名，返回目录
            # os.path.basename()去掉目录，返回文件名(含后缀)
            projectnumber = stem.rsplit('-', 1)
            inputfile = os.path.join(os.path.dirname(FilePath), projectnumber[0]+'-DeviceLabel'+suffix)    # -DeviceLabel.xlsx

            book1 = load_workbook(inputfile)
            sheet1 = book1['Z5_xlsx']
            A1 = []
            B1 = []
            C1 = []
            D1 = []
            for i in range(3, sheet1.max_row + 1):
                A1.append(str(sheet1.cell(row=i, column=1).value))    # Typical列
                B1.append(str(sheet1.cell(row=i, column=2).value))    # DT列
                C1.append(str(sheet1.cell(row=i, column=3).value))    # Zone列
                D1.append(str(sheet1.cell(row=i, column=4).value))    # PartNumber列

            # 筛选后的设备清单数据
            Typical_Device = []
            DT_Device = []
            PartNumber_Device = []

            for i in range(0, len(A1)):    # 项目号-DeviceLabel.xls
                if C1[i] == 'LVD' and (B1[i] != 'LVD' or D1[i] != 'Mounting_Panel'):
                    Typical_Device.append(A1[i])
                    DT_Device.append(B1[i])
                    PartNumber_Device.append(D1[i])

            for i in range(0, len(Typical_Device)):
                Flag1 = 0    # 该物料设备清单有，面板开孔无
                for j in range(0, len(A)):    # 对筛选后的设备清单数据逐个去门板开孔遍历查找
                    if A[j] == Typical_Device[i] and B[j] == DT_Device[i] and C[j] == PartNumber_Device[i]:
                        Flag1 = 1    # 没问题
                        break
                    if A[j] == Typical_Device[i] and B[j] == DT_Device[i] and C[j] != PartNumber_Device[i]:
                        Flag1 = 2    # 标识一致，物料号不一致
                        break
                if Flag1 == 0:
                    if info is not None:
                        for _, devices, typical in info:
                            if Typical_Device[i] == typical:
                                if ('-'+DT_Device[i]) not in devices:
                                    text.insert(tk.INSERT, "%s：设备清单中%s(物料号%s),在低压室面板不存在\n" % (Typical_Device[i], DT_Device[i], PartNumber_Device[i]), 'error')
                                    error_calculator += 1
                    else:
                        text.insert(tk.INSERT, "%s：设备清单中%s(物料号%s),在低压室面板不存在\n" % (Typical_Device[i], DT_Device[i], PartNumber_Device[i]), 'error')
                        error_calculator += 1
                if Flag1 == 2:
                    text.insert(tk.INSERT, "%s：设备清单中%s(物料号%s),与低压室面板标识一致，物料号不一致\n" % (Typical_Device[i], DT_Device[i], PartNumber_Device[i]), 'error')
                    error_calculator += 1
            for i in range(0, len(A)):
                Flag2 = 0    # 该物料面板开孔有，设备清单无
                for j in range(0, len(Typical_Device)):    # 对门板开孔逐个去筛选后的设备清单数据遍历查找
                    if A[i] == Typical_Device[j] and B[i] == DT_Device[j] and C[i] == PartNumber_Device[j]:
                        Flag2 = 1    # 没问题
                        break
                    if A[i] == Typical_Device[j] and B[i] == DT_Device[j] and C[i] != PartNumber_Device[j]:
                        Flag2 = 3    # 标识一致，物料号不一致
                        break
                if Flag2 == 0:
                    text.insert(tk.INSERT, "%s：低压室面板中%s(物料号%s),在设备清单不存在\n" % (A[i], B[i], C[i]), 'error')
                    error_calculator += 1
                if Flag2 == 3:
                    text.insert(tk.INSERT, "%s：低压室面板中%s(物料号%s),与设备清单标识一致，物料号不一致\n" % (A[i], B[i], C[i]), 'error')
                    error_calculator += 1

            end = time()

            text.insert(tk.INSERT, ">>>面板开孔检查完成!  用时%.3f秒\n" % (end - start))

            global checklog_file_path
            checklogbook = load_workbook(checklog_file_path)
            checklogsheet = checklogbook['Sheet']
            project_no = stem.replace('-LVD layout', '')
            current_time = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            data = [project_no, '开孔检查', error_calculator, current_time]
            checklogsheet.append(data)
            checklogbook.save(checklog_file_path)

    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())
