import tkinter as tk
from tkinter import ttk
import tkinter.font as tkFont
import need.tkutils as tku
from PIL import Image, ImageTk
from tkinter.filedialog import askdirectory
import os


import time
import numpy as np
from time import *
import re
import pandas as pd
import xlrd
import xlwt
from xlutils.copy import copy
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import sys
import shutil    # 文件操作
import xml.etree.ElementTree as ET

import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

FilePath = ""    # 设置一个地址变量


def center_window(win, width=None, height=None):
    """ 将窗口屏幕居中 """
    screenwidth = win.winfo_screenwidth()
    screenheight = win.winfo_screenheight()
    if width is None:
        width, height = get_window_size(win)[:2]
    size = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 3)
    win.geometry(size)


def get_window_size(win, update=True):
    """ 获得窗体的尺寸 """
    if update:
        win.update()
    return win.winfo_width(), win.winfo_height(), win.winfo_x(), win.winfo_y()


def tkimg_resized(img, w_box, h_box, keep_ratio=True):
    """对图片进行按比例缩放处理"""
    w, h = img.size

    if keep_ratio:
        if w > h:
            width = w_box
            height = int(h_box * (1.0 * h / w))

        if h >= w:
            height = h_box
            width = int(w_box * (1.0 * w / h))
    else:
        width = w_box
        height = h_box

    img1 = img.resize((width, height), Image.ANTIALIAS)
    tkimg = ImageTk.PhotoImage(img1)
    return tkimg


def image_label(frame, img, width, height, keep_ratio=True):
    """输入图片信息，及尺寸，返回界面组件"""
    if isinstance(img, str):
        _img = Image.open(img)
    else:
        _img = img
    lbl_image = tk.Label(frame, width=width, height=height)

    tk_img = tkimg_resized(_img, width, height, keep_ratio)
    lbl_image.image = tk_img
    lbl_image.config(image=tk_img)
    return lbl_image


def _font(fname="微软雅黑", size=12, bold=tkFont.NORMAL):
    """设置字体"""
    ft = tkFont.Font(family=fname, size=size, weight=bold)
    return ft


def _ft(size=12, bold=False):
    """极简字体设置函数"""
    if bold:
        return _font(size=size, bold=tkFont.BOLD)
    else:
        return _font(size=size, bold=tkFont.NORMAL)


def h_seperator(parent, height=2):  # height 单位为像素值
    """水平分割线, 水平填充 """
    tk.Frame(parent, height=height, bg="whitesmoke").pack(fill=tk.X)


def v_seperator(parent, width, bg="whitesmoke"):  # width 单位为像素值
    """垂直分割线 , fill=tk.Y, 但如何定位不确定，直接返回对象，由容器决定 """
    frame = tk.Frame(parent, width=width, bg=bg)
    return frame


class Window:
    def __init__(self, parent):
        self.root = tk.Toplevel()
        self.parent = parent
        self.root.geometry("%dx%d" % (600, 480))    # 窗体尺寸
        center_window(self.root)    # 将窗体移动到屏幕中央
        self.root.title('MVE CTPT Parameter Export')    # 设置窗口标题
        self.root.iconbitmap("ico\\logo_new.ico")    # 窗口图标
        self.root.resizable(False, False)  # 设置窗体不可改变大小
        self.no_title = False    # 标题有无标志
        self.root.overrideredirect(self.no_title)    # 设置无标题框
        self.body()    # 绘制窗体组件

    def body(self):
        self.title(self.root).pack(fill=tk.X)
        self.main(self.root).pack(expand=tk.YES, fill=tk.BOTH)
        # self.bottom(self.root).pack(fill=tk.X)

    def title(self, parent):
        """ 标题栏 """

        def label(frame, text, size, bold=False):
            return tk.Label(frame, text=text, bg="#c9dbe9", fg="black", height=2, font=_ft(size, bold))

        frame = tk.Frame(parent, bg="#c9dbe9", bd=0)    # 标题栏处的矩形框架(容器)
        label(frame, "      ", 16, True).pack(side=tk.LEFT)
        label(frame, "欢迎使用CTPT参数导出程序", 15, True).pack(side=tk.LEFT, padx=125)

        # 额外增加一个子功能退出按钮
        im = tku.image_label(frame, "ico\\close.png", 20, 20, False)
        im.configure(bg="#c9dbe9")
        im.bind('<Button-1>', self.close)  # 关闭图标绑定动作,用于弹出退出确定框，若确定,则关闭程序
        im.pack(side=tk.RIGHT, padx=10)

        return frame

    def close(self, *arg):    # 点击退出按钮，销毁当前子窗口,显示父窗口
        if tku.show_confirm("确认退出吗 ?"):
            self.root.destroy()
            self.parent.deiconify()

    def bottom(self, parent):
        """ 窗体最下面留空白 """

        frame = tk.Frame(parent, height=10, bg="black", bd=0)    # 底部的矩形框架(容器)
        # frame.propagate(False)    # 自动根据子组件改变自身大小
        return frame

    def main(self, parent):
        """ 窗体主体 """

        frame = tk.Frame(parent, bg="whitesmoke", bd=0)    # 主窗体的矩形框架(容器)

        self.main_top(frame).pack(fill=tk.X, padx=10, pady=10)
        return frame

    def main_top(self, parent):
        def label(frame, text, size=10, bold=False, fg="black", justify=tk.CENTER):
            return tk.Label(frame, text=text, bg="white", fg=fg, font=_ft(size, bold), justify=justify)

        def space(n):
            s = " "
            r = ""
            for i in range(n):
                r += s
            return r

        frame = tk.Frame(parent, bg="white", bd=0)    # 主窗体容器的子容器

        label(frame, "功能实现", 12, True).pack(anchor=tk.W, padx=20, pady=5)

        h_seperator(frame, 10)    # 水平分割

        f1 = tk.Frame(frame, bg="white")
        label(f1, space(8) + "文件来源:").pack(side=tk.LEFT, pady=10)
        label(f1, "①选中所有站号，Options→Administration中选择Grant Authority MVE，给MVE权限，\n②选中项目名，Project→Export导出，选取该.mve文件后进行导出。", size=8, justify=tk.LEFT).pack(side=tk.LEFT, padx=20, pady=10)
        f1.pack(fill=tk.X)

        f2 = tk.Frame(frame, bg="white")
        label(f2, space(5) + "*", fg="red").pack(side=tk.LEFT, pady=10)
        label(f2, "目标路径:").pack(side=tk.LEFT)
        global entry    # 为了确保selectpath函数能正确调用entry,将其全局化
        entry = tk.Entry(f2, bg="white", font=_ft(10), width=25)
        entry.pack(side=tk.LEFT, padx=20)

        button = ttk.Button(f2, text="选取文件", width=12, command=self.selectpath)
        button.pack(side=tk.LEFT, padx=0, pady=10)

        ttk.Button(f2, text="开始处理", width=12, command=self.process).pack(side=tk.LEFT, padx=20, pady=10)
        f2.pack(fill=tk.X)

        f3 = tk.Frame(frame, bg="white")
        label(f3, space(8) + "结果输出:").pack(side=tk.LEFT, anchor=tk.N, pady=10)
        global text
        text = tk.Text(f3, bg="white", font=_ft(9), height=12, width=61)
        text.pack(side=tk.LEFT, padx=20, pady=10, fill=tk.X, expand=tk.N)

        scrollbar = tk.Scrollbar(f3)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        scrollbar.config(command=text.yview)
        text.config(yscrollcommand=scrollbar.set)
        f3.pack(fill=tk.BOTH)

        return frame

    def selectpath(self):
        self.filepath = tk.filedialog.askopenfilename(title=u'请选择文件(XXX.mve)', filetypes=[("Excel", ".mve")])    # 选择打开什么文件，返回文件名
        if len(self.filepath) != 0:
            string_filename = ""
            for i in range(0, len(self.filepath)):
                string_filename += str(self.filepath[i])
            # print("您选择的文件是：" + string_filename)
        # else:
        #     print("您未选择任何文件")    # 点击"取消"，返回空
        text.delete(1.0, tk.END)    # 清空输出结果框
        entry.delete(0, "end")    # 删除entry原始内容
        entry.insert(0, self.filepath)    # 重新填入地址
        global FilePath
        FilePath = self.filepath

    def process(self):
        text.delete(1.0, tk.END)  # 清空输出结果框
        if FilePath == "":
            tk.messagebox.showwarning("提示", "请选择文件！")
        stem, suffix = os.path.splitext(os.path.basename(FilePath))    # stem是文件名,suffix是后缀
        if 'mve' not in suffix:
            tk.messagebox.showwarning("提示", "请选择.mve文件！")

        else:
            TargetPath = 'C:/Temp/target.xml'
            shutil.copy(FilePath, TargetPath)    # 复制.mve文件并命名为.xml，方便后续解析
            text.insert(tk.INSERT, '>>>CT、PT参数正在读取中...\n')    # 进行CT、PT参数表格处理
            start = time()

            # xml文件解析操作
            tree = ET.parse(TargetPath)    # 把xml文件解析为Element tree,调用此函数返回Element tree实例对象
            root = tree.getroot()    # 获取根节点（ProjectXml）
            project = root.find('project')

            # Switchgear下的信息
            SwitchgearnameList = []

            # Typical下的信息
            TypicalnameList = []
            PanelnameList = []
            PanelwidthList = []
            TotalCTParameterList = []
            TotalPTParameterList = []
            Typicalname = ''
            Panelwidth = ''
            Panelname = ''

            for switchgear in project.find('switchgearlist'):    # 遍历得到各站
                for attribute in switchgear.find('attributelist'):
                    for name in attribute.findall('name'):
                        if name.text == 'ObjectName':
                            Switchgearname = attribute.find('value').text    # 得到站号

                for typical in switchgear.find('typicallist'):
                    for attribute in typical.find('attributelist'):
                        for name in attribute.findall('name'):
                            if name.text == 'ObjectName':
                                Typicalname = attribute.find('value').text    # 得到Typicalname
                            if name.text == 'PanelNames':
                                Panelname = attribute.find('value').text    # 得到Panelname
                            if name.text == 'PanelWidth':
                                Panelwidth = attribute.find('value').text    # 得到Panelwidth
                    CTParameterList = []
                    PTParameterList = []
                    for zone in typical.find('zonelist'):
                        for device in zone.find('devicelist'):
                            NumCore = '0'  # CT绕组数
                            ObjectName_CT = ''
                            CT_IP = ''
                            CT_Core1_IS = ''
                            CT_Core1_Class = ''
                            CT_Core1_Burden = ''
                            CT_IP2 = ''
                            CT_Core2_IS = ''
                            CT_Core2_Class = ''
                            CT_Core2_Burden = ''
                            CT_IP3 = ''
                            CT_Core3_IS = ''
                            CT_Core3_Class = ''
                            CT_Core3_Burden = ''
                            CT_IP4 = ''
                            CT_Core4_IS = ''
                            CT_Core4_Class = ''
                            CT_Core4_Burden = ''

                            NumSec = '0'  # PT二次绕组数
                            ObjectName_PT = ''
                            VT_PrimVol = ''
                            VT_Core1_SecVol = ''
                            VT_Core2_SecVol = ''
                            VT_Core3_SecVol = ''
                            VT_Core4_SecVol = ''
                            VT_Core1_Class = ''
                            VT_Core1_Burden = ''
                            VT_Core2_Class = ''
                            VT_Core2_Burden = ''
                            VT_Core3_Class = ''
                            VT_Core3_Burden = ''
                            VT_Core4_Class = ''
                            VT_Core4_Burden = ''
                            for attribute in device.find('attributelist'):
                                for name in attribute.findall('name'):
                                    if name.text == 'ObjectName' and '-BCT' in attribute.find('value').text:
                                        ObjectName_CT = attribute.find('value').text    # 得到CT的ObjectName
                                    if name.text == 'CT.NumCore':
                                        NumCore = attribute.find('value').text    # 记录下该CT的绕组数
                                    if NumCore == '1':    # 每个CT只有一个绕组
                                        for attribute in device.find('attributelist'):
                                            for name in attribute.findall('name'):
                                                if name.text == 'CT.IP':    # 一次电流
                                                    CT_IP = attribute.find('value').text
                                                if name.text == 'CT.Core1.IS':    # 二次电流
                                                    CT_Core1_IS = attribute.find('value').text
                                                if name.text == 'CT.Core1.Class':    # 准确级
                                                    CT_Core1_Class = attribute.find('value').text
                                                if name.text == 'CT.Core1.Burden':    # 容量
                                                    CT_Core1_Burden = attribute.find('value').text
                                    if NumCore == '2':    # 每个CT有二个绕组
                                        for attribute in device.find('attributelist'):
                                            for name in attribute.findall('name'):
                                                if name.text == 'CT.IP':  # 一次电流
                                                    CT_IP = attribute.find('value').text
                                                if name.text == 'CT.Core1.IS':  # 二次电流
                                                    CT_Core1_IS = attribute.find('value').text
                                                if name.text == 'CT.Core1.Class':  # 准确级
                                                    CT_Core1_Class = attribute.find('value').text
                                                if name.text == 'CT.Core1.Burden':  # 容量
                                                    CT_Core1_Burden = attribute.find('value').text
                                                if name.text == 'CT.IP2':
                                                    CT_IP2 = attribute.find('value').text
                                                if name.text == 'CT.Core2.IS':
                                                    CT_Core2_IS = attribute.find('value').text
                                                if name.text == 'CT.Core2.Class':
                                                    CT_Core2_Class = attribute.find('value').text
                                                if name.text == 'CT.Core2.Burden':
                                                    CT_Core2_Burden = attribute.find('value').text
                                    if NumCore == '3':    # 每个CT有三个绕组
                                        for attribute in device.find('attributelist'):
                                            for name in attribute.findall('name'):
                                                if name.text == 'CT.IP':  # 一次电流
                                                    CT_IP = attribute.find('value').text
                                                if name.text == 'CT.Core1.IS':  # 二次电流
                                                    CT_Core1_IS = attribute.find('value').text
                                                if name.text == 'CT.Core1.Class':  # 准确级
                                                    CT_Core1_Class = attribute.find('value').text
                                                if name.text == 'CT.Core1.Burden':  # 容量
                                                    CT_Core1_Burden = attribute.find('value').text
                                                if name.text == 'CT.IP2':
                                                    CT_IP2 = attribute.find('value').text
                                                if name.text == 'CT.Core2.IS':
                                                    CT_Core2_IS = attribute.find('value').text
                                                if name.text == 'CT.Core2.Class':
                                                    CT_Core2_Class = attribute.find('value').text
                                                if name.text == 'CT.Core2.Burden':
                                                    CT_Core2_Burden = attribute.find('value').text
                                                if name.text == 'CT.IP3':
                                                    CT_IP3 = attribute.find('value').text
                                                if name.text == 'CT.Core3.IS':
                                                    CT_Core3_IS = attribute.find('value').text
                                                if name.text == 'CT.Core3.Class':
                                                    CT_Core3_Class = attribute.find('value').text
                                                if name.text == 'CT.Core3.Burden':
                                                    CT_Core3_Burden = attribute.find('value').text
                                    if NumCore == '4':    # 每个CT有四个绕组
                                        for attribute in device.find('attributelist'):
                                            for name in attribute.findall('name'):
                                                if name.text == 'CT.IP':  # 一次电流
                                                    CT_IP = attribute.find('value').text
                                                if name.text == 'CT.Core1.IS':  # 二次电流
                                                    CT_Core1_IS = attribute.find('value').text
                                                if name.text == 'CT.Core1.Class':  # 准确级
                                                    CT_Core1_Class = attribute.find('value').text
                                                if name.text == 'CT.Core1.Burden':  # 容量
                                                    CT_Core1_Burden = attribute.find('value').text
                                                if name.text == 'CT.IP2':
                                                    CT_IP2 = attribute.find('value').text
                                                if name.text == 'CT.Core2.IS':
                                                    CT_Core2_IS = attribute.find('value').text
                                                if name.text == 'CT.Core2.Class':
                                                    CT_Core2_Class = attribute.find('value').text
                                                if name.text == 'CT.Core2.Burden':
                                                    CT_Core2_Burden = attribute.find('value').text
                                                if name.text == 'CT.IP3':
                                                    CT_IP3 = attribute.find('value').text
                                                if name.text == 'CT.Core3.IS':
                                                    CT_Core3_IS = attribute.find('value').text
                                                if name.text == 'CT.Core3.Class':
                                                    CT_Core3_Class = attribute.find('value').text
                                                if name.text == 'CT.Core3.Burden':
                                                    CT_Core3_Burden = attribute.find('value').text
                                                if name.text == 'CT.IP4':
                                                    CT_IP4 = attribute.find('value').text
                                                if name.text == 'CT.Core4.IS':
                                                    CT_Core4_IS = attribute.find('value').text
                                                if name.text == 'CT.Core4.Class':
                                                    CT_Core4_Class = attribute.find('value').text
                                                if name.text == 'CT.Core4.Burden':
                                                    CT_Core4_Burden = attribute.find('value').text

                                    if name.text == 'ObjectName' and '-BAT' in attribute.find('value').text:
                                        ObjectName_PT = attribute.find('value').text    # 得到PT的ObjectName
                                    if name.text == 'VT.NumSec':
                                        NumSec = attribute.find('value').text    # 记录下该PT的二次绕组数
                                    if NumSec == '1':    # 每个PT只有一个二次绕组
                                        for attribute in device.find('attributelist'):
                                            for name in attribute.findall('name'):
                                                if name.text == 'VT.PrimVol':    # 一次额定电压
                                                    VT_PrimVol = attribute.find('value').text
                                                if name.text == 'VT.Core1.SecVol':    # 二次额定电压
                                                    VT_Core1_SecVol = attribute.find('value').text
                                                if name.text == 'VT.Core1.Class':    # 准确级
                                                    VT_Core1_Class = attribute.find('value').text
                                                if name.text == 'VT.Core1.Burden':    # 容量
                                                    VT_Core1_Burden = attribute.find('value').text
                                    if NumSec == '2':    # 每个PT有两个二次绕组
                                        for attribute in device.find('attributelist'):
                                            for name in attribute.findall('name'):
                                                if name.text == 'VT.PrimVol':    # 一次额定电压
                                                    VT_PrimVol = attribute.find('value').text
                                                if name.text == 'VT.Core1.SecVol':    # 第一个绕组的二次额定电压
                                                    VT_Core1_SecVol = attribute.find('value').text
                                                if name.text == 'VT.Core2.SecVol':    # 第二个绕组的二次额定电压
                                                    VT_Core2_SecVol = attribute.find('value').text
                                                if name.text == 'VT.Core1.Class':    # 第一个绕组的准确级
                                                    VT_Core1_Class = attribute.find('value').text
                                                if name.text == 'VT.Core1.Burden':    # 第一个绕组的容量
                                                    VT_Core1_Burden = attribute.find('value').text
                                                if name.text == 'VT.Core2.Class':    # 第二个绕组的准确级
                                                    VT_Core2_Class = attribute.find('value').text
                                                if name.text == 'VT.Core2.Burden':    # 第二个绕组的容量
                                                    VT_Core2_Burden = attribute.find('value').text
                                    if NumSec == '3':    # 每个PT有三个二次绕组
                                        for attribute in device.find('attributelist'):
                                            for name in attribute.findall('name'):
                                                if name.text == 'VT.PrimVol':    # 一次额定电压
                                                    VT_PrimVol = attribute.find('value').text
                                                if name.text == 'VT.Core1.SecVol':    # 第一个绕组的二次额定电压
                                                    VT_Core1_SecVol = attribute.find('value').text
                                                if name.text == 'VT.Core2.SecVol':    # 第二个绕组的二次额定电压
                                                    VT_Core2_SecVol = attribute.find('value').text
                                                if name.text == 'VT.Core3.SecVol':    # 第三个绕组的二次额定电压
                                                    VT_Core3_SecVol = attribute.find('value').text
                                                if name.text == 'VT.Core1.Class':    # 第一个绕组的准确级
                                                    VT_Core1_Class = attribute.find('value').text
                                                if name.text == 'VT.Core1.Burden':    # 第一个绕组的容量
                                                    VT_Core1_Burden = attribute.find('value').text
                                                if name.text == 'VT.Core2.Class':    # 第二个绕组的准确级
                                                    VT_Core2_Class = attribute.find('value').text
                                                if name.text == 'VT.Core2.Burden':    # 第二个绕组的容量
                                                    VT_Core2_Burden = attribute.find('value').text
                                                if name.text == 'VT.Core3.Class':    # 第三个绕组的准确级
                                                    VT_Core3_Class = attribute.find('value').text
                                                if name.text == 'VT.Core3.Burden':    # 第三个绕组的容量
                                                    VT_Core3_Burden = attribute.find('value').text
                                    if NumSec == '4':    # 每个PT有四个二次绕组
                                        for attribute in device.find('attributelist'):
                                            for name in attribute.findall('name'):
                                                if name.text == 'VT.PrimVol':    # 一次额定电压
                                                    VT_PrimVol = attribute.find('value').text
                                                if name.text == 'VT.Core1.SecVol':    # 第一个绕组的二次额定电压
                                                    VT_Core1_SecVol = attribute.find('value').text
                                                if name.text == 'VT.Core2.SecVol':    # 第二个绕组的二次额定电压
                                                    VT_Core2_SecVol = attribute.find('value').text
                                                if name.text == 'VT.Core3.SecVol':    # 第三个绕组的二次额定电压
                                                    VT_Core3_SecVol = attribute.find('value').text
                                                if name.text == 'VT.Core4.SecVol':    # 第四个绕组的二次额定电压
                                                    VT_Core4_SecVol = attribute.find('value').text
                                                if name.text == 'VT.Core1.Class':    # 第一个绕组的准确级
                                                    VT_Core1_Class = attribute.find('value').text
                                                if name.text == 'VT.Core1.Burden':    # 第一个绕组的容量
                                                    VT_Core1_Burden = attribute.find('value').text
                                                if name.text == 'VT.Core2.Class':    # 第二个绕组的准确级
                                                    VT_Core2_Class = attribute.find('value').text
                                                if name.text == 'VT.Core2.Burden':    # 第二个绕组的容量
                                                    VT_Core2_Burden = attribute.find('value').text
                                                if name.text == 'VT.Core3.Class':    # 第三个绕组的准确级
                                                    VT_Core3_Class = attribute.find('value').text
                                                if name.text == 'VT.Core3.Burden':    # 第三个绕组的容量
                                                    VT_Core3_Burden = attribute.find('value').text
                                                if name.text == 'VT.Core4.Class':    # 第四个绕组的准确级
                                                    VT_Core4_Class = attribute.find('value').text
                                                if name.text == 'VT.Core4.Burden':    # 第四个绕组的容量
                                                    VT_Core4_Burden = attribute.find('value').text

                            if NumCore == '1' and CTParameterList:
                                CTParameterList.append('\n' + ObjectName_CT + '\n' + CT_IP + '/' + CT_Core1_IS + ',' + CT_Core1_Class + '/' + CT_Core1_Burden)
                            if NumCore == '1' and not CTParameterList:
                                CTParameterList.append(ObjectName_CT + '\n' + CT_IP + '/' + CT_Core1_IS + ',' + CT_Core1_Class + '/' + CT_Core1_Burden)
                            if NumCore == '2' and CTParameterList:
                                CTParameterList.append('\n' + ObjectName_CT + '\n' + CT_IP + '/' + CT_Core1_IS + ',' + CT_Core1_Class + '/' + CT_Core1_Burden + '\n' + CT_IP2 + '/' + CT_Core2_IS + ',' + CT_Core2_Class + '/' + CT_Core2_Burden)
                            if NumCore == '2' and not CTParameterList:
                                CTParameterList.append(ObjectName_CT + '\n' + CT_IP + '/' + CT_Core1_IS + ',' + CT_Core1_Class + '/' + CT_Core1_Burden + '\n' + CT_IP2 + '/' + CT_Core2_IS + ',' + CT_Core2_Class + '/' + CT_Core2_Burden)
                            if NumCore == '3' and CTParameterList:
                                CTParameterList.append('\n' + ObjectName_CT + '\n' + CT_IP + '/' + CT_Core1_IS + ',' + CT_Core1_Class + '/' + CT_Core1_Burden + '\n' + CT_IP2 + '/' + CT_Core2_IS + ',' + CT_Core2_Class + '/' + CT_Core2_Burden + '\n' + CT_IP3 + '/' + CT_Core3_IS + ',' + CT_Core3_Class + '/' + CT_Core3_Burden)
                            if NumCore == '3' and not CTParameterList:
                                CTParameterList.append(ObjectName_CT + '\n' + CT_IP + '/' + CT_Core1_IS + ',' + CT_Core1_Class + '/' + CT_Core1_Burden + '\n' + CT_IP2 + '/' + CT_Core2_IS + ',' + CT_Core2_Class + '/' + CT_Core2_Burden + '\n' + CT_IP3 + '/' + CT_Core3_IS + ',' + CT_Core3_Class + '/' + CT_Core3_Burden)
                            if NumCore == '4' and CTParameterList:
                                CTParameterList.append('\n' + ObjectName_CT + '\n' + CT_IP + '/' + CT_Core1_IS + ',' + CT_Core1_Class + '/' + CT_Core1_Burden + '\n' + CT_IP2 + '/' + CT_Core2_IS + ',' + CT_Core2_Class + '/' + CT_Core2_Burden + '\n' + CT_IP3 + '/' + CT_Core3_IS + ',' + CT_Core3_Class + '/' + CT_Core3_Burden + '\n' + CT_IP4 + '/' + CT_Core4_IS + ',' + CT_Core4_Class + '/' + CT_Core4_Burden)
                            if NumCore == '4' and not CTParameterList:
                                CTParameterList.append(ObjectName_CT + '\n' + CT_IP + '/' + CT_Core1_IS + ',' + CT_Core1_Class + '/' + CT_Core1_Burden + '\n' + CT_IP2 + '/' + CT_Core2_IS + ',' + CT_Core2_Class + '/' + CT_Core2_Burden + '\n' + CT_IP3 + '/' + CT_Core3_IS + ',' + CT_Core3_Class + '/' + CT_Core3_Burden + '\n' + CT_IP4 + '/' + CT_Core4_IS + ',' + CT_Core4_Class + '/' + CT_Core4_Burden)

                            if NumSec == '1' and PTParameterList:
                                PTParameterList.append('\n' + ObjectName_PT + '\n' + VT_PrimVol + '|' + VT_Core1_SecVol + '\n' + VT_Core1_Class + '/' + VT_Core1_Burden)
                            if NumSec == '1' and not PTParameterList:
                                PTParameterList.append(ObjectName_PT + '\n' + VT_PrimVol + '|' + VT_Core1_SecVol + '\n' + VT_Core1_Class + '/' + VT_Core1_Burden)
                            if NumSec == '2' and PTParameterList:
                                PTParameterList.append('\n' + ObjectName_PT + '\n' + VT_PrimVol + '|' + VT_Core1_SecVol + '|' + VT_Core2_SecVol + '\n' + VT_Core1_Class + '/' + VT_Core1_Burden + '\n' + VT_Core2_Class + '/' + VT_Core2_Burden)
                            if NumSec == '2' and not PTParameterList:
                                PTParameterList.append(ObjectName_PT + '\n' + VT_PrimVol + '|' + VT_Core1_SecVol + '|' + VT_Core2_SecVol + '\n' + VT_Core1_Class + '/' + VT_Core1_Burden + '\n' + VT_Core2_Class + '/' + VT_Core2_Burden)
                            if NumSec == '3' and PTParameterList:
                                PTParameterList.append('\n' + ObjectName_PT + '\n' + VT_PrimVol + '|' + VT_Core1_SecVol + '|' + VT_Core2_SecVol + '|' + VT_Core3_SecVol + '\n' + VT_Core1_Class + '/' + VT_Core1_Burden + '\n' + VT_Core2_Class + '/' + VT_Core2_Burden + '\n' + VT_Core3_Class + '/' + VT_Core3_Burden)
                            if NumSec == '3' and not PTParameterList:
                                PTParameterList.append(ObjectName_PT + '\n' + VT_PrimVol + '|' + VT_Core1_SecVol + '|' + VT_Core2_SecVol + '|' + VT_Core3_SecVol + '\n' + VT_Core1_Class + '/' + VT_Core1_Burden + '\n' + VT_Core2_Class + '/' + VT_Core2_Burden + '\n' + VT_Core3_Class + '/' + VT_Core3_Burden)
                            if NumSec == '4' and PTParameterList:
                                PTParameterList.append('\n' + ObjectName_PT + '\n' + VT_PrimVol + '|' + VT_Core1_SecVol + '|' + VT_Core2_SecVol + '|' + VT_Core3_SecVol + '|' + VT_Core4_SecVol + '\n' + VT_Core1_Class + '/' + VT_Core1_Burden + '\n' + VT_Core2_Class + '/' + VT_Core2_Burden + '\n' + VT_Core3_Class + '/' + VT_Core3_Burden + '\n' + VT_Core4_Class + '/' + VT_Core4_Burden)
                            if NumSec == '4' and not PTParameterList:
                                PTParameterList.append(ObjectName_PT + '\n' + VT_PrimVol + '|' + VT_Core1_SecVol + '|' + VT_Core2_SecVol + '|' + VT_Core3_SecVol + '|' + VT_Core4_SecVol + '\n' + VT_Core1_Class + '/' + VT_Core1_Burden + '\n' + VT_Core2_Class + '/' + VT_Core2_Burden + '\n' + VT_Core3_Class + '/' + VT_Core3_Burden + '\n' + VT_Core4_Class + '/' + VT_Core4_Burden)

                    if not CTParameterList:    # 如果此Typical下无CT，则参数信息置为NAN
                        TotalCTParameterList.append('NAN')
                    else:
                        TotalCTParameterList.append(CTParameterList)

                    if not PTParameterList:    # 如果此Typical下无PT，则参数信息置为NAN
                        TotalPTParameterList.append('NAN')
                    else:
                        TotalPTParameterList.append(PTParameterList)

                    if Panelwidth == '':
                        Panelwidth = 'NAN'
                    if Panelname == '':
                        Panelname = 'NAN'
                    if Typicalname == '':
                        Typicalname = 'NAN'

                    PanelnameList.append(Panelname)
                    PanelwidthList.append(Panelwidth)
                    TypicalnameList.append(Typicalname)
                    SwitchgearnameList.append(Switchgearname)

            stem, suffix = os.path.splitext(os.path.basename(FilePath))    # stem是文件名,suffix是后缀
            # os.path.dirname()去掉文件名，返回目录
            # os.path.basename()去掉目录，返回文件名(含后缀)

            outputfile = os.path.join(os.path.dirname(FilePath), stem+'-Parameter CTPT.xls')    # Parameter CTPT.xls

            book2 = xlwt.Workbook()    # 创建一个空文件对象
            book2.add_sheet('Sheet1')    # 创建一个Sheet页
            book2.save(outputfile)    # 创建-Parameter CTPT.xls文件
            book2 = xlrd.open_workbook(outputfile)    # 加载【-Parameter CTPT.xls】表格

            workbook = copy(book2)    # 使用xlutils.copy将xlrd读取的book对象转为xlwt可操作对象
            worksheet = workbook.get_sheet(0)    # 获取sheet
            worksheet.write(0, 0, '站号')    # 在sheet指定位置写入数据
            worksheet.write(0, 1, '柜型')    # 在sheet指定位置写入数据
            worksheet.write(0, 2, '柜号')    # 在sheet指定位置写入数据
            worksheet.write(0, 3, '柜宽(mm)')    # 在sheet指定位置写入数据
            worksheet.write(0, 4, 'CT参数')    # 在sheet指定位置写入数据
            worksheet.write(0, 5, 'PT参数')    # 在sheet指定位置写入数据

            style = xlwt.XFStyle()
            style.alignment.wrap = 1    # 设置自动换行
            style.alignment.horz = 0x01    # 0x01(左端对齐)、0x02(水平方向上居中对齐)、0x03(右端对齐)
            style.alignment.vert = 0x01    # 0x00(上端对齐)、 0x01(垂直方向上居中对齐)、0x02(底端对齐)
            for i in range(0, len(SwitchgearnameList)):    # 将数据写入到-Parameter CTPT.xls
                worksheet.write(i+1, 0, SwitchgearnameList[i], style)
                worksheet.write(i+1, 1, TypicalnameList[i], style)
                worksheet.write(i+1, 2, PanelnameList[i], style)
                worksheet.write(i+1, 3, int(PanelwidthList[i]), style)
                worksheet.write(i+1, 4, TotalCTParameterList[i], style)
                worksheet.write(i+1, 5, TotalPTParameterList[i], style)
            worksheet.col(0).width = 1200
            worksheet.col(1).width = 1200
            worksheet.col(2).width = 6000
            worksheet.col(3).width = 2000
            worksheet.col(4).width = 6000
            worksheet.col(5).width = 6000
            workbook.save(outputfile)    # 将workbook保存到指定位置

            text.insert(tk.INSERT, ">>>生成%s\n" % outputfile.replace("\\", "/"))
            end = time()

            text.insert(tk.INSERT, ">>>CT、PT参数导出完成!  用时%.3f秒\n" % (end - start))



