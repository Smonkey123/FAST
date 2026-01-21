import tkinter as tk
from tkinter import ttk
from tkinter import StringVar
from tkinter import IntVar
from tkinter.ttk import Treeview
import tkinter.font as tkFont
import need.tkutils as tku
from tkinter.filedialog import askdirectory
from tkinter import Menu
from tkcalendar import CCalendar
from babel import numbers
import ctypes
import calendar
import os
from PIL import Image, ImageTk
import numpy as np
from time import time
from time import localtime
from time import strftime
from playwright.sync_api import Playwright, sync_playwright, expect
import re
import pandas as pd
import xlrd
import xlwt
from xlutils.copy import copy

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
import sys
import shutil  # 文件操作
import warnings
import traceback
import math
from win32com import client
import PyPDF2
import sqlite3


warnings.simplefilter(action='ignore', category=FutureWarning)

FilePath = ""  # 设置一个地址变量


def main(parent, w_ratio, h_ratio):
    global canvas
    canvas = tk.Canvas(parent, width=int(1600 * w_ratio), height=int(640 * h_ratio), bg="#C9DBE9")
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    canvas.update()
    canvas.bind("<MouseWheel>", on_mousewheel)

    scrollbar_v = tk.Scrollbar(master=parent)
    scrollbar_v.pack(side=tk.RIGHT, fill=tk.Y)
    scrollbar_v.config(command=canvas.yview)
    canvas.config(yscrollcommand=scrollbar_v.set)

    content = tk.Frame(canvas)
    # content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    canvas.create_window(0, 1, width=int(1400 * w_ratio), anchor=tk.NW, window=content)

    style = ttk.Style()  # 设计树形表格字体大小
    style.configure("Treeview.Heading", font=("微软雅黑", int(12 * h_ratio)))
    style.configure("Treeview", font=("微软雅黑", int(10 * h_ratio)))

    f1 = tk.Frame(content, bg="#c9dbe9", bd=0)
    # im = tku.image_label(f1, "ico\\help.png", int(30 * h_ratio), int(30 * h_ratio), False)
    # im.configure(bg="#c9dbe9")
    # im.bind('<Button-1>', about_help)  # 帮助图标绑定动作
    # im.pack(side=tk.RIGHT)
    tk.Label(f1, text="欢迎使用CheckList导出功能", bg="#c9dbe9", fg="black", height=int(1 * h_ratio), font=tku._ft(int(20 * h_ratio), True)).pack(fill=tk.X)
    f1.pack(fill=tk.X)

    # tk.Frame(content, height=int(20 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f1 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f1, text='文件来源:                         ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')
    tk.Label(f1, text='(1)选择.mve文件，点击"读取"，获取属性信息。', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    f1.pack(fill=tk.X)

    tk.Frame(content, height=int(20 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f2 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f2, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2, text='目标路径:                       ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global mve_entry    # 为了确保selectpath函数能正确调用entry,将其全局化
    mve_entry = tk.Entry(f2, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(80 * w_ratio))
    mve_entry.pack(side=tk.LEFT, fill=tk.X)

    tk.Button(f2, text="选取文件", width=int(12 * w_ratio), font=tku._ft(int(13 * h_ratio), False), command=selectpath, activebackground='blue').pack(side=tk.LEFT, padx=int(20 * w_ratio))
    # tk.Button(f2, text="开始读取", width=12, command=process).pack(side=tk.LEFT, padx=0)

    global button_read
    button_read = tk.Button(f2, text="开始读取", width=int(12 * w_ratio), command=getmvedata, font=tku._ft(int(13 * h_ratio), False), activebackground='blue')
    button_read.pack(side=tk.LEFT, padx=0)
    button_read['state'] = 'disabled'

    f2.pack(fill=tk.X)

    tk.Frame(content, height=int(20 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    tk.Label(content, text='————————————————————————————————项目基本信息————————————————————————————————', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), True), justify='left').pack(side=tk.TOP, fill=tk.X)
    tk.Label(content, text='*为必填项，#为可选/多选项', bg="white", fg="red", font=tku._ft(int(10 * h_ratio), False), justify='left', anchor='w').pack(side=tk.TOP, fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f20 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f20, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f20, text='项目号:                          ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry_project_number
    entry_project_number = tk.Entry(f20, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(20 * w_ratio))
    entry_project_number.pack(side=tk.LEFT)

    f20.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f21 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f21, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X, padx=0)
    tk.Label(f21, text='项目名:                          ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry_project_name
    entry_project_name = tk.Entry(f21, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(100 * w_ratio))
    entry_project_name.pack(side=tk.LEFT)
    f21.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='区域特殊要求:                ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry_specrequire
    entry_specrequire = tk.Entry(f22, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(100 * w_ratio))
    entry_specrequire.pack(side=tk.LEFT)

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='系统中性点接地方式:      ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_neutralearth_type
    combobox_neutralearth_type_value = StringVar()
    combobox_neutralearth_type_values = ['直接接地', '经低电阻接地', '经高电阻接地', '不接地或经消弧线圈接地']
    combobox_neutralearth_type = ttk.Combobox(master=f22, width=int(28 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_neutralearth_type_value, values=combobox_neutralearth_type_values)
    combobox_neutralearth_type.pack(side=tk.LEFT)

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='适用标准:                       ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global entry_application
    entry_application = tk.Entry(f22, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(10 * w_ratio))
    entry_application.pack(side=tk.LEFT)

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='是否增补:                       ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global combobox_supplement
    combobox_supplement_value = StringVar()
    combobox_supplement_values = ['是', '否']
    combobox_supplement = ttk.Combobox(master=f22, width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_supplement_value, values=combobox_supplement_values)
    combobox_supplement.pack(side=tk.LEFT)

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='ABB母线桥:                    ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_busduct_type
    combobox_busduct_type_value = StringVar()
    combobox_busduct_type_values = ['无', '有']
    combobox_busduct_type = ttk.Combobox(master=f22, width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_busduct_type_value, values=combobox_busduct_type_values)
    combobox_busduct_type.pack(side=tk.LEFT)
    combobox_busduct_type.bind("<<ComboboxSelected>>", on_busduct_combobox_select)

    global frame_busduct1
    frame_busduct1 = tk.Frame(f22, bg="white", bd=0)
    tk.Label(frame_busduct1, text='    *', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(frame_busduct1, text='母线桥额定电流/A:', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global entry_busduct_current
    entry_busduct_current = tk.Entry(frame_busduct1, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(10 * w_ratio))
    entry_busduct_current.pack(side=tk.LEFT)

    tk.Label(frame_busduct1, text='    *', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(frame_busduct1, text='是否需要现场测量:', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global combobox_busduct_measure
    combobox_busduct_measure_value = StringVar()
    combobox_busduct_measure_values = ['是', '否']
    combobox_busduct_measure = ttk.Combobox(master=frame_busduct1, width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_busduct_measure_value, values=combobox_busduct_measure_values)
    combobox_busduct_measure.pack(side=tk.LEFT)

    frame_busduct1.pack_forget()

    global frame_busduct2
    frame_busduct2 = tk.Frame(f22, bg="white", bd=0)

    tk.Label(frame_busduct2, text='    *', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(frame_busduct2, text='母线桥类型:', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global entry_busduct_type
    entry_busduct_type = tk.Entry(frame_busduct2, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(10 * w_ratio))
    entry_busduct_type.pack(side=tk.LEFT)

    tk.Label(frame_busduct2, text='    *', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(frame_busduct2, text='母线桥数量:', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global entry_busduct_amount
    entry_busduct_amount = tk.Entry(frame_busduct2, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(10 * w_ratio))
    entry_busduct_amount.pack(side=tk.LEFT)

    tk.Label(frame_busduct2, text='    *', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(frame_busduct2, text='母线桥长度/m:', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global entry_busduct_length
    entry_busduct_length = tk.Entry(frame_busduct2, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(10 * w_ratio))
    entry_busduct_length.pack(side=tk.LEFT)

    frame_busduct2.pack_forget()

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(15 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    tk.Label(content, text='————————————————————————————————项目设计信息————————————————————————————————', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), True), justify='left').pack(side=tk.TOP, fill=tk.X)
    tk.Label(content, text='一、柜体及一次设备', bg="white", fg="red", font=tku._ft(int(10 * h_ratio), False), justify='left', anchor='w').pack(side=tk.TOP, fill=tk.X)

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='柜体1.1倍温升:               ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_tempture_rise
    combobox_tempture_rise_value = StringVar()
    combobox_tempture_rise_values = ['是', '否']
    combobox_tempture_rise = ttk.Combobox(master=f22, width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_tempture_rise_value, values=combobox_tempture_rise_values)
    combobox_tempture_rise.pack(side=tk.LEFT)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='局放要求:                       ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry_partialdischarge
    entry_partialdischarge = tk.Entry(f22, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(100 * w_ratio))
    entry_partialdischarge.pack(side=tk.LEFT)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='压力释放通道:                ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_gasduct
    combobox_gasduct_value = StringVar()
    combobox_gasduct_values = ['Top Chimneys', 'Side outlet', 'Rear outlet', '无', '其他']
    combobox_gasduct = ttk.Combobox(master=f22, width=int(18 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_gasduct_value, values=combobox_gasduct_values)
    combobox_gasduct.bind("<<ComboboxSelected>>", on_gasduct_combobox_select)
    combobox_gasduct.pack(side=tk.LEFT)

    global frame_gasduct
    frame_gasduct = tk.Frame(f22, bg="white", bd=0)
    tk.Label(frame_gasduct, text='    *', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(frame_gasduct, text='压力释放通道安装柜号:', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global entry_gasduct_panelnumber
    entry_gasduct_panelnumber = tk.Entry(frame_gasduct, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(30 * w_ratio))
    entry_gasduct_panelnumber.pack(side=tk.LEFT)

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='母线桥接口:                    ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_busduct_interface
    combobox_busduct_interface_value = StringVar()
    combobox_busduct_interface_values = ['需要（ABB桥）', '需要（客户桥）', '不需要']
    combobox_busduct_interface = ttk.Combobox(master=f22, width=int(18 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_busduct_interface_value, values=combobox_busduct_interface_values)
    combobox_busduct_interface.bind("<<ComboboxSelected>>", on_busduct_interface_combobox_select)
    combobox_busduct_interface.pack(side=tk.LEFT)

    global frame_busduct_interface
    frame_busduct_interface = tk.Frame(f22, bg="white", bd=0)

    frame_busduct_interface_Right = tk.Frame(frame_busduct_interface, bg="white", bd=0)
    frame_busduct_interface_Right.pack(side=tk.LEFT, fill=tk.X, padx=(int(10 * w_ratio), 0))

    frame_busduct_interface_Left = tk.Frame(frame_busduct_interface, bg="white", bd=0)
    frame_busduct_interface_Left.pack(side=tk.LEFT, fill=tk.X)

    global num_rows1
    num_rows1 = 3
    global num_columns1
    num_columns1 = 1
    global entry_widths1
    entry_widths1 = [int(20 * w_ratio)]
    global entries1
    entries1 = [[None for _ in range(num_columns1)] for _ in range(num_rows1)]

    for row in range(num_rows1):
        row_frame1 = tk.Frame(frame_busduct_interface_Right)
        row_frame1.pack(side=tk.TOP, fill=tk.X)

        for col in range(num_columns1):
            entry1 = tk.Entry(row_frame1, width=entry_widths1[col], relief='solid')
            entry1.pack(side=tk.LEFT)
            entries1[row][col] = entry1
            entry1.config(justify='center')
            if row == 0 and col == 0:
                entry1.insert(0, '柜号')
                entry1['state'] = 'disabled'

    add_row_button1 = tk.Button(frame_busduct_interface_Left, text="增加行", command=lambda: add_row1(content, frame_busduct_interface_Right, canvas), font=tku._ft(int(10 * h_ratio), False), activebackground='blue')
    add_row_button1.pack(side=tk.TOP, pady=(0, int(5 * h_ratio)), padx=(int(10 * w_ratio), int(10 * w_ratio)))

    delete_row_button1 = tk.Button(frame_busduct_interface_Left, text="删除行", command=lambda: delete_row1(content, frame_busduct_interface_Right, canvas), font=tku._ft(int(10 * h_ratio), False), activebackground='blue')
    delete_row_button1.pack(side=tk.TOP, pady=(int(5 * h_ratio), 0), padx=(int(10 * w_ratio), int(10 * w_ratio)))

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='反相序柜号:                    ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_phasesequence
    combobox_phasesequence_value = StringVar()
    combobox_phasesequence_values = ['有', '无']
    combobox_phasesequence = ttk.Combobox(master=f22, width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_phasesequence_value, values=combobox_phasesequence_values)
    combobox_phasesequence.bind("<<ComboboxSelected>>", on_phasesequence_combobox_select)
    combobox_phasesequence.pack(side=tk.LEFT)

    global frame_phasesequence
    frame_phasesequence = tk.Frame(f22, bg="white", bd=0)

    frame_phasesequence_Right = tk.Frame(frame_phasesequence, bg="white", bd=0)
    frame_phasesequence_Right.pack(side=tk.LEFT, fill=tk.X, padx=(int(10 * w_ratio), 0))

    frame_phasesequence_Left = tk.Frame(frame_phasesequence, bg="white", bd=0)
    frame_phasesequence_Left.pack(side=tk.LEFT, fill=tk.X)

    global phasesequence_items
    phasesequence_items = ['整柜反相序', '主开关上下侧相序不一致']
    global phasesequence_items_listbox
    phasesequence_items_listbox = tk.Listbox(frame_phasesequence_Right, height=2, selectmode='single')
    phasesequence_items_listbox.bind("<Double-Button-1>", typeinselection)

    global num_rows2
    num_rows2 = 2
    global num_columns2
    num_columns2 = 2
    global entry_widths2
    entry_widths2 = [int(20 * w_ratio), int(40 * w_ratio)]
    global entries2
    entries2 = [[None for _ in range(num_columns2)] for _ in range(num_rows2)]

    for row in range(num_rows2):
        row_frame2 = tk.Frame(frame_phasesequence_Right)
        row_frame2.pack(side=tk.TOP, fill=tk.X)

        for col in range(num_columns2):
            entry2 = tk.Entry(row_frame2, width=entry_widths2[col], relief='solid')
            entry2.pack(side=tk.LEFT)
            entries2[row][col] = entry2
            entry2.config(justify='center')
            if row == 0 and col == 0:
                entry2.insert(0, '柜号')
                entry2['state'] = 'disabled'
            elif row == 0 and col == 1:
                entry2.insert(0, '反相序类型')
                entry2['state'] = 'disabled'
            elif row != 0 and col == 1:
                entry2.bind("<Double-Button-1>", lambda event, row=row, col=col: show_phasesequence_items(event, row, col))

    add_row_button2 = tk.Button(frame_phasesequence_Left, text="增加行", command=lambda: add_row2(content, frame_phasesequence_Right, canvas), font=tku._ft(int(10 * h_ratio), False), activebackground='blue')
    add_row_button2.pack(side=tk.TOP, pady=(0, int(5 * h_ratio)), padx=(int(10 * w_ratio), int(10 * w_ratio)))

    delete_row_button2 = tk.Button(frame_phasesequence_Left, text="删除行", command=lambda: delete_row2(content, frame_phasesequence_Right, canvas), font=tku._ft(int(10 * h_ratio), False), activebackground='blue')
    delete_row_button2.pack(side=tk.TOP, pady=(int(5 * h_ratio), 0), padx=(int(10 * w_ratio), int(10 * w_ratio)))

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='相序标识字母:                 ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global entry_phasemarkletter
    entry_phasemarkletter = tk.Entry(f22, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(10 * w_ratio))
    entry_phasemarkletter.pack(side=tk.LEFT)

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='相序标识颜色:                 ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global entry_phasemarkcolour
    entry_phasemarkcolour = tk.Entry(f22, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(10 * w_ratio))
    entry_phasemarkcolour.pack(side=tk.LEFT)

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='风机:                              ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_fan
    combobox_fan_value = StringVar()
    combobox_fan_values = ['无', '有']
    combobox_fan = ttk.Combobox(master=f22, width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_fan_value, values=combobox_fan_values)
    combobox_fan.bind("<<ComboboxSelected>>", on_fan_combobox_select)
    combobox_fan.pack(side=tk.LEFT)

    global frame_fan
    frame_fan = tk.Frame(f22, bg="white", bd=0)

    frame_fan_Right = tk.Frame(frame_fan, bg="white", bd=0)
    frame_fan_Right.pack(side=tk.LEFT, fill=tk.X, padx=(int(10 * w_ratio), 0))

    frame_fan_Left = tk.Frame(frame_fan, bg="white", bd=0)
    frame_fan_Left.pack(side=tk.LEFT, fill=tk.X)

    global num_rows3
    num_rows3 = 2
    global num_columns3
    num_columns3 = 2
    global entry_widths3
    entry_widths3 = [int(20 * w_ratio), int(30 * w_ratio)]
    global entries3
    entries3 = [[None for _ in range(num_columns3)] for _ in range(num_rows3)]

    for row in range(num_rows3):
        row_frame3 = tk.Frame(frame_fan_Right)
        row_frame3.pack(side=tk.TOP, fill=tk.X)

        for col in range(num_columns3):
            entry3 = tk.Entry(row_frame3, width=entry_widths3[col], relief='solid')
            entry3.pack(side=tk.LEFT)
            entries3[row][col] = entry3
            entry3.config(justify='center')
            if row == 0 and col == 0:
                entry3.insert(0, '柜号')
                entry3['state'] = 'disabled'
            elif row == 0 and col == 1:
                entry3.insert(0, '风机电源（注明VDC/VAC）')
                entry3['state'] = 'disabled'

    add_row_button3 = tk.Button(frame_fan_Left, text="增加行", command=lambda: add_row3(content, frame_fan_Right, canvas), font=tku._ft(int(10 * h_ratio), False), activebackground='blue')
    add_row_button3.pack(side=tk.TOP, pady=(0, int(5 * h_ratio)), padx=(int(10 * w_ratio), int(10 * w_ratio)))

    delete_row_button3 = tk.Button(frame_fan_Left, text="删除行", command=lambda: delete_row3(content, frame_fan_Right, canvas), font=tku._ft(int(10 * h_ratio), False), activebackground='blue')
    delete_row_button3.pack(side=tk.TOP, pady=(int(5 * h_ratio), 0), padx=(int(10 * w_ratio), int(10 * w_ratio)))

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='计量柜:                          ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_meterpanel
    combobox_meterpanel_value = StringVar()
    combobox_meterpanel_values = ['无', '有']
    combobox_meterpanel = ttk.Combobox(master=f22, width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_meterpanel_value, values=combobox_meterpanel_values)
    combobox_meterpanel.bind("<<ComboboxSelected>>", on_meterpanel_combobox_select)
    combobox_meterpanel.pack(side=tk.LEFT)

    global frame_meterpanel
    frame_meterpanel = tk.Frame(f22, bg="white", bd=0)

    frame_meterpanel_Right = tk.Frame(frame_meterpanel, bg="white", bd=0)
    frame_meterpanel_Right.pack(side=tk.LEFT, fill=tk.X, padx=(int(10 * w_ratio), 0))

    frame_meterpanel_Left = tk.Frame(frame_meterpanel, bg="white", bd=0)
    frame_meterpanel_Left.pack(side=tk.LEFT, fill=tk.X)

    global meterpanelseal_items
    meterpanelseal_items = ['无铅封', '低压室门', '断路器室门', '低压室门+断路器室门']
    global meterpanelseal_items_listbox
    meterpanelseal_items_listbox = tk.Listbox(frame_meterpanel_Right, height=4, selectmode='single')
    meterpanelseal_items_listbox.bind("<Double-Button-1>", meterpanelseal_typeinselection)

    global meterpanellighting_items
    meterpanellighting_items = ['无', '有']
    global meterpanellighting_items_listbox
    meterpanellighting_items_listbox = tk.Listbox(frame_meterpanel_Right, height=2, selectmode='single')
    meterpanellighting_items_listbox.bind("<Double-Button-1>", meterpanellighting_typeinselection)

    global meterpanelmicroswitch_items
    meterpanelmicroswitch_items = ['断路器室', '电缆室', '后门/盖', '断路器室+电缆室', '断路器室+后门/后盖', '电缆室+后门/盖', '断路器室+电缆室+后门/盖']
    global meterpanelmicroswitch_items_listbox
    meterpanelmicroswitch_items_listbox = tk.Listbox(frame_meterpanel_Right, height=7, selectmode='single')
    meterpanelmicroswitch_items_listbox.bind("<Double-Button-1>", meterpanelmicroswitch_typeinselection)

    global meterpaneltruckblockmagnet_items
    meterpaneltruckblockmagnet_items = ['无', '有']
    global meterpaneltruckblockmagnet_items_listbox
    meterpaneltruckblockmagnet_items_listbox = tk.Listbox(frame_meterpanel_Right, height=2, selectmode='single')
    meterpaneltruckblockmagnet_items_listbox.bind("<Double-Button-1>", meterpaneltruckblockmagnet_typeinselection)

    global meterpanelopenoperation_items
    meterpanelopenoperation_items = ['是', '否']
    global meterpanelopenoperation_items_listbox
    meterpanelopenoperation_items_listbox = tk.Listbox(frame_meterpanel_Right, height=2, selectmode='single')
    meterpanelopenoperation_items_listbox.bind("<Double-Button-1>", meterpanelopenoperation_typeinselection)

    global meterpanelopenwindowsize_items
    meterpanelopenwindowsize_items = ['一次指定', 'IE指定']
    global meterpanelopenwindowsize_items_listbox
    meterpanelopenwindowsize_items_listbox = tk.Listbox(frame_meterpanel_Right, height=2, selectmode='single')
    meterpanelopenwindowsize_items_listbox.bind("<Double-Button-1>", meterpanelopenwindowsize_typeinselection)

    global num_rows4
    num_rows4 = 2
    global num_columns4
    num_columns4 = 7
    global entry_widths4
    entry_widths4 = [int(20 * w_ratio), int(25 * w_ratio), int(15 * w_ratio), int(35 * w_ratio), int(20 * w_ratio), int(10 * w_ratio), int(10 * w_ratio)]
    global entries4
    entries4 = [[None for _ in range(num_columns4)] for _ in range(num_rows4)]

    for row in range(num_rows4):
        row_frame4 = tk.Frame(frame_meterpanel_Right)
        row_frame4.pack(side=tk.TOP, fill=tk.X)

        for col in range(num_columns4):
            entry4 = tk.Entry(row_frame4, width=entry_widths4[col], relief='solid')
            entry4.pack(side=tk.LEFT)
            entries4[row][col] = entry4
            entry4.config(justify='center')
            if row == 0 and col == 0:
                entry4.insert(0, '柜号')
                entry4['state'] = 'disabled'
            elif row == 0 and col == 1:
                entry4.insert(0, '铅封')
                entry4['state'] = 'disabled'
            elif row == 0 and col == 2:
                entry4.insert(0, '低压室照明')
                entry4['state'] = 'disabled'
            elif row == 0 and col == 3:
                entry4.insert(0, '微动开关（低压室默认有）')
                entry4['state'] = 'disabled'
            elif row == 0 and col == 4:
                entry4.insert(0, '手车闭锁电磁铁')
                entry4['state'] = 'disabled'
            elif row == 0 and col == 5:
                entry4.insert(0, '开门操作')
                entry4['state'] = 'disabled'
            elif row == 0 and col == 6:
                entry4.insert(0, '开窗尺寸')
                entry4['state'] = 'disabled'
            elif row != 0 and col == 1:
                entry4.bind("<Button-1>", lambda event, row=row, col=col: show_meterpanelseal_items(event, row, col))
            elif row != 0 and col == 2:
                entry4.bind("<Button-1>", lambda event, row=row, col=col: show_meterpanellighting_items(event, row, col))
            elif row != 0 and col == 3:
                entry4.bind("<Button-1>", lambda event, row=row, col=col: show_meterpanelmicroswitch_items(event, row, col))
            elif row != 0 and col == 4:
                entry4.bind("<Button-1>", lambda event, row=row, col=col: show_meterpaneltruckblockmagnet_items(event, row, col))
            elif row != 0 and col == 5:
                entry4.bind("<Button-1>", lambda event, row=row, col=col: show_meterpanelopenoperation_items(event, row, col))
            elif row != 0 and col == 6:
                entry4.bind("<Button-1>", lambda event, row=row, col=col: show_meterpanelopenwindowsize_items(event, row, col))

    add_row_button4 = tk.Button(frame_meterpanel_Left, text="增加行", command=lambda: add_row4(content, frame_meterpanel_Right, canvas), font=tku._ft(int(10 * h_ratio), False), activebackground='blue')
    add_row_button4.pack(side=tk.TOP, pady=(0, int(5 * h_ratio)), padx=(int(10 * w_ratio), int(10 * w_ratio)))

    delete_row_button4 = tk.Button(frame_meterpanel_Left, text="删除行", command=lambda: delete_row4(content, frame_meterpanel_Right, canvas), font=tku._ft(int(10 * h_ratio), False), activebackground='blue')
    delete_row_button4.pack(side=tk.TOP, pady=(int(5 * h_ratio), 0), padx=(int(10 * w_ratio), int(10 * w_ratio)))

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='电缆室照明:                    ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_calighting
    combobox_calighting_value = StringVar()
    combobox_calighting_values = ['有', '无']
    combobox_calighting = ttk.Combobox(master=f22, width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_calighting_value, values=combobox_calighting_values)
    combobox_calighting.pack(side=tk.LEFT)
    tk.Label(f22, text='（若有，ZS1电缆底进底出2个，UG550电缆底进底出1个）', bg="white", fg="red", font=tku._ft(int(10 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='低压室照明:                    ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_lvlighting
    combobox_lvlighting_value = StringVar()
    combobox_lvlighting_values = ['有', '无']
    combobox_lvlighting = ttk.Combobox(master=f22, width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_lvlighting_value, values=combobox_lvlighting_values)
    combobox_lvlighting.pack(side=tk.LEFT)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='加热器:                          ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_heater
    combobox_heater_value = StringVar()
    combobox_heater_values = ['无', '电缆室（国内标准）', '电缆室+断路器室（国内特殊，国外标准）']
    combobox_heater = ttk.Combobox(master=f22, width=int(30 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_heater_value, values=combobox_heater_values)
    combobox_heater.pack(side=tk.LEFT)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='活门剪刀叉闭锁Fail Safe: ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_failsafe
    combobox_failsafe_value = StringVar()
    combobox_failsafe_values = ['有', '无']
    combobox_failsafe = ttk.Combobox(master=f22, width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_failsafe_value, values=combobox_failsafe_values)
    combobox_failsafe.pack(side=tk.LEFT)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='低压室门类型:                 ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_lvdoortype
    combobox_lvdoortype_value = StringVar()
    combobox_lvdoortype_values = ['平门', '微凸门', '凸门']
    combobox_lvdoortype = ttk.Combobox(master=f22, width=int(8 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_lvdoortype_value, values=combobox_lvdoortype_values)
    combobox_lvdoortype.pack(side=tk.LEFT)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='低压室类型:                    ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_lvctype
    combobox_lvctype_value = StringVar()
    combobox_lvctype_values = ['新低压室（覆铝锌板+导轨安装）', '旧低压室（网格板安装）']
    combobox_lvctype = ttk.Combobox(master=f22, width=int(40 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_lvctype_value, values=combobox_lvctype_values)
    combobox_lvctype.pack(side=tk.LEFT)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='低压室高度:                    ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global entry_lvcheight
    entry_lvcheight = tk.Entry(f22, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(8 * w_ratio))
    entry_lvcheight.pack(side=tk.LEFT)

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='低压室特殊二次元器件:    ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_secondarycomponent
    combobox_secondarycomponent_value = StringVar()
    combobox_secondarycomponent_values = ['有', '无']
    combobox_secondarycomponent = ttk.Combobox(master=f22, width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_secondarycomponent_value, values=combobox_secondarycomponent_values)
    combobox_secondarycomponent.bind("<<ComboboxSelected>>", on_secondarycomponent_combobox_select)
    combobox_secondarycomponent.pack(side=tk.LEFT, padx=(0, int(10 * w_ratio)))

    global frame_secondarycomponent
    frame_secondarycomponent = tk.Frame(f22, bg="white", bd=0)

    global checkbox_secondarycomponent
    global secondarycomponents
    secondarycomponents = ['542保护', 'SUE3000', '特殊铭牌']
    global checkbox_secondarycomponent_list
    checkbox_secondarycomponent_list = []
    for i in range(len(secondarycomponents)):
        v = IntVar()
        checkbox_secondarycomponent = tk.Checkbutton(frame_secondarycomponent, command=select_secondarycomponent, text=secondarycomponents[i], variable=v, font=tku._ft(int(13 * h_ratio), False), height=1, background='#FFFFFF')
        checkbox_secondarycomponent.pack(side=tk.TOP, anchor=tk.W, expand=True)
        checkbox_secondarycomponent_list.append(v)

    tk.Label(frame_secondarycomponent, text='(器件汇总，其他特殊请填入下框）', bg="white", fg="red", font=tku._ft(int(10 * h_ratio), False), justify='left').pack(side=tk.TOP, anchor=tk.W, expand=True)
    global entry_select_secondarycomponent
    entry_select_secondarycomponent = tk.Entry(frame_secondarycomponent, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(80 * w_ratio))
    entry_select_secondarycomponent.pack(side=tk.TOP, anchor=tk.W, expand=True)

    tk.Label(frame_secondarycomponent, text='(安装柜号，一次需要安装支架）', bg="white", fg="red", font=tku._ft(int(10 * h_ratio), False), justify='left').pack(side=tk.TOP, anchor=tk.W, expand=True)
    global entry_secondarycomponent_panelnumber
    entry_secondarycomponent_panelnumber = tk.Entry(frame_secondarycomponent, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(30 * w_ratio))
    entry_secondarycomponent_panelnumber.pack(side=tk.TOP, anchor=tk.W, expand=True)

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    # f22 = tk.Frame(content, bg="white", bd=0)
    # tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    # tk.Label(f22, text='电流互感器供应商:           ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    #
    # global combobox_ctcompany
    # combobox_ctcompany_value = StringVar()
    # combobox_ctcompany_values = ['ABB', 'DYH', 'NTK', 'TLEP']
    # combobox_ctcompany = ttk.Combobox(master=f22, width=int(8 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_ctcompany_value, values=combobox_ctcompany_values)
    # combobox_ctcompany.pack(side=tk.LEFT)
    # f22.pack(fill=tk.X)
    #
    # tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线
    #
    # f22 = tk.Frame(content, bg="white", bd=0)
    # tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    # tk.Label(f22, text='电压互感器供应商:           ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    #
    # global combobox_ptcompany
    # combobox_ptcompany_value = StringVar()
    # combobox_ptcompany_values = ['ABB', 'DYH', 'NTK', 'TLEP']
    # combobox_ptcompany = ttk.Combobox(master=f22, width=int(8 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_ptcompany_value, values=combobox_ptcompany_values)
    # combobox_ptcompany.pack(side=tk.LEFT)
    # f22.pack(fill=tk.X)
    #
    # tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='避雷器供应商:                 ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_surgearrestcompany
    combobox_surgearrestcompany_value = StringVar()
    combobox_surgearrestcompany_values = ['神电', '日立', 'GCA']
    combobox_surgearrestcompany = ttk.Combobox(master=f22, width=int(8 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_surgearrestcompany_value, values=combobox_surgearrestcompany_values)
    combobox_surgearrestcompany.pack(side=tk.LEFT)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='内部燃弧保护装置:           ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_faultlimitingdevice
    combobox_faultlimitingdevice_value = StringVar()
    combobox_faultlimitingdevice_values = ['有', '无']
    combobox_faultlimitingdevice = ttk.Combobox(master=f22, width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_faultlimitingdevice_value, values=combobox_faultlimitingdevice_values)
    combobox_faultlimitingdevice.bind("<<ComboboxSelected>>", on_faultlimitingdevice_combobox_select)
    combobox_faultlimitingdevice.pack(side=tk.LEFT, padx=(0, int(10 * w_ratio)))

    global frame_faultlimitingdevice
    frame_faultlimitingdevice = tk.Frame(f22, bg="white", bd=0)


    global combobox_faultlimitingdevice2
    combobox_faultlimitingdevice2_value = StringVar()
    combobox_faultlimitingdevice2_values = ['Ith Limiter（压力释放微动开关）', 'Arc protection in IED', 'REA弧光保护', 'TVOC', 'UFES']
    combobox_faultlimitingdevice2 = ttk.Combobox(master=frame_faultlimitingdevice, width=int(40 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_faultlimitingdevice2_value, values=combobox_faultlimitingdevice2_values)
    combobox_faultlimitingdevice2.bind("<<ComboboxSelected>>", on_faultlimitingdevice2_combobox_select)
    combobox_faultlimitingdevice2.pack(side=tk.LEFT, padx=(0, int(10 * w_ratio)))

    global frame_faultlimitingdevice2
    frame_faultlimitingdevice2 = tk.Frame(f22, bg="white", bd=0)

    global label_faultlimitingdevice2_location
    label_faultlimitingdevice2_location = tk.Label(frame_faultlimitingdevice2, text='(安装位置）', bg="white", fg="red", font=tku._ft(int(10 * h_ratio), False), justify='left')
    label_faultlimitingdevice2_location.pack(side=tk.TOP, anchor=tk.W, expand=True)

    global combobox_faultlimitingdevice2_location
    combobox_faultlimitingdevice2_location_value = StringVar()
    combobox_faultlimitingdevice2_location_values = ['母线室', '断路器室', '电缆室']
    combobox_faultlimitingdevice2_location = ttk.Combobox(master=frame_faultlimitingdevice2, width=int(10 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_faultlimitingdevice2_location_value, values=combobox_faultlimitingdevice2_location_values)
    combobox_faultlimitingdevice2_location.pack(side=tk.TOP, anchor=tk.W, expand=True)

    global label_faultlimitingdevice2_panelnumber
    label_faultlimitingdevice2_panelnumber = tk.Label(frame_faultlimitingdevice2, text='(安装柜号）', bg="white", fg="red", font=tku._ft(int(10 * h_ratio), False), justify='left')
    label_faultlimitingdevice2_panelnumber.pack(side=tk.TOP, anchor=tk.W, expand=True)
    global entry_faultlimitingdevice2_panelnumber
    entry_faultlimitingdevice2_panelnumber = tk.Entry(frame_faultlimitingdevice2, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(30 * w_ratio))
    entry_faultlimitingdevice2_panelnumber.pack(side=tk.TOP, anchor=tk.W, expand=True)

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='电动操作:                        ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_electricaloperation
    combobox_electricaloperation_value = StringVar()
    combobox_electricaloperation_values = ['有', '无']
    combobox_electricaloperation = ttk.Combobox(master=f22, width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_electricaloperation_value, values=combobox_electricaloperation_values)
    combobox_electricaloperation.bind("<<ComboboxSelected>>", on_electricaloperation_combobox_select)
    combobox_electricaloperation.pack(side=tk.LEFT, padx=(0, int(10 * w_ratio)))

    global frame_electricaloperation
    frame_electricaloperation = tk.Frame(f22, bg="white", bd=0)

    global combobox_electricaloperation_type
    combobox_electricaloperation_type_value = StringVar()
    combobox_electricaloperation_type_values = ['断路器手车电操（若带机械锁，则锁安装在侧板，并带微动）', '接地开关电操', '断路器手车电操+接地开关电操']
    combobox_electricaloperation_type = ttk.Combobox(master=frame_electricaloperation, width=int(70 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_electricaloperation_type_value, values=combobox_electricaloperation_type_values)
    combobox_electricaloperation_type.pack(side=tk.TOP, anchor=tk.W, expand=True)

    tk.Label(frame_electricaloperation, text='（电操电源，注明VDC/VAC）', bg="white", fg="red", font=tku._ft(int(10 * h_ratio), False), justify='left').pack(side=tk.TOP, anchor=tk.W, expand=True)
    global entry_frame_electricaloperation_voltage
    entry_frame_electricaloperation_voltage = tk.Entry(frame_electricaloperation, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(30 * w_ratio))
    entry_frame_electricaloperation_voltage.pack(side=tk.TOP, anchor=tk.W, expand=True)

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='手车测温方式（500）:     ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_trucktempturemeasure
    combobox_trucktempturemeasure_value = StringVar()
    combobox_trucktempturemeasure_values = ['无', '铜管测温', '触臂测温']
    combobox_trucktempturemeasure = ttk.Combobox(master=f22, width=int(10 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_trucktempturemeasure_value, values=combobox_trucktempturemeasure_values)
    combobox_trucktempturemeasure.pack(side=tk.LEFT)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='温湿度控制器:                  ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_temptureandhumiditycontroller
    combobox_temptureandhumiditycontroller_value = StringVar()
    combobox_temptureandhumiditycontroller_values = ['无', '百岗CD10-R，传感器在电缆室', '立林WS-7100-1，传感器在电缆室', '立林WS-7100，传感器在电缆室+断路器室']
    combobox_temptureandhumiditycontroller = ttk.Combobox(master=f22, width=int(50 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_temptureandhumiditycontroller_value, values=combobox_temptureandhumiditycontroller_values)
    combobox_temptureandhumiditycontroller.pack(side=tk.LEFT)
    tk.Label(f22, text='（若有，一次需要安装传感器支架）', bg="white", fg="red", font=tku._ft(int(10 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='CB室门紧急合分闸按钮:    ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_cbdooremergencybutton
    combobox_cbdooremergencybutton_value = StringVar()
    combobox_cbdooremergencybutton_values = ['紧急分闸', '紧急合闸+分闸（要配断路器合闸闭锁电磁铁RL1）']
    combobox_cbdooremergencybutton = ttk.Combobox(master=f22, width=int(56 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_cbdooremergencybutton_value, values=combobox_cbdooremergencybutton_values)
    combobox_cbdooremergencybutton.pack(side=tk.LEFT)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='CB室门微动:                    ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_cbdoormicroswitch
    combobox_cbdoormicroswitch_value = StringVar()
    combobox_cbdoormicroswitch_values = ['无', '有']
    combobox_cbdoormicroswitch = ttk.Combobox(master=f22, width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_cbdoormicroswitch_value, values=combobox_cbdoormicroswitch_values)
    combobox_cbdoormicroswitch.pack(side=tk.LEFT)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='CB室航空插闭锁:             ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_cbplugblock
    combobox_cbplugblock_value = StringVar()
    combobox_cbplugblock_values = ['无（国内标准）', '有（国内特殊，国外标准）']
    combobox_cbplugblock = ttk.Combobox(master=f22, width=int(30 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_cbplugblock_value, values=combobox_cbplugblock_values)
    combobox_cbplugblock.pack(side=tk.LEFT)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='电缆夹具:                        ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_cableclamp
    combobox_cableclamp_value = StringVar()
    combobox_cableclamp_values = ['电缆夹', '不锈钢扎带']
    combobox_cableclamp = ttk.Combobox(master=f22, width=int(16 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_cableclamp_value, values=combobox_cableclamp_values)
    combobox_cableclamp.pack(side=tk.LEFT)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='手车接地方式:                  ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_truckearthing
    combobox_truckearthing_value = StringVar()
    combobox_truckearthing_values = ['轮接地', '接地夹接地']
    combobox_truckearthing = ttk.Combobox(master=f22, width=int(16 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_truckearthing_value, values=combobox_truckearthing_values)
    combobox_truckearthing.pack(side=tk.LEFT)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='母线搭接面镀银:               ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_surface
    combobox_surface_value = StringVar()
    combobox_surface_values = ['不镀银', '主母线镀银', '分支母线镀银', '主母线镀银+分支母线镀银']
    combobox_surface = ttk.Combobox(master=f22, width=int(30 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_surface_value, values=combobox_surface_values)
    combobox_surface.pack(side=tk.LEFT)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='母线热缩套:                     ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_heatshrinkable
    combobox_heatshrinkable_value = StringVar()
    combobox_heatshrinkable_values = ['纯ZS1（复合绝缘）：主+分支母线热缩', '纯500/Beni（非复合绝缘）：无热缩', '纯500/Beni（复合绝缘）：主+分支母线热缩', 'ZS1与500/Beni拼柜（非复合绝缘）：仅ZS1上分支母线热缩', 'ZS1与500/Beni拼柜（复合绝缘）：主+分支母线热缩']
    combobox_heatshrinkable = ttk.Combobox(master=f22, width=int(70 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_heatshrinkable_value, values=combobox_heatshrinkable_values)
    combobox_heatshrinkable.pack(side=tk.LEFT)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='二次电缆孔:                     ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_cablehole
    combobox_cablehole_value = StringVar()
    combobox_cablehole_values = ['仅左侧', '仅右侧', '左右两侧', '不需要']
    combobox_cablehole = ttk.Combobox(master=f22, width=int(70 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_cablehole_value, values=combobox_cablehole_values)
    combobox_cablehole.pack(side=tk.LEFT)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='地坪高度:                        ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global entry_floorheight
    entry_floorheight = tk.Entry(f22, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(10 * w_ratio))
    entry_floorheight.pack(side=tk.LEFT)

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='断路器是否特殊:              ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_speccb
    combobox_speccb_value = StringVar()
    combobox_speccb_values = ['是', '否']
    combobox_speccb = ttk.Combobox(master=f22, width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_speccb_value, values=combobox_speccb_values)
    combobox_speccb.bind("<<ComboboxSelected>>", on_speccb_combobox_select)
    combobox_speccb.pack(side=tk.LEFT, padx=(0, int(10 * w_ratio)))

    global frame_speccb
    frame_speccb = tk.Frame(f22, bg="white", bd=0)

    global checkbox_speccb
    global speccbs
    speccbs = ['1.1倍温升', 'PT/HE极柱', 'Unification']
    global checkbox_speccb_list
    checkbox_speccb_list = []
    for i in range(len(speccbs)):
        v = IntVar()
        checkbox_speccb = tk.Checkbutton(frame_speccb, command=select_speccb, text=speccbs[i], variable=v, font=tku._ft(int(13 * h_ratio), False), height=1, background='#FFFFFF')
        checkbox_speccb.pack(side=tk.TOP, anchor=tk.W, expand=True)
        checkbox_speccb_list.append(v)

    tk.Label(frame_speccb, text='(汇总，其他特殊请填入下框）', bg="white", fg="red", font=tku._ft(int(10 * h_ratio), False), justify='left').pack(side=tk.TOP, anchor=tk.W, expand=True)
    global entry_select_speccb
    entry_select_speccb = tk.Entry(frame_speccb, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(80 * w_ratio))
    entry_select_speccb.pack(side=tk.TOP, anchor=tk.W, expand=True)

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='一次铭牌开孔:                  ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_nameplateopenhole
    combobox_nameplateopenhole_value = StringVar()
    combobox_nameplateopenhole_values = ['无', '有']
    combobox_nameplateopenhole = ttk.Combobox(master=f22, width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_nameplateopenhole_value, values=combobox_nameplateopenhole_values)
    combobox_nameplateopenhole.bind("<<ComboboxSelected>>", on_nameplateopenhole_combobox_select)
    combobox_nameplateopenhole.pack(side=tk.LEFT)

    global frame_nameplateopenhole
    frame_nameplateopenhole = tk.Frame(f22, bg="white", bd=0)
    tk.Label(frame_nameplateopenhole, text='    *', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(frame_nameplateopenhole, text='安装位置:', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global entry_nameplateopenhole_panelnumber
    entry_nameplateopenhole_panelnumber = tk.Entry(frame_nameplateopenhole, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(40 * w_ratio))
    entry_nameplateopenhole_panelnumber.pack(side=tk.LEFT)

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='眉头铭牌:                        ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_brownameplate
    combobox_brownameplate_value = StringVar()
    combobox_brownameplate_values = ['无', '有']
    combobox_brownameplate = ttk.Combobox(master=f22, width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_brownameplate_value, values=combobox_brownameplate_values)
    combobox_brownameplate.bind("<<ComboboxSelected>>", on_brownameplate_combobox_select)
    combobox_brownameplate.pack(side=tk.LEFT)

    global frame_brownameplate
    frame_brownameplate = tk.Frame(f22, bg="white", bd=0)
    tk.Label(frame_brownameplate, text='    *', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(frame_brownameplate, text='安装尺寸:', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global entry_brownameplate_panelnumber
    entry_brownameplate_panelnumber = tk.Entry(frame_brownameplate, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(40 * w_ratio))
    entry_brownameplate_panelnumber.pack(side=tk.LEFT)

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='特殊铭牌:                        ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_specnameplate
    combobox_specnameplate_value = StringVar()
    combobox_specnameplate_values = ['无', '有']
    combobox_specnameplate = ttk.Combobox(master=f22, width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_specnameplate_value, values=combobox_specnameplate_values)
    combobox_specnameplate.bind("<<ComboboxSelected>>", on_specnameplate_combobox_select)
    combobox_specnameplate.pack(side=tk.LEFT)

    global frame_specnameplate
    frame_specnameplate = tk.Frame(f22, bg="white", bd=0)
    tk.Label(frame_specnameplate, text='    *', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(frame_specnameplate, text='材质等特殊信息:', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global entry_specnameplate_panelnumber
    entry_specnameplate_panelnumber = tk.Entry(frame_specnameplate, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(40 * w_ratio))
    entry_specnameplate_panelnumber.pack(side=tk.LEFT)

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='二次导线类型:                  ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_secondarywiring
    combobox_secondarywiring_value = StringVar()
    combobox_secondarywiring_values = ['普通PVC导线', '低烟无卤阻燃导线']
    combobox_secondarywiring = ttk.Combobox(master=f22, width=int(20 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_secondarywiring_value, values=combobox_secondarywiring_values)
    combobox_secondarywiring.pack(side=tk.LEFT)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='柜体/主元件特殊接地:       ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global entry_specearthing
    entry_specearthing = tk.Entry(f22, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(70 * w_ratio))
    entry_specearthing.pack(side=tk.LEFT)
    tk.Label(f22, text='（门/盖、活门、地刀、手车、CT、PT、SA、CB）', bg="white", fg="red", font=tku._ft(int(10 * h_ratio), False), justify='left').pack(side=tk.LEFT)

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='低压箱特殊接地:               ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global entry_lvcspecearthing
    entry_lvcspecearthing = tk.Entry(f22, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(70 * w_ratio))
    entry_lvcspecearthing.pack(side=tk.LEFT)
    tk.Label(f22, text='（低压箱、低压元件、通讯、端子）', bg="white", fg="red", font=tku._ft(int(10 * h_ratio), False), justify='left').pack(side=tk.LEFT)

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='备品备件及附件:               ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_accdevice
    combobox_accdevice_value = StringVar()
    combobox_accdevice_values = ['以技术协议为准', '以商务合同为准，PM提供清单']
    combobox_accdevice = ttk.Combobox(master=f22, width=int(36 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_accdevice_value, values=combobox_accdevice_values)
    combobox_accdevice.pack(side=tk.LEFT)
    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='电缆参数:                        ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global entry_cableparameter
    entry_cableparameter = tk.Entry(f22, bg="white", font=tku._ft(int(13 * h_ratio), False), width=int(10 * w_ratio))
    entry_cableparameter.pack(side=tk.LEFT)
    entry_cableparameter.insert(0, '见单线图')

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(10 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='接地开关:                        ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    f2L = tk.Frame(f22, bg="white", bd=0)
    tk.Frame(f2L, bg="white", bd=0).pack(side=tk.LEFT, fill=tk.X, padx=0)

    f2L_Right = tk.Frame(f2L, bg="white", bd=0)
    f2L_Right.pack(side=tk.LEFT, fill=tk.X)

    f2L_Left = tk.Frame(f2L, bg="white", bd=0)
    f2L_Left.pack(side=tk.LEFT, fill=tk.X)

    global num_rows11
    num_rows11 = 1
    global num_columns11
    num_columns11 = 4
    global entry_widths11
    entry_widths11 = [int(10 * w_ratio), int(10 * w_ratio), int(100 * w_ratio), int(10 * w_ratio)]
    global entries11
    entries11 = [[None for _ in range(num_columns11)] for _ in range(num_rows11)]

    for row in range(num_rows11):
        row_frame11 = tk.Frame(f2L_Right)
        row_frame11.pack(side=tk.TOP, fill=tk.X)

        for col in range(num_columns11):
            entry11 = tk.Entry(row_frame11, width=entry_widths11[col], relief='solid')
            entry11.pack(side=tk.LEFT)
            entries11[row][col] = entry11
            entry11.config(justify='center')
            if row == 0 and col == 0:
                entry11.insert(0, '开关段')
                entry11['state'] = 'disabled'
            elif row == 0 and col == 1:
                entry11.insert(0, '柜型')
                entry11['state'] = 'disabled'
            elif row == 0 and col == 2:
                entry11.insert(0, '柜号')
                entry11['state'] = 'disabled'
            elif row == 0 and col == 3:
                entry11.insert(0, '型号')
                entry11['state'] = 'disabled'

    add_row_button11 = tk.Button(f2L_Left, text="增加行", command=lambda: add_row11(content, f2L_Right, canvas), font=tku._ft(int(10 * h_ratio), False), activebackground='blue')
    add_row_button11.pack(side=tk.TOP, pady=(0, int(10 * h_ratio)), padx=(int(10 * w_ratio), int(10 * w_ratio)))

    delete_row_button11 = tk.Button(f2L_Left, text="删除行", command=lambda: delete_row11(content, f2L_Right, canvas), font=tku._ft(int(10 * h_ratio), False), activebackground='blue')
    delete_row_button11.pack(side=tk.TOP, pady=(int(10 * h_ratio), 0), padx=(int(10 * w_ratio), int(10 * w_ratio)))

    f2L.pack(fill=tk.X)

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(20 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="white", bd=0)
    tk.Label(f22, text='*', bg="white", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='零序电流互感器:               ', bg="white", fg="black", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_zeroct
    combobox_zeroct_value = StringVar()
    combobox_zeroct_values = ['无', '有']
    combobox_zeroct = ttk.Combobox(master=f22, width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_zeroct_value, values=combobox_zeroct_values)
    combobox_zeroct.bind("<<ComboboxSelected>>", on_zeroct_combobox_select)
    combobox_zeroct.pack(side=tk.LEFT)

    global frame_zeroct
    frame_zeroct = tk.Frame(f22, bg="white", bd=0)

    frame_zeroct_Right = tk.Frame(frame_zeroct, bg="white", bd=0)
    frame_zeroct_Right.pack(side=tk.LEFT, fill=tk.X, padx=(int(10 * w_ratio), 0))

    frame_zeroct_Left = tk.Frame(frame_zeroct, bg="white", bd=0)
    frame_zeroct_Left.pack(side=tk.LEFT, fill=tk.X)

    global zeroctcompany_items
    zeroctcompany_items = ['ABB', 'DYH', 'NTK', '北京科伟达']
    global zeroctcompany_items_listbox
    zeroctcompany_items_listbox = tk.Listbox(frame_zeroct_Right, height=4, selectmode='single')
    zeroctcompany_items_listbox.bind("<Double-Button-1>", zeroctcompany_typeinselection)

    global zerocttype_items
    zerocttype_items = []
    global zerocttype_items_listbox
    zerocttype_items_listbox = tk.Listbox(frame_zeroct_Right, height=2, selectmode='single')
    zerocttype_items_listbox.bind("<Double-Button-1>", zerocttype_typeinselection)

    global zeroctinstallation_items
    zeroctinstallation_items = ['柜内', '柜外', '电缆延伸箱']
    global zeroctinstallation_items_listbox
    zeroctinstallation_items_listbox = tk.Listbox(frame_zeroct_Right, height=3, selectmode='single')
    zeroctinstallation_items_listbox.bind("<Double-Button-1>", zeroctinstallation_typeinselection)

    global num_rows5
    num_rows5 = 2
    global num_columns5
    num_columns5 = 6
    global entry_widths5
    entry_widths5 = [int(10 * w_ratio), int(10 * w_ratio), int(70 * w_ratio), int(14 * w_ratio), int(24 * w_ratio), int(14 * w_ratio)]
    global entries5
    entries5 = [[None for _ in range(num_columns5)] for _ in range(num_rows5)]

    for row in range(num_rows5):
        row_frame5 = tk.Frame(frame_zeroct_Right)
        row_frame5.pack(side=tk.TOP, fill=tk.X)

        for col in range(num_columns5):
            entry5 = tk.Entry(row_frame5, width=entry_widths5[col], relief='solid')
            entry5.pack(side=tk.LEFT)
            entries5[row][col] = entry5
            entry5.config(justify='center')
            if row == 0 and col == 0:
                entry5.insert(0, '开关段')
                entry5['state'] = 'disabled'
            elif row == 0 and col == 1:
                entry5.insert(0, '柜型')
                entry5['state'] = 'disabled'
            elif row == 0 and col == 2:
                entry5.insert(0, '柜号')
                entry5['state'] = 'disabled'
            elif row == 0 and col == 3:
                entry5.insert(0, '厂家')
                entry5['state'] = 'disabled'
            elif row == 0 and col == 4:
                entry5.insert(0, '型号')
                entry5['state'] = 'disabled'
            elif row == 0 and col == 5:
                entry5.insert(0, '安装方式')
                entry5['state'] = 'disabled'

            elif row != 0 and col == 3:
                entry5.bind("<Button-1>", lambda event, row=row, col=col: show_zeroctcompany_items(event, row, col))
            elif row != 0 and col == 4:
                entry5.bind("<Button-1>", lambda event, row=row, col=col: show_zerocttype_items(event, row, col))
            elif row != 0 and col == 5:
                entry5.bind("<Button-1>", lambda event, row=row, col=col: show_zeroctinstallation_items(event, row, col))


    add_row_button5 = tk.Button(frame_zeroct_Left, text="增加行", command=lambda: add_row5(content, frame_zeroct_Right, canvas), font=tku._ft(int(10 * h_ratio), False), activebackground='blue')
    add_row_button5.pack(side=tk.TOP, pady=(0, int(5 * h_ratio)), padx=(int(10 * w_ratio), int(10 * w_ratio)))

    delete_row_button5 = tk.Button(frame_zeroct_Left, text="删除行", command=lambda: delete_row5(content, frame_zeroct_Right, canvas), font=tku._ft(int(10 * h_ratio), False), activebackground='blue')
    delete_row_button5.pack(side=tk.TOP, pady=(int(5 * h_ratio), 0), padx=(int(10 * w_ratio), int(10 * w_ratio)))

    f22.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线



    f2P = tk.Frame(content, bg="white", bd=0)

    read_button = tk.Button(f2P, text="查阅标准方案", command=read_design_details, font=tku._ft(int(10 * h_ratio), False), activebackground='blue')
    read_button.pack(side=tk.LEFT, padx=int(280 * w_ratio), pady=(int(30 * w_ratio), int(40 * w_ratio)))

    global export_button
    export_button = tk.Button(f2P, text="导出传递表.pdf", command=export_transfer_sheet, font=tku._ft(int(10 * h_ratio), False), activebackground='blue')
    export_button.pack(side=tk.LEFT, pady=(int(30 * w_ratio), int(40 * w_ratio)))
    export_button['state'] = 'disabled'

    f2P.pack(fill=tk.X, expand=True)

    canvas.update_idletasks()
    # content.update_idletasks()
    canvas.config(scrollregion=canvas.bbox('all'))

    if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools'):
        tk.messagebox.showwarning("提示", "请连接内网")


def on_busduct_combobox_select(event):
    selection = combobox_busduct_type.get()
    if selection == '有':
        frame_busduct1.pack(fill=tk.X)
        frame_busduct2.pack(fill=tk.X)

    else:
        frame_busduct1.pack_forget()
        frame_busduct2.pack_forget()


def on_gasduct_combobox_select(event):
    selection = combobox_gasduct.get()
    if selection != '无':
        frame_gasduct.pack(fill=tk.X)

    else:
        frame_gasduct.pack_forget()


def on_busduct_interface_combobox_select(event):
    selection = combobox_busduct_interface.get()
    if selection != '不需要':
        frame_busduct_interface.pack(fill=tk.X)

    else:
        frame_busduct_interface.pack_forget()


def on_phasesequence_combobox_select(event):
    selection = combobox_phasesequence.get()
    if selection == '有':
        frame_phasesequence.pack(fill=tk.X)

    else:
        frame_phasesequence.pack_forget()


def show_phasesequence_items(event, row, col):
    global phasesequence_items_select_row
    global phasesequence_items_select_col

    phasesequence_items_select_row = row
    phasesequence_items_select_col = col

    for i in phasesequence_items:
        phasesequence_items_listbox.insert(tk.END, i)
        phasesequence_items_listbox.pack(fill=tk.X)


def typeinselection(event):
    entries2[phasesequence_items_select_row][phasesequence_items_select_col].delete(0, tk.END)
    curselection = phasesequence_items_listbox.curselection()
    if curselection:
        selected_text = phasesequence_items_listbox.get(curselection)
        entries2[phasesequence_items_select_row][phasesequence_items_select_col].insert(0, selected_text)
        phasesequence_items_listbox.pack_forget()


def on_fan_combobox_select(event):
    selection = combobox_fan.get()
    if selection == '有':
        frame_fan.pack(fill=tk.X)

    else:
        frame_fan.pack_forget()


def on_meterpanel_combobox_select(event):
    selection = combobox_meterpanel.get()
    if selection == '有':
        frame_meterpanel.pack(fill=tk.X)

    else:
        frame_meterpanel.pack_forget()


def show_meterpanelseal_items(event, row, col):
    global meterpanelseal_items_select_row
    global meterpanelseal_items_select_col

    meterpanelseal_items_select_row = row
    meterpanelseal_items_select_col = col

    for i in meterpanelseal_items:
        meterpanelseal_items_listbox.insert(tk.END, i)
        meterpanelseal_items_listbox.pack(fill=tk.X)


def meterpanelseal_typeinselection(event):
    entries4[meterpanelseal_items_select_row][meterpanelseal_items_select_col].delete(0, tk.END)
    curselection = meterpanelseal_items_listbox.curselection()
    if curselection:
        selected_text = meterpanelseal_items_listbox.get(curselection)
        entries4[meterpanelseal_items_select_row][meterpanelseal_items_select_col].insert(0, selected_text)
        meterpanelseal_items_listbox.pack_forget()


def show_meterpanellighting_items(event, row, col):
    global meterpanellighting_items_select_row
    global meterpanellighting_items_select_col

    meterpanellighting_items_select_row = row
    meterpanellighting_items_select_col = col

    for i in meterpanellighting_items:
        meterpanellighting_items_listbox.insert(tk.END, i)
        meterpanellighting_items_listbox.pack(fill=tk.X)


def meterpanellighting_typeinselection(event):
    entries4[meterpanellighting_items_select_row][meterpanellighting_items_select_col].delete(0, tk.END)
    curselection = meterpanellighting_items_listbox.curselection()
    if curselection:
        selected_text = meterpanellighting_items_listbox.get(curselection)
        entries4[meterpanellighting_items_select_row][meterpanellighting_items_select_col].insert(0, selected_text)
        meterpanellighting_items_listbox.pack_forget()


def show_meterpanelmicroswitch_items(event, row, col):
    global meterpanelmicroswitch_items_select_row
    global meterpanelmicroswitch_items_select_col

    meterpanelmicroswitch_items_select_row = row
    meterpanelmicroswitch_items_select_col = col

    for i in meterpanelmicroswitch_items:
        meterpanelmicroswitch_items_listbox.insert(tk.END, i)
        meterpanelmicroswitch_items_listbox.pack(fill=tk.X)


def meterpanelmicroswitch_typeinselection(event):
    entries4[meterpanelmicroswitch_items_select_row][meterpanelmicroswitch_items_select_col].delete(0, tk.END)
    curselection = meterpanelmicroswitch_items_listbox.curselection()
    if curselection:
        selected_text = meterpanelmicroswitch_items_listbox.get(curselection)
        entries4[meterpanelmicroswitch_items_select_row][meterpanelmicroswitch_items_select_col].insert(0, selected_text)
        meterpanelmicroswitch_items_listbox.pack_forget()


def show_meterpaneltruckblockmagnet_items(event, row, col):
    global meterpaneltruckblockmagnet_items_select_row
    global meterpaneltruckblockmagnet_items_select_col

    meterpaneltruckblockmagnet_items_select_row = row
    meterpaneltruckblockmagnet_items_select_col = col

    for i in meterpaneltruckblockmagnet_items:
        meterpaneltruckblockmagnet_items_listbox.insert(tk.END, i)
        meterpaneltruckblockmagnet_items_listbox.pack(fill=tk.X)


def meterpaneltruckblockmagnet_typeinselection(event):
    entries4[meterpaneltruckblockmagnet_items_select_row][meterpaneltruckblockmagnet_items_select_col].delete(0, tk.END)
    curselection = meterpaneltruckblockmagnet_items_listbox.curselection()
    if curselection:
        selected_text = meterpaneltruckblockmagnet_items_listbox.get(curselection)
        entries4[meterpaneltruckblockmagnet_items_select_row][meterpaneltruckblockmagnet_items_select_col].insert(0, selected_text)
        meterpaneltruckblockmagnet_items_listbox.pack_forget()


def show_meterpanelopenoperation_items(event, row, col):
    global meterpanelopenoperation_items_select_row
    global meterpanelopenoperation_items_select_col

    meterpanelopenoperation_items_select_row = row
    meterpanelopenoperation_items_select_col = col

    for i in meterpanelopenoperation_items:
        meterpanelopenoperation_items_listbox.insert(tk.END, i)
        meterpanelopenoperation_items_listbox.pack(fill=tk.X)


def meterpanelopenoperation_typeinselection(event):
    entries4[meterpanelopenoperation_items_select_row][meterpanelopenoperation_items_select_col].delete(0, tk.END)
    curselection = meterpanelopenoperation_items_listbox.curselection()
    if curselection:
        selected_text = meterpanelopenoperation_items_listbox.get(curselection)
        entries4[meterpanelopenoperation_items_select_row][meterpanelopenoperation_items_select_col].insert(0, selected_text)
        meterpanelopenoperation_items_listbox.pack_forget()


def show_meterpanelopenwindowsize_items(event, row, col):
    global meterpanelopenwindowsize_items_select_row
    global meterpanelopenwindowsize_items_select_col

    meterpanelopenwindowsize_items_select_row = row
    meterpanelopenwindowsize_items_select_col = col

    for i in meterpanelopenwindowsize_items:
        meterpanelopenwindowsize_items_listbox.insert(tk.END, i)
        meterpanelopenwindowsize_items_listbox.pack(fill=tk.X)


def meterpanelopenwindowsize_typeinselection(event):
    entries4[meterpanelopenwindowsize_items_select_row][meterpanelopenwindowsize_items_select_col].delete(0, tk.END)
    curselection = meterpanelopenwindowsize_items_listbox.curselection()
    if curselection:
        selected_text = meterpanelopenwindowsize_items_listbox.get(curselection)
        entries4[meterpanelopenwindowsize_items_select_row][meterpanelopenwindowsize_items_select_col].insert(0, selected_text)
        meterpanelopenwindowsize_items_listbox.pack_forget()


def on_secondarycomponent_combobox_select(event):
    selection = combobox_secondarycomponent.get()
    if selection == '有':
        frame_secondarycomponent.pack(fill=tk.X)

    else:
        frame_secondarycomponent.pack_forget()


def select_secondarycomponent():
    global entry_select_secondarycomponent
    global secondarycomponents
    true_checkbox_secondarycomponent_list = ''
    for i in range(0, len(secondarycomponents)):
        if checkbox_secondarycomponent_list[i].get():
            true_checkbox_secondarycomponent_list += secondarycomponents[i] + ';'
    entry_select_secondarycomponent.delete(0, tk.END)
    entry_select_secondarycomponent.insert(tk.END, true_checkbox_secondarycomponent_list.replace(' ', ''))


def on_faultlimitingdevice_combobox_select(event):
    selection = combobox_faultlimitingdevice.get()
    if selection == '有':
        frame_faultlimitingdevice.pack(fill=tk.X)
        frame_faultlimitingdevice2.pack(fill=tk.X)

    else:
        frame_faultlimitingdevice.pack_forget()
        frame_faultlimitingdevice2.pack_forget()
        combobox_faultlimitingdevice2.set('')


def on_faultlimitingdevice2_combobox_select(event):
    selection = combobox_faultlimitingdevice2.get()
    if selection != 'UFES':
        frame_faultlimitingdevice2.pack(fill=tk.X)

    else:
        frame_faultlimitingdevice2.pack_forget()


def on_electricaloperation_combobox_select(event):
    selection = combobox_electricaloperation.get()
    if selection == '有':
        frame_electricaloperation.pack(fill=tk.X)

    else:
        frame_electricaloperation.pack_forget()


def on_speccb_combobox_select(event):
    selection = combobox_speccb.get()
    if selection == '是':
        frame_speccb.pack(fill=tk.X)

    else:
        frame_speccb.pack_forget()


def select_speccb():
    global entry_select_speccb
    global speccbs
    true_checkbox_speccb_list = ''
    for i in range(0, len(speccbs)):
        if checkbox_speccb_list[i].get():
            true_checkbox_speccb_list += speccbs[i] + ';'
    entry_select_speccb.delete(0, tk.END)
    entry_select_speccb.insert(tk.END, true_checkbox_speccb_list.replace(' ', ''))


def on_nameplateopenhole_combobox_select(event):
    selection = combobox_nameplateopenhole.get()
    if selection == '有':
        frame_nameplateopenhole.pack(fill=tk.X)

    else:
        frame_nameplateopenhole.pack_forget()


def on_brownameplate_combobox_select(event):
    selection = combobox_brownameplate.get()
    if selection == '有':
        frame_brownameplate.pack(fill=tk.X)

    else:
        frame_brownameplate.pack_forget()


def on_specnameplate_combobox_select(event):
    selection = combobox_specnameplate.get()
    if selection == '有':
        frame_specnameplate.pack(fill=tk.X)

    else:
        frame_specnameplate.pack_forget()


def on_zeroct_combobox_select(event):
    selection = combobox_zeroct.get()
    if selection == '有':
        frame_zeroct.pack(fill=tk.X)

    else:
        frame_zeroct.pack_forget()


def zeroctcompany_typeinselection(event):
    entries5[zeroctcompany_items_select_row][zeroctcompany_items_select_col].delete(0, tk.END)
    entries5[zeroctcompany_items_select_row][zeroctcompany_items_select_col+1].delete(0, tk.END)
    curselection = zeroctcompany_items_listbox.curselection()
    if curselection:
        selected_text = zeroctcompany_items_listbox.get(curselection)
        entries5[zeroctcompany_items_select_row][zeroctcompany_items_select_col].insert(0, selected_text)
        zeroctcompany_items_listbox.pack_forget()


def show_zeroctcompany_items(event, row, col):
    global zeroctcompany_items_select_row
    global zeroctcompany_items_select_col

    zeroctcompany_items_select_row = row
    zeroctcompany_items_select_col = col

    for i in zeroctcompany_items:
        zeroctcompany_items_listbox.insert(tk.END, i)
        zeroctcompany_items_listbox.pack(fill=tk.X)


def zerocttype_typeinselection(event):
    entries5[zerocttype_items_select_row][zerocttype_items_select_col].delete(0, tk.END)
    curselection = zerocttype_items_listbox.curselection()
    if curselection:
        selected_text = zerocttype_items_listbox.get(curselection)
        entries5[zerocttype_items_select_row][zerocttype_items_select_col].insert(0, selected_text)
        zerocttype_items_listbox.pack_forget()


def show_zerocttype_items(event, row, col):
    global zerocttype_items_select_row
    global zerocttype_items_select_col
    global zerocttype_items
    global zerocttype_items_listbox

    zerocttype_items_select_row = row
    zerocttype_items_select_col = col

    zerocttype_items_listbox.delete(0, tk.END)

    if entries5[zerocttype_items_select_row][zerocttype_items_select_col-1].get() == 'ABB':
        zerocttype_items = ['LMZ165-50-70', 'LMZ215-80-120', 'LMZ255-80-160', 'LMZC-0.5', 'LMBF-0.5']
        zerocttype_items_listbox.configure(height=5)
    elif entries5[zerocttype_items_select_row][zerocttype_items_select_col-1].get() == 'NTK':
        zerocttype_items = ['LMZ165-50-70', 'LMZ215-80-120', 'LMZ255-80-160', 'LMZC-0.5', 'LMZC2-0.5', 'LMZC3-0.5', 'LMZC4-0.5', 'AM2-0.5', 'AM4-0.5']
        zerocttype_items_listbox.configure(height=9)
    elif entries5[zerocttype_items_select_row][zerocttype_items_select_col-1].get() == 'DYH':
        zerocttype_items = ['LMZ165-50-70', 'LMZ215-80-120', 'LMZ255-80-160', 'LMZC-0.5', 'LMZ6-0.5/155X310', 'LMBF-0.5']
        zerocttype_items_listbox.configure(height=6)
    elif entries5[zerocttype_items_select_row][zerocttype_items_select_col-1].get() == '北京科伟达':
        zerocttype_items = ['KLH-100K/方', 'KLH-120K/方', 'KLH-280Z/T']
        zerocttype_items_listbox.configure(height=3)
    else:
        return

    for i in zerocttype_items:
        zerocttype_items_listbox.insert(tk.END, i)
        zerocttype_items_listbox.pack(fill=tk.X)


def zeroctinstallation_typeinselection(event):
    entries5[zeroctinstallation_items_select_row][zeroctinstallation_items_select_col].delete(0, tk.END)
    curselection = zeroctinstallation_items_listbox.curselection()
    if curselection:
        selected_text = zeroctinstallation_items_listbox.get(curselection)
        entries5[zeroctinstallation_items_select_row][zeroctinstallation_items_select_col].insert(0, selected_text)
        zeroctinstallation_items_listbox.pack_forget()


def show_zeroctinstallation_items(event, row, col):
    global zeroctinstallation_items_select_row
    global zeroctinstallation_items_select_col

    zeroctinstallation_items_select_row = row
    zeroctinstallation_items_select_col = col

    for i in zeroctinstallation_items:
        zeroctinstallation_items_listbox.insert(tk.END, i)
        zeroctinstallation_items_listbox.pack(fill=tk.X)


def on_mousewheel(event):
    global canvas
    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
    # print(canvas.winfo_width())
    # canvas.create_text(canvas.winfo_width() - 10, 10, anchor='ne', text='滚动区↓')


def selectpath():
    filepath = tk.filedialog.askopenfilename(title=u'请选择.mve文件', filetypes=[("Excel", ".mve")])  # 选择打开什么文件，返回文件名
    if len(filepath) != 0:
        string_filename = ""
        for i in range(0, len(filepath)):
            string_filename += str(filepath[i])
        button_read['state'] = 'normal'
    else:
        button_read['state'] = 'disabled'


    mve_entry.delete(0, "end")  # 删除entry原始内容
    mve_entry.insert(0, filepath)  # 重新填入地址

    global FilePath

    FilePath = filepath


def getmvedata():
    print('a')


def on_select(event):
    global design_type_flag
    global design_type_last
    selected_value = combobox_design_type.get()
    if selected_value == '提前设计':
        entry_project_number['state'] = 'normal'
        entry_project_number.delete(0, "end")
        entry_project_number.insert(0, '提前设计')
        entry_project_number['state'] = 'disabled'

        button_query_project['state'] = 'disabled'
        button_update_amount['state'] = 'disabled'

        entry_project_name.delete(0, "end")

        text_item['state'] = 'normal'
        text_item.delete(1.0, tk.END)
        text_item.insert(tk.END, '1000')
        text_item['background'] = 'whitesmoke'
        text_item['state'] = 'disabled'

        text_typical.delete(1.0, tk.END)
        text_amount.delete(1.0, tk.END)
        combobox_product_type.set('')

        combobox_location_type.set('')
        combobox_protection_type.set('')
        combobox_drawing_language_type.set('')

        # text_drawing_language['state'] = 'normal'
        # text_drawing_language.delete(1.0, tk.END)
        # text_drawing_language['background'] = 'whitesmoke'
        # text_drawing_language['state'] = 'disabled'

        combobox_customer_wiring.set('')

        # text_customer_wiring['state'] = 'normal'
        # text_customer_wiring.delete(1.0, tk.END)
        # text_customer_wiring['background'] = 'whitesmoke'
        # text_customer_wiring['state'] = 'disabled'

        combobox_customer_requirements.set('')

        # text_customer_requirements['state'] = 'normal'
        # text_customer_requirements.delete(1.0, tk.END)
        # text_customer_requirements['background'] = 'whitesmoke'
        # text_customer_requirements['state'] = 'disabled'

        combobox_MCB_type.set('人民电器')
        combobox_aux_type.set('否')
        combobox_terminal_type.set('瑞联')
        # combobox_charged_display_type.set('立林')
        combobox_switch_type.set('江阴长江')
        combobox_advice_times.set('2')
        # text_switch['state'] = 'normal'
        # text_switch.delete(1.0, tk.END)
        # text_switch['background'] = 'whitesmoke'
        # text_switch['state'] = 'disabled'

        combobox_project_type.set('')
        combobox_frame_type.set('')
        combobox_program_type.set('')
        combobox_language_type.set('')
        combobox_manage_type.set('')

        entry_bom_time['state'] = 'disabled'
        button_select_bom_time['state'] = 'disabled'
        design_type_flag = 1

        if selected_value != design_type_last:
            entry_calculate_time['state'] = 'normal'
            entry_calculate_time.delete(0, "end")
            entry_calculate_time['state'] = 'disabled'
            entry_extra_time['state'] = 'normal'
            entry_extra_time.delete(0, "end")
            entry_extra_time['state'] = 'disabled'
            entry_upload_time['state'] = 'normal'
            entry_upload_time.delete(0, "end")
            entry_upload_time['state'] = 'disabled'
            entry_bom_time['state'] = 'normal'
            entry_bom_time.delete(0, "end")
            entry_bom_time['state'] = 'disabled'
            entry_check_time['state'] = 'normal'
            entry_check_time.delete(0, "end")
            entry_check_time['state'] = 'disabled'

            button_select_upload_time['state'] = 'disabled'
            button_select_bom_time['state'] = 'disabled'
            button_select_check_time['state'] = 'disabled'

    elif selected_value == '图纸设计':
        entry_project_number['state'] = 'normal'
        entry_project_number.delete(0, "end")

        button_query_project['state'] = 'normal'

        entry_project_name.delete(0, "end")

        text_item['state'] = 'normal'
        text_item.delete(1.0, tk.END)
        text_item['background'] = 'white'

        text_typical.delete(1.0, tk.END)
        text_amount.delete(1.0, tk.END)
        combobox_product_type.set('')

        combobox_location_type.set('')
        combobox_protection_type.set('')
        combobox_drawing_language_type.set('')

        # text_drawing_language['state'] = 'normal'
        # text_drawing_language.delete(1.0, tk.END)
        # text_drawing_language['background'] = 'whitesmoke'
        # text_drawing_language['state'] = 'disabled'

        combobox_customer_wiring.set('')

        # text_customer_wiring['state'] = 'normal'
        # text_customer_wiring.delete(1.0, tk.END)
        # text_customer_wiring['background'] = 'whitesmoke'
        # text_customer_wiring['state'] = 'disabled'

        combobox_customer_requirements.set('')

        # text_customer_requirements['state'] = 'normal'
        # text_customer_requirements.delete(1.0, tk.END)
        # text_customer_requirements['background'] = 'whitesmoke'
        # text_customer_requirements['state'] = 'disabled'

        combobox_MCB_type.set('人民电器')
        combobox_aux_type.set('否')
        combobox_terminal_type.set('瑞联')
        # combobox_charged_display_type.set('立林')
        combobox_switch_type.set('江阴长江')
        combobox_advice_times.set('2')
        # text_switch['state'] = 'normal'
        # text_switch.delete(1.0, tk.END)
        # text_switch['background'] = 'whitesmoke'
        # text_switch['state'] = 'disabled'

        combobox_project_type.set('')
        combobox_frame_type.set('')
        combobox_program_type.set('')
        combobox_language_type.set('')
        combobox_manage_type.set('')

        entry_bom_time['state'] = 'disabled'
        button_select_bom_time['state'] = 'normal'
        design_type_flag = 2

        if selected_value != design_type_last:
            entry_calculate_time['state'] = 'normal'
            entry_calculate_time.delete(0, "end")
            entry_calculate_time['state'] = 'disabled'
            entry_extra_time['state'] = 'normal'
            entry_extra_time.delete(0, "end")
            entry_extra_time['state'] = 'disabled'
            entry_upload_time['state'] = 'normal'
            entry_upload_time.delete(0, "end")
            entry_upload_time['state'] = 'disabled'
            entry_bom_time['state'] = 'normal'
            entry_bom_time.delete(0, "end")
            entry_bom_time['state'] = 'disabled'
            entry_check_time['state'] = 'normal'
            entry_check_time.delete(0, "end")
            entry_check_time['state'] = 'disabled'

            button_select_upload_time['state'] = 'disabled'
            button_select_bom_time['state'] = 'disabled'
            button_select_check_time['state'] = 'disabled'

    elif selected_value == '工程设计':
        entry_project_number['state'] = 'normal'
        entry_project_number.delete(0, "end")

        button_query_project['state'] = 'normal'

        entry_project_name.delete(0, "end")
        text_item['state'] = 'normal'
        text_item.delete(1.0, tk.END)
        text_item['background'] = 'white'

        text_typical.delete(1.0, tk.END)
        text_amount.delete(1.0, tk.END)
        combobox_product_type.set('')

        combobox_location_type.set('')
        combobox_protection_type.set('')
        combobox_drawing_language_type.set('')

        # text_drawing_language['state'] = 'normal'
        # text_drawing_language.delete(1.0, tk.END)
        # text_drawing_language['background'] = 'whitesmoke'
        # text_drawing_language['state'] = 'disabled'

        combobox_customer_wiring.set('')

        # text_customer_wiring['state'] = 'normal'
        # text_customer_wiring.delete(1.0, tk.END)
        # text_customer_wiring['background'] = 'whitesmoke'
        # text_customer_wiring['state'] = 'disabled'

        combobox_customer_requirements.set('')

        # text_customer_requirements['state'] = 'normal'
        # text_customer_requirements.delete(1.0, tk.END)
        # text_customer_requirements['background'] = 'whitesmoke'
        # text_customer_requirements['state'] = 'disabled'

        combobox_MCB_type.set('人民电器')
        combobox_aux_type.set('否')
        combobox_terminal_type.set('瑞联')
        # combobox_charged_display_type.set('立林')
        combobox_switch_type.set('江阴长江')
        combobox_advice_times.set('2')
        # text_switch['state'] = 'normal'
        # text_switch.delete(1.0, tk.END)
        # text_switch['background'] = 'whitesmoke'
        # text_switch['state'] = 'disabled'

        combobox_project_type.set('')
        combobox_frame_type.set('')
        combobox_program_type.set('')
        combobox_language_type.set('')
        combobox_manage_type.set('')

        entry_bom_time['state'] = 'disabled'
        button_select_bom_time['state'] = 'normal'
        design_type_flag = 3

        if selected_value != design_type_last:
            entry_calculate_time['state'] = 'normal'
            entry_calculate_time.delete(0, "end")
            entry_calculate_time['state'] = 'disabled'
            entry_extra_time['state'] = 'normal'
            entry_extra_time.delete(0, "end")
            entry_extra_time['state'] = 'disabled'
            entry_upload_time['state'] = 'normal'
            entry_upload_time.delete(0, "end")
            entry_upload_time['state'] = 'disabled'
            entry_bom_time['state'] = 'normal'
            entry_bom_time.delete(0, "end")
            entry_bom_time['state'] = 'disabled'
            entry_check_time['state'] = 'normal'
            entry_check_time.delete(0, "end")
            entry_check_time['state'] = 'disabled'

            button_select_upload_time['state'] = 'disabled'
            button_select_bom_time['state'] = 'disabled'
            button_select_check_time['state'] = 'disabled'

    design_type_last = selected_value


def query_project():
    try:
        if entry_project_number.get() == "" or len(entry_project_number.get()) != 9:
            tk.messagebox.showwarning("提示", "请输入9位合同号")
        elif not entry_project_number.get().isdigit():
            tk.messagebox.showwarning("提示", "合同号必须是9位纯数字")
        else:
            try:
                workbook1 = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\SR_Material.xlsx")
                worksheet1 = workbook1['Sheet1']
            except:
                tk.messagebox.showwarning("提示", "缺少产品型号表")

            worksheet1_max_row = max(ee.row for ee in worksheet1['A'] if ee.value)
            global material_list
            material_list = []
            global product_type_list
            product_type_list = []

            for i in range(1, worksheet1_max_row + 1):
                material_list.append(worksheet1.cell(row=i, column=1).value)
                product_type_list.append(worksheet1.cell(row=i, column=3).value)

            with sync_playwright() as playwright:
                run(playwright)
            update_amount()  # 每次读取完刷新一下
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def run(playwright: Playwright) -> None:
    try:

        entry_calculate_time['state'] = 'normal'
        entry_calculate_time.delete(0, "end")
        entry_calculate_time['state'] = 'disabled'
        entry_extra_time['state'] = 'normal'
        entry_extra_time.delete(0, "end")
        entry_extra_time['state'] = 'disabled'
        entry_upload_time['state'] = 'normal'
        entry_upload_time.delete(0, "end")
        entry_upload_time['state'] = 'disabled'
        entry_bom_time['state'] = 'normal'
        entry_bom_time.delete(0, "end")
        entry_bom_time['state'] = 'disabled'
        entry_check_time['state'] = 'normal'
        entry_check_time.delete(0, "end")
        entry_check_time['state'] = 'disabled'

        button_select_upload_time['state'] = 'disabled'
        button_select_bom_time['state'] = 'disabled'
        button_select_check_time['state'] = 'disabled'

        # 这是下面playwright.chromium.launch()改为playwright.chromium.launch_persistent_context()所定义的两个参数
        user_data_dir = 'ms-playwright/chromium-1033/chrome-win/user data'
        executable_path = 'ms-playwright/chromium-1033/chrome-win/chrome.exe'

        browser = playwright.chromium.launch_persistent_context(headless=True, executable_path=executable_path, user_data_dir=user_data_dir)
        # 修改登录函数后，下面两行改为一行
        # context = browser.new_context()
        # page = context.new_page()
        page = browser.new_page()

        page.goto("https://project-management.epds.cn.abb.com/ABBPMP/")
        page.get_by_role("link", name="项目管理").click()
        page.get_by_role("link", name="工程进度管理").click()
        page.get_by_role("link", name="创建项目").click()
        page.frame_locator("#iframeid").locator("#txtSO").click()
        page.frame_locator("#iframeid").locator("#txtSO").fill("%s" % entry_project_number.get())
        page.frame_locator("#iframeid").get_by_role("button", name="查询").click()
        page.frame_locator("#iframeid").get_by_role("row", name=entry_project_number.get()).get_by_role("link").click()
        project_name = page.frame_locator("#iframeid").locator('xpath=//*[@id="grid"]/div[2]/table/tbody/tr/td[3]').inner_text()
        entry_project_name.delete(0, "end")
        entry_project_name.insert(0, project_name)
        page.frame_locator("#iframeid").get_by_role("link", name="项目模板和客户信息").click()

        # rows = page.frame_locator("#iframeid").locator('xpath=//*[@id="gridItem"]/div[2]')
        # print(rows.count())
        # print(rows.inner_text(), len(rows.inner_text()))

        rows = page.frame_locator("#iframeid").locator('xpath=//*[@id="gridItem"]/div[2]/table/tbody/tr', has=page.get_by_role("gridcell"))

        table_content = []
        for i in range(1, rows.count() + 1):
            for j in range(2, 10):
                table_content.append(page.frame_locator("#iframeid").locator('xpath=//*[@id="gridItem"]/div[2]/table/tbody/tr[%d]/td[%d]' % (i, j)).inner_text())

        # ['1000', '1201', 'UNIGEAR-ZS1-500', '2022-08-31', '14', '833478.76', 'YANNIE-XIAOYAN.CHEN@CN.ABB.COM', 'XIAOQING.GAO@CN.ABB.COM', '2000', '1201', 'UNIGEAR-ZS1-500']
        # print(len(table_content))
        # print(table_content)
        global item_list
        item_list = []
        global amount_list
        amount_list = []
        amount_panel = 0
        global product_type_list1
        product_type_list1 = []
        global product_type_list2
        product_type_list2 = []
        text_item['state'] = 'normal'
        text_item.delete(1.0, tk.END)
        text_amount.delete(1.0, tk.END)

        for i in range(1, rows.count() + 1):
            for j in range(0, len(material_list)):
                if table_content[i * 8 - 6] == material_list[j]:
                    item_list.append(int(table_content[(i - 1) * 8]))
                    amount_list.append(int(table_content[(i - 1) * 8 + 4]))
                    amount_panel += int(table_content[(i - 1) * 8 + 4])
                    product_type_list1.append(material_list[j])
                    product_type_list2.append(product_type_list[j])
                    break

        item_list_copy = item_list.copy()
        item_list.sort()

        amount_list = [amount_list[i] for i in sorted(range(len(item_list_copy)), key=lambda x: item_list_copy[x])]
        product_type_list1 = [product_type_list1[i] for i in sorted(range(len(item_list_copy)), key=lambda x: item_list_copy[x])]
        product_type_list2 = [product_type_list2[i] for i in sorted(range(len(item_list_copy)), key=lambda x: item_list_copy[x])]

        # print(item_list, amount_list, product_type_list1, product_type_list2)
        text_item.insert(tk.END, ';'.join(str(num) for num in item_list))

        text_amount.insert(tk.INSERT, amount_panel)

        if set(product_type_list2) == {'AIS'}:
            combobox_product_type.set('AIS')
        elif set(product_type_list2) == {'GIS'}:
            combobox_product_type.set('GIS')
        else:
            combobox_product_type.set('AIS+GIS')

        button_update_amount['state'] = 'normal'

        # ---------------------
        page.close()
        # 修改登录函数后，context就变成未定义，此行要注释掉
        # context.close()
        browser.close()
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def add_row1(content1, f2L_Right, canvas):
    entry_row1 = len(entries1)
    entries1.append([None for _ in range(num_columns1)])

    row_frame1 = tk.Frame(f2L_Right)
    row_frame1.pack(side=tk.TOP, fill=tk.X)

    for col in range(num_columns1):
        entry1 = tk.Entry(row_frame1, width=entry_widths1[col], relief='solid')
        entry1.pack(side=tk.LEFT)
        entries1[entry_row1][col] = entry1
        entry1.config(justify='center')

    # 更新滚动条
    content1.update_idletasks()
    canvas.config(scrollregion=canvas.bbox('all'))
    # canvas.yview_moveto(1)


def delete_row1(content1, f2L_Right, canvas):
    if len(entries1) > 2:
        last_row_frame1 = f2L_Right.winfo_children()[-1]
        last_row_frame1.destroy()
        entries1.pop()

        # 更新滚动条
        content1.update_idletasks()
        canvas.config(scrollregion=canvas.bbox('all'))
        # canvas.yview_moveto(1)


def add_row2(content1, f2M_Right, canvas):
    entry_row2 = len(entries2)
    entries2.append([None for _ in range(num_columns2)])

    row_frame2 = tk.Frame(f2M_Right)
    row_frame2.pack(side=tk.TOP, fill=tk.X)
    global phasesequence_items_select_row
    global phasesequence_items_select_col
    for col in range(num_columns2):
        entry2 = tk.Entry(row_frame2, width=entry_widths2[col], relief='solid')
        entry2.pack(side=tk.LEFT)
        entries2[entry_row2][col] = entry2
        entry2.config(justify='center')
        entry2.bind("<Button-1>", lambda event, row=entry_row2, col=col: show_phasesequence_items(event, row, col))
    # 更新滚动条
    content1.update_idletasks()
    canvas.config(scrollregion=canvas.bbox('all'))
    # canvas.yview_moveto(1)


def delete_row2(content1, f2M_Right, canvas):
    if len(entries2) > 1:
        last_row_frame2 = f2M_Right.winfo_children()[-1]
        last_row_frame2.destroy()
        entries2.pop()

        # 更新滚动条
        content1.update_idletasks()
        canvas.config(scrollregion=canvas.bbox('all'))
        # canvas.yview_moveto(1)


def add_row3(content1, f2L_Right, canvas):
    entry_row3 = len(entries3)
    entries3.append([None for _ in range(num_columns3)])

    row_frame3 = tk.Frame(f2L_Right)
    row_frame3.pack(side=tk.TOP, fill=tk.X)

    for col in range(num_columns3):
        entry3 = tk.Entry(row_frame3, width=entry_widths3[col], relief='solid')
        entry3.pack(side=tk.LEFT)
        entries3[entry_row3][col] = entry3
        entry3.config(justify='center')

    # 更新滚动条
    content1.update_idletasks()
    canvas.config(scrollregion=canvas.bbox('all'))
    # canvas.yview_moveto(1)


def delete_row3(content1, f2L_Right, canvas):
    if len(entries3) > 2:
        last_row_frame3 = f2L_Right.winfo_children()[-1]
        last_row_frame3.destroy()
        entries3.pop()

        # 更新滚动条
        content1.update_idletasks()
        canvas.config(scrollregion=canvas.bbox('all'))
        # canvas.yview_moveto(1)


def add_row4(content1, f2L_Right, canvas):
    entry_row4 = len(entries4)
    entries4.append([None for _ in range(num_columns4)])

    row_frame4 = tk.Frame(f2L_Right)
    row_frame4.pack(side=tk.TOP, fill=tk.X)
    global meterpanelseal_items_select_row
    global meterpanelseal_items_select_col
    global meterpanellighting_items_select_row
    global meterpanellighting_items_select_col
    global meterpanelmicroswitch_items_select_row
    global meterpanelmicroswitch_items_select_col
    global meterpaneltruckblockmagnet_items_select_row
    global meterpaneltruckblockmagnet_items_select_col
    global meterpanelopenoperation_items_select_row
    global meterpanelopenoperation_items_select_col
    global meterpanelopenwindowsize_items_select_row
    global meterpanelopenwindowsize_items_select_col
    for col in range(num_columns4):
        entry4 = tk.Entry(row_frame4, width=entry_widths4[col], relief='solid')
        entry4.pack(side=tk.LEFT)
        entries4[entry_row4][col] = entry4
        entry4.config(justify='center')
        if col == 1:
            entry4.bind("<Button-1>", lambda event, row=entry_row4, col=col: show_meterpanelseal_items(event, row, col))
        elif col == 2:
            entry4.bind("<Button-1>", lambda event, row=entry_row4, col=col: show_meterpanellighting_items(event, row, col))
        elif col == 3:
            entry4.bind("<Button-1>", lambda event, row=entry_row4, col=col: show_meterpanelmicroswitch_items(event, row, col))
        elif col == 4:
            entry4.bind("<Button-1>", lambda event, row=entry_row4, col=col: show_meterpaneltruckblockmagnet_items(event, row, col))
        elif col == 5:
            entry4.bind("<Button-1>", lambda event, row=entry_row4, col=col: show_meterpanelopenoperation_items(event, row, col))
        elif col == 6:
            entry4.bind("<Button-1>", lambda event, row=entry_row4, col=col: show_meterpanelopenwindowsize_items(event, row, col))


    # 更新滚动条
    content1.update_idletasks()
    canvas.config(scrollregion=canvas.bbox('all'))
    # canvas.yview_moveto(1)


def delete_row4(content1, f2L_Right, canvas):
    if len(entries4) > 2:
        last_row_frame4 = f2L_Right.winfo_children()[-1]
        last_row_frame4.destroy()
        entries4.pop()

        # 更新滚动条
        content1.update_idletasks()
        canvas.config(scrollregion=canvas.bbox('all'))
        # canvas.yview_moveto(1)


def add_row5(content1, f2L_Right, canvas):
    entry_row5 = len(entries5)
    entries5.append([None for _ in range(num_columns5)])

    row_frame5 = tk.Frame(f2L_Right)
    row_frame5.pack(side=tk.TOP, fill=tk.X)
    global zeroctcompany_items_select_row
    global zeroctcompany_items_select_col
    global zerocttype_items_select_row
    global zerocttype_items_select_col
    global zeroctinstallation_items_select_row
    global zeroctinstallation_items_select_col

    for col in range(num_columns5):
        entry5 = tk.Entry(row_frame5, width=entry_widths5[col], relief='solid')
        entry5.pack(side=tk.LEFT)
        entries5[entry_row5][col] = entry5
        entry5.config(justify='center')
        if col == 3:
            entry5.bind("<Button-1>", lambda event, row=entry_row5, col=col: show_zeroctcompany_items(event, row, col))
        elif col == 4:
            entry5.bind("<Button-1>", lambda event, row=entry_row5, col=col: show_zerocttype_items(event, row, col))
        elif col == 5:
            entry5.bind("<Button-1>", lambda event, row=entry_row5, col=col: show_zeroctinstallation_items(event, row, col))

    # 更新滚动条
    content1.update_idletasks()
    canvas.config(scrollregion=canvas.bbox('all'))
    # canvas.yview_moveto(1)


def delete_row5(content1, f2L_Right, canvas):
    if len(entries5) > 2:
        last_row_frame5 = f2L_Right.winfo_children()[-1]
        last_row_frame5.destroy()
        entries5.pop()

        # 更新滚动条
        content1.update_idletasks()
        canvas.config(scrollregion=canvas.bbox('all'))
        # canvas.yview_moveto(1)


def add_row10(content1, f2M_Right, canvas):
    entry_row10 = len(entries10)
    entries10.append([None for _ in range(num_columns10)])

    row_frame10 = tk.Frame(f2M_Right)
    row_frame10.pack(side=tk.TOP, fill=tk.X)

    for col in range(num_columns10):
        entry10 = tk.Entry(row_frame10, width=entry_widths10[col], relief='solid')
        entry10.pack(side=tk.LEFT)
        if col == 0:
            entry10.config(justify='center')
            entry10.insert(0, '%s' % str(entry_row10 + 1))
            entry10['state'] = 'disabled'

        entries10[entry_row10][col] = entry10
        entry10.config()

    # 更新滚动条
    content1.update_idletasks()
    canvas.config(scrollregion=canvas.bbox('all'))
    # canvas.yview_moveto(1)


def delete_row10(content1, f2M_Right, canvas):
    if len(entries10) > 1:
        last_row_frame10 = f2M_Right.winfo_children()[-1]
        last_row_frame10.destroy()
        entries10.pop()

        # 更新滚动条
        content1.update_idletasks()
        canvas.config(scrollregion=canvas.bbox('all'))
        # canvas.yview_moveto(1)


def add_row11(content1, f2L_Right, canvas):
    entry_row11 = len(entries11)
    entries11.append([None for _ in range(num_columns11)])

    row_frame11 = tk.Frame(f2L_Right)
    row_frame11.pack(side=tk.TOP, fill=tk.X)

    for col in range(num_columns11):
        entry11 = tk.Entry(row_frame11, width=entry_widths11[col], relief='solid')
        entry11.pack(side=tk.LEFT)
        entries11[entry_row11][col] = entry11
        entry11.config(justify='center')

    # 更新滚动条
    content1.update_idletasks()
    canvas.config(scrollregion=canvas.bbox('all'))
    canvas.yview_moveto(1)


def delete_row11(content1, f2L_Right, canvas):
    if len(entries11) > 1:
        last_row_frame11 = f2L_Right.winfo_children()[-1]
        last_row_frame11.destroy()
        entries11.pop()

        # 更新滚动条
        content1.update_idletasks()
        canvas.config(scrollregion=canvas.bbox('all'))
        canvas.yview_moveto(1)


def read_design_details():
    if not os.path.exists(os.path.abspath('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\原理图标准设计_red.pdf')):
        tk.messagebox.showwarning("提示", "缺少原理图标准设计说明书.pdf")
    else:
        os.startfile(os.path.abspath('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\原理图标准设计_red.pdf'))
        export_button['state'] = 'normal'


def export_transfer_sheet():
    try:
        if combobox_design_type.get() == '' or entry_project_number.get() == '' or entry_project_name.get() == '' or (combobox_design_type.get() == '提前设计' and (entry_calculate_time.get() == '' or entry_extra_time.get() == '' or entry_upload_time.get() == '')) or \
                (combobox_design_type.get() == '图纸设计' and (entry_calculate_time.get() == '' or entry_extra_time.get() == '' or entry_upload_time.get() == '' or entry_check_time.get() == '')) or \
                (combobox_design_type.get() == '工程设计' and (entry_calculate_time.get() == '' or entry_extra_time.get() == '' or entry_upload_time.get() == '' or entry_check_time.get() == '' or entry_bom_time.get() == '')) or \
                combobox_MCB_type.get() == '' or combobox_aux_type.get() == '' or combobox_terminal_type.get() == '' or combobox_charged_display_type.get() == '' or combobox_switch_type.get() == '' or combobox_project_type.get() == '' or \
                combobox_frame_type.get() == '' or combobox_program_type.get() == '' or combobox_language_type.get() == '' or combobox_manage_type.get() == '':
            tk.messagebox.showwarning("提示", "*标属性未填写完整，无法导出Excel")

        elif combobox_design_type.get() != '提前设计' and (entry_project_number.get() == "" or len(entry_project_number.get()) != 9 or not entry_project_number.get().isdigit()):
            tk.messagebox.showwarning("提示", "请输入9位合同号，且需要为纯数字")

        else:
            try:
                exit_excel = tk.messagebox.askquestion("提示", "Excel、OneNote（右下角任务栏）是否已经退出？")
                if exit_excel == 'yes':
                    template_book = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\Design_Transfer_Sheet.xlsx")
                    template_sheet = template_book['Sheet1']
                    template_sheet_max_row = max(ee.row for ee in template_sheet['A'] if ee.value)
                    Timestamp = strftime('%Y-%m-%d %H:%M:%S', localtime())  # 时间戳
                    template_sheet.cell(row=2, column=7).value = Timestamp
                    template_sheet.cell(row=4, column=2).value = combobox_design_type.get()
                    template_sheet.cell(row=5, column=2).value = entry_project_number.get()
                    template_sheet.cell(row=6, column=2).value = entry_project_name.get()
                    if combobox_design_type.get() == '工程设计':
                        template_sheet.cell(row=4, column=3).value = '项目经理:'
                        template_sheet.cell(row=4, column=7).value = '工程师:'

                    temp = str(text_item.get("1.0", "end-1c")).replace('；', ';')
                    if temp[len(temp) - 1] == ';':
                        temp = temp[0:len(temp) - 1]

                    template_sheet.cell(row=7, column=2).value = temp

                    template_sheet.cell(row=8, column=2).value = text_typical.get("1.0", "end-1c")
                    template_sheet.cell(row=8, column=4).value = text_amount.get("1.0", "end-1c")
                    template_sheet.cell(row=8, column=6).value = combobox_product_type.get()

                    true_checkbox_input_list = ''
                    for i in range(0, 5):
                        if checkbox_input_list[i].get():
                            true_checkbox_input_list += input_files[i] + ';'

                    template_sheet.cell(row=13, column=2).value = true_checkbox_input_list[0:len(true_checkbox_input_list) - 1].replace(' ', '')

                    true_checkbox_drawing_type_list = ''
                    for i in range(0, 10):
                        if checkbox_drawing_type_list[i].get():
                            true_checkbox_drawing_type_list += drawing_types[i] + ';'

                    template_sheet.cell(row=15, column=2).value = true_checkbox_drawing_type_list[0:len(true_checkbox_drawing_type_list) - 1].replace(' ', '')

                    true_checkbox_intelligent_list = ''
                    for i in range(0, 7):
                        if checkbox_intelligent_list[i].get():
                            true_checkbox_intelligent_list += intelligent_needs[i] + ';'

                    template_sheet.cell(row=14, column=2).value = true_checkbox_intelligent_list[0:len(true_checkbox_intelligent_list) - 1].replace(' ', '')

                    template_sheet.cell(row=9, column=6).value = combobox_location_type.get()
                    template_sheet.cell(row=9, column=4).value = combobox_protection_type.get()

                    template_sheet.cell(row=9, column=2).value = combobox_drawing_language_type.get()
                    # if combobox_drawing_language_type.get() != '英文+AS BUILT其他语言(备注)':
                    #     template_sheet.cell(row=16, column=6).value = 'N/A'
                    #     template_sheet.cell(row=16, column=6).font = Font(name='ABBvoice CNSG Light', size=9)
                    # else:
                    #     if len(text_drawing_language.get("1.0", "end-1c")) > 23:
                    #         template_sheet.cell(row=16, column=6).font = Font(name='ABBvoice CNSG Light', size=6)
                    #     template_sheet.cell(row=16, column=6).value = text_drawing_language.get("1.0", "end-1c")

                    template_sheet.cell(row=11, column=2).value = combobox_customer_wiring.get()
                    # if combobox_customer_wiring.get() == '否':
                    #     template_sheet.cell(row=17, column=6).value = 'N/A'
                    #     template_sheet.cell(row=17, column=6).font = Font(name='ABBvoice CNSG Light', size=9)
                    # else:
                    #     if len(text_customer_wiring.get("1.0", "end-1c")) > 23:
                    #         template_sheet.cell(row=17, column=6).font = Font(name='ABBvoice CNSG Light', size=6)
                    #     template_sheet.cell(row=17, column=6).value = text_customer_wiring.get("1.0", "end-1c")

                    template_sheet.cell(row=12, column=2).value = combobox_customer_requirements.get()
                    # if combobox_customer_requirements.get() == '否':
                    #     template_sheet.cell(row=18, column=6).value = 'N/A'
                    #     template_sheet.cell(row=18, column=6).font = Font(name='ABBvoice CNSG Light', size=9)
                    # else:
                    #     if len(text_customer_requirements.get("1.0", "end-1c")) > 23:
                    #         template_sheet.cell(row=18, column=6).font = Font(name='ABBvoice CNSG Light', size=6)
                    #     template_sheet.cell(row=18, column=6).value = text_customer_requirements.get("1.0", "end-1c")

                    template_sheet.cell(row=16, column=2).value = str(entry_calculate_time.get()) + '天'
                    template_sheet.cell(row=16, column=4).value = str(entry_extra_time.get()) + '天'
                    template_sheet.cell(row=16, column=6).value = combobox_advice_times.get() + '次'
                    template_sheet.cell(row=17, column=2).value = entry_upload_time.get()
                    template_sheet.cell(row=17, column=4).value = entry_bom_time.get()
                    template_sheet.cell(row=17, column=6).value = entry_check_time.get()

                    if entry_bom_time.get() == '':
                        template_sheet.cell(row=17, column=4).value = 'N/A'
                        template_sheet.cell(row=17, column=4).font = Font(name='ABBvoice CNSG Light', size=9)
                    if entry_check_time.get() == '':
                        template_sheet.cell(row=17, column=6).value = 'N/A'
                        template_sheet.cell(row=17, column=6).font = Font(name='ABBvoice CNSG Light', size=9)

                    template_sheet.cell(row=10, column=2).value = combobox_MCB_type.get()
                    template_sheet.cell(row=10, column=4).value = combobox_aux_type.get()
                    template_sheet.cell(row=10, column=6).value = combobox_terminal_type.get()
                    template_sheet.cell(row=10, column=8).value = combobox_charged_display_type.get()
                    template_sheet.cell(row=9, column=8).value = combobox_switch_type.get()
                    # if combobox_switch_type.get() != '其他(备注)':
                    #     template_sheet.cell(row=11, column=4).value = 'N/A'
                    #     template_sheet.cell(row=11, column=4).font = Font(name='ABBvoice CNSG Light', size=9)
                    # else:
                    #     template_sheet.cell(row=11, column=4).value = text_switch.get("1.0", "end-1c")

                    template_sheet.cell(row=18, column=2).value = combobox_project_type.get()
                    template_sheet.cell(row=18, column=6).value = combobox_frame_type.get()
                    if len(combobox_frame_type.get()) > 23:
                        template_sheet.cell(row=18, column=6).font = Font(name='ABBvoice CNSG Light', size=6)

                    template_sheet.cell(row=19, column=2).value = combobox_program_type.get()
                    template_sheet.cell(row=19, column=6).value = combobox_language_type.get()
                    template_sheet.cell(row=20, column=2).value = combobox_manage_type.get()
                    if len(combobox_manage_type.get()) > 23:
                        template_sheet.cell(row=20, column=2).font = Font(name='ABBvoice CNSG Light', size=6)

                    error_input1 = 0
                    valid_row_count1 = 0
                    for row in range(1, len(entries1)):
                        if ';' not in entries1[row][0].get() and ';' not in entries1[row][1].get() and ';' not in entries1[row][2].get():
                            if entries1[row][0].get() != '' or entries1[row][1].get() != '' or entries1[row][2].get() != '':
                                valid_row_count1 += 1
                        else:
                            error_input1 = 1

                    error_input2 = 0
                    valid_row_count2 = 0
                    for row in range(0, len(entries2)):
                        if ';' not in entries2[row][1].get():
                            if len(entries2[row][1].get()) > 0:
                                valid_row_count2 += 1
                        else:
                            error_input2 = 1

                    if error_input1 and not error_input2:
                        tk.messagebox.showwarning("提示", "请移除典型柜配置表中的;(分号)")
                        return
                    elif error_input2 and not error_input1:
                        tk.messagebox.showwarning("提示", "请移除备注表中的;(分号)")
                        return
                    elif error_input1 and error_input2:
                        tk.messagebox.showwarning("提示", "请移除典型柜配置表、备注表中的;(分号)")
                    else:
                        if valid_row_count1 >= valid_row_count2 - 1:
                            max_row = valid_row_count1
                        else:
                            max_row = valid_row_count2

                        border_style = Side(border_style='thin', color='000000')

                        for row in range(1, len(entries1)):
                            if entries1[row][0].get() != '' or entries1[row][1].get() != '' or entries1[row][2].get() != '':
                                template_sheet.cell(row=23 + row - 1, column=1).value = entries1[row][0].get()
                                template_sheet.cell(row=23 + row - 1, column=2).value = entries1[row][1].get()
                                template_sheet.cell(row=23 + row - 1, column=4).value = entries1[row][2].get()

                        count_k = 0
                        for row in range(0, len(entries2)):
                            if len(entries2[row][1].get()) > 0:
                                count_k += 1
                                template_sheet.cell(row=23 + row - 1, column=6).value = str(count_k) + '、' + entries2[row][1].get()
                                template_sheet.cell(row=23 + row - 1, column=6).font = Font(name='ABBvoice CNSG Light', size=7)

                        if valid_row_count1 >= valid_row_count2 - 1:
                            for i in range(1, max_row + 1):
                                template_sheet['A%d' % (23 + i - 1)].border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                                template_sheet['A%d' % (23 + i - 1)].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                                template_sheet['A%d' % (23 + i - 1)].font = Font(name='ABBvoice CNSG Light', size=7)

                                template_sheet.merge_cells('B%d:C%d' % (23 + i - 1, 23 + i - 1))
                                template_sheet['B%d' % (23 + i - 1)].border = Border(left=border_style, top=border_style, bottom=border_style)
                                template_sheet['C%d' % (23 + i - 1)].border = Border(right=border_style, top=border_style, bottom=border_style)
                                template_sheet['B%d' % (23 + i - 1)].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                                template_sheet['B%d' % (23 + i - 1)].font = Font(name='ABBvoice CNSG Light', size=7)

                                template_sheet.merge_cells('D%d:E%d' % (23 + i - 1, 23 + i - 1))
                                template_sheet['D%d' % (23 + i - 1)].border = Border(left=border_style, top=border_style, bottom=border_style)
                                template_sheet['E%d' % (23 + i - 1)].border = Border(right=border_style, top=border_style, bottom=border_style)
                                template_sheet['D%d' % (23 + i - 1)].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                                template_sheet['D%d' % (23 + i - 1)].font = Font(name='ABBvoice CNSG Light', size=7)

                                template_sheet.merge_cells('F%d:H%d' % (23 + i - 1, 23 + i - 1))
                                template_sheet['F%d' % (23 + i - 1)].border = Border(left=border_style)
                                template_sheet['H%d' % (23 + i - 1)].border = Border(right=border_style)
                                template_sheet['F%d' % (23 + i - 1)].alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
                                template_sheet['F%d' % (23 + i - 1)].font = Font(name='ABBvoice CNSG Light', size=7)

                            template_sheet['F%d' % (23 + max_row - 1)].border = Border(left=border_style, bottom=border_style)
                            template_sheet['G%d' % (23 + max_row - 1)].border = Border(bottom=border_style)
                            template_sheet['H%d' % (23 + max_row - 1)].border = Border(right=border_style, bottom=border_style)
                        else:
                            for i in range(1, max_row):
                                template_sheet['A%d' % (23 + i - 1)].border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                                template_sheet['A%d' % (23 + i - 1)].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                                template_sheet['A%d' % (23 + i - 1)].font = Font(name='ABBvoice CNSG Light', size=7)

                                template_sheet.merge_cells('B%d:C%d' % (23 + i - 1, 23 + i - 1))
                                template_sheet['B%d' % (23 + i - 1)].border = Border(left=border_style, top=border_style, bottom=border_style)
                                template_sheet['C%d' % (23 + i - 1)].border = Border(right=border_style, top=border_style, bottom=border_style)
                                template_sheet['B%d' % (23 + i - 1)].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                                template_sheet['B%d' % (23 + i - 1)].font = Font(name='ABBvoice CNSG Light', size=7)

                                template_sheet.merge_cells('D%d:E%d' % (23 + i - 1, 23 + i - 1))
                                template_sheet['D%d' % (23 + i - 1)].border = Border(left=border_style, top=border_style, bottom=border_style)
                                template_sheet['E%d' % (23 + i - 1)].border = Border(right=border_style, top=border_style, bottom=border_style)
                                template_sheet['D%d' % (23 + i - 1)].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                                template_sheet['D%d' % (23 + i - 1)].font = Font(name='ABBvoice CNSG Light', size=7)

                                template_sheet.merge_cells('F%d:H%d' % (23 + i - 1, 23 + i - 1))
                                template_sheet['F%d' % (23 + i - 1)].border = Border(left=border_style)
                                template_sheet['H%d' % (23 + i - 1)].border = Border(right=border_style)
                                template_sheet['F%d' % (23 + i - 1)].alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
                                template_sheet['F%d' % (23 + i - 1)].font = Font(name='ABBvoice CNSG Light', size=7)

                            template_sheet['F%d' % (23 + max_row - 2)].border = Border(left=border_style, bottom=border_style)
                            template_sheet['G%d' % (23 + max_row - 2)].border = Border(bottom=border_style)
                            template_sheet['H%d' % (23 + max_row - 2)].border = Border(right=border_style, bottom=border_style)

                        if combobox_design_type.get() != '提前设计':
                            # print(real_item_list.split(';'), real_product_type_list1, real_amount_list)
                            for i in range(0, len(real_item_list.split(';'))):
                                template_sheet.cell(row=39 + i + 1, column=1).value = (real_item_list.split(';'))[i]
                                template_sheet.cell(row=39 + i + 1, column=2).value = real_product_type_list1[i]
                                template_sheet.cell(row=39 + i + 1, column=5).value = real_amount_list[i]
                            for i in range(1, len(real_item_list.split(';')) + 1):
                                template_sheet['A%d' % (39 + i)].border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                                template_sheet['A%d' % (39 + i)].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                                template_sheet['A%d' % (39 + i)].font = Font(name='ABBvoice CNSG Light', size=7)

                                template_sheet.merge_cells('B%d:D%d' % (39 + i, 39 + i))
                                template_sheet['B%d' % (39 + i)].border = Border(left=border_style, top=border_style, bottom=border_style)
                                template_sheet['C%d' % (39 + i)].border = Border(top=border_style, bottom=border_style)
                                template_sheet['D%d' % (39 + i)].border = Border(right=border_style, top=border_style, bottom=border_style)
                                template_sheet['B%d' % (39 + i)].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                                template_sheet['B%d' % (39 + i)].font = Font(name='ABBvoice CNSG Light', size=7)

                                template_sheet['E%d' % (39 + i)].border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                                template_sheet['E%d' % (39 + i)].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                                template_sheet['E%d' % (39 + i)].font = Font(name='ABBvoice CNSG Light', size=7)

                        new_file = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\file\\设计传递表-%s.xlsx' % entry_project_number.get()
                        if os.path.exists(new_file):
                            os.remove(new_file)

                        template_book.save(new_file)

                        home_path = os.path.expanduser("~")
                        desktop_path = os.path.join(home_path, "Desktop")

                        if os.path.exists(os.path.join(desktop_path, '设计传递表-%s.pdf' % entry_project_number.get())):
                            os.remove(os.path.join(desktop_path, '设计传递表-%s.pdf' % entry_project_number.get()))

                        excel_app = client.DispatchEx('Excel.Application')
                        file = excel_app.Workbooks.Open(os.path.abspath(new_file))
                        file.ExportAsFixedFormat(0, os.path.abspath('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\file\\设计传递表-%s.pdf' % entry_project_number.get()))
                        file.Close()
                        excel_app.Quit()

                        input_paths = [os.path.abspath('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\file\\设计传递表-%s.pdf' % entry_project_number.get()), os.path.abspath('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\原理图标准设计.pdf')]
                        merger = PyPDF2.PdfMerger()
                        for path in input_paths:
                            merger.append(path)
                        merger.write(os.path.join(desktop_path, '设计传递表-%s.pdf' % entry_project_number.get()))
                        merger.close()

                        tk.messagebox.showwarning("提示", "设计传递表-%s.pdf\n已保存到桌面" % entry_project_number.get())

                        result = tk.messagebox.askquestion("提示", "确定要将项目信息写入数据库吗？")
                        if result == 'yes':
                            if os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools'):
                                if is_folder_hidden('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb') or os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb'):
                                    if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db'):
                                        tk.messagebox.showwarning("提示", "数据库不存在，请联系管理员")
                                    else:
                                        conn = sqlite3.connect('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db')
                                        cursor = conn.cursor()  # 创建一个Cursor

                                        # 创建project_data表
                                        cursor.execute('''CREATE TABLE IF NOT EXISTS project_data
                                                     (id INTEGER PRIMARY KEY AUTOINCREMENT,
                                                      design_type TEXT,
                                                      project_number TEXT,
                                                      item_info INTEGER,
                                                      project_name TEXT,
                                                      typical_amount INTEGER,
                                                      panel_amount INTEGER,
                                                      panel_type TEXT,
                                                      product_type TEXT,
                                                      PE TEXT,
                                                      DE TEXT,
                                                      intelligence TEXT,
                                                      main_bus_current INTEGER,
                                                      main_protection TEXT,
                                                      arc_light TEXT,
                                                      location_info TEXT,
                                                      drawing_language TEXT,
                                                      standard_time INTEGER,
                                                      extra_time INTEGER,
                                                      advice_times INTEGER,
                                                      upload_time TEXT,
                                                      bom_time TEXT,
                                                      check_time TEXT,
                                                      receive_time TEXT,
                                                      start_time TEXT,
                                                      estimated_time TEXT,
                                                      status_info TEXT,
                                                      actual_time TEXT,
                                                      team_info TEXT,
                                                      abnormal_status TEXT,
                                                      create_time TIMESTAMP,
                                                      project_type TEXT,
                                                      frame_type TEXT,
                                                      program_type TEXT,
                                                      language_info TEXT,
                                                      management_info TEXT,
                                                      input_file TEXT,
                                                      drawing_type TEXT,
                                                      mcb_type TEXT,
                                                      aux_type TEXT,
                                                      terminal_type TEXT,
                                                      charged_display_type TEXT,
                                                      switch_type TEXT,
                                                      special_panel_config TEXT,
                                                      remark TEXT,
                                                      reviser TEXT,
                                                      revise_time TIMESTAMP)''')

                                        content_table1 = ''
                                        for i in range(0, len(entries1)):
                                            for j in range(num_columns1):
                                                content_table1 += entries1[i][j].get() + ';'

                                        content_table2 = ''
                                        for i in range(0, len(entries2)):
                                            for j in range(num_columns2):
                                                content_table2 += entries2[i][j].get() + ';'
                                        # print(content_table1)
                                        # print(content_table2)

                                        if combobox_design_type.get() == '提前设计':
                                            data = {
                                                'design_type': combobox_design_type.get(),
                                                'project_number': entry_project_number.get(),
                                                'item_info': 1000,
                                                'project_name': entry_project_name.get(),
                                                'typical_amount': int(text_typical.get('1.0', 'end-1c')),
                                                'panel_amount': int(text_amount.get('1.0', 'end-1c')),
                                                'product_type': combobox_product_type.get(),
                                                'intelligence': true_checkbox_intelligent_list[0:len(true_checkbox_intelligent_list) - 1].replace(' ', ''),
                                                'location_info': combobox_location_type.get(),
                                                'drawing_language': combobox_drawing_language_type.get(),
                                                'standard_time': int(entry_calculate_time.get()),
                                                'extra_time': int(entry_extra_time.get()),
                                                'advice_times': int(combobox_advice_times.get()),
                                                'upload_time': entry_upload_time.get(),
                                                'bom_time': entry_bom_time.get(),
                                                'check_time': entry_check_time.get(),
                                                'create_time': Timestamp,
                                                'project_type': combobox_project_type.get(),
                                                'frame_type': combobox_frame_type.get(),
                                                'program_type': combobox_program_type.get(),
                                                'language_info': combobox_language_type.get(),
                                                'management_info': combobox_manage_type.get(),
                                                'input_file': true_checkbox_input_list[0:len(true_checkbox_input_list) - 1].replace(' ', ''),
                                                'drawing_type': true_checkbox_drawing_type_list[0:len(true_checkbox_drawing_type_list) - 1].replace(' ', ''),
                                                'mcb_type': combobox_MCB_type.get(),
                                                'aux_type': combobox_aux_type.get(),
                                                'terminal_type': combobox_terminal_type.get(),
                                                'charged_display_type': combobox_charged_display_type.get(),
                                                'switch_type': combobox_switch_type.get(),
                                                'special_panel_config': content_table1,
                                                'remark': content_table2
                                            }
                                            columns = ', '.join(data.keys())
                                            placeholders = ', '.join(['?' for _ in range(len(data))])
                                            query = f"INSERT INTO project_data ({columns}) VALUES ({placeholders})"

                                            values = [data.get(col) for col in data.keys()]
                                            cursor.execute(query, tuple(values))
                                            conn.commit()
                                            tk.messagebox.showwarning("提示", "数据写入完成")

                                        else:
                                            already_exist_item = []
                                            for i in range(0, len(real_item_list.split(';'))):

                                                # 插入之前先查询数据是否已经存在，若存在，则不插入；若不存在，则插入
                                                data = {
                                                    'design_type': combobox_design_type.get(),
                                                    'project_number': entry_project_number.get(),
                                                    'item_info': int((real_item_list.split(';'))[i]),
                                                }
                                                query = "SELECT * FROM project_data WHERE design_type=? AND project_number=? AND item_info=?"
                                                cursor.execute(query, (data['design_type'], data['project_number'], data['item_info']))
                                                result = cursor.fetchone()
                                                if result:
                                                    already_exist_item.append(int((real_item_list.split(';'))[i]))
                                            if len(already_exist_item) > 0:
                                                tk.messagebox.showwarning("提示", "%s行号数据已存在，写入失败" % already_exist_item)
                                            else:
                                                for i in range(0, len(real_item_list.split(';'))):
                                                    data = {
                                                        'design_type': combobox_design_type.get(),
                                                        'project_number': entry_project_number.get(),
                                                        'item_info': int((real_item_list.split(';'))[i]),
                                                        'project_name': entry_project_name.get(),
                                                        'typical_amount': int(text_typical.get('1.0', 'end-1c')),
                                                        'panel_amount': real_amount_list[i],
                                                        'panel_type': real_product_type_list1[i],
                                                        'product_type': combobox_product_type.get(),
                                                        'intelligence': true_checkbox_intelligent_list[0:len(true_checkbox_intelligent_list) - 1].replace(' ', ''),
                                                        'location_info': combobox_location_type.get(),
                                                        'drawing_language': combobox_drawing_language_type.get(),
                                                        'standard_time': int(entry_calculate_time.get()),
                                                        'extra_time': int(entry_extra_time.get()),
                                                        'advice_times': int(combobox_advice_times.get()),
                                                        'upload_time': entry_upload_time.get(),
                                                        'bom_time': entry_bom_time.get(),
                                                        'check_time': entry_check_time.get(),
                                                        'create_time': Timestamp,
                                                        'project_type': combobox_project_type.get(),
                                                        'frame_type': combobox_frame_type.get(),
                                                        'program_type': combobox_program_type.get(),
                                                        'language_info': combobox_language_type.get(),
                                                        'management_info': combobox_manage_type.get(),
                                                        'input_file': true_checkbox_input_list[0:len(true_checkbox_input_list) - 1].replace(' ', ''),
                                                        'drawing_type': true_checkbox_drawing_type_list[0:len(true_checkbox_drawing_type_list) - 1].replace(' ', ''),
                                                        'mcb_type': combobox_MCB_type.get(),
                                                        'aux_type': combobox_aux_type.get(),
                                                        'terminal_type': combobox_terminal_type.get(),
                                                        'charged_display_type': combobox_charged_display_type.get(),
                                                        'switch_type': combobox_switch_type.get(),
                                                        'special_panel_config': content_table1,
                                                        'remark': content_table2
                                                    }
                                                    columns = ', '.join(data.keys())
                                                    placeholders = ', '.join(['?' for _ in range(len(data))])
                                                    query = f"INSERT INTO project_data ({columns}) VALUES ({placeholders})"

                                                    values = [data.get(col) for col in data.keys()]
                                                    cursor.execute(query, tuple(values))

                                                    conn.commit()
                                                tk.messagebox.showwarning("提示", "数据写入完成")
                                        cursor.close()
                                        conn.close()
                                else:
                                    tk.messagebox.showwarning("提示", "数据库不存在，请联系管理员")
                            else:
                                tk.messagebox.showwarning("提示", "请连接内网")

            except:
                tk.messagebox.showwarning("提示", traceback.format_exc())
                # tk.messagebox.showwarning("提示", "缺少模板Excel文件")
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def is_folder_hidden(fpath):
    try:
        attrs = ctypes.windll.kernel32.GetFileAttributesW(fpath)  # attrs值为18表示该文件夹具有以下属性组合：只读 (1)、隐藏 (2) 和 子文件夹 (16)
        # print(attrs)
        if attrs != -1 and attrs & 2 == 2:  # 对于18（二进制为10010）与2（二进制为00010）进行按位与运算，结果为2（二进制为00010）
            return True
    except OSError:
        pass
    return False


def about_help(event):
    # tku.show_info("说明书")
    os.startfile(os.path.abspath('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\原理图标准设计_red.pdf'))


