import tkinter as tk
from tkinter import ttk
from tkinter import StringVar
from tkinter import IntVar
from tkinter import filedialog
from tkinter.filedialog import askdirectory

from tkcalendar import CCalendar
import ctypes
import os
import logging

import shutil
from PIL import Image
from time import localtime
from time import strftime

from openpyxl import load_workbook

import warnings
from cryptography.fernet import Fernet
import base64
import traceback
import pyrfc
import math

import PyPDF2
import sqlite3

from reportlab.pdfgen import canvas as rl_canvas    # 为了防止冲突，为canvas重新命名
from reportlab.platypus import Image, Table
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import Color
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors

import json

warnings.simplefilter(action='ignore', category=FutureWarning)

FilePath = ""  # 设置一个地址变量
from need.custom_dialogs import CustomDialog, center_window, Tooltip, image_label

def main(parent, w_rat, h_rat):
    global query_project1
    query_project1 = tk.PhotoImage(file="ico\\search.png")
    global refresh_project
    refresh_project = tk.PhotoImage(file="ico\\refresh.png")
    global calculate_project
    calculate_project = tk.PhotoImage(file="ico\\calculator.png")
    global select_date_project
    select_date_project = tk.PhotoImage(file="ico\\calendar.png")
    global add_row_project
    add_row_project = tk.PhotoImage(file="ico\\add.png")
    global delete_row_project
    delete_row_project = tk.PhotoImage(file="ico\\minus.png")
    global save_draft
    save_draft = tk.PhotoImage(file="ico\\save.png")
    global view_standard_scheme
    view_standard_scheme = tk.PhotoImage(file="ico\\view.png")
    global pdf_export
    pdf_export = tk.PhotoImage(file="ico\\export.png")

    global w_ratio
    w_ratio = w_rat

    global h_ratio
    h_ratio = h_rat

    global canvas
    canvas = tk.Canvas(parent, width=int(1750 * w_ratio), height=int(640 * h_ratio), bg="#C9DBE9")
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    canvas.update()
    canvas.bind("<MouseWheel>", on_mousewheel)

    scrollbar_v = tk.Scrollbar(master=parent)
    scrollbar_v.pack(side=tk.RIGHT, fill=tk.Y)
    scrollbar_v.config(command=canvas.yview)
    Tooltip(scrollbar_v, "下滑滚动条")
    canvas.config(yscrollcommand=scrollbar_v.set)
    global content
    content = tk.Frame(canvas)
    # content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    canvas.create_window(0, 1, width=int(1750 * w_ratio), anchor=tk.NW, window=content)

    f1 = tk.Frame(content, bg="#c9dbe9", bd=0)
    # im = tku.image_label(f1, "ico\\help.png", int(30 * h_ratio), int(30 * h_ratio), False)
    # im.configure(bg="#c9dbe9")
    # im.bind('<Button-1>', about_help)  # 帮助图标绑定动作
    # im.pack(side=tk.RIGHT)
    tk.Label(f1, text="欢迎使用DesignList导出功能", bg="#c9dbe9", fg="black", height=int(1 * h_ratio), font=("ABBvoice CNSG", int(20 * h_ratio), "bold")).pack(fill=tk.X)
    f1.pack(fill=tk.X)

    # tk.Frame(content, height=int(20 * h_ratio), bg="white").pack(fill=tk.X)  # 水平分割线

    tk.Label(content, text='——————————————————————————————————————项目基本信息——————————————————————————————————————', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio), "bold"), justify='left').pack(side=tk.TOP, fill=tk.X)
    tk.Label(content, text='*为必填项，#为可选/多选项', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(10 * h_ratio)), justify='left').pack(side=tk.TOP, fill=tk.X)

    parent.option_add("*TCombobox*Listbox.font", ("ABBvoice CNSG", int(13 * h_ratio)))

    f200 = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Label(f200, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f200, text='设计需求：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global design_type_flag
    design_type_flag = 0
    global design_type_last
    design_type_last = ''
    global combobox_design_type
    combobox_design_type_value = StringVar()
    combobox_design_type_values = ['提前设计', '图纸设计', '工程设计']
    combobox_design_type = ttk.Combobox(master=f200, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(14 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_design_type_value, values=combobox_design_type_values)
    combobox_design_type.pack(side=tk.LEFT)
    combobox_design_type.bind("<<ComboboxSelected>>", on_select)
    f200.pack(fill=tk.X)
    Tooltip(f200, "*为必填项")

    tk.Frame(content, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f20 = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Label(f20, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f20, text='    项目号：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry_project_number
    entry_project_number = tk.Entry(f20, bg="white", font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(20*w_ratio))
    entry_project_number.pack(side=tk.LEFT)
    entry_project_number['state'] = 'disabled'
    entry_project_number.bind('<Return>', query_project)

    global button_query_project
    # tk.Label(f20, text='', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    button_query_project = tk.Button(master=f20, text='查询', font=("ABBvoice CNSG", int(10 * h_ratio)), bg="#eaf1f6", image=query_project1, compound=tk.LEFT, command=query_project, activebackground='blue')
    button_query_project.pack(side=tk.LEFT, padx=int(10 * w_ratio))
    button_query_project['state'] = 'disabled'

    # tk.Label(f20, text='  (需要连接内网)', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(10 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    f20.pack(fill=tk.X)
    Tooltip(f20, "*为必填项")

    tk.Frame(content, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f21 = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Label(f21, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, padx=0)
    tk.Label(f21, text='    项目名：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry_project_name
    entry_project_name = tk.Entry(f21, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(100*w_ratio))
    entry_project_name.pack(side=tk.LEFT)
    f21.pack(fill=tk.X)
    Tooltip(f21, "*为必填项")

    tk.Frame(content, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f22 = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Label(f22, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f22, text='行号信息：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global text_item
    text_item = tk.Text(f22, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), height=3)
    text_item.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    # text_item['state'] = 'disabled'
    Tooltip(text_item, "行号信息可手动删除，若有修改后，请点击右侧刷新按钮")

    global button_update_amount
    # tk.Label(f22, text='', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    button_update_amount = tk.Button(master=f22, text='刷新', font=("ABBvoice CNSG", int(10 * h_ratio)), bg="#eaf1f6", image=refresh_project, compound=tk.LEFT, command=update_amount, activebackground='blue')
    button_update_amount.pack(side=tk.LEFT, padx=int(10 * w_ratio))
    button_update_amount['state'] = 'disabled'

    tk.Label(f22, text='  (若移除部分行号，请点击刷新按钮，更新台数)', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(10 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, padx=(0, int(100 * w_ratio)))

    f22.pack(fill=tk.X)
    Tooltip(f22, "*为必填项")

    tk.Frame(content, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f28 = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Label(f28, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f28, text='典型图套：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global entry_typical
    entry_typical = tk.Entry(f28, bg="white", font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(14*w_ratio))
    entry_typical.pack(side=tk.LEFT)

    tk.Label(f28, text='          *', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f28, text='实际柜数：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry_amount
    entry_amount = tk.Entry(f28, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(14*w_ratio))
    entry_amount.pack(side=tk.LEFT)

    tk.Label(f28, text='           *', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f28, text='产品类型：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_product_type
    combobox_product_type_value = StringVar()
    combobox_product_type_values = ['AIS', 'GIS', 'AIS+GIS']
    combobox_product_type = ttk.Combobox(master=f28, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(24 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_product_type_value, values=combobox_product_type_values)
    combobox_product_type.pack(side=tk.LEFT)

    f28.pack(fill=tk.X)
    Tooltip(f28, "*为必填项")

    tk.Frame(content, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f2B = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Label(f2B, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2B, text='语言类型：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_drawing_language_type
    combobox_drawing_language_type_value = StringVar()
    combobox_drawing_language_type_values = ['中文', '英文', '中英文']
    combobox_drawing_language_type = ttk.Combobox(master=f2B, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(14 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_drawing_language_type_value, values=combobox_drawing_language_type_values)
    combobox_drawing_language_type.pack(side=tk.LEFT)

    tk.Label(f2B, text='    *', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2B, text='保护类型：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_protection_type
    combobox_protection_type_value = StringVar()
    combobox_protection_type_values = ['无保护', 'ABB保护', '第三方保护']
    combobox_protection_type = ttk.Combobox(master=f2B, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(14 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_protection_type_value, values=combobox_protection_type_values)
    combobox_protection_type.pack(side=tk.LEFT)

    tk.Label(f2B, text='    *', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2B, text='地区要求：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_location_type
    combobox_location_type_value = StringVar()
    combobox_location_type_values = ['UT万华', '港机', 'Building&数据中心', '工业/UT', '核电', '船用', '瑞典DOK80', '其他', '海外']
    combobox_location_type = ttk.Combobox(master=f2B, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(24 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_location_type_value, values=combobox_location_type_values)
    combobox_location_type.pack(side=tk.LEFT)

    tk.Label(f2B, text='    *', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2B, text='意见次数：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_advice_times
    combobox_advice_times_value = StringVar()
    combobox_advice_times_values = ['1', '2', '3', '4']
    combobox_advice_times = ttk.Combobox(master=f2B, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_advice_times_value, values=combobox_advice_times_values)
    combobox_advice_times.pack(side=tk.LEFT)

    f2B.pack(fill=tk.X)
    Tooltip(f2B, "*为必填项")

    tk.Frame(content, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f2J = tk.Frame(content, bg="#eaf1f6", bd=0)

    tk.Label(f2J, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2J, text='选开类型：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_switch_type
    combobox_switch_type_value = StringVar()
    combobox_switch_type_values = ['江阴长江', 'K&N', '其他(备注)']
    combobox_switch_type = ttk.Combobox(master=f2J, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(14 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_switch_type_value, values=combobox_switch_type_values)
    combobox_switch_type.pack(side=tk.LEFT)

    tk.Label(f2J, text='    *', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2J, text='端子类型：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_terminal_type
    combobox_terminal_type_value = StringVar()
    combobox_terminal_type_values = ['瑞联', 'TE', '魏德米勒', '菲尼克斯(*)']
    combobox_terminal_type = ttk.Combobox(master=f2J, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(14 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_terminal_type_value, values=combobox_terminal_type_values)
    combobox_terminal_type.pack(side=tk.LEFT)

    tk.Label(f2J, text='    *', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2J, text='空开类型：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_MCB_type
    combobox_MCB_type_value = StringVar()
    combobox_MCB_type_values = ['人民电器', 'ABB', '熔丝+熔芯']
    combobox_MCB_type = ttk.Combobox(master=f2J, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(14 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_MCB_type_value, values=combobox_MCB_type_values)
    combobox_MCB_type.pack(side=tk.LEFT)

    tk.Label(f2J, text='    *', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2J, text='辅助触点：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_aux_type
    combobox_aux_type_value = StringVar()
    combobox_aux_type_values = ['是', '否']
    combobox_aux_type = ttk.Combobox(master=f2J, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(4 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_aux_type_value, values=combobox_aux_type_values)
    combobox_aux_type.pack(side=tk.LEFT)

    f2J.pack(fill=tk.X)
    Tooltip(f2J, "*为必填项")

    tk.Frame(content, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f2D = tk.Frame(content, bg="#eaf1f6", bd=0)

    tk.Label(f2D, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2D, text='带显类型：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_charged_display_type
    combobox_charged_display_type_value = StringVar()
    combobox_charged_display_type_values = ['立林', '百岗']
    combobox_charged_display_type = ttk.Combobox(master=f2D, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(14 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_charged_display_type_value, values=combobox_charged_display_type_values)
    combobox_charged_display_type.pack(side=tk.LEFT)

    tk.Label(f2D, text='    *', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2D, text='   CT类型：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_ct_type
    combobox_ct_type_value = StringVar()
    combobox_ct_type_values = ['ABB', 'DYH', 'NTK', 'TLEP', '无', '其他']
    combobox_ct_type = ttk.Combobox(master=f2D, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(14 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_ct_type_value, values=combobox_ct_type_values)
    combobox_ct_type.pack(side=tk.LEFT)

    tk.Label(f2D, text='    *', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2D, text='  PT 类型：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_pt_type
    combobox_pt_type_value = StringVar()
    combobox_pt_type_values = ['ABB', 'DYH', 'NTK', 'TLEP', '无', '其他']
    combobox_pt_type = ttk.Combobox(master=f2D, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(14 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_pt_type_value, values=combobox_pt_type_values)
    combobox_pt_type.pack(side=tk.LEFT)

    tk.Label(f2D, text='    *', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2D, text='   SA类型：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_sa_type
    combobox_sa_type_value = StringVar()
    combobox_sa_type_values = ['神电', '日立', 'GCA', '其他']
    combobox_sa_type = ttk.Combobox(master=f2D, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(14 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_sa_type_value, values=combobox_sa_type_values)
    combobox_sa_type.pack(side=tk.LEFT)

    f2D.pack(fill=tk.X)
    Tooltip(f2D, "*为必填项")

    tk.Frame(content, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f2D = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Label(f2D, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2D, text='客户线号：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_customer_wiring
    combobox_customer_wiring_value = StringVar()
    combobox_customer_wiring_values = ['否', '是(备注前期项目号及需更改内容)']
    combobox_customer_wiring = ttk.Combobox(master=f2D, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(52 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_customer_wiring_value, values=combobox_customer_wiring_values)
    combobox_customer_wiring.pack(side=tk.LEFT)

    tk.Label(f2D, text='    *', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2D, text='复杂客户需求：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_customer_requirements
    combobox_customer_requirements_value = StringVar()
    combobox_customer_requirements_values = ['否', '是(备注前期项目号及需更改内容)']
    combobox_customer_requirements = ttk.Combobox(master=f2D, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(52 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_customer_requirements_value, values=combobox_customer_requirements_values)
    combobox_customer_requirements.pack(side=tk.LEFT)

    f2D.pack(fill=tk.X)

    tk.Frame(content, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f29 = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Label(f29, text='#', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f29, text='输入文件：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global checkbox_input_file
    global input_files
    input_files = ['Check List', '签字版单线图', '客户/设计院图纸', '参考项目信息(保护、产品一致，部分参考视为无参考)', '完全按照客户图纸设计']
    global checkbox_input_list
    checkbox_input_list = []
    for i in range(len(input_files)):
        v = IntVar()
        checkbox_input_file = tk.Checkbutton(f29, text=input_files[i], variable=v, font=("ABBvoice CNSG", int(13 * h_ratio)), height=1, background='#eaf1f6')
        checkbox_input_file.pack(side=tk.TOP, anchor=tk.W, expand=True)
        checkbox_input_list.append(v)
    f29.pack(fill=tk.X)
    Tooltip(f29, "#为可选项，单选/多选/不选")

    tk.Frame(content, height=int(5 * h_ratio), bg="whitesmoke").pack(fill=tk.X)  # 水平分割线

    f2A = tk.Frame(content, bg="#eaf1f6", bd=0)

    f2AR = tk.Frame(f2A, bg="#eaf1f6", bd=0)
    f2AR.pack(side=tk.LEFT, fill=tk.X)
    Tooltip(f2AR, "#为可选项，单选/多选/不选")

    f2AL = tk.Frame(f2A, bg="#eaf1f6", bd=0)
    f2AL.pack(side=tk.LEFT, fill=tk.X)
    Tooltip(f2AL, "#为可选项，单选/多选/不选")

    tk.Label(f2AL, text='                                              #', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2AL, text='图纸需求：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global checkbox_drawing_type
    global drawing_types
    drawing_types = ['SLD/FVD/FFD/FFDD/DS/S/E', 'SVD', 'C', 'AFD', 'LOGIC/拓扑图', '客户图框', 'AS BUILT一台柜子一套图', '非标铭牌图', 'BB/BC', '其他特殊图纸', 'AS BUILT其他语言(备注)']
    global checkbox_drawing_type_list
    checkbox_drawing_type_list = []
    for i in range(len(drawing_types)):
        v = IntVar()
        checkbox_drawing_type = tk.Checkbutton(f2AL, text=drawing_types[i], variable=v, font=("ABBvoice CNSG", int(13 * h_ratio)), height=1, background='#eaf1f6')
        checkbox_drawing_type.pack(side=tk.TOP, anchor=tk.W, expand=True)
        checkbox_drawing_type_list.append(v)

    tk.Label(f2A, text='SLD单线图; FVD开关柜排列图; FFD地基排列图; FFDD地基详图;        \nDS技术参数表; S原理图; E设备表; SVD截面图; C低压室布置&开孔图;\nAFD弧光保护布置图; BB/BC柜间小母线/柜间线图; AS BUILT竣工图  ', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(10 * h_ratio))).pack(side=tk.LEFT, fill=tk.X, anchor='nw', padx=10)

    tk.Label(f2AR, text='#', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2AR, text='智能方案：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global checkbox_intelligent
    global intelligent_needs
    intelligent_needs = ['断路器手车电操', '接地开关电操', '温升在线监测', '五防联锁监测', '断路器机械特性', '真空泡VI电寿命', '视频摄像头', '真空泡VI行程']
    global checkbox_intelligent_list
    checkbox_intelligent_list = []
    for i in range(len(intelligent_needs)):
        v = IntVar()
        checkbox_intelligent = tk.Checkbutton(f2AR, text=intelligent_needs[i], variable=v, font=("ABBvoice CNSG", int(13 * h_ratio)), height=1, background='#eaf1f6')
        checkbox_intelligent.pack(side=tk.TOP, anchor=tk.W, expand=True)
        checkbox_intelligent_list.append(v)
    f2A.pack(fill=tk.X)


    tk.Frame(content, height=int(5 * h_ratio), bg="whitesmoke").pack(fill=tk.X)  # 水平分割线
    tk.Frame(content, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f2G = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Label(f2G, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2G, text='标准天数：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry_calculate_time
    entry_calculate_time = tk.Entry(master=f2G, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(20 * w_ratio))
    entry_calculate_time.pack(side=tk.LEFT)
    entry_calculate_time['state'] = 'disabled'

    tk.Label(f2G, text='    *', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2G, text='额外天数：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry_extra_time
    entry_extra_time = tk.Entry(master=f2G, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(20 * w_ratio))
    entry_extra_time.pack(side=tk.LEFT)
    entry_extra_time['state'] = 'disabled'

    button_calculate_time = tk.Button(master=f2G, text='计算', font=("ABBvoice CNSG", int(10 * h_ratio)), bg="#eaf1f6", image=calculate_project, compound=tk.LEFT, command=calculate_time, activebackground='blue')
    button_calculate_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))

    tk.Label(f2G, text='  (根据上面的选项，计算时长)', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(10 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    # tk.Label(f2G, text='  (单位:天)', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(10 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    f2G.pack(fill=tk.X)
    Tooltip(f2G, "*为必填项")

    tk.Frame(content, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f2H = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Label(f2H, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2H, text='提原理图：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry_upload_time
    entry_upload_time = tk.Entry(master=f2H, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(20 * w_ratio))
    entry_upload_time.pack(side=tk.LEFT)
    entry_upload_time['state'] = 'disabled'

    global button_select_upload_time
    button_select_upload_time = tk.Button(master=f2H, text='选择', font=("ABBvoice CNSG", int(10 * h_ratio)), bg="#eaf1f6", image=select_date_project, compound=tk.LEFT, command=lambda: select_update_date(content), activebackground='blue')
    button_select_upload_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))
    button_select_upload_time['state'] = 'disabled'

    tk.Label(f2H, text='    *', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2H, text='发放初级：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry_bom_time
    entry_bom_time = tk.Entry(master=f2H, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(20 * w_ratio))
    entry_bom_time.pack(side=tk.LEFT)
    entry_bom_time['state'] = 'disabled'

    global button_select_bom_time
    button_select_bom_time = tk.Button(master=f2H, text='选择', font=("ABBvoice CNSG", int(10 * h_ratio)), bg="#eaf1f6", image=select_date_project, compound=tk.LEFT, command=lambda: select_bom_date(content), activebackground='blue')
    button_select_bom_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))
    button_select_bom_time['state'] = 'disabled'

    tk.Label(f2H, text='    *', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2H, text='生产检查：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry_check_time
    entry_check_time = tk.Entry(master=f2H, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(20 * w_ratio))
    entry_check_time.pack(side=tk.LEFT)
    entry_check_time['state'] = 'disabled'

    global button_select_check_time
    button_select_check_time = tk.Button(master=f2H, text='选择', font=("ABBvoice CNSG", int(10 * h_ratio)), bg="#eaf1f6", image=select_date_project, compound=tk.LEFT, command=lambda: select_check_date(content), activebackground='blue')
    button_select_check_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))
    button_select_check_time['state'] = 'disabled'

    f2H.pack(fill=tk.X)
    Tooltip(f2H, "*为必填项")

    tk.Frame(content, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f2N = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Label(f2N, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2N, text='项目类型：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_project_type
    combobox_project_type_value = StringVar()
    combobox_project_type_values = ['Domestic', 'Export_fully built', 'EPC', 'Metro', 'Marine', 'Wind power', '供电局项目(含18相反措)']
    combobox_project_type = ttk.Combobox(master=f2N, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(150 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_project_type_value, values=combobox_project_type_values)
    combobox_project_type.pack(side=tk.LEFT)

    f2N.pack(fill=tk.X)
    Tooltip(f2N, "*为必填项")

    tk.Frame(content, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f24 = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Label(f24, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f24, text='框架类型：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_frame_type
    combobox_frame_type_value = StringVar()
    combobox_frame_type_values = ['非框架项目', '重复或增补项目', '框架(阿里、中国移动、浙石化、泉化、万华、雄安、许继或大盛ATS Metro、上海小区变、港机等)']
    combobox_frame_type = ttk.Combobox(master=f24, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(150 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_frame_type_value, values=combobox_frame_type_values)
    combobox_frame_type.pack(side=tk.LEFT)
    f24.pack(fill=tk.X)
    Tooltip(f24, "*为必填项")

    tk.Frame(content, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f25 = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Label(f25, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f25, text='编程难度：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_program_type
    combobox_program_type_value = StringVar()
    combobox_program_type_values = ['不含保护、客供(ABB/非ABB)保护、repeat程序', 'ABB常用保护的国内/出口编程', 'ABB非常用保护编程', 'ABB保护编程+参与车间联调']
    combobox_program_type = ttk.Combobox(master=f25, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(150 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_program_type_value, values=combobox_program_type_values)
    combobox_program_type.pack(side=tk.LEFT)
    f25.pack(fill=tk.X)
    Tooltip(f25, "*为必填项")

    tk.Frame(content, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f26 = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Label(f26, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f26, text='语言难度：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_language_type
    combobox_language_type_value = StringVar()
    combobox_language_type_values = ['完全中文', '一般英语书面沟通', '英语书面+口语沟通', '复杂英语专业术语沟通']
    combobox_language_type = ttk.Combobox(master=f26, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(150 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_language_type_value, values=combobox_language_type_values)
    combobox_language_type.pack(side=tk.LEFT)
    f26.pack(fill=tk.X)
    Tooltip(f26, "*为必填项")

    tk.Frame(content, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f27 = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Label(f27, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f27, text='管理难度：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global combobox_manage_type
    combobox_manage_type_value = StringVar()
    combobox_manage_type_values = ['标准项目', '同一项目号下多种产品、首次生产新柜型项目', '复杂文档要求(客户要求模板)', '需要读复杂Spec文件、FAT要求复杂、客户特殊复杂的技术规格要求']
    combobox_manage_type = ttk.Combobox(master=f27, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(150 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_manage_type_value, values=combobox_manage_type_values)
    combobox_manage_type.pack(side=tk.LEFT)
    f27.pack(fill=tk.X)
    Tooltip(f27, "*为必填项")

    tk.Frame(content, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    # 典型柜配置表单
    tk.Label(content, text='———————————————————————————————————————典型柜配置——————————————————————————————————————', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio), "bold"), justify='left').pack(side=tk.TOP, fill=tk.X)
    tk.Label(content, text='如下表格列宽度控制在7、24、24个中文字符长度（1个中文字符=2英文字符）', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(10 * h_ratio)), justify='left').pack(side=tk.TOP, fill=tk.X)
    tk.Frame(content, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f2L = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Frame(f2L, bg="#eaf1f6", bd=0).pack(side=tk.LEFT, fill=tk.X, padx=0)

    f2L_Left = tk.Frame(f2L, bg="#eaf1f6", bd=0)
    f2L_Left.pack(side=tk.LEFT, fill=tk.X)
    Tooltip(f2L_Left, "可根据实际加/减表格行数")

    global f2L_Right
    f2L_Right = tk.Frame(f2L, bg="#eaf1f6", bd=0)
    f2L_Right.pack(side=tk.LEFT, fill=tk.X)
    Tooltip(f2L_Right, "注意精简，长度不要过长，否则影响导出的报表显示")

    global num_rows1
    num_rows1 = 6
    global num_columns1
    num_columns1 = 3
    global entry_widths1
    entry_widths1 = [int(18*w_ratio), int(60*w_ratio), int(60*w_ratio)]
    global entries1
    entries1 = [[None for _ in range(num_columns1)] for _ in range(num_rows1)]

    for row in range(num_rows1):
        row_frame1 = tk.Frame(f2L_Right)
        row_frame1.pack(side=tk.TOP, fill=tk.X)

        for col in range(num_columns1):
            entry1 = tk.Entry(row_frame1, width=entry_widths1[col], relief='solid', font=("ABBvoice CNSG", int(13 * h_ratio)))
            entry1.pack(side=tk.LEFT)
            entries1[row][col] = entry1
            entry1.config(justify='center')
            if row == 0 and col == 0:
                entry1.insert(0, '柜号/Typical')
                entry1['state'] = 'disabled'
            elif row == 0 and col == 1:
                entry1.insert(0, '保护料号/建号提单号')
                entry1['state'] = 'disabled'
            elif row == 0 and col == 2:
                entry1.insert(0, '多功能表、电度表料号/建号提单号')
                entry1['state'] = 'disabled'

    add_row_button1 = tk.Button(f2L_Left, image=add_row_project, bg="#eaf1f6", command=lambda: add_row1(content, f2L_Right, canvas, h_ratio), activebackground='blue')
    add_row_button1.pack(side=tk.TOP, pady=(0, int(10*h_ratio)), padx=(int(35*w_ratio), int(35*w_ratio)))

    delete_row_button1 = tk.Button(f2L_Left, image=delete_row_project, bg="#eaf1f6", command=lambda: delete_row1(content, f2L_Right, canvas), activebackground='blue')
    delete_row_button1.pack(side=tk.TOP, pady=(int(10*h_ratio), 0), padx=(int(35*w_ratio), int(35*w_ratio)))

    f2L.pack(fill=tk.X)

    tk.Frame(content, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    # 备注表单
    tk.Label(content, text='————————————————————————————————————————备注————————————————————————————————————————', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio), "bold"), justify='left').pack(side=tk.TOP, fill=tk.X)
    tk.Label(content, text='如下表格列宽度控制在54个中文字符长度（1个中文字符=2英文字符）', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(10 * h_ratio)), justify='left').pack(side=tk.TOP, fill=tk.X)
    tk.Frame(content, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f2M = tk.Frame(content, bg="#eaf1f6", bd=0)
    tk.Frame(f2M, bg="#eaf1f6", bd=0).pack(side=tk.LEFT, fill=tk.X, padx=0)

    f2M_Left = tk.Frame(f2M, bg="#eaf1f6", bd=0)
    f2M_Left.pack(side=tk.LEFT, fill=tk.X)
    Tooltip(f2M_Left, "可根据实际加/减表格行数")

    global f2M_Right
    f2M_Right = tk.Frame(f2M, bg="#eaf1f6", bd=0)
    f2M_Right.pack(side=tk.LEFT, fill=tk.X)
    Tooltip(f2M_Right, "注意精简，长度不要过长，否则影响导出的报表显示")

    global num_rows2
    num_rows2 = 6
    global num_columns2
    num_columns2 = 2
    global entry_widths2
    entry_widths2 = [int(3*w_ratio), int(135*w_ratio)]
    global entries2
    entries2 = [[None for _ in range(num_columns2)] for _ in range(num_rows2)]

    for row in range(num_rows2):
        row_frame2 = tk.Frame(f2M_Right)
        row_frame2.pack(side=tk.TOP, fill=tk.X)

        for col in range(num_columns2):
            entry2 = tk.Entry(row_frame2, width=entry_widths2[col], relief='solid', font=("ABBvoice CNSG", int(13 * h_ratio)))
            entry2.pack(side=tk.LEFT)
            if col == 0:
                entry2.config(justify='center')
                entry2.insert(0, '%s' % str(row + 1))
                entry2['state'] = 'disabled'
            entries2[row][col] = entry2

    add_row_button2 = tk.Button(f2M_Left, image=add_row_project, bg="#eaf1f6", command=lambda: add_row2(content, f2M_Right, canvas, h_ratio), activebackground='blue')
    add_row_button2.pack(side=tk.TOP, pady=(0, int(10*h_ratio)), padx=(int(35*w_ratio), int(35*w_ratio)))

    delete_row_button2 = tk.Button(f2M_Left, image=delete_row_project, bg="#eaf1f6", command=lambda: delete_row2(content, f2M_Right, canvas), activebackground='blue')
    delete_row_button2.pack(side=tk.TOP, pady=(int(10*h_ratio), 0), padx=(int(35*w_ratio), int(35*w_ratio)))

    f2M.pack(fill=tk.X)

    f2P = tk.Frame(content, bg="#eaf1f6", bd=0)
    save_button = tk.Button(f2P, image=save_draft, bg="#eaf1f6", text="保存已填内容", compound=tk.LEFT, command=save_state, font=("ABBvoice CNSG", int(13 * h_ratio)), activebackground='blue')
    save_button.pack(side=tk.LEFT, padx=int(220*w_ratio), pady=(int(30*w_ratio), int(40*w_ratio)))

    read_button = tk.Button(f2P, image=view_standard_scheme, bg="#eaf1f6", text="标准方案", compound=tk.LEFT, command=read_design_details, font=("ABBvoice CNSG", int(13 * h_ratio)), activebackground='blue')
    read_button.pack(side=tk.LEFT, padx=0, pady=(int(30*w_ratio), int(40*w_ratio)))

    global checkbox_read, checkbox_v
    checkbox_v = IntVar()
    checkbox_read = tk.Checkbutton(f2P, text='已阅', variable=checkbox_v, font=("ABBvoice CNSG", int(13 * h_ratio)), height=1, background='#eaf1f6')
    checkbox_read.pack(side=tk.LEFT, padx=10, pady=(int(30*w_ratio), int(40*w_ratio)))
    checkbox_read.bind('<Button-1>', checkbox_check)

    global export_button
    export_button = tk.Button(f2P, image=pdf_export, bg="#eaf1f6", text='导出DesignList', compound=tk.LEFT, command=export_transfer_sheet, font=("ABBvoice CNSG", int(13 * h_ratio)), activebackground='blue')
    export_button.pack(side=tk.LEFT, padx=int(200*w_ratio), pady=(int(30*w_ratio), int(40*w_ratio)))
    export_button['state'] = 'disabled'

    f2P.pack(fill=tk.X, expand=True)
    canvas.update_idletasks()
    # content.update_idletasks()
    canvas.config(scrollregion=canvas.bbox('all'))

    if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools'):
        tk.messagebox.showwarning("提示", "请连接内网")
    else:
        parent.after(100, lambda:load_state(h_ratio))


def on_mousewheel(event):
    global canvas
    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
    # print(canvas.winfo_width())
    # canvas.create_text(canvas.winfo_width() - 10, 10, anchor='ne', text='滚动区↓')


def on_select(event):
    global design_type_flag
    global design_type_last

    selected_value = combobox_design_type.get()
    if selected_value == '提前设计':
        entry_project_number['state'] = 'normal'
        entry_project_number.delete(0, "end")
        entry_project_number.insert(0, '提前设计')
        entry_project_number['state'] = 'disabled'

        button_query_project['state'] = 'disabled'
        button_update_amount['state'] = 'disabled'

        entry_project_name['state'] = 'normal'
        entry_project_name.delete(0, "end")
        entry_project_name['bg'] = 'white'
        text_item['state'] = 'normal'
        text_item.delete(1.0, tk.END)

        text_item['background'] = 'whitesmoke'
        text_item.insert(tk.END, '1000')
        text_item['state'] = 'disabled'

        entry_typical.delete(0, "end")
        entry_amount.delete(0, "end")
        entry_amount['bg'] = 'white'
        combobox_product_type.set('')

        combobox_location_type.set('')
        combobox_protection_type.set('')
        combobox_drawing_language_type.set('')

        # text_drawing_language['state'] = 'normal'
        # text_drawing_language.delete(1.0, tk.END)
        # text_drawing_language['background'] = 'whitesmoke'
        # text_drawing_language['state'] = 'disabled'

        combobox_customer_wiring.set('')

        # text_customer_wiring['state'] = 'normal'
        # text_customer_wiring.delete(1.0, tk.END)
        # text_customer_wiring['background'] = 'whitesmoke'
        # text_customer_wiring['state'] = 'disabled'

        combobox_customer_requirements.set('')

        # text_customer_requirements['state'] = 'normal'
        # text_customer_requirements.delete(1.0, tk.END)
        # text_customer_requirements['background'] = 'whitesmoke'
        # text_customer_requirements['state'] = 'disabled'

        combobox_MCB_type.set('人民电器')
        combobox_aux_type.set('否')
        combobox_terminal_type.set('瑞联')
        # combobox_charged_display_type.set('立林')
        combobox_switch_type.set('江阴长江')
        combobox_advice_times.set('2')
        # text_switch['state'] = 'normal'
        # text_switch.delete(1.0, tk.END)
        # text_switch['background'] = 'whitesmoke'
        # text_switch['state'] = 'disabled'

        combobox_project_type.set('')
        combobox_frame_type.set('')
        combobox_program_type.set('')
        combobox_language_type.set('')
        combobox_manage_type.set('')

        entry_bom_time['state'] = 'disabled'
        button_select_bom_time['state'] = 'disabled'
        design_type_flag = 1

        if selected_value != design_type_last:
            entry_calculate_time['state'] = 'normal'
            entry_calculate_time.delete(0, "end")
            entry_calculate_time['state'] = 'disabled'
            entry_extra_time['state'] = 'normal'
            entry_extra_time.delete(0, "end")
            entry_extra_time['state'] = 'disabled'
            entry_upload_time['state'] = 'normal'
            entry_upload_time.delete(0, "end")
            entry_upload_time['state'] = 'disabled'
            entry_bom_time['state'] = 'normal'
            entry_bom_time.delete(0, "end")
            entry_bom_time['state'] = 'disabled'
            entry_check_time['state'] = 'normal'
            entry_check_time.delete(0, "end")
            entry_check_time['state'] = 'disabled'

            button_select_upload_time['state'] = 'disabled'
            button_select_bom_time['state'] = 'disabled'
            button_select_check_time['state'] = 'disabled'

    elif selected_value == '图纸设计':
        entry_project_number['state'] = 'normal'
        entry_project_number.delete(0, "end")

        button_query_project['state'] = 'normal'
        button_update_amount['state'] = 'disabled'

        entry_project_name['state'] = 'normal'
        entry_project_name.delete(0, "end")
        entry_project_name['bg'] = '#eaf1f6'
        text_item['state'] = 'normal'
        text_item.delete(1.0, tk.END)
        text_item['background'] = 'white'

        entry_typical.delete(0, "end")
        entry_amount.delete(0, "end")
        entry_amount['bg'] = '#eaf1f6'
        combobox_product_type.set('')

        combobox_location_type.set('')
        combobox_protection_type.set('')
        combobox_drawing_language_type.set('')

        # text_drawing_language['state'] = 'normal'
        # text_drawing_language.delete(1.0, tk.END)
        # text_drawing_language['background'] = 'whitesmoke'
        # text_drawing_language['state'] = 'disabled'

        combobox_customer_wiring.set('')

        # text_customer_wiring['state'] = 'normal'
        # text_customer_wiring.delete(1.0, tk.END)
        # text_customer_wiring['background'] = 'whitesmoke'
        # text_customer_wiring['state'] = 'disabled'

        combobox_customer_requirements.set('')

        # text_customer_requirements['state'] = 'normal'
        # text_customer_requirements.delete(1.0, tk.END)
        # text_customer_requirements['background'] = 'whitesmoke'
        # text_customer_requirements['state'] = 'disabled'

        combobox_MCB_type.set('人民电器')
        combobox_aux_type.set('否')
        combobox_terminal_type.set('瑞联')
        # combobox_charged_display_type.set('立林')
        combobox_switch_type.set('江阴长江')
        combobox_advice_times.set('2')
        # text_switch['state'] = 'normal'
        # text_switch.delete(1.0, tk.END)
        # text_switch['background'] = 'whitesmoke'
        # text_switch['state'] = 'disabled'

        combobox_project_type.set('')
        combobox_frame_type.set('')
        combobox_program_type.set('')
        combobox_language_type.set('')
        combobox_manage_type.set('')

        entry_bom_time['state'] = 'disabled'
        button_select_bom_time['state'] = 'normal'
        design_type_flag = 2

        if selected_value != design_type_last:
            entry_calculate_time['state'] = 'normal'
            entry_calculate_time.delete(0, "end")
            entry_calculate_time['state'] = 'disabled'
            entry_extra_time['state'] = 'normal'
            entry_extra_time.delete(0, "end")
            entry_extra_time['state'] = 'disabled'
            entry_upload_time['state'] = 'normal'
            entry_upload_time.delete(0, "end")
            entry_upload_time['state'] = 'disabled'
            entry_bom_time['state'] = 'normal'
            entry_bom_time.delete(0, "end")
            entry_bom_time['state'] = 'disabled'
            entry_check_time['state'] = 'normal'
            entry_check_time.delete(0, "end")
            entry_check_time['state'] = 'disabled'

            button_select_upload_time['state'] = 'disabled'
            button_select_bom_time['state'] = 'disabled'
            button_select_check_time['state'] = 'disabled'

    elif selected_value == '工程设计':
        entry_project_number['state'] = 'normal'
        entry_project_number.delete(0, "end")

        button_query_project['state'] = 'normal'
        button_update_amount['state'] = 'disabled'

        entry_project_name['state'] = 'normal'
        entry_project_name.delete(0, "end")
        entry_project_name['bg'] = '#eaf1f6'

        text_item['state'] = 'normal'
        text_item.delete(1.0, tk.END)
        text_item['background'] = 'white'

        entry_typical.delete(0, "end")
        entry_amount.delete(0, "end")
        entry_amount['bg'] = '#eaf1f6'
        combobox_product_type.set('')

        combobox_location_type.set('')
        combobox_protection_type.set('')
        combobox_drawing_language_type.set('')

        # text_drawing_language['state'] = 'normal'
        # text_drawing_language.delete(1.0, tk.END)
        # text_drawing_language['background'] = 'whitesmoke'
        # text_drawing_language['state'] = 'disabled'

        combobox_customer_wiring.set('')

        # text_customer_wiring['state'] = 'normal'
        # text_customer_wiring.delete(1.0, tk.END)
        # text_customer_wiring['background'] = 'whitesmoke'
        # text_customer_wiring['state'] = 'disabled'

        combobox_customer_requirements.set('')

        # text_customer_requirements['state'] = 'normal'
        # text_customer_requirements.delete(1.0, tk.END)
        # text_customer_requirements['background'] = 'whitesmoke'
        # text_customer_requirements['state'] = 'disabled'

        combobox_MCB_type.set('人民电器')
        combobox_aux_type.set('否')
        combobox_terminal_type.set('瑞联')
        # combobox_charged_display_type.set('立林')
        combobox_switch_type.set('江阴长江')
        combobox_advice_times.set('2')
        # text_switch['state'] = 'normal'
        # text_switch.delete(1.0, tk.END)
        # text_switch['background'] = 'whitesmoke'
        # text_switch['state'] = 'disabled'

        combobox_project_type.set('')
        combobox_frame_type.set('')
        combobox_program_type.set('')
        combobox_language_type.set('')
        combobox_manage_type.set('')

        entry_bom_time['state'] = 'disabled'
        button_select_bom_time['state'] = 'normal'
        design_type_flag = 3

        if selected_value != design_type_last:
            entry_calculate_time['state'] = 'normal'
            entry_calculate_time.delete(0, "end")
            entry_calculate_time['state'] = 'disabled'
            entry_extra_time['state'] = 'normal'
            entry_extra_time.delete(0, "end")
            entry_extra_time['state'] = 'disabled'
            entry_upload_time['state'] = 'normal'
            entry_upload_time.delete(0, "end")
            entry_upload_time['state'] = 'disabled'
            entry_bom_time['state'] = 'normal'
            entry_bom_time.delete(0, "end")
            entry_bom_time['state'] = 'disabled'
            entry_check_time['state'] = 'normal'
            entry_check_time.delete(0, "end")
            entry_check_time['state'] = 'disabled'

            button_select_upload_time['state'] = 'disabled'
            button_select_bom_time['state'] = 'disabled'
            button_select_check_time['state'] = 'disabled'

    design_type_last = selected_value


def checkbox_check(event):
    if checkbox_v.get() == 0:
        export_button['state'] = 'normal'
    else:
        export_button['state'] = 'disabled'


def query_project(event=''):
    entry_project_name['state'] = 'normal'
    entry_project_name.delete(0, "end")
    try:
        if entry_project_number.get() == "" or len(entry_project_number.get()) != 9:
            tk.messagebox.showwarning("提示", "请输入9位项目号")
        elif not entry_project_number.get().isdigit():
            tk.messagebox.showwarning("提示", "项目号必须是9位纯数字")
        else:
            try:
                workbook1 = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\SR_Material.xlsx")
                worksheet1 = workbook1['Sheet1']
            except:
                tk.messagebox.showwarning("提示", "缺少产品型号表")

            entry_calculate_time['state'] = 'normal'
            entry_calculate_time.delete(0, "end")
            entry_calculate_time['state'] = 'disabled'
            entry_extra_time['state'] = 'normal'
            entry_extra_time.delete(0, "end")
            entry_extra_time['state'] = 'disabled'
            entry_upload_time['state'] = 'normal'
            entry_upload_time.delete(0, "end")
            entry_upload_time['state'] = 'disabled'
            entry_bom_time['state'] = 'normal'
            entry_bom_time.delete(0, "end")
            entry_bom_time['state'] = 'disabled'
            entry_check_time['state'] = 'normal'
            entry_check_time.delete(0, "end")
            entry_check_time['state'] = 'disabled'

            button_update_amount['state'] = 'disabled'
            button_select_upload_time['state'] = 'disabled'
            button_select_bom_time['state'] = 'disabled'
            button_select_check_time['state'] = 'disabled'

            global item_list
            item_list = []
            global amount_list
            amount_list = []

            amount_panel = 0

            global product_type_list1
            product_type_list1 = []

            global product_type_list2
            product_type_list2 = []

            text_item['state'] = 'normal'
            text_item.delete(1.0, tk.END)
            entry_amount.delete(0, "end")

            worksheet1_max_row = max(ee.row for ee in worksheet1['A'] if ee.value)
            global material_list
            material_list = []
            global product_type_list
            product_type_list = []

            for i in range(1, worksheet1_max_row + 1):
                material_list.append(worksheet1.cell(row=i, column=1).value)
                product_type_list.append(worksheet1.cell(row=i, column=3).value)

            # SAP登录
            sap_key_file = 'J:/Engineering/ShareFolder/new_ABB_Production_Tools/Ps/k/&%&^RD&^T&^WRE^@FEYUGYU@^E@DRYTRYTDFYTWF.txt'
            if not os.path.exists(sap_key_file):
                tk.messagebox.showwarning('提示', 'SAP加密密钥文件不存在')
            else:
                with open(sap_key_file, 'rb') as file:
                    sap_key = file.read()
            sap_cipher = Fernet(sap_key)

            sap_ciphertext_file = 'J:/Engineering/ShareFolder/new_ABB_Production_Tools/Ps/s/^&%&^R^YFT$%FTDTEYTKFY$^$FGKJGHE%#$%#EYD^%#^%DTYD.txt'
            if not os.path.exists(sap_ciphertext_file):
                tk.messagebox.showwarning('提示', 'SAP加密密文文件不存在')
            else:
                with open(sap_ciphertext_file, 'r') as file:
                    sap_ciphertext = file.read()

            sap_account = sap_decrypt_data(sap_ciphertext, sap_cipher, 'z_')
            sap_h = sap_account.split('+')[0]
            sap_c = sap_account.split('+')[1]
            sap_s = sap_account.split('+')[2]
            sap_u = sap_account.split('+')[3]
            sap_p = sap_account.split('+')[4]

            logging.info("Ready to connect to SAP")
            try:
                conn = pyrfc.Connection(ashost=sap_h, sysnr=sap_s, client=sap_c, user=sap_u, passwd=sap_p, lang='EN')
                if conn.alive:
                    logging.info("Connecting to SAP successfully")

                    result = conn.call('ZY_SALES_ORDER_SHIFT', VBELN='0' + entry_project_number.get())
                    # print(result)
                    # if result['EX_CEPTION'] == '' and result['ITAB'][0]['VKBUR'] == '1200':
                    if result['EX_CEPTION'] == '' and result['ITAB'][0]['WERKS'] == '1201':
                        project_name = result['ITAB'][0]['BSTKD']
                        entry_project_name.insert(0, project_name)
                        entry_project_name['state'] = 'disabled'

                        table_content = []
                        # 处理数据
                        for item in result['ITAB']:
                            posnr = item['POSNR'].lstrip('0') or '0'
                            kwmeng = str(int(float(item['KWMENG'])))
                            table_content.append((posnr, item['MATNR'], kwmeng))    # [('1000','UNIGEAR-ZS1-500', '14'), ('2000','UNIGEAR-ZS1', '6')]

                        for j in range(0, len(table_content)):
                            for k in range(0, len(material_list)):
                                if table_content[j][1] == material_list[k]:
                                    item_list.append(int(table_content[j][0]))
                                    amount_list.append(int(table_content[j][2]))
                                    amount_panel += int(table_content[j][2])
                                    product_type_list1.append(material_list[k])
                                    product_type_list2.append(product_type_list[k])
                                    break
                        item_list_copy = item_list.copy()
                        item_list.sort()

                        amount_list = [amount_list[i] for i in sorted(range(len(item_list_copy)), key=lambda x: item_list_copy[x])]
                        product_type_list1 = [product_type_list1[i] for i in sorted(range(len(item_list_copy)), key=lambda x: item_list_copy[x])]
                        product_type_list2 = [product_type_list2[i] for i in sorted(range(len(item_list_copy)), key=lambda x: item_list_copy[x])]

                        # print(item_list, amount_list, product_type_list1, product_type_list2)
                        text_item.insert(tk.END, ';'.join(str(num) for num in item_list))

                        entry_amount.insert(0, amount_panel)

                        if set(product_type_list2) == {'AIS'}:
                            combobox_product_type.set('AIS')
                        elif set(product_type_list2) == {'GIS'}:
                            combobox_product_type.set('GIS')
                        else:
                            combobox_product_type.set('AIS+GIS')
                        if len(item_list) > 1:
                            button_update_amount['state'] = 'normal'

                        update_amount()  # 每次读取完刷新一下

                    else:
                        if result['EX_CEPTION'] != '':
                            tk.messagebox.showwarning("提示", result['EX_CEPTION'])
                        else:
                            tk.messagebox.showwarning("提示", 'CNDMX的1201无此项目')

                    conn.close()
                    if not conn.alive:
                        logging.info("Disconnect from SAP")


            except pyrfc.RFCError as e:
                logging.info(e.key + ', ' + e.message)
                tk.messagebox.showwarning("提示", traceback.format_exc())


    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def sap_decrypt_data(data, data_cipher, prefix):
    # 移除列名前缀prefix，如果有的话
    if data.startswith(prefix):
        data = data[2:]

    # Base64 字符串中可能删除了等号，需要添加回原来的填充字符，因为 base64 编码的输出长度应该是 4 的倍数
    padding_needed = 4 - (len(data) % 4)
    if padding_needed:
        data += "=" * padding_needed

    # base64 解码以获取原始的加密字节串
    encrypted_data = base64.urlsafe_b64decode(data.encode())

    # 使用 data_cipher 对象解密，它应该有和加密时一样的密钥
    decrypted_data = data_cipher.decrypt(encrypted_data)

    # 将字节解码回字符串
    decrypted_string = decrypted_data.decode()
    return decrypted_string


def update_amount():
    global real_item_list    # CI
    real_item_list = []
    global real_product_type_list1    # 产品类型
    real_product_type_list1 = []
    global real_product_type_list2    # 对应的产品类型（AIS,GIS）
    real_product_type_list2 = []
    global real_amount_list    # 每个CI数量
    real_amount_list = []
    global real_amount_panel    # 所有CI总数量
    real_amount_panel = 0
    temp = str(text_item.get("1.0", "end-1c")).replace('；', ';')
    if temp[len(temp)-1] == ';':
        temp = temp[0:len(temp)-1]

    real_item_list = temp
    # print(temp, temp.split(';'))

    for i in range(0, len(item_list)):
        if str(item_list[i]) in temp.split(';'):
            real_product_type_list1.append(product_type_list1[i])
            real_product_type_list2.append(product_type_list2[i])
            real_amount_list.append(amount_list[i])
            real_amount_panel += amount_list[i]
    entry_amount.delete(0, "end")
    entry_amount.insert(0, real_amount_panel)

    if set(real_product_type_list2) == {'AIS'}:
        combobox_product_type.set('AIS')
    elif set(real_product_type_list2) == {'GIS'}:
        combobox_product_type.set('GIS')
    else:
        combobox_product_type.set('AIS+GIS')
    # print(real_item_list, real_product_type_list1, real_amount_list)


def calculate_time():
    try:
        workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\Convert_Coefficient.xlsx")
        worksheet = workbook['Sheet1']
    except:
        tk.messagebox.showwarning("提示", "缺少转换系数表")

    if combobox_design_type.get() == '' or combobox_product_type.get() == '' or entry_typical.get() == '' or entry_amount.get() == '' \
            or combobox_location_type.get() == '' or combobox_protection_type.get() == '' or combobox_drawing_language_type.get() == '' \
            or combobox_customer_wiring.get() == '' or combobox_customer_requirements.get() == '':
        tk.messagebox.showwarning("提示", "*标属性未填写完整")

    elif not entry_typical.get().isdigit():
        tk.messagebox.showwarning("提示", "典型图数应为纯数字")
    elif not entry_amount.get().isdigit():
        tk.messagebox.showwarning("提示", "台数应为纯数字")
    elif int(entry_typical.get()) > int(entry_amount.get()):
        tk.messagebox.showwarning("提示", "图套数不能大于台数")
    else:
        combobox_design_type_coff = 1
        for i in range(2, 5):
            if combobox_design_type.get() == worksheet.cell(row=i, column=1).value:
                combobox_design_type_coff = float(worksheet.cell(row=i, column=2).value)
                break

        combobox_product_type_coff = 1
        for i in range(2, 5):
            if combobox_product_type.get() == worksheet.cell(row=i, column=3).value:
                combobox_product_type_coff = float(worksheet.cell(row=i, column=4).value)
                break

        if float(entry_typical.get()) <= worksheet.cell(row=2, column=5).value:
            entry_typical_coff = float(worksheet.cell(row=2, column=6).value)
        elif worksheet.cell(row=2, column=5).value < float(entry_typical.get()) <= worksheet.cell(row=3, column=5).value:
            entry_typical_coff = float(worksheet.cell(row=3, column=6).value)
        elif worksheet.cell(row=3, column=5).value < float(entry_typical.get()) <= worksheet.cell(row=4, column=5).value:
            entry_typical_coff = float(worksheet.cell(row=4, column=6).value)
        elif worksheet.cell(row=4, column=5).value < float(entry_typical.get()):
            entry_typical_coff = float(worksheet.cell(row=5, column=6).value)

        if float(entry_amount.get()) <= worksheet.cell(row=2, column=7).value:
            entry_amount_coff = float(worksheet.cell(row=2, column=8).value)
        elif worksheet.cell(row=2, column=7).value < float(entry_amount.get()) <= worksheet.cell(row=3, column=7).value:
            entry_amount_coff = float(worksheet.cell(row=3, column=8).value)
        elif worksheet.cell(row=3, column=7).value < float(entry_amount.get()):
            entry_amount_coff = float(worksheet.cell(row=4, column=8).value)

        checkbox_input_file_coff = 1
        if checkbox_input_list[0].get():
            checkbox_input_file_coff1 = float(worksheet.cell(row=2, column=10).value)
        else:
            checkbox_input_file_coff1 = 1.5

        if checkbox_input_list[1].get():
            checkbox_input_file_coff2 = float(worksheet.cell(row=3, column=10).value)
        else:
            checkbox_input_file_coff2 = 1.05

        if checkbox_input_list[2].get():
            checkbox_input_file_coff3 = float(worksheet.cell(row=4, column=10).value)
        else:
            checkbox_input_file_coff3 = 1

        if checkbox_input_list[3].get():
            checkbox_input_file_coff4 = float(worksheet.cell(row=5, column=10).value)
        else:
            checkbox_input_file_coff4 = 1

        if checkbox_input_list[4].get():
            checkbox_input_file_coff5 = float(worksheet.cell(row=6, column=10).value)
        else:
            checkbox_input_file_coff5 = 1
        checkbox_input_file_coff = checkbox_input_file_coff * checkbox_input_file_coff1 * checkbox_input_file_coff2 \
                                   * checkbox_input_file_coff3 * checkbox_input_file_coff4 * checkbox_input_file_coff5

        checkbox_drawing_type_coff = 1
        extra_period = 0

        if combobox_design_type.get() == '提前设计':
            item_length = 1
        else:
            item_length = len(real_amount_list)

        if checkbox_drawing_type_list[0].get():
            checkbox_drawing_type_coff1 = float(worksheet.cell(row=2, column=12).value)
            extra_period += 0
        else:
            checkbox_drawing_type_coff1 = 1

        if checkbox_drawing_type_list[1].get():
            checkbox_drawing_type_coff2 = float(worksheet.cell(row=3, column=12).value)
            extra_period += float(worksheet.cell(row=3, column=13).value)*item_length
        else:
            checkbox_drawing_type_coff2 = 1

        if checkbox_drawing_type_list[2].get():
            checkbox_drawing_type_coff3 = float(worksheet.cell(row=4, column=12).value)
            extra_period += float(worksheet.cell(row=4, column=13).value) * float(entry_typical.get())
        else:
            checkbox_drawing_type_coff3 = 1

        if checkbox_drawing_type_list[3].get():
            checkbox_drawing_type_coff4 = float(worksheet.cell(row=5, column=12).value)
            extra_period += float(worksheet.cell(row=5, column=13).value) * item_length
        else:
            checkbox_drawing_type_coff4 = 1

        if checkbox_drawing_type_list[4].get():
            checkbox_drawing_type_coff5 = float(worksheet.cell(row=6, column=12).value)
            extra_period += float(worksheet.cell(row=6, column=13).value) * item_length
        else:
            checkbox_drawing_type_coff5 = 1

        if checkbox_drawing_type_list[5].get():
            checkbox_drawing_type_coff6 = float(worksheet.cell(row=7, column=12).value)
            extra_period += float(worksheet.cell(row=7, column=13).value) * float(entry_typical.get())
        else:
            checkbox_drawing_type_coff6 = 1

        if checkbox_drawing_type_list[6].get():
            checkbox_drawing_type_coff7 = float(worksheet.cell(row=8, column=12).value)
            extra_period += float(worksheet.cell(row=8, column=13).value) * float(entry_amount.get())
        else:
            checkbox_drawing_type_coff7 = 1

        if checkbox_drawing_type_list[7].get():
            checkbox_drawing_type_coff8 = float(worksheet.cell(row=9, column=12).value)
            extra_period += float(worksheet.cell(row=9, column=13).value) * item_length
        else:
            checkbox_drawing_type_coff8 = 1

        if checkbox_drawing_type_list[8].get():
            checkbox_drawing_type_coff9 = float(worksheet.cell(row=10, column=12).value)
            extra_period += float(worksheet.cell(row=10, column=13).value) * item_length
        else:
            checkbox_drawing_type_coff9 = 1

        if checkbox_drawing_type_list[9].get():
            checkbox_drawing_type_coff10 = float(worksheet.cell(row=11, column=12).value)
            extra_period += float(worksheet.cell(row=11, column=13).value) * item_length
        else:
            checkbox_drawing_type_coff10 = 1

        if checkbox_drawing_type_list[10].get():
            checkbox_drawing_type_coff11 = float(worksheet.cell(row=12, column=12).value)
            extra_period += float(worksheet.cell(row=12, column=13).value) * float(entry_typical.get())
        else:
            checkbox_drawing_type_coff11 = 1

        checkbox_drawing_type_coff = checkbox_drawing_type_coff * checkbox_drawing_type_coff1 * checkbox_drawing_type_coff2 \
                                     * checkbox_drawing_type_coff3 * checkbox_drawing_type_coff4 * checkbox_drawing_type_coff5 \
                                     * checkbox_drawing_type_coff6 * checkbox_drawing_type_coff7 * checkbox_drawing_type_coff8 \
                                     * checkbox_drawing_type_coff9 * checkbox_drawing_type_coff10 * checkbox_drawing_type_coff11

        combobox_location_type_coff = 1
        for i in range(2, 11):
            if combobox_location_type.get() == worksheet.cell(row=i, column=14).value:
                combobox_location_type_coff = float(worksheet.cell(row=i, column=15).value)
                break

        combobox_protection_type_coff = 1
        for i in range(2, 5):
            if combobox_protection_type.get() == worksheet.cell(row=i, column=16).value:
                combobox_protection_type_coff = float(worksheet.cell(row=i, column=17).value)
                break

        combobox_drawing_language_type_coff = 1
        for i in range(2, 5):
            if combobox_drawing_language_type.get() == worksheet.cell(row=i, column=18).value:
                combobox_drawing_language_type_coff = float(worksheet.cell(row=i, column=19).value)
                break

        combobox_customer_wiring_coff = 1
        for i in range(2, 4):
            if combobox_customer_wiring.get() == worksheet.cell(row=i, column=20).value:
                combobox_customer_wiring_coff = float(worksheet.cell(row=i, column=21).value)
                break

        combobox_customer_requirements_coff = 1
        for i in range(2, 4):
            if combobox_customer_requirements.get() == worksheet.cell(row=i, column=22).value:
                combobox_customer_requirements_coff = float(worksheet.cell(row=i, column=23).value)
                break

        checkbox_intelligent_coff = 1
        if checkbox_intelligent_list[0].get():
            checkbox_intelligent_coff1 = float(worksheet.cell(row=2, column=25).value)
        else:
            checkbox_intelligent_coff1 = 1

        if checkbox_intelligent_list[1].get():
            checkbox_intelligent_coff2 = float(worksheet.cell(row=3, column=25).value)
        else:
            checkbox_intelligent_coff2 = 1

        if checkbox_intelligent_list[2].get():
            checkbox_intelligent_coff3 = float(worksheet.cell(row=4, column=25).value)
        else:
            checkbox_intelligent_coff3 = 1

        if checkbox_intelligent_list[3].get():
            checkbox_intelligent_coff4 = float(worksheet.cell(row=5, column=25).value)
        else:
            checkbox_intelligent_coff4 = 1

        if checkbox_intelligent_list[4].get():
            checkbox_intelligent_coff5 = float(worksheet.cell(row=6, column=25).value)
        else:
            checkbox_intelligent_coff5 = 1

        if checkbox_intelligent_list[5].get():
            checkbox_intelligent_coff6 = float(worksheet.cell(row=7, column=25).value)
        else:
            checkbox_intelligent_coff6 = 1

        if checkbox_intelligent_list[6].get():
            checkbox_intelligent_coff7 = float(worksheet.cell(row=8, column=25).value)
        else:
            checkbox_intelligent_coff7 = 1

        checkbox_intelligent_coff = checkbox_intelligent_coff * checkbox_intelligent_coff1 * checkbox_intelligent_coff2 \
                                     * checkbox_intelligent_coff3 * checkbox_intelligent_coff4 * checkbox_intelligent_coff5 \
                                     * checkbox_intelligent_coff6 * checkbox_intelligent_coff7

        # print(combobox_design_type_coff, combobox_product_type_coff, entry_typical_coff, entry_amount_coff, checkbox_input_file_coff, checkbox_drawing_type_coff, combobox_location_type_coff, combobox_protection_type_coff,
        #       combobox_drawing_language_type_coff, combobox_customer_wiring_coff, combobox_customer_requirements_coff, checkbox_intelligent_coff)

        # normal_period = (float(entry_typical.get()) * combobox_design_type_coff * combobox_product_type_coff * entry_typical_coff * entry_amount_coff * checkbox_input_file_coff * checkbox_drawing_type_coff * combobox_location_type_coff \
        #                 * combobox_protection_type_coff * combobox_drawing_language_type_coff * combobox_customer_wiring_coff * combobox_customer_requirements_coff * checkbox_intelligent_coff) / 2.0

        normal_period = (float(entry_typical.get()) * (combobox_design_type_coff + combobox_product_type_coff + entry_typical_coff + entry_amount_coff + checkbox_input_file_coff + checkbox_drawing_type_coff + combobox_location_type_coff \
                        + combobox_protection_type_coff + combobox_drawing_language_type_coff + combobox_customer_wiring_coff + combobox_customer_requirements_coff + checkbox_intelligent_coff)) / 10.0

        entry_calculate_time['state'] = 'normal'
        entry_calculate_time.delete(0, "end")
        entry_calculate_time.insert(0, math.ceil(normal_period))
        entry_calculate_time['state'] = 'disabled'

        entry_extra_time['state'] = 'normal'
        entry_extra_time.delete(0, "end")
        entry_extra_time.insert(0, math.ceil(extra_period))
        entry_extra_time['state'] = 'disabled'

        if design_type_flag == 1 or combobox_design_type.get() == '提前设计':
            button_select_upload_time['state'] = 'normal'
        elif design_type_flag == 2 or combobox_design_type.get() == '图纸设计':
            button_select_upload_time['state'] = 'normal'
            button_select_bom_time['state'] = 'normal'
            button_select_check_time['state'] = 'normal'
        elif design_type_flag == 3 or combobox_design_type.get() == '工程设计':
            button_select_upload_time['state'] = 'normal'
            button_select_bom_time['state'] = 'normal'
            button_select_check_time['state'] = 'normal'


def select_update_date(content):
    entry_upload_time['state'] = 'normal'
    top = tk.Toplevel(content)

    cal = CCalendar(top, selectmode="day", date_pattern="yyyy-mm-dd")
    cal.pack()

    def get_date():
        date = cal.get_date()
        now_time = strftime('%Y-%m-%d', localtime())  # 时间戳
        if date < now_time:
            tk.messagebox.showwarning("提示", "日期不能早于今天")
        elif entry_bom_time.get() != '' and not date <= entry_bom_time.get():
            tk.messagebox.showwarning("提示", "日期需要早于发放初级时间")
        elif entry_check_time.get() != '' and not date <= entry_check_time.get():
            tk.messagebox.showwarning("提示", "日期需要早于生产检查时间")
        else:
            entry_upload_time.delete(0, "end")
            entry_upload_time.insert(0, date)
            top.destroy()
            entry_upload_time['state'] = 'disabled'

    button = tk.Button(top, text="确认", font=("ABBvoice CNSG", int(13 * h_ratio)), command=get_date)
    button.pack()
    center_window(top)
    top.grab_set()

def select_bom_date(content):
    entry_bom_time['state'] = 'normal'
    top = tk.Toplevel(content)
    cal = CCalendar(top, selectmode="day", date_pattern="yyyy-mm-dd")
    cal.pack()

    def get_date():
        date = cal.get_date()
        now_time = strftime('%Y-%m-%d', localtime())  # 时间戳
        if date < now_time:
            tk.messagebox.showwarning("提示", "日期不能早于今天")
        elif entry_upload_time.get() != '' and not date >= entry_upload_time.get():
            tk.messagebox.showwarning("提示", "日期需要晚于提原理图时间")
        elif entry_check_time.get() != '' and not date <= entry_check_time.get():
            tk.messagebox.showwarning("提示", "日期需要早于生产检查时间")
        else:
            entry_bom_time.delete(0, "end")
            entry_bom_time.insert(0, date)
            top.destroy()
            entry_bom_time['state'] = 'disabled'

    button = tk.Button(top, text="确认", font=("ABBvoice CNSG", int(13 * h_ratio)), command=get_date)
    button.pack()
    center_window(top)
    top.grab_set()


def select_check_date(content):
    entry_check_time['state'] = 'normal'
    top = tk.Toplevel(content)
    cal = CCalendar(top, selectmode="day", date_pattern="yyyy-mm-dd")
    cal.pack()

    def get_date():
        date = cal.get_date()
        now_time = strftime('%Y-%m-%d', localtime())  # 时间戳
        if date < now_time:
            tk.messagebox.showwarning("提示", "日期不能早于今天")
        elif entry_upload_time.get() != '' and not date >= entry_upload_time.get():
            tk.messagebox.showwarning("提示", "日期需要晚于提原理图时间")
        elif entry_bom_time.get() != '' and not date >= entry_bom_time.get():
            tk.messagebox.showwarning("提示", "日期需要晚于发放初级时间")
        else:
            entry_check_time.delete(0, "end")
            entry_check_time.insert(0, date)
            top.destroy()
            entry_check_time['state'] = 'disabled'

    button = tk.Button(top, text="确认", font=("ABBvoice CNSG", int(13 * h_ratio)), command=get_date)
    button.pack()
    center_window(top)
    top.grab_set()


def add_row1(content1, f2L_Right, canvas, h_ratio):
    entry_row1 = len(entries1)
    entries1.append([None for _ in range(num_columns1)])

    row_frame1 = tk.Frame(f2L_Right)
    row_frame1.pack(side=tk.TOP, fill=tk.X)

    for col in range(num_columns1):
        entry1 = tk.Entry(row_frame1, width=entry_widths1[col], relief='solid', font=("ABBvoice CNSG", int(13 * h_ratio)))
        entry1.pack(side=tk.LEFT)
        entries1[entry_row1][col] = entry1
        entry1.config(justify='center')

    # 更新滚动条
    content1.update_idletasks()
    canvas.config(scrollregion=canvas.bbox('all'))
    canvas.yview_moveto(1)


def delete_row1(content1, f2L_Right, canvas):
    if len(entries1) > 2:
        last_row_frame1 = f2L_Right.winfo_children()[-1]
        last_row_frame1.destroy()
        entries1.pop()

        # 更新滚动条
        content1.update_idletasks()
        canvas.config(scrollregion=canvas.bbox('all'))
        canvas.yview_moveto(1)


def add_row2(content1, f2M_Right, canvas, h_ratio):
    entry_row2 = len(entries2)
    entries2.append([None for _ in range(num_columns2)])

    row_frame2 = tk.Frame(f2M_Right)
    row_frame2.pack(side=tk.TOP, fill=tk.X)

    for col in range(num_columns2):
        entry2 = tk.Entry(row_frame2, width=entry_widths2[col], relief='solid', font=("ABBvoice CNSG", int(13 * h_ratio)))
        entry2.pack(side=tk.LEFT)
        if col == 0:
            entry2.config(justify='center')
            entry2.insert(0, '%s' % str(entry_row2 + 1))
            entry2['state'] = 'disabled'

        entries2[entry_row2][col] = entry2
        entry2.config()

    # 更新滚动条
    content1.update_idletasks()
    canvas.config(scrollregion=canvas.bbox('all'))
    canvas.yview_moveto(1)


def delete_row2(content1, f2M_Right, canvas):
    if len(entries2) > 1:
        last_row_frame2 = f2M_Right.winfo_children()[-1]
        last_row_frame2.destroy()
        entries2.pop()

        # 更新滚动条
        content1.update_idletasks()
        canvas.config(scrollregion=canvas.bbox('all'))
        canvas.yview_moveto(1)


def save_state():
    try:
        if (entry_project_number.get() == "" or len(entry_project_number.get()) != 9) and entry_project_number.get() != '提前设计':
            tk.messagebox.showwarning("提示", "请输入9位项目号")
        elif not entry_project_number.get().isdigit() and entry_project_number.get() != '提前设计':
            tk.messagebox.showwarning("提示", "项目号必须是9位纯数字")
        else:
            if combobox_design_type.get() == '提前设计':
                state = {
                    'design_type': combobox_design_type.get(),
                    'project_number': entry_project_number.get(),
                    'project_name': entry_project_name.get(),
                    'text_typical': int(entry_typical.get()),
                    'text_amount': int(entry_amount.get()),
                    'product_type': combobox_product_type.get(),
                    'drawing_language_type': combobox_drawing_language_type.get(),
                    'protection_type': combobox_protection_type.get(),
                    'location_type': combobox_location_type.get(),
                    'advice_times': combobox_advice_times.get(),
                    'switch_type': combobox_switch_type.get(),
                    'terminal_type': combobox_terminal_type.get(),
                    'mcb_type': combobox_MCB_type.get(),
                    'aux_type': combobox_aux_type.get(),
                    'charged_display_type': combobox_charged_display_type.get(),
                    'ct_type': combobox_ct_type.get(),
                    'pt_type': combobox_pt_type.get(),
                    'sa_type': combobox_sa_type.get(),
                    'customer_wiring': combobox_customer_wiring.get(),
                    'customer_requirements': combobox_customer_requirements.get(),
                    'input_file': [v.get() for v in checkbox_input_list],
                    'drawing_type': [v.get() for v in checkbox_drawing_type_list],
                    'intelligent': [v.get() for v in checkbox_intelligent_list],
                    'project_type': combobox_project_type.get(),
                    'frame_type': combobox_frame_type.get(),
                    'program_type': combobox_program_type.get(),
                    'language_type': combobox_language_type.get(),
                    'manage_type': combobox_manage_type.get(),
                    'entries1': [[entry.get() for entry in row] for row in entries1],
                    'entries2': [[entry.get() for entry in row] for row in entries2]
                }
            else:
                state = {
                    'design_type': combobox_design_type.get(),
                    'project_number': entry_project_number.get(),
                    'text_typical': int(entry_typical.get()),
                    'drawing_language_type': combobox_drawing_language_type.get(),
                    'protection_type': combobox_protection_type.get(),
                    'location_type': combobox_location_type.get(),
                    'advice_times': combobox_advice_times.get(),
                    'switch_type': combobox_switch_type.get(),
                    'terminal_type': combobox_terminal_type.get(),
                    'mcb_type': combobox_MCB_type.get(),
                    'aux_type': combobox_aux_type.get(),
                    'charged_display_type': combobox_charged_display_type.get(),
                    'ct_type': combobox_ct_type.get(),
                    'pt_type': combobox_pt_type.get(),
                    'sa_type': combobox_sa_type.get(),
                    'customer_wiring': combobox_customer_wiring.get(),
                    'customer_requirements': combobox_customer_requirements.get(),
                    'input_file': [v.get() for v in checkbox_input_list],
                    'drawing_type': [v.get() for v in checkbox_drawing_type_list],
                    'intelligent': [v.get() for v in checkbox_intelligent_list],
                    'project_type': combobox_project_type.get(),
                    'frame_type': combobox_frame_type.get(),
                    'program_type': combobox_program_type.get(),
                    'language_type': combobox_language_type.get(),
                    'manage_type': combobox_manage_type.get(),
                    'entries1': [[entry.get() for entry in row] for row in entries1],
                    'entries2': [[entry.get() for entry in row] for row in entries2]
                }
            with open('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\temp_file\\%s-%s.json'%(entry_project_number.get(), os.path.expanduser("~").split('\\')[-1]), 'w', encoding='utf-8') as f:
                json.dump(state, f, ensure_ascii=False)
            tk.messagebox.showwarning("提示", "所填内容已存档")
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def find_files(folder_path, search_string, file_extension):
    file_list = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if search_string in file and file.endswith(file_extension):
                file_list.append(os.path.join(root, file))
    return file_list


def load_state(h_ratio):
    global combobox_design_type
    global entry_project_number
    global entry_project_name
    global entry_typical
    global entry_amount
    global combobox_product_type
    global combobox_drawing_language_type
    global combobox_protection_type
    global combobox_location_type
    global combobox_advice_times
    global combobox_switch_type
    global combobox_terminal_type
    global combobox_MCB_type
    global combobox_aux_type
    global combobox_charged_display_type
    global combobox_ct_type
    global combobox_pt_type
    global combobox_sa_type
    global combobox_customer_wiring
    global combobox_customer_requirements
    global checkbox_input_list
    global checkbox_drawing_type_list
    global checkbox_intelligent_list
    global combobox_project_type
    global combobox_frame_type
    global combobox_program_type
    global combobox_language_type
    global combobox_manage_type
    global entries1
    global entries2


    try:
        username = '-'+os.path.expanduser("~").split('\\')[-1]
        files = find_files('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\temp_file', username, 'json')

        if len(files) > 0:
            filename = filedialog.askopenfilename(
                title="选择历史填写文件",
                filetypes=(("JSON files", "*.json"), ("All files", "*.*")),
                initialdir="J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\temp_file\\",  # 可以指定默认的起始目录
            )
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    state = json.load(f)
                    # combobox_design_type['state'] = 'normal'
                    # print(state['design_type'], type(state['design_type']), state['project_number'], type(state['project_number']))
                    combobox_design_type.set(state['design_type'])
                    # print(combobox_design_type.get())
                    entry_project_number['state'] = 'normal'
                    entry_project_number.insert(0, state['project_number'])
                    if state['design_type'] == '提前设计':
                        entry_project_name['state'] = 'normal'
                        entry_project_name['bg'] = 'white'
                        entry_project_name.insert(0, state['project_name'])

                        text_item['state'] = 'normal'
                        text_item.insert('1.0', '1000')
                        text_item['state'] = 'disabled'

                        entry_amount['state'] = 'normal'
                        entry_amount['bg'] = 'white'
                        entry_amount.insert(0, state['text_amount'])

                        combobox_product_type.set(state['product_type'])
                    else:
                        button_query_project['state'] = 'normal'
                    entry_typical.insert(0, state['text_typical'])
                    combobox_drawing_language_type.set(state['drawing_language_type'])
                    combobox_protection_type.set(state['protection_type'])
                    combobox_location_type.set(state['location_type'])
                    combobox_advice_times.set(state['advice_times'])
                    combobox_switch_type.set(state['switch_type'])
                    combobox_terminal_type.set(state['terminal_type'])
                    combobox_MCB_type.set(state['mcb_type'])
                    combobox_aux_type.set(state['aux_type'])
                    combobox_charged_display_type.set(state['charged_display_type'])
                    combobox_ct_type.set(state['ct_type'])
                    combobox_pt_type.set(state['pt_type'])
                    combobox_sa_type.set(state['sa_type'])
                    combobox_customer_wiring.set(state['customer_wiring'])
                    combobox_customer_requirements.set(state['customer_requirements'])
                    for v, saved_state in zip(checkbox_input_list, state['input_file']):
                        v.set(saved_state)
                    for v, saved_state in zip(checkbox_drawing_type_list, state['drawing_type']):
                        v.set(saved_state)
                    for v, saved_state in zip(checkbox_intelligent_list, state['intelligent']):
                        v.set(saved_state)
                    combobox_project_type.set(state['project_type'])
                    combobox_frame_type.set(state['frame_type'])
                    combobox_program_type.set(state['program_type'])
                    combobox_language_type.set(state['language_type'])
                    combobox_manage_type.set(state['manage_type'])

                    num_entries1_rows = len(state['entries1'])

                    num_entries1_ui_rows = 6
                    while num_entries1_rows > num_entries1_ui_rows:
                        add_row1(content, f2L_Right, canvas, h_ratio)
                        num_entries1_ui_rows += 1

                    # 加载entries1
                    for i, row in enumerate(entries1):
                        for j, entry in enumerate(row):
                            entry.insert(0, state['entries1'][i][j])

                    num_entries2_rows = len(state['entries2'])

                    num_entries2_ui_rows = 6
                    while num_entries2_rows > num_entries2_ui_rows:
                        add_row2(content, f2M_Right, canvas, h_ratio)
                        num_entries2_ui_rows += 1

                    # 加载entries2
                    for i, row in enumerate(entries2):
                        for j, entry in enumerate(row):
                            entry.insert(0, state['entries2'][i][j])


            except json.JSONDecodeError:
                tk.messagebox.showwarning("提示", traceback.format_exc())

    except FileNotFoundError:
        pass  # 没有找到文件，可能是首次运行


def read_design_details():
    if not os.path.exists(os.path.abspath('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\原理图标准设计_red.pdf')):
        tk.messagebox.showwarning("提示", "缺少原理图标准设计说明书.pdf")
    else:
        os.startfile(os.path.abspath('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\原理图标准设计_red.pdf'))
        checkbox_read.select()
        export_button['state'] = 'normal'


def export_transfer_sheet():
    try:
        if combobox_design_type.get() == '' or entry_project_number.get() == '' or entry_project_name.get() == '' or (combobox_design_type.get() == '提前设计' and (entry_calculate_time.get() == '' or entry_extra_time.get() == '' or entry_upload_time.get() == '')) or \
            (combobox_design_type.get() == '图纸设计' and (entry_calculate_time.get() == '' or entry_extra_time.get() == '' or entry_upload_time.get() == '' or entry_check_time.get() == '')) or \
                (combobox_design_type.get() == '工程设计' and (entry_calculate_time.get() == '' or entry_extra_time.get() == '' or entry_upload_time.get() == '' or entry_check_time.get() == '' or entry_bom_time.get() == '')) or \
                combobox_MCB_type.get() == '' or combobox_aux_type.get() == '' or combobox_terminal_type.get() == '' or combobox_charged_display_type.get() == '' or combobox_switch_type.get() == '' or combobox_project_type.get() == '' or \
                combobox_frame_type.get() == '' or combobox_program_type.get() == '' or combobox_language_type.get() == '' or combobox_manage_type.get() == '' or combobox_ct_type.get() == '' or combobox_pt_type.get() == '' or combobox_sa_type.get() == '':
            tk.messagebox.showwarning("提示", "*标属性未填写完整，无法导出")

        elif combobox_design_type.get() != '提前设计' and (entry_project_number.get() == "" or len(entry_project_number.get()) != 9 or not entry_project_number.get().isdigit()):
            tk.messagebox.showwarning("提示", "请输入9位项目号，且需要为纯数字")

        else:
            try:

                pdfmetrics.registerFont(TTFont('SimSun', 'simsun.ttc'))  # 注册字体

                home_path = os.path.expanduser("~")
                initial_address = os.path.join(home_path, "Desktop")

                # if os.path.exists(os.path.join(desktop_path, 'DesignList-%s.pdf' % entry_project_number.get())):
                #     os.remove(os.path.join(desktop_path, 'DesignList-%s.pdf' % entry_project_number.get()))

                desktop_path = askdirectory(title=u'请选择导出文件夹', initialdir=initial_address)
                if not desktop_path:
                    tk.messagebox.showwarning("提示", "未选择保存路径，导出操作已取消")
                    return

                c = rl_canvas.Canvas(os.path.join(desktop_path, 'DesignList-%s.pdf' % entry_project_number.get()))

                img = Image("ico/ABB_logo.png")
                img.drawWidth = 27.7 * mm
                img.drawHeight = 10.5 * mm
                img.drawOn(c, A4[0] / 2 - img.drawWidth / 2, A4[1] - img.drawHeight - 30)

                c.setFont("SimSun", 16)
                c.setFillColor(Color(0, 0, 0, alpha=1))
                c.drawCentredString(A4[0] / 2, A4[1] - img.drawHeight - 30 - 20, "FAST DesignList")

                c.setFont("SimSun", 16)
                c.setFillColor(Color(0, 0, 0, alpha=1))
                c.drawString(40, A4[1] - img.drawHeight - 30 - 30 - 20, "项目基本信息", 2)

                c.setFont("SimSun", 9)
                c.drawString(40, A4[1] - img.drawHeight - 30 - 30 - 40, '设计需求：' + combobox_design_type.get())

                c.setFont("SimSun", 9)
                c.drawString(40, A4[1] - img.drawHeight - 30 - 30 - 60, '  项目号：'+entry_project_number.get())

                textobject = c.beginText()
                textobject.setFont("SimSun", 9)
                textobject.setTextOrigin(40, A4[1] - img.drawHeight - 30 - 30 - 80)
                Designlist_project_name = entry_project_name.get()
                if len(Designlist_project_name) > 28:
                    textobject.textLine('项目名称：' + Designlist_project_name[:28])
                    textobject.textLine('\n        ' + Designlist_project_name[28:])
                else:

                    textobject.textLine('项目名称：' + Designlist_project_name)
                    textobject.textLine('\n        ')
                c.drawText(textobject)

                data = [
                    ('', '签名', '日期'),
                    ('区域经理', '', ''),
                    ('PM/PE', '', ''),
                    ('绘图员', '', '')
                ]
                col_width = ['50', '70', '70']
                row_height = 23
                style = [
                    ('FONTNAME', (0, 0), (-1, -1), 'SimSun'),  # 字体
                    ('FONTSIZE', (0, 0), (-1, 0), 9),  # 第一行的字体大小
                    ('FONTSIZE', (0, 1), (-1, -1), 9),  # 第二行到最后一行的字体大小
                    # ('BACKGROUND', (0, 0), (-1, 0), '#d5dae6'),  # 设置第一行背景颜色
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # 第一行水平居中
                    ('ALIGN', (0, 1), (-1, -1), 'CENTER'),  # 第二行到最后一行左右左对齐
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # 所有表格上下居中对齐
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),  # 设置表格内文字颜色
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # 设置表格框线为grey色，线宽为0.5
                    # ('SPAN', (0, 1), (2, 1)),  # 合并第二行一二三列
                ]
                t = Table(data, colWidths=col_width, rowHeights=row_height, style=style)

                t.wrap(190, 92)
                t.drawOn(c, A4[0] / 2 + 60, A4[1] - img.drawHeight - 30 - 30 - 30 - 80)

                textobject = c.beginText()
                textobject.setFont("SimSun", 9)
                textobject.setTextOrigin(40, A4[1] - img.drawHeight - 30 - 30 - 30 - 100)

                if str(text_item.get("1.0", "end-1c"))[-1] == '0':
                    Designlist_item = str(text_item.get("1.0", "end-1c"))+';'
                else:
                    Designlist_item = str(text_item.get("1.0", "end-1c")).replace('；', ';')

                if Designlist_item.count(';') > 17:
                    textobject.textLine('行号信息：' + (';'.join(str(num) for num in Designlist_item.split(';')[:17])))
                    for i in range(1, math.ceil(Designlist_item.count(';') / 10)):
                        textobject.textLine('\n        ' + (';'.join(str(num) for num in Designlist_item.split(';')[17 * i:17 * (i + 1)])))
                else:
                    textobject.textLine('行号信息：' + Designlist_item)
                    textobject.textLine('\n        ')

                c.drawText(textobject)

                c.setFont("SimSun", 9)
                c.drawString(40, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5, '图套数量：'+entry_typical.get())

                c.setFont("SimSun", 9)
                c.drawString(230, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5, '柜子数量：'+entry_amount.get())

                c.setFont("SimSun", 9)
                c.drawString(420, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5, '产品类型：'+combobox_product_type.get())

                c.setFont("SimSun", 9)
                c.drawString(40, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 20, '图纸语言：'+combobox_drawing_language_type.get())

                c.setFont("SimSun", 9)
                c.drawString(230, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 20, '保护类型：'+combobox_protection_type.get())

                c.setFont("SimSun", 9)
                c.drawString(420, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 20, '地区类型：'+combobox_location_type.get())

                c.setFont("SimSun", 9)
                c.drawString(40, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 40, '选开类型：'+combobox_switch_type.get())

                c.setFont("SimSun", 9)
                c.drawString(230, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 40, '带显类型：'+combobox_charged_display_type.get())

                c.setFont("SimSun", 9)
                c.drawString(420, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 40, '端子类型：'+combobox_terminal_type.get())

                c.setFont("SimSun", 9)
                c.drawString(40, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 60, '空开类型：'+combobox_MCB_type.get())

                c.setFont("SimSun", 9)
                c.drawString(230, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 60, '辅助触点：'+combobox_aux_type.get())

                c.setFont("SimSun", 9)
                c.drawString(420, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 60, '意见次数：'+combobox_advice_times.get())

                c.setFont("SimSun", 9)
                c.drawString(40, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 80, '  CT类型：'+combobox_ct_type.get())

                c.setFont("SimSun", 9)
                c.drawString(230, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 80, '  PT类型：'+combobox_pt_type.get())

                c.setFont("SimSun", 9)
                c.drawString(420, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 80, '  SA类型：'+combobox_sa_type.get())

                c.setFont("SimSun", 9)
                c.drawString(40, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 100, '客户线号：'+combobox_customer_wiring.get())

                c.setFont("SimSun", 9)
                c.drawString(40, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 120, '复杂客户需求：'+combobox_customer_requirements.get())

                textobject = c.beginText()
                textobject.setFont("SimSun", 9)
                textobject.setTextOrigin(40, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 140 - 10)

                true_checkbox_input_list = ''
                for i in range(0, 5):
                    if checkbox_input_list[i].get():
                        true_checkbox_input_list += input_files[i] + ';'

                Designlist_input_file = true_checkbox_input_list[0:len(true_checkbox_input_list)-1]

                if len(Designlist_input_file) > 55:
                    textobject.textLine('输入文件：' + Designlist_input_file[:55])
                    textobject.textLine('\n        ' + Designlist_input_file[55:])
                else:
                    textobject.textLine('输入文件：' + Designlist_input_file)
                c.drawText(textobject)

                textobject = c.beginText()
                textobject.setFont("SimSun", 9)
                textobject.setTextOrigin(40, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 180 - 10)

                true_checkbox_intelligent_list = ''
                for i in range(0, 7):
                    if checkbox_intelligent_list[i].get():
                        true_checkbox_intelligent_list += intelligent_needs[i] + ';'

                Designlist_intelligence = true_checkbox_intelligent_list[0:len(true_checkbox_intelligent_list)-1]

                if len(Designlist_intelligence) > 55:
                    textobject.textLine('智能方案：' + Designlist_intelligence[:55])
                    textobject.textLine('\n        ' + Designlist_intelligence[55:])
                else:
                    textobject.textLine('智能方案：' + Designlist_intelligence)
                c.drawText(textobject)

                textobject = c.beginText()
                textobject.setFont("SimSun", 9)
                textobject.setTextOrigin(40, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 220 - 10)

                true_checkbox_drawing_type_list = ''
                for i in range(0, 10):
                    if checkbox_drawing_type_list[i].get():
                        true_checkbox_drawing_type_list += drawing_types[i] + ';'

                Designlist_drawing_requirements = true_checkbox_drawing_type_list[0:len(true_checkbox_drawing_type_list)-1]

                if len(Designlist_drawing_requirements) > 80:
                    textobject.textLine('图纸需求：' + Designlist_drawing_requirements[:80])
                    textobject.textLine('\n        ' + Designlist_drawing_requirements[80:])
                else:
                    textobject.textLine('图纸需求：' + Designlist_drawing_requirements)
                c.drawText(textobject)

                c.setFont("SimSun", 9)
                c.drawString(40, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 220 - 10 - 40, '标准周期：'+str(entry_calculate_time.get()))

                c.setFont("SimSun", 9)
                c.drawString(230, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 220 - 10 - 40, '额外周期：'+str(entry_extra_time.get()))

                c.setFont("SimSun", 9)
                c.drawString(40, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 220 - 10 - 60, '提原理图：'+entry_upload_time.get())

                if entry_bom_time.get() == '':
                    c.setFont("SimSun", 9)
                    c.drawString(230, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 220 - 10 - 60, '发放初级：'+'N/A')
                else:
                    c.setFont("SimSun", 9)
                    c.drawString(230, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 220 - 10 - 60, '发放初级：' + entry_bom_time.get())

                if entry_check_time.get() == '':
                    c.setFont("SimSun", 9)
                    c.drawString(420, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 220 - 10 - 60, '生产检查：'+'N/A')
                else:
                    c.setFont("SimSun", 9)
                    c.drawString(420, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 220 - 10 - 60, '生产检查：' + entry_check_time.get())

                c.setFont("SimSun", 9)
                c.drawString(40, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 220 - 10 - 80, '项目类型：'+combobox_project_type.get())

                c.setFont("SimSun", 9)
                c.drawString(40, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 220 - 10 - 100, '框架类型：'+combobox_frame_type.get())

                c.setFont("SimSun", 9)
                c.drawString(40, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 220 - 10 - 120, '编程难度：'+combobox_program_type.get())

                c.setFont("SimSun", 9)
                c.drawString(40, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 220 - 10 - 140, '语言难度：'+combobox_language_type.get())

                c.setFont("SimSun", 9)
                c.drawString(40, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 220 - 10 - 160, '管理难度：'+combobox_manage_type.get())

                c.setFont("SimSun", 16)
                c.setFillColor(Color(0, 0, 0, alpha=1))
                c.drawString(40, A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 220 - 10 - 160 - 30, "典型柜配置", 2)
                current_page = 1

                c.setFont("SimSun", 9)
                c.setFillColor(Color(0, 0, 0, alpha=1))
                c.drawString(A4[0] - 80, 10 * mm, "Page %d" % current_page)
                Designlist_Timestamp = strftime('%Y-%m-%d %H:%M:%S', localtime())
                Timestamp = Designlist_Timestamp
                c.drawCentredString(A4[0] / 2, 10 * mm, '制表时间：' + Designlist_Timestamp)

                error_input1 = 0
                valid_row_count1 = 0
                for row in range(1, len(entries1)):
                    if ';' not in entries1[row][0].get() and ';' not in entries1[row][1].get() and ';' not in entries1[row][2].get():
                        if entries1[row][0].get() != '' or entries1[row][1].get() != '' or entries1[row][2].get() != '':
                            valid_row_count1 += 1
                    else:
                        error_input1 = 1

                error_input2 = 0
                valid_row_count2 = 0
                for row in range(0, len(entries2)):
                    if ';' not in entries2[row][1].get():
                        if len(entries2[row][1].get()) > 0:
                            valid_row_count2 += 1
                    else:
                        error_input2 = 1

                if error_input1 and not error_input2:
                    tk.messagebox.showwarning("提示", "请移除典型柜配置表中的;(分号)")
                    return
                elif error_input2 and not error_input1:
                    tk.messagebox.showwarning("提示", "请移除备注表中的;(分号)")
                    return
                elif error_input1 and error_input2:
                    tk.messagebox.showwarning("提示", "请移除典型柜配置表、备注表中的;(分号)")
                else:
                    data = [('Panel/Typical', '保护料号/提单号', '表计料号/提单号')]

                    for row in range(1, len(entries1)):
                        if entries1[row][0].get() != '' or entries1[row][1].get() != '' or entries1[row][2].get() != '':
                            data.append((entries1[row][0].get().replace('\n', ''), entries1[row][1].get().replace('\n', ''), entries1[row][2].get().replace('\n', '')))

                    style = [
                        ('FONTNAME', (0, 0), (-1, -1), 'SimSun'),  # 字体
                        ('FONTSIZE', (0, 0), (-1, 0), 9),  # 第一行的字体大小
                        ('FONTSIZE', (0, 1), (-1, -1), 9),  # 第二行到最后一行的字体大小
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # 第一行水平居中
                        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),  # 第二行到最后一行左右左对齐
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # 确保垂直顶部对齐
                        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),  # 设置表格内文字颜色
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # 设置表格框线为grey色，线宽为0.5
                    ]
                    col_width = [70, 220, 220]
                    row_height = 15

                    current_height = A4[1] - img.drawHeight - 30 - 30 - 30 - 80 - 20 * math.ceil(Designlist_item.count(';') / 17+1) - 5 - 220 - 10 - 160 - 30 - 10

                    for sub_data in data:
                        table = Table([sub_data], colWidths=col_width, rowHeights=row_height)
                        table.setStyle(style)
                        required_height = row_height
                        table.wrap(500, 15)

                        # # 获取表格实际高度
                        # actual_width, actual_height = table.wrap(0, current_height)
                        # required_height = actual_height  # 使用实际计算的高度

                        if current_height - required_height < 20 * mm:
                            c.showPage()
                            current_height = A4[1] - 10 * mm
                            current_page += 1
                            c.setFont("SimSun", 9)
                            c.setFillColor(Color(0, 0, 0, alpha=1))
                            c.drawString(A4[0] - 80, 10 * mm, "Page %d" % current_page)
                            Designlist_Timestamp = strftime('%Y-%m-%d %H:%M:%S', localtime())
                            c.drawCentredString(A4[0] / 2, 10 * mm, '制表时间：' + Designlist_Timestamp)

                        table.drawOn(c, 40, current_height - row_height)
                        current_height -= row_height

                    c.setFont("SimSun", 16)
                    c.setFillColor(Color(0, 0, 0, alpha=1))
                    c.drawString(40, current_height - 30, "备注", 2)

                    data = []
                    count_k = 0
                    for row in range(0, len(entries2)):
                        if len(entries2[row][1].get()) > 0:
                            count_k += 1
                            data.append(str(count_k) + '、' + entries2[row][1].get().replace('\n', ''))

                    style = [
                        ('FONTNAME', (0, 0), (-1, -1), 'SimSun'),  # 字体
                        ('FONTSIZE', (0, 0), (-1, 0), 9),  # 第一行的字体大小
                        ('FONTSIZE', (0, 1), (-1, -1), 9),  # 第二行到最后一行的字体大小
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # 第一行水平居中
                        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),  # 第二行到最后一行左右左对齐
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # 确保垂直顶部对齐
                        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),  # 设置表格内文字颜色
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # 设置表格框线为grey色，线宽为0.5
                    ]
                    col_width = 510
                    row_height = 15

                    current_height = current_height - 30 - 10

                    for sub_data in data:
                        # print(sub_data)
                        table = Table([[sub_data]], colWidths=col_width, rowHeights=row_height)
                        table.setStyle(style)
                        required_height = row_height
                        table.wrap(500, 15)
                        if current_height - required_height < 20 * mm:
                            c.showPage()
                            current_height = A4[1] - 10 * mm
                            current_page += 1
                            c.setFont("SimSun", 9)
                            c.setFillColor(Color(0, 0, 0, alpha=1))
                            c.drawString(A4[0] - 80, 10 * mm, "Page %d" % current_page)
                            Designlist_Timestamp = strftime('%Y-%m-%d %H:%M:%S', localtime())
                            c.drawCentredString(A4[0] / 2, 10 * mm, '制表时间：' + Designlist_Timestamp)

                        table.drawOn(c, 40, current_height - row_height)
                        current_height -= row_height

                    c.setFont("SimSun", 16)
                    c.setFillColor(Color(0, 0, 0, alpha=1))
                    c.drawString(40, current_height - 30, "项目配置", 2)

                    data = [('CI', '柜型', '数量')]

                    if combobox_design_type.get() != '提前设计':
                        for i in range(0, len(real_item_list.split(';'))):
                            data.append(((real_item_list.split(';'))[i], real_product_type_list1[i], real_amount_list[i]))

                        style = [
                            ('FONTNAME', (0, 0), (-1, -1), 'SimSun'),  # 字体
                            ('FONTSIZE', (0, 0), (-1, 0), 9),  # 第一行的字体大小
                            ('FONTSIZE', (0, 1), (-1, -1), 9),  # 第二行到最后一行的字体大小
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # 第一行水平居中
                            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),  # 第二行到最后一行左右左对齐
                            ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # 确保垂直顶部对齐
                            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),  # 设置表格内文字颜色
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # 设置表格框线为grey色，线宽为0.5
                        ]
                        col_width = [50, 100, 50]
                        row_height = 15

                        current_height = current_height - 30 - 10

                        for sub_data in data:
                            # print(sub_data)
                            table = Table([sub_data], colWidths=col_width, rowHeights=row_height)
                            table.setStyle(style)
                            required_height = row_height
                            table.wrap(200, 15)
                            if current_height - required_height < 20 * mm:
                                c.showPage()
                                current_height = A4[1] - 10 * mm
                                current_page += 1
                                c.setFont("SimSun", 9)
                                c.setFillColor(Color(0, 0, 0, alpha=1))
                                c.drawString(A4[0] - 80, 10 * mm, "Page %d" % current_page)
                                Designlist_Timestamp = strftime('%Y-%m-%d %H:%M:%S', localtime())
                                c.drawCentredString(A4[0] / 2, 10 * mm, '制表时间：' + Designlist_Timestamp)

                            table.drawOn(c, 40, current_height - row_height)
                            current_height -= row_height

                    c.save()

                    # input_paths = [os.path.join(desktop_path, 'DesignList-%s.pdf' % entry_project_number.get()), os.path.abspath('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\原理图标准设计.pdf')]
                    # merger = PyPDF2.PdfMerger()
                    # for path in input_paths:
                    #     merger.append(path)
                    # merger.write(os.path.join(desktop_path, 'DesignList-%s.pdf' % entry_project_number.get()))
                    # merger.close()

                    tk.messagebox.showwarning("提示", "DesignList-%s.pdf已保存到桌面" % entry_project_number.get())

                    save_state()

                    result = tk.messagebox.askquestion("提示", "确定要将项目信息写入数据库吗？")
                    if result == 'yes':
                        if os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools'):
                            if is_folder_hidden('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb') or os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb'):
                                if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db'):
                                    tk.messagebox.showwarning("提示", "数据库不存在，请联系管理员")
                                else:
                                    conn = sqlite3.connect('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db')
                                    cursor = conn.cursor()  # 创建一个Cursor

                                    # 创建project_data表
                                    cursor.execute('''CREATE TABLE IF NOT EXISTS project_data
                                                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                                                  design_type TEXT,
                                                  project_number TEXT,
                                                  item_info INTEGER,
                                                  project_name TEXT,
                                                  typical_amount INTEGER,
                                                  panel_amount INTEGER,
                                                  panel_type TEXT,
                                                  product_type TEXT,
                                                  PE TEXT,
                                                  DE TEXT,
                                                  intelligence TEXT,
                                                  main_bus_current INTEGER,
                                                  main_protection TEXT,
                                                  arc_light TEXT,
                                                  location_info TEXT,
                                                  drawing_language TEXT,
                                                  standard_time INTEGER,
                                                  extra_time INTEGER,
                                                  advice_times INTEGER,
                                                  upload_time TEXT,
                                                  bom_time TEXT,
                                                  check_time TEXT,
                                                  receive_time TEXT,
                                                  start_time TEXT,
                                                  estimated_time TEXT,
                                                  status_info TEXT,
                                                  actual_time TEXT,
                                                  team_info TEXT,
                                                  abnormal_status TEXT,
                                                  create_time TIMESTAMP,
                                                  project_type TEXT,
                                                  frame_type TEXT,
                                                  program_type TEXT,
                                                  language_info TEXT,
                                                  management_info TEXT,
                                                  input_file TEXT,
                                                  drawing_type TEXT,
                                                  mcb_type TEXT,
                                                  aux_type TEXT,
                                                  terminal_type TEXT,
                                                  charged_display_type TEXT,
                                                  switch_type TEXT,
                                                  special_panel_config TEXT,
                                                  remark TEXT,
                                                  reviser TEXT,
                                                  revise_time TIMESTAMP,
                                                  ct_type TEXT,
                                                  pt_type TEXT,
                                                  sa_type TEXT,
                                                  customer_wiring TEXT,
                                                  customer_requirements TEXT)''')

                                    content_table1 = ''
                                    for i in range(0, len(entries1)):
                                        for j in range(num_columns1):
                                            content_table1 += entries1[i][j].get() + ';'

                                    content_table2 = ''
                                    for i in range(0, len(entries2)):
                                        for j in range(num_columns2):
                                            content_table2 += entries2[i][j].get() + ';'
                                    # print(content_table1)
                                    # print(content_table2)

                                    if combobox_design_type.get() == '提前设计':
                                        data = {
                                            'design_type': combobox_design_type.get(),
                                            'project_number': entry_project_number.get(),
                                            'item_info': 1000,
                                            'project_name': entry_project_name.get(),
                                            'typical_amount': int(entry_typical.get()),
                                            'panel_amount': int(entry_amount.get()),
                                            'product_type': combobox_product_type.get(),
                                            'intelligence': true_checkbox_intelligent_list[0:len(true_checkbox_intelligent_list) - 1].replace(' ', ''),
                                            'location_info': combobox_location_type.get(),
                                            'drawing_language': combobox_drawing_language_type.get(),
                                            'standard_time': int(entry_calculate_time.get()),
                                            'extra_time': int(entry_extra_time.get()),
                                            'advice_times': int(combobox_advice_times.get()),
                                            'upload_time': entry_upload_time.get(),
                                            'bom_time': entry_bom_time.get(),
                                            'check_time': entry_check_time.get(),
                                            'create_time': Timestamp,
                                            'project_type': combobox_project_type.get(),
                                            'frame_type': combobox_frame_type.get(),
                                            'program_type': combobox_program_type.get(),
                                            'language_info': combobox_language_type.get(),
                                            'management_info': combobox_manage_type.get(),
                                            'input_file': true_checkbox_input_list[0:len(true_checkbox_input_list) - 1].replace(' ', ''),
                                            'drawing_type': true_checkbox_drawing_type_list[0:len(true_checkbox_drawing_type_list) - 1].replace(' ', ''),
                                            'mcb_type': combobox_MCB_type.get(),
                                            'aux_type': combobox_aux_type.get(),
                                            'terminal_type': combobox_terminal_type.get(),
                                            'charged_display_type': combobox_charged_display_type.get(),
                                            'switch_type': combobox_switch_type.get(),
                                            'special_panel_config': content_table1,
                                            'remark': content_table2,
                                            'ct_type': combobox_ct_type.get(),
                                            'pt_type': combobox_pt_type.get(),
                                            'sa_type': combobox_sa_type.get(),
                                            'customer_wiring': combobox_customer_wiring.get(),
                                            'customer_requirements': combobox_customer_requirements.get(),
                                        }
                                        columns = ', '.join(data.keys())
                                        placeholders = ', '.join(['?' for _ in range(len(data))])
                                        query = f"INSERT INTO project_data ({columns}) VALUES ({placeholders})"

                                        values = [data.get(col) for col in data.keys()]
                                        cursor.execute(query, tuple(values))
                                        conn.commit()
                                        tk.messagebox.showwarning("提示", "数据写入完成")

                                    else:
                                        already_exist_item = []
                                        for i in range(0, len(real_item_list.split(';'))):

                                            # 插入之前先查询数据是否已经存在，若存在，则不插入；若不存在，则插入
                                            data = {
                                                'design_type': combobox_design_type.get(),
                                                'project_number': entry_project_number.get(),
                                                'item_info': int((real_item_list.split(';'))[i]),
                                            }
                                            query = "SELECT * FROM project_data WHERE design_type=? AND project_number=? AND item_info=?"
                                            cursor.execute(query, (data['design_type'], data['project_number'], data['item_info']))
                                            result = cursor.fetchone()
                                            if result:
                                                already_exist_item.append(int((real_item_list.split(';'))[i]))
                                        if len(already_exist_item) > 0:
                                            tk.messagebox.showwarning("提示", "%s行号数据已存在，写入失败" % already_exist_item)
                                        else:
                                            for i in range(0, len(real_item_list.split(';'))):
                                                if i == 0:
                                                    data = {
                                                        'design_type': combobox_design_type.get(),
                                                        'project_number': entry_project_number.get(),
                                                        'item_info': int((real_item_list.split(';'))[i]),
                                                        'project_name': entry_project_name.get(),
                                                        'typical_amount': int(entry_typical.get()),
                                                        'panel_amount': real_amount_list[i],
                                                        'panel_type': real_product_type_list1[i],
                                                        'product_type': combobox_product_type.get(),
                                                        'intelligence': true_checkbox_intelligent_list[0:len(true_checkbox_intelligent_list) - 1].replace(' ', ''),
                                                        'location_info': combobox_location_type.get(),
                                                        'drawing_language': combobox_drawing_language_type.get(),
                                                        'standard_time': int(entry_calculate_time.get()),
                                                        'extra_time': int(entry_extra_time.get()),
                                                        'advice_times': int(combobox_advice_times.get()),
                                                        'upload_time': entry_upload_time.get(),
                                                        'bom_time': entry_bom_time.get(),
                                                        'check_time': entry_check_time.get(),
                                                        'create_time': Timestamp,
                                                        'project_type': combobox_project_type.get(),
                                                        'frame_type': combobox_frame_type.get(),
                                                        'program_type': combobox_program_type.get(),
                                                        'language_info': combobox_language_type.get(),
                                                        'management_info': combobox_manage_type.get(),
                                                        'input_file': true_checkbox_input_list[0:len(true_checkbox_input_list) - 1].replace(' ', ''),
                                                        'drawing_type': true_checkbox_drawing_type_list[0:len(true_checkbox_drawing_type_list) - 1].replace(' ', ''),
                                                        'mcb_type': combobox_MCB_type.get(),
                                                        'aux_type': combobox_aux_type.get(),
                                                        'terminal_type': combobox_terminal_type.get(),
                                                        'charged_display_type': combobox_charged_display_type.get(),
                                                        'switch_type': combobox_switch_type.get(),
                                                        'special_panel_config': content_table1,
                                                        'remark': content_table2,
                                                        'ct_type': combobox_ct_type.get(),
                                                        'pt_type': combobox_pt_type.get(),
                                                        'sa_type': combobox_sa_type.get(),
                                                        'customer_wiring': combobox_customer_wiring.get(),
                                                        'customer_requirements': combobox_customer_requirements.get(),
                                                    }
                                                else:
                                                    data = {
                                                        'design_type': combobox_design_type.get(),
                                                        'project_number': entry_project_number.get(),
                                                        'item_info': int((real_item_list.split(';'))[i]),
                                                        'project_name': entry_project_name.get(),
                                                        'typical_amount': 0,
                                                        'panel_amount': real_amount_list[i],
                                                        'panel_type': real_product_type_list1[i],
                                                        'product_type': combobox_product_type.get(),
                                                        'intelligence': true_checkbox_intelligent_list[0:len(true_checkbox_intelligent_list) - 1].replace(' ', ''),
                                                        'location_info': combobox_location_type.get(),
                                                        'drawing_language': combobox_drawing_language_type.get(),
                                                        'standard_time': int(entry_calculate_time.get()),
                                                        'extra_time': int(entry_extra_time.get()),
                                                        'advice_times': int(combobox_advice_times.get()),
                                                        'upload_time': entry_upload_time.get(),
                                                        'bom_time': entry_bom_time.get(),
                                                        'check_time': entry_check_time.get(),
                                                        'create_time': Timestamp,
                                                        'project_type': combobox_project_type.get(),
                                                        'frame_type': combobox_frame_type.get(),
                                                        'program_type': combobox_program_type.get(),
                                                        'language_info': combobox_language_type.get(),
                                                        'management_info': combobox_manage_type.get(),
                                                        'input_file': true_checkbox_input_list[0:len(true_checkbox_input_list) - 1].replace(' ', ''),
                                                        'drawing_type': true_checkbox_drawing_type_list[0:len(true_checkbox_drawing_type_list) - 1].replace(' ', ''),
                                                        'mcb_type': combobox_MCB_type.get(),
                                                        'aux_type': combobox_aux_type.get(),
                                                        'terminal_type': combobox_terminal_type.get(),
                                                        'charged_display_type': combobox_charged_display_type.get(),
                                                        'switch_type': combobox_switch_type.get(),
                                                        'special_panel_config': content_table1,
                                                        'remark': content_table2,
                                                        'ct_type': combobox_ct_type.get(),
                                                        'pt_type': combobox_pt_type.get(),
                                                        'sa_type': combobox_sa_type.get(),
                                                        'customer_wiring': combobox_customer_wiring.get(),
                                                        'customer_requirements': combobox_customer_requirements.get(),
                                                    }
                                                columns = ', '.join(data.keys())
                                                placeholders = ', '.join(['?' for _ in range(len(data))])
                                                query = f"INSERT INTO project_data ({columns}) VALUES ({placeholders})"

                                                values = [data.get(col) for col in data.keys()]
                                                cursor.execute(query, tuple(values))

                                                conn.commit()
                                            tk.messagebox.showwarning("提示", "数据写入完成")
                                    cursor.close()
                                    conn.close()

                                    source_file = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db'
                                    target_dir = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\backup_db'
                                    target_file = os.path.join(target_dir, os.path.basename(source_file))

                                    if os.path.exists(target_file):
                                        os.remove(target_file)

                                    shutil.copy2(source_file, target_dir)

                            else:
                                tk.messagebox.showwarning("提示", "数据库不存在，请联系管理员")
                        else:
                            tk.messagebox.showwarning("提示", "请连接内网")

            except:
                tk.messagebox.showwarning("提示", traceback.format_exc())
                # tk.messagebox.showwarning("提示", "缺少模板Excel文件")
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def is_folder_hidden(fpath):
    try:
        attrs = ctypes.windll.kernel32.GetFileAttributesW(fpath)  # attrs值为18表示该文件夹具有以下属性组合：只读 (1)、隐藏 (2) 和 子文件夹 (16)
        # print(attrs)
        if attrs != -1 and attrs & 2 == 2:  # 对于18（二进制为10010）与2（二进制为00010）进行按位与运算，结果为2（二进制为00010）
            return True
    except OSError:
        pass
    return False


def about_help(event):
    # tku.show_info("说明书")
    os.startfile(os.path.abspath('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\原理图标准设计_red.pdf'))


