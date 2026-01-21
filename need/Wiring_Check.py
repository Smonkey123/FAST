import tkinter as tk

import need.tkutils as tku

from tkinter.filedialog import askdirectory
import os
import csv

import time
import numpy as np
from time import *
import datetime
import re
import pandas as pd
import xlrd
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm, cm

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors

from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Image, Table, TableStyle, NextPageTemplate, PageBreak
import traceback

import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

FilePath = ""    # 设置一个地址变量

def main(parent, w_ratio, h_ratio, file_path):
    global checklog_file_path
    checklog_file_path = file_path
    global open_folder
    open_folder = tk.PhotoImage(file="ico\\open_folder.png")
    global analyze_file
    analyze_file = tk.PhotoImage(file="ico\\read.png")
    global pdf_export
    pdf_export = tk.PhotoImage(file="ico\\export.png")

    tk.Label(parent, text="欢迎使用线号文件检查功能", bg="#c9dbe9", fg="black", height=int(1*h_ratio), font=("ABBvoice CNSG", int(20 * h_ratio), "bold")).pack(side=tk.TOP, fill=tk.X)
    f1 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f1, text='   说明：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')
    tk.Label(f1, text='(1)EPLAN菜单MVE→Tools:KOMAX-wire production后选择No，选择Export导出；\n(2)选择线号文件【C:/Data/项目号.csv】；\n(3)对EPLAN导出的线号文件进行检查。', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    f1.pack(fill=tk.X)

    tk.Frame(parent, height=int(20*h_ratio), bg="#eaf1f6").pack(fill=tk.X)    # 水平分割线

    f2 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f2, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2, text='路径：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry    # 为了确保selectpath函数能正确调用entry,将其全局化
    entry = tk.Entry(f2, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)))
    entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

    global button_export_report
    button_export_report = tk.Button(f2, image=pdf_export, text="导出报表", font=("ABBvoice CNSG", int(13 * h_ratio)), bg="#eaf1f6", command=export_report, compound=tk.LEFT, state='disabled', activebackground='blue')
    button_export_report.pack(side=tk.RIGHT, padx=(0, int(20 * w_ratio)))

    global button_check
    button_check = tk.Button(f2, text='检查', font=("ABBvoice CNSG", int(13 * h_ratio)), image=analyze_file, bg="#eaf1f6", compound=tk.LEFT, command=process, activebackground='blue')
    button_check.pack(side=tk.RIGHT, padx=(0, int(20*w_ratio)))
    button_check['state'] = 'disabled'

    tk.Button(f2, text='选择', font=("ABBvoice CNSG", int(13 * h_ratio)), image=open_folder, bg="#eaf1f6", compound=tk.LEFT, command=selectpath, activebackground='blue').pack(side=tk.RIGHT, padx=int(20 * w_ratio))

    f2.pack(fill=tk.X)

    tk.Frame(parent, height=int(20*h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f3 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f3, text='   结果：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')

    global text
    text = tk.Text(f3, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), height=int(25 * h_ratio), width=int(65*w_ratio))
    text.pack(side=tk.LEFT, padx=(0, 1), pady=0, fill=tk.BOTH, expand=True)

    text.tag_configure('error', foreground='red')  # 设置tag

    scrollbar = tk.Scrollbar(f3)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, int(20*w_ratio)))
    scrollbar.config(command=text.yview)
    text.config(yscrollcommand=scrollbar.set)
    f3.pack(fill=tk.BOTH, expand=True)

    tk.Frame(parent, height=int(20*h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线


def selectpath():
    filepath = tk.filedialog.askopenfilename(title=u'请选择文件(项目号.csv)', filetypes=[("Excel", ".csv")])    # 选择打开什么文件，返回文件名
    if len(filepath) != 0:
        string_filename = ""
        for i in range(0, len(filepath)):
            string_filename += str(filepath[i])
        button_check['state'] = 'normal'
    else:
        button_check['state'] = 'disabled'
        button_export_report['state'] = 'disabled'

    text.delete(1.0, tk.END)    # 清空输出结果框
    entry.delete(0, "end")    # 删除entry原始内容
    entry.insert(0, filepath)    # 重新填入地址
    button_export_report['state'] = 'disabled'
    global FilePath
    FilePath = filepath

def process_strings(str1, str2):
    # 将字符串按空格分割成列表
    list1 = str1.split()
    list2 = str2.split()

    # 初始化结果变量
    result1 = []
    result2 = []

    # 处理第一种情况：A B 和 B A
    if len(list1) == 1 and len(list2) == 1:
        result1 = [list1[0]]
        result2 = [list2[0]]

    # 处理第二种情况：A B 和 B A
    elif len(list1) == 2 and len(list2) == 2:
        if list1[0] == list2[1] and list1[1] == list2[0]:
            result1 = [list1[0]]
            result2 = [list2[0]]
        else:
            result1 = list1
            result2 = list2

    # 处理第三种情况：A B C 和 C A B
    elif len(list1) == 3 and len(list2) == 3:
        if list1[2] == list2[0] and list1[0] == list2[1] and list1[1] == list2[2]:
            result1 = list1[0:2]
            result2 = [list2[0]]
        else:
            result1 = [list1[0]]
            result2 = list2[:2]

    # 处理第四种情况：A B C D 和 D A B C
    elif len(list1) == 4 and len(list2) == 4:
        if list1[0] == list2[3] and list1[1] == list2[0] and list1[2] == list2[1] and list1[3] == list2[2]:
            result1 = [list1[0]]
            result2 = list2[:3]
        elif list1[:2] == list2[2:] and list1[2:] == list2[:2]:
            result1 = list1[:2]
            result2 = list2[:2]
        else:
            result1 = list1[:3]
            result2 = [list2[0]]

    # 处理第五种情况：A B C D E 和 E A B C D
    elif len(list1) == 5 and len(list2) == 5:
        if list1[0] == list2[4] and list1[1] == list2[0] and list1[2] == list2[1] and list1[3] == list2[2] and list1[4] == list2[3]:
            result1 = [list1[0]]
            result2 = list2[:4]
        elif list1[:2] == list2[3:] and list1[2:] == list2[:3]:
            result1 = list1[:2]
            result2 = list2[:3]
        elif list1[:3] == list2[2:] and list1[3:] == list2[:2]:
            result1 = list1[:3]
            result2 = list2[:2]
        else:
            result1 = list1[:4]
            result2 = [list2[0]]

    # 处理第六种情况：A B C D E F 和 F A B C D E
    elif len(list1) == 6 and len(list2) == 6:
        if list1[0] == list2[5] and list1[1] == list2[0] and list1[2] == list2[1] and list1[3] == list2[2] and list1[4] == list2[3] and list1[5] == list2[4]:
            result1 = [list1[0]]
            result2 = list2[:5]
        elif list1[:2] == list2[4:] and list1[2:] == list2[:4]:
            result1 = list1[:2]
            result2 = list2[:4]
        elif list1[:3] == list2[3:] and list1[3:] == list2[:3]:
            result1 = list1[:3]
            result2 = list2[:3]
        elif list1[:4] == list2[2:] and list1[4:] == list2[:2]:
            result1 = list1[:4]
            result2 = list2[:2]
        else:
            result1 = list1[:5]
            result2 = [list2[0]]

    # 处理第七种情况：A B C D E F G 和 G A B C D E F
    elif len(list1) == 7 and len(list2) == 7:
        if list1[0] == list2[6] and list1[1] == list2[0] and list1[2] == list2[1] and list1[3] == list2[2] and list1[4] == list2[3] and list1[5] == list2[4] and list1[6] == list2[5]:
            result1 = [list1[0]]
            result2 = list2[:6]
        elif list1[:2] == list2[5:] and list1[2:] == list2[:5]:
            result1 = list1[:2]
            result2 = list2[:5]
        elif list1[:3] == list2[4:] and list1[3:] == list2[:4]:
            result1 = list1[:3]
            result2 = list2[:4]
        elif list1[:4] == list2[3:] and list1[4:] == list2[:3]:
            result1 = list1[:4]
            result2 = list2[:3]
        elif list1[:5] == list2[2:] and list1[5:] == list2[:2]:
            result1 = list1[:5]
            result2 = list2[:2]
        else:
            result1 = list1[:6]
            result2 = [list2[0]]

    # 处理第八种情况：A B C D E F G H 和 H A B C D E F G
    elif len(list1) == 8 and len(list2) == 8:
        if list1[0] == list2[7] and list1[1] == list2[0] and list1[2] == list2[1] and list1[3] == list2[2] and list1[4] == list2[3] and list1[5] == list2[4] and list1[6] == list2[5] and list1[7] == list2[6]:
            result1 = [list1[0]]
            result2 = list2[:7]
        elif list1[:2] == list2[6:] and list1[2:] == list2[:6]:
            result1 = list1[:2]
            result2 = list2[:6]
        elif list1[:3] == list2[5:] and list1[3:] == list2[:5]:
            result1 = list1[:3]
            result2 = list2[:5]
        elif list1[:4] == list2[4:] and list1[4:] == list2[:4]:
            result1 = list1[:4]
            result2 = list2[:4]
        elif list1[:5] == list2[3:] and list1[5:] == list2[:3]:
            result1 = list1[:5]
            result2 = list2[:3]
        elif list1[:6] == list2[2:] and list1[6:] == list2[:2]:
            result1 = list1[:6]
            result2 = list2[:2]
        else:
            result1 = list1[:7]
            result2 = [list2[0]]

        # 默认情况：直接返回原始列表
    else:
        result1 = list1
        result2 = list2

        # 将结果列表转换回字符串
    result1_str = ' '.join(result1)
    result2_str = ' '.join(result2)

    return result1_str, result2_str

def process():
    try:
        error1_calculator = 0
        error2_calculator = 0
        error3_calculator = 0
        error4_calculator = 0
        global error5_calculator
        error5_calculator = 0
        error6_calculator = 0
        text.delete(1.0, tk.END)  # 清空输出结果框
        if FilePath == "":
            tk.messagebox.showwarning("提示", "请选择文件！")

        else:
            text.insert(tk.INSERT, '>>>线号文件正在检查中...\n')    # 进行From/To检查
            start = time()
            data = []

            # 数据读入与预处理
            with open(FilePath, 'r', encoding="utf-16") as csv_file:
                # 读取csv文件中每行的列表，将每行读取的值作为列表返回
                csv_read = csv.reader(csv_file)
                for row in csv_read:
                    if len(list(row)) >= 16:    # 正常数据行长度为16，异常线号数据行长度小于16
                        data.append(list(row))    # data考虑了标题行

            if len(data) == 0:
                tk.messagebox.showwarning("提示", "线号csv文件内容为空或被破坏\n注意：csv文件只能用记事本打开修改，不能用Excel打开保存")
            else:
                typical = []
                for i in range(1, len(data)):    # 注意data第一行为列名行，不是有效行
                    data[i][4], data[i][6] = process_strings(data[i][4], data[i][6])
                    if len(data[i]) >= 16 and data[i][1] != 'ZS1':    # 正常数据行长度为16，异常线号数据行长度小于16
                        typical.append(data[i][0])

                Typical = list(set(typical))    # set去重后顺序会改变
                Typical.sort(key=typical.index)    # 保证去重后顺序不变

                panel = []    # 线号文件表格同一个Typical下也会有多个Panel的信息,正常只需要一个Typical的一个Panel的线号即可
                DATA = []

                for i in range(0, len(Typical)):
                    # print(Typical[i])
                    for j in range(1, len(data)):
                        if len(data[j]) >= 16:  # 正常数据行长度为16，异常线号数据行长度小于16
                            if data[j][0] == Typical[i]:
                                panel.append(data[j][14])

                    Panel = list(set(panel))    # set去重后顺序会改变
                    Panel.sort(key=panel.index)    # 保证去重后顺序不变

                    normal_flag = []
                    for j, p_i in enumerate(Panel):
                        normal_flag.append(False)  # 初始化 normal_flag
                        for k in range(1, len(data)):
                            if len(data[k]) >= 16 and data[k][0] == Typical[i] and data[k][14] == p_i and data[k][11] != 'C' and data[k][11] != 'E':
                                normal_flag[j] = True

                    # 检查 normal_flag 中 True 的数量
                    true_indices = [idx for idx, flag in enumerate(normal_flag) if flag]
                    # print(Panel, normal_flag)
                    if len(true_indices) > 0:
                        # 如果 normal_flag 中有超过1个True，则取第一个True对应的Panel的数据
                        # 如果只有一个True，则取这唯一一个True对应的Panel数据
                        selected_panel = Panel[true_indices[0]]

                        for k in range(1, len(data)):
                            if len(data[k]) >= 16 and data[k][14] == selected_panel:
                                DATA.append(data[k])  # 真正符合条件的线号文件是DATA，它已经处理为多个Typical,每个Typical只有一个Panel的线号

                    panel = []  # 清空 panel 以进行下一个 Typical 的处理

                text.insert(tk.INSERT, '>>>线号文件去重(每个Typical只分析1个Panel)，%d条数据变化为%d条...\n' % (np.array(data).shape[0]-1, np.array(DATA).shape[0]))    # 进行From/To检查
                text.insert(tk.INSERT, "\n>>>缺少定位/跨区域接线/线号颠倒正在检查中... \n")
                error_flag = False

                # 进行检查
                for i in range(0, len(DATA)):
                    if len(DATA[i]) >= 16 and DATA[i][1] != 'ZS1':    # 正常数据行长度为16，异常线号数据行长度小于16
                        if DATA[i][1] == '':
                            if 'GND' not in DATA[i][4] and 'GND:' not in DATA[i][4] and 'GND：' not in DATA[i][4] and 'SHIELD:' not in DATA[i][4] and 'PE' not in DATA[i][4] and 'XB1' not in DATA[i][4] and 'XB2' not in DATA[i][4] and 'XB3' not in DATA[i][4] and DATA[i][11] != 'E' and DATA[i][11] != 'C':
                                text.insert(tk.INSERT, '▲ ' + DATA[i][0] + "柜 " + DATA[i][4] + "( " + DATA[i][1] + " )" + "—" + DATA[i][6] + "(" + DATA[i][9] + ")" + "【左端缺少定位】\n", 'error')
                                error1_calculator += 1
                                error_flag = True

                            if DATA[i][9] != '':
                                text.insert(tk.INSERT, '▲ ' + DATA[i][0] + "柜 " + DATA[i][4] + "( " + DATA[i][1] + " )" + "—" + DATA[i][6] + "(" + DATA[i][9] + ")" + "【线号颠倒】\n", 'error')
                                error1_calculator += 1
                                error_flag = True
                        else:
                            if DATA[i][9] != '':
                                if DATA[i][9] != 'LV' and DATA[i][9] != 'LV.F' and DATA[i][9] != 'LV.M1' and DATA[i][9] != 'LV.M2' and DATA[i][9] != 'LV.R1' and DATA[i][9] != 'LV.R2' and DATA[i][9] != 'LV.L1' and DATA[i][9] != 'LV.L2':
                                    if DATA[i][1] == 'LV' or DATA[i][1] == 'LV.F' or DATA[i][1] == 'LV.M1' or DATA[i][1] == 'LV.M2' or DATA[i][1] == 'LV.R1' or DATA[i][1] == 'LV.R2' or DATA[i][1] == 'LV.L1' or DATA[i][1] == 'LV.L2':
                                        text.insert(tk.INSERT, '▲ ' + DATA[i][0] + "柜 " + DATA[i][4] + "(" + DATA[i][1] + ")" + "—" + DATA[i][6] + "(" + DATA[i][9] + ")" + "【线号颠倒】\n", 'error')
                                        error1_calculator += 1
                                        error_flag = True

                                if DATA[i][1] == 'LVD' and DATA[i][9] == 'MV':
                                    text.insert(tk.INSERT, '▲ ' + DATA[i][0] + "柜 " + DATA[i][4] + "(" + DATA[i][1] + ")" + "—" + DATA[i][6] + "(" + DATA[i][9] + ")" + "【跨区域接线且线号颠倒】\n", 'error')
                                    error1_calculator += 1
                                    error_flag = True

                                if DATA[i][1] == 'MV' and DATA[i][9] == 'LVD':
                                    text.insert(tk.INSERT, '▲ ' + DATA[i][0] + "柜 " + DATA[i][4] + "(" + DATA[i][1] + ")" + "—" + DATA[i][6] + "(" + DATA[i][9] + ")" + "【跨区域接线】\n", 'error')
                                    error1_calculator += 1
                                    error_flag = True

                                if DATA[i][1] == 'LVD' and DATA[i][9] == 'TR':
                                    text.insert(tk.INSERT, '▲ ' + DATA[i][0] + "柜 " + DATA[i][4] + "(" + DATA[i][1] + ")" + "—" + DATA[i][6] + "(" + DATA[i][9] + ")" + "【跨区域接线且线号颠倒】\n", 'error')
                                    error1_calculator += 1
                                    error_flag = True

                                if DATA[i][1] == 'TR' and DATA[i][9] == 'LVD':
                                    text.insert(tk.INSERT, '▲ ' + DATA[i][0] + "柜 " + DATA[i][4] + "(" + DATA[i][1] + ")" + "—" + DATA[i][6] + "(" + DATA[i][9] + ")" + "【跨区域接线】\n", 'error')
                                    error1_calculator += 1
                                    error_flag = True

                                if DATA[i][1] == 'LVD' and DATA[i][9] == 'ITR':
                                    text.insert(tk.INSERT, '▲ ' + DATA[i][0] + "柜 " + DATA[i][4] + "(" + DATA[i][1] + ")" + "—" + DATA[i][6] + "(" + DATA[i][9] + ")" + "【跨区域接线且线号颠倒】\n", 'error')
                                    error1_calculator += 1
                                    error_flag = True

                                if DATA[i][1] == 'ITR' and DATA[i][9] == 'LVD':
                                    text.insert(tk.INSERT, '▲ ' + DATA[i][0] + "柜 " + DATA[i][4] + "(" + DATA[i][1] + ")" + "—" + DATA[i][6] + "(" + DATA[i][9] + ")" + "【跨区域接线】\n", 'error')
                                    error1_calculator += 1
                                    error_flag = True

                                # V1.8新增，当本柜高压侧向其他柜的低压室时，虽然满足高压到低压（无跨区域接线），但由于跨柜子，也属于跨区域接线
                                if DATA[i][11] == 'C':
                                    conditions = ['MV', 'TR', 'ITR', 'LVD']  # 将所有可能的条件放在一个列表中
                                    for condition in conditions:
                                        if DATA[i][1] == condition or DATA[i][9] == condition:
                                            text.insert(tk.INSERT, '▲ ' + DATA[i][0] + "柜 " + DATA[i][4] + "(" + DATA[i][1] + ")" + "—" + DATA[i][6] + "(" + DATA[i][9] + ")" + "【跨柜体接线，应先经端子过渡】\n", 'error')
                                            error1_calculator += 1
                                            error_flag = True
                                            break  # 如果找到了匹配的条件，就不需要继续检查其他条件，直接退出循环

                                    lv_values = {'LV', 'LV.F', 'LV.M1', 'LV.M2', 'LV.R1', 'LV.R2', 'LV.L1', 'LV.L2'}

                                    if DATA[i][1] in lv_values and DATA[i][9] in lv_values and ('X' not in DATA[i][4] and 'X' not in DATA[i][6]):
                                        text.insert(tk.INSERT, '▲ ' + DATA[i][0] + "柜 " + DATA[i][4] + "(" + DATA[i][1] + ")" + "—" + DATA[i][6] + "(" + DATA[i][9] + ")" + "【跨柜体接线，应先经端子过渡】\n", 'error')
                                        error1_calculator += 1
                                        error_flag = True

                                else:
                                    if (DATA[i][1] == 'LVD' or DATA[i][1] == 'MV' or DATA[i][1] == 'TR' or DATA[i][1] == 'ITR') and (DATA[i][9] == 'LV' or DATA[i][9] == 'LV.F' or DATA[i][9] == 'LV.M1' or DATA[i][9] == 'LV.M2' or DATA[i][9] == 'LV.R1' or DATA[i][9] == 'LV.R2' or DATA[i][9] == 'LV.L1' or DATA[i][9] == 'LV.L2') and 'X' not in DATA[i][6] and ':RJ12' not in DATA[i][6]:
                                        text.insert(tk.INSERT, '▲ ' + DATA[i][0] + "柜 " + DATA[i][4] + "(" + DATA[i][1] + ")" + "—" + DATA[i][6] + "(" + DATA[i][9] + ")" + "【跨区域接线，应先经端子过渡】\n", 'error')
                                        error1_calculator += 1
                                        error_flag = True
                                # if (DATA[i][1] == 'MV' or DATA[i][1] == 'TR' or DATA[i][1] == 'ITR') and  (DATA[i][9] == 'MV' or DATA[i][9] == 'TR' or DATA[i][9] == 'ITR') and not ('QCE:' in DATA[i][4] and 'QCE:' in DATA[i][4]):
                                #     text.insert(tk.INSERT, DATA[i][0] + "柜 " + DATA[i][4] + "(" + DATA[i][1] + ")" + "—" + DATA[i][6] + "(" + DATA[i][9] + ")" + "【可能有问题】\n", 'error')

                            else:
                                if 'GND' not in DATA[i][6] and 'GND:' not in DATA[i][6] and 'GND：' not in DATA[i][6] and 'SHIELD:' not in DATA[i][6] and 'PE' not in DATA[i][6] and 'XB1' not in DATA[i][6] and 'XB2' not in DATA[i][6] and 'XB3' not in DATA[i][6] and DATA[i][11] != 'E' and DATA[i][11] != 'C':
                                    text.insert(tk.INSERT, '▲ ' + DATA[i][0] + "柜 " + DATA[i][4] + "(" + DATA[i][1] + ")" + "—" + DATA[i][6] + "( " + DATA[i][9] + " )" + "【右端缺少定位】\n", 'error')
                                    error1_calculator += 1
                                    error_flag = True

                if not error_flag:
                    text.insert(tk.INSERT, "无缺少定位/跨区域接线/线号颠倒问题\n")
                # text.insert(tk.INSERT, ">>>缺少定位/跨区域接线/线号颠倒——————检查完成 \n")

                text.insert(tk.INSERT, "\n>>>设备并接线线型一致性正在检查中... \n")
                error_type_index = []
                seen_pairs = set()  # 用于记录已经出现的索引对

                error_flag = False
                for i in range(0, len(DATA)):
                    if len(DATA[i]) >= 16 and DATA[i][1] != 'ZS1':    # 正常数据行长度为16，异常线号数据行长度小于16
                        for j in range(0, len(DATA)):
                            if len(DATA[j]) >= 16:    # 正常数据行长度为16，异常线号数据行长度小于16
                                if j != i:
                                    if DATA[i][0] == DATA[j][0] and DATA[i][4] == DATA[j][4] and DATA[i][17] != DATA[j][17] and DATA[i][4].split(':')[1] != '' and not is_x_digit_format(DATA[i][4][0:2]) and DATA[i][6] != 'GND:' and DATA[j][6] != 'GND:':
                                        if (DATA[i][17] == 'RV' and DATA[j][17] == '') or (DATA[i][17] == '' and DATA[j][17] == 'RV'):
                                            continue
                                        pair1 = (i, j)
                                        pair2 = (j, i)
                                        # 仅当两个对都不在 seen_pairs 中时才添加
                                        if pair1 not in seen_pairs and pair2 not in seen_pairs:
                                            error_type_index.append(pair1)
                                            seen_pairs.add(pair1)

                                    if DATA[i][0] == DATA[j][0] and DATA[i][4] == DATA[j][6] and DATA[i][17] != DATA[j][17] and DATA[i][4].split(':')[1] != '' and not is_x_digit_format(DATA[i][4][0:2]) and DATA[i][6] != 'GND:' and DATA[j][4] != 'GND:':
                                        if (DATA[i][17] == 'RV' and DATA[j][17] == '') or (DATA[i][17] == '' and DATA[j][17] == 'RV'):
                                            continue
                                        pair1 = (i, j)
                                        pair2 = (j, i)
                                        # 仅当两个对都不在 seen_pairs 中时才添加
                                        if pair1 not in seen_pairs and pair2 not in seen_pairs:
                                            error_type_index.append(pair1)
                                            seen_pairs.add(pair1)

                                    if DATA[i][0] == DATA[j][0] and DATA[i][6] == DATA[j][4] and DATA[i][17] != DATA[j][17] and DATA[i][6].split(':')[1] != '' and not is_x_digit_format(DATA[i][6][0:2]) and DATA[i][4] != 'GND:' and DATA[j][6] != 'GND:':
                                        if (DATA[i][17] == 'RV' and DATA[j][17] == '') or (DATA[i][17] == '' and DATA[j][17] == 'RV'):
                                            continue
                                        pair1 = (i, j)
                                        pair2 = (j, i)
                                        # 仅当两个对都不在 seen_pairs 中时才添加
                                        if pair1 not in seen_pairs and pair2 not in seen_pairs:
                                            error_type_index.append(pair1)
                                            seen_pairs.add(pair1)

                                    if DATA[i][0] == DATA[j][0] and DATA[i][6] == DATA[j][6] and DATA[i][17] != DATA[j][17] and DATA[i][6].split(':')[1] != '' and not is_x_digit_format(DATA[i][6][0:2]) and DATA[i][4] != 'GND:' and DATA[j][4] != 'GND:':
                                        if (DATA[i][17] == 'RV' and DATA[j][17] == '') or (DATA[i][17] == '' and DATA[j][17] == 'RV'):
                                            continue
                                        pair1 = (i, j)
                                        pair2 = (j, i)
                                        # 仅当两个对都不在 seen_pairs 中时才添加
                                        if pair1 not in seen_pairs and pair2 not in seen_pairs:
                                            error_type_index.append(pair1)
                                            seen_pairs.add(pair1)

                for w1, w2 in error_type_index:
                    text.insert(tk.INSERT, '▲ ' + DATA[w1][0] + "柜 " + DATA[w1][4] + "—" + DATA[w1][6] + "(" + DATA[w1][17] + ")" + "与" + DATA[w2][4] + "—" + DATA[w2][6] + "(" + DATA[w2][17] + ")" + "【软硬线型不一致】\n", 'error')
                    error2_calculator += 1
                    error_flag = True

                if not error_flag:
                    text.insert(tk.INSERT, "无设备并接线线型不一致问题\n")

                text.insert(tk.INSERT, "\n>>>断路器/三工位航空插多接线正在检查中... \n")
                error_flag = False
                for i in range(0, len(DATA)):
                    if len(DATA[i]) >= 16 and DATA[i][1] != 'ZS1':    # 正常数据行长度为16，异常线号数据行长度小于16
                        if i >= 2 and DATA[i][0] == DATA[i-1][0] and DATA[i][4] == DATA[i-1][4] and "QAB" in DATA[i][4]:
                            text.insert(tk.INSERT, '▲ ' + DATA[i-1][0] + "柜 " + DATA[i-1][4] + "(" + DATA[i-1][1] + ")" + "—" + DATA[i-1][6] + "(" + DATA[i-1][9] + ")" + "【断路器航空插误接两根线】\n", 'error')
                            text.insert(tk.INSERT, '▲ ' + DATA[i][0] + "柜 " + DATA[i][4] + "(" + DATA[i][1] + ")" + "—" + DATA[i][6] + "(" + DATA[i][9] + ")" + "【断路器航空插误接两根线】\n", 'error')
                            error3_calculator += 1
                            error_flag = True

                        if i >= 2 and DATA[i][0] == DATA[i-1][0] and DATA[i][4] == DATA[i-1][4] and "QBS" in DATA[i][4]:
                            text.insert(tk.INSERT, '▲ ' + DATA[i-1][0] + "柜 " + DATA[i-1][4] + "(" + DATA[i-1][1] + ")" + "—" + DATA[i-1][6] + "(" + DATA[i-1][9] + ")" + "【三工位航空插误接两根线】\n", 'error')
                            text.insert(tk.INSERT, '▲ ' + DATA[i][0] + "柜 " + DATA[i][4] + "(" + DATA[i][1] + ")" + "—" + DATA[i][6] + "(" + DATA[i][9] + ")" + "【三工位航空插误接两根线】\n", 'error')
                            error3_calculator += 1
                            error_flag = True

                        if i >= 3 and DATA[i][0] == DATA[i-1][0] == DATA[i-2][0] and DATA[i][4] == DATA[i-1][4] == DATA[i-2][4] and "QAB" in DATA[i][4]:
                            text.insert(tk.INSERT, '▲ ' + DATA[i-2][0] + "柜 " + DATA[i-2][4] + "(" + DATA[i-2][1] + ")" + "—" + DATA[i-2][6] + "(" + DATA[i-2][9] + ")" + "【断路器航空插误接三根线】\n", 'error')
                            text.insert(tk.INSERT, '▲ ' + DATA[i-1][0] + "柜 " + DATA[i-1][4] + "(" + DATA[i-1][1] + ")" + "—" + DATA[i-1][6] + "(" + DATA[i-1][9] + ")" + "【断路器航空插误接三根线】\n", 'error')
                            text.insert(tk.INSERT, '▲ ' + DATA[i][0] + "柜 " + DATA[i][4] + "(" + DATA[i][1] + ")" + "—" + DATA[i][6] + "(" + DATA[i][9] + ")" + "【断路器航空插误接两根线】\n", 'error')
                            error3_calculator += 1
                            error_flag = True

                        if i >= 3 and DATA[i][0] == DATA[i-1][0] == DATA[i-2][0] and DATA[i][4] == DATA[i-1][4] == DATA[i-2][4] and "QBS" in DATA[i][4]:
                            text.insert(tk.INSERT, '▲ ' + DATA[i-2][0] + "柜 " + DATA[i-2][4] + "(" + DATA[i-2][1] + ")" + "—" + DATA[i-2][6] + "(" + DATA[i-2][9] + ")" + "【三工位航空插误接三根线】\n", 'error')
                            text.insert(tk.INSERT, '▲ ' + DATA[i-1][0] + "柜 " + DATA[i-1][4] + "(" + DATA[i-1][1] + ")" + "—" + DATA[i-1][6] + "(" + DATA[i-1][9] + ")" + "【三工位航空插误接三根线】\n", 'error')
                            text.insert(tk.INSERT, '▲ ' + DATA[i][0] + "柜 " + DATA[i][4] + "(" + DATA[i][1] + ")" + "—" + DATA[i][6] + "(" + DATA[i][9] + ")" + "【三工位航空插误接两根线】\n", 'error')
                            error3_calculator += 1
                            error_flag = True

                if not error_flag:
                    text.insert(tk.INSERT, "无断路器/三工位航空插多接线问题\n")
                # text.insert(tk.INSERT, ">>>断路器/三工位航空插多接线——————检查完成 \n")

                # 设备节点多接线检查功能
                text.insert(tk.INSERT, "\n>>>设备节点多接线正在检查中... \n")
                stem, suffix = os.path.splitext(os.path.basename(FilePath))    # stem是文件名,suffix是后缀
                inputfile = "C:/Temp/"+stem+"-Files/"+stem+"-DeviceLabel.xlsx"

                error_flag = False

                if not os.path.exists(inputfile):
                    text.insert(tk.INSERT, '▲ ' + "设备节点多接线检查失败,找不到%s文件! \n" % inputfile, 'error')
                    error_flag = True
                else:
                    book1 = load_workbook(inputfile)
                    sheet1 = book1['Z5_xlsx']
                    A1 = []
                    B1 = []
                    C1 = []
                    D1 = []
                    for i in range(3, sheet1.max_row + 1):
                        A1.append(str(sheet1.cell(row=i, column=1).value))  # Typical列
                        B1.append(str(sheet1.cell(row=i, column=2).value))  # DT列
                        C1.append(str(sheet1.cell(row=i, column=3).value))  # Zone列
                        D1.append(str(sheet1.cell(row=i, column=4).value))  # PartNumber列

                    # 筛选后的设备清单数据
                    Typical_Device = []
                    DT_Device = []
                    PartNumber_Device = []
                    # print(A1,B1,C1,D1)
                    for i in range(0, len(A1)):  # 项目号-DeviceLabel.xlsx
                        if C1[i] != 'None' and D1[i] != 'Mounting_Panel':
                            Typical_Device.append(A1[i])
                            DT_Device.append(B1[i])
                            PartNumber_Device.append(D1[i])

                    from_to = []
                    for i in range(0, len(Typical)):    # 遍历每一个Typical下的线号
                        for j in range(0, len(DATA)):
                            if len(DATA[j]) >= 16 and DATA[j][0] == Typical[i]:    # 正常数据行长度为16，异常线号数据行长度小于16
                                from_to.append(DATA[j][4])
                                from_to.append(DATA[j][6])    # 将每个Typical下的From和To线号都存储起来
                        result = pd.value_counts(from_to)    # 对线号列表进行计数，结果是Series形式(index-value)，一列是线号的种类，一列是对应的数量
                        for k in range(0, len(result)):
                            if result[k] >= 3:    # 对数量列(Value)进行遍历
                                for l in range(0, len(Typical_Device)):
                                    if Typical[i] == Typical_Device[l] and str((result.index.tolist()[k]).split(":")[0]) == DT_Device[l]:
                                        text.insert(tk.INSERT, '▲ ' + Typical[i] + "柜 " + result.index.tolist()[k] + "【节点超过三根线】\n", 'error')
                                        error4_calculator += 1
                                        error_flag = True
                        from_to = []
                    # text.insert(tk.INSERT, ">>>设备节点多接线——————检查完成 \n")
                if not error_flag:
                    text.insert(tk.INSERT, "无设备节点多接线问题\n")

                text.insert(tk.INSERT, "\n>>>设备主功能定位与节点定位一致性正在检查中... \n")
                stem, suffix = os.path.splitext(os.path.basename(FilePath))  # stem是文件名,suffix是后缀
                bomfile = "C:/Temp/" + stem + "-Files/" + stem + "-BOM.xlsx"
                pbomfile = "C:/Temp/" + stem + "-Files/" + stem + "-PBOM.xlsx"

                bomfile_exist = True
                pbomfile_exist = True

                global error_flag_spec
                error_flag_spec = False

                if not os.path.exists(bomfile):
                    bomfile_exist = False
                    text.insert(tk.INSERT, '▲ ' + "设备主功能定位与节点定位一致性检查失败，找不到%s文件! \n" % bomfile, 'error')
                    error_flag_spec = True

                if not os.path.exists(pbomfile):
                    pbomfile_exist = False
                    text.insert(tk.INSERT, '▲ ' + "设备主功能定位与节点定位一致性检查失败，找不到%s文件! \n" % pbomfile, 'error')
                    error_flag_spec = True

                if bomfile_exist and pbomfile_exist:
                    check_data_accuracy(DATA, bomfile, pbomfile)

                if not error_flag_spec:
                    text.insert(tk.INSERT, "无设备主功能定位与节点定位不一致问题\n")

                text.insert(tk.INSERT, "\n>>>设备未接线正在检查中... \n")

                error_flag_lack_wiring = False

                if not os.path.exists(bomfile):
                    text.insert(tk.INSERT, '▲ ' + "设备未接线检查失败，找不到%s文件! \n" % bomfile, 'error')
                    error_flag_lack_wiring = True
                else:
                    bom_book = load_workbook(bomfile)
                    bom_sheet = bom_book['Z6_xlsx']
                    A = []
                    B = []
                    C = []
                    D = []
                    E = []
                    F = []
                    G = []
                    for i in range(2, bom_sheet.max_row + 1):
                        A.append(str(bom_sheet.cell(row=i, column=1).value))  # Hight-level列
                        B.append(str(bom_sheet.cell(row=i, column=2).value))  # Zone列
                        C.append(str(bom_sheet.cell(row=i, column=3).value))  # DT列
                        D.append(str(bom_sheet.cell(row=i, column=4).value))  # PartNumber列
                        E.append(str(bom_sheet.cell(row=i, column=5).value))  # Qty列
                        F.append(str(bom_sheet.cell(row=i, column=6).value))  # Designation列
                        G.append(str(bom_sheet.cell(row=i, column=7).value))  # Type列

                    A_real = []
                    B_real = []
                    C_real = []
                    D_real = []
                    E_real = []
                    F_real = []
                    G_real = []

                    for i in range(0, len(A)):
                        if B[i] != 'A' or C[i] != 'Panel':  # 剔除无效数据
                            A_real.append(A[i])
                            B_real.append(B[i])
                            C_real.append(C[i])
                            D_real.append(D[i])
                            E_real.append(E[i])
                            F_real.append(F[i])
                            G_real.append(G[i])

                    # for i in range(0, len(A_real)):
                    A_real_set = list(set(A_real))  # set去重后顺序会改变
                    A_real_set.sort(key=list(A_real).index)  # 保证去重后顺序不变

                    for i in range(0, len(A_real)):
                        # 检查 A_real[i] 是否包含形如 _1, _2 这样的模式
                        if '_' in A_real[i]:
                            # 分割字符串，并检查下划线后的部分是否为数字
                            parts = A_real[i].split('_')
                            if len(parts) > 1 and parts[1].isdigit():
                                continue  # 如果下划线后面是数字，跳过这个 A_real[i]

                        elif C_real[i] in ['BAD', 'FA', 'FCF']:
                            continue
                        # elif B_real[i] != 'LVD':
                        #     continue

                        exists = False
                        for j in range(0, len(DATA)):
                            if A_real[i] == DATA[j][0]:
                                if C_real[i] in DATA[j][4] or C_real[i] in DATA[j][6]:
                                    exists = True
                                    break
                        if not exists and B_real[i] != 'LVD':
                            text.insert(tk.INSERT, '▲ ' + A_real[i] + "柜 " + C_real[i] + "【设备未接线】\n", 'error')
                            error6_calculator += 1
                            error_flag_lack_wiring = True
                        if not exists and B_real[i] == 'LVD':
                            text.insert(tk.INSERT, '▲ ' + A_real[i] + "柜 " + C_real[i] + "【定位LVD，设备未接线】\n", 'error')
                            error6_calculator += 1
                            error_flag_lack_wiring = True

                if not error_flag_lack_wiring:
                    text.insert(tk.INSERT, "无设备未接线问题\n")

                end = time()
                text.insert(tk.INSERT, "\n>>>线号文件检查完成! 用时%.3f秒\n" % (end - start))
                button_export_report['state'] = 'normal'

                global checklog_file_path
                checklogbook = load_workbook(checklog_file_path)
                checklogsheet = checklogbook['Sheet']
                project_no = stem[0:9]
                current_time = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
                data1 = [project_no, '线号检查-缺少定位/跨区域接线/线号颠倒', error1_calculator, current_time]
                data2 = [project_no, '线号检查-设备并接线线型一致性', error2_calculator, current_time]
                data3 = [project_no, '线号检查-断路器/三工位航空插多接线', error3_calculator, current_time]
                data4 = [project_no, '线号检查-设备节点多接线', error4_calculator, current_time]
                data5 = [project_no, '线号检查-设备主功能定位与节点定位一致性', error5_calculator, current_time]
                data6 = [project_no, '线号检查-设备未接线', error6_calculator, current_time]
                checklogsheet.append(data1)
                checklogsheet.append(data2)
                checklogsheet.append(data3)
                checklogsheet.append(data4)
                checklogsheet.append(data5)
                checklogsheet.append(data6)

                checklogbook.save(checklog_file_path)

    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


# 检查字符串中是否包含中文字符的函数
def contains_chinese(s):
    return any('\u4e00' <= c <= '\u9fff' for c in s)


# 检查字符串是否为X + 数字形式的函数
def is_x_digit_format(s):
    return bool(re.match(r"X\d+", s))


def process_data(value1, value2, value3):
    # 判断value2是否包含 ":" 且 ":" 后有内容
    if ':' in value2:
        val2_split = value2.split(':', 1)  # 只分割一次，以防有多个冒号

        part_before_colon = val2_split[0].strip()

        # 判断冒号后是否有内容，并且value2前半部分不含中文，也不是 'X' 开头后跟数字的形式
        if (len(val2_split) > 1 and val2_split[1].strip() and
                not contains_chinese(part_before_colon) and
                not is_x_digit_format(part_before_colon)):
            a = value1.strip()
            b = part_before_colon
            c = value3.strip()
            return a, b, c
    # 如果没有 ':' 或 ':' 后没有内容，或者前半部分包含中文或者是 'X' 开头后跟数字的形式，返回 None
    return None


def check_data_accuracy(DATA, bomfile, pbomfile):
    global error5_calculator
    global error_flag_spec
    book1 = load_workbook(bomfile)
    sheet1 = book1['Z6_xlsx']
    A1 = []
    B1 = []
    C1 = []
    for i in range(2, sheet1.max_row + 1):
        A1.append(str(sheet1.cell(row=i, column=1).value))
        B1.append(str(sheet1.cell(row=i, column=2).value))
        C1.append(str(sheet1.cell(row=i, column=3).value))
    bomfile_data = list(zip(A1, B1, C1))

    book2 = load_workbook(pbomfile)
    sheet2 = book2['Z7_xlsx']
    A2 = []
    B2 = []
    C2 = []
    for i in range(2, sheet2.max_row + 1):
        A2.append(str(sheet2.cell(row=i, column=1).value))
        B2.append(str(sheet2.cell(row=i, column=2).value))
        C2.append(str(sheet2.cell(row=i, column=3).value))
    pbomfile_data = list(zip(A2, B2, C2))

    # print(bomfile_data+pbomfile_data)

    for row in DATA:
        # 以第五列和第二列数据为筛选条件
        res1 = process_data(row[0], row[4], row[1])
        # print(res1)
        if res1:
            a1, b1, c1 = res1
            for record in bomfile_data+pbomfile_data:
                # 当a与record[0]一致且b与record[2]一致时，检查c与record[1]是否一致
                if record[0].strip() == a1 and record[2].strip() == b1 and record[1].strip() != c1:
                    if record[1].strip() in ['LV', 'LV.F', 'LV.M1', 'LV.M2', 'LV.R1', 'LV.R2', 'LV.L1', 'LV.L2'] and c1 in ['LV', 'LV.F', 'LV.M1', 'LV.M2', 'LV.R1', 'LV.R2', 'LV.L1', 'LV.L2']:
                        continue
                    text.insert(tk.INSERT, '▲ ' + a1 + "柜 " + b1 + "的主功能定位为" + record[1].strip() + "，而节点"+row[4]+"定位为"+c1+'\n', 'error')
                    error5_calculator += 1
                    error_flag_spec = True

        # 以第七列和第十列数据为筛选条件
        res2 = process_data(row[0], row[6], row[9])
        # print(res2)
        if res2:
            a2, b2, c2 = res2
            for record in bomfile_data+pbomfile_data:
                # 当a与record[0]一致且b与record[2]一致时，检查c与record[1]是否一致
                if record[0].strip() == a2 and record[2].strip() == b2 and record[1].strip() != c2:
                    if record[1].strip() in ['LV', 'LV.F', 'LV.M1', 'LV.M2', 'LV.R1', 'LV.R2', 'LV.L1', 'LV.L2'] and c2 in ['LV', 'LV.F', 'LV.M1', 'LV.M2', 'LV.R1', 'LV.R2', 'LV.L1', 'LV.L2']:
                        continue
                    text.insert(tk.INSERT, '▲ ' + a2 + "柜 " + b2 + "的主功能定位为" + record[1].strip() + "，而节点"+row[6]+"定位为"+c2+'\n', 'error')
                    error5_calculator += 1
                    error_flag_spec = True


def export_report():
    try:
        pdfmetrics.registerFont(TTFont('SimSun', 'simsun.ttc'))  # 注册字体
        stem, _ = os.path.splitext(os.path.basename(FilePath))
        home_path = os.path.expanduser("~")
        desktop_path = os.path.join(home_path, "Desktop")
        global project_number
        project_number = stem.replace('-Terminal list', '')
        desktop_path = askdirectory(title=u'请选择导出文件夹', initialdir=desktop_path)
        if not desktop_path:
            tk.messagebox.showwarning("提示", "未选择保存路径，导出操作已取消")
            return

        doc = SimpleDocTemplate(os.path.join(desktop_path, '线号检查报告-%s.pdf' % project_number), pagesize=A4, rightMargin=40, leftMargin=40, topMargin=100, bottomMargin=20)
        story = []
        # 获取样式
        styles = getSampleStyleSheet()
        style = styles['Normal']
        style.fontName = 'SimSun'
        style.wordWrap = 'CJK'
        style.leading = 15  # 行间距

        text_content = text.get("1.0", tk.END)
        start_index = text_content.find(">>>线号文件正在检查中...")
        end_index = find_end_index(text_content, "\n>>>线号文件检查完成!")

        text_content = text_to_html(text, start_index, end_index)

        p = Paragraph(text_content,style)
        story.append(p)


        global page_number
        page_number = 1

        # 通过`SimpleDocTemplate`直接生成文档，传入`add_header`函数绘制每页的头部内容
        doc.build(story, onFirstPage=add_header1, onLaterPages=add_header2)

        tk.messagebox.showwarning("提示", "线号检查报告导出完成")
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def find_end_index(text_content, marker):
    lines = text_content.splitlines(True)  # 使用True保持换行符
    for i, line in enumerate(lines):
        if marker in line:
            return sum(len(l) for l in lines[:i+1])  # 返回到该行末尾的累计长度
    return len(text_content)  # 如果没有找到marker，返回文本的总长度


def text_to_html(text_widget, start_index, end_index):
    # 获取文本内容
    text_content = text_widget.get("1.0", tk.END)

    # 获取 start_index 之前的文本
    before_text = text_content[:start_index-1]
    # 获取 end_index 之后的文本
    after_text = text_content[end_index:]

    # 合并排除了指定范围的文本
    final_text = before_text + after_text

    # 将换行符转换为HTML的换行标签
    html_output = final_text.replace('\n', '<br />')

    # 结束HTML文档
    html_output = f"<html><body><p>{html_output}</p></body></html>"

    return html_output


def add_header1(rl_canvas, doc):
    rl_canvas.saveState()

    # 添加图片（请确保图片路径正确）
    image_path = 'ico/ABB_logo.png'
    img_width = 27.7 * mm
    img_height = 10.5 * mm
    img = Image(image_path, width=img_width, height=img_height)
    img.drawOn(rl_canvas, A4[0] / 2 - img_width / 2, A4[1] - img_height - 30)

    # 添加标题
    rl_canvas.setFont('SimSun', 16)
    title = "FAST 线号检查报告"
    rl_canvas.drawCentredString(A4[0] / 2, A4[1] - img_height - 60, title)

    rl_canvas.setFont('SimSun', 10)
    project_number_text = "项目号：" + project_number
    rl_canvas.drawString(45, A4[1] - img_height - 60 - 10, project_number_text)

    rl_canvas.setFont('SimSun', 8)
    global page_number
    page_number_text = "Page " + str(page_number)
    page_number += 1
    rl_canvas.drawString(A4[0]-80, 5*mm, page_number_text)

    rl_canvas.setFont('SimSun', 8)
    Timestamp = strftime('%Y-%m-%d %H:%M:%S', localtime())
    rl_canvas.drawCentredString(A4[0] / 2, 5 * mm, '制表时间：' + Timestamp)

    rl_canvas.restoreState()


def add_header2(rl_canvas, doc):
    rl_canvas.saveState()

    # 添加图片（请确保图片路径正确）
    image_path = 'ico/ABB_logo.png'
    img_width = 27.7 * mm
    img_height = 10.5 * mm
    img = Image(image_path, width=img_width, height=img_height)
    img.drawOn(rl_canvas, A4[0] / 2 - img_width / 2, A4[1] - img_height - 30)

    # 添加标题
    rl_canvas.setFont('SimSun', 16)
    title = "FAST 线号检查报告"
    rl_canvas.drawCentredString(A4[0] / 2, A4[1] - img_height - 60, title)

    rl_canvas.setFont('SimSun', 8)
    global page_number
    page_number_text = "Page " + str(page_number)
    page_number += 1
    rl_canvas.drawString(A4[0]-80, 5*mm, page_number_text)

    rl_canvas.setFont('SimSun', 8)
    Timestamp = strftime('%Y-%m-%d %H:%M:%S', localtime())
    rl_canvas.drawCentredString(A4[0] / 2, 5 * mm, '制表时间：' + Timestamp)

    rl_canvas.restoreState()