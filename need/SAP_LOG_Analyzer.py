#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EBOM导SAP日志解析工具
功能：解析指定项目号的.log日志文件，提取BOM导入信息并生成Excel报告
"""

import os
import re
import glob
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass
from typing import List, Optional, Dict, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import threading


@dataclass
class LogEntry:
    """日志条目数据类"""
    filename: str
    bom_time: str  # 导BOM时间 (Creating Response file for 的时间)
    sales_order: str
    high_level_item: str
    sales_order_line: str
    quantity: str
    panel_numbers: str
    description: str
    status: str  # "成功" 或 "失败"


class SAPLogParser:
    """EBOM导SAP日志解析器"""

    def __init__(self, base_path: str, log_callback=None):
        self.base_path = Path(base_path)
        self.entries: List[LogEntry] = []
        self.log_callback = log_callback
    
    def log(self, message):
        """日志输出"""
        if self.log_callback:
            self.log_callback(message)
        # print(message)

    def find_log_files(self, project_number: str) -> List[Path]:
        """
        查找指定项目号的日志文件
        只在目标文件夹下查找文件名开头为项目号的log文件
        """
        # 只在目标文件夹下查找文件名开头为项目号的log文件
        log_files = list(self.base_path.glob(f"{project_number}*.log"))
        
        # 排序
        return sorted(log_files)

    def parse_timestamp(self, line: str) -> Optional[str]:
        """从日志行中提取时间戳"""
        # 匹配格式: 2/3/2026 8:12:44 PM
        pattern = r'(\d{1,2}/\d{1,2}/\d{4}\s+\d{1,2}:\d{2}:\d{2}\s+(?:AM|PM))'
        match = re.search(pattern, line)
        if match:
            return match.group(1)
        return None
    
    def parse_datetime(self, timestamp: str) -> Optional[datetime]:
        """将时间戳字符串解析为datetime对象"""
        if not timestamp:
            return None
        try:
            # 格式: 2/3/2026 8:12:44 PM
            return datetime.strptime(timestamp, '%m/%d/%Y %I:%M:%S %p')
        except ValueError:
            return None

    def extract_value(self, text: str, key: str) -> str:
        """从文本中提取 key=value 格式的值"""
        pattern = rf'{re.escape(key)}=(.*?)(?:\r?\n|$)'
        match = re.search(pattern, text)
        if match:
            return match.group(1).strip()
        return ""

    def extract_process_id(self, line: str) -> Optional[str]:
        """从日志行中提取进程号，格式如 '14-->' 或 '5-->'"""
        pattern = r'^(\d+)-->'
        match = re.search(pattern, line)
        if match:
            return match.group(1)
        return None

    def parse_log_file(self, file_path: Path) -> List[LogEntry]:
        """解析单个日志文件，返回多个条目（一个文件可能包含多次操作）
        根据进程号区分不同的操作记录，相同进程号但时间间隔超过2小时视为不同进程"""
        entries = []

        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
        except Exception as e:
            self.log(f">>>读取文件失败 {file_path}: {e}")
            return entries

        lines = content.split('\n')

        # 找到所有 "Creating Response file for" 的位置及其进程号和时间
        start_positions = []  # [(line_index, process_id, timestamp), ...]
        for i, line in enumerate(lines):
            if "Creating Response file for" in line:
                process_id = self.extract_process_id(line)
                timestamp = self.parse_timestamp(line)
                if process_id:
                    start_positions.append((i, process_id, timestamp))

        # 如果没有找到起始点，返回空
        if not start_positions:
            self.log(f">>>未找到有效的BOM记录")
            return entries

        self.log(f"发现 {len(start_positions)} 个BOM操作记录")

        # 处理进程号，对于相同进程号但时间间隔超过2小时的情况，视为不同进程
        # 为每个进程号维护一个计数器和最后时间
        process_info = {}
        processed_positions = []  # [(line_index, unique_process_id, timestamp), ...]
        
        for line_index, process_id, timestamp in start_positions:
            if process_id not in process_info:
                # 新进程号，初始化计数器和时间
                process_info[process_id] = {
                    'counter': 1,
                    'last_time': self.parse_datetime(timestamp)
                }
                unique_process_id = process_id
            else:
                # 相同进程号，检查时间间隔
                current_time = self.parse_datetime(timestamp)
                last_time = process_info[process_id]['last_time']
                if current_time and last_time:
                    time_diff = current_time - last_time
                    if time_diff.total_seconds() > 2 * 3600:  # 超过2小时
                        # 视为不同进程，增加计数器
                        process_info[process_id]['counter'] += 1
                        process_info[process_id]['last_time'] = current_time
                        unique_process_id = f"{process_id}_{process_info[process_id]['counter']}"
                    else:
                        # 同一进程，更新最后时间
                        process_info[process_id]['last_time'] = current_time
                        unique_process_id = process_id
                else:
                    # 时间解析失败，视为同一进程
                    unique_process_id = process_id
            
            processed_positions.append((line_index, unique_process_id, timestamp))

        # 处理每个块
        for idx, (start_idx, unique_process_id, timestamp) in enumerate(processed_positions):
            self.log(f"正在处理第 {idx + 1}/{len(processed_positions)} 条记录 (进程号: {unique_process_id})...")
            
            # 确定块的结束位置：下一个属于相同原始进程号的 "Creating Response file for" 或文件结束
            original_process_id = unique_process_id.split('_')[0]
            end_idx = len(lines)
            
            # 查找下一个属于相同原始进程号的 "Creating Response file for" 记录
            for next_pos in processed_positions[idx + 1:]:
                next_line_idx, next_unique_process_id, _ = next_pos
                next_original_process_id = next_unique_process_id.split('_')[0]
                if next_original_process_id == original_process_id:
                    end_idx = next_line_idx
                    break

            # 提取当前块的内容（只包含相同原始进程号的行）
            block_lines = []
            for line in lines[start_idx:end_idx]:
                line_process_id = self.extract_process_id(line)
                # 只保留相同原始进程号的行或没有进程号的行
                if line_process_id == original_process_id or not line_process_id:
                    block_lines.append(line)
            
            block_text = '\n'.join(block_lines)

            # 获取BOM时间（Creating Response file for 所在行的时间）
            bom_time = timestamp

            # 检查是否包含 "Getting Typicals from file" 和 "Getting Sub Parts Information..."
            if "Getting Typicals from file" not in block_text:
                self.log(f"    跳过：未找到 'Getting Typicals from file'")
                continue

            if "Getting Sub Parts Information..." not in block_text:
                self.log(f"    跳过：未找到 'Getting Sub Parts Information...'")
                continue

            # 提取关键字段 - 在 Getting Typicals from file 和 Getting Sub Parts Information... 之间
            # 找到这两个标记的位置，提取中间的内容
            typicals_match = re.search(
                r'Getting Typicals from file.*?(.*?)Getting Sub Parts Information\.\.\.',
                block_text,
                re.DOTALL
            )

            middle_section = typicals_match.group(1) if typicals_match else block_text
            # print(middle_section)
            # 提取各个字段
            sales_order = self.extract_value(middle_section, "Sales Order")
            high_level_item = self.extract_value(middle_section, "High Level Item")
            sales_order_line = self.extract_value(middle_section, "Sales Order Line")
            quantity = self.extract_value(middle_section, "Quantity")
            panel_numbers = self.extract_value(middle_section, "Panel Numbers")
            description = self.extract_value(middle_section, "Description")

            # 判断状态：检查块中是否包含 "BOM Created(Changed) successfully."
            status = "成功" if "BOM Created(Changed) successfully." in block_text else "失败"

            entry = LogEntry(
                filename=file_path.name,
                bom_time=bom_time,
                sales_order=sales_order,
                high_level_item=high_level_item,
                sales_order_line=sales_order_line,
                quantity=quantity,
                panel_numbers=panel_numbers,
                description=description,
                status=status
            )
            entries.append(entry)
            self.log(f"    成功提取记录: Sales Order={sales_order}, Status={status}")

        return entries

    def parse_all(self, project_number: str) -> List[LogEntry]:
        """解析所有匹配的日志文件"""
        find_start_time = datetime.now()
        log_files = self.find_log_files(project_number)
        find_end_time = datetime.now()
        find_time = (find_end_time - find_start_time).total_seconds()
        
        if not log_files:
            self.log(f">>>未找到项目号 '{project_number}' 相关的日志文件")
            return []

        self.log(f">>>找到 {len(log_files)} 个日志文件")
        for f in log_files:
            self.log(f"  - {f}")
        self.log(f"查找耗时: {find_time:.2f} 秒")
        
        self.log("\n>>>开始解析日志文件...")

        all_entries = []
        for log_file in log_files:
            self.log(f"\n>>>正在解析: {log_file.name}")
            start_time = datetime.now()
            entries = self.parse_log_file(log_file)
            end_time = datetime.now()
            processing_time = (end_time - start_time).total_seconds()
            all_entries.extend(entries)
            self.log(f"提取到 {len(entries)} 条记录")
            self.log(f"文件路径: {log_file}")
            self.log(f"处理时间: {processing_time:.2f} 秒")

        self.entries = all_entries
        return all_entries

    def export_to_excel(self, output_path: str, project_number: str):
        """导出结果到Excel"""
        if not self.entries:
            self.log(">>>没有数据可导出")
            return

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"项目{project_number}日志分析"

        # 设置表头
        headers = ["文件名", "导BOM时间", "Sales Order", "High Level Item",
                   "Sales Order Line", "Quantity", "Panel Numbers", "Description", "状态"]

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 写入数据
        for row, entry in enumerate(self.entries, 2):
            ws.cell(row=row, column=1, value=entry.filename)
            ws.cell(row=row, column=2, value=entry.bom_time)
            ws.cell(row=row, column=3, value=entry.sales_order)
            ws.cell(row=row, column=4, value=entry.high_level_item)
            ws.cell(row=row, column=5, value=entry.sales_order_line)
            ws.cell(row=row, column=6, value=entry.quantity)
            ws.cell(row=row, column=7, value=entry.panel_numbers)
            ws.cell(row=row, column=8, value=entry.description)
            ws.cell(row=row, column=9, value=entry.status)

            # 根据状态设置颜色
            status_cell = ws.cell(row=row, column=9)
            if entry.status == "成功":
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(color="006100")
            else:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = Font(color="9C0006")

        # 调整列宽
        column_widths = [30, 20, 15, 15, 18, 10, 20, 15, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # 添加边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=len(self.entries) + 1, min_col=1, max_col=9):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 保存文件
        wb.save(output_path)
        self.log(f"\nExcel文件已保存: {output_path}")
        self.log(f"共导出 {len(self.entries)} 条记录")


class SAPLogParserGUI:
    """EBOM导SAP日志解析工具GUI界面"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("EBOM导SAP日志解析工具")
        
        # 设置窗口图标
        try:
            self.root.iconbitmap("app.ico")
        except Exception:
            pass  # 如果图标文件不存在，忽略错误
        
        # 设置窗口居中
        window_width = 800
        window_height = 600
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        center_x = int((screen_width - window_width) / 2)
        center_y = int((screen_height - window_height) / 2)
        self.root.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")
        
        self.root.resizable(True, True)
        
        # 配置
        self.LOG_BASE_PATH = r"\\CN-S-041MVE1\temp\exportToSAP\logs"
        
        # 字体配置
        self.font_family = "ABBvoice CNSG"
        self.font_size = 10
        
        # 创建主框架
        main_frame = ttk.Frame(root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 项目号输入区域
        input_frame = tk.LabelFrame(main_frame, text="项目号输入", font=(self.font_family, self.font_size), padx=10, pady=10)
        input_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(input_frame, text="项目号:", font=(self.font_family, self.font_size)).pack(side=tk.LEFT, padx=5)
        self.project_number_var = tk.StringVar()
        self.project_entry = tk.Entry(input_frame, textvariable=self.project_number_var, width=30, font=(self.font_family, self.font_size))
        self.project_entry.pack(side=tk.LEFT, padx=5)
        
        self.parse_button = tk.Button(input_frame, text="解析日志", command=self.start_parse, font=(self.font_family, self.font_size))
        self.parse_button.pack(side=tk.LEFT, padx=5)
        
        # 进度输出区域
        output_frame = tk.LabelFrame(main_frame, text="处理进度", font=(self.font_family, self.font_size), padx=10, pady=10)
        output_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.output_text = scrolledtext.ScrolledText(output_frame, wrap=tk.WORD, font=(self.font_family, self.font_size))
        self.output_text.pack(fill=tk.BOTH, expand=True)
        self.output_text.config(state=tk.DISABLED)
        
        # 状态信息
        self.status_var = tk.StringVar(value="就绪")
        status_bar = tk.Label(root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W, font=(self.font_family, self.font_size))
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
    
    def log(self, message):
        """在输出框中显示日志"""
        def update_gui():
            self.output_text.config(state=tk.NORMAL)
            self.output_text.insert(tk.END, message + "\n")
            self.output_text.see(tk.END)
            self.output_text.config(state=tk.DISABLED)
            # 强制刷新界面以实时显示日志
            self.root.update_idletasks()

        # 使用after确保在主线程中更新GUI
        self.root.after(0, update_gui)
    
    def start_parse(self):
        """开始解析日志"""
        project_number = self.project_number_var.get().strip()
        
        # 验证项目号
        if not project_number:
            messagebox.showwarning("输入错误", "项目号不能为空")
            return
        
        # 验证项目号格式：8或9位纯数字，8位只能是7开头
        if not re.match(r'^[0-9]+$', project_number):
            messagebox.showwarning("输入错误", "项目号必须为纯数字")
            return
        
        if len(project_number) not in [8, 9]:
            messagebox.showwarning("输入错误", "项目号必须为8位或9位")
            return
        
        if len(project_number) == 8 and not project_number.startswith('7'):
            messagebox.showwarning("输入错误", "8位项目号必须以7开头")
            return
        
        # 清空输出
        self.output_text.config(state=tk.NORMAL)
        self.output_text.delete(1.0, tk.END)
        self.output_text.config(state=tk.DISABLED)
        
        # 禁用按钮
        self.parse_button.config(state=tk.DISABLED)
        self.status_var.set("解析中...")
        
        # 使用after方法实现非阻塞解析
        def parse_task():
            try:
                # 创建解析器
                parser = SAPLogParser(self.LOG_BASE_PATH, log_callback=self.log)
                
                # 解析日志
                entries = parser.parse_all(project_number)
                
                if not entries:
                    self.log("未找到有效数据")
                    return

                # 获取桌面路径 ✨ 新增
                desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

                # 生成输出文件名 ✨ 修改
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file = os.path.join(desktop_path, f"EBOM导SAP日志分析_{project_number}_{timestamp}.xlsx")
                
                # # 生成输出文件名
                # timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                # output_file = f"EBOM导SAP日志分析_{project_number}_{timestamp}.xlsx"
                
                # 导出到Excel
                parser.export_to_excel(output_file, project_number)
                
                # 打印统计信息
                success_count = sum(1 for e in entries if e.status == "成功")
                fail_count = len(entries) - success_count
                
                self.log(f"\n统计信息:")
                self.log(f"  总记录数: {len(entries)}")
                self.log(f"  成功: {success_count}")
                self.log(f"  失败: {fail_count}")
                
                self.status_var.set("解析完成")
            except Exception as e:
                self.log(f"解析过程出错: {e}")
                self.status_var.set("解析出错")
            finally:
                # 启用按钮
                self.parse_button.config(state=tk.NORMAL)
        
        # 使用after方法在事件循环中执行解析任务
        self.root.after(100, parse_task)


# def main():
#     """主函数"""
#     # 配置
#     LOG_BASE_PATH = r"\\CN-S-041MVE1\temp\exportToSAP\logs"

#     # 获取用户输入的项目号
#     project_number = input("请输入项目号 (例如: 504652829): ").strip()

#     if not project_number:
#         print("项目号不能为空")
#         return

#     # 创建解析器
#     parser = SAPLogParser(LOG_BASE_PATH)

#     # 解析日志
#     entries = parser.parse_all(project_number)

#     if not entries:
#         print("未找到有效数据")
#         return

#     # 生成输出文件名
#     timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#     output_file = f"EBOM导SAP日志分析_{project_number}_{timestamp}.xlsx"

#     # 导出到Excel
#     parser.export_to_excel(output_file, project_number)

#     # 打印统计信息
#     success_count = sum(1 for e in entries if e.status == "成功")
#     fail_count = len(entries) - success_count

#     print(f"\n统计信息:")
#     print(f"  总记录数: {len(entries)}")
#     print(f"  成功: {success_count}")
#     print(f"  失败: {fail_count}")


def gui_main(parent=None):
    """GUI主函数"""
    # 如果提供了父窗口，则创建Toplevel子窗口，否则创建独立的Tk根窗口
    if parent:
        root = tk.Toplevel(parent)
        # 禁用父窗口
        parent.withdraw()
        # 关闭子窗口时恢复父窗口并退出
        def on_child_close():
            parent.deiconify()
            root.destroy()
        root.protocol('WM_DELETE_WINDOW', on_child_close)
        # 设置窗口始终在最前端
        root.attributes("-topmost", True)
    else:
        root = tk.Tk()
    app = SAPLogParserGUI(root)
    # 如果是子窗口，不启动新的mainloop
    if not parent:
        root.mainloop()


if __name__ == "__main__":
    # 启动GUI界面
    gui_main()