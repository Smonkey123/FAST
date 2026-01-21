import tkinter as tk
from tkinter import ttk
from tkinter import StringVar, Menu
from tkinter.ttk import Treeview, Notebook, Scrollbar, Style
from tkinter.filedialog import askdirectory
import os

import time
from time import *

import xlrd
import xlwt
from xlutils.copy import copy

from lxml import etree
import traceback
import warnings
import logging
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Protection
from PIL import Image, ImageTk
from need.custom_dialogs import CustomDialog, center_window, Tooltip, image_label
warnings.simplefilter(action='ignore', category=FutureWarning)

FilePath = ""  # 设置一个地址变量


def main(parent, root, w_rat, h_rat):
    global open_folder
    open_folder = tk.PhotoImage(file="ico\\open_folder.png")
    global analyze_file
    analyze_file = tk.PhotoImage(file="ico\\read.png")
    global change_switchgear_id
    change_switchgear_id = tk.PhotoImage(file="ico\\mve_switchgear_id.png")
    global change_panel_number
    change_panel_number = tk.PhotoImage(file="ico\\mve_no.png")
    global change_project_sn
    change_project_sn = tk.PhotoImage(file="ico\\mve_sn.png")
    global nav_mve
    nav_mve = tk.PhotoImage(file="ico\\mve_nav.png")
    global project_mve
    project_mve = tk.PhotoImage(file="ico\\mve_project.png")
    global switchgear_mve
    switchgear_mve = tk.PhotoImage(file="ico\\mve_switchgear.png")
    global typical_mve
    typical_mve = tk.PhotoImage(file="ico\\mve_typical.png")
    global panel_mve
    panel_mve = tk.PhotoImage(file="ico\\mve_panel.png")
    global export_file
    export_file = tk.PhotoImage(file="ico\\export.png")
    global attribute_hide
    attribute_hide = tk.PhotoImage(file="ico\\attribute_hide.png")
    global attribute_view
    attribute_view = tk.PhotoImage(file="ico\\attribute_view.png")
    global level_fold
    level_fold = tk.PhotoImage(file="ico\\fold.png")
    global level_unfold
    level_unfold = tk.PhotoImage(file="ico\\unfold.png")
    global edit_panel_in_excel
    edit_panel_in_excel = tk.PhotoImage(file="ico\\edit_panel_in_excel.png")

    # global canvas
    # canvas = tk.Canvas(parent, width=int(1600 * w_ratio), height=int(640 * h_ratio), bg="#C9DBE9")
    # canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    # canvas.update()
    # canvas.bind("<MouseWheel>", on_mousewheel)

    # scrollbar_v = Scrollbar(master=parent)
    # scrollbar_v.pack(side=tk.RIGHT, fill=tk.Y)
    # scrollbar_v.config(command=canvas.yview)
    # canvas.config(yscrollcommand=scrollbar_v.set)

    # content = tk.Frame(canvas)
    # canvas.create_window(0, 1, width=int(1600 * w_ratio), anchor=tk.NW, window=content)

    global root_win
    root_win = root

    global w_ratio
    w_ratio = w_rat
    global h_ratio
    h_ratio = h_rat

    global deletebusbarbridge_flag
    deletebusbarbridge_flag = 0

    tk.Label(parent, text="欢迎使用MVE预配置功能", bg="#c9dbe9", fg="black", height=int(1 * h_ratio), font=("ABBvoice CNSG", int(20 * h_ratio), "bold")).pack(fill=tk.X)

    f1 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f1, text='   说明：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')
    tk.Label(f1, text='(1)选择.mve文件，点击"读取"，获取属性信息，可对节点任意属性进行双击修改，修改完成后导出即可；\n(2)红色属性取自.mve文件，修改后会修改对应.mve文件属性，黑色属性取自.mac文件，修改无效。', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    f1.pack(fill=tk.BOTH, expand=True)

    tk.Frame(parent, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f2 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f2, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2, text='路径：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry  # 为了确保selectpath函数能正确调用entry,将其全局化
    entry = tk.Entry(f2, bg="white", font=("ABBvoice CNSG", int(13 * h_ratio)))
    entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

    global button_select
    button_select = tk.Button(f2, text="选择", bg="#eaf1f6", image=open_folder, font=("ABBvoice CNSG", int(12 * h_ratio)), compound=tk.LEFT, command=selectpath, activebackground='blue')
    button_select.pack(side=tk.LEFT, padx=int(20 * w_ratio))
    # tk.Button(f2, text="开始读取", width=12, command=process).pack(side=tk.LEFT, padx=0)
    Tooltip(button_select, "选择mve文件")

    global button_read
    button_read = tk.Button(f2, text="读取", bg="#eaf1f6", image=analyze_file, font=("ABBvoice CNSG", int(12 * h_ratio)), compound=tk.LEFT, command=read_xml_file, activebackground='blue')
    button_read.pack(side=tk.LEFT, padx=(0, int(20 * w_ratio)))
    button_read['state'] = 'disabled'
    Tooltip(button_read, "读取mve文件")

    global button_export
    button_export = tk.Button(f2, image=export_file, font=("ABBvoice CNSG", int(12 * h_ratio)), text="导出", bg="#eaf1f6", command=export_mve, compound=tk.LEFT, state='disabled', activebackground='blue')
    button_export.pack(side=tk.LEFT, padx=(0, int(20 * w_ratio)))
    Tooltip(button_export, "导出修改后的mve文件，手动导入MVE")

    f2.pack(fill=tk.X)

    tk.Frame(parent, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    fun = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(fun, text='   功能：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, anchor='n')

    global button_level_fold
    button_level_fold = tk.Button(fun, image=level_fold, bg="#eaf1f6", command=on_level_fold, state='disabled', activebackground='blue')
    button_level_fold.pack(side=tk.LEFT, padx=(0, 5))
    Tooltip(button_level_fold, "折叠")

    global button_level_unfold
    button_level_unfold = tk.Button(fun, image=level_unfold, bg="#eaf1f6", command=on_level_unfold, state='disabled', activebackground='blue')
    button_level_unfold.pack(side=tk.LEFT, padx=(0, 5))
    Tooltip(button_level_unfold, "展开")

    global button_change_switchgear_id
    button_change_switchgear_id = tk.Button(fun, image=change_switchgear_id, bg="#eaf1f6", command=multichange_switchgear_id, state='disabled', activebackground='blue')
    button_change_switchgear_id.pack(side=tk.LEFT, padx=(0, 5))
    Tooltip(button_change_switchgear_id, "修改开关段ID")

    global button_change_project_sn
    button_change_project_sn = tk.Button(fun, image=change_project_sn, bg="#eaf1f6", command=multichange_project_sn, state='disabled', activebackground='blue')
    button_change_project_sn.pack(side=tk.LEFT, padx=(0, 5))
    Tooltip(button_change_project_sn, "修改项目系列号")

    global button_change_panel_number
    button_change_panel_number = tk.Button(fun, image=change_panel_number, bg="#eaf1f6", command=multichange_panel_number, state='disabled', activebackground='blue')
    button_change_panel_number.pack(side=tk.LEFT, padx=(0, 5))
    Tooltip(button_change_panel_number, "修改柜号")

    global attribute_hv_flag
    attribute_hv_flag = True

    global button_attribute_hv
    button_attribute_hv = tk.Button(fun, image=attribute_hide, bg="#eaf1f6", command=hv_button_click, state='disabled', activebackground='blue')
    button_attribute_hv.pack(side='right', padx=(0, int(20 * w_ratio)))
    Tooltip(button_attribute_hv, "隐藏/显示非mve属性")

    fun.pack(fill=tk.X)

    tk.Frame(parent, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    global f3
    f3 = tk.Frame(parent, bg="#eaf1f6", bd=0)

    tk.Label(f3, text='   属性：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, anchor='n')

    f31 = tk.Frame(f3, bg="#eaf1f6", bd=0)
    f31.pack(side=tk.LEFT, anchor='n')
    f32 = tk.Frame(f3, bg="#eaf1f6", bd=0)
    f32.pack(side=tk.LEFT, fill='both', expand=True, anchor='n')

    style = Style()
    style.configure('TNotebook', background="#eaf1f6", font=("ABBvoice CNSG", int(10 * h_ratio)))
    style.configure('mve.Treeview', rowheight=30, font=("ABBvoice CNSG", int(13 * h_ratio)))
    style.configure('mve.Treeview.Heading', font=("ABBvoice CNSG", int(13 * h_ratio)), background="#EFF1F5")

    mve_notebook_left = Notebook(f31, width=250)
    pane_left = tk.Frame()
    global mve_level_tree
    mve_level_tree = Treeview(pane_left, show='tree', height=100, style='mve.Treeview', selectmode='browse')
    mve_level_tree.pack(side="top", fill="both", expand=True)
    mve_level_tree.tag_configure('left_no_mention_line', foreground='black', font=("ABBvoice CNSG", int(10 * h_ratio)))
    mve_level_tree.tag_configure('left_mention_line', foreground='black', font=("ABBvoice CNSG", int(10 * h_ratio)))
    mve_level_tree.bind('<<TreeviewSelect>>', on_tree_select)
    mve_level_tree.bind('<ButtonPress-1>', on_tree_press)
    mve_level_tree.bind('<ButtonRelease-1>', on_tree_release)
    mve_level_tree.bind('<B1-Motion>', on_tree_move)
    Tooltip(mve_level_tree, "导航栏(支持右键功能)")

    tree_ybar1 = Scrollbar(pane_left, orient='vertical', command=mve_level_tree.yview)
    tree_ybar1.pack_forget()
    mve_notebook_left.add(pane_left, text='Navigator', image=nav_mve, compound=tk.LEFT)
    mve_notebook_left.pack(side='left', fill='both', expand=True, pady=10)

    # 创建右键菜单
    global popup_menu
    popup_menu = Menu(root, tearoff=0)

    mve_level_tree.bind("<Button-3>", on_right_click)

    global mve_notebook_right
    mve_notebook_right = Notebook(f32)
    global pane_right
    pane_right = tk.Frame()
    pane_right_frame = tk.Frame(pane_right)
    pane_right_frame.pack(side="left", fill="both", expand=True, anchor='w')
    global pane_right_tree
    pane_right_tree = Treeview(pane_right_frame, show="tree", style='mve.Treeview', selectmode='browse', height=100)
    pane_right_tree.pack(side="top", fill="both", expand=True, anchor='w')
    pane_right_tree.tag_configure('right_line', foreground='red', font=("ABBvoice CNSG", int(12 * h_ratio)))
    pane_right_tree.tag_configure('right_mention_line', foreground='red', font=("ABBvoice CNSG", int(12 * h_ratio)))
    pane_right_tree.tag_configure('right_no_mention_line', foreground='black', font=("ABBvoice CNSG", int(12 * h_ratio)))
    Tooltip(pane_right_tree, "属性栏(红色属性值支持双击修改，回车保存)")
    # tree_xbar = Scrollbar(pane_right_frame, orient=tk.HORIZONTAL, command=pane_right_tree.xview)
    # pane_right_tree.configure(xscrollcommand=tree_xbar.set)
    # tree_xbar.pack(side=tk.BOTTOM, fill=tk.X)

    mve_notebook_right.add(pane_right, text='Project', image=project_mve, compound=tk.LEFT)

    mve_notebook_right.pack(side='left', fill='both', expand=True, padx=(0, int(20 * w_ratio)), pady=10)

    f3.pack(fill=tk.X)

    tk.Frame(parent, height=int(2000 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    # tk.Frame(content, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线
    #
    # f4 = tk.Frame(content, bg="#eaf1f6", bd=0)
    # tk.Label(f4, text='   输出：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, anchor='n')
    #
    # global text
    # text = tk.Text(f4, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), height=int(20 * h_ratio))
    # text.pack(side=tk.LEFT, padx=(0, 1), pady=0, fill=tk.BOTH, expand=True)
    #
    # # scrollbar = Scrollbar(f4)
    # # scrollbar.pack(side=tk.LEFT, fill=tk.Y)
    # # scrollbar.config(command=text.yview)
    # # text.config(yscrollcommand=scrollbar.set)
    #
    # f4.pack(fill=tk.BOTH, expand=True)
    #
    # tk.Frame(content, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    # canvas.update_idletasks()
    # canvas.config(scrollregion=canvas.bbox('all'))

# def on_mousewheel(event):
#     global canvas
#     canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")


def selectpath():
    filepath = tk.filedialog.askopenfilename(title=u'请选择.mve文件', filetypes=[("Excel", ".mve")])  # 选择打开什么文件，返回文件名
    if len(filepath) != 0:
        string_filename = ""
        for i in range(0, len(filepath)):
            string_filename += str(filepath[i])
        button_read['state'] = 'normal'
    else:
        button_read['state'] = 'disabled'

    entry.delete(0, "end")  # 删除entry原始内容
    entry.insert(0, filepath)  # 重新填入地址
    button_export['state'] = 'disabled'
    button_attribute_hv['state'] = 'disabled'
    button_level_fold['state'] = 'disabled'
    button_level_unfold['state'] = 'disabled'
    button_change_switchgear_id['state'] = 'disabled'
    button_change_project_sn['state'] = 'disabled'
    button_change_panel_number['state'] = 'disabled'

    # 清空原有数据
    for tree in [mve_level_tree, pane_right_tree]:
        for item in tree.get_children():
            tree.delete(item)

    # text.delete(1.0, "end")  # 清空输出结果框
    global FilePath

    FilePath = filepath


def read_xml_file():
    button_export['state'] = 'disabled'
    # 解析.mve文件，存放到左树
    global parser
    parser = etree.XMLParser(remove_comments=True)
    tree = etree.parse(FilePath, parser)  # 把xml文件解析为Element tree,调用此函数返回Element tree对象
    global xml_root
    xml_root = tree.getroot()  # 获取根节点（ProjectXml）,其标记root.tag为‘ProjectXml', 属性字典root.attrib为{}

    # 清空原有数据
    for tree in [mve_level_tree, pane_right_tree]:
        for item in tree.get_children():
            tree.delete(item)

    parse_xml_file()
    select_mac_file_path()
    button_level_fold['state'] = 'normal'
    button_level_unfold['state'] = 'normal'
    button_change_switchgear_id['state'] = 'normal'
    button_change_project_sn['state'] = 'normal'
    button_change_panel_number['state'] = 'normal'


def hv_button_click():
    global attribute_hv_flag
    if attribute_hv_flag:
        button_attribute_hv.config(image=attribute_view)
    else:
        button_attribute_hv.config(image=attribute_hide)
    attribute_hv_flag = not attribute_hv_flag

    # 清空原有数据
    for tree in [pane_right_tree]:
        for item in tree.get_children():
            tree.delete(item)

    update_tab(display_data, order_index)


def parse_xml_file():
    button_attribute_hv['state'] = 'disabled'
    global productname
    productname = xml_root.find('productname').text

    project = xml_root.find('project')
    project_attributelist = project.find('attributelist')

    # 确保'no'节点存在
    ensure_attribute_with_value(project_attributelist, 'no')

    global project_name, switchgear_data

    project_name = []
    parse_attributes(project_attributelist, 'objectname', project_name)

    switchgear_data = []
    switchgear_list = project.find('switchgearlist')

    for switchgear in switchgear_list.findall('switchgear'):
        switchgear_attributelist = switchgear.find('attributelist')

        # 确保'SerialNo'节点存在
        ensure_attribute_with_value(switchgear_attributelist, 'SerialNo')

        switchgear_number = []
        parse_attributes(switchgear_attributelist, 'objectname', switchgear_number)

        typicallist = switchgear.find('typicallist')
        typical_numbers = []
        for typical in typicallist.findall('typical'):
            typical_attributelist = typical.find('attributelist')
            parse_attributes(typical_attributelist, 'objectname', typical_numbers)

        panellist = switchgear.find('panellist')
        panel_numbers = []
        typical_panels = []
        for panel in panellist.findall('panel'):
            panel_attributelist = panel.find('attributelist')
            parse_attributes(panel_attributelist, 'objectname', panel_numbers)
            parse_attributes(panel_attributelist, 'typicalname', typical_panels)

        switchgear_data.append({'switchgear_number': switchgear_number, 'typical_number': typical_numbers, 'panel_number': panel_numbers, 'typical_panel': typical_panels})

    # print(productname, project_name, switchgear_data)


def ensure_attribute_with_value(parent, attribute_name):
    attribute = parent.xpath(f"attribute[name='{attribute_name}']")
    if not attribute:
        attribute = etree.SubElement(parent, 'attribute')
        name = etree.SubElement(attribute, 'name')
        name.text = attribute_name
        value = etree.SubElement(attribute, 'value')
        value.text = ""  # 默认值为空字符串
    else:
        attribute = attribute[0]
        if attribute.find('value') is None:
            value = etree.SubElement(attribute, 'value')
            value.text = ""  # 默认值为空字符串


def multichange_switchgear_id():
    global xml_root
    switchgear_list = xml_root.xpath('//project/switchgearlist/switchgear')
    for i, switchgear in enumerate(switchgear_list, start=1):
        object_name = switchgear.xpath('attributelist/attribute[name="ObjectName"]/value')
        if object_name:
            object_name[0].text = f"A{i:02d}"
        else:
            attributelist = switchgear.find('attributelist')
            attribute = etree.SubElement(attributelist, 'attribute')
            name = etree.SubElement(attribute, 'name')
            name.text = "ObjectName"
            value = etree.SubElement(attribute, 'value')
            value.text = f"A{i:02d}"

    for tree in [mve_level_tree, pane_right_tree]:
        for item in tree.get_children():
            tree.delete(item)
    parse_xml_file()
    read_mac_file()
    insert_to_mve_level_tree()
    tk.messagebox.showwarning("提示", "开关段ID一键修改完成")
    button_export['state'] = 'normal'


def confirm_projectserialnumber():
    serial_number = entry_projectserialnumber.get()
    if len(serial_number) != 5 or not serial_number.isdigit():
        tk.messagebox.showwarning("错误", "请输入5位数字的系列号")
        top2.lift()  # 使顶层窗口继续处于顶层(超越其它窗口)
    else:
        allow_main_window2()  # 输入合理的项目系列号，将top2销毁，并允许操作主页面
        switchgear_list = xml_root.xpath('//project/switchgearlist/switchgear')
        for switchgear in switchgear_list:
            serial_no = switchgear.xpath('attributelist/attribute[name="SerialNo"]/value')
            if serial_no:
                serial_no[0].text = serial_number

        for tree in [mve_level_tree, pane_right_tree]:
            for item in tree.get_children():
                tree.delete(item)
        parse_xml_file()
        read_mac_file()
        insert_to_mve_level_tree()
        tk.messagebox.showwarning("提示", "项目系列号一键修改完成")
        button_export['state'] = 'normal'


def multichange_project_sn():
    root_win.attributes("-disabled", 1)
    global top2
    top2 = tk.Toplevel()
    winw = 500
    winh = 500
    top2.geometry("%dx%d" % (winw, winh))
    top2.title('一键改项目系列号')
    center_window(top2)
    top2.protocol("WM_DELETE_WINDOW", allow_main_window2)

    tk.Frame(top2, height=1, bg="#eaf1f6").pack(fill=tk.X)

    f1 = tk.Frame(top2, bg="#eaf1f6", bd=0)
    tk.Label(f1, text='项目系列号(限5位)：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(10 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry_projectserialnumber
    entry_projectserialnumber = tk.Entry(f1, font=("ABBvoice CNSG", int(10 * h_ratio)), width=7)
    entry_projectserialnumber.pack(side=tk.LEFT)

    tk.Button(f1, text="确认", font=("ABBvoice CNSG", int(10 * h_ratio)), command=confirm_projectserialnumber).pack(side=tk.LEFT, padx=5)
    f1.pack(fill=tk.X)

    tk.Frame(top2, height=1, bg="#eaf1f6").pack(fill=tk.X)

    f2 = tk.Frame(top2, bg="#eaf1f6", bd=0)
    tk.Label(f2, text='数据来源(下图所示)：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(10 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')
    tk.Label(f2, text='PMP', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(10 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    f2.pack(fill=tk.X)

    tk.Frame(top2, height=1, bg="#eaf1f6").pack(fill=tk.X)

    f3 = tk.Frame(top2, bg="#eaf1f6", bd=0)
    im = image_label(f3, "ico\\getProjectSerialNumber.png", 500, 500, True)
    im.pack(anchor='nw')
    Tooltip(im, "从PMP查询项目系列号")
    f3.pack(fill=tk.X)


def allow_main_window2():
    top2.destroy()
    root_win.attributes("-disabled", 0)


def populate_treeview():
    row_index = 0
    for switchgear in xml_root.findall('.//switchgear'):
        switchgear_name = switchgear.find(".//attribute[name='ObjectName']/value").text if switchgear.find(".//attribute[name='ObjectName']/value") is not None else ""
        row_index += 1

        for panel in switchgear.findall('.//panel'):
            panel_abb_name = panel.find(".//attribute[name='ObjectName']/value").text if panel.find(".//attribute[name='ObjectName']/value") is not None else ""
            customer_panel_name = panel.find(".//attribute[name='CustomerPanelName']/value").text if panel.find(".//attribute[name='CustomerPanelName']/value") is not None else ""
            panel_name1 = panel.find(".//attribute[name='PanelName1']/value").text if panel.find(".//attribute[name='PanelName1']/value") is not None else ""
            typical_name = panel.find(".//attribute[name='TypicalName']/value").text if panel.find(".//attribute[name='TypicalName']/value") is not None else ""

            item_id = panel_info_tree.insert('', 'end', values=(switchgear_name, panel_abb_name, customer_panel_name, panel_name1, typical_name))

            if row_index % 2 == 0:
                panel_info_tree.item(item_id, tags=('even_row',))


def multichange_panel_number():
    global panel_info_change_flag
    panel_info_change_flag = 0
    root_win.attributes("-disabled", 1)
    global top3
    top3 = tk.Toplevel(bg="#eaf1f6")
    winw = 800
    winh = 800
    top3.geometry("%dx%d" % (winw, winh))
    top3.title('一键改柜号(退出当前界面后才刷新)')
    center_window(top3)
    top3.protocol("WM_DELETE_WINDOW", allow_main_window3)

    tk.Frame(top3, height=1, bg="#eaf1f6").pack(fill=tk.X)
    f1 = tk.Frame(top3, bg="#eaf1f6", bd=0)
    f2 = tk.Frame(top3, bg="#eaf1f6", bd=0)
    f3 = tk.Frame(top3, bg="#eaf1f6", bd=0)

    tk.Label(f1, text='方法1(少量修改)：在下表中双击修改，仅[ABB柜号]、[客户柜号]、[柜体名称]列可修改', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2, text='方法2(大量修改)：在Excel中修改，点击右侧图标，在Excel中修改完成后，点击读取', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    global button_edit_excel, button_read_excel
    button_edit_excel = tk.Button(f2, image=edit_panel_in_excel, bg="#eaf1f6", command=create_and_open_excel, activebackground='blue', relief='flat')
    button_edit_excel.pack(side=tk.LEFT, padx=(20, 5))
    Tooltip(button_edit_excel, "创建并打开Excel")

    button_read_excel = tk.Button(f2, text='读取', font=("ABBvoice CNSG", int(10 * h_ratio)), bg="#eaf1f6", command=refresh_excel, activebackground='blue')
    button_read_excel.pack(side=tk.LEFT, padx=(0, 5))

    global panel_info_tree
    panel_info_tree = Treeview(f3, show="headings", columns=['a', 'b', 'c', 'd', 'e'], style='mve.Treeview', selectmode='browse', height=40)
    panel_info_tree.heading('a', text='开关段ID', anchor='center')
    panel_info_tree.heading('b', text='ABB柜号', anchor='center')
    panel_info_tree.heading('c', text='客户柜号', anchor='center')
    panel_info_tree.heading('d', text='柜体名称', anchor='center')
    panel_info_tree.heading('e', text='所属Typical', anchor='center')
    panel_info_tree.column('a', width=100, anchor='center')
    panel_info_tree.column('b', width=100, anchor='center')
    panel_info_tree.column('c', width=150, anchor='center')
    panel_info_tree.column('d', width=200, anchor='center')
    panel_info_tree.column('e', width=100, anchor='center')

    panel_info_tree.pack(side=tk.TOP, fill=tk.X, expand=True, anchor='w', padx=20)
    panel_info_tree.tag_configure('even_row', background="#c9dbe9")
    Tooltip(panel_info_tree, "中间3列数据双击可修改")
    populate_treeview()

    panel_info_tree.bind('<Double-1>', lambda event, tree=panel_info_tree: edit_panel_cell(event, tree))
    f1.pack(side=tk.TOP, fill=tk.X)
    f2.pack(side=tk.TOP, fill=tk.X, pady=(0, 10))
    f3.pack(side=tk.TOP, fill=tk.X)


# def loading_gif():
#     global top4
#     top4 = tk.Toplevel(bg="#eaf1f6", bd=0)  # 创建顶层窗口
#     center_window(top4)  # 将窗体移动到屏幕中央
#     gif_w = 80
#     gif_h = 80
#     top4.geometry("%dx%d" % (gif_w, gif_h))  # 设置顶层窗口大小
#     top4.overrideredirect(True)
#     gif = Image.open(r"C:\\Users\\CNXIGAO13\\Desktop\\loading2.gif")
#     # Create a canvas to display the GIF
#     canvas = tk.Canvas(top4, width=gif_w, height=gif_h, bg='white', bd=0, highlightthickness=0)
#     canvas.pack()
#
#     # Convert the frames of the GIF to PhotoImage for use on the canvas
#     frames = []
#     for frame in range(0, gif.n_frames):
#         gif.seek(frame)
#         frames.append(ImageTk.PhotoImage(gif))
#
#     # Add the frames to the canvas and animate the GIF
#     def animate_gif(frame=0):
#         canvas.itemconfig(image_item, image=frames[frame])
#         top4.after(50, animate_gif, (frame + 1) % len(frames))
#
#     # Add the first frame of the GIF to the canvas
#     image_item = canvas.create_image(gif_w // 2, gif_h // 2, image=frames[0])
#     top4.lift()
#     # Start the animation loop
#     animate_gif()


def edit_panel_cell(event, tree):
    panel_content = ''
    column_t = tree.identify_column(event.x)
    row_t = tree.identify_row(event.y)

    if column_t:
        col = int(str(column_t).replace('#', ''))
        if col in [2, 3, 4]:
            panel_content = tree.item(row_t, 'value')[col - 1]
            x = tree.bbox(row_t, column=column_t)[0]  # 单元格x坐标
            y = tree.bbox(row_t, column=column_t)[1]  # 单元格y坐标
            width = tree.bbox(row_t, column=column_t)[2]  # 单元格宽度
            height = tree.bbox(row_t, column=column_t)[3]  # 单元格高度

            entryedit = ttk.Entry(tree)
            entryedit.place(x=x, y=y, width=width, height=height)
            entryedit.insert(0, panel_content)

            def on_leave(event):
                if event.state == 8:
                    save_panel_edit(event.widget)
                else:
                    event.widget.unbind('<FocusOut>')

            def on_focus_out(event):
                event.widget.bind('<FocusOut>', lambda e: event.widget.destroy())

            entryedit.focus_set()
            entryedit.bind('<Return>', lambda e: save_panel_edit(entryedit))
            entryedit.bind('<Leave>', on_leave)
            entryedit.bind('<FocusIn>', on_focus_out)

            def save_panel_edit(widget):
                new_panel_content = widget.get()
                # base_value = tree.item(row_t, 'values')[col - 1]
                row_first_col_value = tree.item(row_t, 'values')[0]

                # 找到与 row_first_col_value 相同的行
                similar_rows = [row for row in tree.get_children() if tree.item(row, 'values')[0] == row_first_col_value and row != row_t]
                similar_values = [tree.item(row, 'values')[col - 1] for row in similar_rows]

                suffix = 1
                while new_panel_content in similar_values:
                    new_panel_content = f"{new_panel_content} ({suffix})"
                    suffix += 1

                tree.set(row_t, column_t, new_panel_content)
                if panel_content != new_panel_content:
                    global panel_info_change_flag
                    panel_info_change_flag = 1
                    button_edit_excel['state'] = 'disabled'
                    button_read_excel['state'] = 'disabled'
                    update_panel_xml_value(row_t, col, new_panel_content)

                widget.destroy()

            def update_panel_xml_value(row_t, col, new_panel_content):
                item_id = panel_info_tree.item(row_t, 'value')[0]
                for switchgear in xml_root.findall('.//switchgear'):
                    switchgear_name = switchgear.find(".//attribute[name='ObjectName']/value").text if switchgear.find(".//attribute[name='ObjectName']/value") is not None else ""
                    if switchgear_name == item_id:
                        for panel in switchgear.findall('.//panel'):
                            panel_abb_name = panel.find(".//attribute[name='ObjectName']/value").text if panel.find(".//attribute[name='ObjectName']/value") is not None else ""
                            customer_panel_name = panel.find(".//attribute[name='CustomerPanelName']/value").text if panel.find(".//attribute[name='CustomerPanelName']/value") is not None else ""
                            panel_name1 = panel.find(".//attribute[name='PanelName1']/value").text if panel.find(".//attribute[name='PanelName1']/value") is not None else ""
                            if col == 2 and panel_content == panel_abb_name:
                                panel.find(".//attribute[name='ObjectName']/value").text = new_panel_content
                            elif col == 3 and panel_content == customer_panel_name:
                                panel.find(".//attribute[name='CustomerPanelName']/value").text = new_panel_content
                            elif col == 4 and panel_content == panel_name1:
                                panel.find(".//attribute[name='PanelName1']/value").text = new_panel_content


def allow_main_window3():
    top3.destroy()
    root_win.attributes("-disabled", 0)
    if panel_info_change_flag:
        # loading_gif()
        for tree in [mve_level_tree, pane_right_tree]:
            for item in tree.get_children():
                tree.delete(item)
        parse_xml_file()
        read_mac_file()
        insert_to_mve_level_tree()
        # top4.destroy()
        tk.messagebox.showwarning("提示", "柜号一键修改完成")
        button_export['state'] = 'normal'


def create_and_open_excel():
    home_path = os.path.expanduser("~")
    desktop_path = os.path.join(home_path, "Desktop")
    desktop_path = askdirectory(title=u'请选择导出文件夹', initialdir=desktop_path)
    if not desktop_path:
        tk.messagebox.showwarning("提示", "未选择保存路径，创建Excel操作已取消")
        top3.lift()  # 使顶层窗口继续处于顶层(超越其它窗口)
        return
    global configure_table
    configure_table = os.path.join(desktop_path, '柜号配置表.xlsx')

    # 将数据写入 Excel 文件
    def write_data_to_excel(filepath, tree):
        if os.path.exists(filepath):
            # 如果文件存在，则先清空
            os.remove(filepath)

        # 创建一个新的工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # 设置样式
        font_bold = Font(bold=True)
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        fill_default = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        fill_alternate = PatternFill(start_color='c9dbe9', end_color='c9dbe9', fill_type='solid')

        # 写入表头
        columns = [tree.heading(col)['text'] for col in tree['columns']]
        for col_num, column in enumerate(columns, start=1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 30
            ws.row_dimensions[1].height = 25
            cell = ws.cell(row=1, column=col_num, value=column)
            cell.font = font_bold
            cell.border = border
            cell.alignment = alignment
            cell.fill = fill_yellow

        # 写入数据
        prev_value = None
        style_flag = True

        for row_num, row_id in enumerate(tree.get_children(), start=2):
            row_values = tree.item(row_id, 'values')
            first_col_value = row_values[0]  # 第一列的值
            if first_col_value != prev_value:
                style_flag = not style_flag  # 切换样式
            prev_value = first_col_value
            fill = fill_alternate if style_flag else fill_default
            for col_num, value in enumerate(row_values, start=1):
                ws.row_dimensions[row_num].height = 20
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = Font(bold=False)
                cell.border = border
                cell.alignment = alignment
                cell.fill = fill

        # 解锁所有单元格
        for row in ws.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=False)

        # 锁定第1列和第5列
        def lock_column(col_num):
            for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    cell.protection = Protection(locked=True)

        # 锁定第1列和第5列
        lock_column(1)  # 锁定第1列
        lock_column(5)  # 锁定第5列

        # # 启用工作表保护
        ws.protection.sheet = True

        wb.save(filepath)

    try:
        write_data_to_excel(configure_table, panel_info_tree)
        os.startfile(configure_table)
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())
    top3.lift()  # 使顶层窗口继续处于顶层(超越其它窗口)


def refresh_excel():
    # # 选择要读取的 Excel 文件
    # excel_path = tk.filedialog.askopenfilename(title='请选择要读取的Excel文件', filetypes=[('Excel Files', '*.xlsx')])
    # if not excel_path:
    #     tk.messagebox.showwarning("提示", "未选择Excel文件")
    #     top3.lift()  # 使顶层窗口继续处于顶层(超越其它窗口)
    #     return

    # 读取 Excel 文件内容
    try:
        workbook = load_workbook(configure_table)
        sheet = workbook.active
    except Exception as e:
        tk.messagebox.showwarning("错误", f"无法读取Excel文件: {str(e)}")
        top3.lift()  # 使顶层窗口继续处于顶层(超越其它窗口)
        return

    top3.lift()  # 使顶层窗口继续处于顶层(超越其它窗口)

    # 检测 Excel 内容
    excel_data = []
    switchgear_dict = {}
    switchgear_panel_count = {}
    duplicate_error = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始读取，第一行是标题
        switchgear_name, panel_abb_name, customer_panel_name, panel_name1, typical_name = row
        if switchgear_name not in switchgear_dict:
            switchgear_dict[switchgear_name] = {'panel_abb_name': set(), 'customer_panel_name': set()}
            switchgear_panel_count[switchgear_name] = 0  # 初始化panel计数器
        if panel_abb_name in switchgear_dict[switchgear_name]['panel_abb_name']:
            tk.messagebox.showinfo("提示", f"{switchgear_name}对应的ABB柜号出现重复")
            duplicate_error = 1
            os.startfile(configure_table)
            return
        if customer_panel_name in switchgear_dict[switchgear_name]['customer_panel_name']:
            tk.messagebox.showinfo("提示", f"{switchgear_name}对应的客户柜号出现重复")
            duplicate_error = 1
        if duplicate_error:
            top3.lift()  # 使顶层窗口继续处于顶层(超越其它窗口)
            os.startfile(configure_table)
            return
        switchgear_dict[switchgear_name]['panel_abb_name'].add(panel_abb_name)
        switchgear_dict[switchgear_name]['customer_panel_name'].add(customer_panel_name)
        excel_data.append(row)
        switchgear_panel_count[switchgear_name] += 1  # 更新panel计数

    # 获取 panel_info_tree 的当前数据
    tree_data = []
    for item in panel_info_tree.get_children():
        tree_data.append(panel_info_tree.item(item)['values'])

    # 对比 Excel 数据和 panel_info_tree 数据并更新
    def update_panel_info_tree():
        global different_row_index
        different_row_index = []
        global panel_info_change_flag
        panel_info_change_flag = 0
        for i, excel_row in enumerate(excel_data):
            tree_row = tree_data[i]
            if excel_row != tree_row:
                panel_info_tree.item(panel_info_tree.get_children()[i], values=excel_row)
                panel_info_change_flag = 1
                different_row_index.append(i)
        top3.lift()  # 使顶层窗口继续处于顶层(超越其它窗口)

    def map_index_to_switchgear_panel(index):
        for switchgear, count in switchgear_panel_count.items():
            if index < count:
                return switchgear, index
            index -= count
        return None, None

    # 更新 panel_info_tree 的数据
    update_panel_info_tree()

    if panel_info_change_flag == 1:
        for index in different_row_index:
            switchgear_name, panel_index = map_index_to_switchgear_panel(index)
            if switchgear_name is not None:
                excel_row = excel_data[index]
                panel_abb_name = str(excel_row[1])
                customer_panel_name = str(excel_row[2])
                panel_name1 = str(excel_row[3])

                for switchgear in xml_root.findall('.//switchgear'):
                    switchgear_name1 = switchgear.find(".//attribute[name='ObjectName']/value").text if switchgear.find(".//attribute[name='ObjectName']/value") is not None else ""
                    if switchgear_name1 == switchgear_name:
                        panel_element = switchgear.findall('.//panel')[panel_index]
                        if panel_element:
                            panel_element.find(".//attribute[name='ObjectName']/value").text = panel_abb_name
                            panel_element.find(".//attribute[name='CustomerPanelName']/value").text = customer_panel_name
                            panel_element.find(".//attribute[name='PanelName1']/value").text = panel_name1

    top3.lift()  # 使顶层窗口继续处于顶层(超越其它窗口)


def insert_to_mve_level_tree():
    # 插入项目名父项
    parent_item = mve_level_tree.insert("", "end", text=project_name[0], image=project_mve, open=True, tags='left_mention_line')
    # 插入站名子项
    for switchgear_info in switchgear_data:
        switchgear_parent_item = mve_level_tree.insert(parent_item, "end", text=switchgear_info['switchgear_number'][0], image=switchgear_mve, open=False, tags='left_no_mention_line')

        # 创建一个字典，用于将typical_number与树项目项进行关联
        typical_items = {}
        for t_number in switchgear_info['typical_number']:
            typical_item = mve_level_tree.insert(switchgear_parent_item, "end", text=t_number, image=typical_mve, tags='left_no_mention_line')
            typical_items[t_number] = typical_item

        # 插入Panel子项
        for i, p_number in enumerate(switchgear_info['panel_number']):
            # 获取当前Panel对应的Typical子项
            typical_name = switchgear_info['typical_panel'][i]
            mve_level_tree.insert(switchgear_parent_item, "end", text=p_number + f' <{typical_name}>', image=panel_mve, tags='left_no_mention_line')


def select_mac_file_path():
    global mac_file_path
    mac_file_path = None
    mac_mve_matchlist_file = 'J:/Engineering/ShareFolder/new_ABB_Production_Tools/Pd/document/MAC_MVE_matchlist.xlsx'
    if not os.path.exists(mac_mve_matchlist_file):
        tk.messagebox.showwarning("提示", "失败,找不到MAC_MVE匹配表")
    else:
        global template_directory
        template_directory = 'J:/Engineering/ShareFolder/new_ABB_Production_Tools/Pd/template/'
        workbook = load_workbook(mac_mve_matchlist_file)
        worksheet = workbook['Sheet1']

        for row_i in range(2, worksheet.max_row + 1):
            if str(worksheet.cell(row=row_i, column=3).value) == productname:
                mac_file_path = os.path.join(template_directory, str(worksheet.cell(row=row_i, column=2).value) + '.mac')
                break

        if not mac_file_path:
            mac_filename_list = []
            for row_i in range(2, worksheet.max_row + 1):
                mac_filename_list.append(str(worksheet.cell(row=row_i, column=2).value))

            root_win.attributes("-disabled", 1)  # 禁止操作主页面
            global top1
            top1 = tk.Toplevel(bg="#eaf1f6")  # 创建顶层窗口
            center_window(top1)  # 将窗体移动到屏幕中央

            top1.geometry("%dx%d" % (300, 250))  # 设置顶层窗口大小
            top1.overrideredirect(True)

            top1.protocol("WM_DELETE_WINDOW", allow_main_window1)  # 如果窗口关闭，则允许操作主页面，并将top1销毁

            f0 = tk.Frame(top1, bg="#eaf1f6", bd=0)
            tk.Label(f0, text='\n根据产品类型选MVE属性配置文件(.mac)\n', bg="#eaf1f6", fg="black", justify='center').pack(fill=tk.X)
            f0.pack(fill=tk.X)

            f1 = tk.Frame(top1, bg="#eaf1f6", bd=0)
            tk.Label(f1, text='MAC:', bg="#eaf1f6", fg="black", justify='left').pack(side=tk.LEFT, fill=tk.X, padx=(5, 0))

            global combobox_mac
            combobox_mac_value = StringVar()
            combobox_mac_values = mac_filename_list
            combobox_mac = ttk.Combobox(master=f1, width=30, font=("ABBvoice CNSG", int(10 * h_ratio)), state='readonly', cursor='arrow', textvariable=combobox_mac_value, values=combobox_mac_values)
            combobox_mac.pack(side=tk.LEFT, padx=0)

            global button_confirm
            button_confirm = tk.Button(f1, text="确认", command=confirm_mac_file)
            button_confirm.pack(side=tk.LEFT, padx=5)
            f1.pack(fill=tk.X)

            tk.Frame(top1, height=1, bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

            f2 = tk.Frame(top1, bg="#eaf1f6", bd=0)
            tk.Label(f2, text='产品名                     文件名\nUniGear ZS1           ZS1\nUniGear ZS3.2        ZS3\nZX0                          ZX00\nZX0.2                       ZX02\nZX1.2                       ZX12\nZX2                          ZX20\nPrimeGear ZX0       PrimeGear ZX0\nZX1.5-R                   ZX15', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(8 * h_ratio)), justify='left').pack(
                fill=tk.X)
            f2.pack(fill=tk.X)

        else:
            read_mac_file()
            insert_to_mve_level_tree()  # 在选择了MAC文件后再执行插入操作


def confirm_mac_file():
    if not combobox_mac.get():
        tk.messagebox.showwarning("提示", "请选择配置文件")
        top1.lift()  # 使顶层窗口继续处于顶层(超越其它窗口)
    else:
        global mac_file_path
        mac_file_path = os.path.join(template_directory, combobox_mac.get() + '.mac')
        read_mac_file()  # 确认选择后读取MAC文件
        insert_to_mve_level_tree()  # 确认选择后再执行插入操作
        allow_main_window1()


def allow_main_window1():
    top1.destroy()
    root_win.attributes("-disabled", 0)


def read_mac_file():
    if not os.path.exists(mac_file_path):
        tk.messagebox.showwarning("提示", "失败,找不到MAC文件: {}".format(mac_file_path))
    else:
        mac_tree = etree.parse(mac_file_path, parser)
        mac_root = mac_tree.getroot()

        global tabs_data
        tabs_data = {
            "project": {},
            "switchgear": {},
            "typical": {},
            "panel": {},
        }
        global original_attribute_names
        original_attribute_names = {}

        # 解析.MAC文件，存放到右树
        for attribute in mac_root.findall('.//attribute'):
            access_elem = attribute.find('access')
            # 检查access节点并确认其子节点是否仅为管理员的读写权限
            if access_elem is not None:
                read_write_admin_only = len(access_elem.findall('ReadWrite')) == 1 and \
                                        access_elem.find('ReadWrite').text == "Administrator" and \
                                        len(access_elem) == 1
                if read_write_admin_only:
                    continue  # 跳过当access不符合条件时的当前属性
                else:
                    xml_link_elem = attribute.find('XML_Link')
                    xml_link = xml_link_elem.text if xml_link_elem is not None else None

                    if xml_link and any(keyword in xml_link.lower() for keyword in ["project", "switchgear", "typical", "panel"]):
                        defaultvalue_elem = attribute.find('defaultvalue')
                        if defaultvalue_elem is not None:
                            defaultvalue_elem_alias = defaultvalue_elem.find('alias')
                            defaultvalue_alias = defaultvalue_elem_alias.text.strip() if defaultvalue_elem_alias is not None and defaultvalue_elem_alias.text else None

                            defaultvalue_elem_description = defaultvalue_elem.find('description')
                            defaultvalue_description = defaultvalue_elem_description.text.strip() if defaultvalue_elem_description is not None and defaultvalue_elem_description.text else None
                        else:
                            defaultvalue_alias = None
                            defaultvalue_description = None

                        description_elem = attribute.find('description')
                        description = description_elem.text.strip() if description_elem is not None and description_elem.text else None

                        group_elem = attribute.find('group')
                        group = group_elem.text.strip() if group_elem is not None and group_elem.text else "Other"

                        valuelist_elem = attribute.find('valuelist')
                        valuelist = []
                        if valuelist_elem is not None:
                            for value in valuelist_elem.findall('value'):
                                value_alias = value.find('alias').text
                                value_description = value.find('description').text
                                valuelist.append((value_alias, value_description))
                        else:
                            valuelist.append((None, None))

                        # 使用 find_xml_value 函数查找值，如果没找到使用默认值
                        xml_value, xml_exist = find_xml_value(xml_root, xml_link)
                        xml_value = [value if value is not None else defaultvalue_description for value in xml_value]

                        # 遍历xml_value并转换为对应的value_description
                        for i, value in enumerate(xml_value):
                            for value_alias, value_description in valuelist:
                                if value == value_alias:
                                    xml_value[i] = value_description

                        # 将每个属性与其组编号对应起来
                        category = xml_link.lower().split('[')[0].strip()

                        if category in tabs_data:
                            if group not in tabs_data[category]:
                                # 如果该组名还没有被创建，则创建一个新组
                                tabs_data[category][group] = {"attributes": [], "values": [], "valuelist": [], "exist": [], "xml_links": []}

                            # 添加属性说明(description)和属性值(xml_value)到对应分类和组别
                            tabs_data[category][group]["attributes"].append(description)
                            tabs_data[category][group]["values"].append(xml_value)
                            tabs_data[category][group]["valuelist"].append(valuelist)
                            tabs_data[category][group]["exist"].append(xml_exist)
                            tabs_data[category][group]["xml_links"].append(xml_link)
                            # print(description, xml_value, valuelist, xml_exist)


def parse_attributes(attributelist, target_name, target_list):
    if attributelist is not None:
        for attribute in attributelist:
            name = attribute.find('name').text
            value = attribute.find('value').text
            # print(name, value)
            if name.lower() == target_name:
                target_list.append(value)


def find_xml_value(xml_root, xml_link):
    split_link = xml_link.split("[")
    if len(split_link) == 2:
        tab = split_link[0].lower()
        attribute_name = split_link[1].rstrip("]")

        # 获取ObjectName的计数作为列对齐的参考
        base_path = './/project/'
        tab_paths = {
            "project": "attributelist/attribute[name='ObjectName']/value",
            "switchgear": "switchgearlist/switchgear/attributelist/attribute[name='ObjectName']/value",
            "typical": "switchgearlist/switchgear/typicallist/typical/attributelist/attribute[name='ObjectName']/value",
            "panel": "switchgearlist/switchgear/panellist/panel/attributelist/attribute[name='ObjectName']/value"
        }
        object_name_path = base_path + tab_paths.get(tab, "")

        object_names = xml_root.findall(object_name_path)
        object_name_count = len(object_names)

        # 如果没有找到ObjectName，也无法对齐其他属性的值
        if object_name_count == 0:
            return []

        search_paths = {
            "project": "attributelist/attribute[name='{}']/value".format(attribute_name),
            "switchgear": "switchgearlist/switchgear/attributelist/attribute[name='{}']/value".format(attribute_name),
            "typical": "switchgearlist/switchgear/typicallist/typical/attributelist/attribute[name='{}']/value".format(attribute_name),
            "panel": "switchgearlist/switchgear/panellist/panel/attributelist/attribute[name='{}']/value".format(attribute_name)
        }
        search_path = base_path + search_paths.get(tab, "")

        if search_path:
            # 创建一个列表，预先填充占位符，用于确保对齐
            attribute_values_aligned = [None] * object_name_count
            # 查找对应的属性值
            attribute_values_elements = xml_root.findall(search_path)

            # 用于存储该属性是否在xml中存在
            attribute_exist = [False] * object_name_count

            # 遍历属性值节点，并更新预填充列表中对应的位置
            for i, elem in enumerate(attribute_values_elements):
                if i < object_name_count:  # 避免超出ObjectName的数量
                    attribute_values_aligned[i] = elem.text.strip() if elem.text and elem.text.strip() != '' else None
                    # attribute_exist[i] = elem.text is not None and elem.text.strip() != ''
                    attribute_exist[i] = True

            return attribute_values_aligned, attribute_exist

    # XML链接格式不正确或其他原因无法找到属性值时返回对齐的空列表
    return [None] * object_name_count if 'object_name_count' in locals() else [], [False] * object_name_count if 'object_name_count' in locals() else []


def get_item_level(tree, item):
    level = 0
    while tree.parent(item):
        item = tree.parent(item)
        level += 1
    return level


def on_tree_select(event):
    selected_item = mve_level_tree.selection()[0]
    item_text = mve_level_tree.item(selected_item, "text")
    item_level = get_item_level(mve_level_tree, selected_item)
    global order_index
    order_index = calculate_order_index(mve_level_tree, selected_item, item_level)
    global display_data
    if item_level == 0:
        display_data = tabs_data.get("project", {})
        mve_notebook_right.add(pane_right, text='Project', image=project_mve, compound=tk.LEFT)
    elif item_level == 1:
        display_data = tabs_data.get("switchgear", {})
        mve_notebook_right.add(pane_right, text='Switchgear', image=switchgear_mve, compound=tk.LEFT)
    elif item_level == 2:
        if '<' not in item_text and '>' not in item_text:
            display_data = tabs_data.get("typical", {})
            mve_notebook_right.add(pane_right, text='Typical', image=typical_mve, compound=tk.LEFT)
        else:
            display_data = tabs_data.get("panel", {})
            mve_notebook_right.add(pane_right, text='Panel', image=panel_mve, compound=tk.LEFT)
    else:
        display_data = {}  # 其他逻辑
    # print(selected_item,item_text)
    # print(f"选中节点 '{item_text}' 在层级 {item_level} 中的次序为 {order_index}")

    update_tab(display_data, order_index)
    button_attribute_hv['state'] = 'normal'


def calculate_order_index(tree, item, item_level):
    if item_level == 0:
        order_index = 1
    elif item_level == 1:
        # 计算第二层次序
        parent = tree.parent(item)  # 父节点
        all_siblings = tree.get_children(parent)  # 子节点兄弟
        order_index = list(all_siblings).index(item) + 1
    elif item_level == 2:
        # 计算第三层次序
        parent = tree.parent(item)  # 父节点
        all_siblings = tree.get_children(parent)  # 子节点兄弟
        second_level_parent = tree.parent(parent)  # 祖父节点
        all_second_level_siblings = tree.get_children(second_level_parent)  # 父节点兄弟

        order_index = 0
        item_text = tree.item(item, "text")
        contains_special_char = '<' in item_text or '>' in item_text

        for sibling in all_second_level_siblings:
            third_level_children = tree.get_children(sibling)
            for child in third_level_children:
                child_text = tree.item(child, "text")
                child_contains_special_char = '<' in child_text or '>' in child_text

                if contains_special_char:
                    if child_contains_special_char:
                        order_index += 1
                else:
                    if not child_contains_special_char:
                        order_index += 1

                if sibling == parent and child == item:
                    break

            if sibling == parent and item in third_level_children:
                break
    else:
        order_index = -1  # 非第一/二/三层节点

    return order_index


def update_tab(data, order_index):
    # 清空原有数据
    for tree in [pane_right_tree]:
        for item in tree.get_children():
            tree.delete(item)
    pane_right_tree.column("#0", width=300)

    column_names = set()

    # # 确定pane_right_tree的行和列数
    # for attributes in data.values():
    #     for i in range(len(attributes["attributes"])):
    #         for j, value in enumerate(attributes["values"][i]):
    #             column_name = f"col{j+1}"
    #             column_names.add(column_name)

    column_names = ["col1"]  # 只显示1列属性值
    pane_right_tree["columns"] = list(column_names)

    for col in pane_right_tree["columns"]:
        pane_right_tree.heading(col, text=col)
        pane_right_tree.column(col, width=100, anchor="w")

    parent_nodes_to_check = []
    for group, group_data in data.items():
        parent = pane_right_tree.insert("", "end", text=group, open=True, tags='right_line')  # 每个属性组父节点
        parent_nodes_to_check.append((parent, group_data))
        for i in range(len(group_data["attributes"])):  # 属性组中的属性
            attribute = group_data["attributes"][i]
            # values = [str(value) if value is not None else '' for value in group_data["values"][i]]
            value = group_data["values"][i][order_index - 1]
            if not value:
                value = ''
            values = [str(value).replace('\t', '').replace('\n', ';')]
            if group_data["exist"][i][0]:
                pane_right_tree.insert(parent, "end", text=attribute, values=values, tags='right_mention_line')
            elif attribute_hv_flag:
                pane_right_tree.insert(parent, "end", text=attribute, values=values, tags='right_no_mention_line')

    # 检查并清除不符合条件的parent节点
    for parent, group_data in parent_nodes_to_check:
        if not attribute_hv_flag and all(not exist[0] for exist in group_data["exist"]):
            tree.delete(parent)

    for col in pane_right_tree["columns"]:
        max_width = get_display_length(col)
        for item in pane_right_tree.get_children(''):
            for child in pane_right_tree.get_children(item):
                cell_value = pane_right_tree.set(child, col)
                max_width = max(max_width, get_display_length(str(cell_value)))
                # print(cell_value, max_width)
        # max_width = max(100, max_width * 9)
        max_width = max(100, 500)
        pane_right_tree.column(col, width=max_width)
    pane_right_tree.bind('<Double-1>', lambda event, tree=pane_right_tree: edit_cell(event, tree))


def expand_tree(tree, level):
    def expand_children(item, current_level, target_level):
        if current_level < target_level:
            tree.item(item, open=True)
            children = tree.get_children(item)
            for child in children:
                expand_children(child, current_level + 1, target_level)
        else:
            tree.item(item, open=False)

    for item in tree.get_children():
        expand_children(item, 0, level)


def on_level_fold():
    expand_tree(mve_level_tree, 1)


def on_level_unfold():
    expand_tree(mve_level_tree, 2)


def delete_node(tree, item):
    # 获取节点的完整路径
    path = []
    while item:
        path.insert(0, tree.item(item, "text"))
        item = tree.parent(item)

    sure_to_delete_flag = False

    if len(path) == 2:
        if tk.messagebox.askyesno("提示", "确定要移除%s节点及其子节点吗？" % path[-1]):
            sure_to_delete_flag = True
    elif len(path) == 3:
        if tk.messagebox.askyesno("提示", "确定要移除%s节点吗？" % path[-1]):
            sure_to_delete_flag = True
    # print(path)
    if sure_to_delete_flag:
        # 删除树节点及其子节点
        delete_tree_node(tree, tree.selection()[0])
        # 根据路径删除xml_root中的对应内容
        delete_from_xml(path)


def delete_tree_node(tree, item):
    for child in tree.get_children(item):
        delete_tree_node(tree, child)
    tree.delete(item)


def delete_from_xml(path):
    base_path = './/project/'
    # 查找满足第一个条件的project节点
    project_nodes = xml_root.findall(".//project")
    for project in project_nodes:
        project_name = project.find("attributelist/attribute[name='ObjectName']/value")
        if project_name is not None and project_name.text == path[0]:
            if len(path) == 2:
                # 查找满足第二个条件的switchgear节点
                switchgear_nodes = project.findall("switchgearlist/switchgear")
                for switchgear in switchgear_nodes:
                    switchgear_name = switchgear.find("attributelist/attribute[name='ObjectName']/value")
                    if switchgear_name is not None and switchgear_name.text == path[1]:
                        switchgear.getparent().remove(switchgear)
                        tk.messagebox.showwarning("提示", f"Switchgear节点'{'/'.join(path)}'被移除")
                        button_export['state'] = 'normal'
                # 清空原有数据
                for tree in [pane_right_tree]:
                    for item in tree.get_children():
                        tree.delete(item)
                read_mac_file()

            elif len(path) == 3:
                # 查找满足第二个条件的switchgear节点
                switchgear_nodes = project.findall("switchgearlist/switchgear")
                for switchgear in switchgear_nodes:
                    switchgear_name = switchgear.find("attributelist/attribute[name='ObjectName']/value")
                    if switchgear_name is not None and switchgear_name.text == path[1]:
                        # 查找满足第三个条件的typical/panel节点
                        typical_nodes = switchgear.findall("typicallist/typical")
                        panel_nodes = switchgear.findall("panellist/panel")
                        if '<' not in path[2] and '>' not in path[2]:
                            for typical in typical_nodes:
                                typical_name = typical.find("attributelist/attribute[name='ObjectName']/value")
                                if typical_name is not None and typical_name.text == path[2]:
                                    typical.getparent().remove(typical)
                                    tk.messagebox.showwarning("提示", f"Typical节点'{'/'.join(path)}'被移除")
                                    button_export['state'] = 'normal'
                        else:
                            for panel in panel_nodes:
                                panel_name = panel.find("attributelist/attribute[name='ObjectName']/value")
                                if panel_name is not None and panel_name.text == path[2].split(' <')[0]:
                                    panel.getparent().remove(panel)
                                    tk.messagebox.showwarning("提示", f"Panel节点'{'/'.join(path)}'被移除")
                                    button_export['state'] = 'normal'
                for tree in [pane_right_tree]:
                    for item in tree.get_children():
                        tree.delete(item)
                read_mac_file()


def record_node_states(tree):
    global node_states
    node_states = {}

    def record_recursive(item, path=""):
        current_path = f"{path}/{tree.item(item, 'text')}" if path else tree.item(item, 'text')
        node_states[current_path] = tree.item(item, 'open')

        for child in tree.get_children(item):
            record_recursive(child, current_path)

    for item in tree.get_children():
        record_recursive(item)

    return node_states


def apply_node_states(tree, states):
    def apply_recursive(item, path=""):
        current_path = f"{path}/{tree.item(item, 'text')}" if path else tree.item(item, 'text')
        if current_path in states:
            tree.item(item, open=states[current_path])

        for child in tree.get_children(item):
            apply_recursive(child, current_path)

    for item in tree.get_children():
        apply_recursive(item)


def rename_node(tree, item):
    original_name = tree.item(item, "text")
    master_typical = ''
    if '<' in original_name and '>' in original_name:
        master_typical = ' <' + original_name.split(' <')[1]
        original_name = original_name.split(' <')[0]

    x, y, width, height = tree.bbox(item, column="#0")
    entry = tk.Entry(tree)
    entry.insert(0, original_name)
    entry.place(x=x, y=y, width=width, height=height)

    def get_full_name(name, parent):
        """
        生成不重复的名称，如果名称已存在，则自动添加 `(1)`、`(2)` 等
        """
        existing_names = [tree.item(child, "text").split(' <')[0] for child in tree.get_children(parent)]
        full_name = name
        index = 1
        while full_name in existing_names:
            full_name = f"{name}({index})"
            index += 1
        return full_name

    def save_edit(widget, item):
        new_name = widget.get()
        widget.destroy()
        if new_name and new_name != original_name:
            # 获取节点的完整路径
            path = []
            current_item = item
            while current_item:
                path.insert(0, tree.item(current_item, "text"))
                current_item = tree.parent(current_item)

            parent_item = tree.parent(item)
            # 确保新名称唯一
            new_name = get_full_name(new_name, parent_item)
            # 更新树节点名称
            tree.item(item, text=new_name + master_typical)

            # 记录节点状态
            node_states = record_node_states(mve_level_tree)

            # 更新 XML 内容
            # print(path, new_name)
            rename_in_xml(path, new_name)
        else:
            tree.item(item, text=original_name + master_typical)

    def on_leave(event):
        if event.state == 8:
            save_edit(event.widget, item)
        else:
            event.widget.unbind('<FocusOut>')

    def on_focus_out(event):
        event.widget.bind('<FocusOut>', lambda e: event.widget.destroy())

    entry.focus_set()
    entry.bind('<Return>', lambda e: save_edit(entry, item))
    entry.bind('<Leave>', on_leave)
    entry.bind('<FocusIn>', on_focus_out)


def rename_in_xml(path, new_name):
    # 查找满足第一个条件的project节点
    project_nodes = xml_root.findall(".//project")
    for project in project_nodes:
        project_name = project.find("attributelist/attribute[name='ObjectName']/value")
        if project_name is not None and project_name.text == path[0]:
            if len(path) == 1:
                # 查找并重命名project节点
                project_name.text = new_name
                tk.messagebox.showwarning("提示", f"Project节点'{'/'.join(path)}'重命名为'{new_name}'")
                button_export['state'] = 'normal'

            if len(path) == 2:
                # 查找并重命名switchgear节点
                switchgear_nodes = project.findall("switchgearlist/switchgear")
                for switchgear in switchgear_nodes:
                    switchgear_name = switchgear.find("attributelist/attribute[name='ObjectName']/value")
                    if switchgear_name is not None and switchgear_name.text == path[1]:
                        switchgear_name.text = new_name
                        tk.messagebox.showwarning("提示", f"Switchgear节点'{'/'.join(path)}'重命名为'{new_name}'")
                        button_export['state'] = 'normal'

                # for tree in [mve_level_tree, pane_right_tree]:
                #     for item in tree.get_children():
                #         tree.delete(item)
                # parse_xml_file()
                # read_mac_file()
                # insert_to_mve_level_tree()
                # on_level_fold()

            elif len(path) == 3:
                # 查找并重命名typical/panel节点
                switchgear_nodes = project.findall("switchgearlist/switchgear")
                for switchgear in switchgear_nodes:
                    switchgear_name = switchgear.find("attributelist/attribute[name='ObjectName']/value")
                    if switchgear_name is not None and switchgear_name.text == path[1]:

                        typical_nodes = switchgear.findall("typicallist/typical")
                        panel_nodes = switchgear.findall("panellist/panel")
                        if '<' not in path[2] and '>' not in path[2]:
                            for typical in typical_nodes:
                                typical_name = typical.find("attributelist/attribute[name='ObjectName']/value")
                                if typical_name is not None and typical_name.text == path[2]:
                                    typical_name.text = new_name
                                    tk.messagebox.showwarning("提示", f"Typical节点'{'/'.join(path)}'重命名为'{new_name}'")

                                    # 更新panellist中的TypicalName
                                    panellist = switchgear.find("panellist")
                                    for panel in panellist.findall("panel"):
                                        panel_name = panel.find("attributelist/attribute[name='ObjectName']/value").text
                                        typical_name_panel = panel.find("attributelist/attribute[name='TypicalName']/value")
                                        if typical_name_panel is not None and typical_name_panel.text == path[2]:
                                            typical_name_panel.text = new_name
                                            tk.messagebox.showwarning("提示", f"Panel节点{panel_name}所属Typical'{path[2]}'重命名为'{new_name}'")
                                            button_export['state'] = 'normal'
                        else:
                            for panel in panel_nodes:
                                panel_name = panel.find("attributelist/attribute[name='ObjectName']/value")
                                if panel_name is not None and panel_name.text == path[2].split(' <')[0]:
                                    panel_name.text = new_name
                                    tk.messagebox.showwarning("提示", f"Panel节点'{'/'.join(path)}'重命名为'{new_name}'")

            for tree in [mve_level_tree, pane_right_tree]:
                for item in tree.get_children():
                    tree.delete(item)
            parse_xml_file()
            read_mac_file()
            insert_to_mve_level_tree()
            apply_node_states(mve_level_tree, node_states)

            if len(path) == 1:
                for item in mve_level_tree.get_children():
                    if mve_level_tree.item(item, "text") == new_name:
                        mve_level_tree.selection_set(item)
                        break
            elif len(path) == 2:
                for item in mve_level_tree.get_children():
                    if mve_level_tree.item(item, 'text') == path[0]:
                        for subitem in mve_level_tree.get_children(item):
                            if mve_level_tree.item(subitem, 'text') == new_name:
                                mve_level_tree.selection_set(subitem)
                                break
            elif len(path) == 3:
                for item in mve_level_tree.get_children():
                    if mve_level_tree.item(item, 'text') == path[0]:
                        for subitem in mve_level_tree.get_children(item):
                            if mve_level_tree.item(subitem, 'text') == path[1]:
                                for subsubitem in mve_level_tree.get_children(subitem):
                                    if mve_level_tree.item(subsubitem, 'text').split(' <')[0] == new_name:
                                        mve_level_tree.selection_set(subsubitem)
                                        break


def on_tree_press(event):
    global start_item, current_item
    start_item = None
    current_item = None
    selected_items = mve_level_tree.selection()
    if not selected_items:
        return  # 如果没有选中的项，直接返回
    selected_item = selected_items[0]

    item_text = mve_level_tree.item(selected_item, "text")
    item_level = get_item_level(mve_level_tree, selected_item)
    if item_level == 2:
        if '<' in item_text and '>' in item_text:
            start_item = selected_item
            current_item = selected_item


def on_tree_release(event):
    global start_item, current_item
    if start_item and current_item and start_item != current_item:
        # 获取当前项的文本内容
        current_text = mve_level_tree.item(current_item, "text")

        # 检查目标项文本是否符合条件
        if '<' in current_text and '>' in current_text:
            move_node(start_item, current_item)
    start_item = None
    current_item = None


def on_tree_move(event):
    global start_item, current_item
    item = mve_level_tree.identify_row(event.y)
    if item and item != current_item:
        # 获取目标项的文本内容
        target_text = mve_level_tree.item(item, "text")

        # 检查目标项文本是否符合条件
        if '<' in target_text and '>' in target_text:
            current_item = item


def move_node(start_item, end_item):
    if start_item and end_item:
        # 获取开始节点和结束节点的父节点
        start_parent = mve_level_tree.parent(start_item)
        end_parent = mve_level_tree.parent(end_item)

        # 检查是否在同一父节点范围内
        if start_parent == end_parent:
            # 同一父节点内移动
            start_index = mve_level_tree.index(start_item)
            end_index = mve_level_tree.index(end_item)
            mve_level_tree.move(start_item, end_parent, end_index)

            # 统计父节点下的不含 <、> 的 typical 节点数量
            typical_count = 0
            for child in mve_level_tree.get_children(start_parent):
                child_text = mve_level_tree.item(child, "text")
                if '<' not in child_text and '>' not in child_text:
                    typical_count += 1

            # 调整实际的 panel 节点次序
            start_index_adjusted = start_index - typical_count
            end_index_adjusted = end_index - typical_count

            mve_level_tree.move(start_item, end_parent, end_index)

            # 更新XML内容
            path = []
            current_item = start_item
            while current_item:
                path.insert(0, mve_level_tree.item(current_item, "text"))
                current_item = mve_level_tree.parent(current_item)

            # print(path, start_index_adjusted, end_index_adjusted)
            move_in_xml(path, start_index_adjusted, end_index_adjusted)


def move_in_xml(path, start_index, end_index):
    project_nodes = xml_root.findall(".//project")
    for project in project_nodes:
        project_name = project.find("attributelist/attribute[name='ObjectName']/value")
        if project_name is not None and project_name.text == path[0]:
            if len(path) == 3:
                switchgear_nodes = project.findall("switchgearlist/switchgear")
                for switchgear in switchgear_nodes:
                    switchgear_name = switchgear.find("attributelist/attribute[name='ObjectName']/value")
                    if switchgear_name is not None and switchgear_name.text == path[1]:
                        panel_nodes = switchgear.findall("panellist/panel")
                        # 移动panel节点
                        panel_to_move = panel_nodes.pop(start_index)
                        panel_nodes.insert(end_index, panel_to_move)
                        # 清空并重新添加节点
                        switchgear.find("panellist").clear()
                        for panel in panel_nodes:
                            switchgear.find("panellist").append(panel)
                        tk.messagebox.showwarning("提示", f"Panel节点'{'/'.join(path)}'已移动")
                        button_export['state'] = 'normal'
                        read_mac_file()


def copy_node(tree, selected_item):
    parent = tree.parent(selected_item)
    item_text = tree.item(selected_item, "text")

    def generate_unique_name(name, parent):
        # 生成不重复的名称
        names = [tree.item(child, "text") for child in tree.get_children(parent)]
        new_name = name
        index = 1
        while new_name in names:
            new_name = f"{name}({index})"
            index += 1
        return new_name

    new_item_text = generate_unique_name(item_text, parent)

    new_item = tree.insert(parent, tree.index(selected_item) + 1, text=new_item_text, image=typical_mve, tags='left_no_mention_line')

    # 复制XML节点
    path = []
    current_item = selected_item
    while current_item:
        path.insert(0, tree.item(current_item, "text"))
        current_item = tree.parent(current_item)

    copy_xml_node(path, new_item_text)


def copy_xml_node(path, new_item_text):
    project_nodes = xml_root.findall(".//project")
    for project in project_nodes:
        project_name = project.find("attributelist/attribute[name='ObjectName']/value")
        if project_name is not None and project_name.text == path[0]:
            if len(path) == 3:
                switchgear_nodes = project.findall("switchgearlist/switchgear")
                for switchgear in switchgear_nodes:
                    switchgear_name = switchgear.find("attributelist/attribute[name='ObjectName']/value")
                    if switchgear_name is not None and switchgear_name.text == path[1]:
                        typical_nodes = switchgear.findall("typicallist/typical")
                        for typical in typical_nodes:
                            typical_name = typical.find("attributelist/attribute[name='ObjectName']/value")
                            if typical_name is not None and typical_name.text == path[2]:
                                new_typical = etree.fromstring(etree.tostring(typical))
                                new_typical.find("attributelist/attribute[name='ObjectName']/value").text = new_item_text
                                switchgear.find("typicallist").insert(typical_nodes.index(typical) + 1, new_typical)
                                tk.messagebox.showwarning("提示", f"Typical节点'{'/'.join(path)}'已创建")
                                button_export['state'] = 'normal'

                # for tree in [pane_right_tree]:
                #     for item in tree.get_children():
                #         tree.delete(item)
                read_mac_file()


def on_right_click(event):
    item = mve_level_tree.identify_row(event.y)
    if item:
        # 获取节点的完整路径以确定层级
        level = 0
        temp_item = item
        while temp_item:
            level += 1
            temp_item = mve_level_tree.parent(temp_item)

        # 清空右键菜单
        popup_menu.delete(0, tk.END)

        if level in [1, 2, 3]:
            item_text = mve_level_tree.item(item, "text")
            if level == 1:
                # 只有重命名功能项
                popup_menu.add_command(label="重命名", command=lambda: rename_node(mve_level_tree, item))
            if level == 2 or ('<' in item_text and '>' in item_text):
                # 只有重命名和删除功能项
                popup_menu.add_command(label="重命名", command=lambda: rename_node(mve_level_tree, item))
                popup_menu.add_command(label="删除", command=lambda: delete_node(mve_level_tree, item))
            elif level == 3 and '<' not in item_text and '>' not in item_text:
                # 重命名、删除和复制功能项
                popup_menu.add_command(label="重命名", command=lambda: rename_node(mve_level_tree, item))
                popup_menu.add_command(label="删除", command=lambda: delete_node(mve_level_tree, item))
                popup_menu.add_command(label="复制", command=lambda: copy_node(mve_level_tree, item))

            # 显示右键菜单
            mve_level_tree.selection_set(item)
            popup_menu.post(event.x_root, event.y_root)
            popup_menu.grab_release()


def edit_cell(event, tree):
    global left_item, left_item_copy, modify_objectname_flag, typical_panel_flag
    typical_panel_flag = 0
    modify_objectname_flag = False
    if mve_level_tree.selection():
        left_item = mve_level_tree.selection()[0]
        left_item_copy = mve_level_tree.selection()[0]
    else:
        left_item = None
        left_item_copy = None
    column_t = tree.identify_column(event.x)
    row_t = tree.identify_row(event.y)
    # print(left_item, row_t, column_t)

    # 检查是否是parent行
    if not tree.parent(row_t):  # 如果row_t没有parent，说明这是一个parent行
        return

    if column_t:
        col = int(str(column_t).replace('#', ''))
        if col == 1:
            content = tree.item(row_t, 'value')[0]
            x = tree.bbox(row_t, column=column_t)[0]  # 单元格x坐标
            y = tree.bbox(row_t, column=column_t)[1]  # 单元格y坐标
            width = tree.bbox(row_t, column=column_t)[2]  # 单元格宽度
            height = tree.bbox(row_t, column=column_t)[3]  # 单元格高度

            # 获取当前行的attribute
            description = tree.item(row_t, 'text')

            # 获取当前行的valuelist
            valuelist = None

            for category, groups in tabs_data.items():
                # print(category, groups)
                # 判断 mve_notebook_right 的 text 属性（全小写后）是否与 category 一致
                for tab_id in mve_notebook_right.tabs():
                    tab_text = mve_notebook_right.tab(tab_id, option="text")
                    if tab_text.lower() == category.lower():
                        for group, data in groups.items():
                            # print(group, data)
                            if description in data["attributes"]:
                                valuelist = data["valuelist"][data["attributes"].index(description)]
                                break
                if valuelist is not None:
                    break

            if valuelist and any(value != (None, None) for value in valuelist):
                values = [value_description for _, value_description in valuelist]
                comboedit = ttk.Combobox(tree, values=values, font=("ABBvoice CNSG", int(10 * h_ratio)))
                comboedit.place(x=x, y=y, width=width, height=height)
                comboedit.set(content)
            else:
                entryedit = ttk.Entry(tree)
                entryedit.place(x=x, y=y, width=width, height=height)
                entryedit.insert(0, content)

            def on_leave(event):
                if event.state == 8:
                    save_edit(event.widget)
                else:
                    event.widget.unbind('<FocusOut>')

            def on_focus_out(event):
                event.widget.bind('<FocusOut>', lambda e: event.widget.destroy())

            if valuelist and any(value != (None, None) for value in valuelist):
                comboedit.focus_set()
                comboedit.bind('<<ComboboxSelected>>', lambda e: save_edit(comboedit))
                comboedit.bind('<Leave>', on_leave)
                comboedit.bind('<Return>', lambda e: save_edit(comboedit))
                comboedit.bind('<FocusIn>', on_focus_out)
            else:
                entryedit.focus_set()
                entryedit.bind('<Return>', lambda e: save_edit(entryedit))
                entryedit.bind('<Leave>', on_leave)
                entryedit.bind('<FocusIn>', on_focus_out)

            def save_edit(widget):
                global new_content, left_path, left_same_level_node_text, typical_panel_flag
                new_content = widget.get()
                tree.set(row_t, column_t, new_content)
                # print(row_t, column_t, tree.item(row_t, 'value'))
                widget.destroy()

                # 获取节点的完整路径
                if left_item_copy is not None:
                    item_copy = left_item_copy
                    left_path = []
                    left_same_level_node_text = []
                    if '<' not in mve_level_tree.item(item_copy, 'text') and '>' not in mve_level_tree.item(item_copy, 'text'):
                        left_same_level_node_text = [
                            mve_level_tree.item(child, "text").split(' <')[0]
                            for child in mve_level_tree.get_children(mve_level_tree.parent(item_copy))
                            if child != item_copy and '<' not in mve_level_tree.item(child, 'text') and '>' not in mve_level_tree.item(child, 'text')
                        ]
                        typical_panel_flag = 1
                    elif '<' in mve_level_tree.item(item_copy, 'text') and '>' in mve_level_tree.item(item_copy, 'text'):
                        left_same_level_node_text = [
                            mve_level_tree.item(child, "text").split(' <')[0]
                            for child in mve_level_tree.get_children(mve_level_tree.parent(item_copy))
                            if child != item_copy and '<' in mve_level_tree.item(child, 'text') and '>' in mve_level_tree.item(child, 'text')
                        ]
                        typical_panel_flag = 2

                    while item_copy:
                        left_path.insert(0, mve_level_tree.item(item_copy, "text"))
                        item_copy = mve_level_tree.parent(item_copy)

                # print(left_path, left_same_level_node_text)

                # 记录节点状态
                node_states = record_node_states(mve_level_tree)

                # 更新XML数据
                for category, groups in tabs_data.items():
                    for tab_id in mve_notebook_right.tabs():
                        tab_text = mve_notebook_right.tab(tab_id, option="text")
                        if tab_text.lower() == category.lower():
                            for group, data in groups.items():
                                if description in data["attributes"]:
                                    index = data["attributes"].index(description)
                                    xml_link = data["xml_links"][index]
                                    # print(description, index, xml_link)
                                    update_xml_value(xml_root, xml_link, new_content, left_path)
                                    break

                for xtree in [mve_level_tree]:
                    for item in xtree.get_children():
                        xtree.delete(item)

                parse_xml_file()
                read_mac_file()
                insert_to_mve_level_tree()
                apply_node_states(mve_level_tree, node_states)

                # 当 modify_objectname_flag 为 True 时，选择被修改名字的节点
                # print(modify_objectname_flag, new_content)
                if modify_objectname_flag:
                    if len(left_path) == 1:
                        for item in mve_level_tree.get_children():
                            if mve_level_tree.item(item, "text") == new_content:
                                mve_level_tree.selection_set(item)
                                break
                    elif len(left_path) == 2:
                        for item in mve_level_tree.get_children():
                            if mve_level_tree.item(item, 'text') == left_path[0]:
                                for subitem in mve_level_tree.get_children(item):
                                    if mve_level_tree.item(subitem, 'text') not in left_same_level_node_text:
                                        mve_level_tree.selection_set(subitem)
                                        break
                    elif len(left_path) == 3:
                        for item in mve_level_tree.get_children():
                            if mve_level_tree.item(item, 'text') == left_path[0]:
                                for subitem in mve_level_tree.get_children(item):
                                    if mve_level_tree.item(subitem, 'text') == left_path[1]:
                                        if typical_panel_flag == 1:
                                            for subsubitem in mve_level_tree.get_children(subitem):
                                                if mve_level_tree.item(subsubitem, 'text') not in left_same_level_node_text and '<' not in mve_level_tree.item(subsubitem, 'text') and '>' not in mve_level_tree.item(subsubitem, 'text'):
                                                    mve_level_tree.selection_set(subsubitem)
                                                    break
                                        elif typical_panel_flag == 2:
                                            for subsubitem in mve_level_tree.get_children(subitem):
                                                if mve_level_tree.item(subsubitem, 'text').split(' <')[0] not in left_same_level_node_text and '<' in mve_level_tree.item(subsubitem, 'text') and '>' in mve_level_tree.item(subsubitem, 'text'):
                                                    mve_level_tree.selection_set(subsubitem)
                                                    break

                button_export['state'] = 'normal'

            def update_xml_value(xml_root, xml_link, new_content, path):
                global modify_objectname_flag
                project_nodes = xml_root.findall(".//project")
                for project in project_nodes:
                    project_name = project.find("attributelist/attribute[name='ObjectName']/value")
                    if project_name is not None and project_name.text == path[0]:
                        if len(path) == 1:  # 说明修改的是Project层级
                            split_link = xml_link.split("[")
                            if len(split_link) == 2:
                                attribute_name = split_link[1].rstrip("]")
                                if attribute_name == 'ObjectName':
                                    modify_objectname_flag = True
                                attribute_element = project.find(f".//attribute[name='{attribute_name}']/value")
                                if attribute_element is not None:
                                    attribute_element.text = new_content
                        elif len(path) == 2:  # 说明修改的是Switchgear层级
                            switchgear_nodes = project.findall("switchgearlist/switchgear")
                            for switchgear in switchgear_nodes:
                                switchgear_name = switchgear.find("attributelist/attribute[name='ObjectName']/value")
                                if switchgear_name is not None and switchgear_name.text == path[1]:
                                    split_link = xml_link.split("[")
                                    if len(split_link) == 2:
                                        attribute_name = split_link[1].rstrip("]")
                                        if attribute_name == 'ObjectName':
                                            index = 1
                                            while new_content in left_same_level_node_text:
                                                new_content = f'{new_content}({index})'
                                                index += 1
                                            modify_objectname_flag = True
                                        attribute_element = switchgear.find(f".//attribute[name='{attribute_name}']/value")
                                        if attribute_element is not None:
                                            attribute_element.text = new_content
                        elif len(path) == 3:  # 说明修改的是Typical/Panel层级
                            switchgear_nodes = project.findall("switchgearlist/switchgear")
                            for switchgear in switchgear_nodes:
                                switchgear_name = switchgear.find("attributelist/attribute[name='ObjectName']/value")
                                if switchgear_name is not None and switchgear_name.text == path[1]:
                                    typical_nodes = switchgear.findall("typicallist/typical")
                                    panel_nodes = switchgear.findall("panellist/panel")
                                    if '<' not in path[2] and '>' not in path[2]:
                                        for typical in typical_nodes:
                                            typical_name = typical.find("attributelist/attribute[name='ObjectName']/value")
                                            if typical_name is not None and typical_name.text == path[2]:
                                                split_link = xml_link.split("[")
                                                if len(split_link) == 2:
                                                    attribute_name = split_link[1].rstrip("]")
                                                    if attribute_name == 'ObjectName':
                                                        index = 1
                                                        while new_content in left_same_level_node_text:
                                                            new_content = f'{new_content}({index})'
                                                            index += 1
                                                        modify_objectname_flag = True
                                                    attribute_element = typical.find(f".//attribute[name='{attribute_name}']/value")
                                                    if attribute_element is not None:
                                                        attribute_element.text = new_content
                                    else:
                                        for panel in panel_nodes:
                                            panel_name = panel.find("attributelist/attribute[name='ObjectName']/value")
                                            if panel_name is not None and panel_name.text == path[2].split(' <')[0]:
                                                split_link = xml_link.split("[")
                                                if len(split_link) == 2:
                                                    attribute_name = split_link[1].rstrip("]")
                                                    if attribute_name == 'ObjectName':
                                                        index = 1
                                                        while new_content in left_same_level_node_text:
                                                            new_content = f'{new_content}({index})'
                                                            index += 1
                                                        modify_objectname_flag = True
                                                    attribute_element = panel.find(f".//attribute[name='{attribute_name}']/value")
                                                    if attribute_element is not None:
                                                        attribute_element.text = new_content


def get_display_length(s):
    return sum(2 if ord(c) > 127 else 1 for c in s)


def export_mve():  # 导出
    try:
        stem, _ = os.path.splitext(os.path.basename(FilePath))
        home_path = os.path.expanduser("~")
        desktop_path = os.path.join(home_path, "Desktop")
        desktop_path = askdirectory(title=u'请选择导出文件夹', initialdir=desktop_path)
        if not desktop_path:
            tk.messagebox.showwarning("提示", "未选择保存路径，导出操作已取消")
            return

        xmltree = xml_root.getroottree()
        xmltree.write(os.path.join(desktop_path, stem + '-revised.mve'))  # 数据存入mve

        tk.messagebox.showwarning("提示", "修改后的.mve文件导出完成")
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())
