import tkinter as tk

from tkinter.filedialog import askdirectory
import os

import time
import datetime
from time import *
import math
import xlrd

from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm, cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from reportlab.lib import colors

from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Image, Table, TableStyle, NextPageTemplate, PageBreak
import traceback
import pandas as pd
from collections import defaultdict
import need.Terminal_Check_Extension as Terminal_Check_Extension

import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

FilePath = ""  # 设置一个地址变量
from need.custom_dialogs import CustomDialog, center_window, Tooltip, image_label

def main(parent, w_rat, h_rat, file_path):
    global w_ratio
    w_ratio = w_rat

    global h_ratio
    h_ratio = h_rat

    global checklog_file_path
    checklog_file_path = file_path
    global open_folder
    open_folder = tk.PhotoImage(file="ico\\open_folder.png")
    global analyze_file
    analyze_file = tk.PhotoImage(file="ico\\read.png")
    global pdf_export
    pdf_export = tk.PhotoImage(file="ico\\export.png")

    tk.Label(parent, text="欢迎使用端子文件检查功能", bg="#c9dbe9", fg="black", height=int(1*h_ratio), font=("ABBvoice CNSG", int(20 * h_ratio), "bold")).pack(fill=tk.X)

    f1 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f1, text='   说明：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')
    tk.Label(f1, text='(1)EPLAN菜单Tools→Reports:Automated processing中选择Export Files V04导出；\n(2)选择端子文件【C:/Temp/项目号-Files/项目号-Terminal list.xlsx】；\n(3)对EPLAN导出的端子文件进行端子号和端子类型检查。', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    f1.pack(fill=tk.X)

    tk.Frame(parent, height=int(20*h_ratio), bg="#eaf1f6").pack(fill=tk.X)    # 水平分割线

    f2 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f2, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2, text='路径：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry    # 为了确保selectpath函数能正确调用entry,将其全局化
    entry = tk.Entry(f2, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)))
    entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

    global button_export_report
    button_export_report = tk.Button(f2, image=pdf_export, text="导出报表", font=("ABBvoice CNSG", int(13 * h_ratio)), bg="#eaf1f6", command=export_report, compound=tk.LEFT, state='disabled', activebackground='blue')
    button_export_report.pack(side=tk.RIGHT, padx=(0, int(20*w_ratio)))

    global button_check
    button_check = tk.Button(f2, text='检查', bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), image=analyze_file, compound=tk.LEFT, command=process, activebackground='blue')
    button_check.pack(side=tk.RIGHT, padx=(0, int(20 * w_ratio)))
    button_check['state'] = 'disabled'

    tk.Button(f2, text='选择', bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), image=open_folder, compound=tk.LEFT, command=selectpath, activebackground='blue').pack(side=tk.RIGHT, padx=int(20*w_ratio))

    f2.pack(fill=tk.X)

    tk.Frame(parent, height=int(20*h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f3 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f3, text='   结果：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')

    global text
    text = tk.Text(f3, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), height=int(25 * h_ratio), width=int(65*w_ratio))
    text.pack(side=tk.LEFT, padx=(0, 1), pady=0, fill=tk.BOTH, expand=True)

    text.tag_configure('error', foreground='red')    # 设置tag

    scrollbar = tk.Scrollbar(f3)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, int(20*w_ratio)))
    scrollbar.config(command=text.yview)
    text.config(yscrollcommand=scrollbar.set)
    f3.pack(fill=tk.BOTH, expand=True)

    tk.Frame(parent, height=int(20*h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线


def selectpath():
    filepath = tk.filedialog.askopenfilename(initialdir='C:/Temp/', title=u'请选择文件(项目号-Terminal list.xlsx)', filetypes=[("Excel", ".xlsx")])  # 选择打开什么文件，返回文件名
    if len(filepath) != 0:
        string_filename = ""
        for i in range(0, len(filepath)):
            string_filename += str(filepath[i])
        button_check['state'] = 'normal'
    else:
        button_check['state'] = 'disabled'
        button_export_report['state'] = 'disabled'

    text.delete(1.0, tk.END)  # 清空输出结果框
    entry.delete(0, "end")  # 删除entry原始内容
    entry.insert(0, filepath)  # 重新填入地址
    button_export_report['state'] = 'disabled'
    global FilePath
    FilePath = filepath


def process():
    try:
        error1_calculator = 0
        error2_calculator = 0
        global error3_calculator
        error3_calculator = 0
        error4_calculator = 0
        error5_calculator = 0
        error6_calculator = 0

        text.delete(1.0, tk.END)  # 清空输出结果框
        if FilePath == "":
            tk.messagebox.showwarning("提示", "请选择文件！")

        stem, suffix = os.path.splitext(os.path.basename(FilePath))  # stem是文件名,suffix是后缀
        if '-Terminal list' not in stem:
            tk.messagebox.showwarning("提示", "请选择项目号-Terminal list.xlsx文件！")

        else:
            text.insert(tk.INSERT, '>>>端子号正在检查中...\n')  # 进行端子号检查

            if not os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\ABB_FNKS_terminal list.xlsx"):
                tk.messagebox.showwarning("提示", "失败,找不到端子类型表!")
            else:
                start = time()

                workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\ABB_FNKS_terminal list.xlsx")  # 加载ABB_FNKS端子对照表
                worksheet = workbook['TE']
                worksheet1 = workbook['WDML']
                worksheet2 = workbook['RL']
                worksheet3 = workbook['PHOENIX']

                # print("正在处理中...")
                book = load_workbook(FilePath)
                sheet = book['Z1_xlsx']
                A = []
                B = []
                C = []
                D = []
                E = []
                F = []
                G = []

                for i in range(3, sheet.max_row + 1):
                    A.append(str(sheet.cell(row=i, column=1).value))    # Typical列
                    B.append(str(sheet.cell(row=i, column=2).value))    # 左侧Remark
                    C.append(str(sheet.cell(row=i, column=3).value))    # 左侧接线，用来判断是否出现一个端子接两根线(等号数量>3)
                    D.append(str(sheet.cell(row=i, column=4).value))    # Terminal列
                    E.append(str(sheet.cell(row=i, column=5).value))    # Type列
                    F.append(str(sheet.cell(row=i, column=6).value))    # 右侧接线
                    G.append(str(sheet.cell(row=i, column=7).value))    # 右侧Remark

                Typical = []  # 由于部分数据行可能出现端子号异常，另外构造列表，存放正常数据
                Terminal = []
                Type = []
                External = []
                Internal = []
                E_Remark = []
                I_Remark = []

                Terminal_X = []
                Terminal_n = []

                for i in range(0, len(A)):
                    if "RAD" in D[i]:
                        text.insert(tk.INSERT, '▲ ' + '>>>发现二极管端子,忽略二极管端子,请自查...\n', 'error')

                    if ":" in D[i] and "X" in D[i]:  # 处理端子号异常情况,忽略二极管端子和不带:的端子
                        Typical.append(A[i])
                        Terminal.append(D[i])
                        Type.append(E[i])
                        External.append(C[i])
                        Internal.append(F[i])
                        E_Remark.append(B[i])
                        I_Remark.append(G[i])

                        Terminal_X.append(str(D[i].split(":")[0]))  # X*端子排
                        if 'A' not in str(D[i].split(":")[1]) and 'B' not in str(D[i].split(":")[1]) and 'C' not in str(D[i].split(":")[1]) and 'D' not in str(D[i].split(":")[1]):
                            Terminal_n.append(int(D[i].split(":")[1]))  # 数值型端子号
                        else:
                            Terminal_n.append(str(D[i].split(":")[1]))  # 数值+字符型端子号

                # 2.1版本新增Multi_Wiring_Terminal获取，而不是手动指定，以避免后续新增物料导致需要手动变更Multi_Wiring_Terminal
                Multi_Wiring_Terminal = []
                for i in range(2, worksheet.max_row + 1):
                    if str(worksheet.cell(row=i, column=1).value) == '0':
                        Multi_Wiring_Terminal.append(str(worksheet.cell(row=i, column=2).value))
                for i in range(2, worksheet1.max_row + 1):
                    if str(worksheet1.cell(row=i, column=1).value) == '0':
                        Multi_Wiring_Terminal.append(str(worksheet1.cell(row=i, column=2).value))
                for i in range(2, worksheet2.max_row + 1):
                    if str(worksheet2.cell(row=i, column=1).value) == '0':
                        Multi_Wiring_Terminal.append(str(worksheet2.cell(row=i, column=2).value))
                for i in range(2, worksheet3.max_row + 1):
                    if str(worksheet3.cell(row=i, column=1).value) == '0':
                        Multi_Wiring_Terminal.append(str(worksheet3.cell(row=i, column=2).value))

                # Multi_Wiring_Terminal = ['M4/6.D1', 'A4C 2.5', 'AEP 4C 2.5', 'A4C 4', 'AEP 4C 4', 'A3C 6', 'AEP 3C 6', 'UT4-QUATT/2P', 'UK3-MSTB-5.08', 'UT6-QUATTRO/2P', 'PT4-QUATTRO/2P',
                #                          'PT2.5-QUATTRO/2P', 'PT6-QUATTRO/2P', 'PT 2.5-QUATTRO', 'PT 4-QUATTRO', 'PT 6-QUATTRO', 'ST4-QUATTRO/2P', 'TB4-TWIN-I', 'UDK4']

                # V1.8升版：将所有的信息，按照Typical分开
                for i in range(0, len(Typical)):
                    Typical_set = list(set(Typical))  # set去重后顺序会改变
                    Typical_set.sort(key=list(Typical).index)  # 保证去重后顺序不变
                # print(Typical_set)    # ['I1', 'FT1', 'P1', 'I2', 'P2', 'M2', 'FT4', 'FT5']

                Typical_first_index = []
                for i in range(0, len(Typical_set)):
                    Typical_first_index.append(Typical.index(Typical_set[i]))    # 记录下Typical首次出现对应的位置，以确保能够按Typical拆分端子信息
                # print(Typical_first_index)    # [0, 198, 384, 508, 699, 823, 895, 915]
                # I1对应0-197，FT1对应198-383，P1对应384-507，I2对应508-698，P2对应699-822，M2对应823-894，FT4对应895-914，FT5对应915-(len(Typical)-1)

                for i in range(0, len(Typical_first_index)):    # (0,1,2,3,4,5,6,7)   (0-(len(Typical_first_index)-1))
                    if i != len(Typical_first_index)-1:
                        for j in range(int(Typical_first_index[i]), int(Typical_first_index[i+1])):
                            if 'A' in str(Terminal_n[j]) or 'B' in str(Terminal_n[j]) or 'C' in str(Terminal_n[j]) or 'D' in str(Terminal_n[j]):  # 端子号为数字+字符型的端子
                                if str(Internal[j]) == 'None':
                                    text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "对应航空插接线漏引出【错误】\n", 'error')
                                    error1_calculator += 1
                                else:
                                    length_ = len(str(Terminal_n[j]))
                                    if str(Internal[j][-length_:]).upper() != Terminal_n[j]:
                                        text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "端子号与航空插端子号不一致【错误】\n", 'error')
                                        error1_calculator += 1

                            elif Terminal_X[j] in ['X20', 'X30', 'X31', 'X40']:
                                if str(Internal[j]) == 'None':
                                    text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "对应航空插接线漏引出【错误】\n", 'error')
                                    error1_calculator += 1
                                else:
                                    length_ = len(str(Terminal_n[j]))
                                    if str(Internal[j][-length_:]).upper() != str(Terminal_n[j]):
                                        text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "端子号与航空插端子号不一致【错误】\n", 'error')
                                        error1_calculator += 1
                            else:
                                if str(External[j]).count('=') > 3 and str(Type[j]) not in Multi_Wiring_Terminal and (int(str(External[j]).count('=') / 3) != str(E_Remark[j]).count('C')) and (int(str(External[j]).count('=') / 3) != str(E_Remark[j]).count('A')+1):
                                    text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "误接两/多根线【错误】\n", 'error')
                                    error1_calculator += 1

                                if str(Internal[j]).count('=') > 3 and str(Type[j]) not in Multi_Wiring_Terminal and (int(str(Internal[j]).count('=') / 3) != str(I_Remark[j]).count('C')) and (int(str(Internal[j]).count('=') / 3) != str(I_Remark[j]).count('A')+1):
                                    text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "误接两/多根线【错误】\n", 'error')
                                    error1_calculator += 1

                                if j == int(Typical_first_index[i]):  # 每个Typcial的第一个端子
                                    if Terminal_n[j] != 1:
                                        text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "应为" + Terminal_X[j] + ":" + "1" + "【错误】\n", 'error')
                                        error1_calculator += 1

                                else:  # 非首端子
                                    if Terminal_X[j] == Terminal_X[j - 1]:  # 属于同一个端子排时
                                        if not (type(Terminal_n[j]) == type(Terminal_n[j - 1])):
                                            text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j - 1]) + "后的" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "【错误】\n", 'error')
                                            error1_calculator += 1

                                        else:
                                            if Terminal_n[j] - 1 != Terminal_n[j-1]:    # V1.8升版，原来是跟每个端子排的第一个比较，现在是相邻端子比较，编号是否相差1
                                                if Terminal_n[j] - Terminal_n[j - 1] < 0:
                                                    text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j - 1]) + "后的" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "【乱序】\n", 'error')
                                                    error1_calculator += 1

                                                elif Terminal_n[j] - Terminal_n[j - 1] == 0:
                                                    text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j - 1]) + "后的" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "【重复】\n", 'error')
                                                    error1_calculator += 1

                                                else:
                                                    text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j - 1]) + "后的" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "【跳号】\n", 'error')
                                                    error1_calculator += 1

                                    else:    # 新的端子排号时，第一个端子
                                        if Terminal_n[j] != 1:
                                            text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "应为" + Terminal_X[j] + ":" + "1" + "【错误】\n", 'error')
                                            error1_calculator += 1

                    if i == len(Typical_first_index)-1:
                        for j in range(int(Typical_first_index[i]), len(Typical)):
                            if 'A' in str(Terminal_n[j]) or 'B' in str(Terminal_n[j]) or 'C' in str(Terminal_n[j]) or 'D' in str(Terminal_n[j]):  # 端子号为数字+字符型的端子
                                if str(Internal[j]) == 'None':
                                    text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "对应航空插接线漏引出【错误】\n", 'error')
                                    error1_calculator += 1

                                else:
                                    length_ = len(str(Terminal_n[j]))
                                    if str(Internal[j][-length_:]).upper() != Terminal_n[j]:
                                        text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "端子号与航空插端子号不一致【错误】\n", 'error')
                                        error1_calculator += 1

                            elif Terminal_X[j] in ['X20', 'X30', 'X31', 'X40']:
                                if str(Internal[j]) == 'None':
                                    text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "对应航空插接线漏引出【错误】\n", 'error')
                                    error1_calculator += 1
                                else:
                                    length_ = len(str(Terminal_n[j]))
                                    if str(Internal[j][-length_:]).upper() != str(Terminal_n[j]):
                                        text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "端子号与航空插端子号不一致【错误】\n", 'error')
                                        error1_calculator += 1

                            else:
                                if str(External[j]).count('=') > 3 and str(Type[j]) not in Multi_Wiring_Terminal and (int(str(External[j]).count('=') / 3) != str(E_Remark[j]).count('C')) and (int(str(External[j]).count('=') / 3) != str(E_Remark[j]).count('A')+1):
                                    text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "误接两/多根线【错误】\n", 'error')
                                    error1_calculator += 1

                                if str(Internal[j]).count('=') > 3 and str(Type[j]) not in Multi_Wiring_Terminal and (int(str(Internal[j]).count('=') / 3) != str(I_Remark[j]).count('C')) and (int(str(Internal[j]).count('=') / 3) != str(I_Remark[j]).count('A')+1):
                                    text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "误接两/多根线【错误】\n", 'error')
                                    error1_calculator += 1

                                if j == int(Typical_first_index[i]):  # 每个Typcial的第一个端子
                                    if Terminal_n[j] != 1:
                                        text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "应为" + Terminal_X[j] + ":" + "1" + "【错误】\n", 'error')
                                        error1_calculator += 1

                                else:  # 非首端子
                                    if Terminal_X[j] == Terminal_X[j - 1]:  # 属于同一个端子排时
                                        if not (type(Terminal_n[j]) == type(Terminal_n[j - 1])):
                                            text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j - 1]) + "后的" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "【错误】\n", 'error')
                                            error1_calculator += 1

                                        else:
                                            if Terminal_n[j] - 1 != Terminal_n[j-1]:    # V1.8升版，原来是跟每个端子排的第一个比较，现在是相邻端子比较，编号是否相差1
                                                if Terminal_n[j] - Terminal_n[j - 1] < 0:
                                                    text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j - 1]) + "后的" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "【乱序】\n", 'error')
                                                    error1_calculator += 1

                                                elif Terminal_n[j] - Terminal_n[j - 1] == 0:
                                                    text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j - 1]) + "后的" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "【重复】\n", 'error')
                                                    error1_calculator += 1

                                                else:
                                                    text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j - 1]) + "后的" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "【跳号】\n", 'error')
                                                    error1_calculator += 1

                                    else:  # 新的端子排号时，第一个端子
                                        if Terminal_n[j] != 1:
                                            text.insert(tk.INSERT, '▲ ' +  Typical[j] + "——" + Terminal_X[j] + ":" + str(Terminal_n[j]) + "应为" + Terminal_X[j] + ":" + "1" + "【错误】\n", 'error')
                                            error1_calculator += 1

                end = time()
                text.insert(tk.INSERT, ">>>端子号检查完成!  用时%.3f秒\n" % (end - start))

                text.insert(tk.INSERT, '\n>>>端子类型正在检查中...\n')  # 进行端子厂家检查


                start = time()

                index1 = 0  # 端子排的头索引
                index2 = 0  # 端子排的尾索引

                number1 = 0  # 某一端子排内TE端子数量
                number2 = 0  # 某一端子排内WDML端子数量
                number3 = 0  # 某一端子排内RL端子数量
                number4 = 0  # 某一端子排内PHOENIX端子数量

                number11 = 0  # TE端子排数量
                number22 = 0  # WDML端子排数量
                number33 = 0  # RL端子排数量
                number44 = 0  # PHOENIX端子排数量

                index111 = []  # TE端子所在位置
                index222 = []  # WDML端子所在位置
                index333 = []  # RL端子所在位置
                index444 = []  # PHOENIX端子所在位置

                for i in range(0, len(Typical_first_index)):  # (0,1,2,3,4,5,6,7)
                    if i != len(Typical_first_index) - 1:
                        for j in range(int(Typical_first_index[i]), int(Typical_first_index[i + 1])):
                            if Terminal_X[j] != Terminal_X[j+1]:
                                if Terminal_X[j] != Terminal_X[j+1]:
                                    index1 = index2
                                    index2 = j
                                # print(index2)    # 这个输出可以检查是否遍历到最后一个端子排的最后一个端子
                                for k in range(index1, index2 + 1):  # 遍历端子排中的每个端子
                                    for l in range(2, worksheet.max_row + 1):  # 遍历端子类型表
                                        if Type[k] == worksheet.cell(row=l, column=2).value:  # 端子属于TE端子
                                            number1 += 1
                                            continue
                                    for l in range(2, worksheet1.max_row + 1):
                                        if Type[k] == worksheet1.cell(row=l, column=2).value:  # 端子属于WDML端子
                                            number2 += 1
                                            continue
                                    for l in range(2, worksheet2.max_row + 1):
                                        if Type[k] == worksheet2.cell(row=l, column=2).value:  # 端子属于RL端子
                                            number3 += 1
                                            continue
                                    for l in range(2, worksheet3.max_row + 1):
                                        if Type[k] == worksheet3.cell(row=l, column=2).value:  # 端子属于PHOENIX端子
                                            number4 += 1
                                            continue
                                if number1 != 0 and number2 == 0 and number3 == 0 and number4 == 0:  # 此端子排属于TE端子
                                    number11 += 1
                                    index111.append(j)
                                elif number1 == 0 and number2 != 0 and number3 == 0 and number4 == 0:  # 此端子排属于WDML端子
                                    number22 += 1
                                    index222.append(j)
                                elif number1 == 0 and number2 == 0 and number3 != 0 and number4 == 0:  # 此端子排属于RL端子
                                    number33 += 1
                                    index333.append(j)
                                elif number1 == 0 and number2 == 0 and number3 == 0 and number4 != 0:  # 此端子排属于PHOENIX端子
                                    number44 += 1
                                    index444.append(j)
                                elif number1 == 0 and number2 == 0 and number3 == 0 and number4 == 0:  # 此端子排不在端子类型报表中
                                    text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + "端子型号不在端子类型表中,无法识别\n", 'error')
                                else:
                                    text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + "端子出现不同厂家\n", 'error')
                                    error2_calculator += 1
                                number1 = 0    # 每个Typical的每个端子排统计完成要置零一下
                                number2 = 0
                                number3 = 0
                                number4 = 0
                                index2 += 1

                        terminal_x_TE_list = []
                        terminal_x_WDML_list = []
                        terminal_x_RL_list = []
                        terminal_x_PHOENIX_list = []

                        # 处理第一个循环
                        for j in range(0, number11):
                            terminal_x_TE_list.append(Terminal_X[index111[j]])

                        # 处理第二个循环
                        for j in range(0, number22):
                            terminal_x_WDML_list.append(Terminal_X[index222[j]])

                        # 处理第三个循环
                        for j in range(0, number33):
                            terminal_x_RL_list.append(Terminal_X[index333[j]])

                        # 处理第四个循环
                        for j in range(0, number44):
                            terminal_x_PHOENIX_list.append(Terminal_X[index444[j]])

                        # 使用 join 方法拼接列表，并在每个拼接字符串后添加标识
                        output_TE = ",".join(terminal_x_TE_list) + "为TE端子,  " if terminal_x_TE_list else ""
                        output_WDML = ",".join(terminal_x_WDML_list) + "为魏德米勒端子,  " if terminal_x_WDML_list else ""
                        output_RL = ",".join(terminal_x_RL_list) + "为瑞联端子,  " if terminal_x_RL_list else ""
                        output_PHOENIX = ",".join(terminal_x_PHOENIX_list) + "为菲尼克斯端子,  " if terminal_x_PHOENIX_list else ""

                        if len(index111) != 0:
                            output_typical = Typical[index111[0]]
                        elif len(index222) != 0:
                            output_typical = Typical[index222[0]]
                        elif len(index333) != 0:
                            output_typical = Typical[index333[0]]
                        else:
                            output_typical = Typical[index444[0]]

                        # 将所有内容插入到 text 小部件
                        output_text = output_typical + '——' + output_TE + output_WDML + output_RL + output_PHOENIX
                        output_text = output_text.rsplit(',  ', 1)[0] + '\n'
                        text.insert(tk.INSERT, output_text)

                        number11 = 0  # TE端子排数量    # 每个Typcial统计完成，要置零一下
                        number22 = 0  # WDML端子排数量
                        number33 = 0  # RL端子排数量
                        number44 = 0  # PHOENIX端子排数量

                        index111 = []  # TE端子所在位置
                        index222 = []  # WDML端子所在位置
                        index333 = []  # RL端子所在位置
                        index444 = []  # PHOENIX端子所在位置

                    if i == len(Typical_first_index) - 1:
                        for j in range(int(Typical_first_index[i]), len(Typical)-1):
                            if Terminal_X[j] != Terminal_X[j+1] or j == len(Typical) - 2:
                                if Terminal_X[j] != Terminal_X[j+1]:  # 这种会忽略最后一个端子排，因为最后一个端子排的最后一个端子下面没有端子
                                    index1 = index2
                                    index2 = j
                                if j == len(Typical) - 2:  # 将最后一个端子排考虑在内
                                    index1 = index2
                                    index2 = j + 1
                                # print('index2=', index2)    # 这个输出可以检查是否遍历到最后一个端子排的最后一个端子
                                for k in range(index1, index2 + 1):  # 遍历端子排中的每个端子
                                    for l in range(2, worksheet.max_row + 1):  # 遍历端子类型表
                                        if Type[k] == worksheet.cell(row=l, column=2).value:  # 端子属于TE端子
                                            number1 += 1
                                            continue
                                    for l in range(2, worksheet1.max_row + 1):
                                        if Type[k] == worksheet1.cell(row=l, column=2).value:  # 端子属于WDML端子
                                            number2 += 1
                                            continue
                                    for l in range(2, worksheet2.max_row + 1):
                                        if Type[k] == worksheet2.cell(row=l, column=2).value:  # 端子属于RL端子
                                            number3 += 1
                                            continue
                                    for l in range(2, worksheet3.max_row + 1):
                                        if Type[k] == worksheet3.cell(row=l, column=2).value:  # 端子属于PHOENIX端子
                                            number4 += 1
                                            continue
                                if number1 != 0 and number2 == 0 and number3 == 0 and number4 == 0:  # 此端子排属于TE端子
                                    number11 += 1
                                    index111.append(j)
                                elif number1 == 0 and number2 != 0 and number3 == 0 and number4 == 0:  # 此端子排属于WDML端子
                                    number22 += 1
                                    index222.append(j)
                                elif number1 == 0 and number2 == 0 and number3 != 0 and number4 == 0:  # 此端子排属于RL端子
                                    number33 += 1
                                    index333.append(j)
                                elif number1 == 0 and number2 == 0 and number3 == 0 and number4 != 0:  # 此端子排属于PHOENIX端子
                                    number44 += 1
                                    index444.append(j)
                                elif number1 == 0 and number2 == 0 and number3 == 0 and number4 == 0:  # 此端子排不在端子类型报表中
                                    text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + "端子型号不在端子类型表中,无法识别\n", 'error')
                                else:
                                    text.insert(tk.INSERT, '▲ ' + Typical[j] + "——" + Terminal_X[j] + "端子出现不同厂家\n", 'error')
                                    error2_calculator += 1

                                number1 = 0    # 每个Typical的每个端子排统计完成要置零一下
                                number2 = 0
                                number3 = 0
                                number4 = 0
                                index2 += 1

                        terminal_x_TE_list = []
                        terminal_x_WDML_list = []
                        terminal_x_RL_list = []
                        terminal_x_PHOENIX_list = []

                        # 处理第一个循环
                        for j in range(0, number11):
                            terminal_x_TE_list.append(Terminal_X[index111[j]])

                        # 处理第二个循环
                        for j in range(0, number22):
                            terminal_x_WDML_list.append(Terminal_X[index222[j]])

                        # 处理第三个循环
                        for j in range(0, number33):
                            terminal_x_RL_list.append(Terminal_X[index333[j]])

                        # 处理第四个循环
                        for j in range(0, number44):
                            terminal_x_PHOENIX_list.append(Terminal_X[index444[j]])

                        # 使用 join 方法拼接列表，并在每个拼接字符串后添加标识
                        output_TE = ",".join(terminal_x_TE_list) + "为TE端子,  " if terminal_x_TE_list else ""
                        output_WDML = ",".join(terminal_x_WDML_list) + "为魏德米勒端子,  " if terminal_x_WDML_list else ""
                        output_RL = ",".join(terminal_x_RL_list) + "为瑞联端子,  " if terminal_x_RL_list else ""
                        output_PHOENIX = ",".join(terminal_x_PHOENIX_list) + "为菲尼克斯端子,  " if terminal_x_PHOENIX_list else ""

                        if len(index111) != 0:
                            output_typical = Typical[index111[0]]
                        elif len(index222) != 0:
                            output_typical = Typical[index222[0]]
                        elif len(index333) != 0:
                            output_typical = Typical[index333[0]]
                        else:
                            output_typical = Typical[index444[0]]

                        # 将所有内容插入到 text 小部件
                        output_text = output_typical + '——' + output_TE + output_WDML + output_RL + output_PHOENIX
                        output_text = output_text.rsplit(',  ', 1)[0] + '\n'
                        text.insert(tk.INSERT, output_text)

                        number11 = 0  # TE端子排数量    # 每个Typcial统计完成，要置零一下
                        number22 = 0  # WDML端子排数量
                        number33 = 0  # RL端子排数量
                        number44 = 0  # PHOENIX端子排数量

                        index111 = []  # TE端子所在位置
                        index222 = []  # WDML端子所在位置
                        index333 = []  # RL端子所在位置
                        index444 = []  # PHOENIX端子所在位置

                end = time()
                text.insert(tk.INSERT, ">>>端子类型检查完成!  用时%.3f秒\n" % (end - start))

                
                start = time()

                smallbusbar = []
                # 记录所有识别到的 X 端子类型
                detected_X = []

                for i in range(0, len(Typical_first_index)):    # (0,1,2,3,4,5,6,7)   (0-(len(Typical_first_index)-1))
                    if i != len(Typical_first_index)-1:
                        count_X11 = 0
                        count_X12 = 0
                        count_X13 = 0
                        for j in range(int(Typical_first_index[i]), int(Typical_first_index[i+1])):
                            real_Typical = Typical[j]
                            if Terminal_X[j] == 'X11':
                                count_X11 += 1
                            elif Terminal_X[j] == 'X12':
                                count_X12 += 1
                            elif Terminal_X[j] == 'X13':
                                count_X13 += 1
                        smallbusbar.append((real_Typical, count_X11, count_X12, count_X13))

                        # 记录识别到的 X 端子类型
                        if count_X11 > 0 and 'X11' not in detected_X:
                            detected_X.append('X11')
                        if count_X12 > 0 and 'X12' not in detected_X:
                            detected_X.append('X12')
                        if count_X13 > 0 and 'X13' not in detected_X:
                            detected_X.append('X13')

                    if i == len(Typical_first_index)-1:
                        count_X11 = 0
                        count_X12 = 0
                        count_X13 = 0
                        for j in range(int(Typical_first_index[i]), len(Typical)):
                            real_Typical = Typical[j]
                            if Terminal_X[j] == 'X11':
                                count_X11 += 1
                            elif Terminal_X[j] == 'X12':
                                count_X12 += 1
                            elif Terminal_X[j] == 'X13':
                                count_X13 += 1
                        smallbusbar.append((real_Typical, count_X11, count_X12, count_X13))

                        # 记录识别到的 X 端子类型
                        if count_X11 > 0 and 'X11' not in detected_X:
                            detected_X.append('X11')
                        if count_X12 > 0 and 'X12' not in detected_X:
                            detected_X.append('X12')
                        if count_X13 > 0 and 'X13' not in detected_X:
                            detected_X.append('X13')

                if detected_X:
                    detected_text = '/'.join(detected_X)
                    text.insert(tk.INSERT, f'\n>>>小母线端子(识别到{detected_text})一致性正在检查中...\n')  # 进行小母线端子检查
                else:
                    text.insert(tk.INSERT, '\n>>>小母线端子一致性正在检查中...\n')  # 进行小母线端子检查

                check_terminal_counts(smallbusbar)

                end = time()
                # 生成识别到的 X 端子文本

                text.insert(tk.INSERT, ">>>小母线端子一致性检查完成!  用时%.3f秒\n" % (end - start))


                text.insert(tk.INSERT, '\n>>>端子内外侧接线颠倒正在检查中...\n')  # 进行端子内外侧接线颠倒检查
                start = time()

                external_internal_error = []
                for i in range(len(Typical_first_index)):
                    if i != len(Typical_first_index) - 1:
                        start1, end1 = int(Typical_first_index[i]), int(Typical_first_index[i + 1])
                    else:
                        start1, end1 = int(Typical_first_index[i]), len(Typical)

                    for j in range(start1, end1):
                        if Terminal_X[j] != 'X4' and 'XB' not in Terminal[j]:
                            status_flag = check_terminal_external_internal(
                                E_Remark[j], I_Remark[j], Internal[j], External[j])

                            if not status_flag:
                                external_internal_error.append((Typical[j], Terminal[j]))
                                error4_calculator += 1

                grouped = defaultdict(list)
                for typ, term in external_internal_error:
                    grouped[typ].append(term)
                external_internal_error = [f"{k}——{','.join(v)}" for k, v in grouped.items()]

                for output_external_internal in external_internal_error:
                    text.insert(tk.INSERT, '▲ ' + output_external_internal+ "端子内外侧接线颠倒\n", 'error')

                end = time()
                text.insert(tk.INSERT, ">>>端子内外侧接线颠倒检查完成!  用时%.3f秒\n" % (end - start))

                base_name = os.path.basename(FilePath)  # 获取路径的文件名（包括后缀）
                stem, suffix = os.path.splitext(base_name)  # stem是文件名,suffix是后缀

                new_stem = stem.replace('-Terminal list', '')
                pdf_file_path = os.path.join(os.path.dirname(FilePath), new_stem + '.pdf')
                pdf_file_output_path = os.path.join(os.path.dirname(FilePath), new_stem + '-annotated.pdf')

                if not os.path.exists(pdf_file_path):
                    tk.messagebox.showwarning("提示", "文件夹下缺少XXX.pdf文件，无法进行短接片跨端子检查！")
                else:
                    start = time()
                    text.insert(tk.INSERT, '\n>>>端子短接片跨端子正在检查中...\n')  # 进行端子短接片跨端子检查

                    td_list = Terminal_Check_Extension.extract_terminal_diagram_from_pdf(pdf_file_path)

                    result = Terminal_Check_Extension.check_separator_plate(pdf_file_path, td_list, pdf_file_output_path)
                    short_result = result['short']
                    missing_result = result['missing']

                    for res in short_result:
                        if res['strip'] not in ['X4']:
                            short_across_terminal = False
                            for i in range(0, len(Typical_first_index)):  # (0,1,2,3,4,5,6,7)   (0-(len(Typical_first_index)-1))
                                if i != len(Typical_first_index) - 1:
                                    for j in range(int(Typical_first_index[i]), int(Typical_first_index[i + 1])):
                                        if res['typical'] == str(Typical[j]) and res['strip'] == str(Terminal_X[j]) and res['term_pair'].split('-')[0] == str(Terminal_n[j]) and res['term_pair'].split('-')[1] != str(Terminal_n[j+1]):
                                            short_across_terminal = True
                                            break
                                if i == len(Typical_first_index) - 1:
                                    for j in range(int(Typical_first_index[i]), len(Typical)):
                                        if res['typical'] == str(Typical[j]) and res['strip'] == str(Terminal_X[j]) and res['term_pair'].split('-')[0] == str(Terminal_n[j]) and res['term_pair'].split('-')[1] != str(Terminal_n[j+1]):
                                            short_across_terminal = True
                                            break
                                if short_across_terminal:
                                    error5_calculator += 1
                                    text.insert(tk.INSERT, '▲ ' + res['typical'] + '——' + res['strip'] + ':'+ res['term_pair'] +"短接片跨端子【错误】\n", 'error')
                                    break

                    end = time()
                    text.insert(tk.INSERT, ">>>端子短接片跨端子检查完成!  用时%.3f秒\n" % (end - start))

                    start = time()
                    text.insert(tk.INSERT, '\n>>>端子隔板缺失正在检查中...\n')  # 进行端子隔板缺失检查

                    Ordinary_Terminal_List = []
                    for i in range(2, worksheet.max_row + 1):
                        if str(worksheet.cell(row=i, column=3).value) == '普通端子':
                            Ordinary_Terminal_List.append(str(worksheet.cell(row=i, column=2).value))
                    for i in range(2, worksheet1.max_row + 1):
                        if str(worksheet1.cell(row=i, column=3).value) == '普通端子':
                            Ordinary_Terminal_List.append(str(worksheet1.cell(row=i, column=2).value))
                    for i in range(2, worksheet2.max_row + 1):
                        if str(worksheet2.cell(row=i, column=3).value) == '普通端子':
                            Ordinary_Terminal_List.append(str(worksheet2.cell(row=i, column=2).value))
                    for i in range(2, worksheet3.max_row + 1):
                        if str(worksheet3.cell(row=i, column=3).value) == '普通端子':
                            Ordinary_Terminal_List.append(str(worksheet3.cell(row=i, column=2).value))

                    # print(Ordinary_Terminal_List)
                    for res in missing_result:
                        if res['strip'] not in ['X4']:
                            terminal_lack_separator_plate = False
                            for i in range(0, len(Typical_first_index)):  # (0,1,2,3,4,5,6,7)   (0-(len(Typical_first_index)-1))
                                if i != len(Typical_first_index) - 1:
                                    for j in range(int(Typical_first_index[i]), int(Typical_first_index[i + 1])):
                                        if res['typical'] == str(Typical[j]) and res['strip'] == str(Terminal_X[j]) and res['term_pair'] == str(Terminal_n[j]):
                                            if Type[j] in Ordinary_Terminal_List:
                                                if Terminal_X[j] == Terminal_X[j+1] and Type[j] == Type[j+1]:  # 与后续端子是同一个端子排，且是同型号
                                                    terminal_lack_separator_plate = True
                                                    break
                                                if Terminal_X[j] != Terminal_X[j+1] :  #如果是端子排最后一个端子
                                                    terminal_lack_separator_plate = True
                                                    break
                                if i == len(Typical_first_index) - 1:
                                    for j in range(int(Typical_first_index[i]), len(Typical)):
                                        if res['typical'] == str(Typical[j]) and res['strip'] == str(Terminal_X[j]) and res['term_pair'] == str(Terminal_n[j]):
                                            if Type[j] in Ordinary_Terminal_List:
                                                if j != len(Typical) - 1:
                                                    if Terminal_X[j] == Terminal_X[j+1] and Type[j] == Type[j+1]:  # 与后续端子是同一个端子排，且是同型号
                                                        terminal_lack_separator_plate = True
                                                        break
                                                    if Terminal_X[j] != Terminal_X[j+1] :  #如果是端子排最后一个端子
                                                        terminal_lack_separator_plate = True
                                                        break
                                                else:
                                                    terminal_lack_separator_plate = True

                                if terminal_lack_separator_plate:
                                    error6_calculator += 1
                                    text.insert(tk.INSERT, '▲ ' + res['typical'] + '——' + res['strip'] + ':' + res['term_pair'] + "隔板缺失【错误】\n", 'error')
                                    break

                    end = time()
                    text.insert(tk.INSERT, ">>>端子隔板缺失检查完成!  用时%.3f秒\n" % (end - start))


                button_export_report['state'] = 'normal'

                global checklog_file_path
                checklogbook = load_workbook(checklog_file_path)
                checklogsheet = checklogbook['Sheet']
                project_no = stem.replace('-Terminal list', '')
                current_time = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
                data1 = [project_no, '端子检查-端子号检查', error1_calculator, current_time]    # 项目号
                data2 = [project_no, '端子检查-端子型号检查', error2_calculator, current_time]
                data3 = [project_no, '端子检查-小母线端子一致性检查', error3_calculator, current_time]
                data4 = [project_no, '端子检查-端子内外侧接线颠倒检查', error4_calculator, current_time]
                data5 = [project_no, '端子检查-端子短接片跨端子检查', error5_calculator, current_time]
                data6 = [project_no, '端子检查-端子隔板缺失检查', error6_calculator, current_time]
                checklogsheet.append(data1)
                checklogsheet.append(data2)
                checklogsheet.append(data3)
                checklogsheet.append(data4)
                checklogsheet.append(data5)
                checklogsheet.append(data6)
                checklogbook.save(checklog_file_path)

    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def check_terminal_counts(lst):
    # 初始化字典来存储端子的数量信息
    terminal_counts = {'X11': {}, 'X12': {}, 'X13': {}}

    # 遍历列表，收集每种端子在每个柜号的数量
    for cabinet, x11_count, x12_count, x13_count in lst:
        terminal_counts['X11'].setdefault(x11_count, []).append(cabinet)
        terminal_counts['X12'].setdefault(x12_count, []).append(cabinet)
        terminal_counts['X13'].setdefault(x13_count, []).append(cabinet)

    # 统计不一致的情况
    inconsistent_counts = sum(1 for counts in terminal_counts.values() if len(counts) > 1)

    # 报告不一致的端子数量
    for i, (terminal, counts) in enumerate(terminal_counts.items()):
        if len(counts) > 1:  # 有不一致的情况
            text.insert(tk.INSERT, '▲ ' + "%s端子有数量差异：\n" % terminal, 'error')
            global error3_calculator
            error3_calculator += 1
            for count, typical in counts.items():
                typical_str = ', '.join(typical)
                text.insert(tk.INSERT, '▲ ' + "端子数量为%d的柜型有：%s\n" % (count, typical_str), 'error')
            if i < inconsistent_counts - 1:
                text.insert(tk.INSERT, "\n")

def _is_empty(v) -> bool:
    """NaN、None、'' 都算空"""
    if v is None or v == 'None':
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    return str(v).strip() == ''

def check_terminal_external_internal(b_val, g_val, f_val, c_val):
    b = str(b_val).strip()
    b = '' if _is_empty(b_val) else b

    g = str(g_val).strip()
    g = '' if _is_empty(g_val) else g

    f = str(f_val).strip()
    f = '' if _is_empty(f_val) else f

    c = str(c_val).strip()
    c = '' if _is_empty(c_val) else c
    # print(b_val, c_val, f_val, g_val, b, c, f, g)

    # ---------- 分支 1 ----------
    if b in {'C', 'C C', 'C C C', 'C C C C', 'C C C C C', 'E', 'E E', 'E E E', 'E E E E', 'E E E E E'}:
        # print('-> 分支 1：B 是 C/C C/E/E E，直接 True')
        return True

    # ---------- 分支 2 ----------
    elif b == '' and g == '':
        if f != '':  # F 非空
            # print('-> 分支 2a：B 空 & G 空 & F 非空，True')
            return True
        else:  # F 空
            if c != '':  # C 非空
                # print('-> 分支 2b：B 空 & G 空 & F 空 & C 非空，False')
                return False
            else:  # C 也空
                # print('-> 分支 2c：B 空 & G 空 & F 空 & C 空，True')
                return True

    # ---------- 分支 3 ----------
    elif b == '' and g in {'C', 'C C', 'E', 'E E'}:
        # print('-> 分支 3：B 空 & G 是 C/C C/E/E E，False')
        return False

    # ---------- 分支 4（其余） ----------
    else:
        # print('-> 分支 4：其余情况，False')
        return False

def export_report():
    try:
        pdfmetrics.registerFont(TTFont('SimSun', 'simsun.ttc'))  # 注册字体
        stem, _ = os.path.splitext(os.path.basename(FilePath))
        home_path = os.path.expanduser("~")
        desktop_path = os.path.join(home_path, "Desktop")
        global project_number
        project_number = stem.replace('-Terminal list', '')
        desktop_path = askdirectory(title=u'请选择导出文件夹', initialdir=desktop_path)
        if not desktop_path:
            tk.messagebox.showwarning("提示", "未选择保存路径，导出操作已取消")
            return

        doc = SimpleDocTemplate(os.path.join(desktop_path, '端子检查报告-%s.pdf' % project_number), pagesize=A4, rightMargin=40, leftMargin=40, topMargin=100, bottomMargin=20)
        story = []
        # 获取样式
        styles = getSampleStyleSheet()
        style = styles['Normal']
        style.fontName = 'SimSun'
        style.wordWrap = 'CJK'
        style.leading = 15  # 行间距

        text_content = text.get("1.0", tk.END)
        start_index = text_content.find(">>>端子号正在检查中...")
        end_index = find_end_index(text_content, ">>>端子隔板缺失检查完成! ")

        text_content = text_to_html(text, start_index, end_index)

        p = Paragraph(text_content,style)
        story.append(p)

        # table_name = '附表1：端子型号表'
        # table_header_style = styles['Title']
        # table_header_style.leading = 22
        # table_header_style.spaceAfter = 0
        # table_header_style.fontName = 'SimSun'  # 使用SimSun字体
        # table_header_style.fontSize = 10  # 可以设置更大的字体大小
        #
        # # 创建表格名称的段落并添加到story
        # p = Paragraph(table_name, table_header_style)
        # story.append(p)
        #
        # df = extract_terminal_types(text)
        # table_data = df.values.tolist()
        # # table_data = [df.columns.tolist()] + df.values.tolist()
        # # print(table_data)
        #
        # # 这里创建局部样式styleN以避免影响全局样式
        # styleN = styles['Normal'].clone('new_centered_style')  # 创建一个基于Normal的新样式副本
        # styleN.alignment = "JUSTIFY"  # 设置对齐为两端对齐
        # styleN.fontName = 'SimSun'
        # # 构建包含旋转文本段落的单元格数据
        #
        # table_data_rotated = [[Paragraph("<para align=justify>{}</para>".format(cell), styleN) for cell in row] for row in table_data]
        # table = Table(table_data_rotated, spaceAfter=10, colWidths=[1.5 * cm] + [None] * (len(table_data[0]) - 1))
        # table.setStyle(TableStyle([
        #     ('BACKGROUND', (0, 0), (-1, 0), colors.white),
        #     ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        #     ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        #     ('FONTNAME', (0, 0), (-1, -1), 'SimSun'),
        #     ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        #     ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        #     ('GRID', (0, 0), (-1, -1), 1, colors.black),
        # ]))
        #
        # story.append(table)

        global page_number
        page_number = 1

        # 通过`SimpleDocTemplate`直接生成文档，传入`add_header`函数绘制每页的头部内容
        doc.build(story, onFirstPage=add_header1, onLaterPages=add_header2)

        tk.messagebox.showwarning("提示", "端子检查报告导出完成")
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def find_end_index(text_content, marker):
    lines = text_content.splitlines(True)  # 使用True保持换行符
    for i, line in enumerate(lines):
        if marker in line:
            return sum(len(l) for l in lines[:i+1])  # 返回到该行末尾的累计长度
    return len(text_content)  # 如果没有找到marker，返回文本的总长度


# def extract_terminal_types(text_widget):
#     text_content = text_widget.get("1.0", tk.END)
#     start_index = text_content.find(">>>端子类型正在检查中...")
#     end_index = text_content.find(">>>端子类型检查完成")
#     terminal_text = text_content[start_index:end_index]
#     terminal_lines = terminal_text.split("\n")[1:-1]
#     terminal_lines = [line.replace('为', ':').replace('端子', '') for line in terminal_lines]
#     # print(terminal_lines)
#
#     typical_info = [line.split("——")[0] for line in terminal_lines]
#     terminal_info = [line.split("——")[1] for line in terminal_lines]
#
#     df = pd.DataFrame({"Typical": typical_info, "Terminal": terminal_info})
#     df = df.groupby("Typical")["Terminal"].apply(lambda x: '，'.join(x)).reset_index()
#     return df


def text_to_html(text_widget, start_index, end_index):
    # 获取文本内容
    text_content = text_widget.get("1.0", tk.END)

    # 获取 start_index 之前的文本
    before_text = text_content[:start_index-1]
    # 获取 end_index 之后的文本
    after_text = text_content[end_index:]

    # 合并排除了指定范围的文本
    final_text = before_text + after_text

    # 将换行符转换为HTML的换行标签
    html_output = final_text.replace('\n', '<br />')

    # 结束HTML文档
    html_output = f"<html><body><p>{html_output}</p></body></html>"

    return html_output

def add_header1(rl_canvas, doc):
    rl_canvas.saveState()

    # 添加图片（请确保图片路径正确）
    image_path = 'ico/ABB_logo.png'
    img_width = 27.7 * mm
    img_height = 10.5 * mm
    img = Image(image_path, width=img_width, height=img_height)
    img.drawOn(rl_canvas, A4[0] / 2 - img_width / 2, A4[1] - img_height - 30)

    # 添加标题
    rl_canvas.setFont('SimSun', 16)
    title = "FAST 端子检查报告"
    rl_canvas.drawCentredString(A4[0] / 2, A4[1] - img_height - 60, title)

    rl_canvas.setFont('SimSun', 10)
    project_number_text = "项目号：" + project_number
    rl_canvas.drawString(45, A4[1] - img_height - 60 - 10, project_number_text)

    rl_canvas.setFont('SimSun', 8)
    global page_number
    page_number_text = "Page " + str(page_number)
    page_number += 1
    rl_canvas.drawString(A4[0]-80, 5*mm, page_number_text)

    rl_canvas.setFont('SimSun', 8)
    Timestamp = strftime('%Y-%m-%d %H:%M:%S', localtime())
    rl_canvas.drawCentredString(A4[0] / 2, 5 * mm, '制表时间：' + Timestamp)

    rl_canvas.restoreState()


def add_header2(rl_canvas, doc):
    rl_canvas.saveState()

    # 添加图片（请确保图片路径正确）
    image_path = 'ico/ABB_logo.png'
    img_width = 27.7 * mm
    img_height = 10.5 * mm
    img = Image(image_path, width=img_width, height=img_height)
    img.drawOn(rl_canvas, A4[0] / 2 - img_width / 2, A4[1] - img_height - 30)

    # 添加标题
    rl_canvas.setFont('SimSun', 16)
    title = "FAST 端子检查报告"
    rl_canvas.drawCentredString(A4[0] / 2, A4[1] - img_height - 60, title)

    rl_canvas.setFont('SimSun', 8)
    global page_number
    page_number_text = "Page " + str(page_number)
    page_number += 1
    rl_canvas.drawString(A4[0]-80, 5*mm, page_number_text)

    rl_canvas.setFont('SimSun', 8)
    Timestamp = strftime('%Y-%m-%d %H:%M:%S', localtime())
    rl_canvas.drawCentredString(A4[0] / 2, 5 * mm, '制表时间：' + Timestamp)

    rl_canvas.restoreState()