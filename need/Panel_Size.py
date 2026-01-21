import tkinter as tk
from tkinter import ttk

from tkinter.ttk import Treeview, Style

from tkinter.filedialog import askdirectory
import os


import time
import datetime
from time import *

import pandas as pd
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl import load_workbook
import traceback

import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

FilePath = ""    # 设置一个地址变量


def main(parent, w_ratio, h_ratio, file_path):
    global checklog_file_path
    checklog_file_path = file_path
    global open_folder
    open_folder = tk.PhotoImage(file="ico\\open_folder.png")
    global analyze_file
    analyze_file = tk.PhotoImage(file="ico\\read.png")

    tk.Label(parent, text="欢迎使用柜型尺寸检查功能", bg="#c9dbe9", fg="black", height=int(1*h_ratio), font=("ABBvoice CNSG", int(20 * h_ratio), "bold")).pack(side=tk.TOP, fill=tk.X)
    f1 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f1, text='   说明：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')
    tk.Label(f1, text='(1)EPLAN菜单Tools→Reports:Automated processing中选择Export Files V04导出；\n(2)选择柜型文件【C:/Temp/项目号-Files/项目号-Panel.xlsx】；\n(3)输出并导出EPLAN属性设置中的柜型尺寸，做初步检查，并用于后续与单线图核对；\n(4)点击EPLAN菜单栏Project data下Structure identifier management…中为相应Typical赋值，注意宽深属性对应Typical栏，高对应Panel栏。如果是尺寸缺失，需要维护信息；\n如果尺寸不缺失，但出现蓝色地球图标，右键属性，点击Remove translations即可，再次批量导出报表即可。', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    f1.pack(fill=tk.X)

    tk.Frame(parent, height=int(20*h_ratio), bg="#eaf1f6").pack(fill=tk.X)    # 水平分割线

    f2 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f2, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f2, text='路径：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry    # 为了确保selectpath函数能正确调用entry,将其全局化
    entry = tk.Entry(f2, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)))
    entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

    global button_check
    button_check = tk.Button(f2, text='检查', font=("ABBvoice CNSG", int(13 * h_ratio)), image=analyze_file, bg="#eaf1f6", compound=tk.LEFT, command=process, activebackground='blue')
    button_check.pack(side=tk.RIGHT, padx=(0, int(20*w_ratio)))
    button_check['state'] = 'disabled'

    tk.Button(f2, text='选择', font=("ABBvoice CNSG", int(13 * h_ratio)), image=open_folder, bg="#eaf1f6", compound=tk.LEFT, command=selectpath, activebackground='blue').pack(side=tk.RIGHT, padx=int(20*w_ratio))

    f2.pack(fill=tk.X)

    tk.Frame(parent, height=int(20*h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f3 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f3, text='   数据：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')

    style = Style()
    style.configure('panel.Treeview', rowheight=30, font=("ABBvoice CNSG", int(13 * h_ratio)))
    style.configure('panel.Treeview.Heading', font=("ABBvoice CNSG", int(13 * h_ratio)), background="#EFF1F5")

    global Panel_Size_table
    table_ybar = tk.Scrollbar(f3, orient="vertical")
    table_ybar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, int(20*w_ratio)))

    Panel_Size_table = Treeview(f3, show='headings', style='panel.Treeview', columns=('a', 'b', 'c', 'd', 'e', 'f'), yscrollcommand=table_ybar.set, height=int(15*h_ratio))

    Panel_Size_table.tag_configure('error_line', foreground='red')    # 缺失行显示红色

    table_ybar.config(command=Panel_Size_table.yview)
    Panel_Size_table.column('a', width=int(250 * w_ratio), anchor='center')
    Panel_Size_table.column('b', width=int(250 * w_ratio), anchor='center')
    Panel_Size_table.column('c', width=int(250 * w_ratio), anchor='center')
    Panel_Size_table.column('d', width=int(250 * w_ratio), anchor='center')
    Panel_Size_table.column('e', width=int(250 * w_ratio), anchor='center')
    Panel_Size_table.column('f', width=int(250 * w_ratio), anchor='center')

    Panel_Size_table.heading('a', text='站号')
    Panel_Size_table.heading('b', text='柜型')
    Panel_Size_table.heading('c', text='柜号')
    Panel_Size_table.heading('d', text='宽')
    Panel_Size_table.heading('e', text='高')
    Panel_Size_table.heading('f', text='深')

    Panel_Size_table.pack(fill=tk.BOTH, padx=(0, 1), pady=0)
    f3.pack(fill=tk.X)

    tk.Frame(parent, height=int(20*h_ratio), bg="#eaf1f6").pack(fill=tk.X)    # 水平分割线

    f4 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f4, text='   结果：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')

    global text
    text = tk.Text(f4, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), height=int(25 * h_ratio), width=int(65*w_ratio))
    text.pack(side=tk.LEFT, padx=(0, 1), pady=0, fill=tk.BOTH, expand=True)
    text.tag_configure('error', foreground='red')  # 设置tag

    scrollbar = tk.Scrollbar(f4)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, int(20*w_ratio)))
    scrollbar.config(command=text.yview)
    text.config(yscrollcommand=scrollbar.set)
    f4.pack(fill=tk.BOTH, expand=True)

    tk.Frame(parent, height=int(20*h_ratio), bg="#eaf1f6").pack(fill=tk.X)    # 水平分割线


def selectpath():
    filepath = tk.filedialog.askopenfilename(initialdir='C:/Temp/', title=u'请选择文件(项目号-Panel.xlsx)', filetypes=[("Excel", ".xlsx")])    # 选择打开什么文件，返回文件名
    if len(filepath) != 0:
        string_filename = ""
        for i in range(0, len(filepath)):
            string_filename += str(filepath[i])
        button_check['state'] = 'normal'
    else:
        button_check['state'] = 'disabled'

    text.delete(1.0, tk.END)    # 清空输出结果框

    # 清空treeview表格
    table_items = Panel_Size_table.get_children()  # 在插入treeview数据时，需要先清空treeview
    [Panel_Size_table.delete(table_item) for table_item in table_items]

    entry.delete(0, "end")    # 删除entry原始内容
    entry.insert(0, filepath)    # 重新填入地址
    global FilePath
    FilePath = filepath


def process():
    try:
        text.delete(1.0, 'end')  # 清空输出结果框
        # 清空treeview表格
        table_items = Panel_Size_table.get_children()  # 在插入treeview数据时，需要先清空treeview
        [Panel_Size_table.delete(table_item) for table_item in table_items]

        if FilePath == "":
            tk.messagebox.showwarning("提示", "请选择文件！")
        stem, suffix = os.path.splitext(os.path.basename(FilePath))    # stem是文件名,suffix是后缀
        if 'data' in stem or 'Size' in stem or '-Panel' not in stem:
            tk.messagebox.showwarning("提示", "请选择项目号-Panel.xlsx文件！")

        else:
            # text.insert(tk.INSERT, '>>>柜型尺寸正在读取中...\n')    # 进行柜型尺寸表格处理
            start = time()
            # print("正在处理中...")
            book = load_workbook(FilePath)
            sheet = book['Z2_xlsx']
            A = []
            B = []
            C = []
            for i in range(2, sheet.max_row + 1):
                A.append(str(sheet.cell(row=i, column=1).value))    # Order Line列
                B.append(str(sheet.cell(row=i, column=2).value))    # Typical列
                C.append(str(sheet.cell(row=i, column=3).value))    # Panel No列

            if B[0] == 'None':    # 说明是旧版升版2024项目，它的Panel.xlsx报表B列为空，需要用Panel.xls，如果是纯2024项目，只需Panel.xlsx即可
                A = []
                B = []
                C = []
                book = xlrd.open_workbook(os.path.join(os.path.dirname(FilePath), stem+'.xls'))    # Panel.xls
                sheet = book.sheet_by_index(0)
                A = sheet.col_values(colx=0, start_rowx=1, end_rowx=None)    # Order Line列
                B = sheet.col_values(colx=1, start_rowx=1, end_rowx=None)    # Typical列
                C = sheet.col_values(colx=2, start_rowx=1, end_rowx=None)    # Panel No列

            stem, suffix = os.path.splitext(os.path.basename(FilePath))    # stem是文件名,suffix是后缀
            # os.path.dirname()去掉文件名，返回目录
            # os.path.basename()去掉目录，返回文件名(含后缀)
            inputfile = os.path.join(os.path.dirname(FilePath), stem+' data'+'.xlsx')    # Panel data.xlsx
            outputfile = os.path.join(os.path.dirname(FilePath), stem+' Size'+'.xls')    # Panel Size.xls

            book1 = load_workbook(inputfile)
            sheet1 = book1['Z3_xlsx']
            A1 = []
            B1 = []
            C1 = []
            D1 = []
            for i in range(2, sheet1.max_row + 1):
                A1.append(str(sheet1.cell(row=i, column=1).value))    # Typical/Panel列
                B1.append(str(sheet1.cell(row=i, column=2).value))    # 柜宽（Typical）列
                C1.append(str(sheet1.cell(row=i, column=3).value))    # 柜深（Typical）列
                D1.append(str(sheet1.cell(row=i, column=4).value))    # 柜高（Panel）列
                # print(str(sheet1.cell(row=i, column=1).value), str(sheet1.cell(row=i, column=2).value), str(sheet1.cell(row=i, column=3).value), str(sheet1.cell(row=i, column=4).value))
            book2 = xlwt.Workbook()    # 创建一个空文件对象
            book2.add_sheet('Sheet1')    # 创建一个Sheet页
            book2.save(outputfile)    # 创建Panel Size.xls文件
            book2 = xlrd.open_workbook(outputfile)    # 加载【项目号-Panel Size.xls】表格

            workbook = copy(book2)    # 使用xlutils.copy将xlrd读取的book对象转为xlwt可操作对象
            worksheet = workbook.get_sheet(0)    # 获取sheet
            worksheet.write(0, 0, '站号')    # 在sheet指定位置写入数据
            worksheet.write(0, 1, '柜型')    # 在sheet指定位置写入数据
            worksheet.write(0, 2, '柜号')    # 在sheet指定位置写入数据
            worksheet.write(0, 3, '宽(mm)')    # 在sheet指定位置写入数据
            worksheet.write(0, 4, '深(mm)')    # 在sheet指定位置写入数据
            worksheet.write(0, 5, '高(mm)')    # 在sheet指定位置写入数据
            for i in range(0, len(A)):    # 遍历Panel.xls，将Panel.xls数据复制到Panel Size.xls
                worksheet.write(i+1, 0, A[i])
                worksheet.write(i+1, 1, B[i])
                worksheet.write(i+1, 2, C[i])

            flag = 0
            flag1 = 0
            NAN_flag = 0

            def custom_filter(char):
                return char.isdigit() or char == '+'

            for i in range(0, len(A)):    # 遍历Panel.xlsx
                if B[i] not in A1:
                    worksheet.write(i + 1, 3, '--')
                    worksheet.write(i + 1, 4, '--')
                    NAN_flag += 1
                for j in range(0, len(A1)):    # 遍历Panel data.xls，逐行读取尺寸信息
                    if B[i] == A1[j] and B1[j] != 'None' and C1[j] != 'None':
                        worksheet.write(i+1, 3, str(''.join(filter(custom_filter, B1[j]))))    # https://blog.csdn.net/weixin_44606217/article/details/100534834
                        worksheet.write(i+1, 4, str(''.join(filter(custom_filter, C1[j]))))
                        flag = 1
                    if B[i] == A1[j] and B1[j] == 'None' and C1[j] != 'None':
                        NAN_flag += 1
                        worksheet.write(i+1, 3, '--')
                        worksheet.write(i+1, 4, str(''.join(filter(custom_filter, C1[j]))))
                    if B[i] == A1[j] and B1[j] != 'None' and C1[j] == 'None':
                        NAN_flag += 1
                        worksheet.write(i+1, 3, str(''.join(filter(custom_filter, B1[j]))))
                        worksheet.write(i+1, 4, '--')
                    if B[i] == A1[j] and B1[j] == 'None' and C1[j] == 'None':
                        NAN_flag += 1
                        worksheet.write(i+1, 3, '--')
                        worksheet.write(i+1, 4, '--')
                    if C[i] == A1[j] and D1[j] != 'None':
                        worksheet.write(i+1, 5, str(''.join(filter(custom_filter, D1[j]))))
                        flag1 = 1
                    if C[i] == A1[j] and D1[j] == 'None':
                        worksheet.write(i+1, 5, '--')
                        NAN_flag += 1
                        flag1 = 1
                    if flag and flag1:
                        break    # 当宽高深数据都匹配到，就跳出，继续Panel.xls下一行
                flag = 0
                flag1 = 0
            workbook.save(outputfile)    # 将workbook保存到指定位置

            # if NAN_flag:    # 处理出现数据缺失的情况
                # book2 = xlrd.open_workbook(outputfile)  # 加载【项目号-Panel Size.xls】表格
                # sheet2 = book2.sheet_by_index(0)
                # # Panel size.xls中会出现部分Typical（F2A,F2B）缺少数据的情况（默认F2数据是完整的），需要对文件数据进行遍历
                # for i in range(0, len(A)):
                #     if sheet2.cell_value(i+1, 3) == '--' or sheet2.cell_value(i+1, 4) == '--':    # 如果存在空的情况
                #         for j in range(0, len(A)):
                #             if i == j:
                #                 continue
                #             else:
                #                 if B[i][:-1] == B[j] and B[i][-1].isalpha():    # 如果Typical字符串除最后一个字母外都相同,且前者Typical最后一个字符为字母
                #                     worksheet.write(i+1, 3, sheet2.cell_value(j+1, 3))
                #                     worksheet.write(i+1, 4, sheet2.cell_value(j+1, 4))
                #                 if B[i][:-1] == B[j][:-1] and B[i][-1].isalpha() and B[j][-1].isalpha():    # 如果Typical字符串除最后一个字母外都相同,且两者Typical最后一个字符都为字母
                #                     worksheet.write(i+1, 3, sheet2.cell_value(j+1, 3))
                #                     worksheet.write(i+1, 4, sheet2.cell_value(j+1, 4))
                # workbook.save(outputfile)  # 将workbook保存到指定位置

            # 数据表为了方便与单线图核对，应将“宽，深，高”改成“宽，高，深”排列，交换5,6列即可
            book2 = xlrd.open_workbook(outputfile)  # 加载【项目号-Panel Size.xls】表格
            sheet2 = book2.sheet_by_index(0)
            workbook = copy(book2)  # 使用xlutils.copy将xlrd读取的book对象转为xlwt可操作对象
            worksheet = workbook.get_sheet(0)  # 获取sheet
            switchgear_number_panel_size = sheet2.col_values(colx=0, start_rowx=0, end_rowx=None)
            typical_type_panel_size = sheet2.col_values(colx=1, start_rowx=0, end_rowx=None)
            panel_number_panel_size = sheet2.col_values(colx=2, start_rowx=0, end_rowx=None)

            depth = sheet2.col_values(colx=4, start_rowx=0, end_rowx=None)
            for i in range(1, len(depth)):
                if '+' in depth[i]:
                    text.insert(tk.INSERT, f'EPLAN属性：站号{switchgear_number_panel_size[i]},柜型{typical_type_panel_size[i]},柜号{panel_number_panel_size[i]}深度为{depth[i]},请移除+号及其后尺寸,只保留+号前面的尺寸\n', 'error')
            height = sheet2.col_values(colx=5, start_rowx=0, end_rowx=None)
            for i in range(0, len(depth)):
                worksheet.write(i, 4, height[i])    # 将”高“数据放到第5列
                worksheet.write(i, 5, depth[i])    # 将”宽“数据放到第6列
            workbook.save(outputfile)  # 将workbook保存到指定位置

            # 为了方便与单线图核对，将表格按照站号对柜号进行排序
            df = pd.read_excel(outputfile, names=['站号', '柜型', '柜号', '宽(mm)', '高(mm)', '深(mm)'])
            df.sort_values(by=['站号', '柜号'], ascending=True, inplace=True)    # 升序
            df.to_excel(outputfile, index=False)    # 写入excel时不带索引

            # 为treeview表格插入数据
            for i in range(0, len(df.values)):
                if '--' in list(df.values[i]):
                    Panel_Size_table.insert('', 'end', values=list(df.values[i]), tags="error_line")  # 逐行插入到表格中
                    # print('发现--')
                else:
                    Panel_Size_table.insert('', 'end', values=list(df.values[i]))    # 逐行插入到表格中

            # text.insert(tk.INSERT, "站号   柜型   柜号   宽   高   深\n")
            # for i in range(0, len(df.values)):
            #     text.insert(tk.INSERT, "%s  " % df.values[i][0])    # DataFrame行索引属性index,列索引属性columns,值属性values
            #     if len(df.values[i][1]) == 2:
            #         text.insert(tk.INSERT, "%s      " % df.values[i][1])
            #     elif len(df.values[i][1]) == 3:
            #         text.insert(tk.INSERT, "%s    " % df.values[i][1])
            #     elif len(df.values[i][1]) == 4:
            #         text.insert(tk.INSERT, "%s  " % df.values[i][1])
            #     text.insert(tk.INSERT, "%s  " % df.values[i][2])
            #     text.insert(tk.INSERT, "%s  " % df.values[i][3])
            #     text.insert(tk.INSERT, "%s  " % df.values[i][4])
            #     text.insert(tk.INSERT, "%s  \n" % df.values[i][5])
            end = time()
            # text.insert(tk.INSERT, ">>>数据存入%s\n" % outputfile.replace("\\", "/"))

            if NAN_flag > 0:
                text.insert(tk.INSERT, ">>>尺寸数据有缺失，请手动维护EPLAN！！！用时%.3f秒\n" % (end - start), 'error')
            else:
                text.insert(tk.INSERT, ">>>尺寸数据无缺失，用时%.3f秒\n" % (end - start))

            global checklog_file_path
            checklogbook = load_workbook(checklog_file_path)
            checklogsheet = checklogbook['Sheet']
            project_no = stem.replace('-Panel', '')
            current_time = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            data = [project_no, '尺寸检查', NAN_flag, current_time]
            checklogsheet.append(data)
            checklogbook.save(checklog_file_path)

    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())






