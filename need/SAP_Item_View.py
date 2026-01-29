import tkinter as tk
from tkinter import ttk
from tkinter.ttk import Treeview, Style
from tkinter.filedialog import askdirectory
import need.tkutils as tku
import os
import logging
from cryptography.fernet import Fernet
import base64
import traceback
import pyrfc
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
import warnings
import json
import time
import datetime
from openpyxl import load_workbook

from need.custom_dialogs import CustomDialog, center_window, Tooltip, image_label
import datetime as dt

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from tkinter.filedialog import asksaveasfilename
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm

warnings.simplefilter(action='ignore', category=FutureWarning)

FilePath = ""    # 设置一个地址变量

def main(parent, root, w_rat, h_rat, file_path):
    global HISTORY_FILE
    HISTORY_DIR = r'C:\Temp\FAST_log\history'
    HISTORY_FILE = os.path.join(HISTORY_DIR, 'project_history.json')
    os.makedirs(HISTORY_DIR, exist_ok=True)

    global checklog_file_path
    checklog_file_path = file_path

    global project_history, MAX_HISTORY
    MAX_HISTORY = 5  # 最多保留5条
    project_history = []

    global h_ratio
    h_ratio = h_rat
    global w_ratio
    w_ratio = w_rat
    global root_win
    root_win = root

    global level_fold
    level_fold = tk.PhotoImage(file="ico\\fold.png")
    global level_unfold
    level_unfold = tk.PhotoImage(file="ico\\unfold.png")

    global query_project
    query_project = tk.PhotoImage(file="ico\\search.png")

    global sap_export_cbbom
    sap_export_cbbom = tk.PhotoImage(file="ico\\export.png")

    global sap_export_ebom
    sap_export_ebom = tk.PhotoImage(file="ico\\export.png")

    global sap_compare_ebom
    sap_compare_ebom = tk.PhotoImage(file="ico\\sap_compare_ebom.png")

    global sap_compare_pr_ebom
    sap_compare_pr_ebom = tk.PhotoImage(file="ico\\sap_compare_pr_ebom.png")

    global sap_item_select
    sap_item_select = tk.PhotoImage(file="ico\\sap_item_select.png")

    global sap_ebom_compare_export
    sap_ebom_compare_export = tk.PhotoImage(file="ico\\sap_ebom_compare_export.png")

    global sap_ebom_compare
    sap_ebom_compare = tk.PhotoImage(file="ico\\check.png")

    tk.Label(parent, text="欢迎使用SAP的BOM对比功能", bg="#c9dbe9", fg="black", height=int(1 * h_ratio), font=("ABBvoice CNSG", int(20 * h_ratio), "bold")).pack(fill=tk.X)

    tk.Frame(parent, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f1 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f1, text=' *', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f1, text='SO ：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    # global entry_project_number    # 为了确保selectpath函数能正确调用entry_project_number,将其全局化
    # entry_project_number = ttk.Entry(f1, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(20 * w_ratio))
    # entry_project_number.pack(side=tk.LEFT, fill=tk.X)
    #
    # entry_project_number.bind('<Return>', query)

    parent.option_add("*TCombobox*Listbox.font", ("ABBvoice CNSG", int(13 * h_ratio)))

    global combobox_project_number
    combobox_project_number = ttk.Combobox(f1, width=int(20 * w_ratio), font=("ABBvoice CNSG", int(13 * h_ratio)))
    combobox_project_number.pack(side=tk.LEFT)
    combobox_project_number.bind("<<ComboboxSelected>>", on_history_select)
    combobox_project_number.bind("<Return>", query)
    refresh_history_combo()
    Tooltip(combobox_project_number, "点击v，查看最近输入历史")
    tk.Button(f1, text='查询', font=("ABBvoice CNSG", int(13 * h_ratio)), image=query_project, bg="#eaf1f6", compound=tk.LEFT, command=query, activebackground='blue').pack(side=tk.LEFT, padx=int(20 * w_ratio))

    tk.Label(f1, text='  (仅支持9位项目号  |  本功能所有表格数据均支持Ctrl+C复制)', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(10 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    f1.pack(fill=tk.X)

    tk.Frame(parent, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f2 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f2, text='   项目：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry_project_name
    entry_project_name = tk.Entry(f2, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(100*w_ratio))
    entry_project_name.pack(side=tk.LEFT, fill=tk.X)

    f2.pack(fill=tk.X)

    tk.Frame(parent, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    fun = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(fun, text='   功能：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global button_level_fold
    button_level_fold = tk.Button(fun, image=level_fold, bg="#eaf1f6", command=lambda e=0:button_wrapper(e), state='disabled', activebackground='blue')
    button_level_fold.pack(side=tk.LEFT, padx=(0, 5))
    # button_level_fold.bind('<Enter>', lambda e, name='1': on_enter_for_tooltip(e, name))
    # button_level_fold.bind('<Leave>', on_leave_for_tooltip)
    Tooltip(button_level_fold, "折叠下表Item")

    global button_level_unfold
    button_level_unfold = tk.Button(fun, image=level_unfold, bg="#eaf1f6", command=lambda e=1:button_wrapper(e), state='disabled', activebackground='blue')
    button_level_unfold.pack(side=tk.LEFT, padx=(0, 5))
    # button_level_unfold.bind('<Enter>', lambda e, name='2': on_enter_for_tooltip(e, name))
    # button_level_unfold.bind('<Leave>', on_leave_for_tooltip)
    Tooltip(button_level_unfold, "展开下表Item")

    global button_export_item
    button_export_item = tk.Button(fun, text='Item导出', font=("ABBvoice CNSG", int(13 * h_ratio)), compound=tk.LEFT, image=sap_export_ebom, bg="#eaf1f6", command=export_item, state='disabled', activebackground='blue')
    button_export_item.pack(side=tk.LEFT, padx=(0, 5))
    Tooltip(button_export_item, '导出Item信息到Excel')

    global button_compare_ebom
    button_compare_ebom = tk.Button(fun, text='BOM对比', font=("ABBvoice CNSG", int(13 * h_ratio)), compound=tk.LEFT, image=sap_compare_ebom, bg="#eaf1f6", command=compare_ebom, state='disabled', activebackground='blue')
    button_compare_ebom.pack(side=tk.LEFT, padx=(0, 5))
    # Tooltip(button_compare_ebom, 'BOM对比')

    global button_export_ebom
    button_export_ebom = tk.Button(fun, text='BOM导出', font=("ABBvoice CNSG", int(13 * h_ratio)), compound=tk.LEFT, image=sap_export_ebom, bg="#eaf1f6", command=export_ebom, state='disabled', activebackground='blue')
    button_export_ebom.pack(side=tk.LEFT, padx=(0, 5))
    # button_export_ebom.bind('<Enter>', lambda e, name='4': on_enter_for_tooltip(e, name))
    # button_export_ebom.bind('<Leave>', on_leave_for_tooltip)
    # Tooltip(button_export_ebom, '导出EBOM清单')

    global button_export_cbbom
    button_export_cbbom = tk.Button(fun, text='断路器选配导出', font=("ABBvoice CNSG", int(13 * h_ratio)), compound=tk.LEFT, image=sap_export_cbbom, bg="#eaf1f6", command=export_cbbom, state='disabled', activebackground='blue')
    # button_export_cbbom.pack(side=tk.LEFT, padx=(0, 5))
    # button_export_cbbom.bind('<Enter>', lambda e, name='3': on_enter_for_tooltip(e, name))
    # button_export_cbbom.bind('<Leave>', on_leave_for_tooltip)
    # Tooltip(button_export_cbbom, "导出断路器选配")

    # tk.Label(fun, text='  (下表不含SAP中已经被Cancel/Reject的CI/TI; 标红/蓝的是电气工程师需要关注的Item，双击可以显示详细信息)', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(10 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    fun.pack(fill=tk.X)

    tk.Frame(parent, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f3 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f3, text='   数据：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')

    style = Style()
    style.configure('panel1.Treeview', rowheight=25, font=("ABBvoice CNSG", int(13 * h_ratio)))
    style.configure('panel1.Treeview.Heading', font=("ABBvoice CNSG", int(13 * h_ratio)), background="#EFF1F5")

    global Item_Info_table
    ybar = tk.Scrollbar(f3)
    Item_Info_table = Treeview(f3, style='panel1.Treeview', columns=('a', 'b', 'c', 'd', 'e', 'f'), height=int(100 * h_ratio), selectmode='browse', yscrollcommand=ybar.set)
    ybar.config(command=Item_Info_table.yview)
    Item_Info_table.column('#0', width=int(120*w_ratio), anchor='w')
    Item_Info_table.column('a', width=int(200*w_ratio), anchor='w')
    Item_Info_table.column('b', width=int(50*w_ratio), anchor='w')
    Item_Info_table.column('c', width=int(400*w_ratio), anchor='w')
    Item_Info_table.column('d', width=int(600*w_ratio), anchor='w')
    Item_Info_table.column('e', width=int(120 * w_ratio), anchor='w')
    Item_Info_table.column('f', width=int(80 * w_ratio), anchor='w')

    Item_Info_table.heading('#0', text='Item', anchor='w')
    Item_Info_table.heading('a', text='Material', anchor='w')
    Item_Info_table.heading('b', text='数量', anchor='w')
    Item_Info_table.heading('c', text='描述', anchor='w')
    Item_Info_table.heading('d', text='柜号', anchor='w')
    Item_Info_table.heading('e', text='交货期', anchor='w')
    Item_Info_table.heading('f', text='YB/YE', anchor='w')

    Item_Info_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, pady=(0, 10))
    ybar.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10), pady=(0, 10))
    Item_Info_table.tag_configure('attention_row_a', foreground='red', font=("ABBvoice CNSG", int(11 * h_ratio)))
    Item_Info_table.tag_configure('attention_row_b', foreground='blue', font=("ABBvoice CNSG", int(11 * h_ratio)))
    Item_Info_table.tag_configure('no_attention_row', foreground='black', font=("ABBvoice CNSG", int(11 * h_ratio)))
    Item_Info_table.tag_configure('high_level_row', background='whitesmoke', font=("ABBvoice CNSG", int(11 * h_ratio)))
    Item_Info_table.bind('<Double-1>', on_double_click)
    Tooltip(Item_Info_table, "已经被Cancel/Reject的Item不显示；标红/蓝的是电气工程师需要关注的Item，双击可以显示BOM/断路器选配信息")
    f3.pack(fill=tk.X)

    tk.Frame(parent, height=int(2000 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    Item_Info_table.bind("<Control-c>",
                         lambda e: copy_cell_ctrl_c(e, Item_Info_table))

global _float_label
_float_label = None  # type: tk.Label | None

def copy_cell_ctrl_c(event, tree):
    global _float_label

    region = tree.identify("region", event.x, event.y)
    if region != "cell":
        return
    col = tree.identify_column(event.x)
    row = tree.identify_row(event.y)

    # ---- 取文本 & 写剪贴板 ----
    if col == "#0":
        text = tree.item(row, "text")
    else:
        idx = int(col.replace("#", "")) - 1
        text = tree.item(row, "values")[idx]
    root_win.clipboard_clear()
    root_win.clipboard_append(text)

    # ---- 销毁旧 Label（如果有） ----
    if _float_label is not None:
        _float_label.destroy()

    # ---- 计算单元格几何 ----
    x, y, w, h = tree.bbox(row, column=col)

    # ---- 新建 Label ----
    _float_label = tk.Label(tree, text=text, bg="#FFFB00", fg="black",
                            font=("ABBvoice CNSG", int(11 * 0.8)),  # 字号稍小
                            relief="solid", bd=1, anchor="w")
    _float_label.place(x=x, y=y, width=w, height=h)

    # ---- 事件：鼠标离开 或 获得焦点即自杀 ----
    _float_label.bind("<Leave>", lambda e: _float_label.destroy())
    _float_label.bind("<FocusIn>", lambda e: _float_label.destroy())
    # 额外保险：400 ms 后自动消失（防止卡住）
    _float_label.after(400, lambda: _float_label.destroy() if _float_label else None)

def button_wrapper(flag):
    if flag == 0:
        on_level_fold(Item_Info_table)
    elif flag == 1:
        on_level_unfold(Item_Info_table)


def on_level_fold(object):
    expand_tree(object, 0)


def on_level_unfold(object):
    expand_tree(object, 2)


def expand_tree(tree, level):
    def expand_children(item, current_level, target_level):
        if current_level < target_level:
            tree.item(item, open=True)
            children = tree.get_children(item)
            for child in children:
                expand_children(child, current_level + 1, target_level)
        else:
            tree.item(item, open=False)

    for item in tree.get_children():
        expand_children(item, 0, level)


def query(event=''):
    try:
        button_level_fold['state'] = 'disabled'
        button_level_unfold['state'] = 'disabled'
        button_export_item['state'] = 'disabled'
        button_compare_ebom['state'] = 'disabled'
        button_export_cbbom['state'] = 'disabled'
        button_export_ebom['state'] = 'disabled'
        entry_project_name.delete(0, 'end')    # 清空输出结果框
        global projectID, attention_rows, attention_a_rows, attention_b_rows
        projectID = combobox_project_number.get()    # 读取文本框内容

        attention_rows = []
        attention_a_rows = []
        attention_b_rows = []

        # 清空treeview表格
        table_items = Item_Info_table.get_children()    # 在插入treeview数据时，需要先清空treeview
        [Item_Info_table.delete(table_item) for table_item in table_items]

        if projectID == "" or len(projectID) != 9:
            tk.messagebox.showwarning("提示", "请输入正确项目号！")

        else:
            if not projectID.isdigit():
                tk.messagebox.showwarning("提示", "请输入正确项目号！")
            else:
                # SAP登录
                sap_key_file = 'J:/Engineering/ShareFolder/new_ABB_Production_Tools/Ps/k/&%&^RD&^T&^WRE^@FEYUGYU@^E@DRYTRYTDFYTWF.txt'
                if not os.path.exists(sap_key_file):
                    tk.messagebox.showwarning('提示', 'SAP加密密钥文件不存在')
                else:
                    with open(sap_key_file, 'rb') as file:
                        sap_key = file.read()
                sap_cipher = Fernet(sap_key)

                sap_ciphertext_file = 'J:/Engineering/ShareFolder/new_ABB_Production_Tools/Ps/s/^&%&^R^YFT$%FTDTEYTKFY$^$FGKJGHE%#$%#EYD^%#^%DTYD.txt'
                if not os.path.exists(sap_ciphertext_file):
                    tk.messagebox.showwarning('提示', 'SAP加密密文文件不存在')
                else:
                    with open(sap_ciphertext_file, 'r') as file:
                        sap_ciphertext = file.read()
                global sap_h, sap_c, sap_s, sap_u, sap_p
                sap_account = sap_decrypt_data(sap_ciphertext, sap_cipher, 'z_')
                sap_h = sap_account.split('+')[0]
                sap_c = sap_account.split('+')[1]
                sap_s = sap_account.split('+')[2]
                sap_u = sap_account.split('+')[3]
                sap_p = sap_account.split('+')[4]

                logging.info("Ready to connect to SAP")
                try:
                    conn = pyrfc.Connection(ashost=sap_h, sysnr=sap_s, client=sap_c, user=sap_u, passwd=sap_p, lang='EN')
                    if conn.alive:
                        logging.info("Connecting to SAP successfully")

                        result = conn.call('ZY_SALES_ORDER_SHIFT', VBELN='0'+projectID)
                        if result['EX_CEPTION'] == '' and result['ITAB'][0]['WERKS'] == '1201':
                            project_name = result['ITAB'][0]['BSTKD']
                            entry_project_name.insert(0, project_name)

                            item_data = []
                            # 处理数据
                            for item in result['ITAB']:
                                posnr = item['POSNR'].lstrip('0') or '0'
                                kwmeng = str(int(float(item['KWMENG'])))
                                edate = f"{item['EDATU'][:4]}-{item['EDATU'][4:6]}-{item['EDATU'][6:]}"
                                arktx = item['ARKTX']
                                tptx1 = item['TPTX1']
                                tptx2 = item.get('TPTX2', '')
                                ettyp = item['ETTYP']
                                item_data.append((posnr, item['MATNR'], kwmeng, arktx, tptx1 + tptx2, edate, ettyp))

                            # 插入数据
                            parent_nodes = {}
                            # print(item_data)
                            for item in item_data:
                                posnr = int(item[0])

                                if posnr % 1000 == 0:
                                    parent_id = Item_Info_table.insert("", "end", text=str(posnr), values=item[1:6]+(str(''),), tags='high_level_row')
                                    parent_nodes[posnr] = parent_id
                                    attention_rows.append(item)
                                else:
                                    parent_id = parent_nodes.get((posnr // 1000) * 1000)
                                    if parent_id:
                                        # 获取 posnr 的千位数
                                        base = (posnr // 1000) * 1000
                                        # 计算 posnr 的偏移量
                                        offset = posnr - base

                                        # 判断使用哪种tag
                                        if 501 <= offset <= 549 and ('CDX' in str(item[1]) or 'GCE' in str(item[1]) or 'HD4' in str(item[1])):
                                            tag = 'attention_row_a'
                                            attention_a_rows.append(item)
                                            attention_rows.append(item)
                                        # elif 750 <= offset <= 799 and ('NAMEPLATE' == str(item[1]) or 'DZP-' in str(item[1])):
                                        #     tag = 'attention_row_a'
                                        elif 750 <= offset <= 799 and 'AA' in str(item[1]):
                                            tag = 'attention_row_a'
                                            attention_b_rows.append(item)
                                            attention_rows.append(item)
                                        elif (801 <= offset <= 899 and 'EBOM' in str(item[1])) or (951 <= offset <= 980 and 'ACC-EL' == str(item[1])):
                                            tag = 'attention_row_a'
                                            attention_b_rows.append(item)
                                            attention_rows.append(item)
                                        elif 991 <= offset <= 998 and str(item[1]) == 'PR':
                                            tag = 'attention_row_b'
                                            attention_b_rows.append(item)
                                            attention_rows.append(item)
                                        else:
                                            tag = 'no_attention_row'

                                        if base == 999000 and 'ACC-EL' in str(item[1]):
                                            tag = 'attention_row_a'
                                            attention_b_rows.append(item)
                                            attention_rows.append(item)

                                        # 插入数据
                                        if tag:
                                            Item_Info_table.insert(parent_id, "end", text=str(posnr), values=item[1:5]+(str(''),)+item[6:], tags=(tag,))
                                        else:
                                            Item_Info_table.insert(parent_id, "end", text=str(posnr), values=item[1:5]+(str(''),)+item[6:])

                            button_level_fold['state'] = 'normal'
                            button_level_unfold['state'] = 'normal'
                            button_export_item['state'] = 'normal'
                            button_compare_ebom['state'] = 'normal'
                            button_export_cbbom['state'] = 'normal'
                            button_export_ebom['state'] = 'normal'

                            found = False
                            for item in project_history:
                                if item['project'] == projectID:
                                    item['time'] = time.strftime('%Y-%m-%d %H:%M:%S')
                                    found = True
                                    break

                            # 如果不存在则新增
                            if not found:
                                project_history.append({
                                    'project': projectID,
                                    'time': time.strftime('%Y-%m-%d %H:%M:%S')
                                })

                            # if projectID not in [x['project'] for x in project_history]:
                            #     project_history.append({'project': projectID,
                            #                             'time': time.strftime('%Y-%m-%d %H:%M:%S')})
                            save_history()
                            refresh_history_combo()


                        else:
                            if result['EX_CEPTION'] != '':
                                tk.messagebox.showwarning("提示", result['EX_CEPTION'])
                            else:
                                tk.messagebox.showwarning("提示", 'CNDMX的1201无此项目')


                        conn.close()
                        if not conn.alive:
                            logging.info("Disconnect from SAP")


                except pyrfc.RFCError as e:
                    logging.info(e.key + ', ' + e.message)
                    tk.messagebox.showwarning("提示", traceback.format_exc())

    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def export_treeview_to_excel(treeview, title="导出数据", project_id=""):
    """
    将 Treeview 中的数据导出到 Excel 文件
    :param treeview: Treeview 组件
    :param title: 导出对话框标题
    :param project_id: 项目ID，用于命名文件和工作表
    """
    # 获取所有列名
    columns = list(treeview["columns"])
    headers = [treeview.heading(col)["text"] for col in ["#0"] + columns]

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    # 设置工作表名称
    if project_id:
        ws.title = f"{project_id}-SAP_Item数据"
    else:
        ws.title = "SAP_Item数据"

    # 写入表头
    ws.append(headers)

    # 递归函数，用于遍历树状结构
    def traverse_tree(item=""):
        # 获取当前项的值
        values = treeview.item(item, "values")
        text = treeview.item(item, "text")

        # 写入当前项数据
        if text or values:
            row_data = [text] + list(values)
            ws.append(row_data)

        # 遍历子项
        children = treeview.get_children(item)
        for child in children:
            traverse_tree(child)

    # 开始遍历
    traverse_tree()

    # 设置单元格格式
    # 创建边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 遍历所有单元格，设置自动换行和边框
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
            cell.border = thin_border

    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # 最大列宽限制为50
        ws.column_dimensions[column_letter].width = adjusted_width

    # 保存文件
    home_path = os.path.expanduser("~")
    desktop_path = os.path.join(home_path, "Desktop")
    if project_id:
        file_path = os.path.join(desktop_path, f"{project_id}-SAP_Item数据.xlsx")
    else:
        file_path = asksaveasfilename(
            title=title,
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            initialdir=desktop_path
        )

    if file_path:
        wb.save(file_path)
        return True, file_path
    return False, ""


def export_item():
    success, file_path = export_treeview_to_excel(Item_Info_table, "导出 Item 信息", projectID)
    if success:
        tk.messagebox.showinfo("提示", f"数据已成功导出到 Excel 文件：\n{file_path}")
    else:
        tk.messagebox.showinfo("提示", "导出已取消")

def export_cbbom():
    if len(attention_a_rows) == 0:
        tk.messagebox.showwarning("提示", "无断路器Item，无法导出选配清单")
        return

    sel = PosnrSelector(root_win, h_ratio, attention_a_rows, attention_rows, mode='cb')

    root_win.wait_window(sel)  # 阻塞到关闭
    do_export_cbbom(sel)  # 一键分发


def do_export_cbbom(sel):
    rows = [attention_a_rows[i] for i in sel.selected]
    if not rows:
        tk.messagebox.showwarning("提示", "未选择任何条目，导出已取消")
        return

    home_path = os.path.expanduser("~")
    desktop_path = os.path.join(home_path, "Desktop")
    desktop_path = askdirectory(title=u'请选择导出文件夹', initialdir=desktop_path)
    if not desktop_path:
        tk.messagebox.showwarning("提示", "未选择保存路径，导出已取消")
        return

    file_path = os.path.join(desktop_path, f'断路器选配-{projectID}.xlsx')

    wb = Workbook()
    # 删除默认创建的空白表单
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 使用提前保存的行索引
    for item in rows:
        posnr = item[0].zfill(6)
        sheet_name = item[0] + '-' + item[1]

        try:
            conn = pyrfc.Connection(ashost=sap_h, sysnr=sap_s, client=sap_c, user=sap_u, passwd=sap_p, lang='EN')
            if conn.alive:
                cb_configuration_result = conn.call('ZSD_GET_SO_CONFIGURATION', IM_VBELN='0'+projectID, IM_POSNR=posnr)
                if cb_configuration_result['ET_CONF_OUT']:
                    sheet = wb.create_sheet(title=sheet_name)
                    sheet.append(['属性', '值'])  # 表头

                    for confi_item in cb_configuration_result['ET_CONF_OUT']:
                        Char_description = confi_item['ATBEZ']
                        Char_value = confi_item['ATWTB']
                        if 'VDN_' not in Char_description and '_TABLE' not in Char_description:
                            sheet.append([Char_description, Char_value])
            conn.close()
        except pyrfc.RFCError as e:
            logging.error(e.key + ', ' + e.message)
            tk.messagebox.showwarning("提示", traceback.format_exc())

    wb.save(file_path)
    tk.messagebox.showwarning("提示", "断路器选配已成功导出为Excel文件")


def compare_ebom():
    if len(attention_b_rows) == 0:
        tk.messagebox.showwarning("提示", "无BOM(EBOM/PR/ACC-EL/AA)，无法对比")
        return

    sel = PosnrSelector(root_win, h_ratio, attention_b_rows, attention_rows, mode='compare')

    root_win.wait_window(sel)


def do_compare_ebom(sel, new_date, old_date):
    compare_rows = [attention_b_rows[i] for i in sel.selected]
    if not compare_rows:
        tk.messagebox.showwarning("提示", "未选择任何条目，对比已取消", parent=root_win)
        return

    # --------------- 通用子函数 ---------------
    def fetch_bom(item_posnr, date_str):
        """拉指定日期的 BOM 清单，返回以 IDNRK 为 key 的字典"""
        try:
            with pyrfc.Connection(ashost=sap_h, sysnr=sap_s, client=sap_c,
                                  user=sap_u, passwd=sap_p, lang='EN') as conn:
                res = conn.call('ZPP_MES_BOM_EMS', IV_WERKS='1201',
                                IV_FBSTP='0', IV_EXPLO_DATE=date_str,
                                IT_ORDERS=[{'VBELN': '0' + projectID,
                                            'VBPOS': str(item_posnr).zfill(6)}])
                return {row['IDNRK']: row for row in res['ET_COMPONENTS']
                        if row['STUFE'] == 1}
        except Exception as e:
            logging.error(e)
            return {}

    def diff_bom(old_dict, new_dict):
        """返回 (新增列表, 减少列表)"""
        added = [new_dict[k] for k in new_dict if k not in old_dict]
        removed = [old_dict[k] for k in old_dict if k not in new_dict]
        return added, removed

    def export_pdf():
        pdfmetrics.registerFont(TTFont('SimSun', 'simsun.ttc'))  # 注册字体
        home_path = os.path.expanduser("~")
        desktop_path = os.path.join(home_path, "Desktop")
        desktop_path = askdirectory(title=u'请选择导出文件夹', initialdir=desktop_path, parent=top_compare_ebom)
        if not desktop_path:
            tk.messagebox.showwarning("提示", "未选择保存路径，导出操作已取消", parent=top_compare_ebom)
            return

        story = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'Title', parent=styles['Normal'],
            fontName='SimSun', fontSize=8, spaceAfter=0)
        sub_style = ParagraphStyle(
            'Sub', parent=styles['Normal'],
            fontName='SimSun', fontSize=8, spaceAfter=0)

        col_widths = [40 * mm, 80 * mm, 15 * mm, 15 * mm, 20 * mm, 30 * mm, 25 * mm, 25 * mm, 25 * mm]  # 最多 297 mm

        global page_number
        page_number = 1

        material_type = []

        # 逐 Item 生成一页
        for idx, (posnr, material, added, removed) in enumerate(diff_cache):
            if material not in material_type:
                material_type.append(material)
            # 页头
            story.append(Paragraph(f'Item {posnr} 差异对比', title_style))
            story.append(Paragraph(f'旧日期：{old_date}   →   新日期：{new_date}', sub_style))
            # story.append(Spacer(1, 6))

            # ---- 新增表 ----
            story.append(Paragraph('新增物料', sub_style))
            add_data = [['物料号', '描述', '定位', '数量', '创建日期', '变更人', '变更号', '变更生效日期', '变更失效日期']]
            for r in added:
                add_data.append([
                    r['IDNRK'],  # b 物料号
                    r['MAKTX'],  # c 描述
                    r['SORTF'],  # d 定位
                    str(int(r['MENGE'])),  # e 数量
                    r['DATUV1'],  # f 创建日期
                    r['ANNAM'],  # g 变更人
                    r['AENNR'],  # h 变更号
                    r['DATUV'],  # i 生效日期
                    r['DATUB']  # j 失效日期
                    ])
            add_t = Table(add_data, colWidths=col_widths)
            add_t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'SimSun'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            story.append(add_t)
            # story.append(Spacer(1, 12))

            # ---- 减少表 ----
            story.append(Paragraph('减少物料', sub_style))
            rem_data = [['物料号', '描述', '定位', '数量', '创建日期', '变更人', '变更号', '变更生效日期', '变更失效日期']]
            for r in removed:
                rem_data.append([
                    r['IDNRK'], r['MAKTX'], r['SORTF'],
                    str(int(r['MENGE'])), r['DATUV1'], r['ANNAM'],
                    r['AENNR'], r['DATUV'], r['DATUB']
                ])
            rem_t = Table(rem_data, colWidths=col_widths)
            rem_t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightcoral),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'SimSun'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            story.append(rem_t)
            story.append(PageBreak())   # 每 Item 一页

        def footer(r_canvas, doc):
            r_canvas.saveState()
            r_canvas.setFont('SimSun', 8)
            global page_number
            page_number_text = "Page " + str(page_number)
            page_number += 1
            r_canvas.drawString(landscape(A4)[0]/2, 10, page_number_text)

            Timestamp = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
            r_canvas.drawCentredString(landscape(A4)[0]-90, 10, '制表时间：' + Timestamp)
            r_canvas.restoreState()

        result = '_'.join(str(s) for s in material_type)
        file_path = os.path.join(desktop_path, f'SAP的BOM({result})差异对比报表-{projectID}.pdf')
        doc = SimpleDocTemplate(file_path, pagesize=landscape(A4), rightMargin=10 * mm, leftMargin=10 * mm,
                                topMargin=10 * mm, bottomMargin=10 * mm)
        doc.build(story, onFirstPage=footer, onLaterPages=footer)
        tk.messagebox.showwarning('完成', f'PDF 已导出至：\n{file_path}', parent=top_compare_ebom)

    diff_cache = []          # [(item_posnr, added_list, removed_list), ...]
    for row in compare_rows:
        posnr = int(row[0])
        material = (row[1])
        old_bom = fetch_bom(posnr, old_date)
        new_bom = fetch_bom(posnr, new_date)
        added, removed = diff_bom(old_bom, new_bom)
        diff_cache.append((posnr, material, added, removed))

    top_compare_ebom = tk.Toplevel(root_win, bg='#eaf1f6')
    top_compare_ebom.title(f'BOM差异对比(逐Item)')
    winw = 1350
    winh = 800

    top_compare_ebom.geometry("%dx%d" % (winw, winh))
    center_window(top_compare_ebom)
    top_compare_ebom.grab_set()

    def return_grab():
        top_compare_ebom.grab_release()
        global _active_selector
        if '_active_selector' in globals() and _active_selector and _active_selector.winfo_exists():
            _active_selector.grab_set()

    top_compare_ebom.protocol("WM_DELETE_WINDOW",
                              lambda: (return_grab(), top_compare_ebom.destroy()))

    page_idx = 0
    page_var = tk.IntVar(value=1)
    total_pages = len(diff_cache)

    info_lbl = tk.Label(top_compare_ebom, text='', font=("ABBvoice CNSG", int(14 * h_ratio)), bg='#eaf1f6')
    info_lbl.pack(side='top', fill='x', padx=10, pady=2)

    # 新增表
    tk.Label(top_compare_ebom, text='新增物料（新日期有，旧日期无）', font=("ABBvoice CNSG", int(11 * h_ratio)),
             bg='#eaf1f6', fg='green').pack(anchor='w', padx=10, pady=(5, 0))

    add_frame = tk.Frame(top_compare_ebom)
    add_frame.pack(fill='both', expand=True, padx=0, pady=5)
    add_scroll = tk.Scrollbar(add_frame)

    add_tree = Treeview(add_frame, show="headings", style='panel1.Treeview', columns=['b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j'], selectmode='browse', height=10, yscrollcommand=add_scroll.set)
    add_tree.tag_configure('data', font=("ABBvoice CNSG", int(10 * h_ratio)))
    add_scroll.config(command=add_tree.yview)
    add_tree.heading('b', text='物料号', anchor='w')
    add_tree.heading('c', text='描述', anchor='w')
    add_tree.heading('d', text='定位', anchor='w')
    add_tree.heading('e', text='数量', anchor='w')
    add_tree.heading('f', text='创建日期', anchor='w')
    add_tree.heading('g', text='变更人', anchor='w')
    add_tree.heading('h', text='变更号', anchor='w')
    add_tree.heading('i', text='变更生效日期', anchor='w')
    add_tree.heading('j', text='变更失效日期', anchor='w')

    add_tree.column('b', width=200, anchor='w')
    add_tree.column('c', width=400, anchor='w')
    add_tree.column('d', width=50, anchor='w')
    add_tree.column('e', width=50, anchor='w')
    add_tree.column('f', width=100, anchor='w')
    add_tree.column('g', width=120, anchor='w')
    add_tree.column('h', width=120, anchor='w')
    add_tree.column('i', width=100, anchor='w')
    add_tree.column('j', width=100, anchor='w')

    add_tree.tag_configure('bg_color', background="#c9dbe9")
    add_tree.bind("<Control-c>",
                         lambda e: copy_cell_ctrl_c(e, add_tree))
    add_tree.pack(side='left', fill='both', expand=True, padx=(10,0), pady=2)
    add_scroll.pack(side='left', fill='y')
    # 减少表
    tk.Label(top_compare_ebom, text='减少物料（旧日期有，新日期无）', font=("ABBvoice CNSG", int(11 * h_ratio)),
             bg='#eaf1f6', fg='red').pack(anchor='w', padx=10, pady=(5, 0))

    rem_frame = tk.Frame(top_compare_ebom)
    rem_frame.pack(fill='both', expand=True, padx=0, pady=5)
    rem_scroll = tk.Scrollbar(rem_frame)

    rem_tree = Treeview(rem_frame, show="headings", style='panel1.Treeview', columns=['b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j'], selectmode='browse', height=10, yscrollcommand= rem_scroll.set)
    rem_tree.tag_configure('data', font=("ABBvoice CNSG", int(10 * h_ratio)))
    rem_scroll.config(command=rem_tree.yview)

    rem_tree.heading('b', text='物料号', anchor='w')
    rem_tree.heading('c', text='描述', anchor='w')
    rem_tree.heading('d', text='定位', anchor='w')
    rem_tree.heading('e', text='数量', anchor='w')
    rem_tree.heading('f', text='创建日期', anchor='w')
    rem_tree.heading('g', text='变更人', anchor='w')
    rem_tree.heading('h', text='变更号', anchor='w')
    rem_tree.heading('i', text='变更生效日期', anchor='w')
    rem_tree.heading('j', text='变更失效日期', anchor='w')

    rem_tree.column('b', width=200, anchor='w')
    rem_tree.column('c', width=400, anchor='w')
    rem_tree.column('d', width=50, anchor='w')
    rem_tree.column('e', width=50, anchor='w')
    rem_tree.column('f', width=100, anchor='w')
    rem_tree.column('g', width=120, anchor='w')
    rem_tree.column('h', width=120, anchor='w')
    rem_tree.column('i', width=100, anchor='w')
    rem_tree.column('j', width=100, anchor='w')

    rem_tree.tag_configure('bg_color', background="#c9dbe9")
    rem_tree.bind("<Control-c>",
                         lambda e: copy_cell_ctrl_c(e, rem_tree))
    rem_tree.pack(side='left', fill='both', expand=True, padx=(10,0), pady=2)
    rem_scroll.pack(side='left', fill='y')

    bottom_btn_frame = tk.Frame(top_compare_ebom, bg='#eaf1f6')
    bottom_btn_frame.pack(side='bottom', fill='x', padx=10, pady=(0, 10))

    global checklog_file_path
    checklogbook = load_workbook(checklog_file_path)
    checklogsheet = checklogbook['Sheet']

    current_time = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    data1 = [projectID, 'SAP中BOM对比', 0, current_time]
    checklogsheet.append(data1)
    checklogbook.save(checklog_file_path)

    tk.Button(bottom_btn_frame, text='◀ 上一项', bg="#eaf1f6", font=("ABBvoice CNSG", int(12 * h_ratio)),
              command=lambda: change_page(-1)).pack(side='left')
    tk.Label(bottom_btn_frame, textvariable=page_var, font=("ABBvoice CNSG", int(12 * h_ratio)),
             bg='#eaf1f6').pack(side='left', padx=(10, 0))
    tk.Label(bottom_btn_frame, text=f'/ {total_pages}', font=("ABBvoice CNSG", int(12 * h_ratio)),
             bg='#eaf1f6').pack(side='left', padx=(0, 10))
    tk.Button(bottom_btn_frame, text='下一项 ▶', bg="#eaf1f6", font=("ABBvoice CNSG", int(12 * h_ratio)),
              command=lambda: change_page(1)).pack(side='left')
    tk.Button(bottom_btn_frame, text='导出PDF', bg="#eaf1f6", compound=tk.LEFT, image=sap_ebom_compare_export, font=("ABBvoice CNSG", int(12 * h_ratio)),
              command=export_pdf).pack(side='right', padx=20)

    def change_page(delta):
        nonlocal page_idx
        new_idx = page_idx + delta
        if 0 <= new_idx < total_pages:
            page_idx = new_idx
            page_var.set(page_idx + 1)
            show_current_page()

    def show_current_page():
        """清空两表，载入当前页差异"""
        add_tree.delete(*add_tree.get_children())
        rem_tree.delete(*rem_tree.get_children())

        posnr, material, added, removed = diff_cache[page_idx]
        info_lbl.config(text=f'当前Item：{posnr}  （旧日期 {old_date}  →  新日期 {new_date}）')

        # 填新增表
        for row in added:
            add_tree.insert('', 'end', values=(
                row['IDNRK'], row['MAKTX'], row['SORTF'], int(row['MENGE']), row['DATUV1'], row['ANNAM'], row['AENNR'], row['DATUV'], row['DATUB']), tags='data')

        # 填减少表
        for row in removed:
            rem_tree.insert('', 'end', values=(
                row['IDNRK'], row['MAKTX'], row['SORTF'], int(row['MENGE']), row['DATUV1'], row['ANNAM'], row['AENNR'], row['DATUV'], row['DATUB']), tags='data')

    page_var.set(1)
    show_current_page()


def export_ebom():
    if len(attention_b_rows) == 0:
        tk.messagebox.showwarning("提示", "无BOM(EBOM/PR/ACC-EL/AA)，无法导出BOM清单")
        return
    sel = PosnrSelector(root_win, h_ratio, attention_b_rows, attention_rows, mode='ebom')
    root_win.wait_window(sel)  # 等待窗口关闭


def do_export_ebom(sel):
    selec_rows = [attention_b_rows[i] for i in sel.selected]
    if not selec_rows:
        tk.messagebox.showwarning("提示", "未选择任何条目，导出已取消", parent=root_win)
        return

    def get_ebom(item_posnr):
        try:
            with pyrfc.Connection(ashost=sap_h, sysnr=sap_s, client=sap_c, user=sap_u, passwd=sap_p, lang='EN') as conn:
                res = conn.call('ZPP_MES_BOM_EMS', IV_WERKS='1201', IV_FBSTP='0', IT_ORDERS=[{'VBELN': '0' + projectID, 'VBPOS': str(item_posnr).zfill(6)}])
                return [row for row in res['ET_COMPONENTS'] if row['STUFE'] == 1]
        except Exception as e:
            logging.error(e)
            return {}

    def export_pdf():
        pdfmetrics.registerFont(TTFont('SimSun', 'simsun.ttc'))  # 注册字体
        home_path = os.path.expanduser("~")
        desktop_path = os.path.join(home_path, "Desktop")
        desktop_path = askdirectory(title=u'请选择导出文件夹', initialdir=desktop_path, parent=top_export_ebom)
        if not desktop_path:
            tk.messagebox.showwarning("提示", "未选择保存路径，导出操作已取消", parent=top_export_ebom)
            return

        story = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'Title', parent=styles['Normal'],
            fontName='SimSun', fontSize=8, spaceAfter=0)
        sub_style = ParagraphStyle(
            'Sub', parent=styles['Normal'],
            fontName='SimSun', fontSize=8, spaceAfter=0)

        col_widths = [40 * mm, 80 * mm, 15 * mm, 15 * mm, 20 * mm, 30 * mm, 25 * mm, 25 * mm, 25 * mm]  # 最多 297 mm

        global page_number
        page_number = 1

        material_type = []

        # 逐 Item 生成一页
        for idx, (posnr, material, typical, panel_number, data_m) in enumerate(item_cache):
            if material not in material_type:
                material_type.append(material)
            # 页头
            story.append(Paragraph(f'BOM清单——Item：{posnr}， 对应柜型：{typical}， 柜号：{panel_number}', title_style))

            # story.append(Spacer(1, 6))

            col_data = [['物料号', '描述', '定位', '数量', '创建日期', '变更人', '变更号', '变更生效日期', '变更失效日期']]
            for r in data_m:
                col_data.append([
                    r['IDNRK'],  # b 物料号
                    r['MAKTX'],  # c 描述
                    r['SORTF'],  # d 定位
                    str(int(r['MENGE'])),  # e 数量
                    r['DATUV1'],  # f 创建日期
                    r['ANNAM'],  # g 变更人
                    r['AENNR'],  # h 变更号
                    r['DATUV'],  # i 生效日期
                    r['DATUB']  # j 失效日期
                ])
            col_data_t = Table(col_data, colWidths=col_widths)
            col_data_t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'SimSun'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            story.append(col_data_t)
            # story.append(Spacer(1, 12))

            story.append(PageBreak())  # 每 Item 一页

        def footer(r_canvas, doc):
            r_canvas.saveState()
            r_canvas.setFont('SimSun', 8)
            global page_number
            page_number_text = "Page " + str(page_number)
            page_number += 1
            r_canvas.drawString(landscape(A4)[0] / 2, 10, page_number_text)

            Timestamp = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
            r_canvas.drawCentredString(landscape(A4)[0] - 90, 10, '制表时间：' + Timestamp)
            r_canvas.restoreState()

        result = '_'.join(str(s) for s in material_type)
        file_path = os.path.join(desktop_path, f'SAP的BOM({result})报表-{projectID}.pdf')
        doc = SimpleDocTemplate(file_path, pagesize=landscape(A4), rightMargin=10 * mm, leftMargin=10 * mm,
                                topMargin=10 * mm, bottomMargin=10 * mm)
        doc.build(story, onFirstPage=footer, onLaterPages=footer)
        tk.messagebox.showwarning('完成', f'PDF 已导出至：\n{file_path}', parent=top_export_ebom)

    item_cache = []
    for row in selec_rows:
        posnr = int(row[0])
        material = row[1]
        typical = row[3]
        panel_number = row[4]
        data_m = get_ebom(posnr)
        item_cache.append((posnr, material, typical, panel_number, data_m))

    top_export_ebom = tk.Toplevel(root_win, bg='#eaf1f6')
    top_export_ebom.title(f'BOM清单(逐Item)')
    winw = 1350
    winh = 800
    top_export_ebom.geometry('%dx%d' % (winw, winh))
    center_window(top_export_ebom)
    top_export_ebom.grab_set()

    def return_grab():
        top_export_ebom.grab_release()  # 先释放
        global _active_selector
        if '_active_selector' in globals() and _active_selector and _active_selector.winfo_exists():
            _active_selector.grab_set()  # 还给自己

    top_export_ebom.protocol("WM_DELETE_WINDOW",
                             lambda: (return_grab(), top_export_ebom.destroy()))

    page_idx = 0
    page_var = tk.IntVar(value=1)
    total_pages = len(item_cache)

    info_lbl = tk.Label(top_export_ebom, text='', font=("ABBvoice CNSG", int(14 * h_ratio)), bg='#eaf1f6')
    info_lbl.pack(side='top', fill='x', padx=10, pady=2)

    tree_frame = tk.Frame(top_export_ebom)
    tree_frame.pack(fill='both', expand=True, padx=0, pady=5)
    tree_scroll = tk.Scrollbar(tree_frame)

    tree = Treeview(tree_frame, show="headings", style='panel1.Treeview', columns=['b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j'], selectmode='browse', height=20, yscrollcommand=tree_scroll.set)
    tree.tag_configure('data', font=("ABBvoice CNSG", int(10 * h_ratio)))
    tree_scroll.config(command=tree.yview)
    tree.heading('b', text='物料号', anchor='w')
    tree.heading('c', text='描述', anchor='w')
    tree.heading('d', text='定位', anchor='w')
    tree.heading('e', text='数量', anchor='w')
    tree.heading('f', text='创建日期', anchor='w')
    tree.heading('g', text='变更人', anchor='w')
    tree.heading('h', text='变更号', anchor='w')
    tree.heading('i', text='变更生效日期', anchor='w')
    tree.heading('j', text='变更失效日期', anchor='w')

    tree.column('b', width=200, anchor='w')
    tree.column('c', width=400, anchor='w')
    tree.column('d', width=50, anchor='w')
    tree.column('e', width=50, anchor='w')
    tree.column('f', width=100, anchor='w')
    tree.column('g', width=120, anchor='w')
    tree.column('h', width=120, anchor='w')
    tree.column('i', width=100, anchor='w')
    tree.column('j', width=100, anchor='w')

    tree.tag_configure('bg_color', background="#c9dbe9")
    tree.bind("<Control-c>",
                         lambda e: copy_cell_ctrl_c(e, tree))
    tree.pack(side='left', fill='both', expand=True, padx=(10, 0), pady=2)
    tree_scroll.pack(side='left', fill='y')

    bottom_btn_frame = tk.Frame(top_export_ebom, bg='#eaf1f6')
    bottom_btn_frame.pack(side='bottom', fill='x', padx=10, pady=(0, 10))

    tk.Button(bottom_btn_frame, text='◀ 上一项', bg="#eaf1f6", font=("ABBvoice CNSG", int(12 * h_ratio)),
              command=lambda: change_page(-1)).pack(side='left')
    tk.Label(bottom_btn_frame, textvariable=page_var, font=("ABBvoice CNSG", int(12 * h_ratio)),
             bg='#eaf1f6').pack(side='left', padx=(10, 0))
    tk.Label(bottom_btn_frame, text=f'/ {total_pages}', font=("ABBvoice CNSG", int(12 * h_ratio)),
             bg='#eaf1f6').pack(side='left', padx=(0, 10))
    tk.Button(bottom_btn_frame, text='下一项 ▶', bg="#eaf1f6", font=("ABBvoice CNSG", int(12 * h_ratio)),
              command=lambda: change_page(1)).pack(side='left')
    tk.Button(bottom_btn_frame, text='导出PDF', bg="#eaf1f6", compound=tk.LEFT, image=sap_ebom_compare_export, font=("ABBvoice CNSG", int(12 * h_ratio)),
              command=export_pdf).pack(side='right', padx=20)

    def change_page(delta):
        nonlocal page_idx
        new_idx = page_idx + delta
        if 0 <= new_idx < total_pages:
            page_idx = new_idx
            page_var.set(page_idx + 1)
            show_current_page()

    def show_current_page():
        tree.delete(*tree.get_children())
        posnr, material, typical, panel_number, data_m = item_cache[page_idx]
        info_lbl.config(text=f'当前Item：{posnr}， 对应柜型：{typical}， 柜号：{panel_number}')

        for row in data_m:
            tree.insert('', 'end', values=(
                row['IDNRK'], row['MAKTX'], row['SORTF'], int(row['MENGE']), row['DATUV1'], row['ANNAM'], row['AENNR'], row['DATUV'], row['DATUB']), tags='data')

    page_var.set(1)
    show_current_page()


def on_double_click(event):
    item = Item_Info_table.selection()[0]
    text = Item_Info_table.item(item, 'text')
    values = Item_Info_table.item(item, 'values')

    base = (int(text) // 1000) * 1000
    offset = int(text) - base

    if (501 <= offset <= 549 and ('CDX' in str(values[0]) or 'GCE' in str(values[0]) or 'HD4' in str(values[0]))):
        # root_win.attributes("-disabled", 1)
        global top1
        top1 = tk.Toplevel()
        winw = 600
        winh = 400

        top1.geometry("%dx%d" % (winw, winh))
        center_window(top1)
        top1.grab_set()
        # top1.protocol("WM_DELETE_WINDOW", allow_main_window1)
        f1 = tk.Frame(top1, bg="#eaf1f6", bd=0)
        f1.pack(side=tk.TOP, fill=tk.X)

        top1.title(f'{text}_{str(values[0])}_断路器选配，SAP未选配属性不显示')

        scrollbar_cbbom = tk.Scrollbar(f1)

        cbbom_configuration_table = Treeview(f1, style='panel1.Treeview', show="headings", columns=['a', 'b'], selectmode='browse', height=20, yscrollcommand=scrollbar_cbbom.set)
        scrollbar_cbbom.config(command=cbbom_configuration_table.yview)
        cbbom_configuration_table.heading('a', text='属性', anchor='w')
        cbbom_configuration_table.heading('b', text='值', anchor='w')

        cbbom_configuration_table.column('a', width=100, anchor='w')
        cbbom_configuration_table.column('b', width=100, anchor='w')
        cbbom_configuration_table.tag_configure('bg_color', background="#c9dbe9")
        cbbom_configuration_table.pack(side='left', fill='both', expand=True, anchor='w', padx=(10, 0))
        cbbom_configuration_table.bind("<Control-c>",
                             lambda e: copy_cell_ctrl_c(e, cbbom_configuration_table))
        scrollbar_cbbom.pack(side='left', fill='y')

        logging.info("Ready to connect to SAP")
        try:
            conn = pyrfc.Connection(ashost=sap_h, sysnr=sap_s, client=sap_c, user=sap_u, passwd=sap_p, lang='EN')
            if conn.alive:
                logging.info("Connecting to SAP successfully")
                item_id = str(text).zfill(6)

                cb_configuration_result = conn.call('ZSD_GET_SO_CONFIGURATION', IM_VBELN='0'+projectID, IM_POSNR=item_id)

                if cb_configuration_result['ET_CONF_OUT']:
                    for confi_item in cb_configuration_result['ET_CONF_OUT']:
                        Char_description = confi_item['ATBEZ']
                        Char_value = confi_item['ATWTB']
                        if 'VDN_' not in Char_description and '_TABLE' not in Char_description:
                            cbbom_configuration_table.insert('', 'end', values=(Char_description, Char_value), tags='bg_color')

            conn.close()
            if not conn.alive:
                logging.info("Disconnect from SAP")


        except pyrfc.RFCError as e:
            logging.info(e.key + ', ' + e.message)
            tk.messagebox.showwarning("提示", traceback.format_exc())

    elif (801 <= offset <= 899 and 'EBOM' in str(values[0])) or (951 <= offset <= 980 and 'ACC-EL' == str(values[0])) or (991 <= offset <= 998 and str(values[0]) == 'PR') or (750 <= offset <= 799 and str(values[0]) == 'AA') or 'ACC-EL' == str(values[0]):
        # root_win.attributes("-disabled", 1)
        global top2
        top2 = tk.Toplevel()
        winw = 1350
        winh = 800

        top2.geometry("%dx%d" % (winw, winh))
        center_window(top2)
        top2.grab_set()
        # top2.protocol("WM_DELETE_WINDOW", allow_main_window2)
        f1 = tk.Frame(top2, bg="#eaf1f6", bd=0)
        f1.pack(side=tk.TOP, fill=tk.X)

        top2.title(f'{text}_{str(values[0])}_BOM信息')

        scrollbar_bom = tk.Scrollbar(f1)

        bom_table = Treeview(f1, show="headings", style='panel1.Treeview', columns=['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j'], selectmode='browse', height=40, yscrollcommand=scrollbar_bom.set)
        scrollbar_bom.config(command=bom_table.yview)
        bom_table.heading('a', text='Id', anchor='w')
        bom_table.heading('b', text='物料号', anchor='w')
        bom_table.heading('c', text='描述', anchor='w')
        bom_table.heading('d', text='定位', anchor='w')
        bom_table.heading('e', text='数量', anchor='w')
        bom_table.heading('f', text='创建日期', anchor='w')
        bom_table.heading('g', text='变更人', anchor='w')
        bom_table.heading('h', text='变更号', anchor='w')
        bom_table.heading('i', text='变更生效日期', anchor='w')
        bom_table.heading('j', text='变更失效日期', anchor='w')


        bom_table.column('a', width=50, anchor='w')
        bom_table.column('b', width=200, anchor='w')
        bom_table.column('c', width=400, anchor='w')
        bom_table.column('d', width=50, anchor='w')
        bom_table.column('e', width=50, anchor='w')
        bom_table.column('f', width=100, anchor='w')
        bom_table.column('g', width=120, anchor='w')
        bom_table.column('h', width=120, anchor='w')
        bom_table.column('i', width=100, anchor='w')
        bom_table.column('j', width=100, anchor='w')

        bom_table.tag_configure('bg_color', background="#c9dbe9")
        bom_table.bind("<Control-c>",
                             lambda e: copy_cell_ctrl_c(e, bom_table))
        bom_table.pack(side='left', fill='both', expand=True, anchor='w', padx=(10,0))
        scrollbar_bom.pack(side='left', fill='y')

        logging.info("Ready to connect to SAP")
        try:
            conn = pyrfc.Connection(ashost=sap_h, sysnr=sap_s, client=sap_c, user=sap_u, passwd=sap_p, lang='EN')
            if conn.alive:
                logging.info("Connecting to SAP successfully")
                item_id = str(text).zfill(6)
                rfc_table = [{'VBELN': '0' + projectID, 'VBPOS': item_id}]
                bom_result = conn.call('ZPP_MES_BOM_EMS', IV_WERKS='1201', IV_FBSTP='0', IT_ORDERS=rfc_table)    # , IV_EXPLO_DATE='20250928'指定BOM日期
                if bom_result['ET_COMPONENTS']:
                    for bom_item in bom_result['ET_COMPONENTS']:
                        if bom_item['STUFE'] == 1:
                            bom_table.insert('', 'end', values=(bom_item['POSNR'], bom_item['IDNRK'], bom_item['MAKTX'], bom_item['SORTF'], str(int(bom_item['MENGE'])), bom_item['DATUV1'], bom_item['ANNAM'], bom_item['AENNR'], bom_item['DATUV'], bom_item['DATUB']), tags='bg_color')

            conn.close()
            if not conn.alive:
                logging.info("Disconnect from SAP")

        except pyrfc.RFCError as e:
            logging.info(e.key + ', ' + e.message)
            tk.messagebox.showwarning("提示", traceback.format_exc())


def sap_decrypt_data(data, data_cipher, prefix):
    # 移除列名前缀prefix，如果有的话
    if data.startswith(prefix):
        data = data[2:]

    # Base64 字符串中可能删除了等号，需要添加回原来的填充字符，因为 base64 编码的输出长度应该是 4 的倍数
    padding_needed = 4 - (len(data) % 4)
    if padding_needed:
        data += "=" * padding_needed

    # base64 解码以获取原始的加密字节串
    encrypted_data = base64.urlsafe_b64decode(data.encode())

    # 使用 data_cipher 对象解密，它应该有和加密时一样的密钥
    decrypted_data = data_cipher.decrypt(encrypted_data)

    # 将字节解码回字符串
    decrypted_string = decrypted_data.decode()
    return decrypted_string


def save_history():
    global project_history, MAX_HISTORY
    # 先排好序并截断
    project_history.sort(key=lambda x: x['time'], reverse=True)
    project_history[:] = project_history[:MAX_HISTORY]
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(project_history, f, ensure_ascii=False, indent=4)


def refresh_history_combo():
    load_history()
    """把历史记录同步到下拉框"""
    combobox_project_number['values'] = [x['project'] for x in project_history]


def load_history():
    global project_history, MAX_HISTORY
    project_history.clear()
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
            project_history.extend(json.load(f))
    project_history.sort(key=lambda x: x['time'], reverse=True)
    project_history[:] = project_history[:MAX_HISTORY]


def on_history_select(evt):
    """选中历史记录后自动填入并查询"""
    idx = combobox_project_number.current()
    project = combobox_project_number['values'][idx]


class PosnrSelector(tk.Toplevel):
    def __init__(self, parent, h_ratio, a_rows, all_rows, mode='cb'):
        super().__init__(parent)
        self.mode = mode
        self.selected = []  # 返回给主线程的结果
        self.old_date = None
        self.new_date = None

        global _active_selector
        _active_selector = self

        # 根据不同模式决定标题
        title_map = {'cb': '导出断路器选配，请选择Item',
                     'ebom': '导出BOM，请选择Item',
                     'compare': '对比BOM，请选择Item'}
        self.title(title_map.get(mode))
        self.geometry('1350x800')
        center_window(self)
        self.resizable(False, False)
        self.grab_set()  # 模态
        self.configure(bg="#eaf1f6")

        # --- 记录数据 ---
        self.a_rows = a_rows
        self.all_rows = all_rows

        # --- 上侧工具栏 ---
        tool = tk.Frame(self, bg="#eaf1f6")
        tool.pack(side='top', fill='x', padx=10, pady=10)

        self.btn_fold = tk.Button(tool, image=level_fold, bg="#eaf1f6",
                                  command=lambda: self.expand_tree(0))
        self.btn_fold.pack(side='left', padx=(0, 5))
        self.btn_unfold = tk.Button(tool, image=level_unfold, bg="#eaf1f6",
                                    command=lambda: self.expand_tree(2))
        self.btn_unfold.pack(side='left', padx=(0, 5))

        Tooltip(self.btn_fold, "折叠下表Item")
        Tooltip(self.btn_unfold, "展开下表Item")

        # ===== 新增：Item 范围快速选择 =====
        tk.Label(tool, text='  Item 范围：', bg='#eaf1f6', font=("ABBvoice CNSG", int(11 * h_ratio))).pack(side='left', padx=(15,0))
        self.from_var = tk.StringVar()
        self.to_var   = tk.StringVar()
        item_low = tk.Entry(tool, textvariable=self.from_var, width=6, font=("ABBvoice CNSG", int(11 * h_ratio)))
        item_low.pack(side='left')

        tk.Label(tool, text='-', font=("ABBvoice CNSG", int(11 * h_ratio)), bg='#eaf1f6').pack(side='left')

        item_high = tk.Entry(tool, textvariable=self.to_var,   width=6, font=("ABBvoice CNSG", int(11 * h_ratio)))
        item_high.pack(side='left')
        Tooltip(item_low, "从某个行号起")
        Tooltip(item_high, "到某个行号(不能跨行)")

        tk.Button(tool, text='按范围选取Item', compound=tk.LEFT, image=sap_item_select, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)),
                  command=self.select_by_range).pack(side='left', padx=5)

        # # --- 下侧按钮 ---
        # btn_frame = tk.Frame(self, bg="#eaf1f6")
        # btn_frame.pack(fill='x', padx=10, pady=5)
        btn_name_map = {'cb': '导出断路器选配',
                     'ebom': '导出BOM',
                     'compare': '对比BOM'}
        tk.Button(tool, text=btn_name_map.get(mode), bg="#eaf1f6", compound=tk.LEFT, image=sap_ebom_compare, font=("ABBvoice CNSG", int(11 * h_ratio)),
                                                command=self.on_ok, activebackground='blue').pack(side='right', padx=20)

        # --- 中间树 + 滚动条 ---
        list_frame = tk.Frame(self)
        list_frame.pack(fill='both', expand=True, padx=10, pady=5)
        scrollbar = tk.Scrollbar(list_frame)

        self.tree = ttk.Treeview(list_frame, style='panel1.Treeview',
                                 columns=('a', 'b', 'c', 'd', 'e'), height=18,
                                 yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.tree.yview)

        Tooltip(self.tree, "可在'选中状态'列勾选对应的Item")

        self.tree.tag_configure('select_row', background='orange', font=("ABBvoice CNSG", int(11 * h_ratio)))
        self.tree.tag_configure('unselect_row', background='whitesmoke', font=("ABBvoice CNSG", int(11 * h_ratio)))

        self.tree.heading('#0', text='选中状态', anchor='w')
        self.tree.heading('a', text='Item', anchor='w')
        self.tree.heading('b', text='Material', anchor='w')
        self.tree.heading('c', text='描述数量', anchor='w')
        self.tree.heading('d', text='描述', anchor='w')
        self.tree.heading('e', text='柜号', anchor='w')
        self.tree.column('#0', width=int(50 * w_ratio), anchor='w')
        self.tree.column('a', width=int(120 * w_ratio), anchor='w')
        self.tree.column('b', width=int(200 * w_ratio), anchor='w')
        self.tree.column('c', width=int(50 * w_ratio), anchor='w')
        self.tree.column('d', width=int(400 * w_ratio), anchor='w')
        self.tree.column('e', width=int(600 * w_ratio), anchor='w')

        self.tree.bind("<Control-c>",
                             lambda e: copy_cell_ctrl_c(e, self.tree))
        self.tree.pack(side='left', fill='both', expand=True, padx=(0,0))
        scrollbar.pack(side='left', fill='y')
        self.tree.tag_configure('unselect_row', background='whitesmoke')

        # --- 数据初始化 ---
        self.vars = {}  # posnr -> BooleanVar
        self.build_tree()
        self.tree.bind('<ButtonRelease-1>', self.on_click_item)
        self.expand_tree(0)  # 默认折叠

    def select_by_range(self):
        """按填写的 Item 范围批量勾选"""
        frm = self.from_var.get().strip()
        to  = self.to_var.get().strip()
        if not frm.isdigit() or not to.isdigit():
            tk.messagebox.showwarning('提示', '请输入数字！', parent=self)
            return
        frm, to = int(frm), int(to)
        if frm > to:
            tk.messagebox.showwarning('提示', '起始不能大于结束！', parent=self)
            return
        # 仅最后两位不同
        if frm // 100 != to // 100:
            tk.messagebox.showwarning('提示', '范围要求前几位一致，仅最后两位不同！', parent=self)
            return

        # 先全部置空
        # for posnr_str, var in self.vars.items():
        #     var.set(False)

        base_posnr = (frm // 1000) * 1000

        hit = 0
        for posnr in range(frm, to + 1):
            posnr_str = str(posnr)
            if posnr_str in self.vars:  # 可勾选才处理
                self.vars[posnr_str].set(True)
                hit += 1  # 实际命中才计数

        self.refresh_check_display(base_posnr)
        tk.messagebox.showwarning('提示', f'已选中 {hit} 条记录', parent=self)

    def refresh_check_display(self, base_posnr=None):
        if base_posnr is None:
            parent_iids = list(self.tree.get_children())
        else:
            parent_iids = []
            for root_iid in self.tree.get_children():
                vals = self.tree.item(root_iid, 'values')
                if vals and int(vals[0]) == base_posnr:  # ✅ 用 values[0]
                    parent_iids.append(root_iid)
                    break
            if not parent_iids:
                return

        # 刷新这些父节点下的子节点
        for p_iid in parent_iids:
            self.expand_tree_node(p_iid, level=2)

            for leaf_iid in self.tree.get_children(p_iid):
                vals = self.tree.item(leaf_iid, 'values')
                if not vals:
                    continue
                try:
                    posnr_str = str(int(vals[0]))
                except ValueError:
                    continue
                if posnr_str in self.vars:
                    checked = self.vars[posnr_str].get()
                    self.tree.item(leaf_iid, text='✓' if checked else '□')
                    self.tree.item(leaf_iid, tags='select_row' if checked else 'unselect_row')

    def build_tree(self):
        self.parent_nodes = {}

        for item in self.all_rows:
            posnr = int(item[0])

            if posnr % 1000 == 0:
                insert_data = [item[0], item[1], item[2], item[3], item[4]]
                parent_id = self.tree.insert("", "end", values=insert_data, tags='unselect_row')
                self.parent_nodes[posnr] = parent_id

            else:
                parent_id = self.parent_nodes.get((posnr // 1000) * 1000)
                if parent_id:
                    # 获取 posnr 的千位数
                    base = (posnr // 1000) * 1000
                    # 计算 posnr 的偏移量
                    offset = posnr - base

                    # 判断使用哪种tag
                    if self.mode == 'cb':
                        if 501 <= offset <= 549 and ('CDX' in str(item[1]) or 'GCE' in str(item[1]) or 'HD4' in str(item[1])):
                            insert_data = [item[0], item[1], item[2], item[3], item[4]]
                            self.tree.insert(parent_id, "end", text='□', values=insert_data, tags='unselect_row')

                    elif self.mode == 'ebom':
                        if (750 <= offset <= 799 and 'AA' in str(item[1])) or ((801 <= offset <= 899 and 'EBOM' in str(item[1])) or (951 <= offset <= 980 and 'ACC-EL' == str(item[1]))) or (991 <= offset <= 998 and str(item[1]) == 'PR') or (base == 999000 and 'ACC-EL' in str(item[1])):
                            insert_data = [item[0], item[1], item[2], item[3], item[4]]
                            self.tree.insert(parent_id, "end", text='□', values=insert_data, tags='unselect_row')
                    elif self.mode == 'compare':
                        if (750 <= offset <= 799 and 'AA' in str(item[1])) or ((801 <= offset <= 899 and 'EBOM' in str(item[1])) or (951 <= offset <= 980 and 'ACC-EL' == str(item[1]))) or (991 <= offset <= 998 and str(item[1]) == 'PR') or (base == 999000 and 'ACC-EL' in str(item[1])):
                            insert_data = [item[0], item[1], item[2], item[3], item[4]]
                            self.tree.insert(parent_id, "end", text='□', values=insert_data, tags='unselect_row')

            posnr_str = str(int(item[0]))
            if item in self.a_rows:
                self.vars[posnr_str] = tk.BooleanVar()

    def expand_tree(self, level):
        """展开/折叠到指定层级"""
        def expand_children(item, current_level, target_level):
            if current_level < target_level:
                self.tree.item(item, open=True)
                for child in self.tree.get_children(item):
                    expand_children(child, current_level + 1, target_level)
            else:
                self.tree.item(item, open=False)

        for item in self.tree.get_children():
            expand_children(item, 0, level)

    def expand_tree_node(self, node_iid, level=999):
        """只展开指定节点及其子树到 level 层"""

        # 复用你已有的递归函数，只是把“根”换成 node_iid
        def _expand(item, cur, target):
            if cur < target:
                self.tree.item(item, open=True)
                for child in self.tree.get_children(item):
                    _expand(child, cur + 1, target)
            else:
                self.tree.item(item, open=False)

        _expand(node_iid, 0, level)

    def on_click_item(self, event):
        """点击行切换勾选状态"""
        if self.tree.identify('region', event.x, event.y) != 'tree':
            return
        iid = self.tree.identify_row(event.y)

        if not iid:
            return

        # 用“有没有变量”判断能否勾选
        posnr_str = str(int(self.tree.item(iid, 'values')[0]))
        if posnr_str not in self.vars:
            return  # 不可勾选，直接忽略

        # 切换勾选状态
        var = self.vars[posnr_str]
        var.set(not var.get())
        self.tree.item(iid, text='✓' if var.get() else '□')
        self.tree.item(iid, tags='select_row' if var.get() else 'unselect_row')

    def on_ok(self):
        # 收集被选行
        self.selected = [idx for idx, it in enumerate(self.a_rows)
                         if self.vars.get(it[0], tk.BooleanVar()).get()]
        if not self.selected:
            tk.messagebox.showwarning('提示', '未选择任何条目', parent=self)
            return

        # 对比模式 -> 再选日期
        if self.mode == 'compare':
            self.ask_compare_dates()
        if self.mode == 'ebom':
            do_export_ebom(self)

    # --------------------- 选日期 ---------------------
    def ask_compare_dates(self):
        top = tk.Toplevel(self)
        top.title('输入BOM对比日期')
        center_window(top)
        top.geometry('250x150')
        top.grab_set()
        top.configure(bg='#eaf1f6')

        # --- 日期输入区 ---
        input_frame = tk.Frame(top, bg='#eaf1f6')
        input_frame.pack(padx=20, pady=15)

        # 新日期
        line1 = tk.Frame(input_frame, bg='#eaf1f6')
        line1.pack(fill='x', pady=5)
        tk.Label(line1, text='新日期 (yyyymmdd):',
                 font=("ABBvoice CNSG", int(11 * h_ratio)), bg='#eaf1f6').pack(side='left')
        ent1 = ttk.Entry(line1, width=12, font=("ABBvoice CNSG", int(11 * h_ratio)))
        ent1.pack(side='left', padx=(10, 0))
        Tooltip(ent1, "默认今天")

        # 旧日期
        line2 = tk.Frame(input_frame, bg='#eaf1f6')
        line2.pack(fill='x', pady=5)
        tk.Label(line2, text='旧日期 (yyyymmdd):',
                 font=("ABBvoice CNSG", int(11 * h_ratio)), bg='#eaf1f6').pack(side='left')
        ent2 = ttk.Entry(line2, width=12, font=("ABBvoice CNSG", int(11 * h_ratio)))
        ent2.pack(side='left', padx=(10, 0))
        Tooltip(ent2, "默认一个月之前")

        # 默认值
        today = dt.date.today().strftime('%Y%m%d')
        ent2.insert(0, (dt.date.today() - dt.timedelta(days=30)).strftime('%Y%m%d'))
        ent1.insert(0, today)

        # --- 按钮区 ---
        btn_frame = tk.Frame(top, bg='#eaf1f6')
        btn_frame.pack(pady=10)
        btn1 = tk.Button(btn_frame, text='确定', bg='#eaf1f6', font=("ABBvoice CNSG", int(11 * h_ratio)),
                  command=lambda: self._on_date_confirm(ent1, ent2, top))
        btn1.pack()
        Tooltip(btn1, "数据查询可能过慢，请耐心等待窗口弹出")
        # self.wait_window(top)

    # 把确认逻辑拆出来，保持清爽
    def _on_date_confirm(self, ent1, ent2, top):
        try:
            self.new_date = dt.datetime.strptime(ent1.get(), '%Y%m%d').date()
            self.old_date = dt.datetime.strptime(ent2.get(), '%Y%m%d').date()
            if self.new_date < self.old_date:
                raise ValueError('新日期不能早于旧日期')
            top.destroy()
            # self.destroy()
            do_compare_ebom(self, self.new_date, self.old_date)

        except Exception as e:
            tk.messagebox.showerror('错误', str(e), parent=self)