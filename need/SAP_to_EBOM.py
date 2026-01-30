import tkinter as tk
from tkinter import ttk

from tkinter.ttk import Treeview, Style

import need.tkutils as tku
from tkinter.filedialog import askdirectory
from collections import defaultdict
import os

from time import time
from time import *
import datetime
import traceback
import subprocess

import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl import load_workbook

import win32com.client

import win32gui
import win32con

from bs4 import BeautifulSoup
import pandas as pd
import warnings
from cryptography.fernet import Fernet
import base64
import pyrfc
import traceback
import datetime
import logging
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm, cm

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors

from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Image, Table, TableStyle, NextPageTemplate, PageBreak

warnings.simplefilter(action='ignore', category=FutureWarning)

FilePath = ""  # 设置一个地址变量


def main(parent, w_ratio, h_ratio, file_path):
    global checklog_file_path
    checklog_file_path = file_path
    global open_folder
    open_folder = tk.PhotoImage(file="ico\\open_folder.png")
    global view_folder
    view_folder = tk.PhotoImage(file="ico\\view.png")
    global read_file
    read_file = tk.PhotoImage(file="ico\\read.png")
    global export_file
    export_file = tk.PhotoImage(file="ico\\export.png")
    global check_file
    check_file = tk.PhotoImage(file="ico\\check.png")
    global pdf_export
    pdf_export = tk.PhotoImage(file="ico\\export.png")

    f1 = tk.Frame(parent, bg="#c9dbe9", bd=0)
    tk.Label(f1, text="欢迎使用EPLAN与SAP的EBOM对比功能", bg="#c9dbe9", fg="black", height=int(1 * h_ratio), font=("ABBvoice CNSG", int(20 * h_ratio), "bold")).pack(fill=tk.X)
    f1.pack(fill=tk.X)

    # tk.Frame(parent, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f2 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f2, text='   说明：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')
    tk.Label(f2, text='(1)选择EPLAN的BOM文件【C:/Temp/项目号-Files/项目号-BOM.xlsx】；\n(2)读取EPLAN和SAP的配置、BOM信息；\n(3)核对EPLAN和SAP的柜型配置、物料信息及端子排型号差异。', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    f2.pack(fill=tk.X)

    tk.Frame(parent, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f3 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f3, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f3, text='路径：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry  # 为了确保selectpath函数能正确调用entry,将其全局化
    entry = tk.Entry(f3, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)))
    entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

    global button_export_report
    button_export_report = tk.Button(f3, image=pdf_export, text="导出报表", font=("ABBvoice CNSG", int(13 * h_ratio)), bg="#eaf1f6", command=export_report, compound=tk.LEFT, state='disabled', activebackground='blue')
    button_export_report.pack(side=tk.RIGHT, padx=(0, int(20 * w_ratio)))

    global button_open
    button_open = tk.Button(f3, image=view_folder, text="看文件夹", font=("ABBvoice CNSG", int(13 * h_ratio)), bg="#eaf1f6", command=open_filefolder, compound=tk.LEFT, state='disabled', activebackground='blue')
    button_open.pack(side=tk.RIGHT, padx=int(20 * w_ratio))

    tk.Button(f3, text='选择', font=("ABBvoice CNSG", int(13 * h_ratio)), image=open_folder, bg="#eaf1f6", compound=tk.LEFT, command=selectpath, activebackground='blue').pack(side=tk.RIGHT, padx=(int(20 * w_ratio), 0))

    f3.pack(fill=tk.X)

    tk.Frame(parent, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    global f4
    f4 = tk.Frame(parent, bg="#eaf1f6", bd=0)

    tk.Label(f4, text='EPLAN：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')

    global sap_ci
    sap_ci = ''

    style = Style()
    style.configure('panel1.Treeview', rowheight=25, font=("ABBvoice CNSG", int(13 * h_ratio)))
    style.configure('panel1.Treeview.Heading', font=("ABBvoice CNSG", int(13 * h_ratio)), background="#EFF1F5")

    global SO_Item_table1
    table_ybar1 = tk.Scrollbar(f4)

    SO_Item_table1 = Treeview(f4, show='headings', style='panel1.Treeview', selectmode='browse', columns=('a', 'b', 'c', 'd', 'e'), yscrollcommand=table_ybar1.set, height=int(10 * h_ratio))
    table_ybar1.config(command=SO_Item_table1.yview)
    SO_Item_table1.column('a', width=int(70 * w_ratio), anchor='center')
    SO_Item_table1.column('b', width=int(70 * w_ratio), anchor='center')
    SO_Item_table1.column('c', width=int(750 * w_ratio), anchor='center')
    SO_Item_table1.column('d', width=int(70 * w_ratio), anchor='center')
    SO_Item_table1.column('e', width=int(140 * w_ratio), anchor='center')
    SO_Item_table1.heading('a', text='站号')
    SO_Item_table1.heading('b', text='柜型')
    SO_Item_table1.heading('c', text='ABB柜号')
    SO_Item_table1.heading('d', text='柜数')
    SO_Item_table1.heading('e', text='SO Item')
    SO_Item_table1.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 1), pady=0)
    SO_Item_table1.tag_configure('attention_row', foreground='red')
    SO_Item_table1.tag_configure('even_row', background="#c9dbe9")
    SO_Item_table1.tag_configure('odd_row', background="white")

    table_ybar1.pack(side=tk.LEFT, fill=tk.Y)

    global button_read_eplan
    button_read_eplan = tk.Button(f4, text='读取', font=("ABBvoice CNSG", int(13 * h_ratio)), image=read_file, bg="#eaf1f6", compound=tk.LEFT, command=read_panel_and_table, activebackground='blue')
    button_read_eplan.pack(side=tk.LEFT, pady=0, padx=int(20 * w_ratio))
    button_read_eplan['state'] = 'disabled'

    f4.pack(fill=tk.X)

    tk.Frame(parent, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    global f40
    f40 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f40, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')
    tk.Label(f40, text=' SAP：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')

    global SO_Item_table0
    table_ybar0 = tk.Scrollbar(f40)

    SO_Item_table0 = Treeview(f40, show='headings', style='panel1.Treeview', selectmode='browse', columns=('a', 'b', 'c', 'd', 'e'), yscrollcommand=table_ybar0.set, height=8)
    table_ybar0.config(command=SO_Item_table0.yview)
    SO_Item_table0.column('a', width=int(70 * w_ratio), anchor='center')
    SO_Item_table0.column('b', width=int(70 * w_ratio), anchor='center')
    SO_Item_table0.column('c', width=int(750 * w_ratio), anchor='center')
    SO_Item_table0.column('d', width=int(70 * w_ratio), anchor='center')
    SO_Item_table0.column('e', width=int(140 * w_ratio), anchor='center')

    SO_Item_table0.heading('a', text='站号', anchor='center')
    SO_Item_table0.heading('b', text='柜型', anchor='center')
    SO_Item_table0.heading('c', text='ABB柜号', anchor='center')
    SO_Item_table0.heading('d', text='柜数', anchor='center')
    SO_Item_table0.heading('e', text='SO Item', anchor='center')
    SO_Item_table0.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 1), pady=0)
    SO_Item_table0.tag_configure('even_row', background="#c9dbe9")
    SO_Item_table0.tag_configure('odd_row', background="white")

    table_ybar0.pack(side=tk.LEFT, fill=tk.Y)

    global button_read_sap
    button_read_sap = tk.Button(f40, text="查询", font=("ABBvoice CNSG", int(13 * h_ratio)), image=read_file, bg="#eaf1f6", compound=tk.LEFT, command=read_sap_item, activebackground='blue')
    button_read_sap.pack(side=tk.LEFT, pady=0, padx=int(20 * w_ratio))
    button_read_sap['state'] = 'disabled'

    f40.pack(fill=tk.X)

    tk.Frame(parent, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f8 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f8, text='   输出：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')

    global text
    text = tk.Text(f8, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), height=int(50 * h_ratio), width=int(170 * w_ratio))
    text.pack(side=tk.LEFT, padx=(0, 1), pady=0, fill=tk.BOTH, expand=True)

    text.tag_configure('error', foreground='red')  # 设置tag
    text.tag_configure('success', foreground='green')  # 设置tag

    scrollbar = tk.Scrollbar(f8)
    scrollbar.pack(side=tk.LEFT, fill=tk.Y)
    scrollbar.config(command=text.yview)
    text.config(yscrollcommand=scrollbar.set)

    global button_check
    button_check = tk.Button(f8, image=check_file, text="核对", font=("ABBvoice CNSG", int(13 * h_ratio)), bg="#eaf1f6", command=compare_ebom, compound=tk.LEFT, activebackground='blue')
    button_check.pack(side=tk.LEFT, pady=0, padx=int(20 * w_ratio))
    button_check['state'] = 'disabled'

    f8.pack(fill=tk.BOTH, expand=True)

    tk.Frame(parent, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线


def selectpath():
    filepath = tk.filedialog.askopenfilename(initialdir='C:/Temp/', title=u'请选择文件(XXX-BOM.xlsx)', filetypes=[("Excel", ".xlsx")])  # 选择打开什么文件，返回文件名
    stem, suffix = os.path.splitext(os.path.basename(filepath))  # stem是文件名,suffix是后缀
    # print(stem, suffix)
    if len(filepath) != 0:
        string_filename = ""
        for i in range(0, len(filepath)):
            string_filename += str(filepath[i])
        button_open['state'] = 'normal'
        button_read_eplan['state'] = 'normal'

    else:
        button_read_eplan['state'] = 'disabled'
        button_open['state'] = 'disabled'
        button_export_report['state'] = 'disabled'

    button_read_sap['state'] = 'disabled'
    button_check['state'] = 'disabled'
    button_export_report['state'] = 'disabled'

    text.delete(1.0, tk.END)  # 清空输出结果框

    # 清空treeview表格
    table_items = SO_Item_table0.get_children()  # 在插入treeview数据时，需要先清空treeview
    [SO_Item_table0.delete(table_item) for table_item in table_items]

    table_items = SO_Item_table1.get_children()  # 在插入treeview数据时，需要先清空treeview
    [SO_Item_table1.delete(table_item) for table_item in table_items]

    entry.delete(0, "end")  # 删除entry原始内容
    entry.insert(0, filepath)  # 重新填入地址
    global FilePath
    FilePath = filepath


def read_sap_item():
    # 清空treeview表格
    table_items = SO_Item_table0.get_children()  # 在插入treeview数据时，需要先清空treeview
    [SO_Item_table0.delete(table_item) for table_item in table_items]

    stem, suffix = os.path.splitext(os.path.basename(FilePath))  # stem是文件名,suffix是后缀

    # SAP登录
    sap_key_file = 'J:/Engineering/ShareFolder/new_ABB_Production_Tools/Ps/k/&%&^RD&^T&^WRE^@FEYUGYU@^E@DRYTRYTDFYTWF.txt'
    if not os.path.exists(sap_key_file):
        tk.messagebox.showwarning('提示', 'SAP加密密钥文件不存在')
    else:
        with open(sap_key_file, 'rb') as file:
            sap_key = file.read()
    sap_cipher = Fernet(sap_key)

    sap_ciphertext_file = 'J:/Engineering/ShareFolder/new_ABB_Production_Tools/Ps/s/^&%&^R^YFT$%FTDTEYTKFY$^$FGKJGHE%#$%#EYD^%#^%DTYD.txt'
    if not os.path.exists(sap_ciphertext_file):
        tk.messagebox.showwarning('提示', 'SAP加密密文文件不存在')
    else:
        with open(sap_ciphertext_file, 'r') as file:
            sap_ciphertext = file.read()

    sap_account = sap_decrypt_data(sap_ciphertext, sap_cipher, 'z_')
    sap_h = sap_account.split('+')[0]
    sap_c = sap_account.split('+')[1]
    sap_s = sap_account.split('+')[2]
    sap_u = sap_account.split('+')[3]
    sap_p = sap_account.split('+')[4]

    logging.info("Ready to connect to SAP")
    try:
        conn = pyrfc.Connection(ashost=sap_h, sysnr=sap_s, client=sap_c, user=sap_u, passwd=sap_p, lang='EN')
        if conn.alive:
            logging.info("Connecting to SAP successfully")
            project_number = stem.split('-')[0].split('.')[0]
            result = conn.call('ZY_SALES_ORDER_SHIFT', VBELN='0' + project_number)
            global sap_ci

            if result['EX_CEPTION'] == '' and result['ITAB'][0]['WERKS'] == '1201':
                item_data = []
                for item in result['ITAB']:
                    posnr = item['POSNR'].lstrip('0') or '0'
                    matnr = item['MATNR']
                    arktx = item['ARKTX']
                    tptx1 = item['TPTX1']
                    tptx2 = item.get('TPTX2', '')
                    kwmeng = str(int(float(item['KWMENG'])))

                    temp_var = (int(posnr)//1000)*1000//1000
                    temp_str_vr = f"A{temp_var:02d}"
                    if temp_str_vr in sap_ci:
                        item_data.append((posnr, matnr, arktx, tptx1 + tptx2, kwmeng))

                item_data_switchgear_id = []
                item_data_typical = []
                item_data_panel_number = []
                item_data_so_item = []
                item_data_amount = []
                global DZP_item_list
                DZP_item_list = []
                global DZP_list
                DZP_list = []

                parent_nodes = {}

                # 预定义颜色
                colors = ['even_row', 'odd_row']  # 你可以自定义更多颜色标签

                # 记录上一个插入的代码编号和颜色索引
                last_code_number = None
                color_index = 0
                code_color_map = {}  # 保存 code_number 与颜色的映射

                global sap_ebom_item, sap_ebom_typical, sap_ebom_material, sap_ebom_quantity
                sap_ebom_item = []
                sap_ebom_typical = []
                sap_ebom_material = []
                sap_ebom_quantity = []

                for item in item_data:
                    posnr = int(item[0])
                    if posnr % 1000 == 0:
                        parent_nodes[posnr] = 1
                    else:
                        parent_id = parent_nodes.get((int(posnr) // 1000) * 1000)
                        if parent_id:
                            # 获取 posnr 的千位数
                            base = (posnr // 1000) * 1000
                            # 计算 posnr 的偏移量
                            offset = posnr - base

                            if 801 <= offset <= 899 and str(item[1]) == 'EBOM':
                                # 物料数据读取
                                item_id = str(posnr).zfill(6)
                                rfc_table = [{'VBELN': '0' + project_number, 'VBPOS': item_id}]
                                bom_result = conn.call('ZPP_MES_BOM_EMS', IV_WERKS='1201', IV_FBSTP='X', IT_ORDERS=rfc_table)
                                if bom_result['ET_COMPONENTS']:
                                    for bom_item in bom_result['ET_COMPONENTS']:
                                        if bom_item['STUFE'] == 1:
                                            sap_ebom_item.append(posnr)
                                            sap_ebom_typical.append(str(item[2]))
                                            sap_ebom_material.append(bom_item['IDNRK'])
                                            sap_ebom_quantity.append(str(int(bom_item['MENGE'])))

                                # 配置数据写入表格
                                code_number = base // 1000
                                code_str = f"A{code_number:02d}"
                                item_data_switchgear_id.append(f"A{code_number:02d}")
                                item_data_typical.append(str(item[2]))
                                item_data_panel_number.append(str(item[3]))
                                item_data_amount.append(str(item[4]))
                                item_data_so_item.append(posnr % 100)

                                if code_str not in code_color_map:
                                    # 如果这是一个新的 code_number，选择下一个颜色并记录在映射中
                                    code_color_map[code_str] = colors[color_index % len(colors)]
                                    color_index += 1

                                    # 从映射中获取颜色标签
                                tag = code_color_map[code_str]

                                SO_Item_table0.insert("", "end", values=(f"A{code_number:02d}", str(item[2]), str(item[3]), str(item[4]), posnr % 100), tags=tag)

                            if offset == 751 and 'DZP-' in str(item[1]):
                                DZP_item_list.append(str(item[0]))
                                DZP_list.append(str(item[1]))

                # for i in range(0, len(sap_ebom_item)):
                #     print(sap_ebom_item[i], sap_ebom_typical[i],sap_ebom_material[i],sap_ebom_quantity[i])

            conn.close()
            if not conn.alive:
                logging.info("Disconnect from SAP")

            if len(item_data_switchgear_id) == 0:
                tk.messagebox.showwarning("提示", "SAP中尚无EBOM，无法对比")
                button_check['state'] = 'disabled'
            else:
                button_check['state'] = 'normal'


    except pyrfc.RFCError as e:
        logging.info(e.key + ', ' + e.message)
        tk.messagebox.showwarning("提示", traceback.format_exc())


def sap_decrypt_data(data, data_cipher, prefix):
    # 移除列名前缀prefix，如果有的话
    if data.startswith(prefix):
        data = data[2:]

    # Base64 字符串中可能删除了等号，需要添加回原来的填充字符，因为 base64 编码的输出长度应该是 4 的倍数
    padding_needed = 4 - (len(data) % 4)
    if padding_needed:
        data += "=" * padding_needed

    # base64 解码以获取原始的加密字节串
    encrypted_data = base64.urlsafe_b64decode(data.encode())

    # 使用 data_cipher 对象解密，它应该有和加密时一样的密钥
    decrypted_data = data_cipher.decrypt(encrypted_data)

    # 将字节解码回字符串
    decrypted_string = decrypted_data.decode()
    return decrypted_string


def open_filefolder():
    os.startfile(os.path.dirname(FilePath))


# ”读取“按钮
def read_panel_and_table():
    try:
        # 清空treeview表格
        table_items = SO_Item_table1.get_children()  # 在插入treeview数据时，需要先清空treeview
        [SO_Item_table1.delete(table_item) for table_item in table_items]

        if FilePath == "":
            tk.messagebox.showwarning("提示", "请选择文件！")
            button_read_sap['state'] = 'disabled'
        stem, suffix = os.path.splitext(os.path.basename(FilePath))  # stem是文件名,suffix是后缀
        if '-BOM' not in stem:
            tk.messagebox.showwarning("提示", "请选择XXX-BOM.xlsx文件！")
            button_read_sap['state'] = 'disabled'
        else:
            stem, suffix = os.path.splitext(os.path.basename(FilePath))  # stem是文件名,suffix是后缀
            # os.path.dirname()去掉文件名，返回目录
            # os.path.basename()去掉目录，返回文件名(含后缀)
            truename = stem.rsplit("-", 1)[0]
            inputfile = os.path.join(os.path.dirname(FilePath), truename + '-Panel.xlsx')  # XXX-Panel.xlsx
            global outputfile
            outputfile = os.path.join(os.path.dirname(FilePath), truename + '-Table.xls')  # XXX-Table.xls

            if not os.path.exists(inputfile):
                tk.messagebox.showwarning("提示", "失败,找不到%s文件! " % inputfile.replace("\\", "/"))
                button_read_sap['state'] = 'disabled'
            else:
                book = load_workbook(inputfile)
                sheet = book['Z2_xlsx']
                A = []
                B = []
                C = []
                for i in range(2, sheet.max_row + 1):
                    A.append(str(sheet.cell(row=i, column=1).value))  # Order Line列
                    B.append(str(sheet.cell(row=i, column=2).value))  # Typical列
                    C.append(str(sheet.cell(row=i, column=3).value))  # Panel No列

                none_count = B.count('None')

                if none_count == len(B):    # 说明是旧版升版2024项目，它的Panel.xlsx报表B列为空，需要用Panel.xls，如果是纯2024项目，只需Panel.xlsx即可
                    A = []
                    B = []
                    C = []
                    book = xlrd.open_workbook(os.path.join(os.path.dirname(FilePath), truename + '-Panel.xls'))  # Panel.xls
                    sheet = book.sheet_by_index(0)
                    A = sheet.col_values(colx=0, start_rowx=1, end_rowx=None)  # Order Line列
                    B = sheet.col_values(colx=1, start_rowx=1, end_rowx=None)  # Typical列
                    C = sheet.col_values(colx=2, start_rowx=1, end_rowx=None)  # Panel No列
                if 0 < none_count < len(B):
                    tk.messagebox.showwarning('提示', "失败, Panel.xlsx中部分Typical数据缺失, 请补全")
                    return

                new_A = []
                new_B = []
                new_C = []

                for i in range(len(A)):
                    if i < len(B) and (B[i] != '空柜' and 'DUMMY' not in B[i].upper()):
                        new_A.append(A[i])
                        new_B.append(B[i])
                        new_C.append(C[i])

                A = new_A
                B = new_B
                C = new_C
                
                for i in range(0, len(A)):
                    C[i] = C[i] + ';'  # 给ABB柜号后面添加分号

                dataframe_A = pd.DataFrame(A, dtype=str)
                dataframe_B = pd.concat([dataframe_A, pd.DataFrame(B, dtype=str)], axis=1)
                dataframe_C = pd.concat([dataframe_B, pd.DataFrame(C, dtype=str)], axis=1)
                dataframe = dataframe_C
                dataframe.columns = ['OrderLine', 'Typical', 'PanelNo']  # 给各列赋名
                # print(dataframe, '\n\n')

                dataframe_groupby = dataframe.groupby(['OrderLine', 'Typical']).agg({'PanelNo': 'sum'}).reset_index()  # 根据站号和柜型列，将柜列数据进行合并，并重置index
                # print(dataframe_groupby)

                panelamount = []
                for i in range(0, len(dataframe_groupby['OrderLine'])):
                    panelamount.append(dataframe_groupby['PanelNo'][i].count(';'))  # dataframe增加柜数列，数据来源于柜号列中每一个数据中分号(;)个数
                    # print(dataframe_groupby['PanelNo'][i].count(';'))
                dataframe_groupby = pd.concat([dataframe_groupby, pd.DataFrame(panelamount, dtype=str)], axis=1)  # dataframe增加柜数列
                dataframe_groupby.columns = ['OrderLine', 'Typical', 'PanelNo', 'PanelAmount']  # 给各列赋名
                # print(dataframe_groupby, '\n\n')

                orderline_set = list(set(dataframe_groupby['OrderLine']))  # set去重后顺序会改变
                orderline_set.sort(key=list(dataframe_groupby['OrderLine']).index)  # 保证去重后顺序不变

                if not os.path.exists(outputfile):  # 如果配置表格不存在，则SO Item列按默认配置，否则按表格内容配置
                    so_item = [0] * len(dataframe_groupby['OrderLine'])  # 创建指定长度列表
                    for i in range(0, len(orderline_set)):
                        temp = [j for j, k in enumerate(list(dataframe_groupby['OrderLine'])) if k == orderline_set[i]]
                        number = list(range(1, len(temp) + 1))
                        for l in range(0, len(temp)):
                            so_item[temp[l]] = number[l]
                else:
                    book2 = xlrd.open_workbook(outputfile)  # 加载XXX-Table.xls表格
                    worksheet = book2.sheet_by_index(0)

                    table_col_A = worksheet.col_values(colx=0, start_rowx=1, end_rowx=None)
                    table_col_B = worksheet.col_values(colx=1, start_rowx=1, end_rowx=None)
                    table_col_C = worksheet.col_values(colx=2, start_rowx=1, end_rowx=None)
                    table_col_D = worksheet.col_values(colx=3, start_rowx=1, end_rowx=None)
                    table_col_E = worksheet.col_values(colx=4, start_rowx=1, end_rowx=None)

                    tuples_table_col_AB = list(zip(table_col_A, table_col_B))
                    tuples_dataframe_groupby = list(zip(dataframe_groupby['OrderLine'], dataframe_groupby['Typical']))

                    # dataframe_groupby比table_col多的数据及索引
                    add_index = [i for i in range(len(tuples_dataframe_groupby)) if tuples_dataframe_groupby[i] not in tuples_table_col_AB]
                    add_content = [tuples_dataframe_groupby[i] for i in add_index]

                    # dataframe_groupby比table_col少的数据及索引
                    minus_index = [i for i in range(len(tuples_table_col_AB)) if tuples_table_col_AB[i] not in tuples_dataframe_groupby]
                    minus_content = [tuples_table_col_AB[i] for i in minus_index]

                    # print(add_index, add_content)
                    # print(minus_index, minus_content)
                    if len(add_index) == 0 and len(minus_index) == 0:  # 无新增Typical
                        so_item = table_col_E  # 仍取原来的Table中的SO Item
                    elif len(add_index) > 0 and len(minus_index) == 0:  # 新增Typical
                        so_item = table_col_E
                        new_typical = []
                        for n in range(0, len(add_index)):
                            new_typical.append(add_content[n])
                            so_item.insert(add_index[n], '0')  # 新增的Typical的SO Item设置为0

                        tk.messagebox.showwarning("提示", f"新增柜型{new_typical}\n请维护SO Item(默认0)！")

                    elif len(add_index) == 0 and len(minus_index) > 0:  # 减少Typical
                        so_item = table_col_E
                        lack_typical = []
                        for i, n in enumerate(sorted(minus_index, reverse=True)):  # 使用enumerate来跟踪索引和值
                            lack_typical.append(minus_content[i])
                            del so_item[n]

                        tk.messagebox.showwarning("提示", f"减少柜型{lack_typical}\n自动移除对应SO Item！")

                    elif len(add_index) > 0 and len(minus_index) > 0:  # 有增有减Typical
                        so_item = table_col_E
                        new_typical = []
                        lack_typical = []

                        # 首先处理减少部分
                        for i, n in enumerate(sorted(minus_index, reverse=True)):  # 使用enumerate来跟踪索引和值
                            lack_typical.append(minus_content[i])
                            del so_item[n]

                        # 处理新增部分
                        for n in range(len(add_index)):
                            new_typical.append(add_content[n])
                            so_item.insert(add_index[n], '0')  # 新增的Typical的SO Item设置为0

                        tk.messagebox.showwarning("提示", f"新增柜型{new_typical}\n请维护SO Item(默认0)！\n\n减少柜型{lack_typical}\n自动移除对应SO Item！")

                # 完整的dataframe数据
                dataframe_groupby = pd.concat([dataframe_groupby, pd.DataFrame(so_item, dtype=str)], axis=1)  # dataframe增加SO Item列
                dataframe_groupby.columns = ['OrderLine', 'Typical', 'PanelNo', 'PanelAmount', 'SOItem']  # 给各列赋名

                # 清空treeview表格
                table_items = SO_Item_table1.get_children()  # 在插入treeview数据时，需要先清空treeview
                [SO_Item_table1.delete(table_item) for table_item in table_items]

                # 预定义颜色
                colors = ['even_row', 'odd_row']  # 你可以自定义更多颜色标签

                # 记录上一个插入的 OrderLine 和颜色索引
                last_order_line = None
                color_index = 0
                order_line_color_map = {}  # 保存 OrderLine 与颜色的映射

                # 为treeview表格插入数据
                for i in range(0, len(dataframe_groupby['OrderLine'])):
                    order_line = dataframe_groupby['OrderLine'][i]

                    if order_line not in order_line_color_map:
                        # 如果这是一个新的 OrderLine，选择下一个颜色并记录在映射中
                        order_line_color_map[order_line] = colors[color_index % len(colors)]
                        color_index += 1

                    # 从映射中获取颜色标签
                    tag = order_line_color_map[order_line]

                    if dataframe_groupby['SOItem'][i] == 0:
                        SO_Item_table1.insert('', 'end', values=list(dataframe_groupby.loc[i]), tags=('attention_row', tag))  # dataframe逐行插入到表格中
                    else:
                        SO_Item_table1.insert('', 'end', values=list(dataframe_groupby.loc[i]), tags=(tag,))  # dataframe逐行插入到表格中

                # 完整的数据导入Table表格
                book2 = xlwt.Workbook()  # 创建一个空文件对象
                book2.add_sheet('Sheet1')  # 创建一个Sheet页
                book2.save(outputfile)  # XXX-Table.xls
                book2 = xlrd.open_workbook(outputfile)  # 加载XXX-Table.xls表格

                workbook = copy(book2)  # 使用xlutils.copy将xlrd读取的book对象转为xlwt可操作对象
                worksheet = workbook.get_sheet(0)  # 获取sheet
                worksheet.write(0, 0, 'OrderLine')  # 在sheet指定位置写入数据
                worksheet.write(0, 1, 'Typical')  # 在sheet指定位置写入数据
                worksheet.write(0, 2, 'PanelNo')  # 在sheet指定位置写入数据
                worksheet.write(0, 3, 'PanelAmount')  # 在sheet指定位置写入数据
                worksheet.write(0, 4, 'SOItem')  # 在sheet指定位置写入数据

                style = xlwt.XFStyle()
                style.num_format_str = '@'  # 设置数据格式为文本

                for i in range(0, len(dataframe_groupby['OrderLine'])):  # 遍历Panel.xls，将Panel.xls数据复制到Panel Size.xls
                    worksheet.write(i + 1, 0, dataframe_groupby['OrderLine'][i], style=style)
                    worksheet.write(i + 1, 1, dataframe_groupby['Typical'][i], style=style)
                    worksheet.write(i + 1, 2, dataframe_groupby['PanelNo'][i], style=style)
                    worksheet.write(i + 1, 3, dataframe_groupby['PanelAmount'][i], style=style)
                    worksheet.write(i + 1, 4, dataframe_groupby['SOItem'][i], style=style)
                global sap_ci
                sap_ci = list(set(dataframe_groupby['OrderLine']))    # 用该数据控制SAP中查询的行号，以避免带出其他行号数据
                # print(sap_ci)
                workbook.save(outputfile)  # 将workbook保存到指定位置

                button_read_sap['state'] = 'normal'

    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


# ”核对“按钮
def compare_ebom():
    try:
        error1_calculator = 0
        error2_calculator = 0
        error3_calculator = 0
        start = time()
        text.delete(1.0, "end")  # 清空输出结果框
        text.insert(tk.INSERT, ">>>EPLAN与SAP配置信息对比(不考虑SO Item差异)...\n")
        missing, extra = compare_treeviews(SO_Item_table0, SO_Item_table1)    # 第二个表格比第一个表格多了什么，少了什么
        config_error_flag = 0
        if len(extra) > 0:
            for item in extra:
                formatted_item = '—'.join(map(str, item))
                text.insert(tk.INSERT, "▲   EPLAN比SAP多出的配置：%s\n" % formatted_item, 'error')
                error1_calculator += 1
                config_error_flag = 1
        if len(missing) > 0:
            for item in missing:
                formatted_item = '—'.join(map(str, item))
                text.insert(tk.INSERT, "▲   EPLAN比SAP缺少的配置：%s\n" % formatted_item, 'error')
                error1_calculator += 1
                config_error_flag = 1

        if not config_error_flag:
            text.insert(tk.INSERT, "EPLAN与SAP配置信息一致\n")

        text.insert(tk.INSERT, "\n>>>EPLAN与SAP物料信息对比...\n")

        if not os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\CRM.xlsx"):
            text.insert(tk.INSERT, '▲   失败，找不到CRM继电器物料号数据表...\n', 'error')
        else:
            workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\CRM.xlsx")  # 加载CRM避雷器数据表
            worksheet = workbook['Sheet1']

            # 创建一个集合来存储第2、3、4列的所有值
            unique_values_set = set()

            for i in range(1, worksheet.max_row + 1):
                # 获取第2、3、4列的值
                value_2 = worksheet.cell(row=i, column=2).value
                value_3 = worksheet.cell(row=i, column=3).value
                value_4 = worksheet.cell(row=i, column=4).value

                # 检查并添加到集合中
                if value_2 is not None and value_2 != '':
                    unique_values_set.add(str(value_2).rstrip())
                if value_3 is not None and value_3 != '':
                    unique_values_set.add(str(value_3).rstrip())
                if value_4 is not None and value_4 != '':
                    unique_values_set.add(str(value_4).rstrip())

            # 先对BOM.xlsx数据进行处理，将同一个Typical下相同物料号的数量累计
            df = pd.read_excel(FilePath)
            # print(df)
            # 准备放要删除的行号
            drop_rows = set()

            # 逐行判断
            for i in range(len(df)):
                # 先取出原始值，对 NaN 统一转成 ''
                des = '' if pd.isna(df.loc[i, 'Designation']) else str(df.loc[i, 'Designation']).strip()
                typ = '' if pd.isna(df.loc[i, 'Type']) else str(df.loc[i, 'Type']).strip()
                dsc = '' if pd.isna(df.loc[i, 'Description']) else str(df.loc[i, 'Description']).strip()
                exist = '' if pd.isna(df.loc[i, 'Exist']) else str(df.loc[i, 'Exist']).strip().upper()
                exist1 = '' if pd.isna(df.loc[i, 'Exist1']) else str(df.loc[i, 'Exist1']).strip().upper()
                pn = '' if pd.isna(df.loc[i, 'PartNumber']) else str(df.loc[i, 'PartNumber']).strip()

                # 规则1：三列全空 或 Exist 为 X/x
                if (des == '' and typ == '' and dsc == '') or exist == 'X':
                    drop_rows.add(i)

                # 规则2：Exist1 为 YES
                if exist1 == 'YES':
                    drop_rows.add(i)

                # 规则3：PartNumber 为 Mounting Panel 或含 EP-
                if pn == 'Mounting Panel' or 'EP-' in pn:
                    drop_rows.add(i)

            # 真正删除
            df_filtered = df.drop(index=list(drop_rows)).reset_index(drop=True)
            # print(df_filtered)
            # 后续 groupby 不变
            df2 = df_filtered.groupby(['Hight-level', 'PartNumber'], as_index=False)['Qty'].sum()

            # 4. 后面原样处理
            bom_col_A = []
            bom_col_D = []
            bom_col_E = []
            for _, row in df2.iterrows():
                if row['PartNumber'] not in sap_ebom_typical:
                    bom_col_A.append(row['Hight-level'])
                    bom_col_D.append(row['PartNumber'])
                    bom_col_E.append(str(int(row['Qty'])))

            # 1. 把 EPLAN 数据做成 (Typical, PartNumber) -> Qty 的字典
            eplan_typical_bom = defaultdict(list)
            for _, row in df2.iterrows():
                hl = row['Hight-level']
                pn = row['PartNumber']
                qty = int(row['Qty'])
                eplan_typical_bom[hl].append((pn, qty))

            # 2. 把 SAP 数据做成同样格式的字典
            sap_item_bom = defaultdict(list)
            for item, typ, mat, qty in zip(sap_ebom_item,
                                           sap_ebom_typical,
                                           sap_ebom_material,
                                           sap_ebom_quantity):
                sap_item_bom[(item, typ)].append((mat, int(qty)))

            bom_error_flag = 0

            for (item, typ), sap_lst in sap_item_bom.items():
                # 找到 EPLAN 里同 Typical 的那份清单
                eplan_lst = eplan_typical_bom.get(typ, [])
                # 两边都转成 dict 方便比对
                sap_dict = defaultdict(int)
                eplan_dict = defaultdict(int)
                for mat, qty in sap_lst:
                    sap_dict[mat] += qty
                for pn, qty in eplan_lst:
                    eplan_dict[pn] += qty

                # 3. 计算三套差异
                eplan_keys = set(eplan_dict.keys())
                sap_keys = set(sap_dict.keys())

                # 3.1 SAP 有、EPLAN 无（且不在 CRM 白名单）
                for pn in sap_keys - eplan_keys:
                    if pn not in unique_values_set:
                        qty = sap_dict[pn]
                        text.insert(tk.INSERT,
                                    f"▲   SAP item {item} 多余物料：{typ}—{pn}—{qty}个    【EPLAN无此料号】\n", 'error')
                        error2_calculator += 1
                        bom_error_flag = 1

                # 3.2 EPLAN 有、SAP 无
                for pn in eplan_keys - sap_keys:
                    qty = eplan_dict[pn]
                    text.insert(tk.INSERT,
                                f"▲   SAP item {item} 缺少物料：{typ}—{pn}—{qty}个    【SAP无此料号】\n", 'error')
                    error2_calculator += 1
                    bom_error_flag = 1

                # 3.3 同料号数量不一致
                for pn in eplan_keys & sap_keys:
                    if eplan_dict[pn] != sap_dict[pn]:
                        text.insert(tk.INSERT,
                                    f"▲   SAP item {item} 数量不一致：{typ}—{pn}—SAP{sap_dict[pn]}个，EPLAN{eplan_dict[pn]}个\n", 'error')
                        error2_calculator += 1
                        bom_error_flag = 1

            if not bom_error_flag:
                text.insert(tk.INSERT, "所有 sap_ebom_item 物料信息与 EPLAN 一致\n")

            # for i in range(0, len(bom_col_A)):
            #     if bom_col_A[i] not in sap_ebom_typical:
            #         text.insert(tk.INSERT, "▲   SAP缺少物料信息为：%s—%s—%s个    【SAP无此Typical】\n" % (bom_col_A[i], bom_col_D[i], bom_col_E[i]), 'error')
            #         error2_calculator += 1
            #         bom_error_flag = 1
            #     elif bom_col_D[i] not in sap_ebom_material:
            #         text.insert(tk.INSERT, "▲   SAP缺少物料信息为：%s—%s—%s个    【SAP无此物料号】\n" % (bom_col_A[i], bom_col_D[i], bom_col_E[i]), 'error')
            #         error2_calculator += 1
            #         bom_error_flag = 1
            #
            # for i in range(0, len(bom_col_A)):
            #     for j in range(0, len(sap_ebom_typical)):
            #         if bom_col_A[i] == sap_ebom_typical[j]:
            #             print(sap_ebom_typical[j], sap_ebom_material[j], sap_ebom_quantity[j])
            #             if bom_col_D[i] == sap_ebom_material[j] and bom_col_E[i] != sap_ebom_quantity[j]:
            #
            #                 text.insert(tk.INSERT, "▲   SAP物料信息为：%s—%s—%s个，EPLAN物料数量为：%s个    【SAP与EPLAN物料数量不一致】\n" % (sap_ebom_typical[j], sap_ebom_material[j], sap_ebom_quantity[j], bom_col_E[i]), 'error')
            #                 error2_calculator += 1
            #                 bom_error_flag = 1
            #                 break
            #
            # for i in range(0, len(sap_ebom_typical)):
            #     for j in range(0, len(bom_col_A)):
            #         if bom_col_A[j] == sap_ebom_typical[i]:
            #
            #
            #
            #     if sap_ebom_material[i] not in bom_col_D and sap_ebom_material[i] not in unique_values_set:
            #         text.insert(tk.INSERT, "▲   SAP多余的物料信息为：%s—%s—%s个    【EPLAN无此物料号】\n" % (sap_ebom_typical[i], sap_ebom_material[i], sap_ebom_quantity[i]), 'error')
            #         error2_calculator += 1
            #         bom_error_flag = 1
            #
            # if not bom_error_flag:
            #     text.insert(tk.INSERT, "EPLAN与SAP物料信息一致\n")

        text.insert(tk.INSERT, "\n>>>EPLAN与SAP的端子排型号对比...\n")

        stem, suffix = os.path.splitext(os.path.basename(FilePath))  # stem是文件名,suffix是后缀
        truename = stem.rsplit("-", 1)[0]  # 项目号
        DZP_file = (os.path.join(os.path.dirname(FilePath), truename + '-DZP.txt')).replace('\\', '/')    # XXX-DZP.txt

        if not os.path.exists(DZP_file):
            text.insert(tk.INSERT, "▲   无法对比！ 找不到%s文件\n" % DZP_file, 'error')
            SAP_DZP = '\n'.join(["%s, %s" % (a, b) for a, b in zip(DZP_item_list, DZP_list)])
            text.insert(tk.INSERT, "SAP中端子排：%s\n" % SAP_DZP, 'error')
        else:
            DZP_error_flag = 0
            with open(DZP_file, "r") as file:
                EPLAN_DZP = file.read()
            for i in range(0, len(DZP_item_list)):
                if EPLAN_DZP != DZP_list[i]:
                    DZP_error_flag = 1
            if DZP_error_flag:
                SAP_DZP = '\n'.join(["%s, %s" % (a, b) for a, b in zip(DZP_item_list, DZP_list)])
                text.insert(tk.INSERT, "▲   导EBOM时端子排：%s\nSAP中端子排：%s\n" % (EPLAN_DZP, SAP_DZP), 'error')
                error3_calculator += 1
            else:
                text.insert(tk.INSERT, "导EBOM时端子排与SAP中端子排型号一致\n")

        end = time()
        text.insert(tk.INSERT, "\n>>>EPLAN与SAP配置信息、物料信息，核对完成!  用时%.3f秒\n" % (end - start))
        # tk.messagebox.showwarning("提示", "EPLAN与SAP配置信息、物料信息，核对完成")

        button_export_report['state'] = 'normal'

        global checklog_file_path
        checklogbook = load_workbook(checklog_file_path)
        checklogsheet = checklogbook['Sheet']
        project_no = stem.replace('-BOM', '')
        current_time = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        data1 = [project_no, 'EBOM对比-EPLAN与SAP配置一致性', error1_calculator, current_time]
        data2 = [project_no, 'EBOM对比-EPLAN与SAP物料一致性', error2_calculator, current_time]
        data3 = [project_no, 'EBOM对比-EPLAN与SAP端子排型号一致性', error3_calculator, current_time]

        checklogsheet.append(data1)
        checklogsheet.append(data2)
        checklogsheet.append(data3)

        checklogbook.save(checklog_file_path)

    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


# def compare_treeviews(treeview1, treeview2):
#     # 获取第一个 Treeview 的数据
#     data1 = []
#     for row in treeview1.get_children():
#         row_data = treeview1.item(row)['values'][:4]  # 取前4列
#         data1.append(tuple(row_data))
#
#     # 获取第二个 Treeview 的数据
#     data2 = []
#     for row in treeview2.get_children():
#         row_data = treeview2.item(row)['values'][:4]  # 取前4列
#         data2.append(tuple(row_data))
#
#     # 转换为集合以便比较
#     set1 = set(data1)
#     set2 = set(data2)
#
#     # 计算差异
#     missing_in_treeview2 = set1 - set2  # 第一个表格有，第二个表格没有的数据
#     extra_in_treeview2 = set2 - set1  # 第二个表格有，第一个表格没有的数据
#
#     return list(missing_in_treeview2), list(extra_in_treeview2)

def _normalize_col3(value: str) -> str:
    """把第3列按';'拆分、排序后再拼回去，忽略顺序差异。"""
    if not isinstance(value, str):
        return value
    parts = [p.strip() for p in value.split(';') if p.strip()]
    return ';'.join(sorted(parts))


def compare_treeviews(treeview1, treeview2):
    # 取数据并生成“比较键”
    def _get_key(tv):
        keys = []
        for row in tv.get_children():
            raw = tv.item(row)['values'][:4]          # 原始4列
            key = (
                raw[0],
                raw[1],
                _normalize_col3(raw[2]),               # 第3列做顺序无关处理
                raw[3]
            )
            keys.append(key)
        return keys

    keys1 = set(_get_key(treeview1))
    keys2 = set(_get_key(treeview2))

    missing_in_treeview2 = keys1 - keys2
    extra_in_treeview2   = keys2 - keys1

    # 如需返回原始行数据，可在这里把key再映射回raw，目前直接返回差异的“键”
    return list(missing_in_treeview2), list(extra_in_treeview2)

def export_report():
    try:
        pdfmetrics.registerFont(TTFont('SimSun', 'simsun.ttc'))  # 注册字体
        stem, _ = os.path.splitext(os.path.basename(FilePath))
        home_path = os.path.expanduser("~")
        desktop_path = os.path.join(home_path, "Desktop")
        global project_number
        project_number = stem.replace('-BOM', '')
        desktop_path = askdirectory(title=u'请选择导出文件夹', initialdir=desktop_path)
        if not desktop_path:
            tk.messagebox.showwarning("提示", "未选择保存路径，导出操作已取消")
            return

        doc = SimpleDocTemplate(os.path.join(desktop_path, 'EPLAN与SAP的EBOM对比报告-%s.pdf' % project_number), pagesize=A4, rightMargin=40, leftMargin=40, topMargin=100, bottomMargin=20)
        story = []
        # 获取样式
        styles = getSampleStyleSheet()
        style = styles['Normal']
        style.fontName = 'SimSun'
        style.wordWrap = 'CJK'
        style.leading = 15  # 行间距

        text_content = text.get("1.0", tk.END)
        start_index = text_content.find(">>>EPLAN与SAP配置信息对比(不考虑SO Item差异)...")
        end_index = find_end_index(text_content, "\n>>>EPLAN与SAP配置信息、物料信息，核对完成!")

        text_content = text_to_html(text, start_index, end_index)

        p = Paragraph(text_content,style)
        story.append(p)

        global page_number
        page_number = 1

        # 通过`SimpleDocTemplate`直接生成文档，传入`add_header`函数绘制每页的头部内容
        doc.build(story, onFirstPage=add_header1, onLaterPages=add_header2)

        tk.messagebox.showwarning("提示", "EBOM对比报告导出完成")
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())

def find_end_index(text_content, marker):
    lines = text_content.splitlines(True)  # 使用True保持换行符
    for i, line in enumerate(lines):
        if marker in line:
            return sum(len(l) for l in lines[:i+1])  # 返回到该行末尾的累计长度
    return len(text_content)  # 如果没有找到marker，返回文本的总长度


def text_to_html(text_widget, start_index, end_index):
    # 获取文本内容
    text_content = text_widget.get("1.0", tk.END)

    # 获取 start_index 之前的文本
    before_text = text_content[:start_index-1]
    # 获取 end_index 之后的文本
    after_text = text_content[end_index:]

    # 合并排除了指定范围的文本
    final_text = before_text + after_text

    # 将换行符转换为HTML的换行标签
    html_output = final_text.replace('\n', '<br />')

    # 结束HTML文档
    html_output = f"<html><body><p>{html_output}</p></body></html>"

    return html_output


def add_header1(rl_canvas, doc):
    rl_canvas.saveState()

    # 添加图片（请确保图片路径正确）
    image_path = 'ico/ABB_logo.png'
    img_width = 27.7 * mm
    img_height = 10.5 * mm
    img = Image(image_path, width=img_width, height=img_height)
    img.drawOn(rl_canvas, A4[0] / 2 - img_width / 2, A4[1] - img_height - 30)

    # 添加标题
    rl_canvas.setFont('SimSun', 16)
    title = "FAST EPLAN与SAP的EBOM对比报告"
    rl_canvas.drawCentredString(A4[0] / 2, A4[1] - img_height - 60, title)

    rl_canvas.setFont('SimSun', 10)
    project_number_text = "项目号：" + project_number
    rl_canvas.drawString(45, A4[1] - img_height - 60 - 10, project_number_text)

    rl_canvas.setFont('SimSun', 8)
    global page_number
    page_number_text = "Page " + str(page_number)
    page_number += 1
    rl_canvas.drawString(A4[0]-80, 5*mm, page_number_text)

    rl_canvas.setFont('SimSun', 8)
    Timestamp = strftime('%Y-%m-%d %H:%M:%S', localtime())
    rl_canvas.drawCentredString(A4[0] / 2, 5 * mm, '制表时间：' + Timestamp)

    rl_canvas.restoreState()


def add_header2(rl_canvas, doc):
    rl_canvas.saveState()

    # 添加图片（请确保图片路径正确）
    image_path = 'ico/ABB_logo.png'
    img_width = 27.7 * mm
    img_height = 10.5 * mm
    img = Image(image_path, width=img_width, height=img_height)
    img.drawOn(rl_canvas, A4[0] / 2 - img_width / 2, A4[1] - img_height - 30)

    # 添加标题
    rl_canvas.setFont('SimSun', 16)
    title = "FAST EPLAN与SAP的EBOM对比报告"
    rl_canvas.drawCentredString(A4[0] / 2, A4[1] - img_height - 60, title)

    rl_canvas.setFont('SimSun', 8)
    global page_number
    page_number_text = "Page " + str(page_number)
    page_number += 1
    rl_canvas.drawString(A4[0]-80, 5*mm, page_number_text)

    rl_canvas.setFont('SimSun', 8)
    Timestamp = strftime('%Y-%m-%d %H:%M:%S', localtime())
    rl_canvas.drawCentredString(A4[0] / 2, 5 * mm, '制表时间：' + Timestamp)

    rl_canvas.restoreState()