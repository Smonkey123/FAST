import tkinter as tk
from tkinter import ttk
from tkinter.filedialog import askdirectory
from tkinter.ttk import Treeview, Style
import need.tkutils as tku

import os
import sqlite3

import ctypes
import time
from datetime import datetime

import re

from openpyxl import load_workbook

import math
import shutil
import subprocess
import win32com.client

import warnings


warnings.simplefilter(action='ignore', category=FutureWarning)

FilePath = ""  # 设置一个地址变量


def main(parent, w_ratio, h_ratio):
    global query_project
    query_project = tk.PhotoImage(file="ico\\search.png")
    global refresh_project
    refresh_project = tk.PhotoImage(file="ico\\refresh.png")
    global view_project
    view_project = tk.PhotoImage(file="ico\\view.png")
    global reply_project
    reply_project = tk.PhotoImage(file="ico\\edit.png")

    global canvas
    canvas = tk.Canvas(parent, width=int(1600 * w_ratio), height=int(640 * h_ratio), bg="#C9DBE9")
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    canvas.update()
    canvas.bind("<MouseWheel>", on_mousewheel)

    scrollbar_v = tk.Scrollbar(master=parent)
    scrollbar_v.pack(side=tk.RIGHT, fill=tk.Y)
    scrollbar_v.config(command=canvas.yview)
    canvas.config(yscrollcommand=scrollbar_v.set)

    content = tk.Frame(canvas)
    # content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    canvas.create_window(0, 1, width=int(1600 * w_ratio), anchor=tk.NW, window=content)

    global subparent
    subparent = parent

    global sub_w_ratio
    sub_w_ratio = w_ratio

    global sub_h_ratio
    sub_h_ratio = h_ratio

    f1 = tk.Frame(content, bg="#c9dbe9", bd=0)
    # im = tku.image_label(f1, "ico\\help.png", int(30 * h_ratio), int(30 * h_ratio), False)
    # im.configure(bg="#c9dbe9")
    # im.bind('<Button-1>', about_help)
    # im.pack(side=tk.RIGHT)
    tk.Label(f1, text="欢迎使用图纸意见传递功能", bg="#c9dbe9", fg="black", height=int(1 * h_ratio), font=tku._ft(int(20 * h_ratio), True)).pack(fill=tk.X)
    f1.pack(fill=tk.X)

    f2 = tk.Frame(content, bg="#eaf1f6", bd=0)
    f2.pack(fill=tk.BOTH, expand=True)

    f21 = tk.Frame(f2, bg="#eaf1f6", bd=0)
    f21.pack(side=tk.TOP, fill=tk.BOTH, pady=(int(10 * h_ratio), int(10 * h_ratio)))

    f22 = tk.Frame(f2, bg="#eaf1f6", bd=0)
    f22.pack(side=tk.TOP, fill=tk.BOTH, pady=(0, int(10 * h_ratio)))

    f23 = tk.Frame(f2, bg="#eaf1f6", bd=0)
    f23.pack(side=tk.TOP, fill=tk.BOTH, pady=(0, int(10 * h_ratio)))

    f231 = tk.Frame(f23, bg="#eaf1f6", bd=0)
    f231.pack(side=tk.LEFT, fill=tk.BOTH)

    f232 = tk.Frame(f23, bg="#eaf1f6", bd=0)
    f232.pack(side=tk.LEFT, fill=tk.BOTH)

    tk.Frame(content, bg="#eaf1f6", bd=0, height=int(40 * h_ratio)).pack(fill=tk.BOTH, expand=True)

    global f3
    f3 = tk.Frame(content, bg="#eaf1f6", bd=0, height=int(700 * h_ratio))
    f3.pack(fill=tk.BOTH, expand=True)
    tk.Label(f21, text=' *', bg="#eaf1f6", fg="red", font=tku._ft(int(13 * h_ratio), False), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f21, text="SO ：", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=tku._ft(int(13 * h_ratio), False)).pack(side=tk.LEFT, fill=tk.X)
    global project_number_list
    project_number_list = []

    parent.option_add("*TCombobox*Listbox.font", ("ABBvoice CNSG", int(13 * h_ratio)))

    global combobox_project_number
    combobox_project_number = ttk.Combobox(f21, font=tku._ft(int(13 * h_ratio), False), width=int(18 * w_ratio), values=project_number_list)
    combobox_project_number.pack(side=tk.LEFT)
    combobox_project_number.bind("<<ComboboxSelected>>", on_select)
    combobox_project_number.bind("<Return>", entry_project)

    global button_check
    button_check = tk.Button(f21, text='查询', font=tku._ft(int(13 * h_ratio), False), image=query_project, bg="#eaf1f6", compound=tk.LEFT, command=lambda: summary_comments(f3), height=int(1 * h_ratio), fg='black', activebackground='blue')
    button_check.pack(side=tk.LEFT, padx=(int(20 * w_ratio), 0))

    tk.Label(f22, text="   项目：", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=tku._ft(int(13 * h_ratio), False)).pack(side=tk.LEFT)
    global project_name_list
    project_name_list = []

    global text_project_name
    text_project_name = tk.Text(f22, bg="#eaf1f6", font=tku._ft(int(13 * h_ratio), False), height=1, width=int(100 * w_ratio))
    text_project_name.pack(side=tk.LEFT)
    text_project_name['state'] = 'disabled'
    text_project_name['background'] = '#eaf1f6'

    style = Style()
    style.configure('panel1.Treeview', rowheight=25, font=("ABBvoice CNSG", int(13 * h_ratio)))
    style.configure('panel1.Treeview.Heading', font=("ABBvoice CNSG", int(13 * h_ratio)), background="#EFF1F5")

    tk.Label(f231, text="   统计：", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=tku._ft(int(13 * h_ratio), False)).pack(side=tk.LEFT, anchor='n')
    global Comments_Info_table
    table_ybar = tk.Scrollbar(f231)

    Comments_Info_table = Treeview(f231, style='panel1.Treeview', show='headings', selectmode='browse', columns=('a', 'b', 'c'), yscrollcommand=table_ybar.set, height=3)
    table_ybar.config(command=Comments_Info_table.yview)

    Comments_Info_table.column('a', anchor='center')
    Comments_Info_table.column('b', anchor='center')
    Comments_Info_table.column('c', anchor='center')

    Comments_Info_table.heading('a', text='内容')
    Comments_Info_table.heading('b', text='澄清条数')
    Comments_Info_table.heading('c', text='状态')
    Comments_Info_table.bind('<<TreeviewSelect>>', on_treeview_select)
    Comments_Info_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 1), pady=0)

    table_ybar.pack(side=tk.LEFT, fill=tk.Y)

    Comments_Info_table.tag_configure('open_line', foreground='green')
    Comments_Info_table.tag_configure('closed_line', foreground='red')

    global button_refresh
    button_refresh = tk.Button(f232, text='刷新', font=tku._ft(int(13 * h_ratio), False), image=refresh_project, bg="#eaf1f6", compound=tk.LEFT, command=lambda: refresh_table(f3), height=int(1 * h_ratio), fg='black', activebackground='blue')
    button_refresh.pack(side=tk.TOP, padx=int(20 * w_ratio))
    button_refresh['state'] = 'disabled'

    global button_view
    button_view = tk.Button(f232, text='查看', font=tku._ft(int(13 * h_ratio), False), image=view_project, bg="#eaf1f6", compound=tk.LEFT, command=lambda: view_comments(f3, sub_w_ratio, sub_h_ratio, selected_rows), height=int(1 * h_ratio), fg='black', activebackground='blue')
    button_view.pack(side=tk.TOP, padx=int(20 * w_ratio), pady=int(15*h_ratio))
    button_view['state'] = 'disabled'

    global button_create
    button_create = tk.Button(f232, text='编辑', font=tku._ft(int(13 * h_ratio), False), image=reply_project, bg="#eaf1f6", compound=tk.LEFT, command=lambda: create_comments(subparent, sub_w_ratio, sub_h_ratio, selected_rows), height=int(1 * h_ratio), fg='black', activebackground='blue')
    button_create.pack(side=tk.TOP, padx=int(20 * w_ratio))
    button_create['state'] = 'disabled'

    query_all()

    canvas.update_idletasks()
    canvas.config(scrollregion=canvas.bbox('all'))


def about_help(event):
    os.startfile(os.path.abspath('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\二次设计辅助工具【EBOM导SAP功能】答疑V1.2.pdf'))


def on_mousewheel(event):
    global canvas
    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
    # print(canvas.winfo_width())
    # canvas.create_text(canvas.winfo_width() - 10, 10, anchor='ne', text='滚动区↓')


def on_edit_mousewheel(event):
    global edit_canvas
    edit_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")


def refresh_table(parent):
    for widget in parent.winfo_children():
        widget.destroy()

    button_view['state'] = 'disabled'
    button_create['state'] = 'disabled'

    folder_path1 = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools'
    folder_path2 = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc'
    folder_path3 = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % combobox_project_number.get()

    Comments_Info_table.delete(*Comments_Info_table.get_children())  # 清空数据

    if os.path.exists(folder_path1):
        if is_folder_hidden(folder_path2) or os.path.exists(folder_path2):
            if is_folder_hidden(folder_path3) or os.path.exists(folder_path3):
                # print(is_folder_hidden(folder_path3))

                folder_content = {
                    "Technical_Clarification": "技术澄清",
                    "Drawing_Comments_1": "第1次图纸意见",
                    "Drawing_Comments_2": "第2次图纸意见",
                    "Drawing_Comments_3": "第3次图纸意见",
                    "Drawing_Comments_4": "第4次图纸意见",
                    "Drawing_Comments_5": "第5次图纸意见",
                    "Drawing_Comments_6": "第6次图纸意见"
                }

                folders = ["Technical_Clarification", "Drawing_Comments_1", "Drawing_Comments_2", "Drawing_Comments_3", "Drawing_Comments_4", "Drawing_Comments_5", "Drawing_Comments_6"]

                for folder in folders:
                    folder_count = 0
                    folder_status = "Open"

                    folder_path = os.path.join(folder_path3, folder)

                    for root_folder, subfolders, files in os.walk(folder_path):
                        # print(root_folder, subfolders, files)
                        if os.path.basename(root_folder) == folder:
                            folder_content_text = folder_content.get(folder, "")

                            close_folder_path = os.path.join(root_folder, "Close")
                            if os.path.exists(close_folder_path):
                                folder_status = "Closed"

                            for subfolder in subfolders:
                                # print(subfolder)
                                if has_only_numbers_and_underscore(subfolder):
                                    folder_count += 1
                            if folder_status == "Closed":
                                Comments_Info_table.insert('', 'end', values=(folder_content_text, folder_count, folder_status), tags="closed_line")
                            else:
                                Comments_Info_table.insert('', 'end', values=(folder_content_text, folder_count, folder_status), tags="open_line")
                        subfolders[:] = []
            else:
                tk.messagebox.showwarning("提示", "此项目尚无图纸意见传递文件夹\n\n请联系绘图员维护项目信息")
        else:
            tk.messagebox.showwarning("提示", "图纸意见传递文件夹丢失\n\n请联系管理员")
    else:
        tk.messagebox.showwarning("提示", "请连接内网")


def on_treeview_select(event):
    global selected_rows
    selected_rows = []
    selected_items = Comments_Info_table.selection()

    for item in selected_items:
        row_values = Comments_Info_table.item(item)['values']
        selected_rows.append(row_values)
    # print(selected_rows)

    global selected_design_type
    global selected_project_number
    selected_design_type = []
    selected_project_number = []

    # 详情按钮激活
    if len(selected_rows) == 1 and selected_rows[0][2] != 'Closed':
        button_create['state'] = 'normal'
        if int(selected_rows[0][1]) > 0:
            button_view['state'] = 'normal'
        else:
            button_view['state'] = 'disabled'
    elif len(selected_rows) == 1:
        if int(selected_rows[0][1]) > 0:
            button_view['state'] = 'normal'
        else:
            button_view['state'] = 'disabled'
        button_create['state'] = 'disabled'
    else:
        button_create['state'] = 'disabled'
        button_view['state'] = 'disabled'


def query_all():
    global result

    if os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools'):
        if is_folder_hidden('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb') or os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb'):
            if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db'):
                tk.messagebox.showwarning("提示", "数据库不存在，请联系管理员")
            else:
                conn = sqlite3.connect('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db')
                cursor = conn.cursor()  # 创建一个Cursor
                cursor.execute("SELECT * FROM project_data WHERE design_type IN ('图纸设计', '工程设计') ORDER BY CASE WHEN design_type = '图纸设计' OR design_type = '工程设计' THEN project_number END DESC")
                result = cursor.fetchall()

                result = [[element if element is not None else '' for element in row] for row in result]
                cursor.close()
                conn.close()

                text_project_name.delete(1.0, tk.END)
                table_items = Comments_Info_table.get_children()
                [Comments_Info_table.delete(table_item) for table_item in table_items]

                for i in range(0, len(result)):
                    project_number_list.append(result[i][2])
                project_number_list_set = list(set(project_number_list))
                project_number_list_set.sort(key=list(project_number_list).index)
                combobox_project_number.set('')
                combobox_project_number['values'] = project_number_list_set

        else:
            tk.messagebox.showwarning("提示", "数据库不存在，请联系管理员")
    else:
        tk.messagebox.showwarning("提示", "请连接内网")


def on_select(event):
    button_view['state'] = 'disabled'
    button_create['state'] = 'disabled'
    button_refresh['state'] = 'disabled'
    text_project_name.delete(1.0, tk.END)
    table_items = Comments_Info_table.get_children()
    [Comments_Info_table.delete(table_item) for table_item in table_items]

    for widget in f3.winfo_children():
        widget.destroy()


def is_folder_hidden(fpath):
    try:
        attrs = ctypes.windll.kernel32.GetFileAttributesW(fpath)  # attrs值为18表示该文件夹具有以下属性组合：只读 (1)、隐藏 (2) 和 子文件夹 (16)
        # print(attrs)
        if attrs != -1 and attrs & 2 == 2:  # 对于18（二进制为10010）与2（二进制为00010）进行按位与运算，结果为2（二进制为00010）
            return True
    except OSError:
        pass
    return False


def has_only_numbers_and_underscore(string):
    pattern = r'^[\d_]+$'
    match = re.match(pattern, string)
    return bool(match)


def entry_project(event):
    summary_comments(f3)


def summary_comments(parent):
    text_project_name.delete(1.0, tk.END)

    table_items = Comments_Info_table.get_children()  # 在插入treeview数据时，需要先清空treeview
    [Comments_Info_table.delete(table_item) for table_item in table_items]

    for widget in parent.winfo_children():
        widget.destroy()

    if combobox_project_number.get == '' or len(combobox_project_number.get()) != 9:
        tk.messagebox.showwarning("提示", "请选择9位项目号")
    else:
        if not combobox_project_number.get().isdigit():
            tk.messagebox.showwarning("提示", "请选择9位纯数字项目号")
        else:
            if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db'):
                tk.messagebox.showwarning("提示", "数据库不存在，请联系管理员")
            else:
                conn = sqlite3.connect('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db')
                cursor = conn.cursor()  # 创建一个Cursor
                query = "SELECT * FROM project_data WHERE project_number=?"
                cursor.execute(query, (combobox_project_number.get(),))
                result = cursor.fetchone()

                if result:
                    text_project_name['state'] = 'normal'
                    text_project_name.insert(tk.END, result[4])

                    folder_path1 = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools'
                    folder_path2 = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc'
                    folder_path3 = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % combobox_project_number.get()

                    if os.path.exists(folder_path1):
                        if is_folder_hidden(folder_path2) or os.path.exists(folder_path2):
                            if is_folder_hidden(folder_path3) or os.path.exists(folder_path3):
                                # print(is_folder_hidden(folder_path3))

                                folder_content = {
                                    "Technical_Clarification": "技术澄清",
                                    "Drawing_Comments_1": "第1次图纸意见",
                                    "Drawing_Comments_2": "第2次图纸意见",
                                    "Drawing_Comments_3": "第3次图纸意见",
                                    "Drawing_Comments_4": "第4次图纸意见",
                                    "Drawing_Comments_5": "第5次图纸意见",
                                    "Drawing_Comments_6": "第6次图纸意见"
                                }

                                folders = ["Technical_Clarification", "Drawing_Comments_1", "Drawing_Comments_2", "Drawing_Comments_3", "Drawing_Comments_4", "Drawing_Comments_5", "Drawing_Comments_6"]

                                for folder in folders:
                                    folder_count = 0
                                    folder_status = "Open"

                                    folder_path = os.path.join(folder_path3, folder)

                                    for root_folder, subfolders, files in os.walk(folder_path):
                                        # print(root_folder, subfolders, files)
                                        if os.path.basename(root_folder) == folder:
                                            folder_content_text = folder_content.get(folder, "")

                                            close_folder_path = os.path.join(root_folder, "Close")
                                            if os.path.exists(close_folder_path):
                                                folder_status = "Closed"

                                            for subfolder in subfolders:
                                                # print(subfolder)
                                                if has_only_numbers_and_underscore(subfolder):
                                                    folder_count += 1
                                            if folder_status == "Closed":
                                                Comments_Info_table.insert('', 'end', values=(folder_content_text, folder_count, folder_status), tags="closed_line")
                                            else:
                                                Comments_Info_table.insert('', 'end', values=(folder_content_text, folder_count, folder_status), tags="open_line")
                                        subfolders[:] = []
                                button_refresh['state'] = 'normal'
                            else:
                                tk.messagebox.showwarning("提示", "此项目尚无图纸意见传递文件夹\n\n请联系绘图员维护项目信息")
                        else:
                            tk.messagebox.showwarning("提示", "图纸意见传递文件夹丢失\n\n请联系管理员")
                    else:
                        tk.messagebox.showwarning("提示", "请连接内网")
                else:
                    tk.messagebox.showwarning("提示", "数据库中无此项目，请通过设计传递表功能录入该项目")


def on_receiver_select(event):
    entry_email.delete(0, "end")
    name = combobox_receiver.get()
    email_dict = dict(zip(receiver_list, receiver_email_list))
    email = email_dict.get(name, "")
    entry_email.insert(0, email)


def upload_comments(edit_window):
    if combobox_comments_type.get() == '':
        tk.messagebox.showwarning("提示", "请选择意见类型")
    elif text_comments_content.get("1.0", "end-1c") == '':
        tk.messagebox.showwarning("提示", "请填写意见内容")
    elif entry_email.get() == '' or '@cn.abb.com' not in entry_email.get():
        tk.messagebox.showwarning("提示", "目标邮箱必须包含@cn.abb.com")
    else:
        if len(Attachment_table.get_children()) == 0:
            result = tk.messagebox.askquestion("提示", "确定不需要上传相关附件？")
            if result == 'yes':
                with open(r'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\comments_temp_file\\Type.txt', "w", encoding="utf-8") as file:
                    file.write(combobox_comments_type.get())

                with open(r'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\comments_temp_file\\Creator.txt', "w", encoding="utf-8") as file:
                    file.write(os.getlogin())

                with open(r'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\comments_temp_file\\Comments.txt', "w", encoding="utf-8") as file:
                    file.write(text_comments_content.get("1.0", "end-1c"))

                if os.path.exists(folder_path1):
                    if is_folder_hidden(folder_path2) or os.path.exists(folder_path2):
                        if is_folder_hidden(folder_path3) or os.path.exists(folder_path3):
                            if is_folder_hidden(folder_path) or os.path.exists(folder_path):
                                current_time = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
                                parent_folder = os.path.join(folder_path, current_time)
                                os.makedirs(parent_folder)

                                subprocess.call(["attrib", "+h", parent_folder])

                                if is_folder_hidden(parent_folder) or os.path.exists(parent_folder):
                                    shutil.move(r'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\comments_temp_file\\Type.txt', parent_folder)
                                    type_file = os.path.join(parent_folder, 'Type.txt')
                                    subprocess.call(["attrib", "+h", type_file])
                                    type_file_exist = 0

                                    shutil.move(r'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\comments_temp_file\\Creator.txt', parent_folder)
                                    creator_file = os.path.join(parent_folder, 'Creator.txt')
                                    subprocess.call(["attrib", "+h", creator_file])
                                    creator_file_exist = 0

                                    shutil.move(r'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\comments_temp_file\\Comments.txt', parent_folder)
                                    comments_file = os.path.join(parent_folder, 'Comments.txt')
                                    subprocess.call(["attrib", "+h", comments_file])
                                    comments_file_exist = 0

                                    if is_folder_hidden(type_file) or os.path.exists(type_file):
                                        type_file_exist = 1
                                    if is_folder_hidden(creator_file) or os.path.exists(creator_file):
                                        creator_file_exist = 1
                                    if is_folder_hidden(comments_file) or os.path.exists(comments_file):
                                        comments_file_exist = 1
                                    if type_file_exist and creator_file_exist and comments_file_exist:
                                        tk.messagebox.showwarning("提示", "相关文件已经创建")
                                    elif not type_file_exist and not creator_file_exist and not comments_file_exist:
                                        tk.messagebox.showwarning("提示", "相关文件创建失败")
                                    else:
                                        tk.messagebox.showwarning("提示", "相关文件部分创建")
                                    # 检查窗口是否存在并且是否已经被销毁
                                    if edit_window.winfo_exists():
                                        edit_window.withdraw()
                                    create_table(f6, rows, columns, sub_w_ratio, sub_h_ratio)

        else:
            with open(r'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\comments_temp_file\\Type.txt', "w", encoding="utf-8") as file:
                file.write(combobox_comments_type.get())

            with open(r'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\comments_temp_file\\Creator.txt', "w", encoding="utf-8") as file:
                file.write(os.getlogin())

            with open(r'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\comments_temp_file\\Comments.txt', "w", encoding="utf-8") as file:
                file.write(text_comments_content.get("1.0", "end-1c"))

            if os.path.exists(folder_path1):
                if is_folder_hidden(folder_path2) or os.path.exists(folder_path2):
                    if is_folder_hidden(folder_path3) or os.path.exists(folder_path3):
                        if is_folder_hidden(folder_path) or os.path.exists(folder_path):
                            current_time = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
                            parent_folder = os.path.join(folder_path, current_time)
                            os.makedirs(parent_folder)
                            attachment_folder = os.path.join(parent_folder, 'Attachment')
                            os.makedirs(attachment_folder)

                            subprocess.call(["attrib", "+h", parent_folder])
                            subprocess.call(["attrib", "+h", attachment_folder])
                            # subprocess.call(["attrib", "-r", attachment_folder])

                            type_file_exist = 0
                            creator_file_exist = 0
                            comments_file_exist = 0

                            if is_folder_hidden(parent_folder) or os.path.exists(parent_folder):
                                shutil.move(r'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\comments_temp_file\\Type.txt', parent_folder)
                                type_file = os.path.join(parent_folder, 'Type.txt')
                                subprocess.call(["attrib", "+h", type_file])

                                shutil.move(r'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\comments_temp_file\\Creator.txt', parent_folder)
                                creator_file = os.path.join(parent_folder, 'Creator.txt')
                                subprocess.call(["attrib", "+h", creator_file])

                                shutil.move(r'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\comments_temp_file\\Comments.txt', parent_folder)
                                comments_file = os.path.join(parent_folder, 'Comments.txt')
                                subprocess.call(["attrib", "+h", comments_file])

                                if is_folder_hidden(type_file) or os.path.exists(type_file):
                                    type_file_exist = 1
                                if is_folder_hidden(creator_file) or os.path.exists(creator_file):
                                    creator_file_exist = 1
                                if is_folder_hidden(comments_file) or os.path.exists(comments_file):
                                    comments_file_exist = 1

                            attach_number = 0
                            if is_folder_hidden(attachment_folder) or os.path.exists(attachment_folder):
                                for attach_ in FilePath:
                                    # print(attach_)
                                    shutil.copy(attach_, attachment_folder)
                                    attach_file = os.path.join(attachment_folder, os.path.basename(attach_))
                                    if is_folder_hidden(attach_file) or os.path.exists(attach_file):
                                        subprocess.call(["attrib", "+h", attach_file])
                                        attach_number += 1

                            if type_file_exist and creator_file_exist and comments_file_exist and attach_number == len(FilePath):
                                tk.messagebox.showwarning("提示", "相关文件已经创建")
                            elif not type_file_exist and not creator_file_exist and not comments_file_exist and attach_number != len(FilePath):
                                tk.messagebox.showwarning("提示", "相关文件创建失败")
                            else:
                                tk.messagebox.showwarning("提示", "相关文件部分创建")
                            # 检查窗口是否存在并且是否已经被销毁
                            if edit_window.winfo_exists():
                                edit_window.withdraw()
                            create_table(f6, rows, columns, sub_w_ratio, sub_h_ratio)

        outlook = win32com.client.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        receiver = entry_email.get()
        mail.To = receiver

        mail.Subject = str(combobox_project_number.get()) + '-项目的图纸意见传递'
        content = 'Dear ' + re.split(r'[-.@]', receiver)[0].capitalize() + ':\n      你有一个新的图纸意见/答复已被创建，请进入FAST-【信息】-【图纸意见传递】中查阅'
        mail.Body = content
        mail.Importance = 2  # 设置重要性为高
        mail.Send()
        tk.messagebox.showwarning("提示", "邮件发送成功")
        # 确保在销毁窗口前检查它是否存在
        if edit_window.winfo_exists():
            edit_window.destroy()


def selectpath():
    filepath = tk.filedialog.askopenfilenames(title=u'请选择要上传的附件')
    if len(filepath) != 0:
        for file in filepath:
            file_dir = os.path.dirname(file)
            file_name = os.path.basename(file)
            file_size = str(math.ceil(int(os.path.getsize(file)) / 1024)) + ' KB'
            file_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(os.path.getmtime(file)))

            Attachment_table.insert('', 'end', values=(file_dir, file_name, file_size, file_time))
    global FilePath
    FilePath = filepath


def remove_file():
    for selected_file in selected_attachment:
        Attachment_table.delete(selected_file)
    button_remove_attachment['state'] = 'disabled'


def on_attachment_treeview_select(event):
    global selected_attachment
    selected_attachment = []
    selected_attachment = Attachment_table.selection()
    # print(selected_attachment)

    if len(selected_attachment) > 0:
        button_remove_attachment['state'] = 'normal'
    else:
        button_remove_attachment['state'] = 'disabled'


def create_comments(parent, w_ratio, h_ratio, selected_rows):
    edit_window = tk.Toplevel(parent)
    edit_window.grab_set()  # 禁用parent窗口的操作

    global edit_canvas

    edit_canvas = tk.Canvas(edit_window, width=int(1800 * w_ratio), height=int(740 * h_ratio), bg="#C9DBE9", borderwidth=0, highlightthickness=0)
    edit_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    edit_canvas.delete('all')  # 清空画布
    edit_canvas.update()
    edit_canvas.bind("<MouseWheel>", on_edit_mousewheel)

    scrollbar_v = tk.Scrollbar(master=edit_window)
    scrollbar_v.pack(side=tk.RIGHT, fill=tk.Y)
    scrollbar_v.config(command=edit_canvas.yview)
    edit_canvas.config(yscrollcommand=scrollbar_v.set)

    edit_parent = tk.Frame(edit_canvas, bg="#eaf1f6")

    edit_canvas.create_window(0, 0, width=int(1700 * w_ratio), anchor=tk.NW, window=edit_parent)

    f0 = tk.Frame(edit_parent, bg="#c9dbe9", bd=0)
    tk.Label(f0, text=selected_rows[0][0], bg="#c9dbe9", fg="black", height=int(1 * h_ratio), font=tku._ft(int(20 * h_ratio), True)).pack(fill=tk.X)
    f0.pack(side=tk.TOP, fill=tk.X)

    f1 = tk.Frame(edit_parent, bg="#eaf1f6")
    f1.pack(side=tk.TOP, fill=tk.X, expand=True)

    tk.Frame(f1, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

    f2 = tk.Frame(edit_parent, bg="#eaf1f6")
    f2.pack(side=tk.TOP, fill=tk.X, expand=True)

    tk.Frame(f2, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

    f3 = tk.Frame(edit_parent, bg="#eaf1f6")
    f3.pack(side=tk.TOP, fill=tk.X, expand=True)

    tk.Frame(f3, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

    f31 = tk.Frame(f3, bg="#eaf1f6")
    f31.pack(side=tk.LEFT, fill=tk.X, expand=True)

    f32 = tk.Frame(f3, bg="#eaf1f6")
    f32.pack(side=tk.LEFT, fill=tk.X, expand=True)

    tk.Frame(f3, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

    f4 = tk.Frame(edit_parent, bg="#eaf1f6")
    f4.pack(side=tk.TOP, fill=tk.X, expand=True)

    tk.Frame(f4, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

    f5 = tk.Frame(edit_parent, bg="#eaf1f6")
    f5.pack(side=tk.TOP, fill=tk.X, expand=True)

    tk.Frame(f5, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

    global f6
    f6 = tk.Frame(edit_parent, bg="#eaf1f6")
    f6.pack(side=tk.TOP, fill=tk.X, expand=True)

    tk.Frame(f6, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

    tk.Label(f1, text="   意见类型   ", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=tku._ft(int(12 * h_ratio), False)).pack(side=tk.LEFT)
    comments_type_list = ['设计澄清', '信息传递', '图纸设计错误', '客户新增', '客户更改', '图纸提交']
    global combobox_comments_type
    combobox_comments_type = ttk.Combobox(f1, font=tku._ft(int(11 * h_ratio), False), width=int(28 * w_ratio), state='readonly', values=comments_type_list)
    combobox_comments_type.pack(side=tk.LEFT)

    global dynamic_label
    dynamic_label = tk.Label(f1, height=int(1 * h_ratio), bg="#eaf1f6", fg="red", font=tku._ft(int(11 * h_ratio), False))
    dynamic_label.pack(side=tk.LEFT, padx=(int(10*w_ratio), 0))

    combobox_comments_type.bind("<<ComboboxSelected>>", change_label)

    tk.Label(f2, text="   意见内容   ", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=tku._ft(int(12 * h_ratio), False)).pack(side=tk.LEFT, anchor='n')
    global text_comments_content
    text_comments_content = tk.Text(f2, bg="white", font=tku._ft(int(11 * h_ratio), False), height=15, width=int(200 * w_ratio))
    text_comments_content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 1), pady=0)
    scrollbar1 = tk.Scrollbar(f2)
    scrollbar1.pack(side=tk.LEFT, fill=tk.Y, padx=(0, int(100*w_ratio)))
    scrollbar1.config(command=text_comments_content.yview)
    text_comments_content.config(yscrollcommand=scrollbar1.set)

    tk.Label(f31, text="   相关附件   ", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=tku._ft(int(12 * h_ratio), False)).pack(side=tk.LEFT, anchor='n')
    global Attachment_table
    Attachment_table = Treeview(f31, style='panel1.Treeview', show='headings', selectmode='extended', columns=('a', 'b', 'c', 'd'), height=5)
    Attachment_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 1), pady=0)
    scrollbar2 = tk.Scrollbar(f31)
    scrollbar2.pack(side=tk.LEFT, fill=tk.Y)
    scrollbar2.config(command=Attachment_table.yview)
    Attachment_table.config(yscrollcommand=scrollbar2.set)

    Attachment_table.column('a', width=int(700 * w_ratio), anchor='w')
    Attachment_table.column('b', width=int(320 * w_ratio), anchor='w')
    Attachment_table.column('c', width=int(200 * w_ratio), anchor='w')
    Attachment_table.column('d', width=int(220 * w_ratio), anchor='w')
    Attachment_table.heading('a', text='文件夹', anchor='w')
    Attachment_table.heading('b', text='文件名', anchor='w')
    Attachment_table.heading('c', text='文件大小', anchor='w')
    Attachment_table.heading('d', text='文件时间', anchor='w')
    Attachment_table.bind('<<TreeviewSelect>>', on_attachment_treeview_select)

    global button_select_attachment
    button_select_attachment = tk.Button(f32, text="选择", bg="#eaf1f6", command=selectpath, font=tku._ft(int(10 * h_ratio), False), height=int(1 * h_ratio), fg='black', activebackground='blue')
    button_select_attachment.pack(side=tk.TOP, padx=int(20 * w_ratio))

    global button_remove_attachment
    button_remove_attachment = tk.Button(f32, text="移除", bg="#eaf1f6", command=remove_file, font=tku._ft(int(10 * h_ratio), False), height=int(1 * h_ratio), fg='black', activebackground='blue')
    button_remove_attachment.pack(side=tk.TOP, padx=int(20 * w_ratio), pady=(int(20*h_ratio), 0))
    button_remove_attachment['state'] = 'disabled'

    tk.Label(f4, text="   邮件提醒   ", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=tku._ft(int(12 * h_ratio), False)).pack(side=tk.LEFT)
    global receiver_list
    global receiver_email_list
    receiver_list = []
    receiver_email_list = []
    if os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\PE_DE.xlsx"):
        workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\PE_DE.xlsx")
        worksheet = workbook['Sheet1']
        for i in range(1, worksheet.max_row + 1):
            receiver_list.append(worksheet.cell(row=i, column=1).value)
            receiver_email_list.append(worksheet.cell(row=i, column=2).value)
    global combobox_receiver
    combobox_receiver = ttk.Combobox(f4, font=tku._ft(int(11 * h_ratio), False), width=int(15 * w_ratio), state='readonly', values=receiver_list)
    combobox_receiver.pack(side=tk.LEFT)

    global entry_email
    entry_email = tk.Entry(f4, bg="#eaf1f6", font=tku._ft(int(11 * h_ratio), False), width=int(60 * w_ratio))
    entry_email.pack(side=tk.LEFT, padx=int(20 * w_ratio))

    combobox_receiver.bind('<<ComboboxSelected>>', on_receiver_select)

    global button_upload_comments
    button_upload_comments = tk.Button(f4, text="提交&通知", bg="#eaf1f6", command=lambda: upload_comments(edit_window), font=tku._ft(int(10 * h_ratio), False), height=int(1 * h_ratio), fg='black', activebackground='blue')
    button_upload_comments.pack(side=tk.LEFT, padx=0)

    tk.Label(f4, text="   (仅支持@cn.abb.com后缀邮箱)   ", height=int(1 * h_ratio), bg="#eaf1f6", fg="red", font=tku._ft(int(11 * h_ratio), False)).pack(side=tk.LEFT)

    tk.Label(f5, text='————————————————————————————————————————————————————————————————————————————————————————————————', bg="#eaf1f6", fg="grey", font=tku._ft(int(13 * h_ratio), True), justify='left').pack(side=tk.TOP, fill=tk.X, pady=(int(20*h_ratio), int(10*h_ratio)))

    global rows
    rows = 1
    global columns
    columns = 6
    tk.Label(f6, text="   沟通记录   ", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=tku._ft(int(12 * h_ratio), False)).pack(side=tk.LEFT, anchor='nw')
    create_table(f6, rows, columns, w_ratio, h_ratio)

    tk.Frame(f6, height=int(50 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

    edit_canvas.update_idletasks()
    edit_canvas.config(scrollregion=edit_canvas.bbox('all'))


def change_label(event):
    label_dict = {
        "设计澄清": "无法在传递表中列明的一些详细设计需求、针对客户图纸意见进行的澄清",
        "信息传递": "信息漏传递、传递错误或不规范",
        "图纸设计错误": "绘图员未按照输入文件设计、图纸设计错误",
        "客户新增": "客户新增要求",
        "客户更改": "客户更改要求",
        "图纸提交": "绘图员图纸提交",
    }
    for key, val in label_dict.items():
        if key == combobox_comments_type.get():
            dynamic_label.config(text=val)


def create_table(root, rows, cols, w_ratio, h_ratio):
    for widget in root.winfo_children():
        widget.destroy()
    table_frame = tk.Frame(root, bg="#eaf1f6")
    table_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

    global folder_path1
    folder_path1 = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools'
    global folder_path2
    folder_path2 = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc'
    global folder_path3
    folder_path3 = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % combobox_project_number.get()

    if os.path.exists(folder_path1):
        if is_folder_hidden(folder_path2) or os.path.exists(folder_path2):
            if is_folder_hidden(folder_path3) or os.path.exists(folder_path3):
                folder_content = {
                    "Technical_Clarification": "技术澄清",
                    "Drawing_Comments_1": "第1次图纸意见",
                    "Drawing_Comments_2": "第2次图纸意见",
                    "Drawing_Comments_3": "第3次图纸意见",
                    "Drawing_Comments_4": "第4次图纸意见",
                    "Drawing_Comments_5": "第5次图纸意见",
                    "Drawing_Comments_6": "第6次图纸意见"
                }

                for key, val in folder_content.items():
                    if val == selected_rows[0][0]:
                        folder = key
                global folder_path
                folder_path = os.path.join(folder_path3, folder)  # 502954820\Drawing_Comments_1层级

                for root_folder, subfolders, files in os.walk(folder_path):
                    # print(root_folder, subfolders, files)    # 502954820\Drawing_Comments_1、['2023_09_10_12_00_02', '2023_09_10_12_00_03', 'Close']、[]
                    if os.path.basename(root_folder) == folder:
                        close_folder_path = os.path.join(root_folder, "Close")
                        if os.path.exists(close_folder_path):
                            folder_status = "Closed"

                        folder_create_time = []
                        type_list = []
                        creator_list = []
                        comments_list = []
                        attachment_list = []

                        global sorted_folder_create_time
                        sorted_folder_create_time = []
                        global sorted_type_list
                        sorted_type_list = []
                        global sorted_creator_list
                        sorted_creator_list = []
                        global sorted_comments_list
                        sorted_comments_list = []

                        global sorted_attachment_list
                        sorted_attachment_list = []

                        folder_count = 1
                        for subfolder in subfolders:
                            # print(subfolder)
                            if has_only_numbers_and_underscore(subfolder):
                                folder_count += 1
                                folder_create_time.append(subfolder)

                                subfolder_path = os.path.join(folder_path, subfolder)  # 502954820\Drawing_Comments_1\2023_09_10_12_00_03层级

                                for root_folder1, subfolders1, files1 in os.walk(subfolder_path):  # subfolders1=['Attachment'], files1=['Comments.txt', 'Creator.txt', 'Type.txt']
                                    # print(root_folder1, subfolders1, files1)
                                    if os.path.basename(root_folder1) == subfolder:
                                        if 'Creator.txt' in files1:
                                            with open(os.path.join(root_folder1, 'Creator.txt'), 'r', encoding='utf-8') as file:
                                                creator_list.append(file.readline())
                                        else:
                                            creator_list.append('')

                                        if 'Type.txt' in files1:
                                            with open(os.path.join(root_folder1, 'Type.txt'), 'r', encoding='utf-8') as file:
                                                type_list.append(file.readline())
                                        else:
                                            type_list.append('')

                                        if 'Comments.txt' in files1:
                                            with open(os.path.join(root_folder1, 'Comments.txt'), 'r', encoding='utf-8') as file:
                                                comments_list.append(file.readlines())
                                        else:
                                            comments_list.append('')
                                    if len(subfolders1) > 0:
                                        attachment_path = os.path.join(subfolder_path, subfolders1[0])
                                        for root_folder2, subfolders2, files2 in os.walk(attachment_path):
                                            if os.path.basename(root_folder2) == subfolders1[0]:
                                                subfolders2[:] = []  # 忽略Attachment文件夹下方的文件夹，同时也避免对齐进行循环遍历
                                                attachment_list.append(files2)
                                    else:
                                        attachment_list.append('')
                                    subfolders1[:] = [d for d in subfolders1 if d not in ['Attachment']]  # subfolders1中排除Attachment文件夹，以避免再次进入上方循环进行子文件遍历
                                    # subfolders1[:] = []    # 也可以用这一句，直接不对任何子文件夹遍历

                        if folder_count > rows:
                            rows = folder_count

                        if len(folder_create_time) > 0 and len(attachment_list) > 0:
                            zipped = zip(folder_create_time, type_list, creator_list, comments_list, attachment_list)
                            sorted_zipped = sorted(zipped, key=lambda x: datetime.strptime(x[0], "%Y_%m_%d_%H_%M_%S"), reverse=True)  # 使早一些的日期排在列表后面
                            sorted_folder_create_time, sorted_type_list, sorted_creator_list, sorted_comments_list, sorted_attachment_list = zip(*sorted_zipped)
                        elif len(folder_create_time) > 0 and len(attachment_list) == 0:
                            zipped = zip(folder_create_time, type_list, creator_list, comments_list)
                            sorted_zipped = sorted(zipped, key=lambda x: datetime.strptime(x[0], "%Y_%m_%d_%H_%M_%S"), reverse=True)  # 使早一些的日期排在列表后面
                            sorted_folder_create_time, sorted_type_list, sorted_creator_list, sorted_comments_list = zip(*sorted_zipped)

                        # print(sorted_folder_create_time)
                        # print(sorted_type_list)
                        # print(sorted_creator_list)
                        # print(sorted_comments_list)
                        # print(sorted_attachment_list)

                    subfolders[:] = []  # 这样操作，便不会对子文件夹进行循环遍历，提升速度

    if rows > 1:
        for row in range(rows):
            row_frame = tk.Frame(table_frame, bg="grey")
            row_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

            global cell_text_list
            cell_text_list = []  # 存储每个cell_text的列表

            for col in range(cols):
                if col == 0:
                    if row == 0:
                        cell_text = tk.Text(row_frame, width=int(8 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert(tk.END, '序号')
                    else:
                        cell_text = tk.Text(row_frame, width=int(8 * w_ratio), wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert('end', row)
                        cell_text_list.append(cell_text)

                elif col == 1:
                    if row == 0:
                        cell_text = tk.Text(row_frame, width=int(100 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert(tk.END, '内容')
                    else:
                        cell_text = tk.Text(row_frame, width=int(100 * w_ratio), wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert('end', ''.join(sorted_comments_list[row-1]))
                        cell_text_list.append(cell_text)

                elif col == 2:
                    if row == 0:
                        cell_text = tk.Text(row_frame, width=int(60 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert(tk.END, '附件')
                    else:
                        cell_text = tk.Text(row_frame, width=int(60 * w_ratio), wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert('end', '\n'.join(sorted_attachment_list[row-1]))
                        cell_text_list.append(cell_text)

                elif col == 3:
                    if row == 0:
                        cell_text = tk.Text(row_frame, width=int(15 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert(tk.END, '类型')
                    else:
                        cell_text = tk.Text(row_frame, width=int(15 * w_ratio), wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert('end', sorted_type_list[row-1])
                        cell_text_list.append(cell_text)

                elif col == 4:
                    if row == 0:
                        cell_text = tk.Text(row_frame, width=int(15 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert(tk.END, '创建人')
                    else:
                        cell_text = tk.Text(row_frame, width=int(15 * w_ratio), wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert('end', sorted_creator_list[row-1])
                        cell_text_list.append(cell_text)
                else:
                    if row == 0:
                        cell_text = tk.Text(row_frame, width=int(20 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert(tk.END, '创建时间')
                    else:
                        cell_text = tk.Text(row_frame, width=int(20 * w_ratio), wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert('end', datetime.strptime(sorted_folder_create_time[row-1], "%Y_%m_%d_%H_%M_%S").strftime("%Y-%m-%d\n%H:%M:%S"))
                        cell_text_list.append(cell_text)

                if row == rows-1 and col == cols-1:
                    cell_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(1, 1), pady=(1, 1))
                elif row == rows-1:
                    cell_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(1, 0), pady=(1, 1))
                elif col == cols-1:
                    cell_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(1, 1), pady=(1, 0))
                else:
                    cell_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(1, 0), pady=(1, 0))
                cell_text.configure(borderwidth=1, relief='flat', font=tku._ft(int(12 * h_ratio), False))
                root.update()    # 为了获取正确的Text行高，需要update，否则计算的是文Text对象呈现在屏幕之前的字符数量
            max_height = 1
            for cell_text in cell_text_list:
                if cell_text.count("1.0", "end", "displaylines")[0] > max_height:
                    max_height = cell_text.count("1.0", "end", "displaylines")[0]

                cell_text.configure(height=max_height, font=tku._ft(int(12 * h_ratio), False))

    else:
        for row in range(rows):
            row_frame = tk.Frame(table_frame, bg="grey")
            row_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

            for col in range(cols):
                if col == 0:
                    cell_text = tk.Text(row_frame, width=int(8 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                    cell_text.insert(tk.END, '序号')
                elif col == 1:
                    cell_text = tk.Text(row_frame, width=int(100 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                    cell_text.insert(tk.END, '内容')
                elif col == 2:
                    cell_text = tk.Text(row_frame, width=int(60 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                    cell_text.insert(tk.END, '附件')
                elif col == 3:
                    cell_text = tk.Text(row_frame, width=int(15 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                    cell_text.insert(tk.END, '类型')
                elif col == 4:
                    cell_text = tk.Text(row_frame, width=int(15 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                    cell_text.insert(tk.END, '创建人')
                else:
                    cell_text = tk.Text(row_frame, width=int(20 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                    cell_text.insert(tk.END, '创建时间')
                if row == rows - 1 and col == cols - 1:
                    cell_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(1, 1), pady=(1, 1))
                elif row == rows - 1:
                    cell_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(1, 0), pady=(1, 1))
                elif col == cols - 1:
                    cell_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(1, 1), pady=(1, 0))
                else:
                    cell_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(1, 0), pady=(1, 0))

                cell_text.configure(borderwidth=1, relief='flat', font=tku._ft(int(12 * h_ratio), False))


def view_comments(parent, w_ratio, h_ratio, selected_rows):
    for widget in parent.winfo_children():
        widget.destroy()
    f0 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f0, text='                        ' + selected_rows[0][0], bg="#eaf1f6", fg="black", height=int(1 * h_ratio), font=tku._ft(int(15 * h_ratio), True)).pack(side=tk.LEFT, fill=tk.X, expand=True)

    global button_close
    button_close = tk.Button(f0, text="状态→Closed", bg="#eaf1f6", command=lambda: open2closed(parent, selected_rows), font=tku._ft(int(10 * h_ratio), False), height=int(1 * h_ratio), fg='black', activebackground='blue')
    button_close.pack(side=tk.RIGHT, fill=tk.X, padx=(0, int(20 * h_ratio)))
    if selected_rows[0][2] != 'Closed':
        button_close['state'] = 'normal'
    else:
        button_close['state'] = 'disabled'
    button_download_attachment = tk.Button(f0, text="附件", bg="#eaf1f6", command=lambda: download_attachment(parent, rows, columns, w_ratio, h_ratio, selected_rows), font=tku._ft(int(10 * h_ratio), False), height=int(1 * h_ratio), fg='black', activebackground='blue')
    button_download_attachment.pack(side=tk.RIGHT, fill=tk.X, padx=(0, int(20 * h_ratio)))

    f0.pack(side=tk.TOP, fill=tk.X)

    global f1
    f1 = tk.Frame(parent, bg="#eaf1f6")
    f1.pack(side=tk.TOP, fill=tk.X, expand=True)

    tk.Frame(f1, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

    global rows
    rows = 1
    global columns
    columns = 6

    # tk.Label(f1, text="   沟通记录   ", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=tku._ft(int(12 * h_ratio), False)).pack(side=tk.LEFT, anchor='nw')
    view_table(f1, rows, columns, w_ratio, h_ratio)

    tk.Frame(f1, height=int(50 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

    canvas.update_idletasks()
    canvas.config(scrollregion=canvas.bbox('all'))


def open2closed(parent, selected_rows):
    result = tk.messagebox.askquestion("提示", "确定将本项状态改为Closed（关闭）？\n\n一旦关闭，将无法对本项内意见进行答复")
    if result == 'yes':
        folder_path1 = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools'
        global folder_path2
        folder_path2 = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc'
        global folder_path3
        folder_path3 = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % combobox_project_number.get()
        if os.path.exists(folder_path1):
            if is_folder_hidden(folder_path2) or os.path.exists(folder_path2):
                if is_folder_hidden(folder_path3) or os.path.exists(folder_path3):
                    folder_content = {
                        "Technical_Clarification": "技术澄清",
                        "Drawing_Comments_1": "第1次图纸意见",
                        "Drawing_Comments_2": "第2次图纸意见",
                        "Drawing_Comments_3": "第3次图纸意见",
                        "Drawing_Comments_4": "第4次图纸意见",
                        "Drawing_Comments_5": "第5次图纸意见",
                        "Drawing_Comments_6": "第6次图纸意见"
                    }

                    for key, val in folder_content.items():
                        if val == selected_rows[0][0]:
                            folder = key

                    folder_path = os.path.join(folder_path3, folder)  # 502954820\Drawing_Comments_1层级
                    if is_folder_hidden(folder_path) or os.path.exists(folder_path):
                        os.makedirs(os.path.join(folder_path, 'Close'))
                        subprocess.call(["attrib", "+h", os.path.join(folder_path, 'Close')])
                        refresh_table(parent)


def download_attachment(parent, rows, columns, w_ratio, h_ratio, selected_rows):
    download_window = tk.Toplevel(parent, bg="#eaf1f6")
    download_window.grab_set()  # 禁用parent窗口的操作
    tku.center_window(download_window)
    download_window.geometry("%dx%d" % (int(450*w_ratio), int(300*h_ratio)))
    download_window.title('下载附件')

    f1 = tk.Frame(download_window, bg="#eaf1f6", width=int(500*w_ratio))
    f1.pack(side=tk.TOP, fill=tk.X, expand=True, anchor='nw', pady=(int(20 * h_ratio), int(150 * h_ratio)), padx=int(10 * h_ratio))

    tk.Label(f1, text="意见序号：", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=tku._ft(int(12 * h_ratio), False)).pack(side=tk.LEFT)
    global attachment_number_list
    attachment_number_list = []

    # print(sorted_attachment_list)
    for i, item in enumerate(sorted_attachment_list):
        if item != [] and item != '':
            attachment_number_list.append(i+1)

    global combobox_attachment_number
    combobox_attachment_number = ttk.Combobox(f1, font=tku._ft(int(10 * h_ratio), False), width=int(18 * w_ratio), state='readonly', values=attachment_number_list)
    combobox_attachment_number.pack(side=tk.LEFT)
    combobox_attachment_number.bind("<<ComboboxSelected>>", on_select_attachment)

    global button_confirm_download
    button_confirm_download = tk.Button(f1, text="下载", bg="#eaf1f6", command=lambda: confirm_download(download_window, combobox_attachment_number.get(), selected_rows), font=tku._ft(int(10 * h_ratio), False), height=int(1 * h_ratio), fg='black', activebackground='blue')
    button_confirm_download.pack(side=tk.LEFT, fill=tk.X, padx=(int(20 * h_ratio), 0))
    button_confirm_download['state'] = 'disabled'


def confirm_download(download_window, selected_item, selected_rows):
    home_path = os.path.expanduser("~")
    desktop_path = os.path.join(home_path, "Desktop")

    desktop_path = askdirectory(title=u'请选择导出文件夹', initialdir=desktop_path)
    if not desktop_path:
        tk.messagebox.showwarning("提示", "未选择保存路径，导出操作已取消")
        return

    folder_path1 = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools'
    global folder_path2
    folder_path2 = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc'
    global folder_path3
    folder_path3 = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % combobox_project_number.get()
    if os.path.exists(folder_path1):
        if is_folder_hidden(folder_path2) or os.path.exists(folder_path2):
            if is_folder_hidden(folder_path3) or os.path.exists(folder_path3):
                folder_content = {
                    "Technical_Clarification": "技术澄清",
                    "Drawing_Comments_1": "第1次图纸意见",
                    "Drawing_Comments_2": "第2次图纸意见",
                    "Drawing_Comments_3": "第3次图纸意见",
                    "Drawing_Comments_4": "第4次图纸意见",
                    "Drawing_Comments_5": "第5次图纸意见",
                    "Drawing_Comments_6": "第6次图纸意见"
                }

                for key, val in folder_content.items():
                    if val == selected_rows[0][0]:
                        folder = key

                folder_path = os.path.join(folder_path3, folder)  # 502954820\Drawing_Comments_1层级

                attachment_document_list = []
                attachment_document_amount = 0
                if is_folder_hidden(folder_path) or os.path.exists(folder_path):
                    for i, item in enumerate(sorted_attachment_list):
                        if int(selected_item)-1 == i:
                            parent_folder = os.path.join(folder_path, sorted_folder_create_time[i])    # 502954820\Drawing_Comments_1\2023_09_10_12_00_03层级
                            attachment_document_time = sorted_folder_create_time[i]

                    if is_folder_hidden(parent_folder) or os.path.exists(parent_folder):
                        attachment_document_folder = os.path.join(parent_folder, 'Attachment')
                        if is_folder_hidden(attachment_document_folder) or os.path.exists(attachment_document_folder):


                            folder_name = '附件' + attachment_document_time
                            dst = os.path.join(desktop_path, folder_name)

                            if not os.path.exists(dst):
                                os.makedirs(dst)
                                dst = 'C:\\Users\\%s\\Desktop\\%s\\' % (os.getlogin(), '附件'+attachment_document_time)
                            else:
                                i = 1
                                while True:
                                    new_folder_name = f'{folder_name} ({i})'
                                    new_dst = os.path.join(desktop_path, new_folder_name)
                                    if not os.path.exists(new_dst):
                                        os.makedirs(new_dst)
                                        dst = new_dst
                                        break
                                    i += 1

                            for root_folder_attach, subfolders_attach, files_attach in os.walk(attachment_document_folder):
                                if os.path.basename(root_folder_attach) == 'Attachment':
                                    subfolders_attach[:] = []  # 忽略Attachment文件夹下方的文件夹，同时也避免对齐进行循环遍历
                                    attachment_document_list = files_attach

                            for attach_ in attachment_document_list:
                                shutil.copy(os.path.join(attachment_document_folder, attach_), dst)
                                subprocess.call(["attrib", "-h", os.path.join(dst, attach_)])

                                if is_folder_hidden(os.path.join(dst, attach_)) or os.path.exists(os.path.join(dst, attach_)):
                                    attachment_document_amount += 1

                            if attachment_document_amount == len(attachment_document_list):
                                tk.messagebox.showwarning("提示", "相关文件已经创建")
                            elif attachment_document_amount == 0:
                                tk.messagebox.showwarning("提示", "相关文件创建失败")
                            else:
                                tk.messagebox.showwarning("提示", "相关文件部分创建")

                            download_window.destroy()


def on_select_attachment(event):
    if combobox_attachment_number.get != '':
        button_confirm_download['state'] = 'normal'
    else:
        button_confirm_download['state'] = 'disabled'


def view_table(root, rows, cols, w_ratio, h_ratio):
    table_frame = tk.Frame(root, bg="#eaf1f6")
    table_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

    folder_path1 = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools'
    folder_path2 = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc'
    folder_path3 = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % combobox_project_number.get()

    if os.path.exists(folder_path1):
        if is_folder_hidden(folder_path2) or os.path.exists(folder_path2):
            if is_folder_hidden(folder_path3) or os.path.exists(folder_path3):
                folder_content = {
                    "Technical_Clarification": "技术澄清",
                    "Drawing_Comments_1": "第1次图纸意见",
                    "Drawing_Comments_2": "第2次图纸意见",
                    "Drawing_Comments_3": "第3次图纸意见",
                    "Drawing_Comments_4": "第4次图纸意见",
                    "Drawing_Comments_5": "第5次图纸意见",
                    "Drawing_Comments_6": "第6次图纸意见"
                }

                for key, val in folder_content.items():
                    if val == selected_rows[0][0]:
                        folder = key

                folder_path = os.path.join(folder_path3, folder)  # 502954820\Drawing_Comments_1层级

                for root_folder, subfolders, files in os.walk(folder_path):
                    # print(root_folder, subfolders, files)    # 502954820\Drawing_Comments_1、['2023_09_10_12_00_02', '2023_09_10_12_00_03', 'Close']、[]
                    if os.path.basename(root_folder) == folder:
                        close_folder_path = os.path.join(root_folder, "Close")
                        if os.path.exists(close_folder_path):
                            folder_status = "Closed"

                        folder_create_time = []
                        type_list = []
                        creator_list = []
                        comments_list = []
                        attachment_list = []

                        global sorted_folder_create_time
                        sorted_folder_create_time = []
                        global sorted_type_list
                        sorted_type_list = []
                        global sorted_creator_list
                        sorted_creator_list = []
                        global sorted_comments_list
                        sorted_comments_list = []

                        global sorted_attachment_list
                        sorted_attachment_list = []

                        folder_count = 1
                        for subfolder in subfolders:
                            # print(subfolder)
                            if has_only_numbers_and_underscore(subfolder):
                                folder_count += 1
                                folder_create_time.append(subfolder)

                                subfolder_path = os.path.join(folder_path, subfolder)  # 502954820\Drawing_Comments_1\2023_09_10_12_00_03层级

                                for root_folder1, subfolders1, files1 in os.walk(subfolder_path):  # subfolders1=['Attachment'], files1=['Comments.txt', 'Creator.txt', 'Type.txt']
                                    # print(root_folder1, subfolders1, files1)
                                    if os.path.basename(root_folder1) == subfolder:
                                        if 'Creator.txt' in files1:
                                            with open(os.path.join(root_folder1, 'Creator.txt'), 'r', encoding='utf-8') as file:
                                                creator_list.append(file.readline())
                                        else:
                                            creator_list.append('')

                                        if 'Type.txt' in files1:
                                            with open(os.path.join(root_folder1, 'Type.txt'), 'r', encoding='utf-8') as file:
                                                type_list.append(file.readline())
                                        else:
                                            type_list.append('')

                                        if 'Comments.txt' in files1:
                                            with open(os.path.join(root_folder1, 'Comments.txt'), 'r', encoding='utf-8') as file:
                                                comments_list.append(file.readlines())
                                        else:
                                            comments_list.append('')
                                    if len(subfolders1) > 0:
                                        attachment_path = os.path.join(subfolder_path, subfolders1[0])
                                        for root_folder2, subfolders2, files2 in os.walk(attachment_path):
                                            if os.path.basename(root_folder2) == subfolders1[0]:
                                                subfolders2[:] = []  # 忽略Attachment文件夹下方的文件夹，同时也避免对齐进行循环遍历
                                                attachment_list.append(files2)
                                    else:
                                        attachment_list.append('')
                                    subfolders1[:] = [d for d in subfolders1 if d not in ['Attachment']]  # subfolders1中排除Attachment文件夹，以避免再次进入上方循环进行子文件遍历
                                    # subfolders1[:] = []    # 也可以用这一句，直接不对任何子文件夹遍历

                        if folder_count > rows:
                            rows = folder_count

                        if len(folder_create_time) > 0 and len(attachment_list) > 0:
                            zipped = zip(folder_create_time, type_list, creator_list, comments_list, attachment_list)
                            sorted_zipped = sorted(zipped, key=lambda x: datetime.strptime(x[0], "%Y_%m_%d_%H_%M_%S"), reverse=True)  # 使早一些的日期排在列表后面
                            sorted_folder_create_time, sorted_type_list, sorted_creator_list, sorted_comments_list, sorted_attachment_list = zip(*sorted_zipped)
                        elif len(folder_create_time) > 0 and len(attachment_list) == 0:
                            zipped = zip(folder_create_time, type_list, creator_list, comments_list)
                            sorted_zipped = sorted(zipped, key=lambda x: datetime.strptime(x[0], "%Y_%m_%d_%H_%M_%S"), reverse=True)  # 使早一些的日期排在列表后面
                            sorted_folder_create_time, sorted_type_list, sorted_creator_list, sorted_comments_list = zip(*sorted_zipped)

                        # print(sorted_folder_create_time)
                        # print(sorted_type_list)
                        # print(sorted_creator_list)
                        # print(sorted_comments_list)
                        # print(sorted_attachment_list)

                    subfolders[:] = []  # 这样操作，便不会对子文件夹进行循环遍历，提升速度

    if rows > 1:
        for row in range(rows):
            row_frame = tk.Frame(table_frame, bg="grey")
            row_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

            global cell_text_list
            cell_text_list = []  # 存储每个cell_text的列表

            for col in range(cols):
                if col == 0:
                    if row == 0:
                        cell_text = tk.Text(row_frame, width=int(8 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert(tk.END, '序号')
                    else:
                        cell_text = tk.Text(row_frame, width=int(8 * w_ratio), wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert('end', row)
                        cell_text_list.append(cell_text)

                elif col == 1:
                    if row == 0:
                        cell_text = tk.Text(row_frame, width=int(80 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert(tk.END, '内容')
                    else:
                        cell_text = tk.Text(row_frame, width=int(80 * w_ratio), wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert('end', ''.join(sorted_comments_list[row-1]))
                        cell_text_list.append(cell_text)

                elif col == 2:
                    if row == 0:
                        cell_text = tk.Text(row_frame, width=int(40 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert(tk.END, '附件')
                    else:
                        cell_text = tk.Text(row_frame, width=int(40 * w_ratio), wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert('end', '\n'.join(sorted_attachment_list[row-1]))
                        cell_text_list.append(cell_text)

                elif col == 3:
                    if row == 0:
                        cell_text = tk.Text(row_frame, width=int(10 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert(tk.END, '类型')
                    else:
                        cell_text = tk.Text(row_frame, width=int(10 * w_ratio), wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert('end', sorted_type_list[row-1])
                        cell_text_list.append(cell_text)

                elif col == 4:
                    if row == 0:
                        cell_text = tk.Text(row_frame, width=int(10 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert(tk.END, '创建人')
                    else:
                        cell_text = tk.Text(row_frame, width=int(10 * w_ratio), wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert('end', sorted_creator_list[row-1])
                        cell_text_list.append(cell_text)
                else:
                    if row == 0:
                        cell_text = tk.Text(row_frame, width=int(20 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert(tk.END, '创建时间')
                    else:
                        cell_text = tk.Text(row_frame, width=int(20 * w_ratio), wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                        cell_text.insert('end', datetime.strptime(sorted_folder_create_time[row-1], "%Y_%m_%d_%H_%M_%S").strftime("%Y-%m-%d\n%H:%M:%S"))
                        cell_text_list.append(cell_text)

                if row == rows-1 and col == cols-1:
                    cell_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(1, 1), pady=(1, 1))
                elif row == rows-1:
                    cell_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(1, 0), pady=(1, 1))
                elif col == cols-1:
                    cell_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(1, 1), pady=(1, 0))
                else:
                    cell_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(1, 0), pady=(1, 0))
                cell_text.configure(borderwidth=1, relief='flat', font=tku._ft(int(12 * h_ratio), False))
                root.update()    # 为了获取正确的Text行高，需要update，否则计算的是文Text对象呈现在屏幕之前的字符数量
            max_height = 1
            for cell_text in cell_text_list:
                if cell_text.count("1.0", "end", "displaylines")[0] > max_height:
                    max_height = cell_text.count("1.0", "end", "displaylines")[0]

                cell_text.configure(height=max_height, font=tku._ft(int(12 * h_ratio), False))

    else:
        for row in range(rows):
            row_frame = tk.Frame(table_frame, bg="grey")
            row_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

            for col in range(cols):
                if col == 0:
                    cell_text = tk.Text(row_frame, width=int(8 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                    cell_text.insert(tk.END, '序号')
                elif col == 1:
                    cell_text = tk.Text(row_frame, width=int(100 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                    cell_text.insert(tk.END, '内容')
                elif col == 2:
                    cell_text = tk.Text(row_frame, width=int(60 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                    cell_text.insert(tk.END, '附件')
                elif col == 3:
                    cell_text = tk.Text(row_frame, width=int(15 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                    cell_text.insert(tk.END, '类型')
                elif col == 4:
                    cell_text = tk.Text(row_frame, width=int(15 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                    cell_text.insert(tk.END, '创建人')
                else:
                    cell_text = tk.Text(row_frame, width=int(20 * w_ratio), height=1, wrap='word', bg='#eaf1f6', font=tku._ft(int(12 * h_ratio), False))
                    cell_text.insert(tk.END, '创建时间')
                if row == rows - 1 and col == cols - 1:
                    cell_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(1, 1), pady=(1, 1))
                elif row == rows - 1:
                    cell_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(1, 0), pady=(1, 1))
                elif col == cols - 1:
                    cell_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(1, 1), pady=(1, 0))
                else:
                    cell_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(1, 0), pady=(1, 0))

                cell_text.configure(borderwidth=1, relief='flat', font=tku._ft(int(12 * h_ratio), False))