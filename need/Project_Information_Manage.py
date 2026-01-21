import tkinter as tk
from tkinter import ttk

from tkinter import IntVar
from tkinter.ttk import Treeview, Style
import need.tkutils as tku
import shutil
import os
import sqlite3
from tkcalendar import CCalendar
import ctypes
from time import *

from openpyxl import load_workbook
from need.custom_dialogs import CustomDialog, center_window, Tooltip, image_label
from cryptography.fernet import Fernet
import base64
import traceback
import pyrfc
import logging
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

FilePath = ""  # 设置一个地址变量


def main(parent, w_ratio, h_ratio):
    global quit_filter
    quit_filter = tk.PhotoImage(file='ico\\no_filter.png')

    global filter
    filter = tk.PhotoImage(file='ico\\filter.png')

    global detail
    detail = tk.PhotoImage(file='ico\\detail.png')

    global trans
    trans = tk.PhotoImage(file='ico\\trans.png')

    global edit_single
    edit_single = tk.PhotoImage(file='ico\\edit.png')

    global edit_multi
    edit_multi = tk.PhotoImage(file='ico\\edit_multi.png')

    global query_project1
    query_project1 = tk.PhotoImage(file="ico\\search.png")

    global refresh_project
    refresh_project = tk.PhotoImage(file="ico\\refresh.png")

    global confirm_trans
    confirm_trans = tk.PhotoImage(file="ico\\confirm_trans.png")

    global subparent
    subparent = parent

    global sub_w_ratio
    sub_w_ratio = w_ratio

    global sub_h_ratio
    sub_h_ratio = h_ratio

    # global canvas
    # canvas = tk.Canvas(parent, width=int(1700 * w_ratio), height=int(640 * h_ratio), bg="#c9dbe9", borderwidth=0, highlightthickness=0)
    # canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    # canvas.update()
    # canvas.bind("<MouseWheel>", on_mousewheel)
    #
    # scrollbar_v = tk.Scrollbar(master=parent)
    # scrollbar_v.pack(side=tk.RIGHT, fill=tk.Y)
    # scrollbar_v.config(command=canvas.yview)
    # canvas.config(yscrollcommand=scrollbar_v.set)

    # content = tk.Frame(canvas)
    #
    # canvas.create_window(0, 1, width=int(1700 * w_ratio), anchor=tk.NW, window=content)

    f1 = tk.Frame(parent, bg="#c9dbe9", bd=0)
    # im = tku.image_label(f1, "ico\\help.png", int(30 * h_ratio), int(30 * h_ratio), False)
    # im.configure(bg="#c9dbe9")
    # im.bind('<Button-1>', about_help)
    # im.pack(side=tk.RIGHT)
    tk.Label(f1, text="欢迎使用项目信息管理功能", bg="#c9dbe9", fg="black", height=int(1 * h_ratio), font=("ABBvoice CNSG", int(20 * h_ratio), "bold")).pack(fill=tk.X)
    f1.pack(fill=tk.X)

    f2 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    f2.pack(side=tk.TOP, fill=tk.X)

    f3 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    f3.pack(side=tk.TOP, fill=tk.X)

    f21 = tk.Frame(f2, bg="#eaf1f6")
    f21.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(1, 1))

    f21_r = tk.Frame(f2, bg="#eaf1f6")
    f21_r.pack(side=tk.LEFT, fill=tk.BOTH, padx=(1, 1))

    f211 = tk.Frame(f21, bg="#eaf1f6")
    f211.pack(side=tk.TOP, fill=tk.BOTH, expand=True, pady=(int(10 * h_ratio), int(10 * h_ratio)))

    f212 = tk.Frame(f21, bg="#eaf1f6")
    f212.pack(side=tk.TOP, fill=tk.BOTH, expand=True, pady=(0, int(10 * h_ratio)))

    f213 = tk.Frame(f21, bg="#eaf1f6")
    f213.pack(side=tk.TOP, fill=tk.BOTH, expand=True, pady=(0, int(10 * h_ratio)))

    f214 = tk.Frame(f21, bg="#eaf1f6")
    f214.pack(side=tk.TOP, fill=tk.BOTH, expand=True, pady=(0, int(10 * h_ratio)))

    f215 = tk.Frame(f21, bg="#eaf1f6")
    f215.pack(side=tk.TOP, fill=tk.BOTH, expand=True, pady=(0, int(10 * h_ratio)))

    f216 = tk.Frame(f21, bg="#eaf1f6")
    f216.pack(side=tk.TOP, fill=tk.X, expand=True, anchor=tk.E, pady=(0, int(20 * h_ratio)))

    f31 = tk.Frame(f3, bg="#eaf1f6")
    f31.pack(fill=tk.BOTH, expand=True, padx=(1, 1))

    tk.Label(f211, text="设计类型 ", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
    global design_type_list
    design_type_list = []

    parent.option_add("*TCombobox*Listbox.font", ("ABBvoice CNSG", int(13 * h_ratio)))

    global combobox_design_type
    combobox_design_type = ttk.Combobox(f211, font=("ABBvoice CNSG", int(12 * h_ratio)), width=int(20 * w_ratio), state='readonly', values=design_type_list)
    combobox_design_type.pack(side=tk.LEFT, fill=tk.X, expand=True)
    combobox_design_type.bind("<<ComboboxSelected>>", on_select)

    tk.Label(f211, text="    项目号", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
    global project_number_list
    project_number_list = []

    global combobox_project_number
    combobox_project_number = ttk.Combobox(f211, font=("ABBvoice CNSG", int(12 * h_ratio)), width=int(30 * w_ratio), values=project_number_list)
    combobox_project_number.pack(side=tk.LEFT, fill=tk.X, expand=True)
    combobox_project_number.bind("<<ComboboxSelected>>", on_select)

    tk.Label(f211, text="    项目名", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
    global project_name_list
    project_name_list = []

    global combobox_project_name
    combobox_project_name = ttk.Combobox(f211, font=("ABBvoice CNSG", int(12 * h_ratio)), width=int(120 * w_ratio), values=project_name_list)
    combobox_project_name.pack(side=tk.LEFT, fill=tk.X, expand=True)
    combobox_project_name.bind("<<ComboboxSelected>>", on_select)

    tk.Label(f212, text="产品类型 ", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
    global product_type_list
    product_type_list = []

    global combobox_product_type
    combobox_product_type = ttk.Combobox(master=f212, font=("ABBvoice CNSG", int(12 * h_ratio)), width=int(20 * w_ratio), state='readonly', values=product_type_list)
    combobox_product_type.pack(side=tk.LEFT, fill=tk.X, expand=True)
    combobox_product_type.bind("<<ComboboxSelected>>", on_select)

    tk.Label(f212, text="       柜型", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
    global panel_type_list
    panel_type_list = []

    global combobox_panel_type
    combobox_panel_type = ttk.Combobox(master=f212, font=("ABBvoice CNSG", int(12 * h_ratio)), width=int(30 * w_ratio), state='readonly', values=panel_type_list)
    combobox_panel_type.pack(side=tk.LEFT, fill=tk.X, expand=True)
    combobox_panel_type.bind("<<ComboboxSelected>>", on_select)

    tk.Label(f212, text="    工程师", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
    global PE_list
    PE_list = []

    global combobox_PE
    combobox_PE = ttk.Combobox(master=f212, font=("ABBvoice CNSG", int(12 * h_ratio)), width=int(30 * w_ratio), state='readonly', values=PE_list)
    combobox_PE.pack(side=tk.LEFT, fill=tk.X, expand=True)
    combobox_PE.bind("<<ComboboxSelected>>", on_select)

    tk.Label(f212, text="    绘图员", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
    global DE_list
    DE_list = []

    global combobox_DE
    combobox_DE = ttk.Combobox(master=f212, font=("ABBvoice CNSG", int(12 * h_ratio)), width=int(30 * w_ratio), state='readonly', values=DE_list)
    combobox_DE.pack(side=tk.LEFT, fill=tk.X, expand=True)
    combobox_DE.bind("<<ComboboxSelected>>", on_select)

    tk.Label(f212, text="     状态", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
    global status_list
    status_list = []

    global combobox_status
    combobox_status = ttk.Combobox(master=f212, font=("ABBvoice CNSG", int(12 * h_ratio)), width=int(32 * w_ratio), state='readonly', values=status_list)
    combobox_status.pack(side=tk.LEFT, fill=tk.X, expand=True)
    combobox_status.bind("<<ComboboxSelected>>", on_select)

    tk.Label(f213, text="  智能化  ", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
    global intelligent_list
    intelligent_list = []

    global combobox_intelligent
    combobox_intelligent = ttk.Combobox(master=f213, font=("ABBvoice CNSG", int(12 * h_ratio)), width=int(200 * w_ratio), state='readonly', values=intelligent_list)
    combobox_intelligent.pack(side=tk.LEFT, fill=tk.X, expand=True)
    combobox_intelligent.bind("<<ComboboxSelected>>", on_select)

    tk.Label(f214, text="  主保护  ", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
    global protection_list
    protection_list = []

    global combobox_protection
    combobox_protection = ttk.Combobox(master=f214, font=("ABBvoice CNSG", int(12 * h_ratio)), width=int(200 * w_ratio), values=protection_list)
    combobox_protection.pack(side=tk.LEFT, fill=tk.X, expand=True)
    combobox_protection.bind("<<ComboboxSelected>>", on_select)

    tk.Label(f215, text="弧光方案 ", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
    global arclight_list
    arclight_list = []

    global combobox_arclight
    combobox_arclight = ttk.Combobox(master=f215, font=("ABBvoice CNSG", int(12 * h_ratio)), width=int(200 * w_ratio), values=arclight_list)
    combobox_arclight.pack(side=tk.LEFT, fill=tk.X, expand=True)
    combobox_arclight.bind("<<ComboboxSelected>>", on_select)

    tk.Label(f216, text="管理功能 ", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)

    global button_detail
    button_detail = tk.Button(f216, text="详细", font=("ABBvoice CNSG", int(12 * h_ratio)), image=detail, compound=tk.LEFT, bg="#eaf1f6", command=lambda: project_detail(subparent, sub_w_ratio, sub_h_ratio, selected_rows), height=int(1 * h_ratio), fg='black', activebackground='blue')
    button_detail.pack(side=tk.LEFT, padx=(0, int(20 * w_ratio)))
    button_detail['state'] = 'disabled'

    global button_change_single
    button_change_single = tk.Button(f216, text="修改(单行号)", font=("ABBvoice CNSG", int(12 * h_ratio)), image=edit_single, compound=tk.LEFT, bg="#eaf1f6", command=lambda: project_change_single(subparent, sub_w_ratio, sub_h_ratio, selected_rows), height=int(1 * h_ratio), fg='black', activebackground='blue')
    button_change_single.pack(side=tk.LEFT, padx=(0, int(20 * w_ratio)))
    button_change_single['state'] = 'disabled'

    global button_change_multi
    button_change_multi = tk.Button(f216, text="修改(多行号)", font=("ABBvoice CNSG", int(12 * h_ratio)), image=edit_multi, compound=tk.LEFT, bg="#eaf1f6", command=lambda: project_change_multi(subparent, sub_w_ratio, sub_h_ratio, selected_rows), height=int(1 * h_ratio), fg='black', activebackground='blue')
    button_change_multi.pack(side=tk.LEFT, padx=(0, int(20 * w_ratio)))
    button_change_multi['state'] = 'disabled'

    global button_ahead2other
    button_ahead2other = tk.Button(f216, text="提前设计→图纸/工程设计", font=("ABBvoice CNSG", int(12 * h_ratio)), image=trans, compound=tk.LEFT, bg="#eaf1f6", command=lambda: project_ahead2other(subparent, sub_w_ratio, sub_h_ratio, selected_rows), height=int(1 * h_ratio), fg='black', activebackground='blue')
    button_ahead2other.pack(side=tk.LEFT, padx=(0, int(20 * w_ratio)))
    button_ahead2other['state'] = 'disabled'

    tk.Label(f216, text="(项目名、主保护、弧光方案支持输入、模糊查询)", height=int(1 * h_ratio), bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(10 * h_ratio))).pack(side=tk.RIGHT)

    tk.Button(f21_r, text="筛选", bg="#eaf1f6", font=("ABBvoice CNSG", int(12 * h_ratio)), image=filter, compound=tk.LEFT, command=query_some, height=int(1 * h_ratio), fg='black', activebackground='blue').pack(side=tk.TOP, padx=int(20 * w_ratio), pady=int(50 * h_ratio))
    tk.Button(f21_r, text="重置", bg="#eaf1f6", font=("ABBvoice CNSG", int(12 * h_ratio)), image=quit_filter, compound=tk.LEFT, command=query_all, height=int(1 * h_ratio), fg='black', activebackground='blue').pack(side=tk.TOP, padx=int(20 * w_ratio), pady=(0, int(50 * h_ratio)))

    style = Style()
    style.configure('panel1.Treeview', rowheight=25, font=("ABBvoice CNSG", int(13 * h_ratio)))
    style.configure('panel1.Treeview.Heading', font=("ABBvoice CNSG", int(13 * h_ratio)), background="#EFF1F5")

    global Project_Info_table
    table_ybar = tk.Scrollbar(f31, orient="vertical")
    table_xbar = tk.Scrollbar(f31, orient="horizontal")

    Project_Info_table = Treeview(f31, show='headings', style='panel1.Treeview', selectmode='extended', columns=('id', 'a', 'b', 'c', 'd', 'd1', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n',
                                                                                                          'o', 'p', 'p1', 'p2', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 'a1', 'a2',
                                                                                                          'a3', 'a4', 'a5', 'a6', 'a7', 'a8', 'a9', 'a10', 'a11', 'a12', 'a13', 'a13a', 'a13b', 'a13c', 'a13d', 'a13e', 'a14', 'a15', 'a16', 'a17'),
                                  displaycolumns=('id', 'a', 'b', 'c', 'd', 'q', 'u', 'x', 'w', 'v', 'r', 's', 't', 'd1', 'e', 'h', 'i', 'y', 'z', 'f', 'g', 'k', 'j', 'l', 'm', 'n', 'o', 'p', 'p1', 'p2', 'a1', 'a2', 'a3', 'a4', 'a5', 'a6', 'a7',
                                                  'a8', 'a9', 'a10', 'a11', 'a12', 'a13', 'a13a', 'a13b', 'a13c', 'a13d', 'a13e', 'a14', 'a15', 'a16', 'a17'), yscrollcommand=table_ybar.set, xscrollcommand=table_xbar.set, height=int(32*h_ratio))
    table_ybar.pack(side=tk.RIGHT, fill=tk.Y)
    table_xbar.pack(side=tk.BOTTOM, fill=tk.X)

    table_ybar.config(command=Project_Info_table.yview)
    table_xbar.config(command=Project_Info_table.xview)
    Project_Info_table.column('id', width=int(1 * w_ratio), anchor='center')
    Project_Info_table.column('a', width=int(90 * w_ratio), anchor='center')
    Project_Info_table.column('b', width=int(100 * w_ratio), anchor='center')
    Project_Info_table.column('c', width=int(90 * w_ratio), anchor='center')
    Project_Info_table.column('d', width=int(480 * w_ratio), anchor='center')
    Project_Info_table.column('d1', width=int(50 * w_ratio), anchor='center')
    Project_Info_table.column('e', width=int(50 * w_ratio), anchor='center')
    Project_Info_table.column('f', width=int(150 * w_ratio), anchor='center')
    Project_Info_table.column('g', width=int(90 * w_ratio), anchor='center')
    Project_Info_table.column('h', width=int(90 * w_ratio), anchor='center')
    Project_Info_table.column('i', width=int(90 * w_ratio), anchor='center')
    Project_Info_table.column('j', width=int(300 * w_ratio), anchor='center')
    Project_Info_table.column('k', width=int(110 * w_ratio), anchor='center')
    Project_Info_table.column('l', width=int(300 * w_ratio), anchor='center')
    Project_Info_table.column('m', width=int(300 * w_ratio), anchor='center')
    Project_Info_table.column('n', width=int(150 * w_ratio), anchor='center')
    Project_Info_table.column('o', width=int(150 * w_ratio), anchor='center')
    Project_Info_table.column('p', width=int(100 * w_ratio), anchor='center')
    Project_Info_table.column('p1', width=int(100 * w_ratio), anchor='center')
    Project_Info_table.column('p2', width=int(100 * w_ratio), anchor='center')
    Project_Info_table.column('q', width=int(100 * w_ratio), anchor='center')
    Project_Info_table.column('r', width=int(100 * w_ratio), anchor='center')
    Project_Info_table.column('s', width=int(100 * w_ratio), anchor='center')
    Project_Info_table.column('t', width=int(100 * w_ratio), anchor='center')
    Project_Info_table.column('u', width=int(100 * w_ratio), anchor='center')
    Project_Info_table.column('v', width=int(100 * w_ratio), anchor='center')
    Project_Info_table.column('w', width=int(150 * w_ratio), anchor='center')
    Project_Info_table.column('x', width=int(100 * w_ratio), anchor='center')
    Project_Info_table.column('y', width=int(50 * w_ratio), anchor='center')
    Project_Info_table.column('z', width=int(150 * w_ratio), anchor='center')
    Project_Info_table.column('a1', width=int(200 * w_ratio), anchor='center')
    Project_Info_table.column('a2', width=int(150 * w_ratio), anchor='center')
    Project_Info_table.column('a3', width=int(150 * w_ratio), anchor='center')
    Project_Info_table.column('a4', width=int(250 * w_ratio), anchor='center')
    Project_Info_table.column('a5', width=int(150 * w_ratio), anchor='center')
    Project_Info_table.column('a6', width=int(250 * w_ratio), anchor='center')
    Project_Info_table.column('a7', width=int(350 * w_ratio), anchor='center')
    Project_Info_table.column('a8', width=int(350 * w_ratio), anchor='center')
    Project_Info_table.column('a9', width=int(90 * w_ratio), anchor='center')
    Project_Info_table.column('a10', width=int(90 * w_ratio), anchor='center')
    Project_Info_table.column('a11', width=int(90 * w_ratio), anchor='center')
    Project_Info_table.column('a12', width=int(90 * w_ratio), anchor='center')
    Project_Info_table.column('a13', width=int(90 * w_ratio), anchor='center')
    Project_Info_table.column('a13a', width=int(90 * w_ratio), anchor='center')
    Project_Info_table.column('a13b', width=int(90 * w_ratio), anchor='center')
    Project_Info_table.column('a13c', width=int(90 * w_ratio), anchor='center')
    Project_Info_table.column('a13d', width=int(90 * w_ratio), anchor='center')
    Project_Info_table.column('a13e', width=int(90 * w_ratio), anchor='center')
    Project_Info_table.column('a14', width=int(500 * w_ratio), anchor='center')
    Project_Info_table.column('a15', width=int(500 * w_ratio), anchor='center')
    Project_Info_table.column('a16', width=int(90 * w_ratio), anchor='center')
    Project_Info_table.column('a17', width=int(200 * w_ratio), anchor='center')
    Project_Info_table.heading('id', text='id')
    Project_Info_table.heading('a', text='设计类型')
    Project_Info_table.heading('b', text='项目号')
    Project_Info_table.heading('c', text='行号')
    Project_Info_table.heading('d', text='项目名')
    Project_Info_table.heading('d1', text='图套')
    Project_Info_table.heading('e', text='台数')
    Project_Info_table.heading('f', text='柜型')
    Project_Info_table.heading('g', text='产品类型')
    Project_Info_table.heading('h', text='工程师')
    Project_Info_table.heading('i', text='绘图员')
    Project_Info_table.heading('j', text='智能方案')
    Project_Info_table.heading('k', text='主母线电流')
    Project_Info_table.heading('l', text='主保护')
    Project_Info_table.heading('m', text='弧光方案')
    Project_Info_table.heading('n', text='地区')
    Project_Info_table.heading('o', text='图纸语言')
    Project_Info_table.heading('p', text='标准周期')
    Project_Info_table.heading('p1', text='额外周期')
    Project_Info_table.heading('p2', text='意见次数')
    Project_Info_table.heading('q', text='提原理图')
    Project_Info_table.heading('r', text='发放初级')
    Project_Info_table.heading('s', text='生产检查')
    Project_Info_table.heading('t', text='接收时间')
    Project_Info_table.heading('u', text='启动时间')
    Project_Info_table.heading('v', text='预计交付')
    Project_Info_table.heading('w', text='状态')
    Project_Info_table.heading('x', text='实际完成')
    Project_Info_table.heading('y', text='小组')
    Project_Info_table.heading('z', text='异常情况')
    Project_Info_table.heading('a1', text='创建时间')
    Project_Info_table.heading('a2', text='项目类型')
    Project_Info_table.heading('a3', text='框架类型')
    Project_Info_table.heading('a4', text='编程难度')
    Project_Info_table.heading('a5', text='语言难度')
    Project_Info_table.heading('a6', text='管理难度')
    Project_Info_table.heading('a7', text='输入文件')
    Project_Info_table.heading('a8', text='图纸需求')
    Project_Info_table.heading('a9', text='空开厂家')
    Project_Info_table.heading('a10', text='辅助节点')
    Project_Info_table.heading('a11', text='端子厂家')
    Project_Info_table.heading('a12', text='带显厂家')
    Project_Info_table.heading('a13', text='选择开关')
    Project_Info_table.heading('a13a', text='CT类型')
    Project_Info_table.heading('a13b', text='PT类型')
    Project_Info_table.heading('a13c', text='SA类型')
    Project_Info_table.heading('a13d', text='客户线号')
    Project_Info_table.heading('a13e', text='复杂客户需求')
    Project_Info_table.heading('a14', text='典型柜配置')
    Project_Info_table.heading('a15', text='备注')
    Project_Info_table.heading('a16', text='修改人')
    Project_Info_table.heading('a17', text='修改时间')
    Project_Info_table.tag_configure('fontsize', font=("ABBvoice CNSG", int(10 * h_ratio)))

    Project_Info_table.bind('<<TreeviewSelect>>', on_treeview_select)

    Project_Info_table.pack(side=tk.LEFT, fill=tk.Y)

    tk.Frame(f2, height=int(200 * h_ratio), bg="#eaf1f6").pack(fill=tk.BOTH, expand=True)

    # f3 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    # global text
    # text = tk.Text(f3, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), height=int(25 * h_ratio), width=int(65 * w_ratio))
    # text.pack(side=tk.LEFT, padx=(0, 1), pady=0, fill=tk.BOTH, expand=True)
    #
    # text.tag_configure('error', foreground='red')  # 设置tag
    #
    # scrollbar = tk.Scrollbar(f3)
    # scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    # scrollbar.config(command=text.yview)
    # text.config(yscrollcommand=scrollbar.set)
    # f3.pack(fill=tk.BOTH, expand=True)
    #
    # tk.Frame(parent, height=int(80 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

    query_all()

    # canvas.update_idletasks()
    # canvas.config(scrollregion=canvas.bbox('all'))


def about_help(event):
    # tku.show_info("说明书")
    os.startfile(os.path.abspath('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\二次设计辅助工具【EBOM导SAP功能】答疑V1.2.pdf'))


def on_mousewheel(event):
    global canvas
    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")


def on_edit_mousewheel(event):
    global edit_canvas
    edit_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")


def on_treeview_select(event):
    global selected_rows
    selected_rows = []
    selected_items = Project_Info_table.selection()

    for item in selected_items:
        row_values = Project_Info_table.item(item)['values']
        selected_rows.append(row_values)
    # print(selected_rows)

    global selected_design_type
    global selected_project_number
    selected_design_type = []
    selected_project_number = []

    # 详情按钮激活
    if len(selected_rows) == 1:
        button_detail['state'] = 'normal'

    elif len(selected_rows) > 1:
        for i in range(0, len(selected_rows)):
            selected_design_type.append(selected_rows[i][1])
            selected_project_number.append(selected_rows[i][2])

        if len(list(set(selected_design_type))) == 1 and (list(set(selected_design_type))[0] == '工程设计' or list(set(selected_design_type))[0] == '图纸设计'):
            if len(set(selected_project_number)) == 1:
                button_detail['state'] = 'normal'

            else:
                button_detail['state'] = 'disabled'
        else:
            button_detail['state'] = 'disabled'
    else:
        button_detail['state'] = 'disabled'

    if os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\Administrators.xlsx"):
        workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\Administrators.xlsx")
        worksheet = workbook['Sheet1']
        global username
        username = []
        for i in range(2, worksheet.max_row + 1):
            username.append(worksheet.cell(row=i, column=1).value)

        # 修改按钮激活（需要管理员权限才能激活）
        if os.getlogin() in username:
            if len(selected_rows) == 1 and selected_rows[0][1] == '提前设计':
                button_change_single['state'] = 'normal'
                button_ahead2other['state'] = 'normal'
                button_change_multi['state'] = 'disabled'

            elif len(selected_rows) == 1 and selected_rows[0][1] != '提前设计':
                button_change_single['state'] = 'normal'
                button_ahead2other['state'] = 'disabled'
                button_change_multi['state'] = 'disabled'

            elif len(selected_rows) > 1:
                for i in range(0, len(selected_rows)):
                    selected_design_type.append(selected_rows[i][1])
                    selected_project_number.append(selected_rows[i][2])

                button_change_single['state'] = 'disabled'
                button_ahead2other['state'] = 'disabled'

                if len(list(set(selected_design_type))) == 1 and (list(set(selected_design_type))[0] == '工程设计' or list(set(selected_design_type))[0] == '图纸设计'):
                    if len(set(selected_project_number)) == 1:
                        button_change_multi['state'] = 'normal'

                    else:
                        button_change_multi['state'] = 'disabled'
                else:
                    button_change_multi['state'] = 'disabled'
            else:
                button_change_single['state'] = 'disabled'
                button_change_multi['state'] = 'disabled'
                button_ahead2other['state'] = 'disabled'

        else:
            button_change_single['state'] = 'disabled'
            button_change_multi['state'] = 'disabled'
            button_ahead2other['state'] = 'disabled'


def project_detail(parent, w_ratio, h_ratio, selected_rows):
    try:
        if selected_rows[0][1] != '提前设计':
            # 自动选择额外的条目，确保所有条目都被选取
            conn = sqlite3.connect('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db')
            cursor = conn.cursor()  # 创建一个Cursor

            attributes = []

            attributes.append("design_type='" + str(selected_rows[0][1]) + "'")
            attributes.append("project_number='" + str(selected_rows[0][2]) + "'")

            attributes = [attr for attr in attributes if attr != '']

            query = "SELECT * FROM project_data WHERE "
            query += " AND ".join(attributes)
            query += "ORDER BY CASE WHEN design_type = '提前设计' THEN rowid END DESC, CASE WHEN design_type = '图纸设计' OR design_type = '工程设计' THEN project_number END DESC"

            cursor.execute(query)
            selected_rows_all = cursor.fetchall()

            selected_rows_all = [[element if element is not None else '' for element in row] for row in selected_rows_all]

            temp_list = []
            for i in range(0, len(selected_rows_all)):
                temp_list.append(list(selected_rows_all[i])[0:43] + list(selected_rows_all[i])[47:] + list(selected_rows_all[i])[43:47])

            cursor.close()
            conn.close()

            if len(selected_rows) < len(selected_rows_all):
                confirm_yn = tk.messagebox.askquestion("提示", "发现该项目号下有其他行号未选中，是否选择所有的行号？")
                if confirm_yn == 'yes':
                    selected_rows = temp_list

        edit_window = tk.Toplevel(parent)
        edit_window.grab_set()  # 禁用parent窗口的操作

        global edit_canvas

        edit_canvas = tk.Canvas(edit_window, width=int(1800 * w_ratio), height=int(740 * h_ratio), bg="#C9DBE9", borderwidth=0, highlightthickness=0)
        edit_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        edit_canvas.delete('all')  # 清空画布
        edit_canvas.update()
        edit_canvas.bind("<MouseWheel>", on_edit_mousewheel)

        scrollbar_v = tk.Scrollbar(master=edit_window)
        scrollbar_v.pack(side=tk.RIGHT, fill=tk.Y)
        scrollbar_v.config(command=edit_canvas.yview)
        edit_canvas.config(yscrollcommand=scrollbar_v.set)

        edit_content = tk.Frame(edit_canvas, bg="#eaf1f6")

        edit_canvas.create_window(0, 0, width=int(1700 * w_ratio), anchor=tk.NW, window=edit_content)

        f0 = tk.Frame(edit_content, bg="#c9dbe9", bd=0)
        tk.Label(f0, text="项目信息详情", bg="#c9dbe9", fg="black", height=int(1 * h_ratio), font=("ABBvoice CNSG", int(20 * h_ratio), "bold")).pack(fill=tk.X)
        tk.Label(f0, text="(本页编辑无效,仅供查看)", bg="#c9dbe9", fg="red", height=int(1 * h_ratio), font=("ABBvoice CNSG", int(12 * h_ratio), "bold")).pack(fill=tk.X)
        f0.pack(side=tk.TOP, fill=tk.X)

        f1 = tk.Frame(edit_content, bg="#eaf1f6")
        f1.pack(side=tk.TOP)

        f11 = tk.Frame(f1, bg="#eaf1f6")
        f11.pack(side=tk.LEFT, fill=tk.BOTH)

        tk.Frame(f1, width=int(1 * h_ratio), bg="black").pack(side=tk.LEFT, fill=tk.Y, padx=int(20*w_ratio))

        f12 = tk.Frame(f1, bg="#eaf1f6")
        f12.pack(side=tk.LEFT, fill=tk.BOTH)

        f2 = tk.Frame(edit_content, bg="#eaf1f6")
        f2.pack(side=tk.TOP)

        f3 = tk.Frame(edit_content, bg="#eaf1f6")
        f3.pack(side=tk.TOP)

        f4 = tk.Frame(edit_content, bg="#eaf1f6")
        f4.pack(side=tk.TOP)

        f5 = tk.Frame(edit_content, bg="#eaf1f6")
        f5.pack(side=tk.TOP)

        f6 = tk.Frame(edit_content, bg="#eaf1f6")
        f6.pack(side=tk.TOP)

        f7 = tk.Frame(edit_content, bg="#eaf1f6")
        f7.pack(side=tk.TOP)

        # 左半边
        selected_project_name_list = []
        selected_item_list = []
        selected_typical_amount_list = []
        selected_panel_amount_list = []
        selected_typical_type_list = []
        selected_product_type_list = []
        selected_PE_list = []
        selected_DE_list = []
        selected_main_busbar_current_list = []
        selected_main_protection_list = []
        selected_arclight_list = []
        selected_location_list = []
        selected_drawing_language_list = []
        selected_standard_period_list = []
        selected_extra_period_list = []
        selected_advice_times_list = []
        selected_mcb_company_list = []
        selected_aux_list = []
        selected_terminal_company_list = []
        selected_charged_display_company_list = []
        selected_switch_company_list = []
        selected_ct_type_list = []
        selected_pt_type_list = []
        selected_sa_type_list = []
        selected_customer_wiring_list = []
        selected_customer_requirements_list = []
        # 右半边
        selected_upload_time_list = []
        selected_bom_time_list = []
        selected_check_time_list = []
        selected_receive_time_list = []
        selected_start_time_list = []
        selected_estimated_time_list = []
        selected_status_info_list = []
        selected_actual_time_list = []
        selected_team_info_list = []
        selected_abnormal_status_list = []
        selected_create_time_list = []
        selected_project_type_list = []
        selected_frame_type_list = []
        selected_program_type_list = []
        selected_language_info_list = []
        selected_management_info_list = []

        for i in range(0, len(selected_rows)):
            selected_project_name_list.append(selected_rows[i][4])
            selected_item_list.append(selected_rows[i][3])
            selected_typical_amount_list.append(selected_rows[i][5])
            selected_panel_amount_list.append(selected_rows[i][6])
            selected_typical_type_list.append(selected_rows[i][7])
            selected_product_type_list.append(selected_rows[i][8])
            selected_PE_list.append(selected_rows[i][9])
            selected_DE_list.append(selected_rows[i][10])
            selected_main_busbar_current_list.append(selected_rows[i][12])
            selected_main_protection_list.append(selected_rows[i][13])
            selected_arclight_list.append(selected_rows[i][14])
            selected_location_list.append(selected_rows[i][15])
            selected_drawing_language_list.append(selected_rows[i][16])
            selected_standard_period_list.append(selected_rows[i][17])
            selected_extra_period_list.append(selected_rows[i][18])
            selected_advice_times_list.append(selected_rows[i][19])
            selected_mcb_company_list.append(selected_rows[i][38])
            selected_aux_list.append(selected_rows[i][39])
            selected_terminal_company_list.append(selected_rows[i][40])
            selected_charged_display_company_list.append(selected_rows[i][41])
            selected_switch_company_list.append(selected_rows[i][42])
            selected_ct_type_list.append(selected_rows[i][43])
            selected_pt_type_list.append(selected_rows[i][44])
            selected_sa_type_list.append(selected_rows[i][45])
            selected_customer_wiring_list.append(selected_rows[i][46])
            selected_customer_requirements_list.append(selected_rows[i][47])

            selected_upload_time_list.append(selected_rows[i][20])
            selected_bom_time_list.append(selected_rows[i][21])
            selected_check_time_list.append(selected_rows[i][22])
            selected_receive_time_list.append(selected_rows[i][23])
            selected_start_time_list.append(selected_rows[i][24])
            selected_estimated_time_list.append(selected_rows[i][25])
            selected_status_info_list.append(selected_rows[i][26])
            selected_actual_time_list.append(selected_rows[i][27])
            selected_team_info_list.append(selected_rows[i][28])
            selected_abnormal_status_list.append(selected_rows[i][29])
            selected_create_time_list.append(selected_rows[i][30])
            selected_project_type_list.append(selected_rows[i][31])
            selected_frame_type_list.append(selected_rows[i][32])
            selected_program_type_list.append(selected_rows[i][33])
            selected_language_info_list.append(selected_rows[i][34])
            selected_management_info_list.append(selected_rows[i][35])

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_1 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_1, text="设计类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_design_type = tk.Text(f_1, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
        edit_text_design_type.pack(side=tk.LEFT)
        edit_text_design_type.insert(tk.END, selected_rows[0][1])
        f_1.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_2 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_2, text="   项目号", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_project_number = tk.Text(f_2, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
        edit_text_project_number.pack(side=tk.LEFT)
        edit_text_project_number.insert(tk.END, selected_rows[0][2])
        f_2.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_3 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_3, text="   项目名", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_project_name = tk.Text(f_3, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100 * w_ratio))
        edit_text_project_name.pack(side=tk.LEFT)
        edit_text_project_name.insert(tk.END, '\n'.join(str(num) for num in selected_project_name_list))
        f_3.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_4 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_4, text="      行号", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_item = tk.Text(f_4, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_item.pack(side=tk.LEFT)
        edit_text_item.insert(tk.END, '\n'.join(str(num) for num in selected_item_list))
        f_4.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_5 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_5, text="      图套", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_typical_amount = tk.Text(f_5, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100 * w_ratio))
        edit_text_typical_amount.pack(side=tk.LEFT)
        edit_text_typical_amount.insert(tk.END, '\n'.join(str(num) for num in selected_typical_amount_list))
        f_5.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_6 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_6, text="      台数", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_panel_amount = tk.Text(f_6, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_panel_amount.pack(side=tk.LEFT)
        edit_text_panel_amount.insert(tk.END, '\n'.join(str(num) for num in selected_panel_amount_list))
        f_6.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_7 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_7, text="      柜型", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_typical_type = tk.Text(f_7, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_typical_type.pack(side=tk.LEFT)
        edit_text_typical_type.insert(tk.END, '\n'.join(str(num) for num in selected_typical_type_list))
        f_7.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_8 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_8, text="产品类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_product_type = tk.Text(f_8, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_product_type.pack(side=tk.LEFT)
        edit_text_product_type.insert(tk.END, '\n'.join(str(num) for num in selected_product_type_list))
        f_8.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_9 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_9, text="   工程师", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_PE = tk.Text(f_9, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_PE.pack(side=tk.LEFT)
        edit_text_PE.insert(tk.END, '\n'.join(str(num) for num in selected_PE_list))
        f_9.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_10 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_10, text="   绘图员", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_DE = tk.Text(f_10, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_DE.pack(side=tk.LEFT)
        edit_text_DE.insert(tk.END, '\n'.join(str(num) for num in selected_DE_list))
        f_10.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_11 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_11, text="母线电流", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_main_busbar_current = tk.Text(f_11, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_main_busbar_current.pack(side=tk.LEFT)
        edit_text_main_busbar_current.insert(tk.END, '\n'.join(str(num) for num in selected_main_busbar_current_list))
        f_11.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_12 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_12, text="   主保护", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_main_protection = tk.Text(f_12, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_main_protection.pack(side=tk.LEFT)
        edit_text_main_protection.insert(tk.END, '\n'.join(str(num) for num in selected_main_protection_list))
        f_12.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_13 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_13, text="弧光方案", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_arclight = tk.Text(f_13, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_arclight.pack(side=tk.LEFT)
        edit_text_arclight.insert(tk.END, '\n'.join(str(num) for num in selected_arclight_list))
        f_13.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_14 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_14, text="      地区", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_location = tk.Text(f_14, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_location.pack(side=tk.LEFT)
        edit_text_location.insert(tk.END, '\n'.join(str(num) for num in selected_location_list))
        f_14.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_15 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_15, text="图纸语言", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_drawing_language = tk.Text(f_15, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_drawing_language.pack(side=tk.LEFT)
        edit_text_drawing_language.insert(tk.END, '\n'.join(str(num) for num in selected_drawing_language_list))
        f_15.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_16 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_16, text="标准周期", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_standard_period = tk.Text(f_16, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_standard_period.pack(side=tk.LEFT)
        edit_text_standard_period.insert(tk.END, '\n'.join(str(num) for num in selected_standard_period_list))
        f_16.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_17 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_17, text="额外周期", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_extra_period = tk.Text(f_17, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_extra_period.pack(side=tk.LEFT)
        edit_text_extra_period.insert(tk.END, '\n'.join(str(num) for num in selected_extra_period_list))
        f_17.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_18 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_18, text="意见次数", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_advice_times = tk.Text(f_18, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_advice_times.pack(side=tk.LEFT)
        edit_text_advice_times.insert(tk.END, '\n'.join(str(num) for num in selected_advice_times_list))
        f_18.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_19 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_19, text="空开厂家", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_mcb_company = tk.Text(f_19, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_mcb_company.pack(side=tk.LEFT)
        edit_text_mcb_company.insert(tk.END, '\n'.join(str(num) for num in selected_mcb_company_list))
        f_19.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_20 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_20, text="辅助触点", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_aux = tk.Text(f_20, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_aux.pack(side=tk.LEFT)
        edit_text_aux.insert(tk.END, '\n'.join(str(num) for num in selected_aux_list))
        f_20.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_21 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_21, text="端子厂家", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_terminal_company = tk.Text(f_21, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_terminal_company.pack(side=tk.LEFT)
        edit_text_terminal_company.insert(tk.END, '\n'.join(str(num) for num in selected_terminal_company_list))
        f_21.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_22 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_22, text="带显厂家", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_charged_display_company = tk.Text(f_22, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_charged_display_company.pack(side=tk.LEFT)
        edit_text_charged_display_company.insert(tk.END, '\n'.join(str(num) for num in selected_charged_display_company_list))
        f_22.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_23 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_23, text="选择开关", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_switch_company = tk.Text(f_23, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_switch_company.pack(side=tk.LEFT)
        edit_text_switch_company.insert(tk.END, '\n'.join(str(num) for num in selected_switch_company_list))
        f_23.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_24 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_24, text="  CT类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_ct_type = tk.Text(f_24, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_ct_type.pack(side=tk.LEFT)
        edit_text_ct_type.insert(tk.END, '\n'.join(str(num) for num in selected_ct_type_list))
        f_24.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_25 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_25, text="  PT类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_pt_type = tk.Text(f_25, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_pt_type.pack(side=tk.LEFT)
        edit_text_pt_type.insert(tk.END, '\n'.join(str(num) for num in selected_pt_type_list))
        f_25.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_26 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_26, text="  SA类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_sa_type = tk.Text(f_26, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_sa_type.pack(side=tk.LEFT)
        edit_text_sa_type.insert(tk.END, '\n'.join(str(num) for num in selected_sa_type_list))
        f_26.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_27 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_27, text="客户线号", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_customer_wiring = tk.Text(f_27, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_customer_wiring.pack(side=tk.LEFT)
        edit_text_customer_wiring.insert(tk.END, '\n'.join(str(num) for num in selected_customer_wiring_list))
        f_27.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_28 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_28, text="客户需求", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_customer_requirements = tk.Text(f_28, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_customer_requirements.pack(side=tk.LEFT)
        edit_text_customer_requirements.insert(tk.END, '\n'.join(str(num) for num in selected_customer_requirements_list))
        f_28.pack(fill=tk.X)

        tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_1 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_1, text="提原理图", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_upload_time = tk.Text(f_1, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_upload_time.pack(side=tk.LEFT, padx=(0, int(5*w_ratio)))
        edit_text_upload_time.insert(tk.END, '\n'.join(str(num) for num in selected_upload_time_list))
        f_1.pack(fill=tk.X)

        tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_2 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_2, text="发放初级", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_bom_time = tk.Text(f_2, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_bom_time.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
        edit_text_bom_time.insert(tk.END, '\n'.join(str(num) for num in selected_bom_time_list))
        f_2.pack(fill=tk.X)

        tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_3 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_3, text="生产检查", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_check_time = tk.Text(f_3, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_check_time.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
        edit_text_check_time.insert(tk.END, '\n'.join(str(num) for num in selected_check_time_list))
        f_3.pack(fill=tk.X)

        tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_4 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_4, text="接收时间", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_receive_time = tk.Text(f_4, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_receive_time.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
        edit_text_receive_time.insert(tk.END, '\n'.join(str(num) for num in selected_receive_time_list))
        f_4.pack(fill=tk.X)

        tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_5 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_5, text="启动时间", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_start_time = tk.Text(f_5, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_start_time.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
        edit_text_start_time.insert(tk.END, '\n'.join(str(num) for num in selected_start_time_list))
        f_5.pack(fill=tk.X)

        tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_6 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_6, text="预计交付", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_estimated_time = tk.Text(f_6, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_estimated_time.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
        edit_text_estimated_time.insert(tk.END, '\n'.join(str(num) for num in selected_estimated_time_list))
        f_6.pack(fill=tk.X)

        tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_7 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_7, text="      状态", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_status_info = tk.Text(f_7, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_status_info.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
        edit_text_status_info.insert(tk.END, '\n'.join(str(num) for num in selected_status_info_list))
        f_7.pack(fill=tk.X)

        tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_8 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_8, text="实际完成", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_actual_time = tk.Text(f_8, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_actual_time.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
        edit_text_actual_time.insert(tk.END, '\n'.join(str(num) for num in selected_actual_time_list))
        f_8.pack(fill=tk.X)

        tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_9 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_9, text="      小组", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_team_info = tk.Text(f_9, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_team_info.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
        edit_text_team_info.insert(tk.END, '\n'.join(str(num) for num in selected_team_info_list))
        f_9.pack(fill=tk.X)

        tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_10 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_10, text="异常情况", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_abnormal_status = tk.Text(f_10, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_abnormal_status.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
        edit_text_abnormal_status.insert(tk.END, '\n'.join(str(num) for num in selected_abnormal_status_list))
        f_10.pack(fill=tk.X)

        tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_11 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_11, text="创建时间", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_create_time = tk.Text(f_11, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100*w_ratio))
        edit_text_create_time.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
        edit_text_create_time.insert(tk.END, '\n'.join(str(num) for num in selected_create_time_list))
        f_11.pack(fill=tk.X)

        tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_12 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_12, text="项目类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_project_type = tk.Text(f_12, bg="#eaf1f6", font=("ABBvoice CNSG", int(9 * h_ratio)), height=len(selected_rows), width=int(140*w_ratio))
        edit_text_project_type.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
        edit_text_project_type.insert(tk.END, '\n'.join(str(num) for num in selected_project_type_list))
        f_12.pack(fill=tk.X)

        tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_13 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_13, text="框架类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_frame_type = tk.Text(f_13, bg="#eaf1f6", font=("ABBvoice CNSG", int(9 * h_ratio)), height=len(selected_rows), width=int(140*w_ratio))
        edit_text_frame_type.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
        edit_text_frame_type.insert(tk.END, '\n'.join(str(num) for num in selected_frame_type_list))
        f_13.pack(fill=tk.X)

        tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_14 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_14, text="编程难度", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_program_type = tk.Text(f_14, bg="#eaf1f6", font=("ABBvoice CNSG", int(9 * h_ratio)), height=len(selected_rows), width=int(140*w_ratio))
        edit_text_program_type.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
        edit_text_program_type.insert(tk.END, '\n'.join(str(num) for num in selected_program_type_list))
        f_14.pack(fill=tk.X)

        tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_15 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_15, text="语言难度", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_language_info = tk.Text(f_15, bg="#eaf1f6", font=("ABBvoice CNSG", int(9 * h_ratio)), height=len(selected_rows), width=int(140*w_ratio))
        edit_text_language_info.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
        edit_text_language_info.insert(tk.END, '\n'.join(str(num) for num in selected_language_info_list))
        f_15.pack(fill=tk.X)

        tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_16 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_16, text="管理难度", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_text_management_info = tk.Text(f_16, bg="#eaf1f6", font=("ABBvoice CNSG", int(9 * h_ratio)), height=len(selected_rows), width=int(140*w_ratio))
        edit_text_management_info.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
        edit_text_management_info.insert(tk.END, '\n'.join(str(num) for num in selected_management_info_list))
        f_16.pack(fill=tk.X)

        tk.Frame(f2, width=int(200 * w_ratio), bg="black").pack(fill=tk.X, pady=int(10*h_ratio))

        tk.Label(f2, text='输入文件', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT)
        # input_files = ['Check List', '签字版单线图', '客户/设计院图纸', '参考项目信息(保护、产品一致，部分参考视为无参考)', '完全按照客户图纸设计']
        # checkbox_input_list = []
        # for i in range(len(input_files)):
        #     v = IntVar()    # 将各复选框绑定到变量v
        #     checkbox_input_file = tk.Checkbutton(f2, text=input_files[i], variable=v, font=("ABBvoice CNSG", int(13 * h_ratio)), height=1, background='#eaf1f6')
        #     checkbox_input_file.pack(side=tk.TOP, anchor=tk.W, expand=True)
        #     checkbox_input_list.append((v, checkbox_input_file))    # 将各复选框的variable存储到一个列表
        edit_text_input_file = tk.Text(f2, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(160 * w_ratio))
        edit_text_input_file.pack(side=tk.LEFT)
        edit_text_input_file.insert(tk.END, selected_rows[0][36])

        tk.Label(f3, text='图纸需求', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT)

        edit_text_drawing_requirement = tk.Text(f3, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(160 * w_ratio))
        edit_text_drawing_requirement.pack(side=tk.LEFT)
        edit_text_drawing_requirement.insert(tk.END, selected_rows[0][37])

        tk.Label(f4, text='智能方案', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT)

        edit_text_intelligence = tk.Text(f4, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(160 * w_ratio))
        edit_text_intelligence.pack(side=tk.LEFT)
        edit_text_intelligence.insert(tk.END, selected_rows[0][11])

        tk.Label(f5, text='典型柜配置', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.TOP)

        global num_columns1
        num_columns1 = 3
        global num_rows1
        num_rows1 = int(len(selected_rows[0][48].split(';'))/num_columns1)

        global entry_widths1
        entry_widths1 = [int(25 * w_ratio), int(75 * w_ratio), int(75 * w_ratio)]
        global entries1
        entries1 = [[None for _ in range(num_columns1)] for _ in range(num_rows1)]

        selfplus1 = 0
        for row in range(num_rows1):
            row_frame1 = tk.Frame(f5)
            row_frame1.pack(side=tk.TOP, fill=tk.X)

            for col in range(num_columns1):
                entry1 = tk.Entry(row_frame1, width=entry_widths1[col], relief='solid', font=("ABBvoice CNSG", int(11 * h_ratio)))
                entry1.pack(side=tk.LEFT)
                entries1[row][col] = entry1
                entry1.config(justify='center')
                if row == 0 and col == 0:
                    entry1.insert(0, '柜号/Typical')
                    entry1['state'] = 'disabled'
                elif row == 0 and col == 1:
                    entry1.insert(0, '保护料号/建号提单号')
                    entry1['state'] = 'disabled'
                elif row == 0 and col == 2:
                    entry1.insert(0, '多功能表、电度表料号/建号提单号')
                    entry1['state'] = 'disabled'
                else:
                    selfplus1 += 1
                    entry1.insert(0, selected_rows[0][48].split(';')[2+selfplus1])
                    entry1['state'] = 'disabled'

        tk.Label(f6, text='备注', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.TOP)

        global num_columns2
        num_columns2 = 2
        global num_rows2
        num_rows2 = int(len(selected_rows[0][49].split(';')) / num_columns2)

        global entry_widths2
        entry_widths2 = [int(6 * w_ratio), int(169 * w_ratio)]
        global entries2
        entries2 = [[None for _ in range(num_columns2)] for _ in range(num_rows2)]

        selfplus2 = 0
        for row in range(num_rows2):
            row_frame2 = tk.Frame(f6)
            row_frame2.pack(side=tk.TOP, fill=tk.X)

            for col in range(num_columns2):
                entry2 = tk.Entry(row_frame2, width=entry_widths2[col], relief='solid', font=("ABBvoice CNSG", int(11 * h_ratio)))
                entry2.pack(side=tk.LEFT)
                if col == 0:
                    entry2.config(justify='center')
                    entry2.insert(0, '%s' % str(row + 1))
                    entry2['state'] = 'disabled'
                else:
                    entry2.insert(0, selected_rows[0][49].split(';')[1 + selfplus2])
                    selfplus2 += 2
                    entry2['state'] = 'disabled'

                entries2[row][col] = entry2

        tk.Label(f7, text='   修改人', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT)
        edit_text_changer = tk.Text(f7, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(40 * w_ratio))
        edit_text_changer.pack(side=tk.LEFT)
        edit_text_changer.insert(tk.END, selected_rows[0][50])

        tk.Label(f7, text='          修改时间', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT)

        edit_text_change_time = tk.Text(f7, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(40 * w_ratio))
        edit_text_change_time.pack(side=tk.LEFT)
        edit_text_change_time.insert(tk.END, selected_rows[0][51])

        tk.Frame(f7, height=int(80 * h_ratio), bg="#eaf1f6").pack(fill=tk.BOTH, expand=True)

        edit_text_changer['state'] = 'disabled'
        edit_text_change_time['state'] = 'disabled'
        edit_text_design_type['state'] = 'disabled'
        edit_text_drawing_requirement['state'] = 'disabled'
        edit_text_input_file['state'] = 'disabled'
        edit_text_intelligence['state'] = 'disabled'
        edit_text_project_name['state'] = 'disabled'
        edit_text_project_number['state'] = 'disabled'
        edit_text_typical_amount['state'] = 'disabled'
        edit_text_abnormal_status['state'] = 'disabled'
        edit_text_actual_time['state'] = 'disabled'
        edit_text_advice_times['state'] = 'disabled'
        edit_text_arclight['state'] = 'disabled'
        edit_text_aux['state'] = 'disabled'
        edit_text_bom_time['state'] = 'disabled'
        edit_text_charged_display_company['state'] = 'disabled'
        edit_text_check_time['state'] = 'disabled'
        edit_text_DE['state'] = 'disabled'
        edit_text_drawing_language['state'] = 'disabled'
        edit_text_estimated_time['state'] = 'disabled'
        edit_text_extra_period['state'] = 'disabled'
        edit_text_frame_type['state'] = 'disabled'
        edit_text_item['state'] = 'disabled'
        edit_text_language_info['state'] = 'disabled'
        edit_text_location['state'] = 'disabled'
        edit_text_main_busbar_current['state'] = 'disabled'
        edit_text_main_protection['state'] = 'disabled'
        edit_text_management_info['state'] = 'disabled'
        edit_text_mcb_company['state'] = 'disabled'
        edit_text_create_time['state'] = 'disabled'
        edit_text_panel_amount['state'] = 'disabled'
        edit_text_PE['state'] = 'disabled'
        edit_text_product_type['state'] = 'disabled'
        edit_text_program_type['state'] = 'disabled'
        edit_text_project_type['state'] = 'disabled'
        edit_text_receive_time['state'] = 'disabled'
        edit_text_standard_period['state'] = 'disabled'
        edit_text_start_time['state'] = 'disabled'
        edit_text_status_info['state'] = 'disabled'
        edit_text_switch_company['state'] = 'disabled'
        edit_text_ct_type['state'] = 'disabled'
        edit_text_pt_type['state'] = 'disabled'
        edit_text_sa_type['state'] = 'disabled'
        edit_text_customer_wiring['state'] = 'disabled'
        edit_text_customer_requirements['state'] = 'disabled'
        edit_text_team_info['state'] = 'disabled'
        edit_text_terminal_company['state'] = 'disabled'
        edit_text_typical_type['state'] = 'disabled'
        edit_text_upload_time['state'] = 'disabled'

        edit_text_changer['background'] = '#eaf1f6'
        edit_text_change_time['background'] = '#eaf1f6'
        edit_text_design_type['background'] = '#eaf1f6'
        edit_text_drawing_requirement['background'] = '#eaf1f6'
        edit_text_input_file['background'] = '#eaf1f6'
        edit_text_intelligence['background'] = '#eaf1f6'
        edit_text_project_name['background'] = '#eaf1f6'
        edit_text_project_number['background'] = '#eaf1f6'
        edit_text_typical_amount['background'] = '#eaf1f6'
        edit_text_abnormal_status['background'] = '#eaf1f6'
        edit_text_actual_time['background'] = '#eaf1f6'
        edit_text_advice_times['background'] = '#eaf1f6'
        edit_text_arclight['background'] = '#eaf1f6'
        edit_text_aux['background'] = '#eaf1f6'
        edit_text_bom_time['background'] = '#eaf1f6'
        edit_text_charged_display_company['background'] = '#eaf1f6'
        edit_text_check_time['background'] = '#eaf1f6'
        edit_text_DE['background'] = '#eaf1f6'
        edit_text_drawing_language['background'] = '#eaf1f6'
        edit_text_estimated_time['background'] = '#eaf1f6'
        edit_text_extra_period['background'] = '#eaf1f6'
        edit_text_frame_type['background'] = '#eaf1f6'
        edit_text_item['background'] = '#eaf1f6'
        edit_text_language_info['background'] = '#eaf1f6'
        edit_text_location['background'] = '#eaf1f6'
        edit_text_main_busbar_current['background'] = '#eaf1f6'
        edit_text_main_protection['background'] = '#eaf1f6'
        edit_text_management_info['background'] = '#eaf1f6'
        edit_text_mcb_company['background'] = '#eaf1f6'
        edit_text_create_time['background'] = '#eaf1f6'
        edit_text_panel_amount['background'] = '#eaf1f6'
        edit_text_PE['background'] = '#eaf1f6'
        edit_text_product_type['background'] = '#eaf1f6'
        edit_text_program_type['background'] = '#eaf1f6'
        edit_text_project_type['background'] = '#eaf1f6'
        edit_text_receive_time['background'] = '#eaf1f6'
        edit_text_standard_period['background'] = '#eaf1f6'
        edit_text_start_time['background'] = '#eaf1f6'
        edit_text_status_info['background'] = '#eaf1f6'
        edit_text_switch_company['background'] = '#eaf1f6'
        edit_text_ct_type['background'] = '#eaf1f6'
        edit_text_pt_type['background'] = '#eaf1f6'
        edit_text_sa_type['background'] = '#eaf1f6'
        edit_text_customer_wiring['background'] = '#eaf1f6'
        edit_text_customer_requirements['background'] = '#eaf1f6'
        edit_text_team_info['background'] = '#eaf1f6'
        edit_text_terminal_company['background'] = '#eaf1f6'
        edit_text_typical_type['background'] = '#eaf1f6'
        edit_text_upload_time['background'] = '#eaf1f6'

        edit_canvas.update_idletasks()
        edit_canvas.config(scrollregion=edit_canvas.bbox('all'))
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def project_change_single(parent, w_ratio, h_ratio, selected_rows):
    def execute_program():
        edit_window = tk.Toplevel(parent)
        edit_window.grab_set()  # 禁用parent窗口的操作

        global edit_canvas

        edit_canvas = tk.Canvas(edit_window, width=int(1800 * w_ratio), height=int(740 * h_ratio), bg="#C9DBE9", borderwidth=0, highlightthickness=0)
        edit_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        edit_canvas.delete('all')  # 清空画布
        edit_canvas.update()
        edit_canvas.bind("<MouseWheel>", on_edit_mousewheel)

        scrollbar_v = tk.Scrollbar(master=edit_window)
        scrollbar_v.pack(side=tk.RIGHT, fill=tk.Y)
        scrollbar_v.config(command=edit_canvas.yview)
        edit_canvas.config(yscrollcommand=scrollbar_v.set)

        edit_content = tk.Frame(edit_canvas, bg="#eaf1f6")

        edit_canvas.create_window(0, 0, width=int(1700 * w_ratio), anchor=tk.NW, window=edit_content)

        global edit_text_design_type
        global edit_text_project_number
        global edit_text_project_name
        global edit_text_item
        global edit_text_typical_amount
        global edit_text_panel_amount
        global edit_text_typical_type
        global edit_combobox_product_type
        global edit_combobox_PE
        global edit_combobox_DE
        global edit_combobox_main_busbar_current
        global edit_text_main_protection
        global edit_text_arclight
        global edit_combobox_location
        global edit_combobox_drawing_language
        global edit_text_standard_period
        global edit_text_extra_period
        global edit_combobox_mcb_type
        global edit_combobox_aux_type
        global edit_combobox_terminal_type
        global edit_combobox_charged_display_type
        global edit_combobox_switch_type
        global edit_combobox_ct_type
        global edit_combobox_pt_type
        global edit_combobox_sa_type
        global edit_combobox_customer_wiring
        global edit_combobox_customer_requirements
        global edit_text_upload_time
        global edit_text_bom_time
        global edit_text_check_time
        global edit_text_receive_time
        global edit_text_start_time
        global edit_text_estimated_time
        global edit_combobox_status_info
        global edit_text_actual_time
        global edit_combobox_team_info
        global edit_text_abnormal_status
        global edit_text_create_time
        global edit_combobox_project_type
        global edit_combobox_frame_type
        global edit_combobox_program_type
        global edit_combobox_language_info
        global edit_combobox_management_info
        global checkbox_input_list
        global checkbox_drawing_type_list
        global checkbox_intelligence_list
        global edit_text_input_file
        global edit_text_drawing_type
        global edit_text_intelligence
        global num_columns1
        global num_rows1
        global entry_widths1
        global entries1
        global num_columns2
        global num_rows2
        global entry_widths2
        global entries2
        global edit_combobox_advice_times

        if selected_rows[0][1] != '提前设计':
            f0 = tk.Frame(edit_content, bg="#c9dbe9", bd=0)
            tk.Label(f0, text="项目信息修改——单行号", bg="#c9dbe9", fg="black", height=int(1 * h_ratio), font=("ABBvoice CNSG", int(20 * h_ratio), "bold")).pack(fill=tk.X)
            tk.Label(f0, text="(黑色属性无法修改，绿色属性可修改，红色属性为多CI共有属性，在修改(多行号)才可修改)", bg="#c9dbe9", fg="red", height=int(1 * h_ratio), font=("ABBvoice CNSG", int(12 * h_ratio), "bold")).pack(fill=tk.X)
            f0.pack(side=tk.TOP, fill=tk.X)

            f1 = tk.Frame(edit_content, bg="#eaf1f6")
            f1.pack(side=tk.TOP)

            f11 = tk.Frame(f1, bg="#eaf1f6")
            f11.pack(side=tk.LEFT, fill=tk.BOTH)

            tk.Frame(f1, width=int(1 * h_ratio), bg="black").pack(side=tk.LEFT, fill=tk.Y, padx=int(20 * w_ratio))

            f12 = tk.Frame(f1, bg="#eaf1f6")
            f12.pack(side=tk.LEFT, fill=tk.BOTH)

            f2 = tk.Frame(edit_content, bg="#eaf1f6")
            f2.pack(side=tk.TOP)

            f3 = tk.Frame(edit_content, bg="#eaf1f6")
            f3.pack(side=tk.TOP)

            f4 = tk.Frame(edit_content, bg="#eaf1f6")
            f4.pack(side=tk.TOP)

            f5 = tk.Frame(edit_content, bg="#eaf1f6")
            f5.pack(side=tk.TOP)

            f6 = tk.Frame(edit_content, bg="#eaf1f6")
            f6.pack(side=tk.TOP)

            f7 = tk.Frame(edit_content, bg="#eaf1f6")
            f7.pack(side=tk.TOP)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_1 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_1, text="设计类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_design_type = tk.Text(f_1, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_design_type.pack(side=tk.LEFT)
            edit_text_design_type.insert(tk.END, selected_rows[0][1])
            f_1.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_2 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_2, text="   项目号", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_project_number = tk.Text(f_2, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_project_number.pack(side=tk.LEFT)
            edit_text_project_number.insert(tk.END, selected_rows[0][2])
            f_2.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_3 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_3, text="   项目名", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_project_name = tk.Text(f_3, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_project_name.pack(side=tk.LEFT)
            edit_text_project_name.insert(tk.END, selected_rows[0][4])
            f_3.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_4 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_4, text="      行号", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_item = tk.Text(f_4, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_item.pack(side=tk.LEFT)
            edit_text_item.insert(tk.END, selected_rows[0][3])
            f_4.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_5 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_5, text="      图套", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_typical_amount = tk.Text(f_5, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_typical_amount.pack(side=tk.LEFT)
            edit_text_typical_amount.insert(tk.END, selected_rows[0][5])
            f_5.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_6 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_6, text="      台数", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_panel_amount = tk.Text(f_6, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_panel_amount.pack(side=tk.LEFT)
            edit_text_panel_amount.insert(tk.END, selected_rows[0][6])
            f_6.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_7 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_7, text="      柜型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_typical_type = tk.Text(f_7, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_typical_type.pack(side=tk.LEFT)
            edit_text_typical_type.insert(tk.END, selected_rows[0][7])
            f_7.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_8 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_8, text="产品类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_product_type_list = ['AIS', 'GIS']
            edit_combobox_product_type = ttk.Combobox(f_8, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_product_type_list)
            edit_combobox_product_type.pack(side=tk.LEFT)
            edit_combobox_product_type.set(selected_rows[0][8])
            f_8.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_9 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_9, text="   工程师", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_PE_list = []
            if os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\PE_DE.xlsx"):
                workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\PE_DE.xlsx")
                worksheet = workbook['Sheet1']
                for i in range(1, worksheet.max_row+1):
                    edit_PE_list.append(worksheet.cell(row=i, column=1).value)

            edit_combobox_PE = ttk.Combobox(f_9, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), values=edit_PE_list)
            edit_combobox_PE.pack(side=tk.LEFT)
            edit_combobox_PE.set(selected_rows[0][9])
            f_9.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_10 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_10, text="   绘图员", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_DE_list = []
            if os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\PE_DE.xlsx"):
                workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\PE_DE.xlsx")
                worksheet = workbook['Sheet1']
                for i in range(1, worksheet.max_row + 1):
                    edit_DE_list.append(worksheet.cell(row=i, column=1).value)

            edit_combobox_DE = ttk.Combobox(f_10, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), values=edit_DE_list)
            edit_combobox_DE.pack(side=tk.LEFT)
            edit_combobox_DE.set(selected_rows[0][10])
            f_10.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_11 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_11, text="母线电流", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_main_busbar_current_list = []
            if os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\Main_current.xlsx"):
                workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\Main_current.xlsx")
                worksheet = workbook['Sheet1']
                for i in range(1, worksheet.max_row + 1):
                    edit_main_busbar_current_list.append(worksheet.cell(row=i, column=1).value)

            edit_combobox_main_busbar_current = ttk.Combobox(f_11, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_main_busbar_current_list)
            edit_combobox_main_busbar_current.pack(side=tk.LEFT)
            edit_combobox_main_busbar_current.set(selected_rows[0][12])
            f_11.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_12 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_12, text="   主保护", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)

            edit_text_main_protection = tk.Text(f_12, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_main_protection.pack(side=tk.LEFT)
            edit_text_main_protection.insert(tk.END, selected_rows[0][13])
            f_12.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_13 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_13, text="弧光方案", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_arclight = tk.Text(f_13, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_arclight.pack(side=tk.LEFT)
            edit_text_arclight.insert(tk.END, selected_rows[0][14])
            f_13.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_14 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_14, text="      地区", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_location_list = ['UT万华', '港机', 'Building&数据中心', '工业/UT', '核电', '船用', '瑞典DOK80', '其他', '海外']
            edit_combobox_location = ttk.Combobox(f_14, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_location_list)
            edit_combobox_location.pack(side=tk.LEFT)
            edit_combobox_location.set(selected_rows[0][15])
            f_14.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_15 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_15, text="图纸语言", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_drawing_language_list = ['中文', '英文', '中英文']
            edit_combobox_drawing_language = ttk.Combobox(f_15, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_drawing_language_list)
            edit_combobox_drawing_language.pack(side=tk.LEFT)
            edit_combobox_drawing_language.set(selected_rows[0][16])
            f_15.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_16 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_16, text="标准周期", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_standard_period = tk.Text(f_16, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_standard_period.pack(side=tk.LEFT)
            edit_text_standard_period.insert(tk.END, selected_rows[0][17])
            f_16.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_17 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_17, text="额外周期", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_extra_period = tk.Text(f_17, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_extra_period.pack(side=tk.LEFT)
            edit_text_extra_period.insert(tk.END, selected_rows[0][18])
            f_17.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_18 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_18, text="意见次数", height=int(1 * h_ratio), bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_advice_times = tk.Text(f_18, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_advice_times.pack(side=tk.LEFT)
            edit_text_advice_times.insert(tk.END, selected_rows[0][19])
            f_18.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_19 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_19, text="空开厂家", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_mcb_type_list = ['人民电器', 'ABB', '熔丝+熔芯']
            edit_combobox_mcb_type = ttk.Combobox(f_19, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_mcb_type_list)
            edit_combobox_mcb_type.pack(side=tk.LEFT)
            edit_combobox_mcb_type.set(selected_rows[0][38])
            f_19.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_20 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_20, text="辅助触点", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_aux_type_list = ['是', '否']
            edit_combobox_aux_type = ttk.Combobox(f_20, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_aux_type_list)
            edit_combobox_aux_type.pack(side=tk.LEFT)
            edit_combobox_aux_type.set(selected_rows[0][39])
            f_20.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_21 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_21, text="端子厂家", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_terminal_type_list = ['瑞联', 'TE', '魏德米勒', '菲尼克斯(*)']
            edit_combobox_terminal_type = ttk.Combobox(f_21, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_terminal_type_list)
            edit_combobox_terminal_type.pack(side=tk.LEFT)
            edit_combobox_terminal_type.set(selected_rows[0][40])
            f_21.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_22 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_22, text="带显厂家", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_charged_display_type_list = ['立林', '百岗']
            edit_combobox_charged_display_type = ttk.Combobox(f_22, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_charged_display_type_list)
            edit_combobox_charged_display_type.pack(side=tk.LEFT)
            edit_combobox_charged_display_type.set(selected_rows[0][41])
            f_22.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_23 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_23, text="选择开关", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_switch_type_list = ['江阴长江', 'K&N', '其他(备注)']
            edit_combobox_switch_type = ttk.Combobox(f_23, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_switch_type_list)
            edit_combobox_switch_type.pack(side=tk.LEFT)
            edit_combobox_switch_type.set(selected_rows[0][42])
            f_23.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_24 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_24, text="  CT类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_ct_type_list = ['ABB', 'DYH', 'NTK', 'TLEP', '无']
            edit_combobox_ct_type = ttk.Combobox(f_24, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_ct_type_list)
            edit_combobox_ct_type.pack(side=tk.LEFT)
            edit_combobox_ct_type.set(selected_rows[0][43])
            f_24.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_25 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_25, text="  PT类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_pt_type_list = ['ABB', 'DYH', 'NTK', 'TLEP', '无']
            edit_combobox_pt_type = ttk.Combobox(f_25, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_pt_type_list)
            edit_combobox_pt_type.pack(side=tk.LEFT)
            edit_combobox_pt_type.set(selected_rows[0][44])
            f_25.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_26 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_26, text="  SA类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_sa_type_list = ['神电', '日立', 'GCA']
            edit_combobox_sa_type = ttk.Combobox(f_26, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_sa_type_list)
            edit_combobox_sa_type.pack(side=tk.LEFT)
            edit_combobox_sa_type.set(selected_rows[0][45])
            f_26.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_27 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_27, text="客户线号", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_customer_wiring_list = ['否', '是(备注前期项目号及需更改内容)']
            edit_combobox_customer_wiring = ttk.Combobox(f_27, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_customer_wiring_list)
            edit_combobox_customer_wiring.pack(side=tk.LEFT)
            edit_combobox_customer_wiring.set(selected_rows[0][46])
            f_27.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_28 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_28, text="客户需求", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_customer_requirements_list = ['否', '是(备注前期项目号及需更改内容)']
            edit_combobox_customer_requirements = ttk.Combobox(f_28, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_customer_requirements_list)
            edit_combobox_customer_requirements.pack(side=tk.LEFT)
            edit_combobox_customer_requirements.set(selected_rows[0][47])
            f_28.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_1 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_1, text="提原理图", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_upload_time = tk.Text(f_1, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(26 * w_ratio))
            edit_text_upload_time.pack(side=tk.LEFT)
            edit_text_upload_time.insert(tk.END, selected_rows[0][20])
            edit_text_upload_time['state'] = 'disabled'

            button_select_upload_time = tk.Button(master=f_1, bg="#eaf1f6", text='选择', command=lambda et=edit_text_upload_time: select_upload_time(edit_content, et), font=("ABBvoice CNSG", int(10 * h_ratio)), activebackground='blue')
            button_select_upload_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))
            f_1.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_2 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_2, text="发放初级", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_bom_time = tk.Text(f_2, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(26 * w_ratio))
            edit_text_bom_time.pack(side=tk.LEFT)
            edit_text_bom_time.insert(tk.END, selected_rows[0][21])
            edit_text_bom_time['state'] = 'disabled'

            button_select_bom_time = tk.Button(master=f_2, bg="#eaf1f6", text='选择', command=lambda et=edit_text_bom_time: select_bom_time(edit_content, et), font=("ABBvoice CNSG", int(10 * h_ratio)), activebackground='blue')
            button_select_bom_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))
            f_2.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_3 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_3, text="生产检查", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_check_time = tk.Text(f_3, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(26 * w_ratio))
            edit_text_check_time.pack(side=tk.LEFT)
            edit_text_check_time.insert(tk.END, selected_rows[0][22])
            edit_text_check_time['state'] = 'disabled'

            button_select_check_time = tk.Button(master=f_3, bg="#eaf1f6", text='选择', command=lambda et=edit_text_check_time: select_check_time(edit_content, et), font=("ABBvoice CNSG", int(10 * h_ratio)), activebackground='blue')
            button_select_check_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))
            f_3.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_4 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_4, text="接收时间", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_receive_time = tk.Text(f_4, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(26 * w_ratio))
            edit_text_receive_time.pack(side=tk.LEFT)
            edit_text_receive_time.insert(tk.END, selected_rows[0][23])
            edit_text_receive_time['state'] = 'disabled'

            button_select_receive_time = tk.Button(master=f_4, bg="#eaf1f6", text='选择', command=lambda et=edit_text_receive_time: select_receive_time(edit_content, et), font=("ABBvoice CNSG", int(10 * h_ratio)), activebackground='blue')
            button_select_receive_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))

            f_4.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_5 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_5, text="启动时间", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_start_time = tk.Text(f_5, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(26 * w_ratio))
            edit_text_start_time.pack(side=tk.LEFT)
            edit_text_start_time.insert(tk.END, selected_rows[0][24])
            edit_text_start_time['state'] = 'disabled'

            button_select_start_time = tk.Button(master=f_5, bg="#eaf1f6", text='选择', command=lambda et=edit_text_start_time: select_start_time(edit_content, et), font=("ABBvoice CNSG", int(10 * h_ratio)), activebackground='blue')
            button_select_start_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))
            f_5.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_6 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_6, text="预计交付", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_estimated_time = tk.Text(f_6, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(26 * w_ratio))
            edit_text_estimated_time.pack(side=tk.LEFT)
            edit_text_estimated_time.insert(tk.END, selected_rows[0][25])
            edit_text_estimated_time['state'] = 'disabled'

            button_select_estimated_time = tk.Button(master=f_6, bg="#eaf1f6", text='选择', command=lambda et=edit_text_estimated_time: select_estimated_time(edit_content, et), font=("ABBvoice CNSG", int(10 * h_ratio)), activebackground='blue')
            button_select_estimated_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))
            f_6.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_7 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_7, text="      状态", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_status_info_list = ['新项目', '待启动', '单线图及MVE完成', '初版图纸设计', '初版原理图提交', '生产图完成']
            edit_combobox_status_info = ttk.Combobox(f_7, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_status_info_list)
            edit_combobox_status_info.pack(side=tk.LEFT)
            edit_combobox_status_info.set(selected_rows[0][26])
            f_7.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_8 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_8, text="实际完成", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_actual_time = tk.Text(f_8, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(26 * w_ratio))
            edit_text_actual_time.pack(side=tk.LEFT)
            edit_text_actual_time.insert(tk.END, selected_rows[0][27])
            edit_text_actual_time['state'] = 'disabled'

            button_select_actual_time = tk.Button(master=f_8, bg="#eaf1f6", text='选择', command=lambda et=edit_text_actual_time: select_actual_time(edit_content, et), font=("ABBvoice CNSG", int(10 * h_ratio)), activebackground='blue')
            button_select_actual_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))
            f_8.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_9 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_9, text="      小组", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_team_info_list = ['A', 'B', 'C', 'D']
            edit_combobox_team_info = ttk.Combobox(f_9, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_team_info_list)
            edit_combobox_team_info.pack(side=tk.LEFT)
            edit_combobox_team_info.set(selected_rows[0][28])
            f_9.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_10 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_10, text="异常情况", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_abnormal_status = tk.Text(f_10, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_abnormal_status.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
            edit_text_abnormal_status.insert(tk.END, selected_rows[0][29])
            f_10.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_11 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_11, text="创建时间", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_create_time = tk.Text(f_11, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_create_time.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
            edit_text_create_time.insert(tk.END, selected_rows[0][30])
            f_11.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_12 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_12, text="项目类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_project_type_list = ['Domestic', 'Export_fully built', 'EPC', 'Metro', 'Marine', 'Wind power', '供电局项目(含18相反措)']
            edit_combobox_project_type = ttk.Combobox(f_12, font=("ABBvoice CNSG", int(9 * h_ratio)), width=int(130 * w_ratio), state='readonly', values=edit_project_type_list)
            edit_combobox_project_type.pack(side=tk.LEFT)
            edit_combobox_project_type.set(selected_rows[0][31])
            f_12.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_13 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_13, text="框架类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_frame_type_list = ['非框架项目', '重复或增补项目', '框架(阿里、中国移动、浙石化、泉化、万华、雄安、许继或大盛ATS Metro、上海小区变、港机等)']
            edit_combobox_frame_type = ttk.Combobox(f_13, font=("ABBvoice CNSG", int(9 * h_ratio)), width=int(130 * w_ratio), state='readonly', values=edit_frame_type_list)
            edit_combobox_frame_type.pack(side=tk.LEFT)
            edit_combobox_frame_type.set(selected_rows[0][32])
            f_13.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_14 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_14, text="编程难度", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_program_type_list = ['不含保护、客供(ABB/非ABB)保护、repeat程序', 'ABB常用保护的国内/出口编程', 'ABB非常用保护编程', 'ABB保护编程+参与车间联调']
            edit_combobox_program_type = ttk.Combobox(f_14, font=("ABBvoice CNSG", int(9 * h_ratio)), width=int(130 * w_ratio), state='readonly', values=edit_program_type_list)
            edit_combobox_program_type.pack(side=tk.LEFT)
            edit_combobox_program_type.set(selected_rows[0][33])
            f_14.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_15 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_15, text="语言难度", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_language_info_list = ['完全中文', '一般英语书面沟通', '英语书面+口语沟通', '复杂英语专业术语沟通']
            edit_combobox_language_info = ttk.Combobox(f_15, font=("ABBvoice CNSG", int(9 * h_ratio)), width=int(130 * w_ratio), state='readonly', values=edit_language_info_list)
            edit_combobox_language_info.pack(side=tk.LEFT)
            edit_combobox_language_info.set(selected_rows[0][34])
            f_15.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_16 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_16, text="管理难度", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_management_info_list = ['标准项目', '同一项目号下多种产品、首次生产新柜型项目', '复杂文档要求(客户要求模板)', '需要读复杂Spec文件、FAT要求复杂、客户特殊复杂的技术规格要求']
            edit_combobox_management_info = ttk.Combobox(f_16, font=("ABBvoice CNSG", int(9 * h_ratio)), width=int(130 * w_ratio), state='readonly', values=edit_management_info_list)
            edit_combobox_management_info.pack(side=tk.LEFT)
            edit_combobox_management_info.set(selected_rows[0][35])
            f_16.pack(fill=tk.X)

            tk.Frame(f2, width=int(200 * w_ratio), bg="black").pack(fill=tk.X, pady=int(10 * h_ratio))

            tk.Label(f2, text='输入文件', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT)
            edit_text_input_file = tk.Text(f2, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(160 * w_ratio))
            edit_text_input_file.pack(side=tk.LEFT)
            edit_text_input_file.insert(tk.END, selected_rows[0][36])

            tk.Label(f3, text='图纸需求', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT)
            edit_text_drawing_type = tk.Text(f3, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(160 * w_ratio))
            edit_text_drawing_type.pack(side=tk.LEFT)
            edit_text_drawing_type.insert(tk.END, selected_rows[0][37])

            tk.Label(f4, text='智能方案', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT)
            edit_text_intelligence = tk.Text(f4, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(160 * w_ratio))
            edit_text_intelligence.pack(side=tk.LEFT)
            edit_text_intelligence.insert(tk.END, selected_rows[0][11])

            tk.Label(f5, text='典型柜配置', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.TOP)

            num_columns1 = 3
            num_rows1 = int(len(selected_rows[0][48].split(';')) / num_columns1)

            entry_widths1 = [int(25 * w_ratio), int(75 * w_ratio), int(75 * w_ratio)]
            entries1 = [[None for _ in range(num_columns1)] for _ in range(num_rows1)]

            selfplus1 = 0
            for row in range(num_rows1):
                row_frame1 = tk.Frame(f5)
                row_frame1.pack(side=tk.TOP, fill=tk.X)

                for col in range(num_columns1):
                    entry1 = tk.Entry(row_frame1, width=entry_widths1[col], relief='solid', font=("ABBvoice CNSG", int(11 * h_ratio)))
                    entry1.pack(side=tk.LEFT)
                    entries1[row][col] = entry1
                    entry1.config(justify='center')
                    if row == 0 and col == 0:
                        entry1.insert(0, '柜号/Typical')
                        entry1['state'] = 'disabled'
                    elif row == 0 and col == 1:
                        entry1.insert(0, '保护料号/建号提单号')
                        entry1['state'] = 'disabled'
                    elif row == 0 and col == 2:
                        entry1.insert(0, '多功能表、电度表料号/建号提单号')
                        entry1['state'] = 'disabled'
                    else:
                        selfplus1 += 1
                        entry1.insert(0, selected_rows[0][48].split(';')[2 + selfplus1])
                        entry1['state'] = 'disabled'

            tk.Label(f6, text='备注', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.TOP)

            num_columns2 = 2
            num_rows2 = int(len(selected_rows[0][49].split(';')) / num_columns2)

            entry_widths2 = [int(6 * w_ratio), int(169 * w_ratio)]
            entries2 = [[None for _ in range(num_columns2)] for _ in range(num_rows2)]

            selfplus2 = 0
            for row in range(num_rows2):
                row_frame2 = tk.Frame(f6)
                row_frame2.pack(side=tk.TOP, fill=tk.X)

                for col in range(num_columns2):
                    entry2 = tk.Entry(row_frame2, width=entry_widths2[col], relief='solid', font=("ABBvoice CNSG", int(11 * h_ratio)))
                    entry2.pack(side=tk.LEFT)
                    if col == 0:
                        entry2.config(justify='center')
                        entry2.insert(0, '%s' % str(row + 1))
                        entry2['state'] = 'disabled'
                    else:
                        entry2.insert(0, selected_rows[0][49].split(';')[1 + selfplus2])
                        selfplus2 += 2
                        entry2['state'] = 'disabled'

                    entries2[row][col] = entry2

            button_confirm_change_single = tk.Button(master=f7, bg="#eaf1f6", text='确认修改', command=lambda: confirm_change_single(edit_window, selected_rows), font=("ABBvoice CNSG", int(11 * h_ratio)), activebackground='blue')
            button_confirm_change_single.pack(side=tk.TOP, pady=int(20*h_ratio))

            tk.Frame(f7, height=int(80 * h_ratio), bg="#eaf1f6").pack(side=tk.TOP, fill=tk.BOTH, expand=True)

            edit_text_design_type['state'] = 'disabled'
            edit_text_project_number['state'] = 'disabled'
            # edit_text_bom_time['state'] = 'disabled'
            # edit_text_check_time['state'] = 'disabled'
            edit_text_extra_period['state'] = 'disabled'
            edit_text_item['state'] = 'disabled'
            edit_text_standard_period['state'] = 'disabled'
            # edit_text_upload_time['state'] = 'disabled'
            edit_text_drawing_type['state'] = 'disabled'
            edit_text_input_file['state'] = 'disabled'
            edit_text_intelligence['state'] = 'disabled'
            edit_text_create_time['state'] = 'disabled'
            edit_text_advice_times['state'] = 'disabled'

            edit_text_design_type['background'] = '#eaf1f6'
            edit_text_project_number['background'] = '#eaf1f6'
            # edit_text_bom_time['background'] = '#eaf1f6'
            # edit_text_check_time['background'] = '#eaf1f6'
            edit_text_extra_period['background'] = '#eaf1f6'
            edit_text_item['background'] = '#eaf1f6'
            edit_text_standard_period['background'] = '#eaf1f6'
            # edit_text_upload_time['background'] = '#eaf1f6'
            edit_text_drawing_type['background'] = '#eaf1f6'
            edit_text_input_file['background'] = '#eaf1f6'
            edit_text_intelligence['background'] = '#eaf1f6'
            edit_text_create_time['background'] = '#eaf1f6'
            edit_text_advice_times['background'] = '#eaf1f6'

            edit_canvas.update_idletasks()
            edit_canvas.config(scrollregion=edit_canvas.bbox('all'))

        else:
            f0 = tk.Frame(edit_content, bg="#c9dbe9", bd=0)
            tk.Label(f0, text="项目信息修改——单行号", bg="#c9dbe9", fg="black", height=int(1 * h_ratio), font=("ABBvoice CNSG", int(20 * h_ratio), "bold")).pack(fill=tk.X)
            tk.Label(f0, text="(黑色属性无法修改，绿色属性可修改)", bg="#c9dbe9", fg="red", height=int(1 * h_ratio), font=("ABBvoice CNSG", int(12 * h_ratio), "bold")).pack(fill=tk.X)
            f0.pack(side=tk.TOP, fill=tk.X)

            f1 = tk.Frame(edit_content, bg="#eaf1f6")
            f1.pack(side=tk.TOP)

            f11 = tk.Frame(f1, bg="#eaf1f6")
            f11.pack(side=tk.LEFT, fill=tk.BOTH)

            tk.Frame(f1, width=int(1 * h_ratio), bg="black").pack(side=tk.LEFT, fill=tk.Y, padx=int(20 * w_ratio))

            f12 = tk.Frame(f1, bg="#eaf1f6")
            f12.pack(side=tk.LEFT, fill=tk.BOTH)

            f2 = tk.Frame(edit_content, bg="#eaf1f6")
            f2.pack(side=tk.TOP)

            f3 = tk.Frame(edit_content, bg="#eaf1f6")
            f3.pack(side=tk.TOP)

            f4 = tk.Frame(edit_content, bg="#eaf1f6")
            f4.pack(side=tk.TOP)

            f5 = tk.Frame(edit_content, bg="#eaf1f6")
            f5.pack(side=tk.TOP)

            f6 = tk.Frame(edit_content, bg="#eaf1f6")
            f6.pack(side=tk.TOP)

            f7 = tk.Frame(edit_content, bg="#eaf1f6")
            f7.pack(side=tk.TOP)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_1 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_1, text="设计类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_design_type = tk.Text(f_1, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_design_type.pack(side=tk.LEFT)
            edit_text_design_type.insert(tk.END, selected_rows[0][1])
            f_1.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_2 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_2, text="   项目号", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_project_number = tk.Text(f_2, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_project_number.pack(side=tk.LEFT)
            edit_text_project_number.insert(tk.END, selected_rows[0][2])
            f_2.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_3 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_3, text="   项目名", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_project_name = tk.Text(f_3, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_project_name.pack(side=tk.LEFT)
            edit_text_project_name.insert(tk.END, selected_rows[0][4])
            f_3.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_4 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_4, text="      行号", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_item = tk.Text(f_4, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_item.pack(side=tk.LEFT)
            edit_text_item.insert(tk.END, selected_rows[0][3])
            f_4.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_5 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_5, text="      图套", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_typical_amount = tk.Text(f_5, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_typical_amount.pack(side=tk.LEFT)
            edit_text_typical_amount.insert(tk.END, selected_rows[0][5])
            f_5.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_6 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_6, text="      台数", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_panel_amount = tk.Text(f_6, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_panel_amount.pack(side=tk.LEFT)
            edit_text_panel_amount.insert(tk.END, selected_rows[0][6])
            f_6.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_7 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_7, text="      柜型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_typical_type = tk.Text(f_7, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_typical_type.pack(side=tk.LEFT)
            edit_text_typical_type.insert(tk.END, selected_rows[0][7])
            f_7.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_8 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_8, text="产品类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_product_type_list = ['AIS', 'GIS']
            edit_combobox_product_type = ttk.Combobox(f_8, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_product_type_list)
            edit_combobox_product_type.pack(side=tk.LEFT)
            edit_combobox_product_type.set(selected_rows[0][8])
            f_8.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_9 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_9, text="   工程师", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_PE_list = []
            if os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\PE_DE.xlsx"):
                workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\PE_DE.xlsx")
                worksheet = workbook['Sheet1']
                for i in range(1, worksheet.max_row + 1):
                    edit_PE_list.append(worksheet.cell(row=i, column=1).value)
            edit_combobox_PE = ttk.Combobox(f_9, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), values=edit_PE_list)
            edit_combobox_PE.pack(side=tk.LEFT)
            edit_combobox_PE.set(selected_rows[0][9])
            f_9.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_10 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_10, text="   绘图员", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_DE_list = []
            if os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\PE_DE.xlsx"):
                workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\PE_DE.xlsx")
                worksheet = workbook['Sheet1']
                for i in range(1, worksheet.max_row + 1):
                    edit_DE_list.append(worksheet.cell(row=i, column=1).value)
            edit_combobox_DE = ttk.Combobox(f_10, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), values=edit_DE_list)
            edit_combobox_DE.pack(side=tk.LEFT)
            edit_combobox_DE.set(selected_rows[0][10])
            f_10.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_11 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_11, text="母线电流", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_main_busbar_current_list = []
            if os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\Main_current.xlsx"):
                workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\Main_current.xlsx")
                worksheet = workbook['Sheet1']
                for i in range(1, worksheet.max_row + 1):
                    edit_main_busbar_current_list.append(worksheet.cell(row=i, column=1).value)
            edit_combobox_main_busbar_current = ttk.Combobox(f_11, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_main_busbar_current_list)
            edit_combobox_main_busbar_current.pack(side=tk.LEFT)
            edit_combobox_main_busbar_current.set(selected_rows[0][12])
            f_11.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_12 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_12, text="   主保护", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_main_protection = tk.Text(f_12, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_main_protection.pack(side=tk.LEFT)
            edit_text_main_protection.insert(tk.END, selected_rows[0][13])
            f_12.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_13 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_13, text="弧光方案", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_arclight = tk.Text(f_13, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_arclight.pack(side=tk.LEFT)
            edit_text_arclight.insert(tk.END, selected_rows[0][14])
            f_13.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_14 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_14, text="      地区", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_location_list = ['UT万华', '港机', 'Building&数据中心', '工业/UT', '核电', '船用', '瑞典DOK80', '其他', '海外']
            edit_combobox_location = ttk.Combobox(f_14, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_location_list)
            edit_combobox_location.pack(side=tk.LEFT)
            edit_combobox_location.set(selected_rows[0][15])
            f_14.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_15 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_15, text="图纸语言", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_drawing_language_list = ['中文', '英文', '中英文']
            edit_combobox_drawing_language = ttk.Combobox(f_15, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_drawing_language_list)
            edit_combobox_drawing_language.pack(side=tk.LEFT)
            edit_combobox_drawing_language.set(selected_rows[0][16])
            f_15.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_16 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_16, text="标准周期", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_standard_period = tk.Text(f_16, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_standard_period.pack(side=tk.LEFT)
            edit_text_standard_period.insert(tk.END, selected_rows[0][17])
            f_16.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_17 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_17, text="额外周期", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_extra_period = tk.Text(f_17, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_extra_period.pack(side=tk.LEFT)
            edit_text_extra_period.insert(tk.END, selected_rows[0][18])
            f_17.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_18 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_18, text="意见次数", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_advice_times_list = ['1', '2', '3', '4']
            edit_combobox_advice_times = ttk.Combobox(f_18, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_advice_times_list)
            edit_combobox_advice_times.pack(side=tk.LEFT)
            edit_combobox_advice_times.set(selected_rows[0][19])
            f_18.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_19 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_19, text="空开厂家", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_mcb_type_list = ['人民电器', 'ABB', '熔丝+熔芯']
            edit_combobox_mcb_type = ttk.Combobox(f_19, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_mcb_type_list)
            edit_combobox_mcb_type.pack(side=tk.LEFT)
            edit_combobox_mcb_type.set(selected_rows[0][38])
            f_19.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_20 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_20, text="辅助触点", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_aux_type_list = ['是', '否']
            edit_combobox_aux_type = ttk.Combobox(f_20, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_aux_type_list)
            edit_combobox_aux_type.pack(side=tk.LEFT)
            edit_combobox_aux_type.set(selected_rows[0][39])
            f_20.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_21 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_21, text="端子厂家", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_terminal_type_list = ['瑞联', 'TE', '魏德米勒', '菲尼克斯(*)']
            edit_combobox_terminal_type = ttk.Combobox(f_21, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_terminal_type_list)
            edit_combobox_terminal_type.pack(side=tk.LEFT)
            edit_combobox_terminal_type.set(selected_rows[0][40])
            f_21.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_22 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_22, text="带显厂家", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_charged_display_type_list = ['立林', '百岗']
            edit_combobox_charged_display_type = ttk.Combobox(f_22, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_charged_display_type_list)
            edit_combobox_charged_display_type.pack(side=tk.LEFT)
            edit_combobox_charged_display_type.set(selected_rows[0][41])
            f_22.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_23 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_23, text="选择开关", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_switch_type_list = ['江阴长江', 'K&N', '其他(备注)']
            edit_combobox_switch_type = ttk.Combobox(f_23, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_switch_type_list)
            edit_combobox_switch_type.pack(side=tk.LEFT)
            edit_combobox_switch_type.set(selected_rows[0][42])
            f_23.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_24 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_24, text="  CT类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_ct_type_list = ['ABB', 'DYH', 'NTK', 'TLEP', '无']
            edit_combobox_ct_type = ttk.Combobox(f_24, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_ct_type_list)
            edit_combobox_ct_type.pack(side=tk.LEFT)
            edit_combobox_ct_type.set(selected_rows[0][43])
            f_24.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_25 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_25, text="  PT类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_pt_type_list = ['ABB', 'DYH', 'NTK', 'TLEP', '无']
            edit_combobox_pt_type = ttk.Combobox(f_25, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_pt_type_list)
            edit_combobox_pt_type.pack(side=tk.LEFT)
            edit_combobox_pt_type.set(selected_rows[0][44])
            f_25.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_26 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_26, text="  SA类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_sa_type_list = ['神电', '日立', 'GCA']
            edit_combobox_sa_type = ttk.Combobox(f_26, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_sa_type_list)
            edit_combobox_sa_type.pack(side=tk.LEFT)
            edit_combobox_sa_type.set(selected_rows[0][45])
            f_26.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_27 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_27, text="客户线号", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_customer_wiring_list = ['否', '是(备注前期项目号及需更改内容)']
            edit_combobox_customer_wiring = ttk.Combobox(f_27, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_customer_wiring_list)
            edit_combobox_customer_wiring.pack(side=tk.LEFT)
            edit_combobox_customer_wiring.set(selected_rows[0][46])
            f_27.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_28 = tk.Frame(f11, bg="#eaf1f6")
            tk.Label(f_28, text="客户需求", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_customer_requirements_list = ['否', '是(备注前期项目号及需更改内容)']
            edit_combobox_customer_requirements = ttk.Combobox(f_28, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_customer_requirements_list)
            edit_combobox_customer_requirements.pack(side=tk.LEFT)
            edit_combobox_customer_requirements.set(selected_rows[0][47])
            f_28.pack(fill=tk.X)

            tk.Frame(f11, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_1 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_1, text="提原理图", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_upload_time = tk.Text(f_1, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(26 * w_ratio))
            edit_text_upload_time.pack(side=tk.LEFT)
            edit_text_upload_time.insert(tk.END, selected_rows[0][20])
            edit_text_upload_time['state'] = 'disabled'

            button_select_upload_time = tk.Button(master=f_1, bg="#eaf1f6", text='选择', command=lambda et=edit_text_upload_time: select_upload_time(edit_content, et), font=("ABBvoice CNSG", int(10 * h_ratio)), activebackground='blue')
            button_select_upload_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))
            f_1.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_2 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_2, text="发放初级", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_bom_time = tk.Text(f_2, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_bom_time.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
            edit_text_bom_time.insert(tk.END, selected_rows[0][21])
            f_2.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_3 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_3, text="生产检查", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_check_time = tk.Text(f_3, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_check_time.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
            edit_text_check_time.insert(tk.END, selected_rows[0][22])
            f_3.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_4 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_4, text="接收时间", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_receive_time = tk.Text(f_4, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(26 * w_ratio))
            edit_text_receive_time.pack(side=tk.LEFT)
            edit_text_receive_time.insert(tk.END, selected_rows[0][23])
            edit_text_receive_time['state'] = 'disabled'

            button_select_receive_time = tk.Button(master=f_4, bg="#eaf1f6", text='选择', command=lambda et=edit_text_receive_time: select_receive_time(edit_content, et), font=("ABBvoice CNSG", int(10 * h_ratio)), activebackground='blue')
            button_select_receive_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))

            f_4.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_5 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_5, text="启动时间", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_start_time = tk.Text(f_5, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(26 * w_ratio))
            edit_text_start_time.pack(side=tk.LEFT)
            edit_text_start_time.insert(tk.END, selected_rows[0][24])
            edit_text_start_time['state'] = 'disabled'

            button_select_start_time = tk.Button(master=f_5, bg="#eaf1f6", text='选择', command=lambda et=edit_text_start_time: select_start_time(edit_content, et), font=("ABBvoice CNSG", int(10 * h_ratio)), activebackground='blue')
            button_select_start_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))
            f_5.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_6 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_6, text="预计交付", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_estimated_time = tk.Text(f_6, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(26 * w_ratio))
            edit_text_estimated_time.pack(side=tk.LEFT)
            edit_text_estimated_time.insert(tk.END, selected_rows[0][25])
            edit_text_estimated_time['state'] = 'disabled'

            button_select_estimated_time = tk.Button(master=f_6, bg="#eaf1f6", text='选择', command=lambda et=edit_text_estimated_time: select_estimated_time(edit_content, et), font=("ABBvoice CNSG", int(10 * h_ratio)), activebackground='blue')
            button_select_estimated_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))
            f_6.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_7 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_7, text="      状态", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_status_info_list = ['新项目', '待启动', '单线图及MVE完成', '初版图纸设计', '初版原理图提交', '生产图完成']
            edit_combobox_status_info = ttk.Combobox(f_7, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_status_info_list)
            edit_combobox_status_info.pack(side=tk.LEFT)
            edit_combobox_status_info.set(selected_rows[0][26])
            f_7.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_8 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_8, text="实际完成", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_actual_time = tk.Text(f_8, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(26 * w_ratio))
            edit_text_actual_time.pack(side=tk.LEFT)
            edit_text_actual_time.insert(tk.END, selected_rows[0][27])
            edit_text_actual_time['state'] = 'disabled'

            button_select_actual_time = tk.Button(master=f_8, bg="#eaf1f6", text='选择', command=lambda et=edit_text_actual_time: select_actual_time(edit_content, et), font=("ABBvoice CNSG", int(10 * h_ratio)), activebackground='blue')
            button_select_actual_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))
            f_8.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_9 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_9, text="      小组", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_team_info_list = ['A', 'B', 'C', 'D']
            edit_combobox_team_info = ttk.Combobox(f_9, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_team_info_list)
            edit_combobox_team_info.pack(side=tk.LEFT)
            edit_combobox_team_info.set(selected_rows[0][28])
            f_9.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_10 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_10, text="异常情况", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_abnormal_status = tk.Text(f_10, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_abnormal_status.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
            edit_text_abnormal_status.insert(tk.END, selected_rows[0][29])
            f_10.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_11 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_11, text="创建时间", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_text_create_time = tk.Text(f_11, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
            edit_text_create_time.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
            edit_text_create_time.insert(tk.END, selected_rows[0][30])
            f_11.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_12 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_12, text="项目类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_project_type_list = ['Domestic', 'Export_fully built', 'EPC', 'Metro', 'Marine', 'Wind power', '供电局项目(含18相反措)']
            edit_combobox_project_type = ttk.Combobox(f_12, font=("ABBvoice CNSG", int(9 * h_ratio)), width=int(130 * w_ratio), state='readonly', values=edit_project_type_list)
            edit_combobox_project_type.pack(side=tk.LEFT)
            edit_combobox_project_type.set(selected_rows[0][31])
            f_12.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_13 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_13, text="框架类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_frame_type_list = ['非框架项目', '重复或增补项目', '框架(阿里、中国移动、浙石化、泉化、万华、雄安、许继或大盛ATS Metro、上海小区变、港机等)']
            edit_combobox_frame_type = ttk.Combobox(f_13, font=("ABBvoice CNSG", int(9 * h_ratio)), width=int(130 * w_ratio), state='readonly', values=edit_frame_type_list)
            edit_combobox_frame_type.pack(side=tk.LEFT)
            edit_combobox_frame_type.set(selected_rows[0][32])
            f_13.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_14 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_14, text="编程难度", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_program_type_list = ['不含保护、客供(ABB/非ABB)保护、repeat程序', 'ABB常用保护的国内/出口编程', 'ABB非常用保护编程', 'ABB保护编程+参与车间联调']
            edit_combobox_program_type = ttk.Combobox(f_14, font=("ABBvoice CNSG", int(9 * h_ratio)), width=int(130 * w_ratio), state='readonly', values=edit_program_type_list)
            edit_combobox_program_type.pack(side=tk.LEFT)
            edit_combobox_program_type.set(selected_rows[0][33])
            f_14.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_15 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_15, text="语言难度", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_language_info_list = ['完全中文', '一般英语书面沟通', '英语书面+口语沟通', '复杂英语专业术语沟通']
            edit_combobox_language_info = ttk.Combobox(f_15, font=("ABBvoice CNSG", int(9 * h_ratio)), width=int(130 * w_ratio), state='readonly', values=edit_language_info_list)
            edit_combobox_language_info.pack(side=tk.LEFT)
            edit_combobox_language_info.set(selected_rows[0][34])
            f_15.pack(fill=tk.X)

            tk.Frame(f12, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

            f_16 = tk.Frame(f12, bg="#eaf1f6")
            tk.Label(f_16, text="管理难度", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
            edit_management_info_list = ['标准项目', '同一项目号下多种产品、首次生产新柜型项目', '复杂文档要求(客户要求模板)', '需要读复杂Spec文件、FAT要求复杂、客户特殊复杂的技术规格要求']
            edit_combobox_management_info = ttk.Combobox(f_16, font=("ABBvoice CNSG", int(9 * h_ratio)), width=int(130 * w_ratio), state='readonly', values=edit_management_info_list)
            edit_combobox_management_info.pack(side=tk.LEFT)
            edit_combobox_management_info.set(selected_rows[0][35])
            f_16.pack(fill=tk.X)

            tk.Frame(f2, width=int(200 * w_ratio), bg="black").pack(fill=tk.X, pady=int(10 * h_ratio))

            tk.Label(f2, text='输入文件', bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT)
            f21 = tk.Frame(f2, bg="#eaf1f6", relief='solid', borderwidth=1)
            f21.pack(side=tk.LEFT)

            input_files = ['Check List', '签字版单线图', '客户/设计院图纸', '参考项目信息(保护、产品一致，部分参考视为无参考)', '完全按照客户图纸设计']
            checkbox_input_list = []
            edit_text_input_file = tk.Text(f21, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(160 * w_ratio))

            for i in range(len(input_files)):
                v = IntVar()  # 将各复选框绑定到变量v
                checkbox_input_file = tk.Checkbutton(f21, text=input_files[i], variable=v, command=lambda: select_input_file_single(edit_text_input_file, input_files), font=("ABBvoice CNSG", int(13 * h_ratio)), height=1, background='#eaf1f6')
                checkbox_input_file.pack(side=tk.TOP, anchor=tk.W, expand=True)
                checkbox_input_list.append(v)  # 将各复选框的variable存储到一个列表

            edit_text_input_file.pack(side=tk.LEFT)
            edit_text_input_file.insert(tk.END, selected_rows[0][36])

            tk.Label(f3, text='图纸需求', bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT)
            f31 = tk.Frame(f3, bg="#eaf1f6", relief='solid', borderwidth=1)
            f31.pack(side=tk.LEFT)

            drawing_types = ['SLD/FVD/FFD/FFDD/DS/S/E', 'SVD', 'C', 'AFD', 'LOGIC/拓扑图', '客户图框', 'AS BUILT一台柜子一套图', '非标铭牌图', 'BB/BC', '其他特殊图纸', 'AS BUILT其他语言(备注)']
            checkbox_drawing_type_list = []
            edit_text_drawing_type = tk.Text(f31, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(160 * w_ratio))

            for i in range(len(drawing_types)):
                v = IntVar()
                checkbox_drawing_type = tk.Checkbutton(f31, text=drawing_types[i], variable=v, command=lambda: select_drawing_type_single(edit_text_drawing_type, drawing_types), font=("ABBvoice CNSG", int(13 * h_ratio)), height=1, background='#eaf1f6')
                checkbox_drawing_type.pack(side=tk.TOP, anchor=tk.W, expand=True)
                checkbox_drawing_type_list.append(v)

            edit_text_drawing_type.pack(side=tk.LEFT)
            edit_text_drawing_type.insert(tk.END, selected_rows[0][37])

            tk.Label(f4, text='智能方案', bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT)
            f41 = tk.Frame(f4, bg="#eaf1f6", relief='solid', borderwidth=1)
            f41.pack(side=tk.LEFT)

            intelligence_needs = ['断路器手车电操', '接地开关电操', '温升在线监测', '五防联锁监测', '断路器机械特性', '真空泡VI电寿命', '视频摄像头']
            checkbox_intelligence_list = []
            edit_text_intelligence = tk.Text(f41, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(160 * w_ratio))

            for i in range(len(intelligence_needs)):
                v = IntVar()
                checkbox_intelligence = tk.Checkbutton(f41, text=intelligence_needs[i], variable=v, command=lambda: select_intelligence_single(edit_text_intelligence, intelligence_needs), font=("ABBvoice CNSG", int(13 * h_ratio)), height=1, background='#eaf1f6')
                checkbox_intelligence.pack(side=tk.TOP, anchor=tk.W, expand=True)
                checkbox_intelligence_list.append(v)

            edit_text_intelligence.pack(side=tk.LEFT)
            edit_text_intelligence.insert(tk.END, selected_rows[0][11])

            tk.Label(f5, text='典型柜配置', bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.TOP)

            num_columns1 = 3
            num_rows1 = int(len(selected_rows[0][48].split(';')) / num_columns1)

            entry_widths1 = [int(25 * w_ratio), int(75 * w_ratio), int(75 * w_ratio)]
            entries1 = [[None for _ in range(num_columns1)] for _ in range(num_rows1)]

            selfplus1 = 0
            for row in range(num_rows1):
                row_frame1 = tk.Frame(f5)
                row_frame1.pack(side=tk.TOP, fill=tk.X)

                for col in range(num_columns1):
                    entry1 = tk.Entry(row_frame1, width=entry_widths1[col], relief='solid', font=("ABBvoice CNSG", int(11 * h_ratio)))
                    entry1.pack(side=tk.LEFT)
                    entries1[row][col] = entry1
                    entry1.config(justify='center')
                    if row == 0 and col == 0:
                        entry1.insert(0, '柜号/Typical')
                        entry1['state'] = 'disabled'
                    elif row == 0 and col == 1:
                        entry1.insert(0, '保护料号/建号提单号')
                        entry1['state'] = 'disabled'
                    elif row == 0 and col == 2:
                        entry1.insert(0, '多功能表、电度表料号/建号提单号')
                        entry1['state'] = 'disabled'
                    else:
                        selfplus1 += 1
                        entry1.insert(0, selected_rows[0][48].split(';')[2 + selfplus1])

            tk.Label(f6, text='备注', bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.TOP)

            num_columns2 = 2
            num_rows2 = int(len(selected_rows[0][49].split(';')) / num_columns2)

            entry_widths2 = [int(6 * w_ratio), int(169 * w_ratio)]
            entries2 = [[None for _ in range(num_columns2)] for _ in range(num_rows2)]

            selfplus2 = 0
            for row in range(num_rows2):
                row_frame2 = tk.Frame(f6)
                row_frame2.pack(side=tk.TOP, fill=tk.X)

                for col in range(num_columns2):
                    entry2 = tk.Entry(row_frame2, width=entry_widths2[col], relief='solid', font=("ABBvoice CNSG", int(11 * h_ratio)))
                    entry2.pack(side=tk.LEFT)
                    if col == 0:
                        entry2.config(justify='center')
                        entry2.insert(0, '%s' % str(row + 1))
                        entry2['state'] = 'disabled'
                    else:
                        entry2.insert(0, selected_rows[0][49].split(';')[1 + selfplus2])
                        selfplus2 += 2

                    entries2[row][col] = entry2

            button_confirm_change_single = tk.Button(master=f7, bg="#eaf1f6", text='确认修改', command=lambda: confirm_change_single(edit_window, selected_rows), font=("ABBvoice CNSG", int(10 * h_ratio)), activebackground='blue')
            button_confirm_change_single.pack(side=tk.TOP, pady=int(20 * h_ratio))

            tk.Frame(f7, height=int(80 * h_ratio), bg="#eaf1f6").pack(side=tk.TOP, fill=tk.BOTH, expand=True)

            edit_text_design_type['state'] = 'disabled'
            edit_text_project_number['state'] = 'disabled'
            edit_text_bom_time['state'] = 'disabled'
            edit_text_check_time['state'] = 'disabled'
            edit_text_extra_period['state'] = 'disabled'
            edit_text_item['state'] = 'disabled'
            edit_text_standard_period['state'] = 'disabled'
            # edit_text_upload_time['state'] = 'disabled'
            edit_text_create_time['state'] = 'disabled'

            edit_text_design_type['background'] = '#eaf1f6'
            edit_text_project_number['background'] = '#eaf1f6'
            edit_text_bom_time['background'] = '#eaf1f6'
            edit_text_check_time['background'] = '#eaf1f6'
            edit_text_extra_period['background'] = '#eaf1f6'
            edit_text_item['background'] = '#eaf1f6'
            edit_text_standard_period['background'] = '#eaf1f6'
            # edit_text_upload_time['background'] = '#eaf1f6'
            edit_text_create_time['background'] = '#eaf1f6'

            edit_canvas.update_idletasks()
            edit_canvas.config(scrollregion=edit_canvas.bbox('all'))

    try:
        if selected_rows[0][1] != '提前设计':
            # 自动选择额外的条目，确保所有条目都被选取
            conn = sqlite3.connect('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db')
            cursor = conn.cursor()  # 创建一个Cursor

            attributes = []

            attributes.append("design_type='" + str(selected_rows[0][1]) + "'")
            attributes.append("project_number='" + str(selected_rows[0][2]) + "'")

            attributes = [attr for attr in attributes if attr != '']

            query = "SELECT * FROM project_data WHERE "
            query += " AND ".join(attributes)
            query += "ORDER BY CASE WHEN design_type = '提前设计' THEN rowid END DESC, CASE WHEN design_type = '图纸设计' OR design_type = '工程设计' THEN project_number END DESC"

            cursor.execute(query)
            selected_rows_all = cursor.fetchall()

            selected_rows_all = [[element if element is not None else '' for element in row] for row in selected_rows_all]
            cursor.close()
            conn.close()

            if len(selected_rows) == len(selected_rows_all):
                temp_list = []
                for i in range(0, len(selected_rows_all)):
                    temp_list.append(list(selected_rows_all[i])[0:43] + list(selected_rows_all[i])[47:] + list(selected_rows_all[i])[43:47])

                selected_rows = temp_list
                project_change_multi(parent, w_ratio, h_ratio, selected_rows)
            else:
                execute_program()
        else:
            execute_program()
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def project_change_multi(parent, w_ratio, h_ratio, selected_rows):
    try:
        # 自动选择额外的条目，确保所有条目都被选取
        conn = sqlite3.connect('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db')
        cursor = conn.cursor()  # 创建一个Cursor

        attributes = []

        attributes.append("design_type='" + str(selected_rows[0][1]) + "'")
        attributes.append("project_number='" + str(selected_rows[0][2]) + "'")

        attributes = [attr for attr in attributes if attr != '']

        query = "SELECT * FROM project_data WHERE "
        query += " AND ".join(attributes)
        query += "ORDER BY CASE WHEN design_type = '提前设计' THEN rowid END DESC, CASE WHEN design_type = '图纸设计' OR design_type = '工程设计' THEN project_number END DESC"

        cursor.execute(query)
        selected_rows_all = cursor.fetchall()

        selected_rows_all = [[element if element is not None else '' for element in row] for row in selected_rows_all]
        cursor.close()
        conn.close()

        if len(selected_rows) < len(selected_rows_all):
            confirm_yn = tk.messagebox.askquestion("提示", "发现该项目号下有其他行号未选中，是否选择所有的行号？")
            if confirm_yn == 'yes':
                temp_list = []
                for i in range(0, len(selected_rows_all)):
                    temp_list.append(list(selected_rows_all[i])[0:43] + list(selected_rows_all[i])[47:] + list(selected_rows_all[i])[43:47])
                selected_rows = temp_list

        edit_window = tk.Toplevel(parent)
        edit_window.grab_set()  # 禁用parent窗口的操作

        global edit_canvas

        edit_canvas = tk.Canvas(edit_window, width=int(1800 * w_ratio), height=int(740 * h_ratio), bg="#C9DBE9", borderwidth=0, highlightthickness=0)
        edit_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        edit_canvas.delete('all')  # 清空画布
        edit_canvas.update()
        edit_canvas.bind("<MouseWheel>", on_edit_mousewheel)

        scrollbar_v = tk.Scrollbar(master=edit_window)
        scrollbar_v.pack(side=tk.RIGHT, fill=tk.Y)
        scrollbar_v.config(command=edit_canvas.yview)
        edit_canvas.config(yscrollcommand=scrollbar_v.set)

        edit_content = tk.Frame(edit_canvas, bg="#eaf1f6")

        edit_canvas.create_window(0, 0, width=int(1700 * w_ratio), anchor=tk.NW, window=edit_content)

        f0 = tk.Frame(edit_content, bg="#c9dbe9", bd=0)
        tk.Label(f0, text="项目信息修改——多行号批量", bg="#c9dbe9", fg="black", height=int(1 * h_ratio), font=("ABBvoice CNSG", int(20 * h_ratio), "bold")).pack(fill=tk.X)
        tk.Label(f0, text="(黑色属性无法修改，绿色属性可修改，红色属性为多CI共有属性可修改)", bg="#c9dbe9", fg="red", height=int(1 * h_ratio), font=("ABBvoice CNSG", int(12 * h_ratio), "bold")).pack(fill=tk.X)
        f0.pack(side=tk.TOP, fill=tk.X)

        f1 = tk.Frame(edit_content, bg="#eaf1f6")
        f1.pack(side=tk.TOP)

        f11 = tk.Frame(f1, bg="#eaf1f6")
        f11.pack(side=tk.LEFT, fill=tk.BOTH)

        tk.Frame(f1, width=int(1 * h_ratio), bg="black").pack(side=tk.LEFT, fill=tk.Y, padx=int(20 * w_ratio))

        f12 = tk.Frame(f1, bg="#eaf1f6")
        f12.pack(side=tk.LEFT, fill=tk.BOTH)

        f2 = tk.Frame(edit_content, bg="#eaf1f6")
        f2.pack(side=tk.TOP, pady=int(10*h_ratio))

        f3 = tk.Frame(edit_content, bg="#eaf1f6")
        f3.pack(side=tk.TOP, pady=int(10*h_ratio))

        f4 = tk.Frame(edit_content, bg="#eaf1f6")
        f4.pack(side=tk.TOP, pady=int(10*h_ratio))

        f5 = tk.Frame(edit_content, bg="#eaf1f6")
        f5.pack(side=tk.TOP, pady=int(10*h_ratio))

        f6 = tk.Frame(edit_content, bg="#eaf1f6")
        f6.pack(side=tk.TOP, pady=int(10*h_ratio))

        f7 = tk.Frame(edit_content, bg="#eaf1f6")
        f7.pack(side=tk.TOP, pady=int(10*h_ratio))

        # 左半边
        selected_project_name_list = []
        selected_item_list = []
        selected_typical_amount_list = []
        selected_panel_amount_list = []
        selected_typical_type_list = []
        selected_product_type_list = []
        selected_PE_list = []
        selected_DE_list = []
        selected_main_busbar_current_list = []
        selected_main_protection_list = []
        selected_arclight_list = []
        selected_location_list = []
        selected_drawing_language_list = []
        selected_standard_period_list = []
        selected_extra_period_list = []
        selected_advice_times_list = []
        selected_mcb_company_list = []
        selected_aux_list = []
        selected_terminal_company_list = []
        selected_charged_display_company_list = []
        selected_switch_company_list = []
        selected_ct_type_list = []
        selected_pt_type_list = []
        selected_sa_type_list = []
        selected_customer_wiring_list = []
        selected_customer_requirements_list = []
        # 右半边
        selected_upload_time_list = []
        selected_bom_time_list = []
        selected_check_time_list = []
        selected_receive_time_list = []
        selected_start_time_list = []
        selected_estimated_time_list = []
        selected_status_info_list = []
        selected_actual_time_list = []
        selected_team_info_list = []
        selected_abnormal_status_list = []
        selected_create_time_list = []
        selected_project_type_list = []
        selected_frame_type_list = []
        selected_program_type_list = []
        selected_language_info_list = []
        selected_management_info_list = []
        selected_input_file_list = []
        selected_drawing_requirement_list = []
        selected_intelligence_list = []

        for i in range(0, len(selected_rows)):
            selected_project_name_list.append(selected_rows[i][4])
            selected_item_list.append(selected_rows[i][3])
            selected_typical_amount_list.append(selected_rows[i][5])
            selected_panel_amount_list.append(selected_rows[i][6])
            selected_typical_type_list.append(selected_rows[i][7])
            selected_product_type_list.append(selected_rows[i][8])
            selected_PE_list.append(selected_rows[i][9])
            selected_DE_list.append(selected_rows[i][10])
            selected_main_busbar_current_list.append(selected_rows[i][12])
            selected_main_protection_list.append(selected_rows[i][13])
            selected_arclight_list.append(selected_rows[i][14])
            selected_location_list.append(selected_rows[i][15])
            selected_drawing_language_list.append(selected_rows[i][16])
            selected_standard_period_list.append(selected_rows[i][17])
            selected_extra_period_list.append(selected_rows[i][18])
            selected_advice_times_list.append(selected_rows[i][19])
            selected_mcb_company_list.append(selected_rows[i][38])
            selected_aux_list.append(selected_rows[i][39])
            selected_terminal_company_list.append(selected_rows[i][40])
            selected_charged_display_company_list.append(selected_rows[i][41])
            selected_switch_company_list.append(selected_rows[i][42])
            selected_ct_type_list.append(selected_rows[i][43])
            selected_pt_type_list.append(selected_rows[i][44])
            selected_sa_type_list.append(selected_rows[i][45])
            selected_customer_wiring_list.append(selected_rows[i][46])
            selected_customer_requirements_list.append(selected_rows[i][47])

            selected_upload_time_list.append(selected_rows[i][20])
            selected_bom_time_list.append(selected_rows[i][21])
            selected_check_time_list.append(selected_rows[i][22])
            selected_receive_time_list.append(selected_rows[i][23])
            selected_start_time_list.append(selected_rows[i][24])
            selected_estimated_time_list.append(selected_rows[i][25])
            selected_status_info_list.append(selected_rows[i][26])
            selected_actual_time_list.append(selected_rows[i][27])
            selected_team_info_list.append(selected_rows[i][28])
            selected_abnormal_status_list.append(selected_rows[i][29])
            selected_create_time_list.append(selected_rows[i][30])
            selected_project_type_list.append(selected_rows[i][31])
            selected_frame_type_list.append(selected_rows[i][32])
            selected_program_type_list.append(selected_rows[i][33])
            selected_language_info_list.append(selected_rows[i][34])
            selected_management_info_list.append(selected_rows[i][35])
            selected_input_file_list.append(selected_rows[i][36])
            selected_drawing_requirement_list.append(selected_rows[i][37])
            selected_intelligence_list.append(selected_rows[i][11])

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_1 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_1, text="设计类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_text_design_type
        edit_text_design_type = tk.Text(f_1, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
        edit_text_design_type.pack(side=tk.LEFT)
        edit_text_design_type.insert(tk.END, selected_rows[0][1])
        f_1.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_2 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_2, text="   项目号", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_text_project_number
        edit_text_project_number = tk.Text(f_2, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(100 * w_ratio))
        edit_text_project_number.pack(side=tk.LEFT)
        edit_text_project_number.insert(tk.END, selected_rows[0][2])
        f_2.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_3 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_3, text="   项目名", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_text_project_name
        edit_text_project_name = tk.Text(f_3, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100 * w_ratio))
        edit_text_project_name.pack(side=tk.LEFT)
        edit_text_project_name.insert(tk.END, '\n'.join(str(num) for num in selected_project_name_list))
        f_3.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_4 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_4, text="      行号", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_text_item
        edit_text_item = tk.Text(f_4, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100 * w_ratio))
        edit_text_item.pack(side=tk.LEFT)
        edit_text_item.insert(tk.END, '\n'.join(str(num) for num in selected_item_list))
        f_4.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_5 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_5, text="      图套", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_text_typical_amount
        edit_text_typical_amount = tk.Text(f_5, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100 * w_ratio))
        edit_text_typical_amount.pack(side=tk.LEFT)
        edit_text_typical_amount.insert(tk.END, '\n'.join(str(num) for num in selected_typical_amount_list))
        f_5.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_6 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_6, text="      台数", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_text_panel_amount
        edit_text_panel_amount = tk.Text(f_6, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100 * w_ratio))
        edit_text_panel_amount.pack(side=tk.LEFT)
        edit_text_panel_amount.insert(tk.END, '\n'.join(str(num) for num in selected_panel_amount_list))
        f_6.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_7 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_7, text="      柜型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_text_typical_type
        edit_text_typical_type = tk.Text(f_7, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100 * w_ratio))
        edit_text_typical_type.pack(side=tk.LEFT)
        edit_text_typical_type.insert(tk.END, '\n'.join(str(num) for num in selected_typical_type_list))
        f_7.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_8 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_8, text="产品类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_product_type_list
        edit_combobox_product_type_list = []
        edit_product_type_list = ['AIS', 'GIS']
        for i in range(0, len(selected_product_type_list)):
            edit_combobox_product_type = ttk.Combobox(f_8, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), values=edit_product_type_list)
            edit_combobox_product_type.pack(side=tk.TOP, anchor='w')
            edit_combobox_product_type.set(selected_product_type_list[i])
            edit_combobox_product_type_list.append(edit_combobox_product_type)
        f_8.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_9 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_9, text="   工程师", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_PE_list = []
        if os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\PE_DE.xlsx"):
            workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\PE_DE.xlsx")
            worksheet = workbook['Sheet1']
            for i in range(1, worksheet.max_row+1):
                edit_PE_list.append(worksheet.cell(row=i, column=1).value)
        global edit_combobox_PE_list
        edit_combobox_PE_list = []
        for i in range(0, len(selected_PE_list)):
            edit_combobox_PE = ttk.Combobox(f_9, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), values=edit_PE_list)
            edit_combobox_PE.pack(side=tk.TOP, anchor='w')
            edit_combobox_PE.set(selected_PE_list[i])
            edit_combobox_PE_list.append(edit_combobox_PE)
        f_9.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_10 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_10, text="   绘图员", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_DE_list = []
        if os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\PE_DE.xlsx"):
            workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\PE_DE.xlsx")
            worksheet = workbook['Sheet1']
            for i in range(1, worksheet.max_row + 1):
                edit_DE_list.append(worksheet.cell(row=i, column=1).value)
        global edit_combobox_DE_list
        edit_combobox_DE_list = []
        for i in range(0, len(selected_DE_list)):
            edit_combobox_DE = ttk.Combobox(f_10, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), values=edit_DE_list)
            edit_combobox_DE.pack(side=tk.TOP, anchor='w')
            edit_combobox_DE.set(selected_DE_list[i])
            edit_combobox_DE_list.append(edit_combobox_DE)
        f_10.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_11 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_11, text="母线电流", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        edit_main_busbar_current_list = []
        if os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\Main_current.xlsx"):
            workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\Main_current.xlsx")
            worksheet = workbook['Sheet1']
            for i in range(1, worksheet.max_row + 1):
                edit_main_busbar_current_list.append(worksheet.cell(row=i, column=1).value)
        global edit_combobox_main_busbar_current_list
        edit_combobox_main_busbar_current_list = []
        for i in range(0, len(selected_main_busbar_current_list)):
            edit_combobox_main_busbar_current = ttk.Combobox(f_11, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_main_busbar_current_list)
            edit_combobox_main_busbar_current.pack(side=tk.TOP, anchor='w')
            edit_combobox_main_busbar_current.set(selected_main_busbar_current_list[i])
            edit_combobox_main_busbar_current_list.append(edit_combobox_main_busbar_current)
        f_11.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_12 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_12, text="   主保护", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_text_main_protection
        edit_text_main_protection = tk.Text(f_12, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100 * w_ratio))
        edit_text_main_protection.pack(side=tk.LEFT)
        edit_text_main_protection.insert(tk.END, '\n'.join(str(num) for num in selected_main_protection_list))
        f_12.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_13 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_13, text="弧光方案", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_text_arclight
        edit_text_arclight = tk.Text(f_13, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100 * w_ratio))
        edit_text_arclight.pack(side=tk.LEFT)
        edit_text_arclight.insert(tk.END, '\n'.join(str(num) for num in selected_arclight_list))
        f_13.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_14 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_14, text="      地区", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_location_list
        edit_combobox_location_list = []
        edit_location_list = ['UT万华', '港机', 'Building&数据中心', '工业/UT', '核电', '船用', '瑞典DOK80', '其他', '海外']
        for i in range(0, len(selected_location_list)):
            edit_combobox_location = ttk.Combobox(f_14, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_location_list)
            edit_combobox_location.pack(side=tk.TOP, anchor='w')
            edit_combobox_location.set(selected_location_list[i])
            edit_combobox_location_list.append(edit_combobox_location)
        f_14.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_15 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_15, text="图纸语言", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_drawing_language_list
        edit_combobox_drawing_language_list = []
        edit_drawing_language_list = ['中文', '英文', '中英文']
        for i in range(0, len(selected_drawing_language_list)):
            edit_combobox_drawing_language = ttk.Combobox(f_15, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_drawing_language_list)
            edit_combobox_drawing_language.pack(side=tk.TOP, anchor='w')
            edit_combobox_drawing_language.set(selected_drawing_language_list[i])
            edit_combobox_drawing_language_list.append(edit_combobox_drawing_language)
        f_15.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_16 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_16, text="标准周期", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_text_standard_period
        edit_text_standard_period = tk.Text(f_16, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100 * w_ratio))
        edit_text_standard_period.pack(side=tk.LEFT)
        edit_text_standard_period.insert(tk.END, '\n'.join(str(num) for num in selected_standard_period_list))
        f_16.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_17 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_17, text="额外周期", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_text_extra_period
        edit_text_extra_period = tk.Text(f_17, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100 * w_ratio))
        edit_text_extra_period.pack(side=tk.LEFT)
        edit_text_extra_period.insert(tk.END, '\n'.join(str(num) for num in selected_extra_period_list))
        f_17.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_18 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_18, text="意见次数", height=int(1 * h_ratio), bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_advice_times
        edit_advice_times_list = ['1', '2', '3', '4']
        edit_combobox_advice_times = ttk.Combobox(f_18, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_advice_times_list)
        edit_combobox_advice_times.pack(side=tk.LEFT)
        edit_combobox_advice_times.set(';'.join(str(num) for num in set(selected_advice_times_list)))
        f_18.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_19 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_19, text="空开厂家", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_mcb_type_list
        edit_combobox_mcb_type_list = []
        edit_mcb_type_list = ['人民电器', 'ABB', '熔丝+熔芯']
        for i in range(0, len(selected_mcb_company_list)):
            edit_combobox_mcb_type = ttk.Combobox(f_19, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_mcb_type_list)
            edit_combobox_mcb_type.pack(side=tk.TOP, anchor='w')
            edit_combobox_mcb_type.set(selected_mcb_company_list[i])
            edit_combobox_mcb_type_list.append(edit_combobox_mcb_type)
        f_19.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_20 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_20, text="辅助触点", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_aux_type_list
        edit_combobox_aux_type_list = []
        edit_aux_type_list = ['是', '否']
        for i in range(0, len(selected_aux_list)):
            edit_combobox_aux_type = ttk.Combobox(f_20, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_aux_type_list)
            edit_combobox_aux_type.pack(side=tk.TOP, anchor='w')
            edit_combobox_aux_type.set(selected_aux_list[i])
            edit_combobox_aux_type_list.append(edit_combobox_aux_type)
        f_20.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_21 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_21, text="端子厂家", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_terminal_type_list
        edit_combobox_terminal_type_list = []
        edit_terminal_type_list = ['瑞联', 'TE', '魏德米勒', '菲尼克斯(*)']
        for i in range(0, len(selected_terminal_company_list)):
            edit_combobox_terminal_type = ttk.Combobox(f_21, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_terminal_type_list)
            edit_combobox_terminal_type.pack(side=tk.TOP, anchor='w')
            edit_combobox_terminal_type.set(selected_terminal_company_list[i])
            edit_combobox_terminal_type_list.append(edit_combobox_terminal_type)
        f_21.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_22 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_22, text="带显厂家", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_charged_display_type_list
        edit_combobox_charged_display_type_list = []
        edit_charged_display_type_list = ['立林', '百岗']
        for i in range(0, len(selected_charged_display_company_list)):
            edit_combobox_charged_display_type = ttk.Combobox(f_22, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_charged_display_type_list)
            edit_combobox_charged_display_type.pack(side=tk.TOP, anchor='w')
            edit_combobox_charged_display_type.set(selected_charged_display_company_list[i])
            edit_combobox_charged_display_type_list.append(edit_combobox_charged_display_type)
        f_22.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_23 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_23, text="选择开关", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_switch_type_list
        edit_combobox_switch_type_list = []
        edit_switch_type_list = ['江阴长江', 'K&N', '其他(备注)']
        for i in range(0, len(selected_switch_company_list)):
            edit_combobox_switch_type = ttk.Combobox(f_23, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_switch_type_list)
            edit_combobox_switch_type.pack(side=tk.TOP, anchor='w')
            edit_combobox_switch_type.set(selected_switch_company_list[i])
            edit_combobox_switch_type_list.append(edit_combobox_switch_type)
        f_23.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_24 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_24, text="  CT类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_ct_type_list
        edit_combobox_ct_type_list = []
        edit_ct_type_list = ['ABB', 'DYH', 'NTK', 'TLEP', '无']
        for i in range(0, len(selected_ct_type_list)):
            edit_combobox_ct_type = ttk.Combobox(f_24, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_ct_type_list)
            edit_combobox_ct_type.pack(side=tk.TOP, anchor='w')
            edit_combobox_ct_type.set(selected_ct_type_list[i])
            edit_combobox_ct_type_list.append(edit_combobox_ct_type)
        f_24.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_25 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_25, text="  PT类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_pt_type_list
        edit_combobox_pt_type_list = []
        edit_pt_type_list = ['ABB', 'DYH', 'NTK', 'TLEP', '无']
        for i in range(0, len(selected_pt_type_list)):
            edit_combobox_pt_type = ttk.Combobox(f_25, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_pt_type_list)
            edit_combobox_pt_type.pack(side=tk.TOP, anchor='w')
            edit_combobox_pt_type.set(selected_pt_type_list[i])
            edit_combobox_pt_type_list.append(edit_combobox_pt_type)
        f_25.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_26 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_26, text="  SA类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_sa_type_list
        edit_combobox_sa_type_list = []
        edit_sa_type_list = ['神电', '日立', 'GCA']
        for i in range(0, len(selected_sa_type_list)):
            edit_combobox_sa_type = ttk.Combobox(f_26, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_sa_type_list)
            edit_combobox_sa_type.pack(side=tk.TOP, anchor='w')
            edit_combobox_sa_type.set(selected_sa_type_list[i])
            edit_combobox_sa_type_list.append(edit_combobox_sa_type)
        f_26.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_27 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_27, text="客户线号", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_customer_wiring_list
        edit_combobox_customer_wiring_list = []
        edit_customer_wiring_list = ['否', '是(备注前期项目号及需更改内容)']
        for i in range(0, len(selected_customer_wiring_list)):
            edit_combobox_customer_wiring = ttk.Combobox(f_27, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_customer_wiring_list)
            edit_combobox_customer_wiring.pack(side=tk.TOP, anchor='w')
            edit_combobox_customer_wiring.set(selected_customer_wiring_list[i])
            edit_combobox_customer_wiring_list.append(edit_combobox_customer_wiring)
        f_27.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_28 = tk.Frame(f11, bg="#eaf1f6")
        tk.Label(f_28, text="客户需求", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_customer_requirements_list
        edit_combobox_customer_requirements_list = []
        edit_customer_requirements_list = ['否', '是(备注前期项目号及需更改内容)']
        for i in range(0, len(selected_customer_requirements_list)):
            edit_combobox_customer_requirements = ttk.Combobox(f_28, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_customer_requirements_list)
            edit_combobox_customer_requirements.pack(side=tk.TOP, anchor='w')
            edit_combobox_customer_requirements.set(selected_customer_requirements_list[i])
            edit_combobox_customer_requirements_list.append(edit_combobox_customer_requirements)
        f_28.pack(fill=tk.X)

        tk.Frame(f11, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_1 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_1, text="提原理图", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        f_11 = tk.Frame(f_1, bg="#eaf1f6")
        f_11.pack(side=tk.LEFT)
        global edit_text_upload_time_list
        edit_text_upload_time_list = []
        for i in range(0, len(selected_upload_time_list)):
            f_101 = tk.Frame(f_11, bg="#eaf1f6")
            f_101.pack(side=tk.TOP, anchor='w')
            edit_text_upload_time = tk.Text(f_101, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(26 * w_ratio))
            edit_text_upload_time.pack(side=tk.LEFT)
            edit_text_upload_time.insert(tk.END, str(selected_upload_time_list[i]))
            edit_text_upload_time['state'] = 'disabled'
            edit_text_upload_time_list.append(edit_text_upload_time)

            button_select_upload_time = tk.Button(master=f_101, bg="#eaf1f6", text='选择', command=lambda et=edit_text_upload_time: select_upload_time(edit_content, et), font=("ABBvoice CNSG", int(9 * h_ratio)), activebackground='blue')
            button_select_upload_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))
        f_1.pack(fill=tk.X)

        tk.Frame(f12, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_2 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_2, text="发放初级", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        f_21 = tk.Frame(f_2, bg="#eaf1f6")
        f_21.pack(side=tk.LEFT)
        global edit_text_bom_time_list
        edit_text_bom_time_list = []
        for i in range(0, len(selected_bom_time_list)):
            f_201 = tk.Frame(f_21, bg="#eaf1f6")
            f_201.pack(side=tk.TOP, anchor='w')
            edit_text_bom_time = tk.Text(f_201, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(26 * w_ratio))
            edit_text_bom_time.pack(side=tk.LEFT)
            edit_text_bom_time.insert(tk.END, str(selected_bom_time_list[i]))
            edit_text_bom_time['state'] = 'disabled'
            edit_text_bom_time_list.append(edit_text_bom_time)

            button_select_bom_time = tk.Button(master=f_201, bg="#eaf1f6", text='选择', command=lambda et=edit_text_bom_time: select_bom_time(edit_content, et), font=("ABBvoice CNSG", int(9 * h_ratio)), activebackground='blue')
            button_select_bom_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))
        f_2.pack(fill=tk.X)

        tk.Frame(f12, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_3 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_3, text="生产检查", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        f_31 = tk.Frame(f_3, bg="#eaf1f6")
        f_31.pack(side=tk.LEFT)
        global edit_text_check_time_list
        edit_text_check_time_list = []
        for i in range(0, len(selected_check_time_list)):
            f_301 = tk.Frame(f_31, bg="#eaf1f6")
            f_301.pack(side=tk.TOP, anchor='w')
            edit_text_check_time = tk.Text(f_301, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(26 * w_ratio))
            edit_text_check_time.pack(side=tk.LEFT)
            edit_text_check_time.insert(tk.END, str(selected_check_time_list[i]))
            edit_text_check_time['state'] = 'disabled'
            edit_text_check_time_list.append(edit_text_check_time)

            button_select_check_time = tk.Button(master=f_301, bg="#eaf1f6", text='选择', command=lambda et=edit_text_check_time: select_check_time(edit_content, et), font=("ABBvoice CNSG", int(9 * h_ratio)), activebackground='blue')
            button_select_check_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))

        f_3.pack(fill=tk.X)

        tk.Frame(f12, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_4 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_4, text="接收时间", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        f_41 = tk.Frame(f_4, bg="#eaf1f6")
        f_41.pack(side=tk.LEFT)
        global edit_text_receive_time_list
        edit_text_receive_time_list = []
        for i in range(0, len(selected_receive_time_list)):
            f_401 = tk.Frame(f_41, bg="#eaf1f6")
            f_401.pack(side=tk.TOP, anchor='w')
            edit_text_receive_time = tk.Text(f_401, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(26 * w_ratio))
            edit_text_receive_time.pack(side=tk.LEFT)
            edit_text_receive_time.insert(tk.END, str(selected_receive_time_list[i]))
            edit_text_receive_time['state'] = 'disabled'
            edit_text_receive_time_list.append(edit_text_receive_time)

            button_select_receive_time = tk.Button(master=f_401, bg="#eaf1f6", text='选择', command=lambda et=edit_text_receive_time: select_receive_time(edit_content, et), font=("ABBvoice CNSG", int(9 * h_ratio)), activebackground='blue')
            button_select_receive_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))

        f_4.pack(fill=tk.X)

        tk.Frame(f12, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_5 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_5, text="启动时间", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        f_51 = tk.Frame(f_5, bg="#eaf1f6")
        f_51.pack(side=tk.LEFT)
        global edit_text_start_time_list
        edit_text_start_time_list = []
        for i in range(0, len(selected_start_time_list)):
            f_501 = tk.Frame(f_51, bg="#eaf1f6")
            f_501.pack(side=tk.TOP, anchor='w')
            edit_text_start_time = tk.Text(f_501, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(26 * w_ratio))
            edit_text_start_time.pack(side=tk.LEFT)
            edit_text_start_time.insert(tk.END, str(selected_start_time_list[i]))
            edit_text_start_time['state'] = 'disabled'
            edit_text_start_time_list.append(edit_text_start_time)

            button_select_start_time = tk.Button(master=f_501, bg="#eaf1f6", text='选择', command=lambda et=edit_text_start_time: select_start_time(edit_content, et), font=("ABBvoice CNSG", int(9 * h_ratio)), activebackground='blue')
            button_select_start_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))
        f_5.pack(fill=tk.X)

        tk.Frame(f12, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_6 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_6, text="预计交付", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        f_61 = tk.Frame(f_6, bg="#eaf1f6")
        f_61.pack(side=tk.LEFT)
        global edit_text_estimated_time_list
        edit_text_estimated_time_list = []
        for i in range(0, len(selected_estimated_time_list)):
            f_601 = tk.Frame(f_61, bg="#eaf1f6")
            f_601.pack(side=tk.TOP, anchor='w')
            edit_text_estimated_time = tk.Text(f_601, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(26 * w_ratio))
            edit_text_estimated_time.pack(side=tk.LEFT)
            edit_text_estimated_time.insert(tk.END, str(selected_estimated_time_list[i]))
            edit_text_estimated_time['state'] = 'disabled'
            edit_text_estimated_time_list.append(edit_text_estimated_time)

            button_select_estimated_time = tk.Button(master=f_601, bg="#eaf1f6", text='选择', command=lambda et=edit_text_estimated_time: select_estimated_time(edit_content, et), font=("ABBvoice CNSG", int(9 * h_ratio)), activebackground='blue')
            button_select_estimated_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))
        f_6.pack(fill=tk.X)

        tk.Frame(f12, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_7 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_7, text="      状态", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_status_info_list
        edit_combobox_status_info_list = []
        edit_status_info_list = ['新项目', '待启动', '单线图及MVE完成', '初版图纸设计', '初版原理图提交', '生产图完成']
        for i in range(0, len(selected_status_info_list)):
            edit_combobox_status_info = ttk.Combobox(f_7, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_status_info_list)
            edit_combobox_status_info.pack(side=tk.TOP, anchor='w')
            edit_combobox_status_info.set(selected_status_info_list[i])
            edit_combobox_status_info_list.append(edit_combobox_status_info)
        f_7.pack(fill=tk.X)

        tk.Frame(f12, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_8 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_8, text="实际完成", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        f_81 = tk.Frame(f_8, bg="#eaf1f6")
        f_81.pack(side=tk.LEFT)
        global edit_text_actual_time_list
        edit_text_actual_time_list = []
        for i in range(0, len(selected_actual_time_list)):
            f_801 = tk.Frame(f_81, bg="#eaf1f6")
            f_801.pack(side=tk.TOP, anchor='w')
            edit_text_actual_time = tk.Text(f_801, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=1, width=int(26 * w_ratio))
            edit_text_actual_time.pack(side=tk.LEFT)
            edit_text_actual_time.insert(tk.END, str(selected_actual_time_list[i]))
            edit_text_actual_time['state'] = 'disabled'
            edit_text_actual_time_list.append(edit_text_actual_time)

            button_select_actual_time = tk.Button(master=f_801, bg="#eaf1f6", text='选择', command=lambda et=edit_text_actual_time: select_actual_time(edit_content, et), font=("ABBvoice CNSG", int(9 * h_ratio)), activebackground='blue')
            button_select_actual_time.pack(side=tk.LEFT, padx=int(10 * w_ratio))
        f_8.pack(fill=tk.X)

        tk.Frame(f12, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_9 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_9, text="      小组", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_team_info_list
        edit_combobox_team_info_list = []
        edit_team_info_list = ['A', 'B', 'C', 'D']
        for i in range(0, len(selected_team_info_list)):
            edit_combobox_team_info = ttk.Combobox(f_9, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(26 * w_ratio), state='readonly', values=edit_team_info_list)
            edit_combobox_team_info.pack(side=tk.TOP, anchor='w')
            edit_combobox_team_info.set(selected_team_info_list[i])
            edit_combobox_team_info_list.append(edit_combobox_team_info)
        f_9.pack(fill=tk.X)

        tk.Frame(f12, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_10 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_10, text="异常情况", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_text_abnormal_status
        edit_text_abnormal_status = tk.Text(f_10, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100 * w_ratio))
        edit_text_abnormal_status.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
        edit_text_abnormal_status.insert(tk.END, '\n'.join(str(num) for num in selected_abnormal_status_list))
        f_10.pack(fill=tk.X)

        tk.Frame(f12, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_11 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_11, text="创建时间", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_text_create_time
        edit_text_create_time = tk.Text(f_11, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(100 * w_ratio))
        edit_text_create_time.pack(side=tk.LEFT, padx=(0, int(5 * w_ratio)))
        edit_text_create_time.insert(tk.END, '\n'.join(str(num) for num in selected_create_time_list))
        f_11.pack(fill=tk.X)

        tk.Frame(f12, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_12 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_12, text="项目类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_project_type_list
        edit_combobox_project_type_list = []
        edit_project_type_list = ['Domestic', 'Export_fully built', 'EPC', 'Metro', 'Marine', 'Wind power', '供电局项目(含18相反措)']
        for i in range(0, len(selected_project_type_list)):
            edit_combobox_project_type = ttk.Combobox(f_12, font=("ABBvoice CNSG", int(9 * h_ratio)), width=int(130 * w_ratio), state='readonly', values=edit_project_type_list)
            edit_combobox_project_type.pack(side=tk.TOP, anchor='w')
            edit_combobox_project_type.set(selected_project_type_list[i])
            edit_combobox_project_type_list.append(edit_combobox_project_type)
        f_12.pack(fill=tk.X)

        tk.Frame(f12, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_13 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_13, text="框架类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_frame_type_list
        edit_combobox_frame_type_list = []
        edit_frame_type_list = ['非框架项目', '重复或增补项目', '框架(阿里、中国移动、浙石化、泉化、万华、雄安、许继或大盛ATS Metro、上海小区变、港机等)']
        for i in range(0, len(selected_frame_type_list)):
            edit_combobox_frame_type = ttk.Combobox(f_13, font=("ABBvoice CNSG", int(9 * h_ratio)), width=int(130 * w_ratio), state='readonly', values=edit_frame_type_list)
            edit_combobox_frame_type.pack(side=tk.TOP, anchor='w')
            edit_combobox_frame_type.set(selected_frame_type_list[i])
            edit_combobox_frame_type_list.append(edit_combobox_frame_type)
        f_13.pack(fill=tk.X)

        tk.Frame(f12, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_14 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_14, text="编程难度", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_program_type_list
        edit_combobox_program_type_list = []
        edit_program_type_list = ['不含保护、客供(ABB/非ABB)保护、repeat程序', 'ABB常用保护的国内/出口编程', 'ABB非常用保护编程', 'ABB保护编程+参与车间联调']
        for i in range(0, len(selected_program_type_list)):
            edit_combobox_program_type = ttk.Combobox(f_14, font=("ABBvoice CNSG", int(9 * h_ratio)), width=int(130 * w_ratio), state='readonly', values=edit_program_type_list)
            edit_combobox_program_type.pack(side=tk.TOP, anchor='w')
            edit_combobox_program_type.set(selected_program_type_list[i])
            edit_combobox_program_type_list.append(edit_combobox_program_type)
        f_14.pack(fill=tk.X)

        tk.Frame(f12, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_15 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_15, text="语言难度", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_language_info_list
        edit_combobox_language_info_list = []
        edit_language_info_list = ['完全中文', '一般英语书面沟通', '英语书面+口语沟通', '复杂英语专业术语沟通']
        for i in range(0, len(selected_language_info_list)):
            edit_combobox_language_info = ttk.Combobox(f_15, font=("ABBvoice CNSG", int(9 * h_ratio)), width=int(130 * w_ratio), state='readonly', values=edit_language_info_list)
            edit_combobox_language_info.pack(side=tk.TOP, anchor='w')
            edit_combobox_language_info.set(selected_language_info_list[i])
            edit_combobox_language_info_list.append(edit_combobox_language_info)
        f_15.pack(fill=tk.X)

        tk.Frame(f12, height=int(10 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_16 = tk.Frame(f12, bg="#eaf1f6")
        tk.Label(f_16, text="管理难度", height=int(1 * h_ratio), bg="#eaf1f6", fg="green", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global edit_combobox_management_info_list
        edit_combobox_management_info_list = []
        edit_management_info_list = ['标准项目', '同一项目号下多种产品、首次生产新柜型项目', '复杂文档要求(客户要求模板)', '需要读复杂Spec文件、FAT要求复杂、客户特殊复杂的技术规格要求']
        for i in range(0, len(selected_management_info_list)):
            edit_combobox_management_info = ttk.Combobox(f_16, font=("ABBvoice CNSG", int(9 * h_ratio)), width=int(130 * w_ratio), state='readonly', values=edit_management_info_list)
            edit_combobox_management_info.pack(side=tk.TOP, anchor='w')
            edit_combobox_management_info.set(selected_management_info_list[i])
            edit_combobox_management_info_list.append(edit_combobox_management_info)
        f_16.pack(fill=tk.X)

        tk.Frame(f2, width=int(200 * w_ratio), bg="black").pack(fill=tk.X, pady=int(10 * h_ratio))

        tk.Label(f2, text='输入文件', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT)
        f21 = tk.Frame(f2, bg="#eaf1f6", relief='solid', borderwidth=1)
        f21.pack(side=tk.LEFT)

        global input_files
        input_files = ['Check List', '签字版单线图', '客户/设计院图纸', '参考项目信息(保护、产品一致，部分参考视为无参考)', '完全按照客户图纸设计']
        global checkbox_input_list
        checkbox_input_list = []
        global edit_text_input_file
        edit_text_input_file = tk.Text(f21, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(160 * w_ratio))

        for i in range(len(input_files)):
            v = IntVar()    # 将各复选框绑定到变量v
            checkbox_input_file = tk.Checkbutton(f21, text=input_files[i], variable=v, command=lambda et=edit_text_input_file: select_input_file_multi(et, len(selected_rows)), font=("ABBvoice CNSG", int(13 * h_ratio)), height=1, background='#eaf1f6')
            checkbox_input_file.pack(side=tk.TOP, anchor=tk.W, expand=True)
            checkbox_input_list.append(v)    # 将各复选框的variable存储到一个列表

        edit_text_input_file.pack(side=tk.LEFT)
        edit_text_input_file.insert(tk.END, '\n'.join(str(num) for num in selected_input_file_list))

        tk.Label(f3, text='图纸需求', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT)
        f31 = tk.Frame(f3, bg="#eaf1f6", relief='solid', borderwidth=1)
        f31.pack(side=tk.LEFT)

        global drawing_types
        drawing_types = ['SLD/FVD/FFD/FFDD/DS/S/E', 'SVD', 'C', 'AFD', 'LOGIC/拓扑图', '客户图框', 'AS BUILT一台柜子一套图', '非标铭牌图', 'BB/BC', '其他特殊图纸', 'AS BUILT其他语言(备注)']
        global checkbox_drawing_type_list
        checkbox_drawing_type_list = []
        global edit_text_drawing_type
        edit_text_drawing_type = tk.Text(f31, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(160 * w_ratio))

        for i in range(len(drawing_types)):
            v = IntVar()
            checkbox_drawing_type = tk.Checkbutton(f31, text=drawing_types[i], variable=v, command=lambda et=edit_text_drawing_type: select_drawing_type_multi(et, len(selected_rows)), font=("ABBvoice CNSG", int(13 * h_ratio)), height=1, background='#eaf1f6')
            checkbox_drawing_type.pack(side=tk.TOP, anchor=tk.W, expand=True)
            checkbox_drawing_type_list.append(v)

        edit_text_drawing_type.pack(side=tk.LEFT)
        edit_text_drawing_type.insert(tk.END, '\n'.join(str(num) for num in selected_drawing_requirement_list))

        tk.Label(f4, text='智能方案', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT)
        f41 = tk.Frame(f4, bg="#eaf1f6", relief='solid', borderwidth=1)
        f41.pack(side=tk.LEFT)

        global intelligence_needs
        intelligence_needs = ['断路器手车电操', '接地开关电操', '温升在线监测', '五防联锁监测', '断路器机械特性', '真空泡VI电寿命', '视频摄像头']
        global checkbox_intelligence_list
        checkbox_intelligence_list = []
        global edit_text_intelligence
        edit_text_intelligence = tk.Text(f41, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=len(selected_rows), width=int(160 * w_ratio))

        for i in range(len(intelligence_needs)):
            v = IntVar()
            checkbox_intelligence = tk.Checkbutton(f41, text=intelligence_needs[i], variable=v, command=lambda et=edit_text_intelligence: select_intelligence_multi(et, len(selected_rows)), font=("ABBvoice CNSG", int(13 * h_ratio)), height=1, background='#eaf1f6')
            checkbox_intelligence.pack(side=tk.TOP, anchor=tk.W, expand=True)
            checkbox_intelligence_list.append(v)

        edit_text_intelligence.pack(side=tk.LEFT)
        edit_text_intelligence.insert(tk.END, '\n'.join(str(num) for num in selected_intelligence_list))

        tk.Label(f5, text='典型柜配置', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.TOP)

        global num_columns1
        num_columns1 = 3
        global num_rows1
        num_rows1 = int(len(selected_rows[0][48].split(';')) / num_columns1)

        global entry_widths1
        entry_widths1 = [int(25 * w_ratio), int(75 * w_ratio), int(75 * w_ratio)]
        global entries1
        entries1 = [[None for _ in range(num_columns1)] for _ in range(num_rows1)]

        selfplus1 = 0
        for row in range(num_rows1):
            row_frame1 = tk.Frame(f5)
            row_frame1.pack(side=tk.TOP, fill=tk.X)

            for col in range(num_columns1):
                entry1 = tk.Entry(row_frame1, width=entry_widths1[col], relief='solid', font=("ABBvoice CNSG", int(11 * h_ratio)))
                entry1.pack(side=tk.LEFT)
                entries1[row][col] = entry1
                entry1.config(justify='center')
                if row == 0 and col == 0:
                    entry1.insert(0, '柜号/Typical')
                    entry1['state'] = 'disabled'
                elif row == 0 and col == 1:
                    entry1.insert(0, '保护料号/建号提单号')
                    entry1['state'] = 'disabled'
                elif row == 0 and col == 2:
                    entry1.insert(0, '多功能表、电度表料号/建号提单号')
                    entry1['state'] = 'disabled'
                else:
                    selfplus1 += 1
                    entry1.insert(0, selected_rows[0][48].split(';')[2 + selfplus1])

        tk.Label(f6, text='备注', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.TOP)

        global num_columns2
        num_columns2 = 2
        global num_rows2
        num_rows2 = int(len(selected_rows[0][49].split(';')) / num_columns2)

        global entry_widths2
        entry_widths2 = [int(6 * w_ratio), int(169 * w_ratio)]
        global entries2
        entries2 = [[None for _ in range(num_columns2)] for _ in range(num_rows2)]

        selfplus2 = 0
        for row in range(num_rows2):
            row_frame2 = tk.Frame(f6)
            row_frame2.pack(side=tk.TOP, fill=tk.X)

            for col in range(num_columns2):
                entry2 = tk.Entry(row_frame2, width=entry_widths2[col], relief='solid', font=("ABBvoice CNSG", int(11 * h_ratio)))
                entry2.pack(side=tk.LEFT)
                if col == 0:
                    entry2.config(justify='center')
                    entry2.insert(0, '%s' % str(row + 1))
                    entry2['state'] = 'disabled'
                else:
                    entry2.insert(0, selected_rows[0][49].split(';')[1 + selfplus2])
                    selfplus2 += 2

                entries2[row][col] = entry2

        button_confirm_change_single = tk.Button(master=f7, bg="#eaf1f6", text='确认修改', command=lambda: confirm_change_multi(edit_window), font=("ABBvoice CNSG", int(11 * h_ratio)), activebackground='blue')
        button_confirm_change_single.pack(side=tk.TOP, pady=int(20*h_ratio))

        tk.Frame(f7, height=int(80 * h_ratio), bg="#eaf1f6").pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        edit_text_design_type['state'] = 'disabled'
        edit_text_project_number['state'] = 'disabled'
        # edit_text_bom_time['state'] = 'disabled'
        # edit_text_check_time['state'] = 'disabled'
        edit_text_extra_period['state'] = 'disabled'
        edit_text_item['state'] = 'disabled'
        edit_text_standard_period['state'] = 'disabled'
        # edit_text_upload_time['state'] = 'disabled'
        edit_text_create_time['state'] = 'disabled'

        edit_text_design_type['background'] = '#eaf1f6'
        edit_text_project_number['background'] = '#eaf1f6'
        # edit_text_bom_time['background'] = '#eaf1f6'
        # edit_text_check_time['background'] = '#eaf1f6'
        edit_text_extra_period['background'] = '#eaf1f6'
        edit_text_item['background'] = '#eaf1f6'
        edit_text_standard_period['background'] = '#eaf1f6'
        # edit_text_upload_time['background'] = '#eaf1f6'
        edit_text_create_time['background'] = '#eaf1f6'

        edit_canvas.update_idletasks()
        edit_canvas.config(scrollregion=edit_canvas.bbox('all'))
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def select_upload_time(edit_content, edit_text_upload_time):
    edit_text_upload_time['state'] = 'normal'
    top = tk.Toplevel(edit_content)
    cal = CCalendar(top, selectmode="day", date_pattern="yyyy-mm-dd")
    cal.pack()

    def get_date():
        date = cal.get_date()
        edit_text_upload_time.delete(1.0, tk.END)
        edit_text_upload_time.insert(tk.END, date)
        top.destroy()
        edit_text_upload_time['state'] = 'disabled'

    button = tk.Button(top, text="确认", command=get_date)
    button.pack()


def select_bom_time(edit_content, edit_text_bom_time):
    edit_text_bom_time['state'] = 'normal'
    top = tk.Toplevel(edit_content)
    cal = CCalendar(top, selectmode="day", date_pattern="yyyy-mm-dd")
    cal.pack()

    def get_date():
        date = cal.get_date()
        edit_text_bom_time.delete(1.0, tk.END)
        edit_text_bom_time.insert(tk.END, date)
        top.destroy()
        edit_text_bom_time['state'] = 'disabled'

    button = tk.Button(top, text="确认", command=get_date)
    button.pack()


def select_check_time(edit_content, edit_text_check_time):
    edit_text_check_time['state'] = 'normal'
    top = tk.Toplevel(edit_content)
    cal = CCalendar(top, selectmode="day", date_pattern="yyyy-mm-dd")
    cal.pack()

    def get_date():
        date = cal.get_date()
        edit_text_check_time.delete(1.0, tk.END)
        edit_text_check_time.insert(tk.END, date)
        top.destroy()
        edit_text_check_time['state'] = 'disabled'

    button = tk.Button(top, text="确认", command=get_date)
    button.pack()


def select_receive_time(edit_content, edit_text_receive_time):
    edit_text_receive_time['state'] = 'normal'
    top = tk.Toplevel(edit_content)
    cal = CCalendar(top, selectmode="day", date_pattern="yyyy-mm-dd")
    cal.pack()

    def get_date():
        date = cal.get_date()
        edit_text_receive_time.delete(1.0, tk.END)
        edit_text_receive_time.insert(tk.END, date)
        top.destroy()
        edit_text_receive_time['state'] = 'disabled'

    button = tk.Button(top, text="确认", command=get_date)
    button.pack()


def select_start_time(edit_content, edit_text_start_time):
    edit_text_start_time['state'] = 'normal'
    top = tk.Toplevel(edit_content)
    cal = CCalendar(top, selectmode="day", date_pattern="yyyy-mm-dd")
    cal.pack()

    def get_date():
        date = cal.get_date()
        edit_text_start_time.delete(1.0, tk.END)
        edit_text_start_time.insert(tk.END, date)
        top.destroy()
        edit_text_start_time['state'] = 'disabled'

    button = tk.Button(top, text="确认", command=get_date)
    button.pack()


def select_estimated_time(edit_content, edit_text_estimated_time):
    edit_text_estimated_time['state'] = 'normal'
    top = tk.Toplevel(edit_content)
    cal = CCalendar(top, selectmode="day", date_pattern="yyyy-mm-dd")
    cal.pack()

    def get_date():
        date = cal.get_date()
        edit_text_estimated_time.delete(1.0, tk.END)
        edit_text_estimated_time.insert(tk.END, date)
        top.destroy()
        edit_text_estimated_time['state'] = 'disabled'

    button = tk.Button(top, text="确认", command=get_date)
    button.pack()


def select_actual_time(edit_content, edit_text_actual_time):
    edit_text_actual_time['state'] = 'normal'
    top = tk.Toplevel(edit_content)
    cal = CCalendar(top, selectmode="day", date_pattern="yyyy-mm-dd")
    cal.pack()

    def get_date():
        date = cal.get_date()
        edit_text_actual_time.delete(1.0, tk.END)
        edit_text_actual_time.insert(tk.END, date)
        top.destroy()
        edit_text_actual_time['state'] = 'disabled'

    button = tk.Button(top, text="确认", command=get_date)
    button.pack()


def confirm_change_single(edit_window, selected_rows):
    try:
        result = tk.messagebox.askquestion("提示", "确定要将项目信息写入数据库吗？")
        if result == 'yes':
            if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db'):
                tk.messagebox.showwarning("提示", "数据库不存在，请联系管理员")
            else:
                conn = sqlite3.connect('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db')
                cursor = conn.cursor()  # 创建一个Cursor
                # 创建project_data表
                cursor.execute('''CREATE TABLE IF NOT EXISTS project_data
                             (id INTEGER PRIMARY KEY AUTOINCREMENT,
                              design_type TEXT,
                              project_number TEXT,
                              item_info INTEGER,
                              project_name TEXT,
                              typical_amount INTEGER,
                              panel_amount INTEGER,
                              panel_type TEXT,
                              product_type TEXT,
                              PE TEXT,
                              DE TEXT,
                              intelligence TEXT,
                              main_bus_current INTEGER,
                              main_protection TEXT,
                              arc_light TEXT,
                              location_info TEXT,
                              drawing_language TEXT,
                              standard_time INTEGER,
                              extra_time INTEGER,
                              advice_times INTEGER,
                              upload_time TEXT,
                              bom_time TEXT,
                              check_time TEXT,
                              receive_time TEXT,
                              start_time TEXT,
                              estimated_time TEXT,
                              status_info TEXT,
                              actual_time TEXT,
                              team_info TEXT,
                              abnormal_status TEXT,
                              create_time TIMESTAMP,
                              project_type TEXT,
                              frame_type TEXT,
                              program_type TEXT,
                              language_info TEXT,
                              management_info TEXT,
                              input_file TEXT,
                              drawing_type TEXT,
                              mcb_type TEXT,
                              aux_type TEXT,
                              terminal_type TEXT,
                              charged_display_type TEXT,
                              switch_type TEXT,
                              special_panel_config TEXT,
                              remark TEXT,
                              reviser TEXT,
                              revise_time TIMESTAMP,
                              ct_type TEXT,
                              pt_type TEXT,
                              sa_type TEXT,
                              customer_wiring TEXT,
                              customer_requirements TEXT)''')

                data = {
                    'design_type': edit_text_design_type.get("1.0", "end-1c"),
                    'project_number': edit_text_project_number.get("1.0", "end-1c"),
                    'item_info': edit_text_item.get("1.0", "end-1c"),
                }
                query = "SELECT * FROM project_data WHERE design_type=? AND project_number=? AND item_info=? AND id=?"
                cursor.execute(query, (data['design_type'], data['project_number'], data['item_info'], selected_rows[0][0]))
                result = cursor.fetchone()
                # print(result)
                chinese_username = ''
                if os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\Administrators.xlsx"):
                    workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\Administrators.xlsx")
                    worksheet = workbook['Sheet1']

                    for i in range(2, worksheet.max_row + 1):
                        if os.getlogin() == worksheet.cell(row=i, column=1).value:
                            chinese_username = worksheet.cell(row=i, column=2).value

                content_table1 = ''
                for i in range(0, len(entries1)):
                    for j in range(num_columns1):
                        content_table1 += entries1[i][j].get() + ';'

                content_table2 = ''
                for i in range(0, len(entries2)):
                    for j in range(num_columns2):
                        content_table2 += entries2[i][j].get() + ';'

                if result:    # 数据已存在，执行更新操作
                    if selected_rows[0][1] != '提前设计':
                        update_data = {
                            'project_name': edit_text_project_name.get("1.0", "end-1c"),
                            'typical_amount': int(edit_text_typical_amount.get('1.0', 'end-1c')) if edit_text_typical_amount.get('1.0', 'end-1c') else None,
                            'panel_amount': int(edit_text_panel_amount.get('1.0', 'end-1c')) if edit_text_panel_amount.get('1.0', 'end-1c') else None,
                            'panel_type': edit_text_typical_type.get("1.0", "end-1c"),
                            'product_type': edit_combobox_product_type.get(),
                            'PE': edit_combobox_PE.get(),
                            'DE': edit_combobox_DE.get(),
                            'main_bus_current': int(edit_combobox_main_busbar_current.get().replace('A', '')) if edit_combobox_main_busbar_current.get() else None,
                            'main_protection': edit_text_main_protection.get("1.0", "end-1c"),
                            'arc_light': edit_text_arclight.get("1.0", "end-1c"),
                            'location_info': edit_combobox_location.get(),
                            'drawing_language': edit_combobox_drawing_language.get(),
                            'upload_time': edit_text_upload_time.get("1.0", "end-1c"),
                            'bom_time': edit_text_bom_time.get("1.0", "end-1c"),
                            'check_time': edit_text_check_time.get("1.0", "end-1c"),
                            'receive_time': edit_text_receive_time.get("1.0", "end-1c"),
                            'start_time': edit_text_start_time.get("1.0", "end-1c"),
                            'estimated_time': edit_text_estimated_time.get("1.0", "end-1c"),
                            'status_info': edit_combobox_status_info.get(),
                            'actual_time': edit_text_actual_time.get("1.0", "end-1c"),
                            'team_info': edit_combobox_team_info.get(),
                            'abnormal_status': edit_text_abnormal_status.get("1.0", "end-1c"),
                            'project_type': edit_combobox_project_type.get(),
                            'frame_type': edit_combobox_frame_type.get(),
                            'program_type': edit_combobox_program_type.get(),
                            'language_info': edit_combobox_language_info.get(),
                            'management_info': edit_combobox_management_info.get(),
                            'mcb_type': edit_combobox_mcb_type.get(),
                            'aux_type': edit_combobox_aux_type.get(),
                            'terminal_type': edit_combobox_terminal_type.get(),
                            'charged_display_type': edit_combobox_charged_display_type.get(),
                            'switch_type': edit_combobox_switch_type.get(),
                            'remark': content_table2,
                            'reviser': chinese_username,
                            'revise_time': strftime('%Y-%m-%d %H:%M:%S', localtime()),
                            'ct_type': edit_combobox_ct_type.get(),
                            'pt_type': edit_combobox_pt_type.get(),
                            'sa_type': edit_combobox_sa_type.get(),
                            'customer_wiring': edit_combobox_customer_wiring.get(),
                            'customer_requirements': edit_combobox_customer_requirements.get(),
                        }
                    else:
                        update_data = {
                            'project_name': edit_text_project_name.get("1.0", "end-1c"),
                            'typical_amount': int(edit_text_typical_amount.get('1.0', 'end-1c')) if edit_text_typical_amount.get('1.0', 'end-1c') else None,
                            'panel_amount': int(edit_text_panel_amount.get('1.0', 'end-1c')) if edit_text_panel_amount.get('1.0', 'end-1c') else None,
                            'panel_type': edit_text_typical_type.get("1.0", "end-1c"),
                            'product_type': edit_combobox_product_type.get(),
                            'PE': edit_combobox_PE.get(),
                            'DE': edit_combobox_DE.get(),
                            'main_bus_current': int(edit_combobox_main_busbar_current.get().replace('A', '')) if edit_combobox_main_busbar_current.get() else None,
                            'main_protection': edit_text_main_protection.get("1.0", "end-1c"),
                            'arc_light': edit_text_arclight.get("1.0", "end-1c"),
                            'location_info': edit_combobox_location.get(),
                            'drawing_language': edit_combobox_drawing_language.get(),
                            'upload_time': edit_text_upload_time.get("1.0", "end-1c"),
                            'receive_time': edit_text_receive_time.get("1.0", "end-1c"),
                            'start_time': edit_text_start_time.get("1.0", "end-1c"),
                            'estimated_time': edit_text_estimated_time.get("1.0", "end-1c"),
                            'status_info': edit_combobox_status_info.get(),
                            'actual_time': edit_text_actual_time.get("1.0", "end-1c"),
                            'team_info': edit_combobox_team_info.get(),
                            'abnormal_status': edit_text_abnormal_status.get("1.0", "end-1c"),
                            'project_type': edit_combobox_project_type.get(),
                            'frame_type': edit_combobox_frame_type.get(),
                            'program_type': edit_combobox_program_type.get(),
                            'language_info': edit_combobox_language_info.get(),
                            'management_info': edit_combobox_management_info.get(),
                            'mcb_type': edit_combobox_mcb_type.get(),
                            'aux_type': edit_combobox_aux_type.get(),
                            'terminal_type': edit_combobox_terminal_type.get(),
                            'charged_display_type': edit_combobox_charged_display_type.get(),
                            'switch_type': edit_combobox_switch_type.get(),
                            'advice_times': edit_combobox_advice_times.get(),
                            'intelligence': edit_text_intelligence.get("1.0", "end-1c").split("\n")[0],
                            'input_file': edit_text_input_file.get("1.0", "end-1c").split("\n")[0],
                            'drawing_type': edit_text_drawing_type.get("1.0", "end-1c").split("\n")[0],
                            'special_panel_config': content_table1,
                            'remark': content_table2,
                            'reviser': chinese_username,
                            'revise_time': strftime('%Y-%m-%d %H:%M:%S', localtime()),
                            'ct_type': edit_combobox_ct_type.get(),
                            'pt_type': edit_combobox_pt_type.get(),
                            'sa_type': edit_combobox_sa_type.get(),
                            'customer_wiring': edit_combobox_customer_wiring.get(),
                            'customer_requirements': edit_combobox_customer_requirements.get(),
                        }

                    # 过滤出非空值
                    update_values = [v for v in update_data.values() if v is not None]
                    update_columns = ', '.join([f"{col} = ?" for col in update_data.keys() if update_data[col] is not None])

                    # 判断非空值是否存在
                    if update_values:
                        update_query = f"UPDATE project_data SET {update_columns} WHERE design_type=? AND project_number=? AND item_info=? AND id=?"
                        update_values.append(data['design_type'])
                        update_values.append(data['project_number'])
                        update_values.append(data['item_info'])
                        update_values.append(selected_rows[0][0])
                        cursor.execute(update_query, tuple(update_values))
                        conn.commit()
                        tk.messagebox.showwarning("提示", "数据更新完成")

                    cursor.close()
                    conn.close()

                    source_file = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db'
                    target_dir = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\backup_db'
                    target_file = os.path.join(target_dir, os.path.basename(source_file))

                    if os.path.exists(target_file):
                        os.remove(target_file)

                    shutil.copy2(source_file, target_dir)

                    edit_window.destroy()
                    query_all()
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def confirm_change_multi(edit_window):
    try:
        if ';' in edit_combobox_advice_times.get():
            tk.messagebox.showwarning("提示", "请修改意见次数")
        else:
            result = tk.messagebox.askquestion("提示", "确定要将项目信息写入数据库吗？")
            if result == 'yes':
                if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db'):
                    tk.messagebox.showwarning("提示", "数据库不存在，请联系管理员")
                else:
                    conn = sqlite3.connect('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db')
                    cursor = conn.cursor()  # 创建一个Cursor
                    # 创建project_data表
                    cursor.execute('''CREATE TABLE IF NOT EXISTS project_data
                                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                                  design_type TEXT,
                                  project_number TEXT,
                                  item_info INTEGER,
                                  project_name TEXT,
                                  typical_amount INTEGER,
                                  panel_amount INTEGER,
                                  panel_type TEXT,
                                  product_type TEXT,
                                  PE TEXT,
                                  DE TEXT,
                                  intelligence TEXT,
                                  main_bus_current INTEGER,
                                  main_protection TEXT,
                                  arc_light TEXT,
                                  location_info TEXT,
                                  drawing_language TEXT,
                                  standard_time INTEGER,
                                  extra_time INTEGER,
                                  advice_times INTEGER,
                                  upload_time TEXT,
                                  bom_time TEXT,
                                  check_time TEXT,
                                  receive_time TEXT,
                                  start_time TEXT,
                                  estimated_time TEXT,
                                  status_info TEXT,
                                  actual_time TEXT,
                                  team_info TEXT,
                                  abnormal_status TEXT,
                                  create_time TIMESTAMP,
                                  project_type TEXT,
                                  frame_type TEXT,
                                  program_type TEXT,
                                  language_info TEXT,
                                  management_info TEXT,
                                  input_file TEXT,
                                  drawing_type TEXT,
                                  mcb_type TEXT,
                                  aux_type TEXT,
                                  terminal_type TEXT,
                                  charged_display_type TEXT,
                                  switch_type TEXT,
                                  special_panel_config TEXT,
                                  remark TEXT,
                                  reviser TEXT,
                                  revise_time TIMESTAMP,
                                  ct_type TEXT,
                                  pt_type TEXT,
                                  sa_type TEXT,
                                  customer_wiring TEXT,
                                  customer_requirements TEXT)''')

                    data = {
                        'design_type': edit_text_design_type.get("1.0", "end-1c"),
                        'project_number': edit_text_project_number.get("1.0", "end-1c"),
                    }
                    item_info_lines = edit_text_item.get("1.0", "end-1c").split('\n')
                    result_list = []
                    for item_info in item_info_lines:
                        query = "SELECT * FROM project_data WHERE design_type=? AND project_number=? AND item_info=?"
                        cursor.execute(query, (data['design_type'], data['project_number'], item_info))
                        result = cursor.fetchone()
                        result_list.append(result)

                    chinese_username = ''
                    if os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\Administrators.xlsx"):
                        workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\Administrators.xlsx")
                        worksheet = workbook['Sheet1']

                        for i in range(2, worksheet.max_row + 1):
                            if os.getlogin() == worksheet.cell(row=i, column=1).value:
                                chinese_username = worksheet.cell(row=i, column=2).value

                    content_table1 = ''
                    for i in range(0, len(entries1)):
                        for j in range(num_columns1):
                            content_table1 += entries1[i][j].get() + ';'

                    content_table2 = ''
                    for i in range(0, len(entries2)):
                        for j in range(num_columns2):
                            content_table2 += entries2[i][j].get() + ';'

                    flag_k = 0
                    for index, item in enumerate(edit_text_item.get("1.0", "end-1c").split("\n")):
                        if result_list[index]:
                            update_data = {
                                'project_name': edit_text_project_name.get("1.0", "end-1c").split("\n")[index],
                                'typical_amount': int(edit_text_typical_amount.get('1.0', 'end-1c').split("\n")[index]) if edit_text_typical_amount.get('1.0', 'end-1c').split("\n")[index] else None,
                                'panel_amount': int(edit_text_panel_amount.get('1.0', 'end-1c').split("\n")[index]) if edit_text_panel_amount.get('1.0', 'end-1c').split("\n")[index] else None,
                                'panel_type': edit_text_typical_type.get("1.0", "end-1c").split("\n")[index],
                                'product_type': edit_combobox_product_type_list[index].get(),
                                'PE': edit_combobox_PE_list[index].get(),
                                'DE': edit_combobox_DE_list[index].get(),
                                'main_bus_current': int(edit_combobox_main_busbar_current_list[index].get().replace('A', '')) if edit_combobox_main_busbar_current_list[index].get() else None,
                                'main_protection': edit_text_main_protection.get("1.0", "end-1c").split("\n")[index],
                                'arc_light': edit_text_arclight.get("1.0", "end-1c").split("\n")[index],
                                'location_info': edit_combobox_location_list[index].get(),
                                'drawing_language': edit_combobox_drawing_language_list[index].get(),
                                'upload_time': edit_text_upload_time_list[index].get("1.0", "end-1c"),
                                'bom_time': edit_text_bom_time_list[index].get("1.0", "end-1c"),
                                'check_time': edit_text_check_time_list[index].get("1.0", "end-1c"),
                                'receive_time': edit_text_receive_time_list[index].get("1.0", "end-1c"),
                                'start_time': edit_text_start_time_list[index].get("1.0", "end-1c"),
                                'estimated_time': edit_text_estimated_time_list[index].get("1.0", "end-1c"),
                                'status_info': edit_combobox_status_info_list[index].get(),
                                'actual_time': edit_text_actual_time_list[index].get("1.0", "end-1c"),
                                'team_info': edit_combobox_team_info_list[index].get(),
                                'abnormal_status': edit_text_abnormal_status.get("1.0", "end-1c").split("\n")[index],
                                'project_type': edit_combobox_project_type_list[index].get(),
                                'frame_type': edit_combobox_frame_type_list[index].get(),
                                'program_type': edit_combobox_program_type_list[index].get(),
                                'language_info': edit_combobox_language_info_list[index].get(),
                                'management_info': edit_combobox_management_info_list[index].get(),
                                'mcb_type': edit_combobox_mcb_type_list[index].get(),
                                'aux_type': edit_combobox_aux_type_list[index].get(),
                                'terminal_type': edit_combobox_terminal_type_list[index].get(),
                                'charged_display_type': edit_combobox_charged_display_type_list[index].get(),
                                'switch_type': edit_combobox_switch_type_list[index].get(),
                                'advice_times': edit_combobox_advice_times.get(),
                                'intelligence': edit_text_intelligence.get("1.0", "end-1c").split("\n")[0],
                                'input_file': edit_text_input_file.get("1.0", "end-1c").split("\n")[0],
                                'drawing_type': edit_text_drawing_type.get("1.0", "end-1c").split("\n")[0],
                                'special_panel_config': content_table1,
                                'remark': content_table2,
                                'reviser': chinese_username,
                                'revise_time': strftime('%Y-%m-%d %H:%M:%S', localtime()),
                                'ct_type': edit_combobox_ct_type_list[index].get(),
                                'pt_type': edit_combobox_pt_type_list[index].get(),
                                'sa_type': edit_combobox_sa_type_list[index].get(),
                                'customer_wiring': edit_combobox_customer_wiring_list[index].get(),
                                'customer_requirements': edit_combobox_customer_requirements_list[index].get(),
                            }
                            # print(index, update_data)

                            # 过滤出非空值
                            update_values = [v for v in update_data.values() if v is not None]
                            update_columns = ', '.join([f"{col} = ?" for col in update_data.keys() if update_data[col] is not None])

                            # 判断非空值是否存在
                            if update_values:
                                update_query = f"UPDATE project_data SET {update_columns} WHERE design_type=? AND project_number=? AND item_info=?"
                                update_values.append(data['design_type'])
                                update_values.append(data['project_number'])
                                update_values.append(item)
                                cursor.execute(update_query, tuple(update_values))
                                conn.commit()
                                flag_k += 1

                    create_file = 0
                    # 如下三行在V2.0版本注释掉，以确保设计传递表填写后就能创建相应文件夹
                    # for i in range(0, len(edit_text_start_time_list)):
                        # if edit_text_start_time_list[i].get("1.0", "end-1c") is not None:
                        #     create_file = 1
                    create_file = 1

                    if create_file:
                        if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc'):
                            tk.messagebox.showwarning("提示", "请连接内网，确保能够读取J盘\n\n否则无法创建/更新图纸意见文件夹")
                        if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc') and os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools'):
                            os.mkdir('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc')

                        if edit_combobox_advice_times.get() is not None:
                            if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % edit_text_project_number.get("1.0", "end-1c")) and os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc'):
                                os.mkdir('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % edit_text_project_number.get("1.0", "end-1c"))
                            if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s\\Technical_Clarification' % edit_text_project_number.get("1.0", "end-1c")) and os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % edit_text_project_number.get("1.0", "end-1c")):
                                os.mkdir('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s\\Technical_Clarification' % edit_text_project_number.get("1.0", "end-1c"))
                            if os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % edit_text_project_number.get("1.0", "end-1c")):
                                file_names = os.listdir('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % edit_text_project_number.get("1.0", "end-1c"))
                                if int(edit_combobox_advice_times.get()) > len(file_names)-1:
                                    if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s\\Drawing_Comments_1' % edit_text_project_number.get("1.0", "end-1c")) and os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % edit_text_project_number.get("1.0", "end-1c")):
                                        os.mkdir('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s\\Drawing_Comments_1' % edit_text_project_number.get("1.0", "end-1c"))
                                    file_names = os.listdir('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % edit_text_project_number.get("1.0", "end-1c"))
                                    if int(edit_combobox_advice_times.get()) > len(file_names) - 1:
                                        if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s\\Drawing_Comments_2' % edit_text_project_number.get("1.0", "end-1c")) and os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % edit_text_project_number.get("1.0", "end-1c")):
                                            os.mkdir('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s\\Drawing_Comments_2' % edit_text_project_number.get("1.0", "end-1c"))
                                        file_names = os.listdir('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % edit_text_project_number.get("1.0", "end-1c"))
                                        if int(edit_combobox_advice_times.get()) > len(file_names) - 1:
                                            if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s\\Drawing_Comments_3' % edit_text_project_number.get("1.0", "end-1c")) and os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % edit_text_project_number.get("1.0", "end-1c")):
                                                os.mkdir('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s\\Drawing_Comments_3' % edit_text_project_number.get("1.0", "end-1c"))
                                            file_names = os.listdir('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % edit_text_project_number.get("1.0", "end-1c"))
                                            if int(edit_combobox_advice_times.get()) > len(file_names) - 1:
                                                if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s\\Drawing_Comments_4' % edit_text_project_number.get("1.0", "end-1c")) and os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % edit_text_project_number.get("1.0", "end-1c")):
                                                    os.mkdir('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s\\Drawing_Comments_4' % edit_text_project_number.get("1.0", "end-1c"))
                                                file_names = os.listdir('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % edit_text_project_number.get("1.0", "end-1c"))
                                                if int(edit_combobox_advice_times.get()) > len(file_names) - 1:
                                                    if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s\\Drawing_Comments_5' % edit_text_project_number.get("1.0", "end-1c")) and os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % edit_text_project_number.get("1.0", "end-1c")):
                                                        os.mkdir('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s\\Drawing_Comments_5' % edit_text_project_number.get("1.0", "end-1c"))
                                                    file_names = os.listdir('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % edit_text_project_number.get("1.0", "end-1c"))
                                                    if int(edit_combobox_advice_times.get()) > len(file_names) - 1:
                                                        if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s\\Drawing_Comments_6' % edit_text_project_number.get("1.0", "end-1c")) and os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s' % edit_text_project_number.get("1.0", "end-1c")):
                                                            os.mkdir('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pc\\%s\\Drawing_Comments_6' % edit_text_project_number.get("1.0", "end-1c"))

                    if flag_k:
                        tk.messagebox.showwarning("提示", "数据更新完成，成功数/总数：%d/%d" % (flag_k, len(edit_text_item.get("1.0", "end-1c").split("\n"))))
                        cursor.close()
                        conn.close()

                        source_file = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db'
                        target_dir = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\backup_db'
                        target_file = os.path.join(target_dir, os.path.basename(source_file))

                        if os.path.exists(target_file):
                            os.remove(target_file)

                        shutil.copy2(source_file, target_dir)

                        edit_window.destroy()
                        query_all()
                    else:
                        tk.messagebox.showwarning("提示", "数据更新失败")
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def select_input_file_single(edit_text_input_file, input_files):
    true_checkbox_input_list = ''
    for i in range(0, len(input_files)):
        if checkbox_input_list[i].get():
            true_checkbox_input_list += input_files[i] + ';'
    edit_text_input_file.delete(1.0, tk.END)
    edit_text_input_file.insert(tk.END, true_checkbox_input_list.replace(' ', ''))


def select_drawing_type_single(edit_text_drawing_type, drawing_types):
    true_checkbox_drawing_type = ''
    for i in range(0, len(drawing_types)):
        if checkbox_drawing_type_list[i].get():
            true_checkbox_drawing_type += drawing_types[i] + ';'
    edit_text_drawing_type.delete(1.0, tk.END)
    edit_text_drawing_type.insert(tk.END, true_checkbox_drawing_type.replace(' ', ''))


def select_intelligence_single(edit_text_intelligence, intelligence_needs):
    true_checkbox_intelligence = ''
    for i in range(0, len(intelligence_needs)):
        if checkbox_intelligence_list[i].get():
            true_checkbox_intelligence += intelligence_needs[i] + ';'
    edit_text_intelligence.delete(1.0, tk.END)
    edit_text_intelligence.insert(tk.END, true_checkbox_intelligence.replace(' ', ''))


def select_input_file_multi(edit_text_input_file, number):
    true_checkbox_input_list = ''
    for i in range(0, len(input_files)):
        if checkbox_input_list[i].get():
            true_checkbox_input_list += input_files[i] + ';'
    edit_text_input_file.delete(1.0, tk.END)
    for i in range(0, number):
        edit_text_input_file.insert(tk.END, true_checkbox_input_list.replace(' ', '')+'\n')


def select_drawing_type_multi(edit_text_drawing_type, number):
    true_checkbox_drawing_type = ''
    for i in range(0, len(drawing_types)):
        if checkbox_drawing_type_list[i].get():
            true_checkbox_drawing_type += drawing_types[i] + ';'
    edit_text_drawing_type.delete(1.0, tk.END)
    for i in range(0, number):
        edit_text_drawing_type.insert(tk.END, true_checkbox_drawing_type.replace(' ', '')+'\n')


def select_intelligence_multi(edit_text_intelligence, number):
    true_checkbox_intelligence = ''
    for i in range(0, len(intelligence_needs)):
        if checkbox_intelligence_list[i].get():
            true_checkbox_intelligence += intelligence_needs[i] + ';'
    edit_text_intelligence.delete(1.0, tk.END)
    for i in range(0, number):
        edit_text_intelligence.insert(tk.END, true_checkbox_intelligence.replace(' ', '')+'\n')


def project_ahead2other(parent, w_ratio, h_ratio, selected_rows):
    try:
        edit_window = tk.Toplevel(parent)
        edit_window.grab_set()  # 禁用parent窗口的操作

        global edit_canvas

        edit_canvas = tk.Canvas(edit_window, width=int(1000 * w_ratio), height=int(500 * h_ratio), bg="#C9DBE9", borderwidth=0, highlightthickness=0)
        edit_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        edit_canvas.delete('all')  # 清空画布
        # edit_canvas.update()
        # edit_canvas.bind("<MouseWheel>", on_edit_mousewheel)
        #
        # scrollbar_v = tk.Scrollbar(master=edit_window)
        # scrollbar_v.pack(side=tk.RIGHT, fill=tk.Y)
        # scrollbar_v.config(command=edit_canvas.yview)
        # edit_canvas.config(yscrollcommand=scrollbar_v.set)

        edit_content = tk.Frame(edit_canvas, bg="#eaf1f6")

        edit_canvas.create_window(0, 0, width=int(1000 * w_ratio), anchor=tk.NW, window=edit_content)

        global ahead_typical_amount
        ahead_typical_amount = selected_rows[0][5]
        global ahead_PE
        ahead_PE = selected_rows[0][9]
        global ahead_DE
        ahead_DE = selected_rows[0][10]
        global ahead_main_busbar_current
        ahead_main_busbar_current = selected_rows[0][12]
        global ahead_main_protection
        ahead_main_protection = selected_rows[0][13]
        global ahead_arclight
        ahead_arclight = selected_rows[0][14]
        global ahead_location
        ahead_location = selected_rows[0][15]
        global ahead_drawing_language
        ahead_drawing_language = selected_rows[0][16]
        global ahead_standard_period
        ahead_standard_period = selected_rows[0][17]
        global ahead_extra_period
        ahead_extra_period = selected_rows[0][18]
        global ahead_mcb_type
        ahead_mcb_type = selected_rows[0][38]
        global ahead_aux_type
        ahead_aux_type = selected_rows[0][39]
        global ahead_terminal_type
        ahead_terminal_type = selected_rows[0][40]
        global ahead_charged_display_type
        ahead_charged_display_type = selected_rows[0][41]
        global ahead_switch_type
        ahead_switch_type = selected_rows[0][42]
        global ahead_ct_type
        ahead_ct_type = selected_rows[0][43]
        global ahead_pt_type
        ahead_pt_type = selected_rows[0][44]
        global ahead_sa_type
        ahead_sa_type = selected_rows[0][45]
        global ahead_customer_wiring
        ahead_customer_wiring = selected_rows[0][46]
        global ahead_customer_requirements
        ahead_customer_requirements = selected_rows[0][47]

        global ahead_upload_time
        ahead_upload_time = selected_rows[0][20]
        global ahead_bom_time
        ahead_bom_time = selected_rows[0][21]
        global ahead_check_time
        ahead_check_time = selected_rows[0][22]
        global ahead_receive_time
        ahead_receive_time = selected_rows[0][23]
        global ahead_start_time
        ahead_start_time = selected_rows[0][24]
        global ahead_estimated_time
        ahead_estimated_time = selected_rows[0][25]
        global ahead_status_info
        ahead_status_info = selected_rows[0][26]
        global ahead_actual_time
        ahead_actual_time = selected_rows[0][27]
        global ahead_team_info
        ahead_team_info = selected_rows[0][28]
        global ahead_abnormal_status
        ahead_abnormal_status = selected_rows[0][29]
        global ahead_create_time
        ahead_create_time = selected_rows[0][30]
        global ahead_project_type
        ahead_project_type = selected_rows[0][31]
        global ahead_frame_type
        ahead_frame_type = selected_rows[0][32]
        global ahead_program_type
        ahead_program_type = selected_rows[0][33]
        global ahead_language_info
        ahead_language_info = selected_rows[0][34]
        global ahead_management_info
        ahead_management_info = selected_rows[0][35]
        global ahead_input_file
        ahead_input_file = selected_rows[0][36]
        global ahead_drawing_type
        ahead_drawing_type = selected_rows[0][37]
        global ahead_intelligence
        ahead_intelligence = selected_rows[0][11]
        global ahead_entries1
        ahead_entries1 = selected_rows[0][48]
        global ahead_entries2
        ahead_entries2 = selected_rows[0][49]
        global ahead_advice_times
        ahead_advice_times = selected_rows[0][19]

        f0 = tk.Frame(edit_content, bg="#c9dbe9", bd=0)
        tk.Label(f0, text="提前设计→图纸/工程设计", bg="#c9dbe9", fg="black", height=int(1 * h_ratio), font=("ABBvoice CNSG", int(20 * h_ratio), "bold")).pack(fill=tk.X)
        # tk.Label(f0, text="(黑色属性无法修改，绿色属性可修改)", bg="#c9dbe9", fg="black", height=int(1 * h_ratio), font=("ABBvoice CNSG", int(12 * h_ratio), "bold")).pack(fill=tk.X)
        f0.pack(side=tk.TOP, fill=tk.X)

        f1 = tk.Frame(edit_content, bg="#eaf1f6")
        f1.pack(side=tk.TOP)

        f_1 = tk.Frame(f1, bg="#eaf1f6")
        global ahead_combobox_design_type
        tk.Label(f_1, text="设计类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        design_type_list = ['图纸设计', '工程设计']
        ahead_combobox_design_type = ttk.Combobox(f_1, font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(12 * w_ratio), state='readonly', values=design_type_list)
        ahead_combobox_design_type.pack(side=tk.LEFT)
        ahead_combobox_design_type.set(selected_rows[0][1])
        ahead_combobox_design_type.bind("<<ComboboxSelected>>", ahead_on_select)
        f_1.pack(fill=tk.X)

        tk.Frame(f1, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_2 = tk.Frame(f1, bg="#eaf1f6")
        tk.Label(f_2, text="   项目号", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global ahead_entry_project_number
        ahead_entry_project_number = tk.Entry(f_2, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(20 * w_ratio))
        ahead_entry_project_number.pack(side=tk.LEFT)
        ahead_entry_project_number.insert(tk.END, selected_rows[0][2])
        ahead_entry_project_number['state'] = 'disabled'
        ahead_entry_project_number['background'] = '#eaf1f6'
        ahead_entry_project_number.bind('<Return>', query_project)

        global button_query_project
        button_query_project = tk.Button(master=f_2, bg="#eaf1f6", text='查询', image=query_project1, compound=tk.LEFT, command=query_project, font=("ABBvoice CNSG", int(11 * h_ratio)), activebackground='blue')
        button_query_project.pack(side=tk.LEFT, padx=int(10 * w_ratio))
        button_query_project['state'] = 'disabled'

        # tk.Label(f_2, text='  (需要连接内网)', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(10 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
        f_2.pack(fill=tk.X)

        tk.Frame(f1, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_3 = tk.Frame(f1, bg="#eaf1f6")
        tk.Label(f_3, text="   项目名", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global ahead_entry_project_name
        ahead_entry_project_name = tk.Entry(f_3, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), width=int(100 * w_ratio))
        ahead_entry_project_name.pack(side=tk.LEFT)
        ahead_entry_project_name.insert(tk.END, selected_rows[0][4])
        ahead_entry_project_name['state'] = 'disabled'
        ahead_entry_project_name['background'] = '#eaf1f6'
        f_3.pack(fill=tk.X)

        tk.Frame(f1, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_4 = tk.Frame(f1, bg="#eaf1f6")
        tk.Label(f_4, text="      行号", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global ahead_text_item
        ahead_text_item = tk.Text(f_4, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=2, width=int(100 * w_ratio))
        ahead_text_item.pack(side=tk.LEFT)
        ahead_text_item.insert(tk.END, selected_rows[0][3])
        ahead_text_item['state'] = 'disabled'
        ahead_text_item['background'] = '#eaf1f6'

        global button_update_amount
        button_update_amount = tk.Button(f_4, text='刷新', font=("ABBvoice CNSG", int(11 * h_ratio)), bg="#eaf1f6", image=refresh_project, compound=tk.LEFT, command=update_amount, activebackground='blue')
        button_update_amount.pack(side=tk.LEFT, padx=int(10 * w_ratio))
        button_update_amount['state'] = 'disabled'

        f_4.pack(fill=tk.X)

        tk.Frame(f1, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_6 = tk.Frame(f1, bg="#eaf1f6")
        tk.Label(f_6, text="      台数", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global ahead_text_panel_amount
        ahead_text_panel_amount = tk.Text(f_6, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=2, width=int(100 * w_ratio))
        ahead_text_panel_amount.pack(side=tk.LEFT)
        ahead_text_panel_amount.insert(tk.END, selected_rows[0][6])
        ahead_text_panel_amount['state'] = 'disabled'
        ahead_text_panel_amount['background'] = '#eaf1f6'
        f_6.pack(fill=tk.X)

        tk.Frame(f1, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_7 = tk.Frame(f1, bg="#eaf1f6")
        tk.Label(f_7, text="      柜型", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global ahead_text_typical_type
        ahead_text_typical_type = tk.Text(f_7, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=3, width=int(100 * w_ratio))
        ahead_text_typical_type.pack(side=tk.LEFT)
        ahead_text_typical_type.insert(tk.END, selected_rows[0][7])
        ahead_text_typical_type['state'] = 'disabled'
        ahead_text_typical_type['background'] = '#eaf1f6'
        f_7.pack(fill=tk.X)

        tk.Frame(f1, height=int(5 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)

        f_8 = tk.Frame(f1, bg="#eaf1f6")
        tk.Label(f_8, text="产品类型", height=int(1 * h_ratio), bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio))).pack(side=tk.LEFT)
        global ahead_text_product_type
        ahead_text_product_type = tk.Text(f_8, bg="#eaf1f6", font=("ABBvoice CNSG", int(11 * h_ratio)), height=2, width=int(100 * w_ratio))
        ahead_text_product_type.pack(side=tk.LEFT)
        ahead_text_product_type.insert(tk.END, selected_rows[0][8])
        ahead_text_product_type['state'] = 'disabled'
        ahead_text_product_type['background'] = '#eaf1f6'
        f_8.pack(fill=tk.X)

        tk.Frame(f1, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)
        global button_confirm_change_ahead2other
        button_confirm_change_ahead2other = tk.Button(master=f1, bg="#eaf1f6", text='转换', image=confirm_trans, compound=tk.LEFT, command=lambda: confirm_change_ahead2other(edit_window, selected_rows), font=("ABBvoice CNSG", int(11 * h_ratio)), activebackground='blue')
        button_confirm_change_ahead2other.pack(side=tk.TOP, pady=int(20 * h_ratio))
        button_confirm_change_ahead2other['state'] = 'disabled'

        tk.Frame(f1, height=int(40 * h_ratio), bg="#eaf1f6").pack(side=tk.TOP, fill=tk.BOTH, expand=True)
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def ahead_on_select(event):
    if ahead_combobox_design_type.get() != '提前设计':
        button_query_project['state'] = 'normal'
        ahead_entry_project_number['state'] = 'normal'
        ahead_entry_project_number['bg'] = 'white'
        ahead_entry_project_number.delete(0, "end")
        ahead_entry_project_name['state'] = 'normal'
        ahead_entry_project_name.delete(0, "end")
        ahead_text_item['state'] = 'normal'
        ahead_text_item.delete(1.0, tk.END)
        ahead_text_panel_amount['state'] = 'normal'
        ahead_text_panel_amount.delete(1.0, tk.END)
        ahead_text_typical_type['state'] = 'normal'
        ahead_text_typical_type.delete(1.0, tk.END)
        ahead_text_product_type['state'] = 'normal'
        ahead_text_product_type.delete(1.0, tk.END)


def on_select(event):
    global result
    button_detail['state'] = 'disabled'
    button_change_single['state'] = 'disabled'
    button_change_multi['state'] = 'disabled'
    button_ahead2other['state'] = 'disabled'
    if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db'):
        tk.messagebox.showwarning("提示", "数据库不存在，请联系管理员")
    else:

        conn = sqlite3.connect('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db')
        cursor = conn.cursor()  # 创建一个Cursor

        attributes = []
        if combobox_design_type.get():
            attributes.append("design_type='" + combobox_design_type.get() + "'")

        if combobox_project_number.get():
            attributes.append("project_number='" + combobox_project_number.get() + "'")

        if combobox_project_name.get():
            attributes.append("project_name LIKE '%" + combobox_project_name.get() + "%'")

        if combobox_product_type.get():
            attributes.append("product_type='" + combobox_product_type.get() + "'")

        if combobox_PE.get():
            attributes.append("PE='" + combobox_PE.get() + "'")

        if combobox_DE.get():
            attributes.append("DE='" + combobox_DE.get() + "'")

        if combobox_status.get():
            attributes.append("status_info='" + combobox_status.get() + "'")

        if combobox_panel_type.get():
            attributes.append("panel_type='" + combobox_panel_type.get() + "'")

        if combobox_intelligent.get():
            attributes.append("intelligence='" + combobox_intelligent.get() + "'")

        if combobox_protection.get():
            attributes.append("main_protection LIKE '%" + combobox_protection.get() + "%'")

        if combobox_arclight.get():
            attributes.append("arc_light LIKE '%" + combobox_arclight.get() + "%'")

        attributes = [attr for attr in attributes if attr != '']

        if not attributes:
            tk.messagebox.showwarning("提示", "至少指定一项属性不为空")
            return

        query = "SELECT * FROM project_data WHERE "
        query += " AND ".join(attributes)
        query += "ORDER BY CASE WHEN design_type = '提前设计' THEN rowid END DESC, CASE WHEN design_type = '图纸设计' OR design_type = '工程设计' THEN project_number END DESC"

        cursor.execute(query)
        result = cursor.fetchall()

        result = [[element if element is not None else '' for element in row] for row in result]
        cursor.close()
        conn.close()

        # 清空treeview表格
        table_items = Project_Info_table.get_children()  # 在插入treeview数据时，需要先清空treeview
        [Project_Info_table.delete(table_item) for table_item in table_items]

        # 为treeview表格插入数据
        temp_list = []
        for i in range(0, len(result)):
            temp_list = list(result[i])[0:43] + list(result[i])[47:] + list(result[i])[43:47]
            Project_Info_table.insert('', 'end', values=temp_list, tags='fontsize')  # dataframe逐行插入到表格中

        design_type_list = []
        for i in range(0, len(result)):
            design_type_list.append(result[i][1])
        design_type_list_set = list(set(design_type_list))
        design_type_list_set.sort(key=list(design_type_list).index)
        # combobox_design_type.set('')
        combobox_design_type['values'] = design_type_list_set

        project_number_list = []
        for i in range(0, len(result)):
            project_number_list.append(result[i][2])
        project_number_list_set = list(set(project_number_list))
        project_number_list_set.sort(key=list(project_number_list).index)
        # combobox_project_number.set('')
        combobox_project_number['values'] = project_number_list_set

        project_name_list = []
        for i in range(0, len(result)):
            project_name_list.append(result[i][4])
        project_name_list_set = list(set(project_name_list))
        project_name_list_set.sort(key=list(project_name_list).index)
        # combobox_project_name.set('')
        combobox_project_name['values'] = project_name_list_set

        product_type_list = []
        for i in range(0, len(result)):
            product_type_list.append(result[i][8])
        product_type_list_set = list(set(product_type_list))
        product_type_list_set.sort(key=list(product_type_list).index)
        # combobox_product_type.set('')
        combobox_product_type['values'] = product_type_list_set

        PE_list = []
        for i in range(0, len(result)):
            PE_list.append(result[i][9])
        PE_list_set = list(set(PE_list))
        PE_list_set.sort(key=list(PE_list).index)
        # combobox_PE.set('')
        combobox_PE['values'] = PE_list_set

        DE_list = []
        for i in range(0, len(result)):
            DE_list.append(result[i][10])
        DE_list_set = list(set(DE_list))
        DE_list_set.sort(key=list(DE_list).index)
        # combobox_DE.set('')
        combobox_DE['values'] = DE_list_set

        status_list = []
        for i in range(0, len(result)):
            status_list.append(result[i][26])
        status_list_set = list(set(status_list))
        status_list_set.sort(key=list(status_list).index)
        # combobox_status.set('')
        combobox_status['values'] = status_list_set

        panel_type_list = []
        for i in range(0, len(result)):
            panel_type_list.append(result[i][7])
        panel_type_list_set = list(set(panel_type_list))
        panel_type_list_set.sort(key=list(panel_type_list).index)
        # combobox_panel_type.set('')
        combobox_panel_type['values'] = panel_type_list_set

        intelligent_list = []
        for i in range(0, len(result)):
            intelligent_list.append(result[i][11])
        intelligent_list_set = list(set(intelligent_list))
        intelligent_list_set.sort(key=list(intelligent_list).index)
        # combobox_intelligent.set('')
        combobox_intelligent['values'] = intelligent_list_set

        protection_list = []
        for i in range(0, len(result)):
            protection_list.append(result[i][13])
        protection_list_set = list(set(protection_list))
        protection_list_set.sort(key=list(protection_list).index)
        # combobox_protection.set('')
        combobox_protection['values'] = protection_list_set

        arclight_list = []
        for i in range(0, len(result)):
            arclight_list.append(result[i][14])
        arclight_list_set = list(set(arclight_list))
        arclight_list_set.sort(key=list(arclight_list).index)
        # combobox_arclight.set('')
        combobox_arclight['values'] = arclight_list_set


def query_some():
    global result
    button_detail['state'] = 'disabled'
    button_change_single['state'] = 'disabled'
    button_change_multi['state'] = 'disabled'
    button_ahead2other['state'] = 'disabled'
    if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db'):
        tk.messagebox.showwarning("提示", "数据库不存在，请联系管理员")
    else:

        conn = sqlite3.connect('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db')
        cursor = conn.cursor()  # 创建一个Cursor

        attributes = []
        if combobox_design_type.get():
            attributes.append("design_type='" + combobox_design_type.get() + "'")

        if combobox_project_number.get():
            attributes.append("project_number='" + combobox_project_number.get() + "'")

        if combobox_project_name.get():
            attributes.append("project_name LIKE '%" + combobox_project_name.get() + "%'")

        if combobox_product_type.get():
            attributes.append("product_type='" + combobox_product_type.get() + "'")

        if combobox_PE.get():
            attributes.append("PE='" + combobox_PE.get() + "'")

        if combobox_DE.get():
            attributes.append("DE='" + combobox_DE.get() + "'")

        if combobox_panel_type.get():
            attributes.append("panel_type='" + combobox_panel_type.get() + "'")

        if combobox_intelligent.get():
            attributes.append("intelligence='" + combobox_intelligent.get() + "'")

        if combobox_protection.get():
            attributes.append("main_protection LIKE '%" + combobox_protection.get() + "%'")

        if combobox_arclight.get():
            attributes.append("arc_light LIKE '%" + combobox_arclight.get() + "%'")

        attributes = [attr for attr in attributes if attr != '']

        if not attributes:
            tk.messagebox.showwarning("提示", "至少指定一项属性不为空")
            return

        query = "SELECT * FROM project_data WHERE "
        query += " AND ".join(attributes)
        query += "ORDER BY CASE WHEN design_type = '提前设计' THEN rowid END DESC, CASE WHEN design_type = '图纸设计' OR design_type = '工程设计' THEN project_number END DESC"

        cursor.execute(query)
        result = cursor.fetchall()

        result = [[element if element is not None else '' for element in row] for row in result]
        cursor.close()
        conn.close()

        # 清空treeview表格
        table_items = Project_Info_table.get_children()  # 在插入treeview数据时，需要先清空treeview
        [Project_Info_table.delete(table_item) for table_item in table_items]

        # 为treeview表格插入数据

        temp_list = []
        for i in range(0, len(result)):
            temp_list = list(result[i])[0:43] + list(result[i])[47:] + list(result[i])[43:47]
            Project_Info_table.insert('', 'end', values=temp_list, tags='fontsize')  # dataframe逐行插入到表格中

        design_type_list = []
        for i in range(0, len(result)):
            design_type_list.append(result[i][1])
        design_type_list_set = list(set(design_type_list))
        design_type_list_set.sort(key=list(design_type_list).index)
        # combobox_design_type.set('')
        combobox_design_type['values'] = design_type_list_set

        project_number_list = []
        for i in range(0, len(result)):
            project_number_list.append(result[i][2])
        project_number_list_set = list(set(project_number_list))
        project_number_list_set.sort(key=list(project_number_list).index)
        # combobox_project_number.set('')
        combobox_project_number['values'] = project_number_list_set

        project_name_list = []
        for i in range(0, len(result)):
            project_name_list.append(result[i][4])
        project_name_list_set = list(set(project_name_list))
        project_name_list_set.sort(key=list(project_name_list).index)
        # combobox_project_name.set('')
        combobox_project_name['values'] = project_name_list_set

        product_type_list = []
        for i in range(0, len(result)):
            product_type_list.append(result[i][8])
        product_type_list_set = list(set(product_type_list))
        product_type_list_set.sort(key=list(product_type_list).index)
        # combobox_product_type.set('')
        combobox_product_type['values'] = product_type_list_set

        PE_list = []
        for i in range(0, len(result)):
            PE_list.append(result[i][9])
        PE_list_set = list(set(PE_list))
        PE_list_set.sort(key=list(PE_list).index)
        # combobox_PE.set('')
        combobox_PE['values'] = PE_list_set

        DE_list = []
        for i in range(0, len(result)):
            DE_list.append(result[i][10])
        DE_list_set = list(set(DE_list))
        DE_list_set.sort(key=list(DE_list).index)
        # combobox_DE.set('')
        combobox_DE['values'] = DE_list_set

        status_list = []
        for i in range(0, len(result)):
            status_list.append(result[i][26])
        status_list_set = list(set(status_list))
        status_list_set.sort(key=list(status_list).index)
        # combobox_status.set('')
        combobox_status['values'] = status_list_set

        panel_type_list = []
        for i in range(0, len(result)):
            panel_type_list.append(result[i][7])
        panel_type_list_set = list(set(panel_type_list))
        panel_type_list_set.sort(key=list(panel_type_list).index)
        # combobox_panel_type.set('')
        combobox_panel_type['values'] = panel_type_list_set

        intelligent_list = []
        for i in range(0, len(result)):
            intelligent_list.append(result[i][11])
        intelligent_list_set = list(set(intelligent_list))
        intelligent_list_set.sort(key=list(intelligent_list).index)
        # combobox_intelligent.set('')
        combobox_intelligent['values'] = intelligent_list_set

        protection_list = []
        for i in range(0, len(result)):
            protection_list.append(result[i][13])
        protection_list_set = list(set(protection_list))
        protection_list_set.sort(key=list(protection_list).index)
        # combobox_protection.set('')
        combobox_protection['values'] = protection_list_set

        arclight_list = []
        for i in range(0, len(result)):
            arclight_list.append(result[i][14])
        arclight_list_set = list(set(arclight_list))
        arclight_list_set.sort(key=list(arclight_list).index)
        # combobox_arclight.set('')
        combobox_arclight['values'] = arclight_list_set


def query_all():
    global result
    button_detail['state'] = 'disabled'
    button_change_single['state'] = 'disabled'
    button_change_multi['state'] = 'disabled'
    button_ahead2other['state'] = 'disabled'

    if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools'):
        tk.messagebox.showwarning("提示", "请连接内网")

    else:
        if is_folder_hidden('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb') or os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb'):
            if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db'):
                tk.messagebox.showwarning("提示", "数据库不存在，请联系管理员")
            else:
                conn = sqlite3.connect('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db')
                cursor = conn.cursor()  # 创建一个Cursor
                cursor.execute('SELECT * FROM project_data ORDER BY CASE WHEN design_type = "提前设计" THEN rowid END DESC, CASE WHEN design_type = "图纸设计" OR design_type = "工程设计" THEN project_number END DESC')
                result = cursor.fetchall()

                result = [[element if element is not None else '' for element in row] for row in result]
                cursor.close()
                conn.close()

                # 清空treeview表格
                table_items = Project_Info_table.get_children()  # 在插入treeview数据时，需要先清空treeview
                [Project_Info_table.delete(table_item) for table_item in table_items]

                # 为treeview表格插入数据
                temp_list = []
                for i in range(0, len(result)):
                    temp_list = list(result[i])[0:43] + list(result[i])[47:] + list(result[i])[43:47]
                    Project_Info_table.insert('', 'end', values=temp_list, tags='fontsize')  # dataframe逐行插入到表格中

                for i in range(0, len(result)):
                    design_type_list.append(result[i][1])
                design_type_list_set = list(set(design_type_list))
                design_type_list_set.sort(key=list(design_type_list).index)
                combobox_design_type.set('')
                combobox_design_type['values'] = design_type_list_set

                for i in range(0, len(result)):
                    project_number_list.append(result[i][2])
                project_number_list_set = list(set(project_number_list))
                project_number_list_set.sort(key=list(project_number_list).index)
                combobox_project_number.set('')
                combobox_project_number['values'] = project_number_list_set

                for i in range(0, len(result)):
                    project_name_list.append(result[i][4])
                project_name_list_set = list(set(project_name_list))
                project_name_list_set.sort(key=list(project_name_list).index)
                combobox_project_name.set('')
                combobox_project_name['values'] = project_name_list_set

                for i in range(0, len(result)):
                    product_type_list.append(result[i][8])
                product_type_list_set = list(set(product_type_list))
                product_type_list_set.sort(key=list(product_type_list).index)
                combobox_product_type.set('')
                combobox_product_type['values'] = product_type_list_set

                for i in range(0, len(result)):
                    PE_list.append(result[i][9])
                PE_list_set = list(set(PE_list))
                PE_list_set.sort(key=list(PE_list).index)
                combobox_PE.set('')
                combobox_PE['values'] = PE_list_set

                for i in range(0, len(result)):
                    DE_list.append(result[i][10])
                DE_list_set = list(set(DE_list))
                DE_list_set.sort(key=list(DE_list).index)
                combobox_DE.set('')
                combobox_DE['values'] = DE_list_set

                for i in range(0, len(result)):
                    status_list.append(result[i][26])
                status_list_set = list(set(status_list))
                status_list_set.sort(key=list(status_list).index)
                combobox_status.set('')
                combobox_status['values'] = status_list_set

                for i in range(0, len(result)):
                    panel_type_list.append(result[i][7])
                panel_type_list_set = list(set(panel_type_list))
                panel_type_list_set.sort(key=list(panel_type_list).index)
                combobox_panel_type.set('')
                combobox_panel_type['values'] = panel_type_list_set

                for i in range(0, len(result)):
                    intelligent_list.append(result[i][11])
                intelligent_list_set = list(set(intelligent_list))
                intelligent_list_set.sort(key=list(intelligent_list).index)
                combobox_intelligent.set('')
                combobox_intelligent['values'] = intelligent_list_set

                for i in range(0, len(result)):
                    protection_list.append(result[i][13])
                protection_list_set = list(set(protection_list))
                protection_list_set.sort(key=list(protection_list).index)
                combobox_protection.set('')
                combobox_protection['values'] = protection_list_set

                for i in range(0, len(result)):
                    arclight_list.append(result[i][14])
                arclight_list_set = list(set(arclight_list))
                arclight_list_set.sort(key=list(arclight_list).index)
                combobox_arclight.set('')
                combobox_arclight['values'] = arclight_list_set

        else:
            tk.messagebox.showwarning("提示", "数据库不存在，请联系管理员")


def is_folder_hidden(fpath):
    try:
        attrs = ctypes.windll.kernel32.GetFileAttributesW(fpath)  # attrs值为18表示该文件夹具有以下属性组合：只读 (1)、隐藏 (2) 和 子文件夹 (16)
        # print(attrs)
        if attrs != -1 and attrs & 2 == 2:  # 对于18（二进制为10010）与2（二进制为00010）进行按位与运算，结果为2（二进制为00010）
            return True
    except OSError:
        pass
    return False


def query_project(event=''):
    try:
        if ahead_entry_project_number.get() == "" or len(ahead_entry_project_number.get()) != 9:
            tk.messagebox.showwarning("提示", "请输入9位项目号")
        elif not ahead_entry_project_number.get().isdigit():
            tk.messagebox.showwarning("提示", "项目号必须是9位纯数字")
        else:
            try:
                workbook1 = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\SR_Material.xlsx")
                worksheet1 = workbook1['Sheet1']
            except:
                tk.messagebox.showwarning("提示", "缺少产品型号表")

            worksheet1_max_row = max(ee.row for ee in worksheet1['A'] if ee.value)
            global material_list
            material_list = []
            global product_type_list
            product_type_list = []

            for i in range(1, worksheet1_max_row + 1):
                material_list.append(worksheet1.cell(row=i, column=1).value)
                product_type_list.append(worksheet1.cell(row=i, column=3).value)

            ahead_entry_project_name['state'] = 'normal'
            ahead_entry_project_name.delete(0, "end")
            ahead_text_item['state'] = 'normal'
            ahead_text_item.delete(1.0, tk.END)
            ahead_text_panel_amount['state'] = 'normal'
            ahead_text_panel_amount.delete(1.0, tk.END)
            ahead_text_typical_type['state'] = 'normal'
            ahead_text_typical_type.delete(1.0, tk.END)
            ahead_text_product_type['state'] = 'normal'
            ahead_text_product_type.delete(1.0, tk.END)

            button_update_amount['state'] = 'disabled'

            global item_list
            item_list = []
            global amount_list
            amount_list = []
            amount_panel = 0
            global product_type_list1
            product_type_list1 = []
            global product_type_list2
            product_type_list2 = []

            # SAP登录
            sap_key_file = 'J:/Engineering/ShareFolder/new_ABB_Production_Tools/Ps/k/&%&^RD&^T&^WRE^@FEYUGYU@^E@DRYTRYTDFYTWF.txt'
            if not os.path.exists(sap_key_file):
                tk.messagebox.showwarning('提示', 'SAP加密密钥文件不存在')
            else:
                with open(sap_key_file, 'rb') as file:
                    sap_key = file.read()
            sap_cipher = Fernet(sap_key)

            sap_ciphertext_file = 'J:/Engineering/ShareFolder/new_ABB_Production_Tools/Ps/s/^&%&^R^YFT$%FTDTEYTKFY$^$FGKJGHE%#$%#EYD^%#^%DTYD.txt'
            if not os.path.exists(sap_ciphertext_file):
                tk.messagebox.showwarning('提示', 'SAP加密密文文件不存在')
            else:
                with open(sap_ciphertext_file, 'r') as file:
                    sap_ciphertext = file.read()

            sap_account = sap_decrypt_data(sap_ciphertext, sap_cipher, 'z_')
            sap_h = sap_account.split('+')[0]
            sap_c = sap_account.split('+')[1]
            sap_s = sap_account.split('+')[2]
            sap_u = sap_account.split('+')[3]
            sap_p = sap_account.split('+')[4]

            logging.info("Ready to connect to SAP")
            try:
                conn = pyrfc.Connection(ashost=sap_h, sysnr=sap_s, client=sap_c, user=sap_u, passwd=sap_p, lang='EN')
                if conn.alive:
                    logging.info("Connecting to SAP successfully")

                    result = conn.call('ZY_SALES_ORDER_SHIFT', VBELN='0' + ahead_entry_project_number.get())
                    if result['EX_CEPTION'] == '' and result['ITAB'][0]['VKBUR'] == '1200':
                        project_name = result['ITAB'][0]['BSTKD']
                        ahead_entry_project_name.insert(0, project_name)
                        ahead_entry_project_name['state'] = 'disabled'

                        table_content = []
                        # 处理数据
                        for item in result['ITAB']:
                            posnr = item['POSNR'].lstrip('0') or '0'
                            kwmeng = str(int(float(item['KWMENG'])))
                            table_content.append((posnr, item['MATNR'], kwmeng))  # [('1000','UNIGEAR-ZS1-500', '14'), ('2000','UNIGEAR-ZS1', '6')]

                        for j in range(0, len(table_content)):
                            for k in range(0, len(material_list)):
                                if table_content[j][1] == material_list[k]:
                                    item_list.append(int(table_content[j][0]))
                                    amount_list.append(int(table_content[j][2]))
                                    amount_panel += int(table_content[j][2])
                                    product_type_list1.append(material_list[k])
                                    product_type_list2.append(product_type_list[k])
                                    break
                        item_list_copy = item_list.copy()
                        item_list.sort()

                        amount_list = [amount_list[i] for i in sorted(range(len(item_list_copy)), key=lambda x: item_list_copy[x])]
                        product_type_list1 = [product_type_list1[i] for i in sorted(range(len(item_list_copy)), key=lambda x: item_list_copy[x])]
                        product_type_list2 = [product_type_list2[i] for i in sorted(range(len(item_list_copy)), key=lambda x: item_list_copy[x])]

                        ahead_text_item.insert(tk.END, ';'.join(str(num) for num in item_list))
                        ahead_text_panel_amount.insert(tk.END, ';'.join(str(num) for num in amount_list))
                        ahead_text_typical_type.insert(tk.END, ';'.join(str(num) for num in product_type_list1))
                        ahead_text_product_type.insert(tk.END, ';'.join(str(num) for num in product_type_list2))

                        if len(item_list) > 1:
                            button_update_amount['state'] = 'normal'

                        button_confirm_change_ahead2other['state'] = 'normal'
                        update_amount()  # 每次读取完刷新一下

                    else:
                        if result['EX_CEPTION'] != '':
                            tk.messagebox.showwarning("提示", result['EX_CEPTION'])
                        else:
                            tk.messagebox.showwarning("提示", 'CNDMX无此项目')

                    conn.close()
                    if not conn.alive:
                        logging.info("Disconnect from SAP")


            except pyrfc.RFCError as e:
                logging.info(e.key + ', ' + e.message)
                tk.messagebox.showwarning("提示", traceback.format_exc())

    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def update_amount():
    global real_item_list  # CI
    real_item_list = []
    global real_product_type_list1  # 产品类型
    real_product_type_list1 = []
    global real_product_type_list2  # 对应的产品类型（AIS,GIS）
    real_product_type_list2 = []
    global real_amount_list  # 每个CI数量
    real_amount_list = []

    temp = str(ahead_text_item.get("1.0", "end-1c")).replace('；', ';')
    if temp[len(temp) - 1] == ';':
        temp = temp[0:len(temp) - 1]
    # print(temp)
    real_item_list = temp

    for i in range(0, len(item_list)):
        if str(item_list[i]) in temp.split(';'):
            real_product_type_list1.append(product_type_list1[i])
            real_product_type_list2.append(product_type_list2[i])
            real_amount_list.append(amount_list[i])

    ahead_text_item.delete(1.0, tk.END)
    ahead_text_panel_amount.delete(1.0, tk.END)
    ahead_text_typical_type.delete(1.0, tk.END)
    ahead_text_product_type.delete(1.0, tk.END)

    ahead_text_item.insert(tk.END, real_item_list)
    ahead_text_panel_amount.insert(tk.END, ';'.join(str(num) for num in real_amount_list))
    ahead_text_typical_type.insert(tk.END, ';'.join(str(num) for num in real_product_type_list1))
    ahead_text_product_type.insert(tk.END, ';'.join(str(num) for num in real_product_type_list2))


def sap_decrypt_data(data, data_cipher, prefix):
    # 移除列名前缀prefix，如果有的话
    if data.startswith(prefix):
        data = data[2:]

    # Base64 字符串中可能删除了等号，需要添加回原来的填充字符，因为 base64 编码的输出长度应该是 4 的倍数
    padding_needed = 4 - (len(data) % 4)
    if padding_needed:
        data += "=" * padding_needed

    # base64 解码以获取原始的加密字节串
    encrypted_data = base64.urlsafe_b64decode(data.encode())

    # 使用 data_cipher 对象解密，它应该有和加密时一样的密钥
    decrypted_data = data_cipher.decrypt(encrypted_data)

    # 将字节解码回字符串
    decrypted_string = decrypted_data.decode()
    return decrypted_string


def confirm_change_ahead2other(edit_window, selected_rows):
    try:
        result = tk.messagebox.askquestion("提示", "确定要将项目信息写入数据库吗？\n将移除对应的提前设计信息行，转换为若干图纸/工程设计信息行")
        if result == 'yes':
            if not os.path.exists('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db'):
                tk.messagebox.showwarning("提示", "数据库不存在，请联系管理员")
            else:
                if ahead_text_item.get("1.0", "end-1c")[-1] == ';' or ahead_text_item.get("1.0", "end-1c")[-1] == '；':
                    tk.messagebox.showwarning("提示", "请点击刷新")
                else:
                    conn = sqlite3.connect('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db')
                    cursor = conn.cursor()  # 创建一个Cursor
                    # 创建project_data表
                    cursor.execute('''CREATE TABLE IF NOT EXISTS project_data
                                     (id INTEGER PRIMARY KEY AUTOINCREMENT,
                                      design_type TEXT,
                                      project_number TEXT,
                                      item_info INTEGER,
                                      project_name TEXT,
                                      typical_amount INTEGER,
                                      panel_amount INTEGER,
                                      panel_type TEXT,
                                      product_type TEXT,
                                      PE TEXT,
                                      DE TEXT,
                                      intelligence TEXT,
                                      main_bus_current INTEGER,
                                      main_protection TEXT,
                                      arc_light TEXT,
                                      location_info TEXT,
                                      drawing_language TEXT,
                                      standard_time INTEGER,
                                      extra_time INTEGER,
                                      advice_times INTEGER,
                                      upload_time TEXT,
                                      bom_time TEXT,
                                      check_time TEXT,
                                      receive_time TEXT,
                                      start_time TEXT,
                                      estimated_time TEXT,
                                      status_info TEXT,
                                      actual_time TEXT,
                                      team_info TEXT,
                                      abnormal_status TEXT,
                                      create_time TIMESTAMP,
                                      project_type TEXT,
                                      frame_type TEXT,
                                      program_type TEXT,
                                      language_info TEXT,
                                      management_info TEXT,
                                      input_file TEXT,
                                      drawing_type TEXT,
                                      mcb_type TEXT,
                                      aux_type TEXT,
                                      terminal_type TEXT,
                                      charged_display_type TEXT,
                                      switch_type TEXT,
                                      special_panel_config TEXT,
                                      remark TEXT,
                                      reviser TEXT,
                                      revise_time TIMESTAMP,
                                      ct_type TEXT,
                                      pt_type TEXT,
                                      sa_type TEXT,
                                      customer_wiring TEXT,
                                      customer_requirements TEXT)''')
                    already_exist_item = []
                    for i in range(0, len(ahead_text_item.get("1.0", "end-1c").split(';'))):
                        print(int((ahead_text_item.get("1.0", "end-1c").split(';'))[i]))
                        data = {
                            'design_type': ahead_combobox_design_type.get(),
                            'project_number': ahead_entry_project_number.get(),
                            'item_info': int((ahead_text_item.get("1.0", "end-1c").split(';'))[i]),
                        }
                        query = "SELECT * FROM project_data WHERE project_number=? AND item_info=?"
                        cursor.execute(query, (data['project_number'], data['item_info']))
                        result = cursor.fetchone()
                        if result:
                            already_exist_item.append(int((ahead_text_item.get("1.0", "end-1c").split(';'))[i]))

                    if len(already_exist_item) > 0:
                        tk.messagebox.showwarning("提示", "%s行号数据已存在，写入失败" % already_exist_item)
                        cursor.close()
                        conn.close()
                        return
                    else:
                        chinese_username = ''
                        if os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\Administrators.xlsx"):
                            workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\Administrators.xlsx")
                            worksheet = workbook['Sheet1']

                            for i in range(2, worksheet.max_row + 1):
                                if os.getlogin() == worksheet.cell(row=i, column=1).value:
                                    chinese_username = worksheet.cell(row=i, column=2).value
                        # print(chinese_username)
                        for i in range(0, len(ahead_text_item.get("1.0", "end-1c").split(';'))):
                            data = {
                                'design_type': ahead_combobox_design_type.get(),
                                'project_number': ahead_entry_project_number.get(),
                                'item_info': int((ahead_text_item.get("1.0", "end-1c").split(';'))[i]),
                                'project_name': ahead_entry_project_name.get(),
                                'typical_amount': int(ahead_typical_amount),
                                'panel_amount': int((ahead_text_panel_amount.get("1.0", "end-1c").split(';'))[i]),
                                'panel_type': ahead_text_typical_type.get("1.0", "end-1c").split(';')[i],
                                'product_type': ahead_text_product_type.get("1.0", "end-1c").split(';')[i],
                                'PE': ahead_PE,
                                'DE': ahead_DE,
                                'intelligence': ahead_intelligence,
                                'main_bus_current': ahead_main_busbar_current,
                                'main_protection': ahead_main_protection,
                                'arc_light': ahead_arclight,
                                'location_info': ahead_location,
                                'drawing_language': ahead_drawing_language,
                                'standard_time': ahead_standard_period,
                                'extra_time': ahead_extra_period,
                                'advice_times': int(ahead_advice_times),
                                'upload_time': ahead_upload_time,
                                'bom_time': ahead_bom_time,
                                'check_time': ahead_check_time,
                                'receive_time': ahead_receive_time,
                                'start_time': ahead_start_time,
                                'estimated_time': ahead_estimated_time,
                                'status_info': ahead_status_info,
                                'actual_time': ahead_actual_time,
                                'team_info': ahead_team_info,
                                'abnormal_status': ahead_abnormal_status,
                                'create_time': ahead_create_time,
                                'project_type': ahead_project_type,
                                'frame_type': ahead_frame_type,
                                'program_type': ahead_program_type,
                                'language_info': ahead_language_info,
                                'management_info': ahead_management_info,
                                'input_file': ahead_input_file,
                                'drawing_type': ahead_drawing_type,
                                'mcb_type': ahead_mcb_type,
                                'aux_type': ahead_aux_type,
                                'terminal_type': ahead_terminal_type,
                                'charged_display_type': ahead_charged_display_type,
                                'switch_type': ahead_switch_type,
                                'special_panel_config': ahead_entries1,
                                'remark': ahead_entries2,
                                'reviser': chinese_username,
                                'revise_time': strftime('%Y-%m-%d %H:%M:%S', localtime()),
                                'ct_type': ahead_ct_type,
                                'pt_type': ahead_pt_type,
                                'sa_type': ahead_sa_type,
                                'customer_wiring': ahead_customer_wiring,
                                'customer_requirements': ahead_customer_requirements,
                            }
                            columns = ', '.join(data.keys())
                            placeholders = ', '.join(['?' for _ in range(len(data))])
                            query = f"INSERT INTO project_data ({columns}) VALUES ({placeholders})"

                            values = [data.get(col) for col in data.keys()]
                            cursor.execute(query, tuple(values))

                            conn.commit()

                            query = "DELETE FROM project_data WHERE design_type=? AND project_number=? AND item_info=? AND id=?"
                            cursor.execute(query, ('提前设计', '提前设计', '1000', selected_rows[0][0]))
                            conn.commit()

                        tk.messagebox.showwarning("提示", "数据更新完成")
                        cursor.close()
                        conn.close()

                        source_file = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pb\\FAST_Project_Database.db'
                        target_dir = 'J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\backup_db'
                        target_file = os.path.join(target_dir, os.path.basename(source_file))

                        if os.path.exists(target_file):
                            os.remove(target_file)

                        shutil.copy2(source_file, target_dir)

                        edit_window.destroy()
                        query_all()
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())
