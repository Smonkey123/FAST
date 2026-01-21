import tkinter as tk
from tkinter import ttk
from tkinter import StringVar
from tkinter.ttk import Treeview, Style

import need.tkutils as tku
from tkinter.filedialog import askdirectory

import os
import re
import numpy as np
from time import time
from time import localtime
from time import strftime

import pandas as pd
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl import load_workbook

import shutil  # 文件操作
from xml.dom.minidom import Document  # 使用最小化dom实现元素树创建
from xml.dom.minidom import parse
import warnings
import traceback
from cryptography.fernet import Fernet
import base64
import pyrfc
import traceback
import datetime
import logging
from need.custom_dialogs import CustomDialog, center_window, Tooltip, image_label

warnings.simplefilter(action='ignore', category=FutureWarning)

FilePath = ""  # 设置一个地址变量


# xml节点构造函数
def makexml(Nodename, Text, FatherNodeElement):
    NodeElement = xml_doc.createElement(Nodename)  # 创建节点
    if Text != '':
        NodeElementtext = xml_doc.createTextNode(Text)  # 创建节点的文本（节点）
        NodeElement.appendChild(NodeElementtext)  # 将文本插入节点
    FatherNodeElement.appendChild(NodeElement)  # 将节点插入父节点


# xml节点构造函数(带1个节点属性)
def makexml1(Nodename, attri_name, attri_value, Text, FatherNodeElement):
    NodeElement = xml_doc.createElement(Nodename)  # 创建节点
    NodeElement.setAttribute(attri_name, attri_value)
    if Text != '':
        NodeElementtext = xml_doc.createTextNode(Text)  # 创建节点的文本（节点）
        NodeElement.appendChild(NodeElementtext)  # 将文本插入节点
    FatherNodeElement.appendChild(NodeElement)  # 将节点插入父节点


# xml节点构造函数(带2个节点属性)
def makexml2(Nodename, attri_name1, attri_value1, attri_name2, attri_value2, Text, FatherNodeElement):
    NodeElement = xml_doc.createElement(Nodename)  # 创建节点
    NodeElement.setAttribute(attri_name1, attri_value1)
    NodeElement.setAttribute(attri_name2, attri_value2)
    if Text != '':
        NodeElementtext = xml_doc.createTextNode(Text)  # 创建节点的文本（节点）
        NodeElement.appendChild(NodeElementtext)  # 将文本插入节点
    FatherNodeElement.appendChild(NodeElement)  # 将节点插入父节点


def main(parent, root, w_rat, h_rat, file_path):
    global checklog_file_path
    checklog_file_path = file_path
    global open_table
    open_table = tk.PhotoImage(file="ico\\down.png")
    global close_table
    close_table = tk.PhotoImage(file="ico\\up.png")
    global open_folder
    open_folder = tk.PhotoImage(file="ico\\open_folder.png")
    global analyze_file
    analyze_file = tk.PhotoImage(file="ico\\read.png")
    global view_folder
    view_folder = tk.PhotoImage(file="ico\\view.png")
    global create_file_document
    create_file_document = tk.PhotoImage(file="ico\\create.png")
    global import_file_document
    import_file_document = tk.PhotoImage(file="ico\\import.png")
    global refresh_state
    refresh_state = tk.PhotoImage(file="ico\\refresh.png")

    # global canvas
    # canvas = tk.Canvas(parent, width=int(1600 * w_ratio), height=int(640 * h_ratio), bg="#C9DBE9")
    # canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    # canvas.update()
    # canvas.bind("<MouseWheel>", on_mousewheel)
    #
    # scrollbar_v = tk.Scrollbar(master=parent)
    # scrollbar_v.pack(side=tk.RIGHT, fill=tk.Y)
    # scrollbar_v.config(command=canvas.yview)
    # canvas.config(yscrollcommand=scrollbar_v.set)
    #
    # content = tk.Frame(canvas)
    # # content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    #
    # canvas.create_window(0, 1, width=int(1600 * w_ratio), anchor=tk.NW, window=content)

    global root_win
    root_win = root

    global w_ratio
    w_ratio = w_rat
    global h_ratio
    h_ratio = h_rat

    f1 = tk.Frame(parent, bg="#c9dbe9", bd=0)
    # im = tku.image_label(f1, "ico\\abb_help-circle-2_32.png", int(30 * h_ratio), int(30 * h_ratio), False)
    # im.configure(bg="#c9dbe9")
    # im.bind('<Button-1>', about_help)  # 帮助图标绑定动作
    # im.pack(side=tk.RIGHT)
    tk.Label(f1, text="欢迎使用EBOM导入SAP功能", bg="#c9dbe9", fg="black", height=int(1 * h_ratio), font=("ABBvoice CNSG", int(20 * h_ratio), "bold")).pack(fill=tk.X)
    f1.pack(fill=tk.X)

    f2 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f2, text='   说明：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')
    tk.Label(f2, text='(1)EPLAN菜单Tools→Reports:Automated processing中选择Export Files V04导出；\n(2)选择BOM文件【C:/Temp/项目号-Files/项目号-BOM.xlsx】；\n(3)如果需要修改"柜型-柜号"关系，需要在单线图对应柜子双击，修改Parts→Part number，重新使用V04导出报表。', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    f2.pack(fill=tk.BOTH, expand=True)

    tk.Frame(parent, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f3 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f3, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
    tk.Label(f3, text='路径：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)

    global entry  # 为了确保selectpath函数能正确调用entry,将其全局化
    entry = tk.Entry(f3, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)))
    entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

    global button_open
    button_open = tk.Button(f3, image=view_folder, bg="#eaf1f6", text="看文件夹", font=("ABBvoice CNSG", int(13 * h_ratio)), command=open_filefolder, compound=tk.LEFT, state='disabled', activebackground='blue')
    button_open.pack(side=tk.RIGHT, padx=int(20 * w_ratio))

    tk.Button(f3, text="选取", image=open_folder, font=("ABBvoice CNSG", int(13 * h_ratio)), bg="#eaf1f6", compound=tk.LEFT, command=selectpath, activebackground='blue').pack(side=tk.RIGHT, padx=(int(20 * w_ratio), 0))

    f3.pack(fill=tk.X)

    tk.Frame(parent, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    global f4
    f4 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    # tk.Label(f4, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')
    tk.Label(f4, text='    SAP：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')

    style = Style()
    style.configure('panel1.Treeview', rowheight=25, font=("ABBvoice CNSG", int(13 * h_ratio)))
    style.configure('panel1.Treeview.Heading', font=("ABBvoice CNSG", int(13 * h_ratio)), background="#EFF1F5")

    global SO_Item_table0
    table_ybar0 = tk.Scrollbar(f4)

    SO_Item_table0 = Treeview(f4, show='headings', style='panel1.Treeview', selectmode='browse', columns=('a', 'b', 'c', 'd', 'e'), yscrollcommand=table_ybar0.set, height=8)
    table_ybar0.config(command=SO_Item_table0.yview)
    SO_Item_table0.column('a', width=int(70 * w_ratio), anchor='center')
    SO_Item_table0.column('b', width=int(70 * w_ratio), anchor='center')
    SO_Item_table0.column('c', width=int(750 * w_ratio), anchor='center')
    SO_Item_table0.column('d', width=int(70 * w_ratio), anchor='center')
    SO_Item_table0.column('e', width=int(140 * w_ratio), anchor='center')

    SO_Item_table0.heading('a', text='站号', anchor='center')
    SO_Item_table0.heading('b', text='柜型', anchor='center')
    SO_Item_table0.heading('c', text='ABB柜号', anchor='center')
    SO_Item_table0.heading('d', text='柜数', anchor='center')
    SO_Item_table0.heading('e', text='SO Item', anchor='center')
    SO_Item_table0.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 1), pady=0)
    SO_Item_table0.tag_configure('even_row', background="#c9dbe9")
    SO_Item_table0.tag_configure('odd_row', background="white")

    table_ybar0.pack(side=tk.LEFT, fill=tk.Y)

    global button_read_sap
    button_read_sap = tk.Button(f4, text="查询", font=("ABBvoice CNSG", int(13 * h_ratio)), image=analyze_file, bg="#eaf1f6", compound=tk.LEFT, command=read_sap_item, activebackground='blue')
    button_read_sap.pack(side=tk.LEFT, pady=0, padx=int(20 * w_ratio))
    button_read_sap['state'] = 'disabled'

    f4.pack(fill=tk.X)

    tk.Frame(parent, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f6 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f6, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')
    tk.Label(f6, text='端子：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')

    f6.option_add("*TCombobox*Listbox.font", ("ABBvoice CNSG", int(13 * h_ratio)))
    global combobox
    combobox_value = StringVar()
    combobox_value.set('DZP-TE-AIS')
    combobox_values = ['DZP-TE-AIS', 'DZP-TE-GIS', 'DZP-WEIDMULLER-AIS', 'DZP-WEIDMULLER-GIS', 'DZP-PHOENIX-AIS', 'DZP-PHOENIX-GIS', 'DZP-RELIANCE-AIS', 'DZP-RELIANCE-GIS']
    combobox = ttk.Combobox(master=f6, font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(40 * w_ratio), state='readonly', cursor='arrow', textvariable=combobox_value, values=combobox_values)
    combobox.pack(side=tk.LEFT, padx=0)
    Tooltip(combobox, "正常会自动识别端子排型号，无需手动选择")

    tk.Label(f6, text='      *', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n', padx=0)
    tk.Label(f6, text='SO：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')
    global entry_projectnumber
    entry_projectnumber = tk.Entry(f6, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), width=int(40 * w_ratio))
    entry_projectnumber.pack(side=tk.LEFT, padx=0)
    Tooltip(entry_projectnumber, "支持9位项目、8位样柜项目号")
    tk.Label(f6, text='（支持9位项目、8位样柜项目号）', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(10 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')
    f6.pack(fill=tk.X)

    tk.Frame(parent, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    global f5
    f5 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    # tk.Label(f5, text='*', bg="#eaf1f6", fg="red", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')
    tk.Label(f5, text='EPLAN：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')

    global SO_Item_table1
    table_ybar1 = tk.Scrollbar(f5)

    SO_Item_table1 =Treeview(f5, show='headings', style='panel1.Treeview', selectmode='browse', columns=('a', 'b', 'c', 'd', 'e'), yscrollcommand=table_ybar1.set, height=8)
    table_ybar1.config(command=SO_Item_table1.yview)
    SO_Item_table1.column('a', width=int(70*w_ratio), anchor='center')
    SO_Item_table1.column('b', width=int(70*w_ratio), anchor='center')
    SO_Item_table1.column('c', width=int(750*w_ratio), anchor='center')
    SO_Item_table1.column('d', width=int(70*w_ratio), anchor='center')
    SO_Item_table1.column('e', width=int(140*w_ratio), anchor='center')

    SO_Item_table1.heading('a', text='站号', anchor='center')
    SO_Item_table1.heading('b', text='柜型', anchor='center')
    SO_Item_table1.heading('c', text='ABB柜号', anchor='center')
    SO_Item_table1.heading('d', text='柜数', anchor='center')
    SO_Item_table1.heading('e', text='SO Item(↓可编辑)', anchor='center')
    SO_Item_table1.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 1), pady=0)
    SO_Item_table1.bind('<Double-1>', lambda event, tree=SO_Item_table1: edit_cell(event, tree))
    SO_Item_table1.tag_configure('attention_row', foreground='red')
    SO_Item_table1.tag_configure('even_row', background="#c9dbe9")
    SO_Item_table1.tag_configure('odd_row', background="white")
    table_ybar1.pack(side=tk.LEFT, fill=tk.Y)

    Tooltip(SO_Item_table1, "SO Item列可双击编辑，请保持与上表(SAP)中SO Item一致")

    global button_read_eplan
    button_read_eplan = tk.Button(f5, text="读取", font=("ABBvoice CNSG", int(13 * h_ratio)), image=analyze_file, bg="#eaf1f6", compound=tk.LEFT, command=read_panel_and_table, activebackground='blue')
    button_read_eplan.pack(side=tk.LEFT, pady=0, padx=int(20 * w_ratio))
    button_read_eplan['state'] = 'disabled'

    f5.pack(fill=tk.X)

    tk.Frame(parent, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    f7 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f7, text='   输出：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')

    global text_ebom_import
    text_ebom_import = tk.Text(f7, bg="#eaf1f6", font=("ABBvoice CNSG", int(13 * h_ratio)), height=int(1 * h_ratio))
    text_ebom_import.pack(side=tk.LEFT, padx=(0, 1), pady=0, fill=tk.BOTH, expand=True)

    scrollbar = tk.Scrollbar(f7)
    scrollbar.pack(side=tk.LEFT, fill=tk.Y)
    scrollbar.config(command=text_ebom_import.yview)
    text_ebom_import.config(yscrollcommand=scrollbar.set)

    f71 = tk.Frame(f7, bg="#eaf1f6", bd=0)
    f71.pack(side=tk.LEFT, fill=tk.Y, pady=0, padx=int(20 * w_ratio))


    global button_create_xml
    button_create_xml = tk.Button(f71, text="创建", font=("ABBvoice CNSG", int(13 * h_ratio)), image=create_file_document, bg="#eaf1f6", compound=tk.LEFT, command=create_xml_file, activebackground='blue')
    button_create_xml.pack(side=tk.TOP, pady=(5, 5))
    button_create_xml['state'] = 'disabled'

    # global button_view_xml
    # button_view_xml = tk.Button(f71, text="查看", image=view_folder, bg="#eaf1f6", compound=tk.LEFT, command=open_filefolder, state='disabled', activebackground='blue')
    # button_view_xml.pack(side=tk.TOP, pady=(0, 5))
    # button_view_xml['state'] = 'disabled'

    global button_import_xml
    button_import_xml = tk.Button(f71, text="导入", font=("ABBvoice CNSG", int(13 * h_ratio)), image=import_file_document, bg="#eaf1f6", compound=tk.LEFT, command=import_xml_file, state='disabled', activebackground='blue')
    button_import_xml.pack(side=tk.TOP, pady=(0, 5))
    button_import_xml['state'] = 'disabled'

    global button_view_imported_xml
    button_view_imported_xml = tk.Button(f71, text="查看", font=("ABBvoice CNSG", int(13 * h_ratio)), image=view_folder, bg="#eaf1f6", compound=tk.LEFT, command=open_sendfolder, activebackground='blue')
    button_view_imported_xml.pack(side=tk.TOP, pady=(0, 5))
    Tooltip(button_view_imported_xml, "查看BOM.xml导入情况")

    f7.pack(fill=tk.X)

    tk.Frame(parent, height=int(20 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    global f8
    f8 = tk.Frame(parent, bg="#eaf1f6", bd=0)
    tk.Label(f8, text='   状态：', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(13 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X, anchor='n')

    global SO_Item_table2
    table_ybar2 = tk.Scrollbar(f8)

    # SO_Item_table2 = Treeview(f8, show='headings', columns=('a', 'b', 'c', 'd', 'e'), yscrollcommand=table_ybar2.set, height=8)
    SO_Item_table2 = Treeview(f8, show='headings', style='panel1.Treeview', selectmode='browse', columns=('b', 'c', 'd', 'e'), yscrollcommand=table_ybar2.set, height=6)

    SO_Item_table2.tag_configure('fail_line', foreground='red')  # 失败行显示红色
    SO_Item_table2.tag_configure('pass_line', foreground='green')  # 成功行显示绿色

    table_ybar2.config(command=SO_Item_table2.yview)
    # SO_Item_table2.column('a', width=125, anchor='w')
    SO_Item_table2.column('b', width=int(200 * w_ratio), anchor='center')
    SO_Item_table2.column('c', width=int(60 * w_ratio), anchor='center')
    SO_Item_table2.column('d', width=int(125 * w_ratio), anchor='center')
    SO_Item_table2.column('e', width=int(180 * w_ratio), anchor='center')
    # SO_Item_table2.heading('a', text='XML文件')
    SO_Item_table2.heading('b', text='响应XML文件', anchor='center')
    SO_Item_table2.heading('c', text='状态', anchor='center')
    SO_Item_table2.heading('d', text='时间', anchor='center')
    SO_Item_table2.heading('e', text='消息', anchor='center')
    SO_Item_table2.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 1), pady=0)
    table_ybar2.pack(side=tk.LEFT, fill=tk.Y)
    Tooltip(SO_Item_table2, "如果某个Item导入失败，请点击本表右侧'查看'按钮，找到对应log文件(项目号_站号_Typical.log)")

    f81 = tk.Frame(f8, bg="#eaf1f6", bd=0)
    f81.pack(side=tk.LEFT, fill=tk.Y, pady=0, padx=int(20 * w_ratio))

    global button6
    button6 = tk.Button(f81, text="刷新", font=("ABBvoice CNSG", int(13 * h_ratio)), image=refresh_state, bg="#eaf1f6", compound=tk.LEFT, command=update_bom_import_state, activebackground='blue')
    button6.pack(side=tk.TOP, pady=(60, 5))
    button6['state'] = 'disabled'

    SO_Item_table2.bind('<Button-1>', show_detail)  # 进入控件，查看详细信息

    global button7
    button7 = tk.Button(f81, text="查看", font=("ABBvoice CNSG", int(13 * h_ratio)), image=view_folder, bg="#eaf1f6", compound=tk.LEFT, command=open_logsfolder, activebackground='blue')
    button7.pack(side=tk.TOP, pady=(0, 20))
    Tooltip(button7, "查看BOM.xml导入日志")

    f8.pack(fill=tk.X)
    tk.Frame(parent, height=int(2000 * h_ratio), bg="#eaf1f6").pack(fill=tk.X)  # 水平分割线

    # canvas.update_idletasks()
    # # content.update_idletasks()
    # canvas.config(scrollregion=canvas.bbox('all'))


def on_mousewheel(event):
    global canvas
    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")


def open_sendfolder():
    os.startfile('\\\\CN-S-041MVE1\\temp\exportToSAP\send\\')

def open_logsfolder():
    os.startfile('\\\\CN-S-041MVE1\\temp\exportToSAP\logs\\')

def open_filefolder():
    os.startfile(os.path.dirname(FilePath))


def edit_cell(event, tree):
    column_t = tree.identify_column(event.x)
    row_t = tree.identify_row(event.y)

    if column_t:
        col = int(str(column_t).replace('#', ''))
        # 只有选中SO Item所在列时，才会出现可编辑的文本框
        if col == 5:
            if row_t != '':    # 标题栏，返回的信息是空，所以要排除
                content = tree.item(row_t, 'value')[col - 1]

                x = tree.bbox(row_t, column=column_t)[0]  # 单元格x坐标
                y = tree.bbox(row_t, column=column_t)[1]  # 单元格y坐标
                width = tree.bbox(row_t, column=column_t)[2]  # 单元格宽度
                height = tree.bbox(row_t, column=column_t)[3]  # 单元格高度

                entryedit = ttk.Entry(tree)
                entryedit.place(x=x, y=y, width=width, height=height)
                entryedit.insert(0, content)

                def on_leave(event):
                    if event.state == 8:
                        save_edit(event.widget)
                    else:
                        event.widget.unbind('<FocusOut>')

                def on_focus_out(event):
                    event.widget.bind('<FocusOut>', lambda e: event.widget.destroy())

                entryedit.focus_set()
                entryedit.bind('<Return>', lambda e: save_edit(entryedit))
                entryedit.bind('<Leave>', on_leave)
                entryedit.bind('<FocusIn>', on_focus_out)

                def save_edit(widget):
                    new_content = widget.get().replace('\n', '')
                    tree.set(row_t, column_t, new_content)
                    widget.destroy()

                    book2 = xlrd.open_workbook(outputfile)  # 加载XXX-Table.xls表格
                    workbook = copy(book2)  # 使用xlutils.copy将xlrd读取的book对象转为xlwt可操作对象
                    worksheet = workbook.get_sheet(0)  # 获取sheet

                    # 遍历treeview的每一行数据，存入Table.xls
                    for index, table_item in enumerate(tree.get_children()):  # enumerate（容器变量）返回位置及元素
                        table_item_list = tree.item(table_item, 'values')
                        # print(index, table.item(table_item, 'values'))
                        worksheet.write(index + 1, 4, table_item_list[4])
                        # 注意，由于前几列数据都是根据Panel得到，因此这些数据无法修改，因此不需要重新写入，要修改需要调整Panel.xlsx（EPLAN）
                    workbook.save(outputfile)  # 将workbook保存到指定位置


def show_detail(event):
    table = SO_Item_table2
    global column_t1
    global row_t1
    global column_t2
    global row_t2
    column_t2 = table.identify_column(event.x)  # 点击的列列号,#0(不显示),#1,#2
    row_t2 = table.identify_row(event.y)  # 点击的行行号,I001,I002,I003

    # print(row_t2, column_t2, type(row_t2), type(column_t2))    # 结果：I003 #2 <class 'str'> <class 'str'>
    col = int(str(column_t2).replace('#', ''))  # 移除列号中的#号,改字符串为整数
    global detail_label
    if col == 4:
        if table.item(row_t2, 'value') != '':  # 标题栏，返回的信息是空，所以要排除
            text = table.item(row_t2, 'value')[col - 1]  # 单元格内容
            number_space = 0

            for i in text:
                if i == ' ' or i == '\n':
                    number_space += 1

            if number_space < len(text):  # 当存在数据信息时，
                try:
                    detail_label.destroy()
                except (NameError, AssertionError):
                    pass

                if len(text) // 30 == 1:
                    text = text[:30] + '\n' + text[30:]
                    height_flag = 2
                elif len(text) // 30 == 2:
                    text = text[:30] + '\n' + text[30:60] + '\n' + text[60:]
                    height_flag = 3
                elif len(text) // 30 == 3:
                    text = text[:30] + '\n' + text[30:60] + '\n' + text[60:90] + '\n' + text[90:]
                    height_flag = 4
                elif len(text) // 30 == 4:
                    text = text[:30] + '\n' + text[30:60] + '\n' + text[60:90] + '\n' + text[90:120] + '\n' + text[120:]
                    height_flag = 5
                else:
                    height_flag = 2

                x = table.bbox(row_t2, column=column_t2)[0]  # 单元格x坐标
                y = table.bbox(row_t2, column=column_t2)[1]  # 单元格y坐标
                width = table.bbox(row_t2, column=column_t2)[2]  # 单元格宽度
                height = table.bbox(row_t2, column=column_t2)[3]  # 单元格高度
                # global detail_label
                detail_label = tk.Label(f8, text=text, justify='left')
                detail_label.place(x=x + 85, y=y, width=width, height=height * height_flag)  # 设置输入条位置及长宽

            else:
                try:
                    detail_label.destroy()
                except (NameError, AssertionError):
                    pass

        else:
            try:
                detail_label.destroy()
            except (NameError, AssertionError):
                pass

    else:
        try:
            detail_label.destroy()
        except (NameError, AssertionError):
            pass


def about_help(event):
    # tku.show_info("说明书")
    os.startfile(os.path.abspath('J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\二次设计辅助工具【EBOM导SAP功能】答疑V1.2.pdf'))


def selectpath():
    entry_projectnumber.delete(0, "end")  # 删除entry_projectnumber原始内容
    filepath = tk.filedialog.askopenfilename(initialdir='C:/Temp/', title=u'请选择文件(XXX-BOM.xlsx)', filetypes=[("Excel", ".xlsx")])  # 选择打开什么文件，返回文件名
    stem, suffix = os.path.splitext(os.path.basename(filepath))  # stem是文件名,suffix是后缀
    # print(stem, suffix)

    if len(filepath) != 0:
        string_filename = ""
        for i in range(0, len(filepath)):
            string_filename += str(filepath[i])
        button_read_sap['state'] = 'normal'
        button_read_eplan['state'] = 'normal'
        button6['state'] = 'normal'
        # button_view_xml['state'] = 'normal'
        button_open['state'] = 'normal'
        entry_projectnumber.insert(0, stem.split('-')[0].split('.')[0])  # 将BOM文件的项目名插入项目号
        combobox.set('')

        # 如下代码用于根据X5端子确定端子厂家，缩小筛选范围
        global truename
        truename = stem.rsplit("-", 1)[0]
        terminalfile = os.path.join(os.path.dirname(filepath), truename + '-Terminal list.xlsx')  # XXX-Terminal list.xlsx

        sld_file = os.path.join(os.path.dirname(filepath), truename + '-SLD tables.xls')
        if os.path.exists(sld_file):
            sld_book = xlrd.open_workbook(sld_file)  # 加载XXX-SLD tables.xls表格
            sld_worksheet = sld_book.sheet_by_index(0)

            table_col_A = sld_worksheet.col_values(colx=0, start_rowx=1, end_rowx=None)
        else:
            table_col_A = ['', '']

        if os.path.exists(terminalfile):
            book = load_workbook(terminalfile)
            sheet = book['Z1_xlsx']
            D = []
            E = []
            for i in range(3, sheet.max_row + 1):
                D.append(str(sheet.cell(row=i, column=4).value))
                E.append(str(sheet.cell(row=i, column=5).value))

            X5_terminal_type_list = []

            for i in range(0, len(D)):
                if D[i][0:3] == 'X5:':
                    X5_terminal_type_list.append(E[i])

            X5_terminal_type_set_max_number = 0
            X5_terminal_type = ''
            for i in list(set(X5_terminal_type_list)):
                k = X5_terminal_type_list.count(i)
                if X5_terminal_type_set_max_number <= k:
                    X5_terminal_type_set_max_number = k
                    X5_terminal_type = i

            # print(set(X5_terminal_type_list))

            if os.path.exists("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\ABB_FNKS_terminal list.xlsx"):
                workbook = load_workbook("J:\\Engineering\\ShareFolder\\new_ABB_Production_Tools\\Pd\\document\\ABB_FNKS_terminal list.xlsx")  # 加载ABB_FNKS端子对照表
                worksheet = workbook['TE']
                worksheet1 = workbook['WDML']
                worksheet2 = workbook['RL']
                worksheet3 = workbook['PHOENIX']
                flag1 = 0
                flag2 = 0
                flag3 = 0
                flag4 = 0

                for i in range(2, worksheet.max_row + 1):  # 遍历端子类型表
                    if X5_terminal_type == worksheet.cell(row=i, column=2).value:  # 端子属于TE端子
                        flag1 = 1
                        continue
                for i in range(2, worksheet1.max_row + 1):
                    if X5_terminal_type == worksheet1.cell(row=i, column=2).value:  # 端子属于WDML端子
                        flag2 = 1
                        continue
                for i in range(2, worksheet2.max_row + 1):
                    if X5_terminal_type == worksheet2.cell(row=i, column=2).value:  # 端子属于RL端子
                        flag3 = 1
                        continue
                for i in range(2, worksheet3.max_row + 1):
                    if X5_terminal_type == worksheet3.cell(row=i, column=2).value:  # 端子属于PHOENIX端子
                        flag4 = 1
                        continue

                # print(flag1, flag2, flag3, flag4)

                if flag1:
                    if 'ZS' in table_col_A[1] or '500' in table_col_A[1] or '550' in table_col_A[1] or 'Beni' in table_col_A[1]:
                        combobox.configure(values=['DZP-TE-AIS'])
                    else:
                        combobox.configure(values=['DZP-TE-AIS', 'DZP-TE-GIS'])
                elif flag2:
                    if 'ZS' in table_col_A[1] or '500' in table_col_A[1] or '550' in table_col_A[1] or 'Beni' in table_col_A[1]:
                        combobox.configure(values=['DZP-WEIDMULLER-AIS'])
                    else:
                        combobox.configure(values=['DZP-WEIDMULLER-AIS', 'DZP-WEIDMULLER-GIS'])
                elif flag3:
                    if 'ZS' in table_col_A[1] or '500' in table_col_A[1] or '550' in table_col_A[1] or 'Beni' in table_col_A[1]:
                        combobox.configure(values=['DZP-RELIANCE-AIS'])
                    else:
                        combobox.configure(values=['DZP-RELIANCE-AIS', 'DZP-RELIANCE-GIS'])
                elif flag4:
                    if 'ZS' in table_col_A[1] or '500' in table_col_A[1] or '550' in table_col_A[1] or 'Beni' in table_col_A[1]:
                        combobox.configure(values=['DZP-PHOENIX-AIS'])
                    else:
                        combobox.configure(values=['DZP-PHOENIX-AIS', 'DZP-PHOENIX-GIS'])

    else:
        button_read_sap['state'] = 'disabled'
        button_read_eplan['state'] = 'disabled'
        button_create_xml['state'] = 'disabled'
        button_import_xml['state'] = 'disabled'
        button6['state'] = 'disabled'
        # button_view_xml['state'] = 'disabled'
        button_open['state'] = 'disabled'
        entry_projectnumber.delete(0, "end")  # 删除entry_projectnumber原始内容
        combobox.set('')
        combobox.configure(values=['DZP-TE-AIS', 'DZP-TE-GIS', 'DZP-WEIDMULLER-AIS', 'DZP-WEIDMULLER-GIS', 'DZP-PHOENIX-AIS', 'DZP-PHOENIX-GIS', 'DZP-RELIANCE-AIS', 'DZP-RELIANCE-GIS'])

    text_ebom_import.delete(1.0, tk.END)  # 清空输出结果框

    table_items = SO_Item_table0.get_children()  # 在插入treeview数据时，需要先清空treeview
    [SO_Item_table0.delete(table_item) for table_item in table_items]

    table_items = SO_Item_table1.get_children()  # 在插入treeview数据时，需要先清空treeview
    [SO_Item_table1.delete(table_item) for table_item in table_items]

    state_table_items = SO_Item_table2.get_children()  # 在插入treeview数据时，需要先清空treeview
    [SO_Item_table2.delete(state_table_item) for state_table_item in state_table_items]

    entry.delete(0, "end")  # 删除entry原始内容
    entry.insert(0, filepath)  # 重新填入地址

    global FilePath
    FilePath = filepath


def read_sap_item():
    table_items = SO_Item_table0.get_children()  # 在插入treeview数据时，需要先清空treeview
    [SO_Item_table0.delete(table_item) for table_item in table_items]

    stem, suffix = os.path.splitext(os.path.basename(FilePath))  # stem是文件名,suffix是后缀

    # SAP登录
    sap_key_file = 'J:/Engineering/ShareFolder/new_ABB_Production_Tools/Ps/k/&%&^RD&^T&^WRE^@FEYUGYU@^E@DRYTRYTDFYTWF.txt'
    if not os.path.exists(sap_key_file):
        tk.messagebox.showwarning('提示', 'SAP加密密钥文件不存在')
    else:
        with open(sap_key_file, 'rb') as file:
            sap_key = file.read()
    sap_cipher = Fernet(sap_key)

    sap_ciphertext_file = 'J:/Engineering/ShareFolder/new_ABB_Production_Tools/Ps/s/^&%&^R^YFT$%FTDTEYTKFY$^$FGKJGHE%#$%#EYD^%#^%DTYD.txt'
    if not os.path.exists(sap_ciphertext_file):
        tk.messagebox.showwarning('提示', 'SAP加密密文文件不存在')
    else:
        with open(sap_ciphertext_file, 'r') as file:
            sap_ciphertext = file.read()

    sap_account = sap_decrypt_data(sap_ciphertext, sap_cipher, 'z_')
    sap_h = sap_account.split('+')[0]
    sap_c = sap_account.split('+')[1]
    sap_s = sap_account.split('+')[2]
    sap_u = sap_account.split('+')[3]
    sap_p = sap_account.split('+')[4]

    logging.info("Ready to connect to SAP")
    try:
        conn = pyrfc.Connection(ashost=sap_h, sysnr=sap_s, client=sap_c, user=sap_u, passwd=sap_p, lang='EN')
        if conn.alive:
            logging.info("Connecting to SAP successfully")
            project_number = stem.split('-')[0].split('.')[0]
            result = conn.call('ZY_SALES_ORDER_SHIFT', VBELN='0' + project_number)

            item_data_switchgear_id = []
            item_data_typical = []
            item_data_panel_number = []
            item_data_so_item = []
            item_data_amount = []
            item_data_terminal = []

            if result['EX_CEPTION'] == '' and result['ITAB'][0]['WERKS'] == '1201':
                item_data = []
                for item in result['ITAB']:
                    posnr = item['POSNR'].lstrip('0') or '0'
                    matnr = item['MATNR']
                    arktx = item['ARKTX']
                    tptx1 = item['TPTX1']
                    tptx2 = item.get('TPTX2', '')
                    kwmeng = str(int(float(item['KWMENG'])))
                    item_data.append((posnr, matnr, arktx, tptx1 + tptx2, kwmeng))

                parent_nodes = {}

                # 预定义颜色
                colors = ['even_row', 'odd_row']  # 你可以自定义更多颜色标签

                # 记录上一个插入的代码编号和颜色索引
                last_code_number = None
                color_index = 0
                code_color_map = {}  # 保存 code_number 与颜色的映射

                for item in item_data:
                    posnr = int(item[0])
                    if posnr % 1000 == 0:
                        parent_nodes[posnr] = 1
                    else:
                        parent_id = parent_nodes.get((int(posnr) // 1000) * 1000)
                        if parent_id:
                            # 获取 posnr 的千位数
                            base = (posnr // 1000) * 1000
                            # 计算 posnr 的偏移量
                            offset = posnr - base

                            if 801 <= offset <= 899 and str(item[1]) == 'EBOM':
                                code_number = base // 1000
                                code_str = f"A{code_number:02d}"
                                item_data_switchgear_id.append(f"A{code_number:02d}")
                                item_data_typical.append(str(item[2]))
                                item_data_panel_number.append(str(item[3]))
                                item_data_amount.append(str(item[4]))
                                item_data_so_item.append(posnr % 100)

                                if code_str not in code_color_map:
                                    # 如果这是一个新的 code_number，选择下一个颜色并记录在映射中
                                    code_color_map[code_str] = colors[color_index % len(colors)]
                                    color_index += 1

                                    # 从映射中获取颜色标签
                                tag = code_color_map[code_str]

                                SO_Item_table0.insert("", "end", values=(f"A{code_number:02d}", str(item[2]), str(item[3]), str(item[4]), posnr % 100), tags=tag)

                            if offset == 751 and 'DZP-' in str(item[1]):
                                item_data_terminal.append(str(item[1]))

            conn.close()
            if not conn.alive:
                logging.info("Disconnect from SAP")

            if len(item_data_switchgear_id) == 0:
                tk.messagebox.showwarning("提示", "SAP中尚无EBOM，可随时导入EBOM")
            if len(item_data_terminal) > 0 and len(set(item_data_terminal)) == 1:
                # combobox.configure(values=[item_data_terminal[0]])
                if item_data_terminal[0] not in combobox['values']:
                    tk.messagebox.showwarning("提示", f"端子排类型{combobox['values']}\n\nSAP已有类型{item_data_terminal[0]}，\n\n以端子表中的端子排类型为准")
                else:
                    combobox.set(item_data_terminal[0])
    except pyrfc.RFCError as e:
        logging.info(e.key + ', ' + e.message)
        tk.messagebox.showwarning("提示", traceback.format_exc())


def sap_decrypt_data(data, data_cipher, prefix):
    # 移除列名前缀prefix，如果有的话
    if data.startswith(prefix):
        data = data[2:]

    # Base64 字符串中可能删除了等号，需要添加回原来的填充字符，因为 base64 编码的输出长度应该是 4 的倍数
    padding_needed = 4 - (len(data) % 4)
    if padding_needed:
        data += "=" * padding_needed

    # base64 解码以获取原始的加密字节串
    encrypted_data = base64.urlsafe_b64decode(data.encode())

    # 使用 data_cipher 对象解密，它应该有和加密时一样的密钥
    decrypted_data = data_cipher.decrypt(encrypted_data)

    # 将字节解码回字符串
    decrypted_string = decrypted_data.decode()
    return decrypted_string


# ”读取“按钮
def read_panel_and_table():
    try:
        text_ebom_import.delete(1.0, "end")  # 清空输出结果框

        # 清空treeview表格
        table_items = SO_Item_table1.get_children()  # 在插入treeview数据时，需要先清空treeview
        [SO_Item_table1.delete(table_item) for table_item in table_items]

        if FilePath == "":
            tk.messagebox.showwarning("提示", "请选择文件！")
        stem, suffix = os.path.splitext(os.path.basename(FilePath))  # stem是文件名,suffix是后缀
        if '-BOM' not in stem:
            tk.messagebox.showwarning("提示", "请选择XXX-BOM.xlsx文件！")

        else:
            button_create_xml['state'] = 'normal'
            stem, suffix = os.path.splitext(os.path.basename(FilePath))  # stem是文件名,suffix是后缀
            # os.path.dirname()去掉文件名，返回目录
            # os.path.basename()去掉目录，返回文件名(含后缀)
            truename = stem.rsplit("-", 1)[0]
            inputfile = os.path.join(os.path.dirname(FilePath), truename + '-Panel.xlsx')  # XXX-Panel.xlsx
            global outputfile
            outputfile = os.path.join(os.path.dirname(FilePath), truename + '-Table.xls')  # XXX-Table.xls

            if not os.path.exists(inputfile):
                tk.messagebox.showwarning("提示", "失败,找不到%s文件! " % inputfile.replace("\\", "/"))
            else:
                book = load_workbook(inputfile)
                sheet = book['Z2_xlsx']
                A, B, C = [], [], []
                for i in range(2, sheet.max_row + 1):
                    if str(sheet.cell(row=i, column=1).value) != 'None':
                        A.append(str(sheet.cell(row=i, column=1).value))  # Order Line列
                        B.append(str(sheet.cell(row=i, column=2).value))  # Typical列
                        C.append(str(sheet.cell(row=i, column=3).value))  # Panel No列

                none_count = B.count('None')
                base_dir = os.path.dirname(FilePath)

                rep_path = os.path.join(base_dir, truename + '-Panel-replaced.xlsx')
                replaced_dict = {}  # (A,C):B
                if os.path.exists(rep_path):
                    wb_rep = load_workbook(rep_path)
                    ws_rep = wb_rep['Z2_xlsx']
                    for r in range(2, ws_rep.max_row + 1):
                        a = str(ws_rep.cell(row=r, column=1).value or '').strip()
                        c = str(ws_rep.cell(row=r, column=3).value or '').strip()
                        b = str(ws_rep.cell(row=r, column=2).value or 'None').strip()
                        replaced_dict[(a, c)] = b

                if none_count == len(B):    # 说明是旧版升版2024项目，它的Panel.xlsx报表B列为空，需要用Panel.xls，如果是纯2024项目，只需Panel.xlsx即可
                    xls_path = os.path.join(base_dir, f'{truename}-Panel.xls')

                    if os.path.exists(xls_path):
                        book = xlrd.open_workbook(xls_path)
                        sheet = book.sheet_by_index(0)
                        A = sheet.col_values(colx=0, start_rowx=1, end_rowx=None)  # Order Line列
                        B = sheet.col_values(colx=1, start_rowx=1, end_rowx=None)  # Typical列
                        C = sheet.col_values(colx=2, start_rowx=1, end_rowx=None)  # Panel No列

                    else:
                        if not replaced_dict:
                            tk.messagebox.showwarning('提示', 'Panel.xlsx的B列为空，且Panel.xls与Panel-replaced.xlsx均不存在！请手动补充数据')
                            return

                        for idx in range(len(A)):
                            key = (A[idx], C[idx])
                            B[idx] = replaced_dict.get(key, 'None')
                        tk.messagebox.showwarning('提示', 'Panel.xlsx的B列为空，且Panel.xls不存在，利用Panel-replaced.xlsx补足数据，请仔细核查柜号信息')

                elif 0 < none_count < len(B):
                    if not replaced_dict:
                        tk.messagebox.showwarning('提示', 'Panel.xlsx中部分Typical数据缺失，但Panel-replaced.xlsx不存在！请手动补充数据')
                        return

                    for idx in range(len(A)):
                        if B[idx] == 'None':
                            key = (A[idx], C[idx])
                            B[idx] = replaced_dict.get(key, 'None')
                    tk.messagebox.showwarning('提示', 'Panel.xlsx中部分Typical数据缺失，且Panel.xls不存在，利用Panel-replaced.xlsx补足数据，请仔细核查柜号信息')


                new_A = []
                new_B = []
                new_C = []

                for i in range(len(A)):
                    if i < len(B) and ('空柜' not in B[i] and 'DUMMY' not in B[i].upper() and 'K1' not in B[i].upper()):
                        new_A.append(A[i])
                        new_B.append(B[i])
                        new_C.append(C[i])

                A = new_A
                B = new_B
                C = new_C

                for i in range(0, len(A)):
                    C[i] = C[i] + ';'  # 给ABB柜号后面添加分号

                dataframe_A = pd.DataFrame(A, dtype=str)
                dataframe_B = pd.concat([dataframe_A, pd.DataFrame(B, dtype=str)], axis=1)
                dataframe_C = pd.concat([dataframe_B, pd.DataFrame(C, dtype=str)], axis=1)
                dataframe = dataframe_C
                dataframe.columns = ['OrderLine', 'Typical', 'PanelNo']  # 给各列赋名
                # print(dataframe, '\n\n')

                dataframe_groupby = dataframe.groupby(['OrderLine', 'Typical']).agg({'PanelNo': 'sum'}).reset_index()  # 根据站号和柜型列，将柜列数据进行合并，并重置index
                # print(dataframe_groupby)

                panelamount = []
                for i in range(0, len(dataframe_groupby['OrderLine'])):
                    panelamount.append(dataframe_groupby['PanelNo'][i].count(';'))  # dataframe增加柜数列，数据来源于柜号列中每一个数据中分号(; )个数
                    # print(dataframe_groupby['PanelNo'][i].count('; '))
                dataframe_groupby = pd.concat([dataframe_groupby, pd.DataFrame(panelamount, dtype=str)], axis=1)  # dataframe增加柜数列
                dataframe_groupby.columns = ['OrderLine', 'Typical', 'PanelNo', 'PanelAmount']  # 给各列赋名

                # print(dataframe_groupby, '\n\n')

                orderline_set = list(set(dataframe_groupby['OrderLine']))  # set去重后顺序会改变
                orderline_set.sort(key=list(dataframe_groupby['OrderLine']).index)  # 保证去重后顺序不变

                if not os.path.exists(outputfile):  # 如果配置表格不存在，则SO Item列按默认配置，否则按表格内容配置
                    so_item = [0] * len(dataframe_groupby['OrderLine'])  # 创建指定长度列表
                    for i in range(0, len(orderline_set)):
                        temp = [j for j, k in enumerate(list(dataframe_groupby['OrderLine'])) if k == orderline_set[i]]
                        number = list(range(1, len(temp) + 1))
                        for l in range(0, len(temp)):
                            so_item[temp[l]] = number[l]
                else:
                    book2 = xlrd.open_workbook(outputfile)  # 加载XXX-Table.xls表格
                    worksheet = book2.sheet_by_index(0)

                    table_col_A = worksheet.col_values(colx=0, start_rowx=1, end_rowx=None)
                    table_col_B = worksheet.col_values(colx=1, start_rowx=1, end_rowx=None)
                    table_col_C = worksheet.col_values(colx=2, start_rowx=1, end_rowx=None)
                    table_col_D = worksheet.col_values(colx=3, start_rowx=1, end_rowx=None)
                    table_col_E = worksheet.col_values(colx=4, start_rowx=1, end_rowx=None)

                    tuples_table_col_AB = list(zip(table_col_A, table_col_B))
                    tuples_dataframe_groupby = list(zip(dataframe_groupby['OrderLine'], dataframe_groupby['Typical']))

                    # dataframe_groupby比table_col多的数据及索引
                    add_index = [i for i in range(len(tuples_dataframe_groupby)) if tuples_dataframe_groupby[i] not in tuples_table_col_AB]
                    add_content = [tuples_dataframe_groupby[i] for i in add_index]

                    # dataframe_groupby比table_col少的数据及索引
                    minus_index = [i for i in range(len(tuples_table_col_AB)) if tuples_table_col_AB[i] not in tuples_dataframe_groupby]
                    minus_content = [tuples_table_col_AB[i] for i in minus_index]

                    # print(add_index, add_content)
                    # print(minus_index, minus_content)
                    if len(add_index) == 0 and len(minus_index) == 0:    # 无新增Typical
                        so_item = table_col_E    # 仍取原来的Table中的SO Item
                    elif len(add_index) > 0 and len(minus_index) == 0:    # 新增Typical
                        so_item = table_col_E
                        new_typical = []
                        for n in range(0, len(add_index)):
                            new_typical.append(add_content[n])
                            so_item.insert(add_index[n], '0')    # 新增的Typical的SO Item设置为0

                        tk.messagebox.showwarning("提示", f"新增柜型{new_typical}\n请维护SO Item(默认0)！")

                    elif len(add_index) == 0 and len(minus_index) > 0:    # 减少Typical
                        so_item = table_col_E
                        lack_typical = []
                        for i, n in enumerate(sorted(minus_index, reverse=True)):  # 使用enumerate来跟踪索引和值
                            lack_typical.append(minus_content[i])
                            del so_item[n]

                        tk.messagebox.showwarning("提示", f"减少柜型{lack_typical}\n自动移除对应SO Item！")

                    elif len(add_index) > 0 and len(minus_index) > 0:  # 有增有减Typical
                        so_item = table_col_E
                        new_typical = []
                        lack_typical = []

                        # 首先处理减少部分
                        for i, n in enumerate(sorted(minus_index, reverse=True)):  # 使用enumerate来跟踪索引和值
                            lack_typical.append(minus_content[i])
                            del so_item[n]

                        # 处理新增部分
                        for n in range(len(add_index)):
                            new_typical.append(add_content[n])
                            so_item.insert(add_index[n], '0')  # 新增的Typical的SO Item设置为0

                        tk.messagebox.showwarning("提示", f"新增柜型{new_typical}\n请维护SO Item(默认0)！\n\n减少柜型{lack_typical}\n自动移除对应SO Item！")

                # 完整的dataframe数据
                dataframe_groupby = pd.concat([dataframe_groupby, pd.DataFrame(so_item, dtype=str)], axis=1)  # dataframe增加SO Item列
                dataframe_groupby.columns = ['OrderLine', 'Typical', 'PanelNo', 'PanelAmount', 'SOItem']  # 给各列赋名

                # 清空treeview表格
                table_items = SO_Item_table1.get_children()  # 在插入treeview数据时，需要先清空treeview
                [SO_Item_table1.delete(table_item) for table_item in table_items]

                # 预定义颜色
                colors = ['even_row', 'odd_row']  # 你可以自定义更多颜色标签

                # 记录上一个插入的 OrderLine 和颜色索引
                last_order_line = None
                color_index = 0
                order_line_color_map = {}  # 保存 OrderLine 与颜色的映射

                # 为treeview表格插入数据
                for i in range(0, len(dataframe_groupby['OrderLine'])):
                    order_line = dataframe_groupby['OrderLine'][i]

                    if order_line not in order_line_color_map:
                        # 如果这是一个新的 OrderLine，选择下一个颜色并记录在映射中
                        order_line_color_map[order_line] = colors[color_index % len(colors)]
                        color_index += 1

                    # 从映射中获取颜色标签
                    tag = order_line_color_map[order_line]

                    if dataframe_groupby['SOItem'][i] == 0:
                        SO_Item_table1.insert('', 'end', values=list(dataframe_groupby.loc[i]), tags=('attention_row', tag))  # dataframe逐行插入到表格中
                    else:
                        SO_Item_table1.insert('', 'end', values=list(dataframe_groupby.loc[i]), tags=(tag,))  # dataframe逐行插入到表格中

                # 完整的数据导入Table表格
                book2 = xlwt.Workbook()  # 创建一个空文件对象
                book2.add_sheet('Sheet1')  # 创建一个Sheet页
                book2.save(outputfile)  # XXX-Table.xls
                book2 = xlrd.open_workbook(outputfile)  # 加载XXX-Table.xls表格

                workbook = copy(book2)  # 使用xlutils.copy将xlrd读取的book对象转为xlwt可操作对象
                worksheet = workbook.get_sheet(0)  # 获取sheet
                worksheet.write(0, 0, 'OrderLine')  # 在sheet指定位置写入数据
                worksheet.write(0, 1, 'Typical')  # 在sheet指定位置写入数据
                worksheet.write(0, 2, 'PanelNo')  # 在sheet指定位置写入数据
                worksheet.write(0, 3, 'PanelAmount')  # 在sheet指定位置写入数据
                worksheet.write(0, 4, 'SOItem')  # 在sheet指定位置写入数据

                style = xlwt.XFStyle()
                style.num_format_str = '@'  # 设置数据格式为文本

                for i in range(0, len(dataframe_groupby['OrderLine'])):
                    worksheet.write(i + 1, 0, dataframe_groupby['OrderLine'][i], style=style)
                    worksheet.write(i + 1, 1, dataframe_groupby['Typical'][i], style=style)
                    worksheet.write(i + 1, 2, dataframe_groupby['PanelNo'][i], style=style)
                    worksheet.write(i + 1, 3, dataframe_groupby['PanelAmount'][i], style=style)
                    worksheet.write(i + 1, 4, dataframe_groupby['SOItem'][i], style=style)

                workbook.save(outputfile)  # 将workbook保存到指定位置
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())

def load_replaced(rep_path, old_A, old_C):
    """完全用 replaced 文件内容替换"""
    wb = load_workbook(rep_path)
    ws = wb.active
    new_A, new_B, new_C = [], [], []
    for r in range(2, ws.max_row + 1):
        a_val = ws.cell(row=r, column=1).value
        if a_val is None:
            continue
        new_A.append(str(a_val).strip())
        new_B.append(str(ws.cell(row=r, column=2).value).strip()
                     if ws.cell(row=r, column=2).value is not None else 'None')
        new_C.append(str(ws.cell(row=r, column=3).value).strip()
                     if ws.cell(row=r, column=3).value is not None else 'None')
    return new_A, new_B, new_C

def fill_missing_from_replaced(rep_path, A, B, C):
    """仅补全 B 列为 None 的行"""
    wb = load_workbook(rep_path)
    ws = wb.active
    # 先把 replaced 建成 (A,C)→B 的字典
    lookup = {}
    for r in range(2, ws.max_row + 1):
        a_val = str(ws.cell(row=r, column=1).value or '').strip()
        c_val = str(ws.cell(row=r, column=3).value or '').strip()
        b_val = str(ws.cell(row=r, column=2).value or 'None').strip()
        lookup[(a_val, c_val)] = b_val

    # 遍历当前缺失行
    for idx in range(len(B)):
        if B[idx] == 'None':
            key = (A[idx], C[idx])
            if key in lookup:
                B[idx] = lookup[key]
    return A, B, C

def discontinue_command():
    global double_check_bom_no_problem_flag
    double_check_bom_no_problem_flag = False
    top.destroy()
    root_win.attributes("-disabled", 0)

def continue_command():
    global double_check_bom_no_problem_flag
    double_check_bom_no_problem_flag = True
    top.destroy()
    root_win.attributes("-disabled", 0)
    generate_xml_files()

def allow_main_window():
    top.destroy()
    root_win.attributes("-disabled", 0)


# ”创建“按钮
def create_xml_file():
    try:
        text_ebom_import.delete(1.0, "end")  # 清空输出结果框

        # 清空treeview表格
        state_table_items = SO_Item_table2.get_children()  # 在插入treeview数据时，需要先清空treeview
        [SO_Item_table2.delete(state_table_item) for state_table_item in state_table_items]

        if FilePath == "":
            tk.messagebox.showwarning("提示", "请选择文件！")
        stem, suffix = os.path.splitext(os.path.basename(FilePath))  # stem是文件名,suffix是后缀
        if '-BOM' not in stem:
            tk.messagebox.showwarning("提示", "请选择XXX-BOM.xlsx文件！")

        else:
            # os.path.dirname()去掉文件名，返回目录
            # os.path.basename()去掉目录，返回文件名(含后缀)

            truename = stem.rsplit("-", 1)[0]  # 项目号
            outputfile = os.path.join(os.path.dirname(FilePath), truename + '-Table.xls')  # XXX-Table.xls

            # BOM表
            wbook1 = load_workbook(FilePath)
            wsheet1 = wbook1['Z6_xlsx']
            global wA1, wB1, wC1, wD1, wE1, wF1, wG1, wH1, wI1, wJ1
            global wA2, wB2, wC2, wD2, wE2

            wA1 = []
            wB1 = []
            wC1 = []
            wD1 = []
            wE1 = []
            wF1 = []
            wG1 = []    # 新增
            wH1 = []    # 新增
            wI1 = []    # 新增
            wJ1 = []    # 新增
            for i in range(2, wsheet1.max_row + 1):
                wA1.append(str(wsheet1.cell(row=i, column=1).value))  # Hight-level列
                wB1.append(str(wsheet1.cell(row=i, column=2).value))  # Zone列
                wC1.append(str(wsheet1.cell(row=i, column=3).value))  # DT列
                wD1.append(str(wsheet1.cell(row=i, column=4).value))  # PartNumber列
                wE1.append(str(wsheet1.cell(row=i, column=5).value))  # Qty列
                wF1.append(str(wsheet1.cell(row=i, column=6).value))  # Designation列
                wG1.append(str(wsheet1.cell(row=i, column=7).value))  # Type列
                wH1.append(str(wsheet1.cell(row=i, column=8).value))  # Description列
                wI1.append(str(wsheet1.cell(row=i, column=9).value))  # Exist列
                wJ1.append(str(wsheet1.cell(row=i, column=10).value))  # Exist1列

            # 配置信息表
            wbook2 = xlrd.open_workbook(outputfile, formatting_info=False)  # XXX-Table.xls
            wsheet2 = wbook2.sheet_by_index(0)
            wA2 = wsheet2.col_values(colx=0, start_rowx=1, end_rowx=None)  # OrderLine列
            wB2 = wsheet2.col_values(colx=1, start_rowx=1, end_rowx=None)  # Typical列
            wC2 = wsheet2.col_values(colx=2, start_rowx=1, end_rowx=None)  # PanelNo列
            wD2 = wsheet2.col_values(colx=3, start_rowx=1, end_rowx=None)  # PanelAmount列
            wE2 = wsheet2.col_values(colx=4, start_rowx=1, end_rowx=None)  # SOItem列
            global projectID
            projectID = entry_projectnumber.get()  # 读取文本框内的项目号

            Error_flag1 = 0  # 同站下SO Item重复（错误）标志位
            for ii in range(0, len(wE2)):
                for jj in range(0, len(wE2)):
                    if ii != jj:
                        if wE2[ii] == wE2[jj] and wA2[ii] == wA2[jj]:  # 出现了同站下SO Item重复
                            Error_flag1 = 1
            Error_flag2 = 0  # SO Item非法（错误）标志位
            for ii in range(0, len(wE2)):
                if wE2[ii] == '0' or '-' in wE2[ii] or '.' in wE2[ii] or (not wE2[ii].isdigit()) or wE2[ii][0] == '0':
                    Error_flag2 = 1

            content_exist_list = []
            for ii in range(0, len(wA2)):
                content_exist = 0
                for jj in range(0, len(wA1)):
                    if wB2[ii] == wA1[jj]:
                        content_exist = 1
                content_exist_list.append(content_exist)

            if '-BOM' not in stem:
                tk.messagebox.showwarning("提示", "请选择XXX-BOM.xlsx文件！")
            elif combobox.get() == '':  # 未选取端子排类型
                tk.messagebox.showwarning("提示", "请选择端子排类型，用于创建Item:1751...！")
            elif not os.path.exists(outputfile):  # 如果Table.xls不存在（通常还未导初级BOM）
                tk.messagebox.showwarning("提示", "请读取配置信息！")
            elif strftime('%Y-%m-%d %H:%M', localtime(os.stat(outputfile).st_mtime)) != strftime('%Y-%m-%d %H:%M', localtime()):  # 读取XXX-Table.xls修改时间，点击“创建”的时间，如果Table.xls修改时间是旧的，跟点击“创建”的实际时间有差异
                tk.messagebox.showwarning("提示", "请读取配置信息！")
            elif Error_flag2:
                tk.messagebox.showwarning("提示", "出现SO Item为0/负数/小数/非数值的情况，请维护为正整数！")
            elif Error_flag1:
                tk.messagebox.showwarning("提示", "出现同站SO Item重复的情况，请维护为同站非重复的正整数！")
            elif projectID == "" or len(projectID) < 8 or len(projectID) > 9:
                tk.messagebox.showwarning("提示", "请输入9位/8位样柜项目号！")
            elif len(projectID) == 8 and projectID[0] != '7':
                tk.messagebox.showwarning("提示", "8位样柜项目号，只允许数字7打头！")
            elif not projectID.isdigit():
                tk.messagebox.showwarning("提示", "请输入纯数字项目号！")
            elif 0 in content_exist_list:
                result = [wB2[i] for i, data in enumerate(content_exist_list) if data == 0]
                tk.messagebox.showwarning("提示", "BOM表格中,Typical：%s的物料为空\n\n请查看并重新调整EPLAN结构，确保物料能够刷出" % ', '.join(list(set(result))))

            else:
                switchgearnumber_no_problem_flag = False
                if len(projectID) == 8 and projectID[0] == '7':
                    confirm_switchgearnumber = tk.messagebox.askquestion("提示", "请确认配置信息中站号是否正确？\n\nA01代表1000行，A02代表2000行，如75150135.14对应的应是A14\n\n如果错误，点击No，修改EPLAN结构重新导BOM\n\n如果正确，点击Yes")
                    if confirm_switchgearnumber == 'yes':
                        switchgearnumber_no_problem_flag = True
                else:
                    switchgearnumber_no_problem_flag = True

                if switchgearnumber_no_problem_flag:
                    global double_check_bom_no_problem_flag
                    double_check_bom_no_problem_flag = False
                    root_win.attributes("-disabled", 1)
                    global top
                    top = tk.Toplevel(bg="#eaf1f6", bd=2, borderwidth=2)
                    top.protocol("WM_DELETE_WINDOW", allow_main_window)
                    winw = 600
                    winh = 800
                    top.geometry("%dx%d" % (winw, winh))
                    top.title('导EBOM二次确认')
                    center_window(top)

                    tk.Frame(top, height=1, bg="#eaf1f6").pack(fill=tk.X)
                    f1 = tk.Frame(top, bg="#eaf1f6", bd=0)
                    f2 = tk.Frame(top, bg="#eaf1f6", bd=0)
                    f3 = tk.Frame(top, bg="#eaf1f6", bd=0)
                    f4 = tk.Frame(top, bg="#eaf1f6", bd=0)
                    f5 = tk.Frame(top, bg="#eaf1f6", bd=0)

                    tk.Label(f1, text='下表物料不在EPLAN/SAP库中 or 已经打删除标记，请替换', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
                    tk.Label(f3, text='下表物料80属性有标记（不采购 or 已删除 or 属于一次柜体BOM），将不会导入EBOM，请核对', bg="#eaf1f6", fg="black", font=("ABBvoice CNSG", int(12 * h_ratio)), justify='left').pack(side=tk.LEFT, fill=tk.X)
                    global deletion_material_tree
                    deletion_material_tree = Treeview(f2, show="headings", columns=['a', 'b', 'c', 'd', 'e'], style='panel1.Treeview', selectmode='browse', height=9)
                    deletion_material_tree.heading('a', text='Typical', anchor='center')
                    deletion_material_tree.heading('b', text='定位', anchor='center')
                    deletion_material_tree.heading('c', text='标识', anchor='center')
                    deletion_material_tree.heading('d', text='物料号', anchor='center')
                    deletion_material_tree.heading('e', text='数量', anchor='center')
                    deletion_material_tree.column('a', width=100, anchor='center')
                    deletion_material_tree.column('b', width=100, anchor='center')
                    deletion_material_tree.column('c', width=100, anchor='center')
                    deletion_material_tree.column('d', width=150, anchor='center')
                    deletion_material_tree.column('e', width=100, anchor='center')
                    deletion_material_tree.pack(side=tk.TOP, fill=tk.X, expand=True, anchor='w', padx=20)
                    Tooltip(deletion_material_tree, '该表格内物料号均有删除标记，需要更换物料号')

                    global no_import_sap_tree
                    no_import_sap_tree = Treeview(f4, show="headings", columns=['a', 'b', 'c', 'd', 'e'], style='panel1.Treeview', selectmode='browse', height=15)
                    no_import_sap_tree.heading('a', text='Typical', anchor='center')
                    no_import_sap_tree.heading('b', text='定位', anchor='center')
                    no_import_sap_tree.heading('c', text='标识', anchor='center')
                    no_import_sap_tree.heading('d', text='物料号', anchor='center')
                    no_import_sap_tree.heading('e', text='数量', anchor='center')
                    no_import_sap_tree.column('a', width=100, anchor='center')
                    no_import_sap_tree.column('b', width=100, anchor='center')
                    no_import_sap_tree.column('c', width=100, anchor='center')
                    no_import_sap_tree.column('d', width=150, anchor='center')
                    no_import_sap_tree.column('e', width=100, anchor='center')
                    no_import_sap_tree.pack(side=tk.TOP, fill=tk.X, expand=True, anchor='w', padx=20)
                    Tooltip(no_import_sap_tree, '该表格内物料不会导入EBOM，请核对')

                    f1.pack(side=tk.TOP, fill=tk.X, pady=10, padx=20)
                    f2.pack(side=tk.TOP, fill=tk.X)
                    f3.pack(side=tk.TOP, fill=tk.X, pady=10, padx=20)
                    f4.pack(side=tk.TOP, fill=tk.X)
                    f5.pack(side=tk.TOP, fill=tk.X, pady=10)
                    global continue_flag
                    continue_flag = False

                    continue_button = tk.Button(f5, text='核对完成，确定创建BOM.xml', compound=tk.LEFT, bg="#eaf1f6", font=("ABBvoice CNSG", int(12 * h_ratio)), cursor='hand2', state='disabled', command=continue_command)
                    continue_button.pack(side=tk.RIGHT, fill=tk.BOTH, padx=20)
                    Tooltip(continue_button, '确保第一个表中数据为空，该按钮才能被点击')

                    discontinue_button = tk.Button(f5, text='返回', compound=tk.LEFT, bg="#eaf1f6", font=("ABBvoice CNSG", int(12 * h_ratio)), cursor='hand2', command=discontinue_command)
                    discontinue_button.pack(side=tk.RIGHT, fill=tk.BOTH, padx=50)

                    deletion_index = []
                    no_import_index = []
                    real_no_import_index = []
                    for i in range(0, len(wA1)):
                        if (wF1[i] == '' and wG1[i] == '' and wH1[i] == '') or (wI1[i] == 'X' or wI1[i] == 'x'):
                            deletion_index.append(i)
                            continue

                        if wJ1[i] == 'YES':    # J列为YES（或不为空），物料不导入
                            real_no_import_index.append(i)
                            continue

                        if wD1[i] == 'Mounting_Panel' or 'EP-' in wD1[i]:
                            real_no_import_index.append(i)
                            continue

                        # if wJ1[i] != 'None' and wD1[i] != 'Mounting_Panel' and 'EP-' not in wD1[i]:
                        #     no_import_index.append(i)    # J列为YES（或不为空）且D列不为Mounting_Panel且D列不含EP-开头，这一类的物料是断路器，地刀，PT车等，虽然确定也不导入，但要提示核对
                        # # no_import_index是real_no_import_index的子集

                    if deletion_index:    # 只要有条目，就无法往下进行，需要EPLAN物料修改或删除
                        for i in deletion_index:
                            deletion_material_tree.insert('', 'end', values=(wA1[i], wB1[i], wC1[i], wD1[i], wE1[i]))
                        continue_button['state'] = 'disabled'
                    else:
                        continue_button['state'] = 'normal'

                    if real_no_import_index:
                        for i in real_no_import_index:
                            no_import_sap_tree.insert('', 'end', values=(wA1[i], wB1[i], wC1[i], wD1[i], wE1[i]))

                    if real_no_import_index:
                        # 1. 先排序：把索引从大到小排列，防止 pop 后索引错位
                        real_no_import_index.sort(reverse=True)
                        # 2. 倒序删除
                        for idx in real_no_import_index:
                            if 0 <= idx < len(wA1):  # 安全保护
                                wA1.pop(idx)
                                wB1.pop(idx)
                                wC1.pop(idx)
                                wD1.pop(idx)
                                wE1.pop(idx)
                                wF1.pop(idx)
                                wG1.pop(idx)
                                wH1.pop(idx)
                                wI1.pop(idx)
                                wJ1.pop(idx)

                    # for i in range(0, len(wA1)):
                    #     print(wA1[i], wB1[i], wC1[i], wD1[i], wE1[i],
                    #                     wF1[i], wG1[i], wH1[i], wI1[i], wJ1[i])

    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


def generate_xml_files():
    text_ebom_import.insert(tk.INSERT, ">>>已选择端子排类型：%s\n" % combobox.get())
    text_ebom_import.insert(tk.INSERT, ">>>正在创建EBOM的xml文件...\n")
    start = time()

    # 根据outputfile的每一行信息创建对应的xml文件，内容结合对应柜型所对应的BOM
    # 其他变量
    Creatername = os.getlogin()  # 用户名

    global xml_TargetPath_list
    xml_TargetPath_list = []

    for i in range(0, len(wA2)):
        global truename
        xml_TargetPath = os.path.join(os.path.dirname(FilePath), truename + '_' + wA2[i] + '_' + wB2[i] + '.xml')  # XXX_A01_I1.xml
        xml_TargetPath_list.append(xml_TargetPath)  # 将EBOM的xml文件地址存入列表，以方便导入到指定文件夹
        if os.path.exists(xml_TargetPath):  # 如果对应.xml文件存在，则删除
            os.remove(xml_TargetPath)
        # print(xml_TargetPath)

        # BOM信息变量
        Zonelist = []
        DTlist = []
        PartNumberlist = []
        Qtylist = []
        Designationlist = []

        Zonelist_sort = []
        DTlist_sort = []
        PartNumberlist_sort = []
        Qtylist_sort = []
        Designationlist_sort = []

        for j in range(0, len(wA1)):
            if wA1[j] == wB2[i]:
                if wB1[j] != 'A':
                    Zonelist.append(wB1[j])
                    DTlist.append(wC1[j])
                    PartNumberlist.append(wD1[j])
                    Qtylist.append(wE1[j])
                    Designationlist.append(wF1[j])
        PartNumberlist_index = np.argsort(PartNumberlist)  # 对一个列表进行排序，返回排序后的索引
        # print(PartNumberlist_index)
        for idx in PartNumberlist_index:  # 根据字母从小到大排序索引，对所有序列进行排序，新列表为XXX_sort
            PartNumberlist_sort.append(PartNumberlist[idx])
            Zonelist_sort.append(Zonelist[idx])
            DTlist_sort.append(DTlist[idx])
            Qtylist_sort.append(Qtylist[idx])
            Designationlist_sort.append(Designationlist[idx])
        # print(PartNumberlist_sort)

        # 创建根元素BDD
        global xml_doc
        xml_doc = Document()  # 创建DOM文档对象
        xml_BDD = xml_doc.createElement('BDD')  # 创建根元素
        xml_BDD.setAttribute('xmlns:xsi', 'http://www.w3.org/2001/XMLSchema-instance')
        xml_BDD.setAttribute('xmlns:xsd', 'http://www.w3.org/2001/XMLSchema')
        xml_BDD.setAttribute('xmlns:msxsl', 'urn:schemas-microsoft-com:xslt')
        xml_BDD.setAttribute('xmlns:mve', 'urn:mve-scripts')
        xml_doc.appendChild(xml_BDD)

        # 创建BDD的子元素BSR
        xml_BSR = xml_doc.createElement('BSR')
        xml_BDD.appendChild(xml_BSR)

        Timestamp = strftime('%Y-%m-%dT%H:%M:%S', localtime())  # 时间戳

        # 创建BSR的子元素Timestamp/Verb/Noun/Message_sender/Creator
        makexml('Timestamp', Timestamp, xml_BSR)

        # 等价于
        # xml_Timestamp = xml_doc.createElement('Timestamp')
        # xml_Timestamptext = xml_doc.createTextNode(Timestamp)
        # xml_Timestamp.appendChild(xml_Timestamptext)
        # xml_BSR.appendChild(xml_Timestamp)

        makexml('Verb', 'Load', xml_BSR)
        makexml('Noun', 'BOM', xml_BSR)
        makexml('Message_sender', 'MVE', xml_BSR)
        makexml('Creator', str(Creatername), xml_BSR)

        # 创建BDD的子元素item
        xml_item = xml_doc.createElement('item')
        xml_BDD.appendChild(xml_item)

        # 创建item的子元素SalesOrder/Plant/CI/C750/C751
        makexml('SalesOrder', str(projectID), xml_item)
        makexml('Plant', '1201', xml_item)
        makexml('CI', str(wA2[i]), xml_item)
        makexml('C750', 'NAMEPLATE', xml_item)
        makexml('C751', str(combobox.get()), xml_item)

        # 创建BDD的子元素Typical
        xml_Typical = xml_doc.createElement('Typical')
        xml_BDD.appendChild(xml_Typical)

        # 创建Typical的子元素Delivery
        xml_Delivery = xml_doc.createElement('Delivery')
        xml_Typical.appendChild(xml_Delivery)

        # 创建Delivery的子元素DeliveryId/DeliveryInfo
        makexml1('DeliveryId', 'Domain', 'BOMRevision', '22222', xml_Delivery)
        makexml1('DeliveryId', 'Domain', 'BOMCategory', 'M', xml_Delivery)
        makexml1('DeliveryId', 'Domain', 'BOMUsage', '1', xml_Delivery)
        makexml1('DeliveryId', 'Domain', 'BOMPlant', '1201', xml_Delivery)
        makexml1('DeliveryId', 'Domain', 'EBOMParent', 'EBOM', xml_Delivery)
        makexml1('DeliveryId', 'Domain', 'BOMDescription2', str(wB2[i]), xml_Delivery)
        makexml1('DeliveryId', 'Domain', 'PanelAmount', str(wD2[i]), xml_Delivery)
        makexml1('DeliveryId', 'Domain', 'BOMDescription1', str(wC2[i]), xml_Delivery)
        makexml1('DeliveryId', 'Domain', 'SalesOrder', str(projectID), xml_Delivery)
        makexml1('DeliveryId', 'Domain', 'SalesOrderItem', str(wE2[i]), xml_Delivery)
        makexml1('DeliveryId', 'Domain', 'CNID', str(wA2[i]), xml_Delivery)
        makexml1('DeliveryInfo', 'Dimension', 'switchgear', '', xml_Delivery)
        makexml1('DeliveryInfo', 'Dimension', 'typical', str(wB2[i]), xml_Delivery)

        # 创建Typical的子元素Parts
        xml_Parts = xml_doc.createElement('Parts')
        xml_Typical.appendChild(xml_Parts)

        for k in range(0, len(PartNumberlist_sort)):
            # 创建Parts的子元素Part
            xml_Part = xml_doc.createElement('Part')
            xml_Parts.appendChild(xml_Part)

            # 创建Part子元素PartId
            xml_PartID = xml_doc.createElement('PartId')
            xml_Part.appendChild(xml_PartID)
            makexml1('IdentNo', 'Domain', 'LineNo', str((k + 1) * 10), xml_PartID)
            makexml1('IdentNo', 'Domain', 'Material', str(PartNumberlist_sort[k]), xml_PartID)
            makexml1('IdentNo', 'Domain', 'ItemCategory', 'N', xml_PartID)

            # 创建Part子元素PartInfo
            xml_PartInfo = xml_doc.createElement('PartInfo')
            xml_Part.appendChild(xml_PartInfo)
            makexml1('Value', 'Dimension', 'MaterialType', 'FERT', xml_PartInfo)
            makexml1('Value', 'Dimension', 'FactoryID', 'ABB', xml_PartInfo)
            makexml2('Value', 'Dimension', 'Qty', 'Unit', 'Piece', str(Qtylist_sort[k]), xml_PartInfo)
            makexml1('Value', 'Dimension', 'DescriptionLanguage', 'ZH', xml_PartInfo)
            makexml2('Value', 'Dimension', 'Dimension', 'Type', 'DeviceTags', '-' + str(DTlist_sort[k]), xml_PartInfo)
            makexml2('Value', 'Dimension', 'Dimension', 'Type', 'Zone', str(Zonelist_sort[k]), xml_PartInfo)
            makexml1('Value', 'Dimension', 'DescriptionLocal', str(Designationlist_sort[k]), xml_PartInfo)
            makexml1('Value', 'Dimension', 'ERPNumber', '', xml_PartInfo)
        with open(xml_TargetPath, 'w', encoding='utf-8') as f:
            xml_doc.writexml(f, indent='', addindent='  ', newl='\n', encoding='utf-8')
            f.close()

        # writer: 文件对象
        # indent: 是每个tag前填充的字符，如“    ”表示每个tag前有四个空格
        # addindent: 是每个子结点的缩进字符
        # newl: 是每个tag后填充的字符，如“\n”表示每个tag后面有一个回车
        # encoding: 是生成的XML信息头中的encoding属性值，在输出时minidom并不真正进行编码的处理,若保存的文本内容有汉字，则需要进行编码转换

        text_ebom_import.insert(tk.INSERT, "%s已创建!\n" % xml_TargetPath)

    end = time()
    text_ebom_import.insert(tk.INSERT, ">>>所有xml文件均已创建，请确认导入!  用时%.3f秒\n" % (end - start))
    button_import_xml['state'] = 'normal'  # 当所有文件创建成功，导入按钮可以点击

    # V1.9增加，将端子排存入txt文档
    stem, suffix = os.path.splitext(os.path.basename(FilePath))  # stem是文件名,suffix是后缀
    truename = stem.rsplit("-", 1)[0]  # 项目号
    outputfile = os.path.join(os.path.dirname(FilePath), truename + '-DZP.txt')  # XXX-DZP.txt
    with open(outputfile, "w") as file:
        file.write(combobox.get())


# ”导入“按钮
def import_xml_file():
    try:
        if tk.messagebox.askyesno("提示", "要将所有的BOM.xml导入send文件夹吗？"):
            text_ebom_import.delete(1.0, "end")  # 清空输出结果框
            start = time()
            text_ebom_import.insert(tk.INSERT, ">>>准备移入\\\\CN-S-041MVE1\\temp\exportToSAP\send\n")
            button_import_xml['state'] = 'disabled'  # 当所有文件移入公共盘，导入按钮不可点击
            for src in xml_TargetPath_list:
                stem, suffix = os.path.splitext(os.path.basename(src))  # stem是文件名,suffix是后缀
                dst = '\\\\CN-S-041MVE1\\temp\exportToSAP\send\\' + stem + suffix  # send文件夹地址
                shutil.move(src, dst)
                text_ebom_import.insert(tk.INSERT, "%s移入成功！\n" % dst)
            end = time()
            text_ebom_import.insert(tk.INSERT, ">>>所有xml文件均已移入，请等待导入SAP!  用时%.3f秒\n" % (end - start))

            global checklog_file_path
            checklogbook = load_workbook(checklog_file_path)
            checklogsheet = checklogbook['Sheet']
            project_no = stem.replace('-BOM', '')
            current_time = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            data1 = [project_no, 'BOM导入SAP', 0, current_time]
            checklogsheet.append(data1)
            checklogbook.save(checklog_file_path)
    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())


# ”刷新“按钮
def update_bom_import_state():
    try:
        # 清空treeview表格
        state_table_items = SO_Item_table2.get_children()  # 在插入treeview数据时，需要先清空treeview
        [SO_Item_table2.delete(state_table_item) for state_table_item in state_table_items]

        if FilePath == "":
            tk.messagebox.showwarning("提示", "请选择文件！")

        stem, suffix = os.path.splitext(os.path.basename(FilePath))  # stem是文件名,suffix是后缀
        truename = stem.rsplit("-", 1)[0]  # 项目号
        outputfile = os.path.join(os.path.dirname(FilePath), truename + '-Table.xls')  # XXX-Table.xls

        if '-BOM' not in stem:
            tk.messagebox.showwarning("提示", "请选择XXX-BOM.xlsx文件！")

        elif not os.path.exists(outputfile):  # 如果Table.xls不存在（通常还未导初级BOM）
            tk.messagebox.showwarning("提示", "请读取配置信息！")

        else:

            # 读取配置信息表
            wbook2 = xlrd.open_workbook(outputfile, formatting_info=False)  # XXX-Table.xls
            wsheet2 = wbook2.sheet_by_index(0)
            wA2 = wsheet2.col_values(colx=0, start_rowx=1, end_rowx=None)  # OrderLine列
            wB2 = wsheet2.col_values(colx=1, start_rowx=1, end_rowx=None)  # Typical列
            wC2 = wsheet2.col_values(colx=2, start_rowx=1, end_rowx=None)  # PanelNo列
            wD2 = wsheet2.col_values(colx=3, start_rowx=1, end_rowx=None)  # PanelAmount列
            wE2 = wsheet2.col_values(colx=4, start_rowx=1, end_rowx=None)  # SOItem列

            global xml_TargetName_list  # xml文件名列表
            xml_TargetName_list = []
            global xml_ResponseName_list  # xml的响应文件名列表
            xml_ResponseName_list = []
            global xml_ResponsePath_list  # xml的响应文件地址列表
            xml_ResponsePath_list = []

            for i in range(0, len(wA2)):
                xml_TargetName = str(truename + '_' + wA2[i] + '_' + wB2[i] + '.xml')  # XXX_A01_I1.xml
                xml_TargetName_list.append(xml_TargetName.replace('.xml', ''))  # 将EBOM的xml文件名存入列表，用于表格显示

                xml_ResponseName = str('response_' + truename + '_' + wA2[i] + '_' + wB2[i] + '.xml')  # response_XXX_A01_I1.xml
                xml_ResponseName_list.append(xml_ResponseName)  # 将EBOM的响应xml文件名存入列表，用于表格显示

                xml_ResponsePath = '\\\\CN-S-041MVE1\\temp\exportToSAP\\receive\\' + xml_ResponseName  # receive文件夹地址
                xml_ResponsePath_list.append(xml_ResponsePath)  # 将EBOM的响应xml文件地址存入列表，用于文件监视

            # state_dataframe_A = pd.DataFrame(xml_TargetName_list, dtype=str)
            # state_dataframe_B = pd.concat([state_dataframe_A, pd.DataFrame(xml_ResponseName_list, dtype=str)], axis=1)    # 将xml文件名和xml响应文件名装入
            state_dataframe_B = pd.DataFrame(xml_ResponseName_list, dtype=str)

            state_list = []
            time_list = []
            message_list = []
            for i in range(0, len(xml_TargetName_list)):
                if not os.path.exists(xml_ResponsePath_list[i]):
                    state_list.append('-')
                    time_list.append('')
                    message_list.append('')
                else:

                    with open(xml_ResponsePath_list[i], 'r', encoding='utf-8') as fo:
                        xml_response_doc = parse(fo)
                        fo.close()
                    rootNode = xml_response_doc.documentElement  # 根元素BDD

                    if rootNode.childNodes:
                        for j in range(0, len(rootNode.childNodes)):
                            first_sonNode = rootNode.childNodes[j]
                            if first_sonNode.nodeName == 'BSR':  # 返回儿子元素BSR
                                for k in range(0, len(first_sonNode.childNodes)):
                                    first_grandsonNode = first_sonNode.childNodes[k]
                                    second_grandsonNode = first_sonNode.childNodes[k]
                                    if first_grandsonNode.nodeName == 'Timestamp':
                                        time_list.append(str(first_grandsonNode.childNodes[0].data.replace('T', ' ')))

                                    if second_grandsonNode.nodeName == 'Verb':
                                        # print(second_grandsonNode.attributes.items())    # [('Qualifier', 'Fail')]   [('Qualifier', 'Pass')]
                                        state_list.append(second_grandsonNode.attributes.items()[0][1])
                                        if second_grandsonNode.childNodes:
                                            message_list.append(second_grandsonNode.childNodes[0].data)
                                        else:
                                            message_list.append('')

                    else:
                        state_list.append('-')
                        time_list.append('')
                        message_list.append('')

            state_dataframe_C = pd.concat([state_dataframe_B, pd.DataFrame(state_list, dtype=str)], axis=1)  # 将状态装入
            state_dataframe_D = pd.concat([state_dataframe_C, pd.DataFrame(time_list, dtype=str)], axis=1)  # 将时间戳装入
            state_dataframe_E = pd.concat([state_dataframe_D, pd.DataFrame(message_list, dtype=str)], axis=1)  # 将消息装入
            state_dataframe = state_dataframe_E
            # state_dataframe.columns = ['xmlTargetName', 'xmlResponseName', 'xmlState', 'xmlTimestamp', 'xmlMessage']  # 给各列赋名
            state_dataframe.columns = ['xmlResponseName', 'xmlState', 'xmlTimestamp', 'xmlMessage']  # 给各列赋名

            # 为treeview表格插入数据
            for i in range(0, len(state_dataframe['xmlResponseName'])):
                if 'Fail' in list(state_dataframe.loc[i]):
                    SO_Item_table2.insert('', 'end', values=list(state_dataframe.loc[i]), tags='fail_line')  # state_dataframe逐行插入到表格中
                elif 'Pass' in list(state_dataframe.loc[i]):
                    SO_Item_table2.insert('', 'end', values=list(state_dataframe.loc[i]), tags='pass_line')  # state_dataframe逐行插入到表格中
                else:
                    SO_Item_table2.insert('', 'end', values=list(state_dataframe.loc[i]))

    except:
        tk.messagebox.showwarning("提示", traceback.format_exc())
